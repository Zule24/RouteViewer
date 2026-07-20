"""
add_farm.py  —  Add a new farm to the RouteViewer distance matrices
====================================================================

Run with:  python add_farm.py

Requires:  Python 3.9+, openpyxl, requests  (pip install openpyxl requests)

What it does
------------
1. Reads all existing node coordinates from extracted.xlsx (id / latitude / longitude).
2. Lets you enter the IRMA #, farm name, and address (or lat/lon directly).
3. Geocodes the address via OpenStreetMap Nominatim (free, no API key needed).
4. Queries your OSRM server for road distances and travel times from the new
   farm to every existing node in one batch request.
5. Appends the new pairs to distance_matrix.csv and duration_matrix.csv.
6. Appends the new farm's coordinates to extracted.xlsx so the registry stays current.

Place the updated CSVs next to viewer160.exe — no rebuild required.

──────────────────────────────────────────────────────────────────────────────
CONFIGURATION  ←  edit these if anything changes
──────────────────────────────────────────────────────────────────────────────
"""

# All files are expected in the same folder as this script.
COORDS_EXCEL = "extracted.xlsx"    # id / latitude / longitude
COORDS_SHEET = 0                   # sheet index (0 = first sheet)
HEADER_ROW   = 1                   # row number of the column headers
COL_IRMA     = "A"                 # column: node key (IRMA # or processor key)
COL_LAT      = "B"                 # column: latitude  (decimal degrees)
COL_LON      = "C"                 # column: longitude (decimal degrees)

# OSRM server. Change to your internal server address if needed.
OSRM_HOST = "https://router.project-osrm.org"

DIST_CSV = "distance_matrix.csv"   # road distances in km
DUR_CSV  = "duration_matrix.csv"   # travel durations in minutes

# ── Nothing below this line normally needs changing ──────────────────────────

import csv
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path
from threading import Thread

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("ERROR: requests not installed.  Run:  pip install requests")
    sys.exit(1)

HERE = Path(__file__).parent


# ── Helpers ───────────────────────────────────────────────────────────────────

def col_index(col):
    """Column letter ('A', 'B', ...) or 1-based int → 1-based int."""
    if isinstance(col, int):
        return col
    idx = 0
    for ch in str(col).strip().upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


COL_IRMA_I = col_index(COL_IRMA)
COL_LAT_I  = col_index(COL_LAT)
COL_LON_I  = col_index(COL_LON)


def normalise_key(k):
    """Strip trailing .0 from integer-like floats ('14247.0' → '14247')."""
    s = str(k).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s


def load_coords(path):
    """
    Load extracted.xlsx → {key: (lat, lon)}.
    Skips rows where id, lat, or lon are missing.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Coordinates file not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[COORDS_SHEET]
    nodes = {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        irma_raw = row[COL_IRMA_I - 1]
        lat_raw  = row[COL_LAT_I  - 1]
        lon_raw  = row[COL_LON_I  - 1]
        if irma_raw is None or lat_raw is None or lon_raw is None:
            continue
        try:
            lat = float(lat_raw)
            lon = float(lon_raw)
        except (TypeError, ValueError):
            continue
        nodes[normalise_key(irma_raw)] = (lat, lon)
    wb.close()
    return nodes


def append_to_coords_excel(path, irma, lat, lon):
    """Append a new row (irma, lat, lon) to extracted.xlsx."""
    path = Path(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[COORDS_SHEET]
    ws.append([irma, lat, lon])
    wb.save(path)
    wb.close()


def load_csv_matrix(path):
    """Load a distance/duration CSV → (dict {(rk,ck): float}, ordered key list)."""
    path = Path(path)
    dm, keys = {}, []
    if not path.exists():
        return dm, keys
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        if not headers:
            return dm, keys
        keys = [normalise_key(h) for h in headers[1:]]
        for row in reader:
            if not row:
                continue
            rk = normalise_key(row[0])
            for j, ck in enumerate(keys):
                try:
                    raw = row[j + 1].strip()
                    if raw:
                        v = float(raw)
                        dm[(rk, ck)] = v
                        dm[(ck, rk)] = v
                except (ValueError, IndexError):
                    pass
    return dm, keys


def write_csv_matrix(path, dm, keys):
    """Write a full symmetric matrix CSV."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([""] + keys)
        for rk in keys:
            row = [rk]
            for ck in keys:
                v = dm.get((rk, ck), "")
                row.append(f"{v:.6f}" if isinstance(v, float) else "")
            writer.writerow(row)


def geocode(address):
    """Geocode via OSM Nominatim. Returns (lat, lon) or raises ValueError."""
    resp = requests.get(
        "https://nominatim.openstreetmap.org/search",
        params={"q": address, "format": "json", "limit": 1},
        headers={"User-Agent": "RouteViewerAdmin/1.0"},
        timeout=10,
    )
    resp.raise_for_status()
    results = resp.json()
    if not results:
        raise ValueError(f"Address not found: {address!r}")
    return float(results[0]["lat"]), float(results[0]["lon"])


def osrm_table(new_lat, new_lon, existing_nodes):
    """
    Query OSRM table API: new farm (source 0) → all existing nodes.
    existing_nodes: [(key, lat, lon), ...]
    Returns (dist_km {key: float}, dur_min {key: float}).
    """
    coords = [f"{new_lon},{new_lat}"] + [f"{lon},{lat}" for _, lat, lon in existing_nodes]
    url = (f"{OSRM_HOST}/table/v1/driving/{';'.join(coords)}"
           f"?sources=0&destinations=all&annotations=duration,distance")
    resp = requests.get(url, headers={"User-Agent": "RouteViewerAdmin/1.0"}, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != "Ok":
        raise RuntimeError(f"OSRM error: {data.get('message', data.get('code'))}")

    durations_s = data["durations"][0]
    distances_m = data["distances"][0]
    dist_km, dur_min = {}, {}
    for i, (key, _, _) in enumerate(existing_nodes):
        d_s = durations_s[i + 1]
        d_m = distances_m[i + 1]
        if d_s is None or d_m is None:
            continue
        dur_min[key] = round(d_s / 60.0, 6)
        dist_km[key] = round(d_m / 1000.0, 6)
    return dist_km, dur_min


# ── GUI ───────────────────────────────────────────────────────────────────────

class AddFarmApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("RouteViewer — Add Farm")
        self.minsize(640, 560)
        self.resizable(True, True)
        self._coords_cache = None
        self._build_ui()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        P = {"padx": 10, "pady": 4}

        # File paths
        fp = ttk.LabelFrame(self, text="File Paths")
        fp.pack(fill="x", **P)
        self._coords_var = self._file_row(fp, "Coordinates Excel:", COORDS_EXCEL, 0)
        self._dist_var   = self._file_row(fp, "distance_matrix.csv:", DIST_CSV, 1)
        self._dur_var    = self._file_row(fp, "duration_matrix.csv:", DUR_CSV, 2)
        fp.columnconfigure(1, weight=1)

        # Farm details
        fd = ttk.LabelFrame(self, text="New Farm Details")
        fd.pack(fill="x", **P)
        self._irma_var    = self._text_row(fd, "IRMA #:", 0)
        self._name_var    = self._text_row(fd, "Farm name (optional):", 1)
        self._address_var = self._text_row(fd, "Address:", 2)
        fd.columnconfigure(1, weight=1)

        # Lat / lon + geocode
        ttk.Label(fd, text="Latitude:").grid(row=3, column=0, sticky="w", padx=6, pady=3)
        ll = ttk.Frame(fd)
        ll.grid(row=3, column=1, sticky="ew", padx=6)
        self._lat_var = tk.StringVar()
        self._lon_var = tk.StringVar()
        ttk.Entry(ll, textvariable=self._lat_var, width=16).pack(side="left")
        ttk.Label(ll, text="  Longitude:").pack(side="left")
        ttk.Entry(ll, textvariable=self._lon_var, width=16).pack(side="left", padx=(4, 0))
        ttk.Button(ll, text="⌖ Geocode address →",
                   command=self._geocode).pack(side="left", padx=(12, 0))

        # Buttons
        bf = ttk.Frame(self)
        bf.pack(fill="x", **P)
        self._add_btn = ttk.Button(bf, text="✚  Compute & Add Farm",
                                   command=self._start_add)
        self._add_btn.pack(side="left", padx=4)
        ttk.Button(bf, text="⟳  Reload coordinate list",
                   command=self._reload_coords).pack(side="left", padx=4)
        self._status = ttk.Label(bf, text="", foreground="#555")
        self._status.pack(side="left", padx=12)

        # Log
        lf = ttk.LabelFrame(self, text="Log")
        lf.pack(fill="both", expand=True, **P)
        self._log = scrolledtext.ScrolledText(lf, height=14, state="disabled",
                                              font=("Consolas", 9), wrap="word")
        self._log.pack(fill="both", expand=True, padx=4, pady=4)
        self._log_write(
            "Ready.  Enter farm details above and click 'Compute & Add Farm'.\n"
            f"Coordinates file : {HERE / COORDS_EXCEL}\n"
            f"Distance matrix  : {HERE / DIST_CSV}\n"
            f"Duration matrix  : {HERE / DUR_CSV}\n"
        )

    def _file_row(self, parent, label, default, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=3)
        var = tk.StringVar(value=str(HERE / default))
        ttk.Entry(parent, textvariable=var, width=50).grid(
            row=row, column=1, sticky="ew", padx=6)
        ttk.Button(parent, text="…",
                   command=lambda v=var: v.set(filedialog.askopenfilename() or v.get())
                   ).grid(row=row, column=2, padx=4)
        return var

    def _text_row(self, parent, label, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=3)
        var = tk.StringVar()
        ttk.Entry(parent, textvariable=var, width=50).grid(
            row=row, column=1, sticky="ew", padx=6)
        return var

    # ── Logging / status ──────────────────────────────────────────────────────

    def _log_write(self, msg):
        self._log.configure(state="normal")
        self._log.insert("end", msg)
        self._log.see("end")
        self._log.configure(state="disabled")
        self.update_idletasks()

    def _set_status(self, msg, color="#555"):
        self._status.configure(text=msg, foreground=color)
        self.update_idletasks()

    # ── Geocoding ─────────────────────────────────────────────────────────────

    def _geocode(self):
        address = self._address_var.get().strip()
        if not address:
            messagebox.showwarning("No address", "Enter an address first.")
            return
        self._set_status("Geocoding…")
        try:
            lat, lon = geocode(address)
            self._lat_var.set(f"{lat:.6f}")
            self._lon_var.set(f"{lon:.6f}")
            self._log_write(f"Geocoded '{address}' → {lat:.6f}, {lon:.6f}\n")
            self._set_status(f"Geocoded: {lat:.6f}, {lon:.6f}", "#1a7a1a")
        except Exception as exc:
            self._set_status(f"Geocode failed: {exc}", "red")
            self._log_write(f"ERROR geocoding: {exc}\n")

    # ── Coordinate list ───────────────────────────────────────────────────────

    def _reload_coords(self):
        self._coords_cache = None
        self._load_coords()

    def _load_coords(self):
        if self._coords_cache is not None:
            return self._coords_cache
        self._set_status("Loading coordinates…")
        try:
            self._coords_cache = load_coords(self._coords_var.get())
            self._log_write(f"Loaded {len(self._coords_cache)} nodes from coordinates file.\n")
            self._set_status(f"{len(self._coords_cache)} nodes loaded", "#1a7a1a")
        except Exception as exc:
            self._set_status(f"Could not load coords: {exc}", "red")
            self._log_write(f"ERROR loading coordinates: {exc}\n")
            self._coords_cache = {}
        return self._coords_cache

    # ── Main action ───────────────────────────────────────────────────────────

    def _start_add(self):
        irma  = normalise_key(self._irma_var.get().strip())
        lat_s = self._lat_var.get().strip()
        lon_s = self._lon_var.get().strip()

        if not irma:
            messagebox.showwarning("Missing field", "IRMA # is required.")
            return
        if not lat_s or not lon_s:
            messagebox.showwarning("Missing coordinates",
                                   "Enter lat/lon, or geocode an address first.")
            return
        try:
            lat, lon = float(lat_s), float(lon_s)
        except ValueError:
            messagebox.showerror("Invalid coordinates",
                                 "Latitude and longitude must be decimal numbers.")
            return

        self._add_btn.configure(state="disabled")
        self._set_status("Working…")
        Thread(target=self._run, args=(irma, lat, lon), daemon=True).start()

    def _run(self, irma, lat, lon):
        try:
            self._do_add(irma, lat, lon)
        except Exception as exc:
            self.after(0, lambda e=exc: self._set_status(f"Failed: {e}", "red"))
            self.after(0, lambda e=exc: self._log_write(f"\nERROR: {e}\n"))
        finally:
            self.after(0, lambda: self._add_btn.configure(state="normal"))

    def _do_add(self, irma, lat, lon):
        self.after(0, lambda: self._log_write(f"\n{'─'*60}\n"))
        self.after(0, lambda: self._log_write(
            f"Adding:  {irma}  ({lat:.6f}, {lon:.6f})\n"))

        # Load coordinates
        coords = self._load_coords()
        if not coords:
            raise RuntimeError("No node coordinates loaded — cannot compute distances.")

        if irma in coords:
            confirmed = [False]
            def ask():
                confirmed[0] = messagebox.askyesno(
                    "Already exists",
                    f"IRMA {irma!r} is already in the coordinate file.\n\n"
                    "Recompute its distances anyway?")
            self.after(0, ask)
            # wait for dialog
            import time
            for _ in range(100):
                time.sleep(0.1)
                if confirmed[0] is not False:
                    break
            if not confirmed[0]:
                return

        # OSRM query
        existing = [(k, v[0], v[1]) for k, v in coords.items() if k != irma]
        self.after(0, lambda: self._log_write(
            f"Querying OSRM ({len(existing)} nodes)…\n"))
        self.after(0, lambda: self._set_status("Querying OSRM…"))

        dist_km, dur_min = osrm_table(lat, lon, existing)

        routed = len(dist_km)
        missed = len(existing) - routed
        self.after(0, lambda: self._log_write(
            f"OSRM: {routed} pairs computed, {missed} unreachable/skipped.\n"))

        if not dist_km:
            raise RuntimeError("OSRM returned no valid distances. "
                               "Check server URL and coordinates.")

        # Update CSVs
        self.after(0, lambda: self._set_status("Updating CSVs…"))

        dm_dist, dist_keys = load_csv_matrix(self._dist_var.get())
        dm_dur,  dur_keys  = load_csv_matrix(self._dur_var.get())

        for key, d in dist_km.items():
            dm_dist[(irma, key)] = d;  dm_dist[(key, irma)] = d
        dm_dist[(irma, irma)] = 0.0

        for key, d in dur_min.items():
            dm_dur[(irma, key)] = d;   dm_dur[(key, irma)] = d
        dm_dur[(irma, irma)] = 0.0

        all_keys = list(dict.fromkeys(
            (dist_keys if dist_keys else dur_keys) + [irma]))

        dist_path = Path(self._dist_var.get())
        dur_path  = Path(self._dur_var.get())

        self.after(0, lambda: self._log_write(f"Writing {dist_path.name}…\n"))
        write_csv_matrix(dist_path, dm_dist, all_keys)
        self.after(0, lambda: self._log_write(f"Writing {dur_path.name}…\n"))
        write_csv_matrix(dur_path, dm_dur, all_keys)

        # Append to extracted.xlsx
        coords_path = Path(self._coords_var.get())
        if irma not in coords:
            self.after(0, lambda: self._log_write(f"Updating {coords_path.name}…\n"))
            append_to_coords_excel(coords_path, irma, lat, lon)
            # Invalidate cache so next add sees the new farm
            self._coords_cache = None

        summary = (
            f"\n✓  Done!  Farm {irma} added.\n"
            f"   Pairs computed : {routed}\n"
            f"   Matrix size    : {len(all_keys)} × {len(all_keys)} nodes\n"
            f"\n   Copy the updated CSVs and extracted.xlsx next to viewer160.exe\n"
            f"   and relaunch — no rebuild required.\n"
        )
        self.after(0, lambda: self._log_write(summary))
        self.after(0, lambda: self._set_status(
            f"✓  Farm {irma} added ({routed} pairs)", "#1a7a1a"))


if __name__ == "__main__":
    AddFarmApp().mainloop()
