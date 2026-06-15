"""
viewer.py  ?  Route Sheet Viewer
"""

import re, sys, csv, copy, random, math, time, uuid, logging
from pathlib import Path
from datetime import time as dt_time, datetime, date, timedelta

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QComboBox, QPushButton, QTableWidget, QTableWidgetItem,
    QFrame, QHeaderView, QAbstractItemView, QTabWidget, QSplitter,
    QScrollBar, QScrollArea, QDoubleSpinBox, QSpinBox, QProgressBar,
    QGroupBox, QCheckBox, QTextEdit, QSizePolicy, QDialog,
    QLineEdit, QMessageBox, QFileDialog
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QMimeData, QByteArray
from PyQt5.QtGui import QFont, QColor, QDrag, QPainter, QPen

import openpyxl

# -- Logging -------------------------------------------------------------------
# Module logger. A NullHandler keeps libraries quiet by default; main() attaches
# a real handler so parsing/data-load problems surface instead of vanishing.
logger = logging.getLogger("viewer")
logger.addHandler(logging.NullHandler())

# -- Constants -----------------------------------------------------------------

IRMA_RE     = re.compile(r"^\d{2}-\d{3}$")
SHEET_RE    = re.compile(r"^\d{4}$")  # kept for display logic
EXCLUDE_SHEETS = {"INFO", "SHEET1", "SHEET2"}  # sheets with no route data
MONTH_ORDER = ["january","february","march","april","may","june",
               "july","august","september","october","november","december"]
VOL_LIMIT   = 41500

# Litres assumed to be on a preload trailer at the start of the day. This is the
# starting load that gets offloaded before the day's collection begins.
# NOTE: intentionally distinct from VOL_LIMIT (the truck's collection capacity).
PRELOAD_VOL = 40000

C_IRMA=1; C_TRAIN=4; C_M1_START=5; C_M1_FINISH=8; C_M2_START=11
C_M2_FINISH=14; C_EDPU=17; C_ROUTE=21; C_LOCATION=32; C_PRIOR_VOL=51
C_DRIVER_START = 52   # AZ1 ? driver start time (datetime.time)
C_DAY_COLOUR   = 62   # BJ1 ? day colour string (RED/BLUE/GRASSFED/A2 etc.)

DRIVE_SPEED_KMH = 50.0   # km/h average
ONSITE_MIN      = 15.0   # fixed on-site setup minutes per stop
PUMP_RATE_LPM   = 750.0  # litres per minute

VEDDER_DEPART_EXTRA_MINS = 40   # extra minutes added to shift start (Vedder departure)
PRELOAD_WASH_MINS        = 75   # wash time added after a preload offload (1h 15m)
INTER_PROCESSOR_BREAK    = 10   # break minutes inserted between processor stops

# Farms whose milking windows can be suppressed (e.g. robots / continuous milking)
NO_MILKING_WINDOW_FARMS = {"37-874", "14-247", "92-545"}

# Three-window farm data loaded from JSON at startup
def _load_three_window_farms():
    import json
    p = get_data_dir() / "three_window_farms.json"
    if p.exists():
        try:
            with open(p, encoding="utf-8") as _f:
                return json.load(_f)
        except Exception as ex:
            logger.warning("Could not load %s: %s", p, ex)
    return {}

THREE_WINDOW_FARMS: dict = {}   # populated in main() after get_data_dir() is available

COLS = [
    ("IRMA #",       "irma"),
    ("Proc ID",      "proc_id"),
    ("Train",        "train"),
    ("M1 Start",     "m1_start"),
    ("M1 Finish",    "m1_finish"),
    ("M2 Start",     "m2_start"),
    ("M2 Finish",    "m2_finish"),
    ("EDPU",         "edpu"),
    ("Name / Location","location"),
    ("Prior Vol (L)","prior_vol"),
    ("Dist to",      "dist"),
    ("Arr.",         "arr_time"),
    ("Wait",         "wait_time"),
    ("Depart",       "dep_time"),
    ("MWO",          "_mwo"),
]

MWO_COL = next(i for i, (_, k) in enumerate(COLS) if k == "_mwo")

# Sheet names (exact, case-insensitive) that the solver leaves untouched.
SOLVER_SKIP_SHEETS = {"1603", "1604",
                      "1531", "1021", "1031", "1071", "1125",
                      "1081", "1451", "1281", "1441", "1121", "1561", "1211"}

# Default plant receiving windows (open HH:MM, close HH:MM).
# None means 24/7 ? no restriction.  Overnight windows (close < open) are
# handled by time_in_window().  These are hard-coded from the
# "Plant_Receiving_Windows" reference sheet; the user can override them in
# the Solver tab at run time.
PLANT_RECEIVING_WINDOWS = {
    "909312": ("00:00", "23:59"),  # Agropur Burnaby        ? 24/7
    "972711": ("05:00", "23:00"),  # Saputo Port Coquitlam  ? 5am?11pm
    "902011": ("06:00", "18:00"),  # Avalon Dairy           ? 6am?6pm
    "907011": ("10:00", "16:00"),  # Birchwood Dairy        ? 10am?4pm
    "906011": ("06:00", "18:00"),  # Dhaliwal Dairy         ? 6am?6pm
    "965713": ("06:00", "18:00"),  # First Choice           ? 6am?6pm
    "911011": ("18:00", "20:00"),  # Golden Ears PM         ? 6pm?8pm
    "918011": ("06:00", "18:00"),  # Khalsa FY              ? 6am?6pm
    "901012": ("06:00", "18:00"),  # Olympic Dairy          ? 6am?6pm
    "905011": ("06:00", "18:00"),  # Pinnacle Dairy         ? 6am?6pm
    "916011": ("06:00", "18:00"),  # Prabu Foods            ? 6am?6pm
    "951305": ("06:00", "18:00"),  # Reva Foods             ? 6am?6pm
    "917011": ("08:00", "11:00"),  # Ridgecrest             ? 8am?11am
    "981301": ("08:00", "11:00"),  # WOW Foods              ? 8am?11am
    "912011": ("05:00", "17:00"),  # Meadowfresh (regular)  ? 5am?5pm
    "908011": ("05:00", "00:30"),  # Punjab                 ? 5am?12:30am (overnight)
    "915011": ("00:00", "23:59"),  # Vitalus Abbotsford     ? 24/7
    "972712": ("06:00", "23:59"),  # Saputo Abbotsford      ? 6am?midnight
    "902013": ("06:00", "18:00"),  # GRASSFED Avalon        ? 6am?6pm
    "907012": ("06:00", "18:00"),  # GRASSFED Birchwood     ? 6am?6pm
    "929011": ("06:00", "18:00"),  # Earth's Own / A2       ? 6am?6pm
    "912015": ("05:00", "17:00"),  # GRASSFED Meadowfresh   ? 5am?5pm
    "913011": ("06:00", "08:00"),  # Farmhouse Agassiz      ? 6am?8am
}

# Tray uses the same columns plus a "From Route" column
TRAY_COLS = COLS + [("Sheet / Route", "_from_route"), ("Type", "_day_colour")]

CLR_HEADER    = QColor("#1a3a5c")
CLR_HEADER_FG = QColor("#ffffff")
CLR_ROUTE_HDR = QColor("#2e6da4")
CLR_ROUTE_FG  = QColor("#ffffff")
CLR_SUBTOTAL  = QColor("#d0e4f7")
CLR_TOTAL     = QColor("#a8c8f0")
CLR_ALT       = QColor("#f0f6fc")
CLR_WHITE     = QColor("#ffffff")
CLR_DEPOT     = QColor("#e8f5e9")
CLR_DEST      = QColor("#fff3e0")
CLR_DEST_WARN = QColor("#ffccbc")   # orange-red: dest arrival outside plant window
CLR_REMOVED   = QColor("#fff8e1")
CLR_RED       = QColor("#c0392b")
CLR_RED_BG    = QColor("#fdecea")
CLR_CHANGED   = QColor("#fff9c4")   # highlight for changed farms in comparison

# -- Helpers -------------------------------------------------------------------

def get_exe_dir():
    """Directory of the executable (or script). Used for the browse-default root."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent

def get_data_dir():
    """Directory where bundled data files live.
    When frozen by PyInstaller --onefile, data files are unpacked to
    sys._MEIPASS at runtime, not next to the exe."""
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            return Path(meipass)
        return Path(sys.executable).parent
    return Path(__file__).parent

def fmt_time(v):
    if isinstance(v, dt_time): return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, str): return v
    return ""

def extract_year(name):
    m = re.search(r"\d{4}", name)
    return m.group(0) if m else name

def month_key(name):
    n = name.lower()
    for i, m in enumerate(MONTH_ORDER):
        if m in n: return i
    return 99

# -- Milking window conflict --------------------------------------------------

def parse_hhmm(s):
    """Parse 'HH:MM' string -> datetime.time, or return None.

    Also accepts:
    - datetime.time directly (pass-through) ? handles cells that openpyxl
      returns as time objects rather than strings on some Excel files.
    - float Excel time serial (fraction of a day) ? converts to time.
    """
    if s is None: return None
    if isinstance(s, dt_time): return s
    if isinstance(s, float):
        # Excel serial: fraction of 24h
        total_mins = round(s * 24 * 60)
        return dt_time(total_mins // 60 % 24, total_mins % 60)
    if not s or s == "-": return None
    try:
        h, m = str(s).split(":")
        return dt_time(int(h), int(m))
    except (ValueError, AttributeError):
        return None

def time_in_window(t, start_s, finish_s):
    """Return True if time t falls within [start_s, finish_s] (HH:MM strings).
    Handles overnight windows (finish < start)."""
    if t is None: return False
    ts = parse_hhmm(start_s); tf = parse_hhmm(finish_s)
    if ts is None or tf is None: return False
    if ts <= tf:
        return ts <= t <= tf
    else:  # overnight
        return t >= ts or t <= tf

def arrives_during_milking(arr_time, row_data, suppress_no_milking=True):
    """Return True if arr_time (datetime.time) falls within any milking window.
    Supports w1/w2 from the sheet and w3 from THREE_WINDOW_FARMS.
    If suppress_no_milking is True, farms in NO_MILKING_WINDOW_FARMS always return False."""
    if arr_time is None: return False
    irma = row_data.get("irma", "")
    if suppress_no_milking and irma in NO_MILKING_WINDOW_FARMS:
        return False
    for start_key, finish_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
        if time_in_window(arr_time, row_data.get(start_key,""), row_data.get(finish_key,"")):
            return True
    # Third window from THREE_WINDOW_FARMS
    w3 = THREE_WINDOW_FARMS.get(irma)
    if w3:
        w3_start = w3.get("w3", [None, None])[0]
        w3_finish = w3.get("w3", [None, None])[1]
        if w3_start and w3_finish and time_in_window(arr_time, w3_start, w3_finish):
            return True
    return False


# -- Distance matrix -----------------------------------------------------------

def normalise_key(k):
    s = str(k).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): s = s[:-2]
    return s

def load_distance_matrix(path):
    dm = {}
    if not path.exists():
        logger.warning("Distance matrix not found at %s ? distances unavailable", path)
        return dm
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        if not headers: return dm
        col_keys = [normalise_key(h) for h in headers[1:]]
        for row in reader:
            if not row: continue
            rk = normalise_key(row[0])
            for j, ck in enumerate(col_keys):
                try:
                    raw = row[j+1].strip()
                    if not raw: continue
                    val = float(raw)
                except (ValueError, IndexError):
                    continue
                dm[(rk, ck)] = val
                dm[(ck, rk)] = val
    return dm

def lookup(dm, a, b):
    return dm.get((normalise_key(a), normalise_key(b)))

def _block_dest_keys(block):
    """Return ordered list of destination keys for a block (may be multiple)."""
    dests = block.get("dests") or []
    if dests:
        return [d["key"] for d in dests if d.get("key")]
    # Legacy fallback
    dk = block.get("dest_key", "")
    return [dk] if dk else []


def _block_last_dest_key(block):
    """The final destination key of a block (used as origin of next block)."""
    keys = _block_dest_keys(block)
    return keys[-1] if keys else ""


def _dest_stop_index(block, d_i, b_idx, blocks):
    """Return the btimes/dists index for the d_i-th dest in block,
    respecting split_after. Preload blocks always put dest at index 1+d_i."""
    if block.get("preload") and not block.get("rows"):
        return 1 + d_i
    is_last = (b_idx == len(blocks) - 1)
    origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
    stops   = _build_block_stops(block, origin, is_last)
    dest_stops = [s for s in stops if s["type"] == "dest"]
    if d_i < len(dest_stops):
        return dest_stops[d_i]["_si"]
    return len(block.get("rows",[])) + 1 + d_i   # fallback


def _farm_stop_index(block, f_i, b_idx, blocks):
    """Return the btimes/dists index for the f_i-th farm in block,
    respecting any split_after dests that appear before this farm."""
    is_last = (b_idx == len(blocks) - 1)
    origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
    stops   = _build_block_stops(block, origin, is_last)
    farm_stops = [s for s in stops if s["type"] == "farm"]
    if f_i < len(farm_stops):
        return farm_stops[f_i]["_si"]
    return f_i + 1   # fallback


def _block_has_split(block):
    """Return True if any dest in this block has a split_after index set,
    meaning it should be visited mid-route rather than after all farms."""
    dests = block.get("dests") or []
    return any(d.get("split_after") is not None for d in dests)


def _build_block_stops(block, origin, is_last):
    """Build the ordered stop sequence for a block, respecting split_after.

    Returns a list of dicts, each with:
      type:   'origin' | 'farm' | 'dest' | 'vedder'
      key:    irma or dest_key string (for distance lookup)
      farm:   farm row dict (type=='farm' only)
      dest:   dest dict (type=='dest' only)

    split_after on a dest dict means "insert this dest after farm index N".
    None / missing means after all farms (legacy behaviour).
    """
    farms = block.get("rows", [])
    dests = block.get("dests") or []
    if not dests:
        dk = block.get("dest_key","")
        dn = block.get("dest_name","") or dk
        dests = [{"key": dk, "name": dn, "vol_partial": None}] if dk else []

    stops = [{"type": "origin", "key": origin}]

    # Group dests by their split_after index.
    # split_after=None means after all farms (index = len(farms)).
    from collections import defaultdict
    dests_by_split = defaultdict(list)
    for d in dests:
        sa = d.get("split_after")
        if sa is None:
            sa = len(farms)   # after all farms
        dests_by_split[sa].append(d)

    for f_idx, farm in enumerate(farms):
        stops.append({"type": "farm", "key": farm.get("irma",""), "farm": farm})
        # Insert any dests that split after this farm.
        # Skip f_idx+1 == len(farms) ? those are handled by the "after all farms"
        # loop below to avoid double-insertion.
        if f_idx + 1 < len(farms):
            for d in dests_by_split.get(f_idx + 1, []):
                stops.append({"type": "dest", "key": d.get("key",""), "dest": d})

    # Dests with split_after=len(farms) (after all farms)
    for d in dests_by_split.get(len(farms), []):
        stops.append({"type": "dest", "key": d.get("key",""), "dest": d})

    if is_last:
        stops.append({"type": "vedder", "key": "VEDDER"})

    # Tag each stop with its sequence index for btimes/dists lookups
    for i, s in enumerate(stops):
        s["_si"] = i

    return stops


def _is_holdover_block(block):
    """Return True if every destination in this block is a 'Yard for ...' location.

    These blocks end at a yard trailer ? the farms are collected the same day
    but the milk sits overnight before being picked up.  The FARM ORDER is
    still optimisable; only the final yard destination must stay constant.
    Use _is_preload_block() to detect the next-morning pickup blocks that must
    be held completely frozen.
    """
    dests = block.get("dests") or []
    if not dests:
        dk = block.get("dest_key", "")
        dn = block.get("dest_name", "")
        dests = [{"name": dn, "key": dk}] if (dk or dn) else []
    if not dests:
        return False
    return all("yard for" in (d.get("name", "") or "").lower() for d in dests)


def _is_preload_block(block):
    """Return True if this block is a preload offload ? no farms, just a
    delivery to a processor at the start of the day (previous night's load).

    These blocks must be held completely constant: no farms added, destination
    not changed.  Identified by the 'preload' flag set during parsing.
    """
    return bool(block.get("preload")) and not block.get("rows")


def _is_fixed_vol_block(block):
    """Return True if this block has a fixed partial-volume delivery on its
    last destination (no catch-all remainder dest).

    When every destination in a block has an explicit vol_partial, the total
    deliverable volume is capped at their sum.  Any farm volume above that cap
    is silently dropped from accounting.  Rather than trying to optimise around
    this, the solver holds these blocks completely frozen ? the dispatcher has
    explicitly specified the volumes and the solver has no business changing the
    farm assignment.
    """
    dests = block.get("dests") or []
    if not dests:
        return False
    return dests[-1].get("vol_partial") is not None


def calc_distances(blocks, dm):
    """
    Routing rules (multi-destination aware, split_after supported):
      - First block starts from VEDDER.
      - Each subsequent block starts from the last destination of the previous block.
      - Within a block: stops follow _build_block_stops order, which interleaves
        partial-dropoff dests at their split_after position.
      - Last block appends a return leg to VEDDER after the final destination.
    Returns list of per-block distance lists (one entry per stop, None sentinel last).
    """
    n = len(blocks)
    all_dists = []
    for b_idx, block in enumerate(blocks):
        is_last = (b_idx == n - 1)
        origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        stops   = _build_block_stops(block, origin, is_last)
        keys    = [s["key"] for s in stops]
        dists   = [lookup(dm, keys[i], keys[i+1]) if i < len(keys)-1 else None
                   for i in range(len(keys))]
        all_dists.append(dists)
    return all_dists

def calc_durations(blocks, dm_dur):
    """
    Parallel to calc_distances but returns drive-time minutes from the duration
    matrix.  Respects split_after via _build_block_stops.
    """
    n = len(blocks)
    all_durs = []
    for b_idx, block in enumerate(blocks):
        is_last = (b_idx == n - 1)
        origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        stops   = _build_block_stops(block, origin, is_last)
        keys    = [s["key"] for s in stops]
        durs    = [lookup(dm_dur, keys[i], keys[i+1]) if i < len(keys)-1 else None
                   for i in range(len(keys))]
        all_durs.append(durs)
    return all_durs


# -- Time estimation ----------------------------------------------------------

def mins_to_time(start_time, offset_mins):
    """Add offset_mins (float) to a datetime.time. Returns datetime.time."""
    dt = datetime.combine(date.today(), start_time) + timedelta(minutes=offset_mins)
    return dt.time()

def fmt_hhmm(t):
    if t is None: return "-"
    return f"{t.hour:02d}:{t.minute:02d}"

def day_colour_style(day_colour):
    """Return (bg QColor, fg QColor, display_text) for a day colour string."""
    dc = day_colour.upper().replace("\n", " ").strip()
    if "GRASSFED"  in dc or "GRASS FED" in dc or "GRASS" in dc:
        return QColor("#43a047"), QColor("#ffffff"), day_colour.replace("\n"," ")
    if "A2"        in dc: return QColor("#b3e5fc"), QColor("#0d47a1"), day_colour
    if "RED"       in dc: return QColor("#e53935"), QColor("#ffffff"), day_colour
    if "BLUE"      in dc: return QColor("#1e88e5"), QColor("#ffffff"), day_colour
    if "CREAM"     in dc: return QColor("#fff9c4"), QColor("#5d4037"), day_colour
    if "SKIM"      in dc: return QColor("#e0f7fa"), QColor("#006064"), day_colour
    if "WPC"       in dc: return QColor("#f3e5f5"), QColor("#6a1b9a"), day_colour
    if "HOLD"      in dc: return QColor("#ffccbc"), QColor("#bf360c"), day_colour.replace("\n"," ")
    if dc:                return QColor("#9e9e9e"), QColor("#ffffff"), day_colour
    return None, None, ""


def stop_duration(vol_litres):
    """Minutes spent at a stop: setup + pump time."""
    pump = (vol_litres / PUMP_RATE_LPM) if vol_litres else 0.0
    return ONSITE_MIN + pump

def drive_mins(km):
    """Minutes to drive km at DRIVE_SPEED_KMH."""
    if km is None: return None
    return (km / DRIVE_SPEED_KMH) * 60.0

def _dest_vol_partial(dest_dict, total_farm_vol, already_delivered):
    """
    Work out how many litres are offloaded at this destination stop.
    dest_dict: {name, key, vol_partial}  vol_partial=None means "rest of load".
    already_delivered: total already dropped at earlier dests in same block.
    """
    remaining = max(0.0, total_farm_vol - already_delivered)
    vp = dest_dict.get("vol_partial")
    if vp is None:
        return remaining
    return min(float(vp), remaining)


def _block_dest_offloads(block):
    """Litres offloaded per destination key for a single block -> {key: litres}.

    Walks the block's destinations in order, assigning each its vol_partial
    (or the remaining farm volume when vol_partial is None), and aggregates by
    destination key.  When the block has no explicit dests, falls back to
    dest_key / dest_name.  This is the single source of truth for the
    volume-accounting loop shared by the cost, penalty, and diagnostic code.
    """
    dests_b = block.get("dests") or []
    if not dests_b:
        dk = block.get("dest_key") or block.get("dest_name") or "?"
        dests_b = [{"key": dk, "vol_partial": None}]
    farm_vol = sum((r.get("prior_vol") or 0) for r in block.get("rows", [])
                   if isinstance(r.get("prior_vol"), (int, float)))
    out = {}
    already = 0.0
    for d in dests_b:
        dk = d.get("key") or "?"
        vp = d.get("vol_partial")
        rem = max(0.0, farm_vol - already)
        offload = min(float(vp), rem) if vp is not None else rem
        already += offload
        out[dk] = out.get(dk, 0.0) + offload
    return out


def calc_times(blocks, dm, start_time, dm_dur=None, suppress_no_milking=True,
               precomputed_dists=None):
    """
    Returns list of per-block time lists in parallel with calc_distances.
    Each block: list of dicts {arr, dep, wait} per stop, ordered by
    _build_block_stops (which respects split_after for mid-route partial dropoffs).

    Multi-destination with split_after: a dest with split_after=N is visited
    after farm N in the sequence, allowing partial dropoff mid-route.

    precomputed_dists: optional output of calc_distances(blocks, dm) for the same
    blocks.  Passing it avoids recomputing the distance legs when the caller has
    already done so (e.g. _sheet_cost), which roughly halves the distance work in
    the solver's hot path.
    """
    if start_time is None:
        return None

    all_dists    = precomputed_dists if precomputed_dists is not None \
                   else calc_distances(blocks, dm)
    all_dur_lists = calc_durations(blocks, dm_dur) if dm_dur else None
    all_times    = []
    base   = datetime.combine(date.today(), start_time)
    cursor = base

    for b_idx, block in enumerate(blocks):
        dists  = all_dists[b_idx]
        durs   = all_dur_lists[b_idx] if all_dur_lists else None
        is_last = (b_idx == len(blocks) - 1)
        origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        stops   = _build_block_stops(block, origin, is_last)
        block_times = []

        def _leg_mins(s_idx):
            if durs is not None and s_idx < len(durs) and durs[s_idx] is not None:
                return durs[s_idx]
            d = dists[s_idx] if s_idx < len(dists) else None
            return drive_mins(d)

        # Preload block: no farms, drive straight to dests then wash
        if block.get("preload") and not block.get("rows"):
            already_del  = 0.0
            real_dests   = [s for s in stops if s["type"] == "dest"
                            and "yard for" not in (s["dest"].get("name","") or "").lower()]
            n_real       = len(real_dests)
            real_seen    = 0
            for s_idx, stop in enumerate(stops):
                if stop["type"] == "origin":
                    if b_idx == 0:
                        dep_dt = cursor + timedelta(minutes=VEDDER_DEPART_EXTRA_MINS)
                        block_times.append({"arr": cursor.time(), "dep": dep_dt.time(), "wait": None})
                        cursor = dep_dt
                    else:
                        block_times.append({"arr": cursor.time(), "dep": cursor.time(), "wait": None})
                    continue
                if stop["type"] == "dest":
                    d      = stop["dest"]
                    is_yard = "yard for" in (d.get("name","") or "").lower()
                    dm_m   = 0.0 if is_yard else _leg_mins(s_idx - 1)
                    if dm_m is None:
                        block_times.append({"arr": None, "dep": None, "wait": None})
                        continue
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    vp     = d.get("vol_partial")
                    rem    = max(0.0, PRELOAD_VOL - already_del)
                    offload = min(float(vp), rem) if vp is not None else rem
                    already_del += offload
                    dep_dt = arr_dt + timedelta(minutes=stop_duration(offload))
                    if not is_yard:
                        real_seen += 1
                        if real_seen < n_real:
                            dep_dt += timedelta(minutes=INTER_PROCESSOR_BREAK)
                    block_times.append({"arr": arr_dt.time(), "dep": dep_dt.time(), "wait": None})
                    cursor = dep_dt
            cursor += timedelta(minutes=PRELOAD_WASH_MINS)
            all_times.append(block_times)
            continue

        # Normal block: farms possibly interleaved with partial-dropoff dests
        total_farm_vol  = sum((r.get("prior_vol") or 0) for r in block.get("rows",[])
                               if isinstance(r.get("prior_vol"), (int, float)))
        already_del     = 0.0
        real_dests      = [s for s in stops if s["type"] == "dest"
                           and "yard for" not in (s["dest"].get("name","") or "").lower()]
        n_real_dests    = len(real_dests)
        real_dest_seen  = 0

        for s_idx, stop in enumerate(stops):
            stype = stop["type"]

            if stype == "origin":
                if b_idx == 0:
                    dep_dt = cursor + timedelta(minutes=VEDDER_DEPART_EXTRA_MINS)
                    block_times.append({"arr": cursor.time(), "dep": dep_dt.time(), "wait": None})
                    cursor = dep_dt
                else:
                    block_times.append({"arr": cursor.time(), "dep": cursor.time(), "wait": None})

            elif stype == "farm":
                farm   = stop["farm"]
                dm_m   = _leg_mins(s_idx - 1)
                if dm_m is None or cursor is None:
                    block_times.append({"arr": None, "dep": None, "wait": None})
                else:
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    vol    = farm.get("prior_vol")
                    # Zero-vol farms in a paired set (same IRMA as another farm
                    # in the block with non-zero vol) contribute 0 minutes ?
                    # the truck arrives, connects the trailer, and leaves with
                    # no pump time and no setup time.  arr == dep, no wait check.
                    if isinstance(vol, (int, float)) and vol == 0:
                        block_times.append({"arr": arr_dt.time(), "dep": arr_dt.time(),
                                            "wait": None})
                        cursor = arr_dt
                        continue
                    vol    = vol or 0
                    dur    = stop_duration(vol if isinstance(vol, (int, float)) else 0)
                    irma   = farm.get("irma","")
                    wait_mins = 0.0

                    # MWO (Milking Window Override): skip all milking window
                    # checks for this farm ? truck arrives and pumps immediately.
                    if farm.get("_mwo"):
                        pass
                    elif suppress_no_milking and irma in NO_MILKING_WINDOW_FARMS:
                        pass
                    else:
                        for s_key, f_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
                            if time_in_window(arr_dt.time(), farm.get(s_key,""), farm.get(f_key,"")):
                                tf = parse_hhmm(farm.get(f_key,""))
                                if tf:
                                    end_w = datetime.combine(date.today(), tf)
                                    if end_w <= arr_dt: end_w += timedelta(days=1)
                                    wait_mins = (end_w - arr_dt).total_seconds() / 60.0
                                break
                        if wait_mins == 0.0:
                            w3data = THREE_WINDOW_FARMS.get(irma)
                            if w3data:
                                w3s, w3f = w3data.get("w3",[None,None])
                                if w3s and w3f and time_in_window(arr_dt.time(), w3s, w3f):
                                    tf3 = parse_hhmm(w3f)
                                    if tf3:
                                        end_w3 = datetime.combine(date.today(), tf3)
                                        if end_w3 <= arr_dt: end_w3 += timedelta(days=1)
                                        wait_mins = (end_w3 - arr_dt).total_seconds() / 60.0

                    dep_dt = arr_dt + timedelta(minutes=wait_mins + dur)
                    block_times.append({"arr": arr_dt.time(), "dep": dep_dt.time(),
                                        "wait": wait_mins if wait_mins > 0 else None})
                    cursor = dep_dt

            elif stype == "dest":
                d       = stop["dest"]
                is_yard = "yard for" in (d.get("name","") or "").lower()
                dm_m    = 0.0 if is_yard else _leg_mins(s_idx - 1)
                # Volume collected so far up to this split point
                if is_yard:
                    offload = 0.0
                else:
                    offload = _dest_vol_partial(d, total_farm_vol, already_del)
                already_del += offload
                if dm_m is None or cursor is None:
                    block_times.append({"arr": None, "dep": None, "wait": None})
                else:
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    dep_dt = arr_dt + timedelta(minutes=stop_duration(offload))
                    if not is_yard:
                        real_dest_seen += 1
                        if real_dest_seen < n_real_dests:
                            dep_dt += timedelta(minutes=INTER_PROCESSOR_BREAK)
                    block_times.append({"arr": arr_dt.time(), "dep": dep_dt.time(), "wait": None})
                    cursor = dep_dt

            elif stype == "vedder":
                dm_m = _leg_mins(s_idx - 1)
                if dm_m is None or cursor is None:
                    block_times.append({"arr": None, "dep": None, "wait": None})
                else:
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    block_times.append({"arr": arr_dt.time(), "dep": arr_dt.time(), "wait": None})

        all_times.append(block_times)

    if not all_times:
        return None
    return all_times, cursor if cursor is not None else base


def total_route_dist(blocks, dm):
    """Sum all known distances across blocks. Returns (total_km, all_known bool)."""
    total = 0.0; ok = True
    for dists in calc_distances(blocks, dm):
        for d in dists[:-1]:
            if d is None: ok = False
            else: total += d
    return total, ok

# -- Excel parsing -------------------------------------------------------------

# -- Delivery-info column constants (1-based) ---------------------------------
# Col 1 = row number label (1., 2., 3.)
# Col 2 = partial volume (e.g. 6900) or "Full Load" or empty
# Col 6 = processor name
# Col 20 = processor key (numeric)
C_DEST_VOL  = 2    # partial volume column in delivery rows
C_DEST_NAME = 6    # processor name column
C_DEST_KEY  = 20   # processor key column


def _parse_dest_row(ws, r, ws_formula=None):
    """
    Parse one numbered delivery row.  Returns a dest dict:
      {name, key, vol_partial}
    vol_partial is a float if a specific partial volume was given,
    or None meaning "rest of load" (blank or "Full Load").
    Returns None if the row has no meaningful content.
    """
    dn_raw = ws.cell(r, C_DEST_NAME).value
    dk_raw = ws.cell(r, C_DEST_KEY).value
    # Must have at least a name or a key to be a real dest row
    dn = str(dn_raw).strip() if dn_raw else ""
    dk = (str(int(dk_raw)).strip()
          if isinstance(dk_raw, (int, float))
          else str(dk_raw or "").strip())
    if not dn and not dk:
        return None
    # "WASH AT VTL" and similar wash instructions are operational notes,
    # not processor destinations ? exclude them.
    if "wash" in dn.lower():
        return None
    vol_raw = ws.cell(r, C_DEST_VOL).value
    if vol_raw is None and ws_formula is not None:
        vol_raw = _try_eval_formula(ws_formula.cell(r, C_DEST_VOL).value)
    if vol_raw is None or (isinstance(vol_raw, str)
                           and vol_raw.strip().lower() in ("", "full load")):
        vol_partial = None
    else:
        try:
            vol_partial = float(vol_raw)
        except (ValueError, TypeError):
            vol_partial = None
    return {"name": dn, "key": dk, "vol_partial": vol_partial}


def parse_sheet(ws, ws_formula=None):
    """Parse one worksheet.  ws is the data_only workbook sheet;
    ws_formula (optional) is the same sheet from a formula workbook ? used
    to recover numeric values when the cached data value is None (e.g. the
    file was saved without recalculating formulas)."""

    # In read_only mode openpyxl yields EmptyCell objects for empty cells.
    # EmptyCell has .value=None but no .row/.column attributes ? normalise
    # them to None so the rest of the parser never sees them.
    def _cell_value(cell):
        try:
            return cell.value
        except AttributeError:
            return None

    def _cell(r, c):
        """Read a cell, falling back to formula evaluation if the cached value is None."""
        val = ws.cell(r, c).value
        if val is None and ws_formula is not None:
            val = _try_eval_formula(ws_formula.cell(r, c).value)
        return val
    # Read driver start time from AZ1 (col 52, row 1)
    raw_start = ws.cell(1, C_DRIVER_START).value
    if isinstance(raw_start, dt_time):
        driver_start = raw_start
    else:
        driver_start = None

    # Read day colour from BJ1 (col 62, row 1)
    raw_colour = ws.cell(1, C_DAY_COLOUR).value
    day_colour = str(raw_colour).strip().upper() if raw_colour else ""

    # Determine the effective row range without walking every cell.
    # ws.max_row can be inflated by phantom rows (e.g. from Excel scrolling
    # or accidental edits far down the sheet).  Instead of iter_rows() which
    # materialises every cell, scan only the columns the parser cares about
    # to find the last row that actually has content.  This is O(real_data)
    # rather than O(max_row * max_col) and avoids stalls on large June files.
    MAX_SCAN_ROWS = min(ws.max_row, 5000)   # safety cap ? no route sheet has 5000 rows
    SCAN_COLS = {C_IRMA, 2, 6}             # IRMA, col-2 (delivery/route headers), col-6 (dest)
    last_data_row = 0
    for r in range(1, MAX_SCAN_ROWS + 1):
        for c in SCAN_COLS:
            if ws.cell(r, c).value is not None:
                last_data_row = r
                break
    all_row_nums = list(range(1, last_data_row + 1))
    irma_header_rows = []
    # dest_list_rows maps irma_header_row -> [dest_dict, ...]
    dest_list_rows = {}
    # Delivery info found before any IRMA# row -> preload dests
    preload_dests = []

    i = 0
    while i < len(all_row_nums):
        r    = all_row_nums[i]
        val0 = ws.cell(r, C_IRMA).value
        c2   = ws.cell(r, 2).value

        if isinstance(val0, str) and val0.strip().upper() == "IRMA#":
            irma_header_rows.append(r)

        # Detect "Delivery Information:" block ? grab numbered rows that follow
        if isinstance(c2, str) and "delivery" in c2.lower():
            hdr_key = irma_header_rows[-1] if irma_header_rows else None
            if hdr_key is not None:
                dest_list_rows.setdefault(hdr_key, [])
                j = i + 1
                while j < len(all_row_nums):
                    dr = all_row_nums[j]
                    # Stop at empty row or another section header
                    c2j = ws.cell(dr, 2).value
                    if c2j is None and ws.cell(dr, C_DEST_NAME).value is None:
                        break
                    d = _parse_dest_row(ws, dr, ws_formula=ws_formula)
                    if d:
                        dest_list_rows[hdr_key].append(d)
                    j += 1
                i = j
                continue
            else:
                # Delivery info before any IRMA# row ? this is a preload offload block.
                # Capture dests into preload_dests (only the first such section).
                if not preload_dests:
                    j = i + 1
                    while j < len(all_row_nums):
                        dr  = all_row_nums[j]
                        c2j = ws.cell(dr, 2).value
                        if c2j is None and ws.cell(dr, C_DEST_NAME).value is None:
                            break
                        d = _parse_dest_row(ws, dr, ws_formula=ws_formula)
                        if d:
                            preload_dests.append(d)
                        j += 1
                    i = j
                    continue

        # Legacy "Destination:" label in col 6 (single-dest fallback)
        c6 = ws.cell(r, 6).value
        if isinstance(c6, str) and "destination" in c6.lower():
            hdr_key = irma_header_rows[-1] if irma_header_rows else None
            if hdr_key is not None and hdr_key not in dest_list_rows:
                nxt = i + 1
                if nxt < len(all_row_nums):
                    dr = all_row_nums[nxt]
                    d = _parse_dest_row(ws, dr, ws_formula=ws_formula)
                    if d:
                        dest_list_rows[hdr_key] = [d]

        i += 1

    blocks  = []
    current = None

    # If we found a preload delivery section before any IRMA# block, prepend it.
    if preload_dests:
        first_p = preload_dests[0]
        blocks.append({
            "route":     "",
            "dests":     preload_dests,
            "dest_name": first_p["name"],
            "dest_key":  first_p["key"],
            "rows":      [],
            "preload":   True,
        })
    for r in all_row_nums:
        val0 = ws.cell(r, C_IRMA).value
        if isinstance(val0, str) and val0.strip().upper() == "IRMA#":
            route_val = ws.cell(r, C_ROUTE).value or ""
            dests = dest_list_rows.get(r, [])
            # Legacy compat: if dests is empty check old dest_rows approach
            # Provide backward-compat fields dest_name/dest_key from first dest
            first = dests[0] if dests else {"name": "", "key": "", "vol_partial": None}
            current = {
                "route":     str(route_val).strip(),
                "dests":     dests,
                # Legacy aliases kept for compatibility with existing code:
                "dest_name": first["name"],
                "dest_key":  first["key"],
                "rows":      [],
            }
            blocks.append(current)
            continue
        if current is not None and isinstance(val0, str) and IRMA_RE.match(val0.strip()):
            # Named per-farm fields the rest of the code knows about. Anything
            # else non-empty on this row is captured into _extra_cells so it
            # travels with the farm when the solver moves it to a different row.
            # Without this, columns like farm name (R) and street address (AM)
            # would stay put while the IRMA in column A moved, leaving the
            # exported sheet showing farm A's name next to farm B's IRMA.
            NAMED_COLS = {C_IRMA, C_TRAIN, C_M1_START, C_M1_FINISH,
                          C_M2_START, C_M2_FINISH, C_EDPU, C_LOCATION,
                          C_PRIOR_VOL, C_ROUTE}
            extras = {}
            for c in range(1, ws.max_column + 1):
                if c in NAMED_COLS:
                    continue
                cv = ws.cell(r, c).value
                if cv is None:
                    continue
                if isinstance(cv, str) and cv.strip() == "":
                    continue
                extras[c] = cv
            current["rows"].append({
                "_uid":         str(uuid.uuid4()),
                "irma":         val0.strip(),
                "train":        ws.cell(r, C_TRAIN).value or "",
                "m1_start":     fmt_time(ws.cell(r, C_M1_START).value),
                "m1_finish":    fmt_time(ws.cell(r, C_M1_FINISH).value),
                "m2_start":     fmt_time(ws.cell(r, C_M2_START).value),
                "m2_finish":    fmt_time(ws.cell(r, C_M2_FINISH).value),
                "edpu":         ws.cell(r, C_EDPU).value or "",
                "location":     ws.cell(r, C_LOCATION).value or "",
                "prior_vol":    _cell(r, C_PRIOR_VOL),
                "_extra_cells": extras,
            })
    # Mark any remaining zero-farm blocks with a dest as preload
    # (handles edge case where IRMA# exists but no farms follow before next block)
    for block in blocks:
        if not block["rows"] and block.get("dests") and not block.get("preload"):
            block["preload"] = True

    return blocks, driver_start, day_colour

# -- Table item factories ------------------------------------------------------

def make_header_item(text, bg=CLR_HEADER, fg=CLR_HEADER_FG, bold=True, draggable=False):
    item = QTableWidgetItem(str(text))
    item.setBackground(bg); item.setForeground(fg)
    if bold:
        f = item.font(); f.setBold(True); item.setFont(f)
    flags = Qt.ItemIsEnabled
    if draggable:
        flags |= Qt.ItemIsDragEnabled | Qt.ItemIsSelectable
    item.setFlags(flags)
    item.setTextAlignment(Qt.AlignCenter)
    return item

def make_data_item(text, bg=CLR_WHITE, align=Qt.AlignCenter, fg=None, draggable=False):
    item = QTableWidgetItem(str(text) if text is not None else "")
    item.setBackground(bg)
    flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
    if draggable: flags |= Qt.ItemIsDragEnabled
    item.setFlags(flags)
    item.setTextAlignment(align)
    if fg: item.setForeground(fg)
    return item

# -- Route table renderer ------------------------------------------------------

def populate_table(table, blocks, dm, editable=False, start_time=None, dm_dur=None,
                   suppress_no_milking=True, plant_windows=None):
    table.clearSpans()
    table.clear()
    table.setColumnCount(len(COLS))
    table.setHorizontalHeaderLabels([c[0] for c in COLS])

    if not blocks:
        table.setRowCount(0)
        return

    all_dists  = calc_distances(blocks, dm)
    _ct = calc_times(blocks, dm, start_time, dm_dur=dm_dur, suppress_no_milking=suppress_no_milking)   # may be None
    all_times, _end_cursor = _ct if _ct is not None else (None, None)
    n_blocks   = len(blocks)
    # per block: banner + col_hdr + origin + farms + dest + subtotal
    # last block gets one extra row for VEDDER return
    # per block: banner(1) + col_hdr(1) + origin(1) + farms + dests + subtotal(1)
    # last block: +1 for VEDDER return
    def _ndests(b):
        d = b.get("dests") or []
        return max(len(d), 1)  # always at least one dest row
    total_rows = sum(2 + 1 + len(b["rows"]) + _ndests(b) + 1 for b in blocks) + 2
    if n_blocks > 0: total_rows += 1  # VEDDER return on last block
    table.setRowCount(total_rows)

    r = 0; day_dist = 0.0; day_vol = 0.0; day_ok = True

    for b_idx, block in enumerate(blocks):
        dists     = all_dists[b_idx]
        farms     = block["rows"]
        dests     = block.get("dests") or []
        if not dests:
            dk = block.get("dest_key",""); dn = block.get("dest_name","") or dk or "Destination"
            dests = [{"name": dn, "key": dk, "vol_partial": None}] if dk else []
        # legacy compat
        dest_key  = dests[0]["key"]  if dests else ""
        dest_name = dests[0]["name"] if dests else "Destination"

        route_dist = 0.0; route_ok = True
        for d in dists[:-1]:
            if d is None: route_ok = False
            else: route_dist += d

        route_vol = sum((row["prior_vol"] or 0) for row in farms
                        if isinstance(row.get("prior_vol"), (int, float)))

        # Banner
        table.setSpan(r, 0, 1, len(COLS))
        banner_item = make_header_item(f"  Route: {block['route']}",
                                       bg=CLR_ROUTE_HDR, fg=CLR_ROUTE_FG,
                                       draggable=editable)
        banner_item.setData(Qt.UserRole + 2, b_idx)
        table.setItem(r, 0, banner_item)
        r += 1

        # Column sub-headers
        for c_idx, (hdr, _) in enumerate(COLS):
            table.setItem(r, c_idx, make_header_item(hdr))
        r += 1

        # Origin row (VEDDER for first block; previous processor for subsequent)
        if b_idx == 0:
            origin_key = "VEDDER"; origin_sub = "Depot"
        else:
            origin_key = _block_last_dest_key(blocks[b_idx-1]) or "VEDDER"
            prev_dests = blocks[b_idx-1].get("dests") or []
            origin_sub = (prev_dests[-1]["name"] if prev_dests else
                          blocks[b_idx-1].get("dest_name","")) or origin_key
        od = dists[0]
        od_str = f"{od:.1f}" if od is not None else "-"
        # Time entries for this block: [origin, farm0, farm1, ..., dest]
        btimes = all_times[b_idx] if all_times else None
        origin_dep_str = fmt_hhmm(btimes[0]["dep"]) if btimes else "-"
        for c_idx, (_, key) in enumerate(COLS):
            if key == "irma":       item = make_data_item(origin_key, bg=CLR_DEPOT)
            elif key == "location": item = make_data_item(origin_sub, bg=CLR_DEPOT)
            elif key == "dist":     item = make_data_item(od_str, bg=CLR_DEPOT,
                                                          align=Qt.AlignRight|Qt.AlignVCenter)
            elif key == "dep_time": item = make_data_item(origin_dep_str, bg=CLR_DEPOT,
                                                          align=Qt.AlignCenter)
            else:                   item = make_data_item("", bg=CLR_DEPOT)
            table.setItem(r, c_idx, item)
        r += 1

        # -- Stop rows in actual visit order ------------------------------
        # Walk _build_block_stops to interleave farms and mid-route dest dropoffs
        # in the order the truck actually visits them.
        is_last_block = (b_idx == len(blocks) - 1)
        origin_key_for_stops = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        block_stops = _build_block_stops(block, origin_key_for_stops, is_last_block)

        # For preload blocks (no farms) use PRELOAD_VOL as the notional total
        PRELOAD_VOL_DISPLAY = 40000
        total_farm_vol = (PRELOAD_VOL_DISPLAY if block.get("preload") and not farms
                          else route_vol)
        already_del    = 0.0
        last_dest_dep_s = "-"   # for subtotal end-time
        farm_alt_idx   = 0       # for alternating row colors

        # Map dest dict id -> position in block.dests list (for d_i lookups)
        dests_by_id = {id(d): i for i, d in enumerate(dests)}
        # Track farm position in block.rows for f_i lookups
        farms_by_id = {id(f): i for i, f in enumerate(farms)}

        # Build list of (dest_d, partial_offload) by walking dests in stop order to
        # properly compute "remaining" and split tagging
        for stop in block_stops:
            si = stop["_si"]
            stype = stop["type"]
            # Skip origin and vedder ? handled separately above and below
            if stype in ("origin", "vedder"):
                continue

            ft = btimes[si] if btimes and si < len(btimes) else None
            arr_t = ft["arr"] if ft else None
            dep_t = ft["dep"] if ft else None
            arr_s = fmt_hhmm(arr_t)
            dep_s = fmt_hhmm(dep_t)
            # Distance leg from THIS stop to the next stop
            leg = dists[si] if si < len(dists) else None
            ds  = f"{leg:.1f}" if leg is not None else "-"

            if stype == "farm":
                row_data = stop["farm"]
                f_i      = farms_by_id.get(id(row_data), 0)
                bg       = CLR_ALT if farm_alt_idx % 2 == 0 else CLR_WHITE
                farm_alt_idx += 1
                milking_conflict = arrives_during_milking(arr_t, row_data, suppress_no_milking=suppress_no_milking)
                for c_idx, (_, key) in enumerate(COLS):
                    if key == "dist":
                        item = make_data_item(ds, bg=bg, align=Qt.AlignRight|Qt.AlignVCenter,
                                             draggable=editable)
                    elif key == "prior_vol":
                        v = row_data.get(key)
                        item = make_data_item(f"{int(v):,}" if isinstance(v,(int,float)) else "",
                                             bg=bg, align=Qt.AlignRight|Qt.AlignVCenter,
                                             draggable=editable)
                    elif key == "arr_time":
                        item = make_data_item(arr_s,
                                             bg=CLR_RED_BG if milking_conflict else bg,
                                             fg=CLR_RED    if milking_conflict else None,
                                             align=Qt.AlignCenter, draggable=editable)
                    elif key == "wait_time":
                        wait = ft["wait"] if ft else None
                        if wait and wait > 0:
                            wm = int(round(wait))
                            wait_s = f"{wm}m" if wm < 60 else f"{wm//60}h{wm%60:02d}m"
                            item = make_data_item(wait_s, bg=QColor("#fff3e0"),
                                                 fg=QColor("#e65100"),
                                                 align=Qt.AlignCenter, draggable=editable)
                        else:
                            item = make_data_item("", bg=bg, align=Qt.AlignCenter, draggable=editable)
                    elif key == "dep_time":
                        item = make_data_item(dep_s, bg=bg, align=Qt.AlignCenter, draggable=editable)
                    elif key == "_mwo":
                        checked = bool(row_data.get("_mwo", False))
                        item = QTableWidgetItem()
                        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                        item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
                        item.setBackground(QColor("#bbdefb") if checked else bg)
                        if editable:
                            item.setToolTip(
                                "Check to give this farm a 2-hour pickup window\n"
                                "starting from its current arrival time in this schedule.\n"
                                "Reflects across both Original and Modified views.")
                        item.setData(Qt.UserRole, (b_idx, f_i))
                    else:
                        # For farm rows, show the farm name (from _extra_cells
                        # col R=18) in the Location column instead of the raw
                        # location string.  Processors keep their name as-is.
                        if key == "location":
                            farm_name = (row_data.get("_extra_cells") or {}).get(18, "")
                            display_val = farm_name if farm_name else row_data.get(key, "")
                        else:
                            display_val = row_data.get(key, "")
                        item = make_data_item(display_val, bg=bg, draggable=editable)
                    item.setData(Qt.UserRole, (b_idx, f_i))
                    table.setItem(r, c_idx, item)
                r += 1

            elif stype == "dest":
                dest_d = stop["dest"]
                d_i    = dests_by_id.get(id(dest_d), 0)
                dest_arr_s   = arr_s
                dest_dep_s_d = dep_s
                last_dest_dep_s = dest_dep_s_d
                # partial vol for this dest
                vp = dest_d.get("vol_partial")
                remaining = max(0.0, total_farm_vol - already_del)
                offload = min(float(vp), remaining) if vp is not None else remaining
                already_del += offload
                vol_s = f"{int(offload):,}" if offload else ""
                # Detect if this is a mid-route partial dropoff (truck continues after)
                is_yard_dest = "yard for" in (dest_d.get("name","") or "").lower()
                # A dest is a "mid-route partial" if it has split_after set (not None)
                # AND there are more stops of type 'farm' after it in block_stops
                is_partial_split = (dest_d.get("split_after") is not None)

                # Check if arrival is outside the plant's receiving window
                dest_win_conflict = False
                if plant_windows and not is_yard_dest and ft and ft.get("arr"):
                    dk_chk = normalise_key(dest_d.get("key","") or "")
                    win_chk = plant_windows.get(dk_chk)
                    if win_chk:
                        dest_win_conflict = not time_in_window(ft["arr"], win_chk[0], win_chk[1])

                # Use a distinct background for partial mid-route dropoffs so they're visually obvious
                if is_partial_split:
                    dest_bg = QColor("#fff8e1")   # soft amber - "partial dropoff"
                elif dest_win_conflict:
                    dest_bg = CLR_DEST_WARN
                else:
                    dest_bg = CLR_DEST

                # Compose location string with PARTIAL tag where appropriate
                loc_str = dest_d.get("name","") or dest_d.get("key","")
                if is_partial_split:
                    loc_str = f"v PARTIAL ? {loc_str}"

                for c_idx, (_, key) in enumerate(COLS):
                    if key == "irma":
                        item = make_data_item("", bg=dest_bg, draggable=editable)
                    elif key == "proc_id":
                        item = make_data_item(dest_d.get("key",""), bg=dest_bg,
                                              align=Qt.AlignCenter, draggable=editable)
                    elif key == "location":
                        item = make_data_item(loc_str, bg=dest_bg, draggable=editable)
                    elif key == "prior_vol":
                        item = make_data_item(vol_s, bg=dest_bg,
                                              align=Qt.AlignRight|Qt.AlignVCenter, draggable=editable)
                    elif key == "dist":
                        item = make_data_item("0.0" if is_yard_dest else ds,
                                              bg=dest_bg, align=Qt.AlignRight|Qt.AlignVCenter)
                    elif key == "arr_time":
                        item = make_data_item(dest_arr_s, bg=dest_bg, align=Qt.AlignCenter)
                    elif key == "dep_time":
                        item = make_data_item(dest_dep_s_d, bg=dest_bg, align=Qt.AlignCenter)
                    else:
                        item = make_data_item("", bg=dest_bg)
                    # tag dest rows so drag resolver knows they're dest rows
                    item.setData(Qt.UserRole+1, ("dest", b_idx, d_i))
                    table.setItem(r, c_idx, item)
                r += 1

        # No-processor warning row if block has no dests
        if not dests:
            CLR_WARN = QColor("#fff3cd")
            for c_idx, (_, key) in enumerate(COLS):
                if c_idx == 0:
                    item = make_header_item("(!) No processor assigned",
                                            bg=CLR_WARN, fg=QColor("#856404"))
                else:
                    item = make_data_item("", bg=CLR_WARN)
                    item.setFlags(Qt.ItemIsEnabled)
                table.setItem(r, c_idx, item)
            r += 1

        # VEDDER return row ? only on the last block
        if is_last_block:
            vedder_stops = [s for s in block_stops if s["type"] == "vedder"]
            vedder_idx   = vedder_stops[0]["_si"] if vedder_stops else (len(farms) + len(dests) + 1)
            vedder_ret_t = btimes[vedder_idx] if btimes and vedder_idx < len(btimes) else None
            vedder_arr_s = fmt_hhmm(vedder_ret_t["arr"]) if vedder_ret_t else "-"
            for c_idx, (_, key) in enumerate(COLS):
                if key == "irma":       item = make_data_item("VEDDER", bg=CLR_DEPOT)
                elif key == "location": item = make_data_item("Depot", bg=CLR_DEPOT)
                elif key == "dist":     item = make_data_item("-", bg=CLR_DEPOT,
                                                              align=Qt.AlignRight|Qt.AlignVCenter)
                elif key == "arr_time": item = make_data_item(vedder_arr_s, bg=CLR_DEPOT,
                                                              align=Qt.AlignCenter)
                else:                   item = make_data_item("", bg=CLR_DEPOT)
                table.setItem(r, c_idx, item)
            r += 1

        # Route subtotal
        vol_over = route_vol > VOL_LIMIT
        ds2 = f"{route_dist:.1f} km" if route_ok else f"~{route_dist:.1f} km*"
        end_s = last_dest_dep_s
        for c_idx, (_, key) in enumerate(COLS):
            if c_idx == 0:
                item = make_header_item("Subtotal",
                                        bg=CLR_SUBTOTAL, fg=QColor("#000000"))
            elif key == "dist":
                item = make_header_item(ds2, bg=CLR_SUBTOTAL, fg=QColor("#000000"))
            elif key == "prior_vol":
                item = make_header_item(f"{int(route_vol):,} L",
                                        bg=CLR_RED_BG if vol_over else CLR_SUBTOTAL,
                                        fg=CLR_RED    if vol_over else QColor("#000000"))
            elif key == "dep_time":
                item = make_header_item(f"End: {end_s}", bg=CLR_SUBTOTAL, fg=QColor("#000000"))
            else:
                item = make_data_item("", bg=CLR_SUBTOTAL)
                item.setFlags(Qt.ItemIsEnabled)
            table.setItem(r, c_idx, item)
        r += 1

        day_dist += route_dist; day_vol += route_vol
        if not route_ok: day_ok = False

    # Day total ? no red colouring here
    ds3 = f"{day_dist:.1f} km" if day_ok else f"~{day_dist:.1f} km*"
    # Shift end = arrival back at VEDDER (last entry in last block's times)
    shift_end = fmt_hhmm(_end_cursor.time()) if _end_cursor is not None else "-"
    for c_idx, (_, key) in enumerate(COLS):
        if c_idx == 0:
            item = make_header_item("Total", bg=CLR_TOTAL, fg=QColor("#000000"))
        elif key == "dist":
            item = make_header_item(ds3, bg=CLR_TOTAL, fg=QColor("#000000"))
        elif key == "prior_vol":
            item = make_header_item(f"{int(day_vol):,} L", bg=CLR_TOTAL, fg=QColor("#000000"))
        elif key == "dep_time":
            item = make_header_item(f"Shift end: {shift_end}", bg=CLR_TOTAL, fg=QColor("#000000"))
        else:
            item = make_data_item("", bg=CLR_TOTAL)
            item.setFlags(Qt.ItemIsEnabled)
        table.setItem(r, c_idx, item)
    r += 1

    # Shift length row
    shift_len_str = "-"
    if start_time and _end_cursor is not None:
        from datetime import datetime, date
        base  = datetime.combine(date.today(), start_time)
        delta = _end_cursor - base
        total_mins = int(delta.total_seconds() / 60)
        hours, mins = divmod(total_mins, 60)
        shift_len_str = f"{hours}h {mins:02d}m"
    CLR_SHIFT = QColor("#e8eaf6")
    table.setSpan(r, 0, 1, len(COLS))
    lbl = f"  Estimated Shift Length:  {fmt_hhmm(start_time)} -> {shift_end}  =  {shift_len_str}"
    item = make_header_item(lbl, bg=CLR_SHIFT, fg=QColor("#1a237e"))
    table.setItem(r, 0, item)

    table.resizeColumnsToContents()
    hh = table.horizontalHeader()
    for c_idx, (_, key) in enumerate(COLS):
        if key == "irma":
            hh.setSectionResizeMode(c_idx, QHeaderView.Fixed)
            table.setColumnWidth(c_idx, 76)
        elif key == "dist":
            hh.setSectionResizeMode(c_idx, QHeaderView.Fixed)
            table.setColumnWidth(c_idx, 88)
        elif key == "_mwo":
            hh.setSectionResizeMode(c_idx, QHeaderView.Fixed)
            table.setColumnWidth(c_idx, 44)

# -- File loader thread --------------------------------------------------------

def _try_eval_formula(formula_str):
    """
    Attempt to evaluate a simple Excel formula string to a float.
    Handles bare numbers, simple arithmetic (+,-,*,/,()), and
    SUM() / sum() with only numeric literals.
    Returns a float, or None if the formula is too complex
    (contains cell references or unsupported functions).
    """
    if not isinstance(formula_str, str):
        return None
    s = formula_str.strip()
    if s.startswith("="):
        s = s[1:].strip()
    # Strip a leading SUM() wrapper with no cell refs
    import re as _re
    sum_m = _re.match(r'^[Ss][Uu][Mm]\(([^)]+)\)$', s)
    if sum_m:
        s = sum_m.group(1).replace(",", "+")
    # Reject anything that still looks like a cell reference (letters followed by digits)
    if _re.search(r'[A-Za-z]', s):
        return None
    # Only allow digits, spaces, arithmetic ops, dots, parens
    if not _re.match(r'^[\d\s\+\-\*\/\.\(\)]+$', s):
        return None
    try:
        result = eval(s, {"__builtins__": {}}, {})  # no builtins ? safe for pure arithmetic
        return float(result)
    except Exception:
        return None


class FileLoader(QThread):
    done   = pyqtSignal(str, dict)
    failed = pyqtSignal(str, str)
    # Non-fatal per-sheet parse warnings ? accumulated and shown in a dialog
    # after load completes, and also written to the debug log.
    sheet_warning = pyqtSignal(str, str, str)
    # Informational/timing log messages ? written to the debug log only,
    # never shown in the warning dialog.
    log = pyqtSignal(str)

    def __init__(self, fname, fpath):
        super().__init__()
        self.fname = fname; self.fpath = fpath

    def run(self):
        try:
            import time as _time
            _t0 = _time.time()
            self.log.emit(f"[{self.fname}] Opening workbook (data pass)...")
            wb_data = openpyxl.load_workbook(self.fpath, read_only=False, data_only=True)
            self.log.emit(
                f"[{self.fname}] Data workbook opened in {_time.time()-_t0:.1f}s, "
                f"opening formula pass...")
            _t1 = _time.time()
            wb_form = openpyxl.load_workbook(self.fpath, read_only=False, data_only=False)
            self.log.emit(
                f"[{self.fname}] Formula workbook opened in {_time.time()-_t1:.1f}s")
            sheets = {}
            for n in wb_data.sheetnames:
                if n.strip().upper() in EXCLUDE_SHEETS:
                    continue
                try:
                    _ts = _time.time()
                    self.log.emit(f"[{self.fname} / {n}] Parsing...")
                    blocks, start_time, day_colour = parse_sheet(
                        wb_data[n], ws_formula=wb_form[n])
                    self.log.emit(
                        f"[{self.fname} / {n}] Done in {_time.time()-_ts:.1f}s - "
                        f"{len(blocks)} block(s), start={start_time}, colour={day_colour!r}")

                    # -- Post-parse diagnostics --------------------------------
                    # Each check emits a specific warning so the user knows
                    # *why* a sheet looks empty or wrong, rather than just
                    # seeing a blank tab.
                    warnings = []

                    if not blocks:
                        warnings.append(
                            "No blocks found ? no 'IRMA#' header rows were "
                            "detected.  Check that the sheet follows the "
                            "expected layout (IRMA# in column A).")

                    if not start_time:
                        warnings.append(
                            f"No driver start time found in cell "
                            f"{openpyxl.utils.get_column_letter(C_DRIVER_START)}1"
                            f" (column {C_DRIVER_START}).  Times will not be "
                            f"calculated for this sheet.")

                    if not day_colour:
                        warnings.append(
                            f"No day colour found in cell "
                            f"{openpyxl.utils.get_column_letter(C_DAY_COLOUR)}1"
                            f" (column {C_DAY_COLOUR}).  Expected RED, BLUE, "
                            f"or GRASSFED.  The solver may skip this sheet.")

                    # Check for formula cells whose cached value is None ?
                    # the most common cause of 'looks fine in Excel, missing
                    # data in the viewer'.  Report the first few occurrences.
                    formula_nones = []
                    for block in blocks:
                        for row in block.get("rows", []):
                            if row.get("prior_vol") is None:
                                formula_nones.append(row.get("irma", "?"))
                    if formula_nones:
                        sample = ", ".join(formula_nones[:5])
                        extra  = f" (and {len(formula_nones)-5} more)" \
                                 if len(formula_nones) > 5 else ""
                        warnings.append(
                            f"prior_vol is None for farm(s): {sample}{extra}.  "
                            f"These cells may contain uncalculated formulas - "
                            f"open the file in Excel, press Ctrl+Alt+F9 to "
                            f"force-recalculate, then save and reload.")

                    # Check for blocks with no destination
                    no_dest = [str(i+1) for i, b in enumerate(blocks)
                               if not b.get("dests") and not b.get("dest_key")
                               and not b.get("preload")]
                    if no_dest:
                        warnings.append(
                            f"Block(s) {', '.join(no_dest)} have no destination.  "
                            f"The 'Delivery Information:' section may be missing "
                            f"or in an unexpected column.")

                    for w in warnings:
                        logger.warning("[%s / %s] %s", self.fname, n, w)
                        self.sheet_warning.emit(self.fname, n, w)

                    if blocks or start_time:
                        sheets[n] = {"blocks": blocks, "start_time": start_time,
                                     "day_colour": day_colour}
                    else:
                        # Emit a warning so it's visible even when no blocks
                        # or start_time were found (sheet would be silently
                        # dropped from the cache without this).
                        if not warnings:
                            msg = ("Sheet produced no usable data and no "
                                   "specific warnings ? the layout may not "
                                   "match the expected format.")
                            logger.warning("[%s / %s] %s", self.fname, n, msg)
                            self.sheet_warning.emit(self.fname, n, msg)

                except Exception as sheet_err:
                    # Per-sheet failure: log it and emit a warning rather than
                    # aborting the whole file load.  The sheet is skipped but
                    # other sheets in the same workbook still load.
                    import traceback
                    detail = traceback.format_exc()
                    msg = (f"Failed to parse sheet '{n}': {sheet_err}\n{detail}")
                    logger.error("[%s / %s] %s", self.fname, n, msg)
                    self.sheet_warning.emit(self.fname, n,
                        f"Parse error ? {sheet_err}  "
                        f"(see console / log for full traceback)")

            wb_data.close(); wb_form.close()
            self.done.emit(self.fname, sheets)
        except Exception as e:
            self.failed.emit(self.fname, str(e))

# -- Drag-aware route table ----------------------------------------------------

MIME_FARM      = "application/x-farm"
MIME_TRAY_FARM = "application/x-tray-farm"
MIME_DEST      = "application/x-dest"
MIME_TRAY_DEST = "application/x-tray-dest"
MIME_BLOCK     = "application/x-block"     # drag entire block by its banner row

class EditableRouteTable(QTableWidget):
    """Editable route table: drag farms OUT (to tray) or accept drags IN (from tray).
    Also supports internal row reordering via drag."""
    farm_removed  = pyqtSignal(int, int)        # b_idx, f_idx  ? removed to tray
    farm_inserted = pyqtSignal(int, int, int)   # tray_idx, b_idx, insert_before_f_idx
    farm_reorder  = pyqtSignal(int, int, int, int)  # src_b, src_f, dst_b, dst_f
    dest_removed  = pyqtSignal(int, int)        # b_idx, d_idx  ? dest removed to tray
    dest_reorder  = pyqtSignal(int, int, int, int)  # src_b, src_d, dst_b, dst_d
    dest_inserted = pyqtSignal(int, int, int)   # tray_idx, b_idx, insert_before_d_idx
    block_reorder = pyqtSignal(int, int)        # src_b_idx, insert_before_b_idx

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.setShowGrid(True)
        self.setDropIndicatorShown(False)   # we draw our own
        self._drop_indicator_row = -1       # visual row to draw indicator above

    def dragMoveEvent(self, e):
        m = e.mimeData()
        if (m.hasFormat(MIME_FARM) or m.hasFormat(MIME_TRAY_FARM) or
                m.hasFormat(MIME_DEST) or m.hasFormat(MIME_TRAY_DEST) or
                m.hasFormat(MIME_BLOCK)):
            row = self.rowAt(e.pos().y())
            if m.hasFormat(MIME_BLOCK):
                # Snap indicator to the nearest banner boundary
                target_b = self._resolve_block_drop(row)
                if target_b >= 0:
                    # Find the visual row of that banner
                    snap_row = row  # fallback
                    for ri in range(self.rowCount()):
                        it = self.item(ri, 0)
                        if it is not None and it.data(Qt.UserRole + 2) == target_b:
                            snap_row = ri
                            break
                    self._drop_indicator_row = snap_row
                else:
                    # Append at very end
                    self._drop_indicator_row = self.rowCount()
            else:
                self._drop_indicator_row = row
            self.viewport().update()
            e.acceptProposedAction()
        else:
            e.ignore()

    def dragLeaveEvent(self, e):
        self._drop_indicator_row = -1
        self.viewport().update()
        super().dragLeaveEvent(e)

    def paintEvent(self, e):
        super().paintEvent(e)
        if self._drop_indicator_row < 0: return
        painter = QPainter(self.viewport())
        pen = QPen(QColor("#000000"), 2)
        painter.setPen(pen)
        row = self._drop_indicator_row
        if row < self.rowCount():
            rect = self.visualRect(self.model().index(row, 0))
            y = rect.top()
        else:
            rect = self.visualRect(self.model().index(self.rowCount()-1, 0))
            y = rect.bottom()
        painter.drawLine(0, y, self.viewport().width(), y)
        # Draw small arrow indicators at both ends
        size = 6
        for x in [0, self.viewport().width() - size]:
            painter.setBrush(QColor("#000000"))
            painter.setPen(Qt.NoPen)
            from PyQt5.QtGui import QPolygon
            from PyQt5.QtCore import QPoint
            tri = QPolygon([QPoint(x, y-size//2), QPoint(x+size, y), QPoint(x, y+size//2)])
            painter.drawPolygon(tri)
        painter.end()

    def startDrag(self, actions):
        item = self.currentItem()
        if item is None: return

        # Block banner drag ? UserRole+2 holds b_idx
        block_idx = item.data(Qt.UserRole + 2)
        if block_idx is not None:
            mime = QMimeData()
            mime.setData(MIME_BLOCK, QByteArray(str(block_idx).encode()))
            drag = QDrag(self)
            drag.setMimeData(mime)
            drag.exec_(Qt.MoveAction)
            return

        # Dest drag ? UserRole+1 holds ("dest", b_idx, d_idx)
        dest_data = item.data(Qt.UserRole + 1)
        if dest_data is not None and dest_data[0] == "dest":
            _, b_idx, d_idx = dest_data
            mime = QMimeData()
            mime.setData(MIME_DEST, QByteArray(f"{b_idx},{d_idx}".encode()))
            drag = QDrag(self)
            drag.setMimeData(mime)
            drag.exec_(Qt.MoveAction)
            return

        # Farm drag ? UserRole holds (b_idx, f_idx)
        farm_data = item.data(Qt.UserRole)
        if farm_data is None: return
        b_idx, f_idx = farm_data
        mime = QMimeData()
        mime.setData(MIME_FARM, QByteArray(f"{b_idx},{f_idx}".encode()))
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec_(Qt.MoveAction)

    def dragEnterEvent(self, e):
        m = e.mimeData()
        if (m.hasFormat(MIME_FARM) or m.hasFormat(MIME_TRAY_FARM) or
                m.hasFormat(MIME_DEST) or m.hasFormat(MIME_TRAY_DEST) or
                m.hasFormat(MIME_BLOCK)):
            e.acceptProposedAction()
        else:
            e.ignore()

    def dropEvent(self, e):
        self._drop_indicator_row = -1
        self.viewport().update()
        mime = e.mimeData()
        row_at_drop = self.rowAt(e.pos().y())

        if mime.hasFormat(MIME_BLOCK):
            src_b = int(bytes(mime.data(MIME_BLOCK)).decode())
            dst_b = self._resolve_block_drop(row_at_drop)
            if dst_b != src_b:
                self.block_reorder.emit(src_b, dst_b)
            e.acceptProposedAction()
            return

        if mime.hasFormat(MIME_DEST):
            raw = bytes(mime.data(MIME_DEST)).decode()
            src_b, src_d = map(int, raw.split(","))
            target_b, target_d = self._resolve_dest_drop(row_at_drop)
            self.dest_reorder.emit(src_b, src_d, target_b, target_d)
            e.acceptProposedAction()
            return

        if mime.hasFormat(MIME_TRAY_DEST):
            raw = bytes(mime.data(MIME_TRAY_DEST)).decode()
            tray_idx = int(raw)
            target_b, target_d = self._resolve_dest_drop(row_at_drop)
            self.dest_inserted.emit(tray_idx, target_b, target_d)
            e.acceptProposedAction()
            return

        # Resolve what block/farm-position the drop landed in
        target_b_idx, target_f_idx = self._resolve_drop_target(row_at_drop)

        if mime.hasFormat(MIME_FARM):
            raw = bytes(mime.data(MIME_FARM)).decode()
            src_b, src_f = map(int, raw.split(","))
            self.farm_reorder.emit(src_b, src_f, target_b_idx, target_f_idx)
            e.acceptProposedAction()

        elif mime.hasFormat(MIME_TRAY_FARM):
            raw = bytes(mime.data(MIME_TRAY_FARM)).decode()
            tray_idx = int(raw)
            self.farm_inserted.emit(tray_idx, target_b_idx, target_f_idx)
            e.acceptProposedAction()

    def _resolve_drop_target(self, row_at_drop):
        """Return (b_idx, insert_before_f_idx) for the given visual row.

        When a block has no farms every row in it is a header/origin/dest row
        and carries no Qt.UserRole farm data.  We detect the enclosing block by
        scanning backward for the nearest banner row (UserRole+2) so that drops
        on empty blocks are correctly attributed to that block at position 0.
        """
        if row_at_drop < 0:
            if self.rowCount() == 0:
                return (0, 0)
            # Dropped below all rows ? find the last block
            for r in range(self.rowCount() - 1, -1, -1):
                it = self.item(r, 0)
                if it is not None:
                    bd = it.data(Qt.UserRole + 2)
                    if bd is not None:
                        return (bd, 0)
                    fd = it.data(Qt.UserRole)
                    if fd is not None:
                        return (fd[0], fd[1] + 1)
            return (0, 0)

        item = self.item(row_at_drop, 0)
        if item is None:
            return (0, 0)

        # Landed directly on a farm row
        data = item.data(Qt.UserRole)
        if data is not None:
            return (data[0], data[1])

        # Dropped onto a dest (processor) row ? insert AFTER the last farm
        # in that block, i.e. immediately before the processor.
        dest_tag = item.data(Qt.UserRole + 1)
        if dest_tag is not None and isinstance(dest_tag, tuple) and dest_tag[0] == "dest":
            b_idx = dest_tag[1]
            # Scan backward for the last farm row in this block
            for rb in range(row_at_drop - 1, -1, -1):
                itb = self.item(rb, 0)
                if itb is None:
                    continue
                fdb = itb.data(Qt.UserRole)
                if fdb is not None and fdb[0] == b_idx:
                    return (b_idx, fdb[1] + 1)   # after last farm
                bdb = itb.data(Qt.UserRole + 2)
                if bdb is not None:
                    return (b_idx, 0)             # empty block, insert at 0
            return (b_idx, 0)

        # Scan forward for the next farm row in this block
        for r in range(row_at_drop, self.rowCount()):
            it = self.item(r, 0)
            if it is None:
                continue
            fd = it.data(Qt.UserRole)
            if fd is not None:
                return (fd[0], fd[1])
            # Hit the next block's banner ? stop; drop is at start of that block
            bd = it.data(Qt.UserRole + 2)
            if bd is not None and r > row_at_drop:
                # Insert at position 0 of whichever block we're currently in.
                # Find that block by scanning back for nearest banner.
                for rb in range(row_at_drop, -1, -1):
                    itb = self.item(rb, 0)
                    if itb is not None:
                        bdb = itb.data(Qt.UserRole + 2)
                        if bdb is not None:
                            return (bdb, 0)
                return (bd, 0)

        # Nothing found forward ? scan backward for containing block banner
        for r in range(row_at_drop, -1, -1):
            it = self.item(r, 0)
            if it is None:
                continue
            fd = it.data(Qt.UserRole)
            if fd is not None:
                return (fd[0], fd[1] + 1)
            bd = it.data(Qt.UserRole + 2)
            if bd is not None:
                return (bd, 0)

        return (0, 0)

    def _resolve_dest_drop(self, row_at_drop):
        """Return (b_idx, insert_before_d_idx) for a dest drop.
        Falls back to the enclosing block (via banner) when no dest rows exist."""
        if row_at_drop < 0:
            return (0, 0)
        item = self.item(row_at_drop, 0)
        if item is None: return (0, 0)
        dd = item.data(Qt.UserRole + 1)
        if dd is not None and dd[0] == "dest":
            return (dd[1], dd[2])
        # Scan forward for a dest row
        for ri in range(row_at_drop, self.rowCount()):
            it = self.item(ri, 0)
            if it:
                dd2 = it.data(Qt.UserRole + 1)
                if dd2 and dd2[0] == "dest":
                    return (dd2[1], dd2[2])
                # Hit next block banner ? drop is in current block
                bd = it.data(Qt.UserRole + 2)
                if bd is not None and ri > row_at_drop:
                    break
        # Scan backward for nearest dest or banner
        for ri in range(row_at_drop, -1, -1):
            it = self.item(ri, 0)
            if it:
                dd2 = it.data(Qt.UserRole + 1)
                if dd2 and dd2[0] == "dest":
                    return (dd2[1], dd2[2] + 1)
                # Found the banner for this block ? append as first dest
                bd = it.data(Qt.UserRole + 2)
                if bd is not None:
                    return (bd, 0)
        return (0, 0)

    def _resolve_block_drop(self, row_at_drop):
        """Return the b_idx to insert the dragged block *before*.
        Scans for the nearest banner row (UserRole+2 set) to use as the target."""
        if row_at_drop < 0:
            return -1   # caller interprets as "append at end"
        # Scan forward for a banner row
        for ri in range(row_at_drop, self.rowCount()):
            it = self.item(ri, 0)
            if it is not None:
                bd = it.data(Qt.UserRole + 2)
                if bd is not None:
                    return bd
        # Nothing found forward ? append after last block
        return -1

# -- Farm tray (removed farms, table-based, drag back) ------------------------

class FarmTray(QTableWidget):
    """Holds removed farms. Farms can be dragged back to the editable route table."""
    farm_incoming = pyqtSignal(int, int)   # b_idx, f_idx ? farm dropped onto tray from route
    dest_incoming = pyqtSignal(int, int)   # b_idx, d_idx ? dest dropped onto tray from route

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self.setColumnCount(len(TRAY_COLS))
        self.setHorizontalHeaderLabels([c[0] for c in TRAY_COLS])
        self.setRowCount(0)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.setShowGrid(True)
        self.setAlternatingRowColors(True)

    def keyPressEvent(self, event):
        from PyQt5.QtCore import Qt as _Qt
        if event.key() in (_Qt.Key_Delete, _Qt.Key_Backspace):
            # Bubble up to MainWindow via parent chain
            parent = self.parent()
            while parent:
                if hasattr(parent, '_on_tray_delete'):
                    parent._on_tray_delete()
                    return
                parent = parent.parent()
        super().keyPressEvent(event)

    def startDrag(self, actions):
        row = self.currentRow()
        if row < 0: return
        # Check if this tray row is a dest or a farm
        item0 = self.item(row, 0)
        is_dest = item0 and item0.data(Qt.UserRole) == "dest"
        mime = QMimeData()
        mime_type = MIME_TRAY_DEST if is_dest else MIME_TRAY_FARM
        mime.setData(mime_type, QByteArray(str(row).encode()))
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec_(Qt.MoveAction)

    def dragEnterEvent(self, e):
        m = e.mimeData()
        if m.hasFormat(MIME_FARM) or m.hasFormat(MIME_DEST): e.acceptProposedAction()
        else: e.ignore()

    def dragMoveEvent(self, e):
        m = e.mimeData()
        if m.hasFormat(MIME_FARM) or m.hasFormat(MIME_DEST): e.acceptProposedAction()
        else: e.ignore()

    def dropEvent(self, e):
        m = e.mimeData()
        if m.hasFormat(MIME_FARM):
            raw = bytes(m.data(MIME_FARM)).decode()
            b_idx, f_idx = map(int, raw.split(","))
            self.farm_incoming.emit(b_idx, f_idx)
            e.acceptProposedAction()
        elif m.hasFormat(MIME_DEST):
            raw = bytes(m.data(MIME_DEST)).decode()
            b_idx, d_idx = map(int, raw.split(","))
            self.dest_incoming.emit(b_idx, d_idx)
            e.acceptProposedAction()

    def add_farm(self, farm_dict, route_label, farm_colour="", current_colour=""):
        """Add a farm row. farm_colour is the day type it was removed from.
        current_colour is the active sheet's day type ? if different, row is bold."""
        r = self.rowCount(); self.insertRow(r)
        mismatch = bool(farm_colour and current_colour and
                        farm_colour.upper().strip() != current_colour.upper().strip())
        bold_font = QFont(); bold_font.setBold(True)
        normal_font = QFont()
        row_font = bold_font if mismatch else normal_font

        for c_idx, (_, key) in enumerate(TRAY_COLS):
            if key == "_from_route":
                val = route_label
                bg  = CLR_REMOVED
                fg  = None
            elif key == "_day_colour":
                bg2, fg2, label = day_colour_style(farm_colour)
                val = label or farm_colour or "-"
                bg  = bg2 if bg2 else CLR_REMOVED
                fg  = fg2
            elif key == "prior_vol":
                v = farm_dict.get(key)
                val = f"{int(v):,}" if isinstance(v, (int,float)) else ""
                bg  = CLR_REMOVED; fg = None
            elif key == "dist":
                val = ""; bg = CLR_REMOVED; fg = None
            elif key == "proc_id":
                val = ""; bg = CLR_REMOVED; fg = None
            elif key == "_mwo":
                val = "OK" if farm_dict.get("_mwo") else ""
                bg = CLR_REMOVED; fg = None
            else:
                if key == "location":
                    farm_name = (farm_dict.get("_extra_cells") or {}).get(18, "")
                    val = farm_name if farm_name else farm_dict.get(key, "")
                else:
                    val = farm_dict.get(key, "")
                bg  = CLR_REMOVED; fg = None

            item = QTableWidgetItem(str(val) if val is not None else "")
            item.setBackground(bg)
            if fg: item.setForeground(fg)
            item.setFont(row_font)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsDragEnabled)
            align = Qt.AlignRight|Qt.AlignVCenter if key == "prior_vol" else Qt.AlignCenter
            item.setTextAlignment(align)
            self.setItem(r, c_idx, item)
        self.setColumnWidth(0, 68)

    def add_dest(self, dest_dict, route_label, farm_colour="", current_colour=""):
        """Add a destination row to the tray. Marked with UserRole='dest'."""
        r = self.rowCount(); self.insertRow(r)
        mismatch = bool(farm_colour and current_colour and
                        farm_colour.upper().strip() != current_colour.upper().strip())
        bold_font = QFont(); bold_font.setBold(True)
        normal_font = QFont()
        row_font = bold_font if mismatch else normal_font
        CLR_DEST_TRAY = QColor("#e8f5e9")  # greenish to distinguish from farm rows
        for c_idx, (_, key) in enumerate(TRAY_COLS):
            if key == "_from_route":
                val = route_label; bg = CLR_DEST_TRAY; fg = None
            elif key == "_day_colour":
                bg2, fg2, label = day_colour_style(farm_colour)
                val = label or farm_colour or "-"
                bg = bg2 if bg2 else CLR_DEST_TRAY; fg = fg2
            elif key == "irma":
                val = ""; bg = CLR_DEST_TRAY; fg = None
            elif key == "proc_id":
                val = dest_dict.get("key",""); bg = CLR_DEST_TRAY; fg = None
            elif key == "location":
                val = dest_dict.get("name","") or dest_dict.get("key","")
                bg = CLR_DEST_TRAY; fg = None
            elif key == "prior_vol":
                vp = dest_dict.get("vol_partial")
                val = f"{int(vp):,}" if isinstance(vp,(int,float)) else "rest"
                bg = CLR_DEST_TRAY; fg = None
            else:
                val = ""; bg = CLR_DEST_TRAY; fg = None
            item = QTableWidgetItem(str(val) if val is not None else "")
            item.setBackground(bg)
            if fg: item.setForeground(fg)
            item.setFont(row_font)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsDragEnabled)
            item.setTextAlignment(Qt.AlignCenter)
            # Tag first cell so we know it's a dest row
            if c_idx == 0: item.setData(Qt.UserRole, "dest")
            self.setItem(r, c_idx, item)

    def refresh_bold_state(self, current_colour):
        """Bold rows whose farm type differs from current_colour."""
        bold_font   = QFont(); bold_font.setBold(True)
        normal_font = QFont()
        type_col = next((i for i,(_, k) in enumerate(TRAY_COLS) if k == "_day_colour"), None)
        for r in range(self.rowCount()):
            farm_colour = ""
            if type_col is not None:
                it = self.item(r, type_col)
                if it: farm_colour = it.text()
            mismatch = bool(farm_colour and current_colour and
                            farm_colour.upper().strip() != current_colour.upper().strip())
            font = bold_font if mismatch else normal_font
            for c in range(self.columnCount()):
                it = self.item(r, c)
                if it: it.setFont(font)

# -- Synchronised scroll helper ------------------------------------------------

def sync_scroll(src_bar, dst_bar, value):
    dst_bar.setValue(value)

# ??????????????????????????????????????????????????????????????????????????????
# ALNS Solver
# ??????????????????????????????????????????????????????????????????????????????

def _sheet_colour_bucket(day_colour):
    dc = day_colour.upper().strip()
    if "RED"  in dc: return "RED"
    if "BLUE" in dc: return "BLUE"
    return "OTHER"


def _route_km_simple(block, dm, origin="VEDDER"):
    """Quick distance estimate: origin -> farms -> all dests (no VEDDER return).
    Pure km only ? used for 2-opt / or-opt intra-block resequencing where
    milking windows don't change between candidate orderings of the same farms."""
    farms     = [r["irma"] for r in block["rows"]]
    dest_keys = _block_dest_keys(block)
    stops     = [origin] + farms + dest_keys
    total     = 0.0
    for i in range(len(stops) - 1):
        d = lookup(dm, stops[i], stops[i + 1])
        if d is not None:
            total += d
    return total


def _route_cost_with_milking(block, dm, origin="VEDDER",
                              milking_weight=1.0, suppress_no_milking=True,
                              shift_start_mins=300):
    """Distance cost + milking-wait penalty for a block.

    Used by _best_insert_cost so the greedy repair heuristic avoids placing
    farms into slots that cause long waits, not just long drives.

    Milking wait estimation (fast, no full calc_times):
      Tracks an absolute datetime cursor starting at shift_start_mins past
      midnight on 'today'.  Each leg advances the cursor by drive time.
      If arrival falls inside a milking window (w1, w2, or w3), the wait
      until window-end is added as a penalty of wait_mins * milking_weight
      (in km-equivalent units, since 1 km-eq ~ 1.2 min at 50 km/h ? close
      enough for ranking purposes).  The cursor advances by the wait so
      downstream farms see a realistic arrival time, correctly handling
      midnight-crossing routes without modular wrap errors.

    shift_start_mins: minutes past midnight for departure from the block origin.
      When no start time is available the sheet is excluded from insertion
      entirely (see _best_insert_cost), so this value always comes from a
      real entry.start_time.
    """
    if milking_weight <= 0.0:
        return _route_km_simple(block, dm, origin=origin)

    from datetime import timedelta as _td, datetime as _dt, date as _d

    farms     = block["rows"]
    dest_keys = _block_dest_keys(block)
    stops     = [origin] + [r["irma"] for r in farms] + dest_keys

    total_km = 0.0
    wait_pen = 0.0
    # Absolute datetime cursor ? no modular wrap, handles overnight routes correctly
    cursor   = _dt.combine(_d.today(), _dt.min.time()) + _td(minutes=shift_start_mins)

    for i in range(len(stops) - 1):
        leg = lookup(dm, stops[i], stops[i + 1])
        if leg is not None:
            total_km += leg
            cursor   += _td(minutes=(leg / DRIVE_SPEED_KMH) * 60.0)

        # Check milking window for farm stops only (not origin or dest)
        if i < len(farms):
            farm = farms[i]
            irma = farm.get("irma", "")
            if suppress_no_milking and irma in NO_MILKING_WINDOW_FARMS:
                continue

            arr_t  = cursor.time()
            wait_m = 0.0

            # w1 / w2
            for s_key, f_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
                if time_in_window(arr_t, farm.get(s_key,""), farm.get(f_key,"")):
                    tf = parse_hhmm(farm.get(f_key,""))
                    if tf:
                        end_w = _dt.combine(cursor.date(), tf)
                        if end_w <= cursor:
                            end_w += _td(days=1)
                        wait_m = (end_w - cursor).total_seconds() / 60.0
                    break
            # w3
            if wait_m == 0.0:
                w3data = THREE_WINDOW_FARMS.get(irma)
                if w3data:
                    w3s, w3f = w3data.get("w3", [None, None])
                    if w3s and w3f and time_in_window(arr_t, w3s, w3f):
                        tf3 = parse_hhmm(w3f)
                        if tf3:
                            end_w3 = _dt.combine(cursor.date(), tf3)
                            if end_w3 <= cursor:
                                end_w3 += _td(days=1)
                            wait_m = (end_w3 - cursor).total_seconds() / 60.0

            # Penalty: wait_mins * milking_weight (same units as km at ~1.2 min/km)
            wait_pen += wait_m * milking_weight
            # Advance cursor by the wait so downstream farms see a later arrival
            cursor   += _td(minutes=wait_m)

    return total_km + wait_pen


def _group_dest_catalogue(sheets):
    """Return {dest_key: dest_name} for every processor in the colour group."""
    cat = {}
    for _sname, entry in sheets:
        for block in entry.get("blocks", []):
            for d in (block.get("dests") or []):
                dk = d.get("key"); dn = d.get("name","")
                if dk: cat[dk] = dn or dk
            # legacy fallback
            dk = block.get("dest_key")
            if dk and dk not in cat:
                cat[dk] = block.get("dest_name","") or dk
    return cat


def _sheet_cost(blocks, dm, start_time, cfg):
    """
    Scalar cost for one truck's day (one sheet's worth of blocks).

    cfg keys used:
      orig_dest_vols  ? {dest_key: original_litres}  (group-wide)
      vol_tol         ? fractional tolerance  (0.15 -> +/-15 %)
      vol_penalty     ? penalty per litre outside tolerance
      milking_weight  ? multiplier on milking-wait km-equivalent
      max_shift_h     ? maximum shift hours before penalty
      shift_penalty   ? penalty per hour over max_shift_h
    """
    # -- distance --------------------------------------------------------------
    total_km  = 0.0
    all_dists = calc_distances(blocks, dm)
    for dists in all_dists:
        for d in dists[:-1]:
            if d is not None:
                total_km += d

    # -- shift time & milking-window waits -------------------------------------
    shift_hours   = 0.0
    milking_mins  = 0.0

    if start_time:
        _suppress = cfg.get("suppress_no_milking", True)
        _ct2 = calc_times(blocks, dm, start_time, suppress_no_milking=_suppress,
                          precomputed_dists=all_dists)
        all_times  = _ct2[0] if _ct2 is not None else None
        _end_cur2  = _ct2[1] if _ct2 is not None else None
        if _end_cur2 is not None:
            base = datetime.combine(date.today(), start_time)
            shift_hours = (_end_cur2 - base).total_seconds() / 3600.0
        if all_times:

            for b_idx, block in enumerate(blocks):
                btimes = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes:
                    continue
                for f_i, farm in enumerate(block["rows"]):
                    f_stop = _farm_stop_index(block, f_i, b_idx, blocks)
                    ft = btimes[f_stop] if f_stop < len(btimes) else None
                    if ft is None or ft["arr"] is None:
                        continue
                    arr = ft["arr"]
                    arr_dt = datetime.combine(date.today(), arr)

                    # Skip milking penalty for suppressed farms
                    if _suppress and farm.get("irma","") in NO_MILKING_WINDOW_FARMS:
                        continue

                    # MWO: skip milking penalty entirely for flagged farms
                    if farm.get("_mwo"):
                        continue

                    for s_key, f_key in [("m1_start", "m1_finish"),
                                         ("m2_start", "m2_finish")]:
                        s_str = farm.get(s_key, "")
                        f_str = farm.get(f_key, "")
                        if time_in_window(arr, s_str, f_str):
                            tf = parse_hhmm(f_str)
                            if tf:
                                end_w = datetime.combine(date.today(), tf)
                                if end_w < arr_dt:
                                    end_w += timedelta(days=1)
                                milking_mins += (end_w - arr_dt).total_seconds() / 60.0
                            break
                    # w3 penalty
                    w3data = THREE_WINDOW_FARMS.get(farm.get("irma",""))
                    if w3data:
                        w3pair = w3data.get("w3", [None, None])
                        w3s, w3f = w3pair[0], w3pair[1]
                        if w3s and w3f and time_in_window(arr, w3s, w3f):
                            tf3 = parse_hhmm(w3f)
                            if tf3:
                                end_w3 = datetime.combine(date.today(), tf3)
                                if end_w3 < arr_dt:
                                    end_w3 += timedelta(days=1)
                                milking_mins += (end_w3 - arr_dt).total_seconds() / 60.0

    # -- plant volume penalty (group-wide vols supplied via cfg) ---------------
    # NOTE: vol penalty is computed at group level in _group_cost to avoid
    # double-counting; here we just return the non-volume portion so that
    # _best_insert_pos can use a fast per-sheet cost without needing group state.
    # When called from _group_cost, vol_pen is added there instead.
    vol_pen = 0.0
    if cfg.get("_include_vol_pen", False):
        vol_tol          = cfg.get("vol_tol", 0.15)
        vol_penalty_rate = cfg.get("vol_penalty", 1.0)
        orig_dest_vols   = cfg.get("orig_dest_vols", {})
        dest_vols        = {}
        for block in blocks:
            for dk, off in _block_dest_offloads(block).items():
                dest_vols[dk] = dest_vols.get(dk, 0.0) + off
        for dk, orig_vol in orig_dest_vols.items():
            cur_vol = dest_vols.get(dk, 0.0)
            lo = orig_vol * (1.0 - vol_tol)
            hi = orig_vol * (1.0 + vol_tol)
            if cur_vol < lo:
                vol_pen += (lo - cur_vol) * vol_penalty_rate
            elif cur_vol > hi:
                vol_pen += (cur_vol - hi) * vol_penalty_rate

    # -- truck capacity penalty ------------------------------------------------
    # For multi-dest routes with split_after, the truck offloads mid-route so
    # the total farm volume overstates the actual peak load.  We walk the stop
    # sequence and track the running load: +farm vol at farm stops, -offload vol
    # at dest stops.  The peak load is what actually matters for capacity.
    hard_cap      = cfg.get("hard_vol_cap", VOL_LIMIT)
    cap_pen_rate  = cfg.get("cap_penalty", 2.0)
    cap_pen       = 0.0
    for b_idx, block in enumerate(blocks):
        if _is_preload_block(block):
            continue   # preload blocks start empty ? no cap issue
        dests = block.get("dests") or []
        if not dests:
            dk = block.get("dest_key","")
            dests = [{"key": dk, "vol_partial": None}] if dk else []
        farms = block.get("rows", [])
        total_farm_vol = sum((r.get("prior_vol") or 0) for r in farms
                             if isinstance(r.get("prior_vol"), (int, float)))

        # If any dest has a split_after, compute peak load along the sequence
        has_split = _block_has_split(block)
        if has_split:
            is_last  = (b_idx == len(blocks) - 1)
            origin   = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
            stops    = _build_block_stops(block, origin, is_last)
            running  = 0.0
            peak     = 0.0
            already_delivered = 0.0
            for stop in stops:
                if stop["type"] == "farm":
                    vol = stop["farm"].get("prior_vol") or 0
                    running += vol if isinstance(vol, (int, float)) else 0
                    peak = max(peak, running)
                elif stop["type"] == "dest":
                    offload = _dest_vol_partial(stop["dest"], total_farm_vol, already_delivered)
                    already_delivered += offload
                    running = max(0.0, running - offload)
            if peak > hard_cap:
                cap_pen += (peak - hard_cap) * cap_pen_rate
        else:
            # No mid-route dropoff ? total farm vol is the peak load
            if total_farm_vol > hard_cap:
                cap_pen += (total_farm_vol - hard_cap) * cap_pen_rate

    # -- shift overage penalty -------------------------------------------------
    max_shift      = cfg.get("max_shift_h", 14.0)
    shift_pen_rate = cfg.get("shift_penalty", 200.0)
    shift_pen      = max(0.0, shift_hours - max_shift) * shift_pen_rate

    # -- total shift hours cost ------------------------------------------------
    # A small per-hour cost on the full shift duration (not just the overage).
    # This gives the solver a continuous gradient toward shorter days ? without
    # it, any route that stays under max_shift_h looks equally good regardless
    # of whether it finishes in 10 hours or 13.5.
    # Default: 5.0 km-equivalent per hour (small enough not to override distance
    # or milking objectives, large enough to break ties in favour of earlier ends).
    shift_hours_rate = cfg.get("shift_hours_weight", 0.0)
    shift_hours_cost = shift_hours * shift_hours_rate if shift_hours > 0 else 0.0

    # -- milking wait penalty --------------------------------------------------
    # milking_mins * milking_weight: 1 min wait = 1 km-equivalent at weight=1.
    milking_equiv = milking_mins * cfg.get("milking_weight", 1.0)

    # -- plant receiving-window penalty ----------------------------------------
    # Both components expressed in km-equivalent per hour ? same scale as routing.
    # 1. OUTSIDE penalty: arrival before open or after close.
    #    Rate: plant_win_penalty (default 200 km/h).
    # 2. MARGIN penalty: arrival inside the window but within the last
    #    plant_win_margin_mins minutes before close ? gradient toward earlier arrivals.
    #    Rate: plant_win_margin_rate (default = plant_win_penalty * 0.5 km/h).
    plant_windows       = cfg.get("plant_windows", {})
    plant_win_rate      = cfg.get("plant_win_penalty", 200.0)          # km per hour outside
    plant_margin_mins   = cfg.get("plant_win_margin_mins", 30.0)
    plant_margin_rate   = cfg.get("plant_win_margin_rate",
                                  plant_win_rate * 0.5)                # km per hour inside margin
    plant_win_cost = 0.0
    if plant_windows and start_time and all_times:
        for b_idx3, block3 in enumerate(blocks):
            btimes3 = all_times[b_idx3] if b_idx3 < len(all_times) else None
            if not btimes3:
                continue
            dests3 = block3.get("dests") or []
            if not dests3:
                dk3 = block3.get("dest_key", "")
                dests3 = [{"key": dk3}] if dk3 else []
            n_farms3 = len(block3["rows"])
            for d_i3, dest_d3 in enumerate(dests3):
                # Yard-for destinations are overnight parking ? no receiving window
                if "yard for" in (dest_d3.get("name","") or "").lower():
                    continue
                dk3 = normalise_key(dest_d3.get("key", "") or "")
                window3 = plant_windows.get(dk3)
                if window3 is None:
                    continue
                open_str3, close_str3 = window3
                if block3.get("preload"):
                    t_idx3 = 1
                else:
                    t_idx3 = _dest_stop_index(block3, d_i3, b_idx3, blocks)
                ft3 = btimes3[t_idx3] if t_idx3 < len(btimes3) else None
                if ft3 is None or ft3.get("arr") is None:
                    continue
                arr3     = ft3["arr"]
                arr_dt3  = datetime.combine(date.today(), arr3)
                close_t3 = parse_hhmm(close_str3)
                open_t3  = parse_hhmm(open_str3)

                if not time_in_window(arr3, open_str3, close_str3):
                    # Outside window ? penalise by hours until next open
                    if open_t3 is not None:
                        open_dt3 = datetime.combine(date.today(), open_t3)
                        if open_t3 > arr3:
                            wait_h = (open_dt3 - arr_dt3).total_seconds() / 3600.0
                        else:
                            open_dt3 += timedelta(days=1)
                            wait_h = (open_dt3 - arr_dt3).total_seconds() / 3600.0
                        plant_win_cost += wait_h * plant_win_rate
                    else:
                        plant_win_cost += 1.0 * plant_win_rate   # flat 1-hour penalty
                elif plant_margin_mins > 0 and close_t3 is not None and plant_margin_rate > 0:
                    close_dt3 = datetime.combine(date.today(), close_t3)
                    if open_t3 is not None and close_t3 < open_t3:
                        close_dt3 += timedelta(days=1)
                    mins_to_close = (close_dt3 - arr_dt3).total_seconds() / 60.0
                    if mins_to_close < 0:
                        close_dt3 += timedelta(days=1)
                        mins_to_close = (close_dt3 - arr_dt3).total_seconds() / 60.0
                    if mins_to_close < plant_margin_mins:
                        depth = (plant_margin_mins - mins_to_close) / plant_margin_mins
                        # depth * margin expressed as hours * rate
                        plant_win_cost += depth * plant_margin_rate * (plant_margin_mins / 60.0)

    return total_km + milking_equiv + vol_pen + shift_pen + shift_hours_cost + cap_pen + plant_win_cost


def _sheet_cost_breakdown(blocks, dm, start_time, cfg):
    """Same as _sheet_cost but returns a dict of cost components instead of a scalar.
    Used by the Full Cost Report so it always uses exactly the same logic as the solver."""
    # Re-use _sheet_cost internals by running both and computing the breakdown.
    # We compute each component independently using the same logic as _sheet_cost.

    # -- distance -------------------------------------------------------------
    total_km = 0.0
    for dists in calc_distances(blocks, dm):
        for d in dists[:-1]:
            if d is not None:
                total_km += d

    # -- shift time & milking -------------------------------------------------
    shift_hours  = 0.0
    milking_mins = 0.0
    all_times    = None
    _suppress    = cfg.get("suppress_no_milking", True)
    if start_time:
        _ct = calc_times(blocks, dm, start_time, suppress_no_milking=_suppress)
        if _ct is not None:
            all_times = _ct[0]
            base = datetime.combine(date.today(), start_time)
            shift_hours = (_ct[1] - base).total_seconds() / 3600.0
        if all_times:
            for b_idx, block in enumerate(blocks):
                btimes = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes: continue
                for f_i, farm in enumerate(block["rows"]):
                    f_stop = _farm_stop_index(block, f_i, b_idx, blocks)
                    ft = btimes[f_stop] if f_stop < len(btimes) else None
                    if ft is None or ft.get("arr") is None: continue
                    arr = ft["arr"]
                    arr_dt = datetime.combine(date.today(), arr)
                    if _suppress and farm.get("irma","") in NO_MILKING_WINDOW_FARMS:
                        continue
                    if farm.get("_mwo"):
                        continue
                    for s_key, f_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
                        if time_in_window(arr, farm.get(s_key,""), farm.get(f_key,"")):
                            tf = parse_hhmm(farm.get(f_key,""))
                            if tf:
                                end_w = datetime.combine(date.today(), tf)
                                if end_w < arr_dt: end_w += timedelta(days=1)
                                milking_mins += (end_w - arr_dt).total_seconds() / 60.0
                            break
                    w3data = THREE_WINDOW_FARMS.get(farm.get("irma",""))
                    if w3data:
                        w3s, w3f = w3data.get("w3",[None,None])
                        if w3s and w3f and time_in_window(arr, w3s, w3f):
                            tf3 = parse_hhmm(w3f)
                            if tf3:
                                end_w3 = datetime.combine(date.today(), tf3)
                                if end_w3 < arr_dt: end_w3 += timedelta(days=1)
                                milking_mins += (end_w3 - arr_dt).total_seconds() / 60.0

    # -- cap -------------------------------------------------------------------
    hard_cap     = cfg.get("hard_vol_cap", VOL_LIMIT)
    cap_pen_rate = cfg.get("cap_penalty", 2.0)
    cap_pen      = 0.0
    for b_idx, block in enumerate(blocks):
        if _is_preload_block(block): continue
        dests = block.get("dests") or []
        if not dests:
            dk = block.get("dest_key","")
            dests = [{"key": dk, "vol_partial": None}] if dk else []
        farms = block.get("rows", [])
        total_fv = sum((r.get("prior_vol") or 0) for r in farms
                       if isinstance(r.get("prior_vol"),(int,float)))
        if any(d.get("split_after") is not None for d in dests):
            is_last = (b_idx == len(blocks)-1)
            origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
            stops   = _build_block_stops(block, origin, is_last)
            running = peak = adel = 0.0
            for stop in stops:
                if stop["type"] == "farm":
                    v = stop["farm"].get("prior_vol") or 0
                    running += v if isinstance(v,(int,float)) else 0
                    peak = max(peak, running)
                elif stop["type"] == "dest":
                    off = _dest_vol_partial(stop["dest"], total_fv, adel)
                    adel += off; running = max(0.0, running - off)
            if peak > hard_cap: cap_pen += (peak - hard_cap) * cap_pen_rate
        else:
            if total_fv > hard_cap: cap_pen += (total_fv - hard_cap) * cap_pen_rate

    # -- overtime --------------------------------------------------------------
    max_shift    = cfg.get("max_shift_h", 14.0)
    shift_pen    = max(0.0, shift_hours - max_shift) * cfg.get("shift_penalty", 200.0)
    shift_cost   = shift_hours * cfg.get("shift_hours_weight", 0.0)

    # -- milking ---------------------------------------------------------------
    milking_equiv = milking_mins * cfg.get("milking_weight", 1.0)

    # -- plant window ----------------------------------------------------------
    plant_windows     = cfg.get("plant_windows", {})
    plant_win_rate    = cfg.get("plant_win_penalty", 200.0)
    plant_margin_mins = cfg.get("plant_win_margin_mins", 30.0)
    plant_margin_rate = cfg.get("plant_win_margin_rate", plant_win_rate * 0.5)
    plant_win_cost    = 0.0
    if plant_windows and all_times:
        for b_idx3, block3 in enumerate(blocks):
            btimes3 = all_times[b_idx3] if b_idx3 < len(all_times) else None
            if not btimes3: continue
            dests3 = block3.get("dests") or []
            if not dests3:
                dk3 = block3.get("dest_key","")
                dests3 = [{"key": dk3}] if dk3 else []
            for d_i3, dest_d3 in enumerate(dests3):
                if "yard for" in (dest_d3.get("name","") or "").lower(): continue
                dk3 = normalise_key(dest_d3.get("key","") or "")
                window3 = plant_windows.get(dk3)
                if not window3: continue
                open_str3, close_str3 = window3
                t_idx3 = (1 if block3.get("preload")
                          else _dest_stop_index(block3, d_i3, b_idx3, blocks))
                ft3 = btimes3[t_idx3] if t_idx3 < len(btimes3) else None
                if ft3 is None or ft3.get("arr") is None: continue
                arr3 = ft3["arr"]
                arr_dt3 = datetime.combine(date.today(), arr3)
                open_t3 = parse_hhmm(open_str3)
                if not time_in_window(arr3, open_str3, close_str3):
                    if open_t3:
                        open_dt3 = datetime.combine(date.today(), open_t3)
                        if open_t3 <= arr3: open_dt3 += timedelta(days=1)
                        plant_win_cost += (open_dt3 - arr_dt3).total_seconds()/3600.0 * plant_win_rate
                    else:
                        plant_win_cost += plant_win_rate
                else:
                    close_t3 = parse_hhmm(close_str3)
                    if close_t3:
                        close_dt3 = datetime.combine(date.today(), close_t3)
                        if open_t3 and close_t3 < open_t3: close_dt3 += timedelta(days=1)
                        mtc = (close_dt3 - arr_dt3).total_seconds()/60.0
                        if mtc < 0:
                            close_dt3 += timedelta(days=1)
                            mtc = (close_dt3 - arr_dt3).total_seconds()/60.0
                        if mtc < plant_margin_mins:
                            depth = (plant_margin_mins - mtc) / plant_margin_mins
                            plant_win_cost += depth * plant_margin_rate * (plant_margin_mins / 60.0)

    return {
        "km":        total_km,
        "milking":   milking_equiv,
        "shift":     shift_cost,
        "overtime":  shift_pen,
        "cap":       cap_pen,
        "plant_win": plant_win_cost,
        "total":     total_km + milking_equiv + shift_cost + shift_pen + cap_pen + plant_win_cost,
    }


def _optimize_split_positions(blocks, dm, start_time, cfg, dm_dur=None):
    """For each block with a mid-route partial-dropoff dest (fixed vol_partial,
    not the last dest, has a plant_window constraint), find the farm insertion
    position that minimises _sheet_cost.  Mutates blocks in-place.
    Returns True if any position changed."""
    plant_windows = cfg.get("plant_windows", {})
    changed = False
    for b_idx, block in enumerate(blocks):
        if _is_preload_block(block) or _is_fixed_vol_block(block):
            continue
        dests   = block.get("dests") or []
        n_farms = len(block["rows"])
        if n_farms == 0 or len(dests) < 2:
            continue
        for d_i, dest_d in enumerate(dests):
            if d_i == len(dests) - 1:
                continue   # last dest is never split
            if dest_d.get("vol_partial") is None:
                continue   # not a fixed partial volume
            dk = normalise_key(dest_d.get("key","") or "")
            if dk not in plant_windows:
                continue   # only optimise window-constrained dests
            best_pos  = dest_d.get("split_after") if dest_d.get("split_after") is not None else n_farms
            dest_d["split_after"] = best_pos
            best_cost = _sheet_cost(blocks, dm, start_time, cfg)
            for pos in range(n_farms + 1):
                if pos == best_pos:
                    continue
                dest_d["split_after"] = pos
                c = _sheet_cost(blocks, dm, start_time, cfg)
                if c < best_cost:
                    best_cost = c
                    best_pos  = pos
            if dest_d.get("split_after") != best_pos:
                changed = True
            dest_d["split_after"] = best_pos
    return changed


def _sheet_cost_breakdown_state(state, dm, cache, fname, cfg):
    """Aggregate _sheet_cost_breakdown across all sheets in a solver state."""
    totals = {"km":0.0,"milking":0.0,"shift":0.0,"overtime":0.0,"cap":0.0,"plant_win":0.0,"total":0.0}
    for sname, blocks in state:
        entry = cache.get(fname, {}).get(sname, {})
        st    = entry.get("start_time") if isinstance(entry, dict) else None
        if not st: continue
        bd = _sheet_cost_breakdown(blocks, dm, st, cfg)
        for k in totals:
            totals[k] += bd.get(k, 0.0)
    return totals


def _group_vol_penalty(state, orig_dest_vols, cfg):
    """
    Compute the volume penalty for the entire colour group at once.
    Sums delivered litres to each processor across ALL sheets, then
    penalises deviations from orig_dest_vols.
    """
    vol_tol          = cfg.get("vol_tol", 0.15)
    vol_penalty_rate = cfg.get("vol_penalty", 1.0)
    dest_vols = {}
    for _sname, blocks in state:
        for block in blocks:
            for dk, off in _block_dest_offloads(block).items():
                dest_vols[dk] = dest_vols.get(dk, 0.0) + off
    pen = 0.0
    for dk, orig_vol in orig_dest_vols.items():
        cur_vol = dest_vols.get(dk, 0.0)
        lo = orig_vol * (1.0 - vol_tol)
        hi = orig_vol * (1.0 + vol_tol)
        if cur_vol < lo:
            pen += (lo - cur_vol) * vol_penalty_rate
        elif cur_vol > hi:
            pen += (cur_vol - hi) * vol_penalty_rate
    return pen


def _group_sheets_by_colour(cache, fname):
    """Return {"RED": [(sname, entry), ...], "BLUE": [...]} for the loaded file."""
    groups = {"RED": [], "BLUE": []}
    if fname not in cache:
        return groups
    for sname, entry in cache[fname].items():
        if not isinstance(entry, dict):
            continue
        bucket = _sheet_colour_bucket(entry.get("day_colour", ""))
        if bucket in groups:
            groups[bucket].append((sname, entry))
    return groups


def _highs_verify_processor_assignment(colour, sheets, state, dm, cfg, log_fn):
    """
    Post-optimality check: given the ALNS result (fixed farm sequences, fixed
    processor assignments), ask HiGHS whether reassigning processors across
    routes would reduce total cost.

    The MIP variables:
      x[i,j] in {0,1}  ?  route i is assigned to processor j

    Objective: minimise  sum_{i,j} x[i,j] * last_leg_cost(i,j)
                       + vol_deviation_penalties
                       + shift_overage_penalties   (encoded as big-M terms)
                       + cap_overage_penalties

    Constraints:
      Each route assigned to exactly one processor (from that route's
      original candidate set ? processors that appeared in the file for
      this colour group).
      Processor volume balance within +/-vol_tol of original.
      Truck capacity hard cap (as a big-M penalty rather than a hard cut,
      to keep the MIP feasible when data itself exceeds cap).

    Returns a string summary (multi-line) to be appended to the solver log.
    """
    try:
        from scipy.optimize import linprog, milp, LinearConstraint, Bounds
        import numpy as np
    except ImportError:
        return "  [HiGHS check] scipy not available ? skipping verification."

    # -- gather route data from ALNS result -----------------------------------
    # state is list of (sname, blocks); map to flat list of route descriptors
    route_records = []   # {sname, block_idx, block, vol, start_time}
    for sname, blocks in state:
        entry = {}
        for _sn, e in sheets:
            if _sn == sname: entry = e; break
        st = entry.get("start_time") if isinstance(entry, dict) else None
        for bi, block in enumerate(blocks):
            vol = sum((r["prior_vol"] or 0) for r in block.get("rows", [])
                      if isinstance(r.get("prior_vol"), (int, float)))
            route_records.append({"sname": sname, "bi": bi, "block": block,
                                  "vol": vol, "start_time": st})

    # catalogue of all processors seen in this colour group
    dest_keys = sorted({
        d.get("key") or "?"
        for _sn, blocks in state
        for block in blocks
        for d in (block.get("dests") or [{"key": block.get("dest_key","?")}])
        if d.get("key")
    })
    if not dest_keys or not route_records:
        return f"  [{colour}] HiGHS: no routes or processors ? skipping."

    n_routes = len(route_records)
    n_procs  = len(dest_keys)
    proc_idx = {dk: j for j, dk in enumerate(dest_keys)}

    # original processor assignment from ALNS result
    def _route_current_dest(rec):
        dests = rec["block"].get("dests") or []
        if dests: return dests[0].get("key") or "?"
        return rec["block"].get("dest_key") or "?"

    # original processor volumes (from ALNS result ? used for tolerance bounds)
    orig_dest_vols = {}
    for sname, entry_orig in sheets:
        for block in entry_orig.get("blocks", []):
            dests_b = block.get("dests") or []
            if not dests_b:
                dk = block.get("dest_key") or "?"
                dests_b = [{"key": dk, "vol_partial": None}]
            fv = sum((r["prior_vol"] or 0) for r in block.get("rows", [])
                     if isinstance(r.get("prior_vol"), (int, float)))
            already = 0.0
            for d in dests_b:
                dk = d.get("key") or "?"
                vp = d.get("vol_partial")
                rem = max(0.0, fv - already)
                off = min(float(vp), rem) if vp is not None else rem
                already += off
                orig_dest_vols[dk] = orig_dest_vols.get(dk, 0.0) + off

    # -- cost coefficients c[i*n_procs + j] ----------------------------------
    # Cost = last-farm-to-processor leg distance (km via dm, or duration if dm_dur)
    # We use the same units as _sheet_cost (km-equivalent).
    # Shift penalty and cap penalty encoded as additive per-assignment costs.
    vol_tol      = cfg.get("vol_tol", 0.15)
    vol_pen_rate = cfg.get("vol_penalty", 1.0)
    hard_cap     = cfg.get("hard_vol_cap", VOL_LIMIT)
    cap_pen_rate = cfg.get("cap_penalty", 2.0)

    c = []
    for rec in route_records:
        block = rec["block"]
        farms = [r["irma"] for r in block.get("rows", [])]
        last_farm = farms[-1] if farms else "VEDDER"
        for dk in dest_keys:
            leg = lookup(dm, last_farm, dk)
            leg_cost = leg if leg is not None else 9999.0
            # Cap penalty contribution for this route if vol > hard_cap
            cap_contrib = max(0.0, rec["vol"] - hard_cap) * cap_pen_rate
            c.append(leg_cost + cap_contrib)

    # -- integrality: all binary -----------------------------------------------
    integrality = [1] * (n_routes * n_procs)  # 1 = integer

    # -- bounds: 0 ? x ? 1 ----------------------------------------------------
    bounds = Bounds(lb=0.0, ub=1.0)

    # -- equality constraint: each route assigned to exactly one processor -----
    # sum_j x[i,j] = 1  for each i
    A_eq_rows, b_eq = [], []
    for i in range(n_routes):
        row = [0.0] * (n_routes * n_procs)
        for j in range(n_procs):
            row[i * n_procs + j] = 1.0
        A_eq_rows.append(row)
        b_eq.append(1.0)

    # -- inequality constraints: volume tolerance per processor ----------------
    # sum_i vol_i * x[i,j]  <=  orig_vol_j * (1 + vol_tol)
    # sum_i vol_i * x[i,j]  >=  orig_vol_j * (1 - vol_tol)   [as <= with negation]
    A_ineq_rows, b_ineq = [], []
    for j, dk in enumerate(dest_keys):
        orig_v = orig_dest_vols.get(dk, 0.0)
        if orig_v <= 0:
            continue
        hi = orig_v * (1.0 + vol_tol)
        lo = orig_v * (1.0 - vol_tol)
        # upper bound
        row_hi = [0.0] * (n_routes * n_procs)
        for i, rec in enumerate(route_records):
            row_hi[i * n_procs + j] = rec["vol"]
        A_ineq_rows.append(row_hi); b_ineq.append(hi)
        # lower bound (negate)
        row_lo = [-x for x in row_hi]
        A_ineq_rows.append(row_lo); b_ineq.append(-lo)

    import numpy as np
    from scipy.optimize import milp, LinearConstraint, Bounds

    c_arr     = np.array(c, dtype=float)
    bounds_   = Bounds(lb=0.0, ub=1.0)

    constraints = []
    if A_eq_rows:
        A_eq = np.array(A_eq_rows, dtype=float)
        constraints.append(LinearConstraint(A_eq, lb=np.array(b_eq), ub=np.array(b_eq)))
    if A_ineq_rows:
        A_iq = np.array(A_ineq_rows, dtype=float)
        neg_inf = np.full(len(b_ineq), -np.inf)
        constraints.append(LinearConstraint(A_iq, lb=neg_inf, ub=np.array(b_ineq)))

    try:
        result = milp(c_arr, constraints=constraints, integrality=np.array(integrality),
                      bounds=bounds_, options={"disp": False, "time_limit": 30.0})
    except Exception as ex:
        return f"  [{colour}] HiGHS MIP failed: {ex}"

    if not result.success:
        return f"  [{colour}] HiGHS: no feasible processor assignment found ({result.message})."

    # -- compute current (ALNS) last-leg cost for comparison ------------------
    current_cost = 0.0
    for i, rec in enumerate(route_records):
        cur_dk = _route_current_dest(rec)
        j_cur  = proc_idx.get(cur_dk, 0)
        current_cost += c[i * n_procs + j_cur]

    mip_cost = result.fun
    improvement = current_cost - mip_cost

    lines = [f"\n-- [{colour}] HiGHS Processor Assignment Verification --"]
    lines.append(f"   ALNS last-leg + cap cost : {current_cost:,.1f} km-eq")
    lines.append(f"   MIP optimal cost          : {mip_cost:,.1f} km-eq")

    THRESHOLD = 0.5   # ignore sub-km differences (floating point noise)
    if improvement <= THRESHOLD:
        lines.append("   OK Processor assignment is OPTIMAL ? no improvement possible.")
    else:
        lines.append(f"   * Improvement available   : {improvement:,.1f} km-eq")
        lines.append("   Suggested reassignments:")
        x_sol = result.x
        for i, rec in enumerate(route_records):
            cur_dk   = _route_current_dest(rec)
            best_j   = int(np.argmax(x_sol[i*n_procs:(i+1)*n_procs]))
            best_dk  = dest_keys[best_j]
            if best_dk != cur_dk:
                farms = [r["irma"] for r in rec["block"].get("rows", [])]
                last  = farms[-1] if farms else "?"
                cur_cost_leg  = c[i * n_procs + proc_idx.get(cur_dk, 0)]
                best_cost_leg = c[i * n_procs + best_j]
                lines.append(
                    f"     Sheet {rec['sname']!r} block {rec['bi']+1}: "
                    f"{cur_dk} -> {best_dk}  "
                    f"(last farm {last}, "
                    f"leg {cur_cost_leg:.1f} -> {best_cost_leg:.1f} km-eq)")
    return "\n".join(lines)


class IntraRouteOptimiser(QThread):
    """Applies 2-opt + or-opt within every route until convergence.
    No cross-route moves ? pure within-block reordering."""
    progress = pyqtSignal(int, int, str)   # cur, total, status
    finished = pyqtSignal(dict)            # {(fname,sname): improved_blocks}
    log      = pyqtSignal(str)

    def __init__(self, fname, cache, dm, cfg, sheet_mods, parent=None):
        super().__init__(parent)
        self.fname      = fname
        self.cache      = cache
        self.dm         = dm
        self.cfg        = cfg
        self.sheet_mods = dict(sheet_mods)

    def run(self):
        results     = {}
        all_snames  = sorted(self.cache.get(self.fname, {}).keys())
        total       = len(all_snames)
        n_improved  = 0

        for cur, sname in enumerate(all_snames):
            entry = self.cache[self.fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue

            key    = (self.fname, sname)
            blocks = copy.deepcopy(
                self.sheet_mods.get(key, entry.get("blocks", [])))

            self.progress.emit(cur, total, f"Optimising {sname}...")

            improved = True
            changed  = False
            while improved:
                improved = False
                # 2-opt
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    base_c    = _sheet_cost(blocks, self.dm, start_time, self.cfg)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n - 1):
                        for j in range(i + 1, n):
                            trial = rows[:i] + rows[i:j+1][::-1] + rows[j+1:]
                            blocks[b_idx] = dict(block, rows=trial)
                            c = _sheet_cost(blocks, self.dm, start_time, self.cfg)
                            if c < best_c:
                                best_c = c; best_rows = trial[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True; changed = True

                # or-opt (single farm relocation)
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    base_c    = _sheet_cost(blocks, self.dm, start_time, self.cfg)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n):
                        farm = rows[i]
                        rest = rows[:i] + rows[i+1:]
                        for j in range(len(rest) + 1):
                            trial = rest[:j] + [farm] + rest[j:]
                            blocks[b_idx] = dict(block, rows=trial)
                            c = _sheet_cost(blocks, self.dm, start_time, self.cfg)
                            if c < best_c:
                                best_c = c; best_rows = trial[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True; changed = True

            if changed:
                results[key] = blocks
                n_improved += 1

        self.progress.emit(total, total, "Done")
        self.log.emit(f"Intra-route optimisation complete ? {n_improved} route(s) improved")
        self.finished.emit(results)


def _apply_prob_floors_ceilings(probs, key_floors, key_ceilings):
    """Redistribute a probability vector in place so each entry respects its
    per-index floor and ceiling, iterating until stable.

    Below-floor entries are raised to their floor and the deficit is taken
    proportionally from entries above their floor; above-ceiling entries are
    clamped and the surplus is redistributed proportionally to those below
    their ceiling.  Returns the same list (mutated).  Shared by _roulette
    (selection) and the solver's probability-display logging so the two can
    never drift.
    """
    n = len(probs)
    if not (any(f > 0.0 for f in key_floors) or any(c < 1.0 for c in key_ceilings)):
        return probs
    for _ in range(n * 2):   # extra iterations for combined floor+ceiling cascades
        changed = False
        # Raise below-floor keys
        below = [i for i, p in enumerate(probs) if p < key_floors[i]]
        above_floor = [i for i, p in enumerate(probs) if p >= key_floors[i]]
        if below and above_floor:
            deficit   = sum(key_floors[i] - probs[i] for i in below)
            above_sum = sum(probs[i] for i in above_floor)
            for i in below:
                probs[i] = key_floors[i]
            if above_sum > 0:
                for i in above_floor:
                    probs[i] -= deficit * (probs[i] / above_sum)
            changed = True
        # Clamp above-ceiling keys
        over = [i for i, p in enumerate(probs) if p > key_ceilings[i] + 1e-9]
        over_set = set(over)
        under_ceil = [i for i, p in enumerate(probs)
                      if p <= key_ceilings[i] and i not in over_set]
        if over and under_ceil:
            surplus   = sum(probs[i] - key_ceilings[i] for i in over)
            under_sum = sum(probs[i] for i in under_ceil)
            for i in over:
                probs[i] = key_ceilings[i]
            if under_sum > 0:
                for i in under_ceil:
                    probs[i] += surplus * (probs[i] / under_sum)
            changed = True
        if not changed:
            break
    return probs


class ALNSSolver(QThread):
    """
    Adaptive Large Neighbourhood Search.

    Operates on RED and BLUE colour groups independently.
    Within each group, farms may move between any sheets/routes of the same
    colour, and block destinations (dest_key) may also be reassigned.

    Three paired move types, each with its own adaptive weight:
      1. Farm move   ? destroy (random | worst) + repair (best-insert | regret)
      2. Dest move   ? strip dest_keys from n blocks + regret-order reassignment
      3. Combined    ? both farm and dest moves together

    Simulated-annealing acceptance; alpha auto-computed from user's
    target_cool_frac and iteration count.

    Signals
    -------
    progress(cur_iter, total_iters, status_str)
    finished({sname: new_blocks, ...})
    log(message_str)
    """

    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(dict)
    log      = pyqtSignal(str)

    def __init__(self, fname, cache, dm, cfg, parent=None, sheet_mods=None,
                 locked_sheets=None):
        super().__init__(parent)
        self.fname         = fname
        self.cache         = cache
        self.dm            = dm
        self.cfg           = cfg
        self._stop         = False
        self.sheet_mods    = sheet_mods or {}
        # locked_sheets: set of sname strings the solver must not modify
        self.locked_sheets = {str(s).strip() for s in (locked_sheets or set())}
        # Accumulated across all colour groups during run(); read by MainWindow
        # after solving to write the Zero Vol Farms sheet on export.
        self.zero_vol_farms = []   # list of (sname, b_idx, farm_dict)

    def stop(self):
        self._stop = True

    # -- group-level cost ------------------------------------------------------

    def _group_cost(self, state, orig_dest_vols, sheet_cost_cache=None):
        """Full cost: sum per-sheet costs + group-wide volume penalty.

        sheet_cost_cache: optional dict {sname: cost} ? if provided, sheets
        present in the cache are not recomputed.  Callers that know which sheets
        changed can pass a partial cache to skip unchanged sheets entirely.
        """
        total = 0.0
        for sname, blocks in state:
            if sheet_cost_cache is not None and sname in sheet_cost_cache:
                total += sheet_cost_cache[sname]
            else:
                entry = self.cache[self.fname].get(sname, {})
                st    = entry.get("start_time") if isinstance(entry, dict) else None
                c     = _sheet_cost(blocks, self.dm, st, self.cfg)
                total += c
                if sheet_cost_cache is not None:
                    sheet_cost_cache[sname] = c
        total += _group_vol_penalty(state, orig_dest_vols, self.cfg)
        return total

    def _make_sheet_cost_cache(self, state):
        """Compute and return a full per-sheet cost cache for the given state."""
        cache = {}
        for sname, blocks in state:
            entry = self.cache[self.fname].get(sname, {})
            st    = entry.get("start_time") if isinstance(entry, dict) else None
            cache[sname] = _sheet_cost(blocks, self.dm, st, self.cfg)
        return cache

    # -- flat farm list --------------------------------------------------------

    def _flatten_farms(self, state):
        out = []
        for s_idx, (sname, blocks) in enumerate(state):
            for b_idx, block in enumerate(blocks):
                for farm in block["rows"]:
                    out.append((s_idx, b_idx, farm))
        return out

    # ??????????????????????????????????????????????????????????????????????????
    # FARM destroy / repair
    # ??????????????????????????????????????????????????????????????????????????

    def _destroy_random(self, state, n_remove):
        """Remove n_remove random farms.
        - Preload blocks (previous-day delivery, no farms): completely frozen.
        - Holdover blocks (yard-for dest): farms CAN be moved; dest stays fixed.
        """
        flat = []
        for s_idx, (sname, blocks) in enumerate(state):
            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block):
                    continue   # frozen ? previous-day load, no farms anyway
                if _is_fixed_vol_block(block):
                    continue   # frozen ? explicit vol_partial on every dest
                for f_idx in range(len(block["rows"])):
                    flat.append((s_idx, b_idx, f_idx))
        if not flat:
            return copy.deepcopy(state), []
        n_remove = min(n_remove, len(flat))
        chosen   = set(map(tuple, random.sample(flat, n_remove)))

        removed   = []
        new_state = []
        for s_idx, (sname, blocks) in enumerate(state):
            new_blocks = []
            for b_idx, block in enumerate(blocks):
                keep = []
                for f_idx, farm in enumerate(block["rows"]):
                    if (s_idx, b_idx, f_idx) in chosen:
                        removed.append((s_idx, b_idx, copy.deepcopy(farm)))
                    else:
                        keep.append(copy.deepcopy(farm))
                new_blocks.append(dict(block, rows=keep))
            new_state.append((sname, new_blocks))
        return new_state, removed

    def _destroy_worst(self, state, n_remove):
        """Remove the n farms whose individual removal saves the most sheet-km.
        Preload blocks are completely frozen; holdover-block farms are fair game."""
        savings = []
        for s_idx, (sname, blocks) in enumerate(state):
            entry = self.cache[self.fname].get(sname, {})
            st    = entry.get("start_time") if isinstance(entry, dict) else None
            base  = _sheet_cost(blocks, self.dm, st, self.cfg)
            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block):
                    continue   # frozen
                if _is_fixed_vol_block(block):
                    continue   # frozen ? explicit vol_partial on every dest
                for f_idx in range(len(block["rows"])):
                    trial = copy.deepcopy(blocks)
                    trial[b_idx]["rows"].pop(f_idx)
                    new_c = _sheet_cost(trial, self.dm, st, self.cfg)
                    savings.append((base - new_c, s_idx, b_idx, f_idx))
        savings.sort(reverse=True)

        # Collect top-n by (s_idx, b_idx, f_idx), at most one f_idx per block
        # to avoid index-shifting conflicts
        seen    = {}
        for saving, s_idx, b_idx, f_idx in savings:
            key = (s_idx, b_idx)
            seen.setdefault(key, []).append(f_idx)
            if sum(len(v) for v in seen.values()) >= n_remove:
                break

        # Now deepcopy state and remove by index (high-to-low to preserve indices)
        new_state = copy.deepcopy(state)
        removed   = []
        for (s_idx, b_idx), idxs in seen.items():
            block = new_state[s_idx][1][b_idx]
            for f_idx in sorted(set(idxs), reverse=True):
                if f_idx < len(block["rows"]):
                    removed.append((s_idx, b_idx, block["rows"].pop(f_idx)))
        return new_state, removed

    def _best_insert_cost(self, blocks, farm, dm, shift_start=None,
                          baseline=None):
        """
        Return (b_idx, pos, marginal_cost) for cheapest insertion of farm.

        Computes baseline _sheet_cost once (or accepts a pre-computed one),
        then evaluates only the delta for each candidate position.  Saves one
        _sheet_cost call per (farm, sheet) pair when the caller caches baselines
        across multiple farms on the same sheet.

        Skips preload blocks and fixed-vol blocks.
        """
        if baseline is None:
            baseline = _sheet_cost(blocks, dm, shift_start, self.cfg)

        best_b, best_pos, best_cost = None, None, float("inf")

        for b_idx, block in enumerate(blocks):
            if _is_preload_block(block):
                continue
            if _is_fixed_vol_block(block):
                continue
            rows = block["rows"]
            for pos in range(len(rows) + 1):
                trial_rows   = rows[:pos] + [farm] + rows[pos:]
                trial_block  = dict(block, rows=trial_rows)
                trial_blocks = blocks[:b_idx] + [trial_block] + blocks[b_idx+1:]
                delta = _sheet_cost(trial_blocks, dm, shift_start, self.cfg) - baseline
                if delta < best_cost:
                    best_b, best_pos, best_cost = b_idx, pos, delta

        return best_b, best_pos, best_cost

    def _repair_best(self, state, removed):
        """Greedy best-insertion: each farm -> cheapest slot across whole group.

        Sheets with no start_time are frozen ? reliable arrival estimates
        require a real shift start; without one the sheet is skipped entirely.
        Farms are inserted most-constrained-first (highest minimum insertion
        cost) so hard-to-place farms claim their preferred slot before easier
        ones, reducing cascades of suboptimal placements.
        """
        state = copy.deepcopy(state)
        # Per-sheet start times. None -> no start time -> sheet is frozen.
        start_map = {}
        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            start_map[sname] = entry.get("start_time") if isinstance(entry, dict) else None

        # Only sheets with a real start time are eligible for insertion.
        eligible = [(s_idx, sname, blocks)
                    for s_idx, (sname, blocks) in enumerate(state)
                    if start_map.get(sname) is not None]

        # Pre-compute baseline _sheet_cost for each eligible sheet.
        # This saves one _sheet_cost call per (farm x sheet) combination ?
        # with 20 removed farms and 27 sheets that's 540 saved calls per repair.
        baseline_cache = {
            sname: _sheet_cost(blocks, self.dm, start_map[sname], self.cfg)
            for _, sname, blocks in eligible
        }

        def _min_cost(item):
            _, _, farm = item
            best_c = float("inf")
            for _, sname, blocks in eligible:
                b, pos, c = self._best_insert_cost(blocks, farm, self.dm,
                                                    shift_start=start_map[sname],
                                                    baseline=baseline_cache[sname])
                if b is not None and c < best_c:
                    best_c = c
            return best_c

        ordered = sorted(removed, key=_min_cost, reverse=True)
        cross_route = 0
        for s_hint, b_hint, farm in ordered:
            best_s, best_b, best_pos, best_c = None, None, None, float("inf")
            for s_idx, sname, blocks in eligible:
                b_idx, pos, c = self._best_insert_cost(blocks, farm, self.dm,
                                                        shift_start=start_map[sname],
                                                        baseline=baseline_cache[sname])
                if b_idx is not None and c < best_c:
                    best_s, best_b, best_pos, best_c = s_idx, b_idx, pos, c
            if best_s is not None:
                if best_s != s_hint:
                    cross_route += 1
                state[best_s][1][best_b]["rows"].insert(best_pos, copy.deepcopy(farm))
                # Invalidate baseline for the modified sheet so subsequent farms
                # on the same sheet see the updated cost.
                sname_mod = state[best_s][0]
                baseline_cache[sname_mod] = _sheet_cost(
                    state[best_s][1], self.dm, start_map[sname_mod], self.cfg)
        return state, cross_route

    def _repair_regret(self, state, removed, k=2):
        """k-regret insertion: always insert the farm with the highest regret first.

        Sheets with no start_time are frozen and excluded from candidate slots.
        """
        state   = copy.deepcopy(state)
        pending = list(removed)
        # Per-sheet start times. None -> frozen.
        start_map = {}
        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            start_map[sname] = entry.get("start_time") if isinstance(entry, dict) else None

        # Only sheets with a real start time are eligible for insertion.
        eligible_idxs = {s_idx for s_idx, (sname, _) in enumerate(state)
                         if start_map.get(sname) is not None}

        # Pre-compute baseline _sheet_cost for each eligible sheet.
        baseline_cache = {
            sname: _sheet_cost(blocks, self.dm, start_map[sname], self.cfg)
            for s_idx, (sname, blocks) in enumerate(state)
            if s_idx in eligible_idxs
        }

        cross_route = 0
        while pending:
            best_farm_i = None
            best_regret = -float("inf")
            best_slot   = None
            for i, (s_hint, b_hint, farm) in enumerate(pending):
                slot_costs = []
                for s_idx, (sname, blocks) in enumerate(state):
                    if s_idx not in eligible_idxs:
                        continue   # sheet has no start time ? frozen
                    b_idx, pos, c = self._best_insert_cost(blocks, farm, self.dm,
                                                            shift_start=start_map[sname],
                                                            baseline=baseline_cache[sname])
                    if b_idx is not None:
                        slot_costs.append((c, s_idx, b_idx, pos))
                slot_costs.sort()
                if not slot_costs:
                    continue
                regret = (slot_costs[1][0] - slot_costs[0][0]) if len(slot_costs) >= k else 0.0
                if regret > best_regret:
                    best_regret = regret
                    best_farm_i = i
                    best_slot   = slot_costs[0]
            if best_farm_i is None:
                break
            _, s_idx, b_idx, pos = best_slot
            s_hint, b_hint, farm = pending.pop(best_farm_i)
            if s_idx != s_hint:
                cross_route += 1
            state[s_idx][1][b_idx]["rows"].insert(pos, copy.deepcopy(farm))
            # Invalidate baseline for the modified sheet
            sname_mod = state[s_idx][0]
            baseline_cache[sname_mod] = _sheet_cost(
                state[s_idx][1], self.dm, start_map[sname_mod], self.cfg)
        return state, cross_route

    # ??????????????????????????????????????????????????????????????????????????
    # DEST destroy / repair
    # ??????????????????????????????????????????????????????????????????????????

    @staticmethod
    def _acc_block_vols(block, vols_dict):
        """Accumulate delivered litres per dest-key from one block into vols_dict."""
        dests_b  = block.get("dests") or []
        farm_vol = sum((r.get("prior_vol") or 0) for r in block["rows"]
                       if isinstance(r.get("prior_vol"), (int, float)))
        already  = 0.0
        for d in dests_b:
            dk  = d.get("key") or "?"
            vp  = d.get("vol_partial")
            rem = max(0.0, farm_vol - already)
            off = min(float(vp), rem) if vp is not None else rem
            already += off
            vols_dict[dk] = vols_dict.get(dk, 0.0) + off

    def _destroy_dest(self, state, n_strip, orig_dest_vols):
        """
        Randomly select n_strip blocks and strip their entire dest list.
        Works even when every block only has a single destination.
        Returns (new_state, stripped: [(s_idx, b_idx, dests_list), ...]).
        """
        state = copy.deepcopy(state)
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sname, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if (block.get("dests") or block.get("dest_key"))
               and not _is_holdover_block(block)    # yard-for dest stays fixed
               and not _is_preload_block(block)      # preload block fully frozen
               and not _is_fixed_vol_block(block)    # fixed vol_partial ? operational instruction
        ]
        if not candidates:
            return state, []
        n_strip  = min(n_strip, len(candidates))
        chosen   = random.sample(candidates, n_strip)
        stripped = []
        for s_idx, b_idx in chosen:
            block = state[s_idx][1][b_idx]
            dests = block.get("dests") or []
            if not dests:
                dk = block.get("dest_key",""); dn = block.get("dest_name","")
                dests = [{"key": dk, "name": dn, "vol_partial": None}] if dk else []
            stripped.append((s_idx, b_idx, copy.deepcopy(dests)))
            block["dests"]     = []
            block["dest_key"]  = ""
            block["dest_name"] = ""
        return state, stripped

    def _repair_dest(self, state, stripped, dest_catalogue, orig_dest_vols):
        """
        Re-assign stripped dest lists to blocks using regret-order greedy.
        stripped: [(s_idx, b_idx, dests_list), ...]

        For each stripped item we find the cheapest block (any block, not just
        original) to attach the dest list to. This lets the solver redistribute
        processors across routes.

        Cost of attaching dests_list to block (s2, b2) =
          route_km(block with dests_list appended) + volume_penalty_delta
        """
        state   = copy.deepcopy(state)
        pending = list(stripped)   # [(s_idx, b_idx, dests_list)]

        # running vol totals for all already-assigned dests
        cur_vols = {}
        for _sname, blocks in state:
            for block in blocks:
                ALNSSolver._acc_block_vols(block, cur_vols)

        vol_tol  = self.cfg.get("vol_tol", 0.15)
        vol_rate = self.cfg.get("vol_penalty", 1.0)

        def _block_origin(s2, b2):
            """True origin of block (s2,b2): VEDDER if first, else last dest of prev block."""
            if b2 == 0:
                return "VEDDER"
            prev = state[s2][1][b2 - 1]
            prev_dests = prev.get("dests") or []
            if prev_dests:
                return prev_dests[-1].get("key","") or "VEDDER"
            return prev.get("dest_key","") or "VEDDER"

        plant_windows       = self.cfg.get("plant_windows", {})
        plant_win_rate      = self.cfg.get("plant_win_penalty", 500.0)
        plant_margin_mins   = self.cfg.get("plant_win_margin_mins", 30.0)
        plant_margin_rate   = self.cfg.get("plant_win_margin_rate", plant_win_rate * 0.5)

        def _attach_cost(s2, b2, dests_list):
            """Cost of giving dests_list to block (s2,b2), using correct block origin."""
            block     = state[s2][1][b2]

            # Guard: if dests_list is fully capped (every dest has fixed vol_partial,
            # i.e. no catch-all remainder), the total deliverable volume is fixed at
            # the sum of all vol_partials.  If the block's farm_vol exceeds that cap,
            # assigning this dest list here silently drops the overflow ? identical to
            # the _is_fixed_vol_block problem.  Apply a prohibitive overflow penalty.
            if dests_list and dests_list[-1].get("vol_partial") is not None:
                dest_cap = sum(float(d.get("vol_partial") or 0) for d in dests_list)
                farm_vol = sum((r.get("prior_vol") or 0) for r in block["rows"]
                               if isinstance(r.get("prior_vol"), (int, float)))
                overflow = max(0.0, farm_vol - dest_cap)
                if overflow > 0:
                    cap_pen_rate = self.cfg.get("cap_penalty", 2.0)
                    return overflow * cap_pen_rate * 1000  # prohibitive

            farm_keys = [r["irma"] for r in block["rows"]]
            dest_keys = [d.get("key","") for d in dests_list if d.get("key")]
            blocks_in_sheet = state[s2][1]
            is_last   = (b2 == len(blocks_in_sheet) - 1)
            origin    = _block_origin(s2, b2)
            stops     = [origin] + farm_keys + dest_keys
            if is_last:
                stops = stops + ["VEDDER"]
            leg_km    = sum(lookup(self.dm, stops[i], stops[i+1]) or 0.0
                            for i in range(len(stops)-1))
            # volume penalty delta
            farm_vol  = sum((r.get("prior_vol") or 0) for r in block["rows"]
                            if isinstance(r.get("prior_vol"), (int, float)))
            already   = 0.0
            vol_pen   = 0.0
            for d in dests_list:
                dk  = d.get("key","")
                vp  = d.get("vol_partial")
                rem = max(0.0, farm_vol - already)
                off = min(float(vp), rem) if vp is not None else rem
                already += off
                new_v = cur_vols.get(dk, 0.0) + off
                orig  = orig_dest_vols.get(dk, 0.0)
                lo = orig * (1.0 - vol_tol); hi = orig * (1.0 + vol_tol)
                vol_pen += (max(0.0, lo - new_v) + max(0.0, new_v - hi)) * vol_rate

            # Plant window penalty ? skip for yard-for destinations (24/7 parking)
            win_pen = 0.0
            is_yard_dests = all("yard for" in (d.get("name","") or "").lower()
                                for d in dests_list if d.get("name"))
            if plant_windows and not is_yard_dests:
                sname_s2 = state[s2][0]
                entry_s2 = self.cache.get(self.fname, {}).get(sname_s2, {})
                st_s2    = entry_s2.get("start_time") if isinstance(entry_s2, dict) else None
                if st_s2:
                    # Estimate time to reach first dest: drive from origin through
                    # all farms then to each dest in sequence.
                    cursor = datetime.combine(date.today(), st_s2)
                    for i in range(len(stops) - 1):
                        leg = lookup(self.dm, stops[i], stops[i+1])
                        if leg is None:
                            break
                        cursor += timedelta(minutes=(leg / DRIVE_SPEED_KMH) * 60.0)
                        stop = stops[i + 1]
                        dk_stop = normalise_key(stop)
                        window = plant_windows.get(dk_stop)
                        if window is None:
                            continue
                        arr_t = cursor.time()
                        if not time_in_window(arr_t, window[0], window[1]):
                            open_t = parse_hhmm(window[0])
                            if open_t:
                                open_dt = datetime.combine(date.today(), open_t)
                                if open_t <= arr_t:
                                    open_dt += timedelta(days=1)
                                wait_h = (open_dt - cursor).total_seconds() / 3600.0
                            else:
                                wait_h = 1.0
                            win_pen += wait_h * plant_win_rate
                        elif plant_margin_mins > 0 and plant_margin_rate > 0:
                            close_t = parse_hhmm(window[1])
                            if close_t:
                                close_dt = datetime.combine(date.today(), close_t)
                                open_t2 = parse_hhmm(window[0])
                                if open_t2 and close_t < open_t2:
                                    close_dt += timedelta(days=1)
                                mins_to_close = (close_dt - cursor).total_seconds() / 60.0
                                if 0 < mins_to_close < plant_margin_mins:
                                    depth = (plant_margin_mins - mins_to_close) / plant_margin_mins
                                    win_pen += depth * plant_margin_rate * (plant_margin_mins / 60.0)

            return leg_km + vol_pen + win_pen

        # flat list of all (s_idx, b_idx) slots
        all_slots = [(s_idx, b_idx)
                     for s_idx, (_sname, blocks) in enumerate(state)
                     for b_idx in range(len(blocks))]

        while pending:
            best_pi   = None
            best_regret = -float("inf")
            best_slot = None

            for pi, (_, _, dests_list) in enumerate(pending):
                costs = sorted(
                    ((_attach_cost(s2, b2, dests_list), (s2, b2))
                     for s2, b2 in all_slots
                     if not (state[s2][1][b2].get("dests") or
                             state[s2][1][b2].get("dest_key"))),
                    key=lambda x: x[0]
                )
                if not costs:
                    # Every slot already has dests ? just pick cheapest overall
                    costs = sorted(
                        ((_attach_cost(s2, b2, dests_list), (s2, b2))
                         for s2, b2 in all_slots),
                        key=lambda x: x[0]
                    )
                if not costs:
                    continue
                regret = (costs[1][0] - costs[0][0]) if len(costs) >= 2 else 0.0
                if regret > best_regret:
                    best_regret = regret; best_pi = pi; best_slot = costs[0][1]

            if best_pi is None:
                break

            s_src, b_src, dests_list = pending.pop(best_pi)
            s2, b2 = best_slot
            block = state[s2][1][b2]
            # Subtract any volumes the slot already had before overwriting
            old_vols = {}
            ALNSSolver._acc_block_vols(block, old_vols)
            for dk, v in old_vols.items():
                cur_vols[dk] = cur_vols.get(dk, 0.0) - v
            block["dests"]     = copy.deepcopy(dests_list)
            block["dest_key"]  = dests_list[0].get("key","")  if dests_list else ""
            block["dest_name"] = dests_list[0].get("name","") if dests_list else ""
            # Add new volumes
            ALNSSolver._acc_block_vols(block, cur_vols)

        # Any still-pending items (no empty slots): put back on original block
        for s_src, b_src, dests_list in pending:
            block = state[s_src][1][b_src]
            block["dests"]     = copy.deepcopy(dests_list)
            block["dest_key"]  = dests_list[0].get("key","")  if dests_list else ""
            block["dest_name"] = dests_list[0].get("name","") if dests_list else ""

        return state

    # ??????????????????????????????????????????????????????????????????????????
    # DEST permutation & volume-split operators
    # ??????????????????????????????????????????????????????????????????????????

    def _shuffle_dests(self, state, orig_dest_vols):
        """
        Pick one random multi-dest block and try all permutations of its dest
        list.  Returns the state with the best permutation applied to that
        block only ? SA in the main loop decides whether to accept.
        Single-block change keeps the move small and well-defined for SA.
        """
        from itertools import permutations as _perms
        vol_tol  = self.cfg.get("vol_tol", 0.15)
        vol_rate = self.cfg.get("vol_penalty", 1.0)

        # Collect all blocks with >=2 dests
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sn, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if len(block.get("dests") or []) >= 2
               and len(block.get("dests") or []) <= 5
        ]
        if not candidates:
            return copy.deepcopy(state)

        s_idx, b_idx = random.choice(candidates)
        state = copy.deepcopy(state)
        blocks   = state[s_idx][1]
        block    = blocks[b_idx]
        dests    = block["dests"]
        farm_keys = [r["irma"] for r in block["rows"]]
        farm_vol  = sum((r.get("prior_vol") or 0) for r in block["rows"]
                        if isinstance(r.get("prior_vol"), (int, float)))
        is_last   = (b_idx == len(blocks) - 1)
        origin    = ("VEDDER" if b_idx == 0 else
                     ((blocks[b_idx-1].get("dests") or [{}])[-1].get("key","") or "VEDDER"))

        # Group vol snapshot excluding this block
        base_vols = {}
        for si2, (_sn2, blks2) in enumerate(state):
            for bi2, blk2 in enumerate(blks2):
                if si2 == s_idx and bi2 == b_idx:
                    continue
                ALNSSolver._acc_block_vols(blk2, base_vols)

        def _perm_cost(perm):
            dest_keys = [d.get("key","") for d in perm if d.get("key")]
            stops = [origin] + farm_keys + dest_keys + (["VEDDER"] if is_last else [])
            km = sum(lookup(self.dm, stops[i], stops[i+1]) or 0.0
                     for i in range(len(stops)-1))
            vol_pen = 0.0; already = 0.0; vols = dict(base_vols)
            for d in perm:
                dk = d.get("key",""); vp = d.get("vol_partial")
                rem = max(0.0, farm_vol - already)
                off = min(float(vp), rem) if vp is not None else rem
                already += off
                new_v = vols.get(dk, 0.0) + off
                orig_v = orig_dest_vols.get(dk, 0.0)
                lo = orig_v * (1.0 - vol_tol); hi = orig_v * (1.0 + vol_tol)
                vol_pen += (max(0.0, lo-new_v) + max(0.0, new_v-hi)) * vol_rate
            return km + vol_pen

        best_perm = min(_perms(dests), key=_perm_cost)
        block["dests"]     = list(best_perm)
        block["dest_key"]  = block["dests"][0].get("key","")
        block["dest_name"] = block["dests"][0].get("name","")
        return state

    # ??????????????????????????????????????????????????????????????????????????
    # Intra-block resequencing operators  (2-opt and Or-opt)
    # ??????????????????????????????????????????????????????????????????????????

    def _two_opt_single(self, state):
        """Pick one random eligible block; find and apply its best 2-opt swap.

        2-opt reverses a sub-sequence of the farm list.  This reorders farms
        already assigned to a block, which the destroy/repair operators cannot
        do.  SA in the main loop decides whether to accept the result.

        Returns a candidate state (may be identical to input if no swap helps).
        """
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sn, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if not _is_preload_block(block) and not _is_fixed_vol_block(block)
            and len(block["rows"]) >= 2
        ]
        if not candidates:
            return copy.deepcopy(state)

        s_idx, b_idx = random.choice(candidates)
        state = copy.deepcopy(state)
        blocks = state[s_idx][1]
        block  = blocks[b_idx]
        rows   = block["rows"]
        n      = len(rows)
        origin = "VEDDER" if b_idx == 0 else (
            _block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"
        )

        base_cost = _route_km_simple(block, self.dm, origin=origin)
        best_rows = rows[:]
        best_cost = base_cost

        for i in range(n - 1):
            for j in range(i + 1, n):
                new_rows = rows[:i] + rows[i:j + 1][::-1] + rows[j + 1:]
                trial    = dict(block, rows=new_rows)
                c = _route_km_simple(trial, self.dm, origin=origin)
                if c < best_cost:
                    best_cost = c
                    best_rows = new_rows[:]

        state[s_idx][1][b_idx] = dict(block, rows=best_rows)
        return state

    def _or_opt_single(self, state, seg_len=2):
        """Pick one random eligible block; relocate the best segment of seg_len
        consecutive farms to a cheaper position in the same block.

        Or-opt with seg_len=2 moves pairs of farms; with seg_len=1 it is a
        simple relocation.  SA in the main loop decides whether to accept.
        """
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sn, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if not _is_preload_block(block) and not _is_fixed_vol_block(block)
            and len(block["rows"]) >= seg_len + 1
        ]
        if not candidates:
            return copy.deepcopy(state)

        s_idx, b_idx = random.choice(candidates)
        state = copy.deepcopy(state)
        blocks = state[s_idx][1]
        block  = blocks[b_idx]
        rows   = block["rows"]
        n      = len(rows)
        origin = "VEDDER" if b_idx == 0 else (
            _block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"
        )

        base_cost = _route_km_simple(block, self.dm, origin=origin)
        best_rows = rows[:]
        best_cost = base_cost

        for i in range(n - seg_len + 1):
            seg  = rows[i:i + seg_len]
            rest = rows[:i] + rows[i + seg_len:]
            for j in range(len(rest) + 1):
                new_rows = rest[:j] + seg + rest[j:]
                trial    = dict(block, rows=new_rows)
                c = _route_km_simple(trial, self.dm, origin=origin)
                if c < best_cost:
                    best_cost = c
                    best_rows = new_rows[:]

        state[s_idx][1][b_idx] = dict(block, rows=best_rows)
        return state

    def _two_opt_all_blocks(self, state):
        """Exhaustive 2-opt polish applied to every block until convergence.

        Called once after the ALNS loop on best_state.  Guaranteed to return a
        solution at least as good as the input; never worsens.
        """
        state   = copy.deepcopy(state)
        changed = True
        while changed:
            changed = False
            for s_idx, (sname, blocks) in enumerate(state):
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block):
                        continue
                    if _is_fixed_vol_block(block):
                        continue
                    rows   = block["rows"]
                    n      = len(rows)
                    if n < 2:
                        continue
                    origin = "VEDDER" if b_idx == 0 else (
                        _block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"
                    )
                    improved = True
                    while improved:
                        improved = False
                        base_cost = _route_km_simple(
                            dict(block, rows=rows), self.dm, origin=origin
                        )
                        for i in range(n - 1):
                            for j in range(i + 1, n):
                                new_rows = rows[:i] + rows[i:j + 1][::-1] + rows[j + 1:]
                                c = _route_km_simple(
                                    dict(block, rows=new_rows), self.dm, origin=origin
                                )
                                if c < base_cost - 0.001:
                                    rows      = new_rows
                                    base_cost = c
                                    improved  = True
                                    changed   = True
                    state[s_idx][1][b_idx] = dict(block, rows=rows)
                    # Keep local blocks reference in sync for subsequent b_idx origin lookups
                    blocks[b_idx] = state[s_idx][1][b_idx]
        return state

    # ??????????????????????????????????????????????????????????????????????????
    # Top-level roulette
    # ??????????????????????????????????????????????????????????????????????????

    def _roulette(self, scores, min_prob=0.0, floors=None, ceilings=None):
        """
        Weighted random selection from scores dict.

        floors (dict, optional): per-key minimum probability.  Takes precedence
          over min_prob.  Any key not in floors falls back to min_prob.

        ceilings (dict, optional): per-key maximum probability.  Keys exceeding
          their ceiling are clamped and the surplus is redistributed to keys
          below their ceiling.  Applied after floors.

        Both floors and ceilings can be combined.  The floor/ceiling pass
        iterates until stable to handle cascading interactions.
        """
        keys   = list(scores.keys())
        n      = len(keys)
        raw    = [max(scores[k], 1e-9) for k in keys]
        total  = sum(raw)
        probs  = [v / total for v in raw]

        # Build per-key floor and ceiling arrays
        key_floors   = [floors.get(k, min_prob) for k in keys]   if floors   else [min_prob] * n
        key_ceilings = [ceilings.get(k, 1.0)    for k in keys]   if ceilings else [1.0]      * n

        _apply_prob_floors_ceilings(probs, key_floors, key_ceilings)

        r   = random.random()
        cum = 0.0
        for k, p in zip(keys, probs):
            cum += p
            if r <= cum:
                return k
        return keys[-1]

    # ??????????????????????????????????????????????????????????????????????????
    # Main ALNS loop
    # ??????????????????????????????????????????????????????????????????????????

    def _solve_group(self, colour, sheets, total_iters, iter_offset):
        if not sheets:
            return {}
        try:
            return self._solve_group_inner(colour, sheets, total_iters, iter_offset)
        except Exception as ex:
            import traceback
            tb = traceback.format_exc()
            self.log.emit(f"\n[{colour}] SOLVER CRASHED:\n{tb}")
            return {}

    def _solve_group_inner(self, colour, sheets, total_iters, iter_offset):

        dest_catalogue = _group_dest_catalogue(sheets)

        # initial state ? start from mod_blocks if available (preserves _mwo
        # and _orig_arr flags set by the user); fall back to cache blocks.
        def _initial_blocks(sname, entry):
            mod_key = (self.fname, sname)
            if mod_key in self.sheet_mods:
                return copy.deepcopy(self.sheet_mods[mod_key])
            return copy.deepcopy(entry.get("blocks", []))

        # Filter out sheets that must not be touched by the solver FIRST,
        # then build orig_dest_vols only from sheets that are actually in the
        # solver's control.  Including skip/locked sheets in the target inflates
        # it with volumes the solver can never deliver, making targets impossible
        # and causing the vol penalty to dominate and never improve.
        all_skip = SOLVER_SKIP_SHEETS | self.locked_sheets
        locked_in_group = [sn for sn, _ in sheets if sn.strip() in self.locked_sheets]
        sheets = [(sn, e) for sn, e in sheets
                  if sn.strip() not in all_skip]
        if locked_in_group:
            self.log.emit(f"  [{colour}] Locked (held constant): {locked_in_group}")
        if not sheets:
            return {}

        # -- Strip zero-vol farms before solving -------------------------------
        # Farms with prior_vol == 0 (or falsy numeric) contribute nothing to
        # collection volume and skew the solver's cost function.  They are
        # removed from the working state, the solver ignores them entirely, and
        # -- Zero-vol farm pairing ---------------------------------------------
        # Farms with prior_vol == 0 that share an IRMA with a non-zero farm in
        # the same block are a "paired set" (e.g. T1 / T2 trailers at the same
        # farm).  The solver must treat the pair as an atomic unit ? if one moves,
        # both move together.  Zero-vol members contribute 0 minutes to the
        # schedule (arr == dep, no setup, no pump, no milking wait).
        #
        # Strategy: strip zero-vol members before solving; after solving,
        # re-insert them adjacent to their non-zero partner in the same block.
        # The non-zero partner's block may have changed (solver moved it);
        # we track by _uid so we always find it in the solved state.
        #
        # zero_vol_farms: list of (sname, b_idx, partner_uid, zero_farm_dict)
        # where partner_uid is the _uid of the non-zero farm with the same IRMA.
        zero_vol_farms = []   # (sname, b_idx, partner_uid, zero_farm)

        def _find_partner_uid(blocks, b_idx, zero_irma, zero_uid):
            """Return the _uid of the non-zero farm with the same IRMA in this block."""
            for farm in blocks[b_idx].get("rows", []):
                if farm.get("irma","") == zero_irma and farm.get("_uid") != zero_uid:
                    pv = farm.get("prior_vol")
                    if not (isinstance(pv, (int, float)) and pv == 0):
                        return farm.get("_uid")
            return None

        def _strip_zero_vol(sname, blocks):
            """Strip zero-vol farms that have a non-zero partner with the same IRMA."""
            stripped = []
            for b_idx, block in enumerate(blocks):
                irma_counts = {}
                for farm in block.get("rows", []):
                    irma_counts[farm.get("irma","")] = \
                        irma_counts.get(farm.get("irma",""), 0) + 1
                new_rows = []
                for farm in block.get("rows", []):
                    pv   = farm.get("prior_vol")
                    irma = farm.get("irma","")
                    # Only sideline if: vol==0 AND the same IRMA appears more
                    # than once (i.e. there is a non-zero partner in this block)
                    if (isinstance(pv, (int, float)) and pv == 0
                            and irma_counts.get(irma, 0) > 1):
                        partner_uid = _find_partner_uid(blocks, b_idx,
                                                        irma, farm.get("_uid"))
                        zero_vol_farms.append(
                            (sname, b_idx, partner_uid, copy.deepcopy(farm)))
                    else:
                        new_rows.append(farm)
                stripped.append(dict(block, rows=new_rows))
            return stripped

        def _reinsert_zero_vol(result):
            """Re-insert zero-vol farms adjacent to their partner in the solved state."""
            if not zero_vol_farms:
                return result
            # Build a uid -> (sname, b_idx, f_idx) map across the solved state
            uid_loc = {}
            for sname, blocks in result.items():
                for b_idx, block in enumerate(blocks):
                    for f_idx, farm in enumerate(block.get("rows", [])):
                        uid = farm.get("_uid")
                        if uid:
                            uid_loc[uid] = (sname, b_idx, f_idx)

            for _orig_sname, _orig_b_idx, partner_uid, zero_farm in zero_vol_farms:
                if partner_uid and partner_uid in uid_loc:
                    sname, b_idx, f_idx = uid_loc[partner_uid]
                    # Insert immediately after partner
                    result[sname][b_idx]["rows"].insert(f_idx + 1, zero_farm)
                    # Update uid_loc so subsequent insertions see the new indices
                    for uid, (s, b, fi) in list(uid_loc.items()):
                        if s == sname and b == b_idx and fi > f_idx:
                            uid_loc[uid] = (s, b, fi + 1)
                    uid_loc[zero_farm.get("_uid","")] = (sname, b_idx, f_idx + 1)
                else:
                    # Partner not found in solved state (shouldn't happen) ?
                    # fall back to original sheet/block
                    sname = _orig_sname
                    if sname in result and _orig_b_idx < len(result[sname]):
                        result[sname][_orig_b_idx]["rows"].append(zero_farm)
            return result

        # Build volume targets from the non-skipped sheets only
        orig_dest_vols = {}
        for _sname, entry in sheets:
            for block in entry.get("blocks", []):
                for dk, off in _block_dest_offloads(block).items():
                    orig_dest_vols[dk] = orig_dest_vols.get(dk, 0.0) + off

        state = [(sname, _strip_zero_vol(sname, _initial_blocks(sname, entry)))
                 for sname, entry in sheets]

        if zero_vol_farms:
            self.log.emit(
                f"  [{colour}] {len(zero_vol_farms)} zero-vol farm(s) held with "
                f"their partner (solver treats pairs as atomic): "
                f"{', '.join(f['irma'] for _, _, _, f in zero_vol_farms)}")

        # Optimize split positions for partial-dropoff dests on initial state
        if self.cfg.get("split_opt", False):
            for sname, blocks in state:
                entry = self.cache.get(self.fname, {}).get(sname, {})
                st    = entry.get("start_time") if isinstance(entry, dict) else None
                if st:
                    _optimize_split_positions(blocks, self.dm, st, self.cfg)

        cfg_no_win = dict(self.cfg, plant_win_penalty=0.0, plant_windows={},
                          win_miss_penalty=0.0, cap_penalty=0.0)

        # Compute frozen block cost offset BEFORE using it in best_cost
        frozen_cost_offset = 0.0
        frozen_cost_offset_no_win = 0.0
        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            st    = entry.get("start_time") if isinstance(entry, dict) else None
            for block in blocks:
                if _is_preload_block(block) or _is_fixed_vol_block(block):
                    frozen_cost_offset        += _sheet_cost([block], self.dm, st, self.cfg)
                    frozen_cost_offset_no_win += _sheet_cost([block], self.dm, st, cfg_no_win)

        best_state = copy.deepcopy(state)
        cur_sheet_cache  = self._make_sheet_cost_cache(state)
        best_cost  = sum(cur_sheet_cache.values()) + _group_vol_penalty(state, orig_dest_vols, self.cfg) - frozen_cost_offset
        cur_cost   = best_cost
        best_sheet_cache = dict(cur_sheet_cache)

        cost_no_win = sum(
            _sheet_cost(blocks, self.dm,
                        (self.cache.get(self.fname, {}).get(sname, {}) or {}).get("start_time"),
                        cfg_no_win)
            for sname, blocks in state
        ) + _group_vol_penalty(state, orig_dest_vols, self.cfg)
        cost_no_win -= frozen_cost_offset_no_win
        T0               = max(cost_no_win * 0.05, 10.0)
        T                = T0
        target_cool_frac = self.cfg.get("target_cool_frac", 0.001)
        if total_iters > 1 and target_cool_frac > 0:
            alpha = target_cool_frac ** (1.0 / total_iters)
        else:
            alpha = 0.9999

        # Move-type scores with hard probability floors.
        #
        # The adaptive roulette learns which moves improve cost and rewards them.
        # Without a floor, farm/combined moves get penalised whenever they fail
        # to beat the plant-window penalty (which dominates early on), causing
        # intra-block 2-opt/or-opt to dominate by iteration 200 ? exactly the
        # behaviour we observed.  Hard floors guarantee a minimum budget for
        # cross-route moves regardless of adaptive history.
        #
        # Floors (min_prob passed to _roulette):
        #   farm:     0.25  ? guaranteed 25% of iterations move farms cross-route
        #   combined: 0.15
        #   dest:     0.10
        #   2opt:     0.05
        #   or_opt:   0.05
        FARM_FLOOR     = 0.25
        COMBINED_FLOOR = 0.15
        DEST_FLOOR     = 0.10
        INTRA_FLOOR    = 0.05
        MOVE_FLOORS = {
            "farm":     FARM_FLOOR,
            "combined": COMBINED_FLOOR,
            "dest":     DEST_FLOOR,
            "2opt":     INTRA_FLOOR,
            "or_opt":   INTRA_FLOOR,
        }
        # Ceilings prevent intra-block ops from hoarding probability even when
        # their adaptive scores are high (they improve easily but do nothing
        # for cross-route optimisation).
        MOVE_CEILINGS = {
            "2opt":   0.12,
            "or_opt": 0.12,
        }
        OP_MIN_PROB   = 0.25   # floor for destroy/repair sub-roulette (2 options each)

        move_scores = {
            "farm":     3.0,   # primary cross-route move
            "combined": 3.0,   # cross-route + dest reassignment
            "dest":     1.5,   # dest-only reassignment
            "2opt":     0.5,   # intra-block only ? useful but deprioritised
            "or_opt":   0.5,
        }

        # Farm sub-operator scores
        d_scores = {"random": 1.0, "worst": 1.0}
        r_scores = {"best":   1.0, "regret": 1.0}

        DECAY    = 0.97
        REWARD   = 2.0
        ACCEPT_REWARD = 0.5
        seg_size = self.cfg.get("segment_size", 100)

        n_farms  = sum(len(b["rows"]) for _, blocks in state for b in blocks)
        n_blocks = sum(len(blocks)    for _, blocks in state)
        n_remove_dests = max(1, min(int(n_blocks * 0.20), 5))

        self.log.emit(
            f"[{colour}]  sheets={len(sheets)}  farms={n_farms}  "
            f"blocks={n_blocks}  cost0={best_cost:.1f}  (cost_no_win={cost_no_win:.1f})\n"
            f"          alpha={alpha:.6f}  T0={T0:.1f} (from cost_no_win)  "
            f"cool_target={target_cool_frac*100:.4f}%\n"
            f"          remove/move: farms=T-scaled[3->30]  dests={n_remove_dests}  "
            f"floors: farm={FARM_FLOOR:.0%} combined={COMBINED_FLOOR:.0%} "
            f"dest={DEST_FLOOR:.0%} intra={INTRA_FLOOR:.0%}"
        )

        iters_no_improve  = 0
        diag_accepted  = {k: 0 for k in move_scores}
        diag_tried     = {k: 0 for k in move_scores}
        diag_cross_route = 0

        for it in range(total_iters):
            if self._stop:
                break

            # Decay all scores each segment ? just multiply, no hard floor
            # on raw scores since the probability floor handles balance.
            if it % seg_size == 0 and it > 0:
                for d in (move_scores, d_scores, r_scores):
                    for k in d:
                        d[k] = max(0.01, d[k] * DECAY)

            move_type = self._roulette(move_scores, floors=MOVE_FLOORS,
                                        ceilings=MOVE_CEILINGS)

            # n_remove scales with temperature: large disruptions when hot
            # (broad exploration), small when cold (fine-tuning).
            # Range: [3, 30] linearly interpolated by T/T0.
            t_frac = min(1.0, T / T0) if T0 > 0 else 0.0
            n_remove_farms = max(3, int(3 + (30 - 3) * t_frac))

            # Track which sheets are modified so we only recompute their costs.
            changed_sheets = set()

            # -- execute chosen move -------------------------------------------
            if move_type == "farm":
                d_op = self._roulette(d_scores, min_prob=OP_MIN_PROB)
                r_op = self._roulette(r_scores, min_prob=OP_MIN_PROB)
                if d_op == "random":
                    new_state, removed = self._destroy_random(state, n_remove_farms)
                else:
                    new_state, removed = self._destroy_worst(state, n_remove_farms)
                changed_sheets = {state[s][0] for s, _, _ in removed}
                if r_op == "best":
                    new_state, _cr = self._repair_best(new_state, removed)
                else:
                    new_state, _cr = self._repair_regret(new_state, removed)
                # Repair may place farms on any sheet ? track destinations too
                for s_idx, (sn, _) in enumerate(new_state):
                    if any(sn == state[s][0] for s, _, _ in removed):
                        changed_sheets.add(sn)
                    # Also detect sheets that gained farms from other routes
                    orig_count = sum(len(b["rows"]) for b in state[s_idx][1])
                    new_count  = sum(len(b["rows"]) for b in new_state[s_idx][1])
                    if orig_count != new_count:
                        changed_sheets.add(sn)
                diag_cross_route += _cr

            elif move_type == "dest":
                new_state, stripped = self._destroy_dest(state, n_remove_dests,
                                                          orig_dest_vols)
                new_state = self._repair_dest(new_state, stripped,
                                              dest_catalogue, orig_dest_vols)
                changed_sheets = {state[s][0] for s, _, _ in stripped}

            elif move_type == "2opt":
                new_state = self._two_opt_single(state)
                # 2-opt touches one block on one sheet ? detect by row count change
                for s_idx, (sn, blocks) in enumerate(new_state):
                    if blocks != state[s_idx][1]:
                        changed_sheets.add(sn)

            elif move_type == "or_opt":
                seg = 2 if random.random() < 0.7 else 3
                new_state = self._or_opt_single(state, seg_len=seg)
                for s_idx, (sn, blocks) in enumerate(new_state):
                    if blocks != state[s_idx][1]:
                        changed_sheets.add(sn)

            else:  # combined
                d_op = self._roulette(d_scores, min_prob=OP_MIN_PROB)
                r_op = self._roulette(r_scores, min_prob=OP_MIN_PROB)
                if d_op == "random":
                    new_state, removed = self._destroy_random(state, n_remove_farms)
                else:
                    new_state, removed = self._destroy_worst(state, n_remove_farms)
                changed_sheets = {state[s][0] for s, _, _ in removed}
                new_state, stripped = self._destroy_dest(new_state, n_remove_dests,
                                                          orig_dest_vols)
                changed_sheets |= {state[s][0] for s, _, _ in stripped}
                if r_op == "best":
                    new_state, _cr = self._repair_best(new_state, removed)
                else:
                    new_state, _cr = self._repair_regret(new_state, removed)
                for s_idx, (sn, _) in enumerate(new_state):
                    orig_count = sum(len(b["rows"]) for b in state[s_idx][1])
                    new_count  = sum(len(b["rows"]) for b in new_state[s_idx][1])
                    if orig_count != new_count:
                        changed_sheets.add(sn)
                diag_cross_route += _cr
                new_state = self._repair_dest(new_state, stripped,
                                              dest_catalogue, orig_dest_vols)

            # Recompute costs only for changed sheets; reuse cache for the rest.
            new_sheet_cache = dict(cur_sheet_cache)
            for s_idx, (sn, blocks) in enumerate(new_state):
                if sn in changed_sheets:
                    entry = self.cache[self.fname].get(sn, {})
                    st    = entry.get("start_time") if isinstance(entry, dict) else None
                    new_sheet_cache[sn] = _sheet_cost(blocks, self.dm, st, self.cfg)
            new_cost = sum(new_sheet_cache.values()) + _group_vol_penalty(new_state, orig_dest_vols, self.cfg) - frozen_cost_offset
            delta    = new_cost - cur_cost

            # SA acceptance ? compute probability explicitly for diagnostics
            if delta < 0:
                accept_prob = 1.0
            elif T > 1e-12:
                accept_prob = math.exp(-delta / T)
            else:
                accept_prob = 0.0
            accepted = accept_prob >= 1.0 or random.random() < accept_prob

            diag_tried[move_type] = diag_tried.get(move_type, 0) + 1

            if accepted:
                diag_accepted[move_type] = diag_accepted.get(move_type, 0) + 1
                state           = new_state
                cur_cost        = new_cost
                cur_sheet_cache = new_sheet_cache

                # Re-optimize split positions on any changed sheet (if enabled)
                split_changed = False
                if self.cfg.get("split_opt", False):
                    for s_idx, (sname, blocks) in enumerate(state):
                        if sname in changed_sheets:
                            entry = self.cache.get(self.fname, {}).get(sname, {})
                            st    = entry.get("start_time") if isinstance(entry, dict) else None
                            if st and _optimize_split_positions(blocks, self.dm, st, self.cfg):
                                split_changed = True
                                new_sheet_cache[sname] = _sheet_cost(blocks, self.dm, st, self.cfg)
                if split_changed:
                    cur_cost        = sum(new_sheet_cache.values()) + _group_vol_penalty(state, orig_dest_vols, self.cfg) - frozen_cost_offset
                    cur_sheet_cache = new_sheet_cache

                improved = cur_cost < best_cost
                if improved:
                    reward = REWARD
                    best_cost        = cur_cost
                    best_state       = copy.deepcopy(state)
                    best_sheet_cache = dict(cur_sheet_cache)
                    iters_no_improve = 0
                else:
                    reward = ACCEPT_REWARD
                    iters_no_improve += 1
                move_scores[move_type] = move_scores[move_type] * DECAY + reward
                if move_type in ("farm", "combined"):
                    d_scores[d_op] = d_scores[d_op] * DECAY + reward
                    r_scores[r_op] = r_scores[r_op] * DECAY + reward
            else:
                iters_no_improve += 1

            T *= alpha

            # -- Every 50 iterations: detailed diagnostic output --------------
            if it % 50 == 49:
                bd_best = _sheet_cost_breakdown_state(best_state, self.dm, self.cache, self.fname, self.cfg)
                vol_pen_d = _group_vol_penalty(best_state, orig_dest_vols, self.cfg)

                # Per-processor vol deviation (uses the same offload accounting
                # as the volume penalty so the displayed numbers always match).
                dest_vols_d = {}
                for _, blocks_d in best_state:
                    for block_d in blocks_d:
                        for dk_d, off_d in _block_dest_offloads(block_d).items():
                            dest_vols_d[dk_d] = dest_vols_d.get(dk_d, 0.0) + off_d

                vol_tol_d = self.cfg.get("vol_tol", 0.15)
                vol_lines = []
                for dk_d, orig_v in sorted(orig_dest_vols.items()):
                    cur_v = dest_vols_d.get(dk_d, 0.0)
                    pct = (cur_v / orig_v * 100) if orig_v else 0
                    flag = " (!)" if abs(cur_v - orig_v) > orig_v * vol_tol_d else ""
                    vol_lines.append(
                        f"      {dk_d}: {cur_v:>8,.0f} / {orig_v:>8,.0f} L  ({pct:5.1f}%){flag}")

                t50 = T * math.log(2)
                t10 = T * math.log(10)
                acc_parts = []
                for k in move_scores:
                    tried = diag_tried.get(k, 0)
                    acc   = diag_accepted.get(k, 0)
                    acc_parts.append(f"{k}={acc}/{tried}" if tried else f"{k}=?")

                self.log.emit(
                    f"  [{colour}] -- it={it+1} --  best={best_cost:.1f}\n"
                    f"    Cost breakdown: km={bd_best['km']:.1f}  shift={bd_best['shift']:.1f}"
                    f"  shift_pen={bd_best['overtime']:.1f}  milking={bd_best['milking']:.1f}"
                    f"  cap={bd_best['cap']:.1f}  vol_pen={vol_pen_d:.1f}\n"
                    f"    T={T:.2f}  50% accept if deltaZ<{t50:.1f}  10% if deltaZ<{t10:.1f}\n"
                    f"    Accepted/tried (last 50): {('  '.join(acc_parts))}\n"
                    f"    Cross-route placements (last 50 farm/combined moves): {diag_cross_route}\n"
                    f"    Processor volumes:\n" + "\n".join(vol_lines)
                )
                diag_accepted    = {k: 0 for k in move_scores}
                diag_tried       = {k: 0 for k in move_scores}
                diag_cross_route = 0

            # progress report
            if it % 25 == 0 or it == total_iters - 1:
                global_it = iter_offset + it
                total_all = iter_offset + total_iters
                self.progress.emit(
                    global_it, total_all,
                    f"{colour}: {it+1}/{total_iters}  "
                    f"best={best_cost:.1f}  cur={cur_cost:.1f}  T={T:.4f}"
                )
                if it % 100 == 0:
                    # Show effective probabilities using the same floor/ceiling
                    # redistribution _roulette applies, so the display matches.
                    _raw = [max(move_scores[k], 1e-9) for k in move_scores]
                    _tot = sum(_raw)
                    _keys = list(move_scores.keys())
                    _pv   = [m / _tot for m in _raw]
                    _fl   = [MOVE_FLOORS.get(k, 0.0) for k in _keys]
                    _cl   = [MOVE_CEILINGS.get(k, 1.0) for k in _keys]
                    _apply_prob_floors_ceilings(_pv, _fl, _cl)
                    _pt = sum(_pv)
                    prob_str = "  ".join(f"{k}={_pv[i]/_pt:.0%}"
                                         for i, k in enumerate(_keys))
                    self.log.emit(f"  [{colour}] it={it+1:4d}  best={best_cost:.1f}"
                                  f"  T={T:.4f}  prob: {prob_str}")

        def _state_vol(st):
            return sum(
                (f.get("prior_vol") or 0)
                for _, blks in st for b in blks
                for f in b["rows"]
                if isinstance(f.get("prior_vol"), (int, float))
            )
        input_farms = sum(len(b["rows"]) for _, blks in state for b in blks)
        input_vol   = _state_vol(state)

        output_farms = sum(len(b["rows"]) for _, blocks in best_state for b in blocks)
        output_vol   = _state_vol(best_state)
        farm_ok = "OK" if output_farms == input_farms else f"(!) LOST {input_farms - output_farms}"
        vol_ok  = "OK" if abs(output_vol - input_vol) < 1 else f"(!) LOST {input_vol - output_vol:,.0f}L"

        # Full cost breakdown for the final best state
        _km_f = _shift_f = _shift_pen_f = _milking_f = _cap_f = _win_f = 0.0
        for sname_f, blocks_f in best_state:
            entry_f = self.cache.get(self.fname, {}).get(sname_f, {})
            st_f    = entry_f.get("start_time") if isinstance(entry_f, dict) else None
            for dists_f in calc_distances(blocks_f, self.dm):
                for d_f in dists_f[:-1]:
                    if d_f is not None: _km_f += d_f
            if st_f:
                ct_f = calc_times(blocks_f, self.dm, st_f,
                                  suppress_no_milking=self.cfg.get("suppress_no_milking", True))
                if ct_f:
                    _base_f = datetime.combine(date.today(), st_f)
                    _sh_f   = (ct_f[1] - _base_f).total_seconds() / 3600.0
                    _max_sh = self.cfg.get("max_shift_h", 12.0)
                    _shift_f     += _sh_f * self.cfg.get("shift_hours_weight", 0.0)
                    _shift_pen_f += max(0.0, _sh_f - _max_sh) * self.cfg.get("shift_penalty", 200.0)
                    for b_f, block_f in enumerate(blocks_f):
                        bt_f = ct_f[0][b_f] if b_f < len(ct_f[0]) else None
                        if not bt_f: continue
                        for fi_f, farm_f in enumerate(block_f["rows"]):
                            ft_f = bt_f[fi_f+1] if (fi_f+1) < len(bt_f) else None
                            if ft_f and ft_f.get("wait"):
                                _milking_f += ft_f["wait"] * self.cfg.get("milking_weight", 1.0)
            _hc_f = self.cfg.get("hard_vol_cap", VOL_LIMIT)
            _cr_f = self.cfg.get("cap_penalty", 2.0)
            for block_f in blocks_f:
                _rv_f = sum((r.get("prior_vol") or 0) for r in block_f["rows"]
                            if isinstance(r.get("prior_vol"), (int, float)))
                if _rv_f > _hc_f:
                    _cap_f += (_rv_f - _hc_f) * _cr_f
        _vol_pen_f   = _group_vol_penalty(best_state, orig_dest_vols, self.cfg)
        _plant_win_f = best_cost - _km_f - _shift_f - _shift_pen_f - _milking_f - _cap_f - _vol_pen_f

        self.log.emit(
            f"[{colour}] Done ? best={best_cost:.1f}  "
            f"(started {self._group_cost([(sname, copy.deepcopy(entry.get('blocks',[]))) for sname,entry in sheets], orig_dest_vols):.1f})\n"
            f"  farms: {input_farms}->{output_farms} {farm_ok}  "
            f"vol: {input_vol:,.0f}->{output_vol:,.0f}L {vol_ok}\n"
            f"  Cost breakdown:\n"
            f"    km={_km_f:.1f}  milking={_milking_f:.1f}  shift={_shift_f:.1f}"
            f"  overtime={_shift_pen_f:.1f}  cap={_cap_f:.1f}"
            f"  plant_win={_plant_win_f:.1f}  vol_pen={_vol_pen_f:.1f}"
        )

        # -- Farm displacement summary -----------------------------------------
        # Compare each farm's original (sname, b_idx) placement against the
        # solved best_state.  A farm is identified by its _uid if available,
        # otherwise by (irma, prior_vol) as a fallback key.
        def _farm_key(farm):
            uid = farm.get("_uid")
            if uid:
                return uid
            return (farm.get("irma",""), farm.get("prior_vol", 0))

        # Build original placement map: farm_key -> (sname, b_idx)
        orig_placement = {}
        for sname, entry in sheets:
            for b_idx, block in enumerate(entry.get("blocks", [])):
                for farm in block.get("rows", []):
                    orig_placement[_farm_key(farm)] = (sname, b_idx)

        # Build solved placement map: farm_key -> (sname, b_idx)
        solved_placement = {}
        for sname, blocks in best_state:
            for b_idx, block in enumerate(blocks):
                for farm in block.get("rows", []):
                    solved_placement[_farm_key(farm)] = (sname, b_idx)

        n_total = len(orig_placement)
        n_diff_block  = 0  # different block (may be same route)
        n_diff_route  = 0  # different route entirely
        for key, (orig_sname, orig_b) in orig_placement.items():
            solved = solved_placement.get(key)
            if solved is None:
                continue
            sol_sname, sol_b = solved
            if sol_sname != orig_sname:
                n_diff_route += 1
                n_diff_block += 1
            elif sol_b != orig_b:
                n_diff_block += 1

        pct_block = n_diff_block / n_total * 100 if n_total else 0
        pct_route = n_diff_route / n_total * 100 if n_total else 0
        self.log.emit(
            f"  [{colour}] Farm displacement vs original:\n"
            f"    Different block (incl. different route): {n_diff_block:3d} / {n_total}  ({pct_block:.1f}%)\n"
            f"    Different route only:                   {n_diff_route:3d} / {n_total}  ({pct_route:.1f}%)"
        )

        # Post-solve: apply exhaustive 2-opt to every block in best_state.
        # This is deterministic, never worsens the solution, and recovers
        # any intra-block ordering improvements the ALNS loop may have missed.
        self.log.emit(f"[{colour}] Applying 2-opt polish pass...")
        best_state_list = [(sn, blks) for sn, blks in best_state]
        polished = self._two_opt_all_blocks(best_state_list)
        polished_cost = self._group_cost(polished, orig_dest_vols)
        if polished_cost < best_cost:
            self.log.emit(
                f"[{colour}] 2-opt improved {best_cost:.1f} -> {polished_cost:.1f}"
            )
            best_state = polished
            best_cost  = polished_cost

        # Post-solve: exhaustively optimise destination order on every multi-dest
        # block.  _shuffle_dests was previously a stochastic SA move but is
        # deterministic ? it always returns the best permutation.  Running it
        # in the SA loop just burned iterations for free improvements and inflated
        # the adaptive scores of non-exploratory moves.  Here it runs once over
        # every eligible block after the search is complete.
        self.log.emit(f"[{colour}] Applying destination-order polish pass...")
        shuffled = copy.deepcopy([(sn, blks) for sn, blks in best_state])
        for s_idx, (sn, blocks) in enumerate(shuffled):
            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block) or _is_fixed_vol_block(block):
                    continue
                dests = block.get("dests") or []
                if len(dests) < 2 or len(dests) > 5:
                    continue
                # _shuffle_dests picks a random block ? call it directly per block
                trial_state = shuffled
                improved_state = self._shuffle_dests(trial_state, orig_dest_vols)
                # Only keep if it improved
                shuffled_cost = self._group_cost(improved_state, orig_dest_vols)
                if shuffled_cost < best_cost:
                    shuffled = improved_state
                    best_cost = shuffled_cost
        shuffled_cost = self._group_cost(shuffled, orig_dest_vols)
        if shuffled_cost < self._group_cost([(sn, blks) for sn, blks in best_state],
                                             orig_dest_vols):
            self.log.emit(
                f"[{colour}] Dest-order polish improved -> {shuffled_cost:.1f}"
            )
            best_state = shuffled

        # Accumulate zero-vol farms for logging; they are already re-inserted
        # into the result below so they appear on their original route sheets.
        self.zero_vol_farms.extend(zero_vol_farms)

        result = {sname: blocks for sname, blocks in best_state}
        _reinsert_zero_vol(result)
        return result

    def run(self):
        fname  = self.fname
        cache  = self.cache
        # Seed the RNG when the user supplied one, so a given (input, settings,
        # seed) triple reproduces exactly. None -> leave the global RNG alone.
        seed = self.cfg.get("seed")
        if seed is not None:
            random.seed(seed)
            self.log.emit(f"Random seed: {seed} (reproducible run)")
        groups = _group_sheets_by_colour(cache, fname)
        iters  = self.cfg.get("iterations", 300)

        # Log holdover blocks found in the file so the user can verify them
        holdover_info = []
        for colour_g, sheet_list in groups.items():
            for sname, entry in sheet_list:
                for bi, block in enumerate(entry.get("blocks", [])):
                    dnames = [d.get("name","?") for d in (block.get("dests") or [])]
                    if _is_preload_block(block):
                        holdover_info.append(
                            f"  [{colour_g}] Sheet {sname} block {bi+1}: "
                            f"PRELOAD (fully frozen) -> {', '.join(dnames)}")
                    elif _is_holdover_block(block):
                        holdover_info.append(
                            f"  [{colour_g}] Sheet {sname} block {bi+1}: "
                            f"holdover (farms free, dest fixed) -> {', '.join(dnames)}")
                    elif _is_fixed_vol_block(block):
                        vps = ", ".join(
                            f"{d.get('name','?')}={d.get('vol_partial',0):,.0f}L"
                            for d in (block.get("dests") or [])
                        )
                        holdover_info.append(
                            f"  [{colour_g}] Sheet {sname} block {bi+1}: "
                            f"fixed-vol (fully frozen) -> {vps}")
        if holdover_info:
            self.log.emit("Preload / holdover blocks detected:\n" +
                          "\n".join(holdover_info))
        else:
            self.log.emit("No preload or holdover blocks found.")

        red_result  = self._solve_group("RED",  groups["RED"],
                                        iters, iter_offset=0)
        blue_result = self._solve_group("BLUE", groups["BLUE"],
                                        iters, iter_offset=iters)

        results = {}
        results.update(red_result)
        results.update(blue_result)

        # -- HiGHS post-optimality processor assignment check -----------------
        self.log.emit("\nRunning HiGHS processor-assignment verification...")
        for colour, sheets in [("RED", groups["RED"]), ("BLUE", groups["BLUE"])]:
            if not sheets:
                continue
            # Build state list from the solved results
            solved_state = [(sn, results[sn]) for sn, _e in sheets if sn in results]
            if not solved_state:
                continue
            summary = _highs_verify_processor_assignment(
                colour, sheets, solved_state, self.dm, self.cfg, self.log.emit)
            self.log.emit(summary)

        self.finished.emit(results)


# -- IRMA Lookup Dialog --------------------------------------------------------

class IRMALookupDialog(QDialog):
    """Cross-file, cross-sheet IRMA farm lookup.

    Given a full or partial IRMA number, searches every loaded file and sheet
    (both original parsed data and solver-modified routes) and lists every
    block that contains a matching farm.  Double-clicking a result navigates
    the main window to that file/sheet.

    Results table columns:
        File | Sheet | Route | Block # | Farm position | IRMA | Source
    where Source is 'Original', 'Modified', or 'Both'.
    """

    # Emitted when the user double-clicks a result row so MainWindow can
    # navigate without the dialog needing a direct reference to it.
    # Payload: (fname, sname)
    navigate_requested = pyqtSignal(str, str)

    def __init__(self, cache, sheet_mods, parent=None):
        """
        cache       : MainWindow._cache  ? {fname: {sname: {blocks, ...}}}
        sheet_mods  : MainWindow._sheet_mods ? {(fname,sname): mod_blocks}
        """
        super().__init__(parent)
        self.setWindowTitle("IRMA Farm Lookup")
        self.setMinimumSize(820, 480)
        self._cache      = cache
        self._sheet_mods = sheet_mods
        self._results    = []   # list of result dicts, parallel to table rows

        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 10, 10, 10)

        # -- Search bar -------------------------------------------------------
        bar = QHBoxLayout()
        bar.addWidget(QLabel("IRMA number:"))
        self._query = QLineEdit()
        self._query.setPlaceholderText(
            "Full (71-117) or partial (71)  ?  Enter to search")
        self._query.setMinimumWidth(200)
        bar.addWidget(self._query, stretch=1)

        self._search_btn = QPushButton("Search")
        self._search_btn.setFixedHeight(28)
        self._search_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:3px; padding: 0 12px; } "
            "QPushButton:disabled { background:#90caf9; }")
        bar.addWidget(self._search_btn)

        self._chk_orig = QCheckBox("Original")
        self._chk_orig.setChecked(True)
        self._chk_mod  = QCheckBox("Modified")
        self._chk_mod.setChecked(True)
        bar.addWidget(self._chk_orig)
        bar.addWidget(self._chk_mod)
        layout.addLayout(bar)

        # -- Status label ----------------------------------------------------
        self._status = QLabel("Enter an IRMA number and press Search or Enter.")
        layout.addWidget(self._status)

        # -- Results table ----------------------------------------------------
        RESULT_COLS = ["File", "Sheet", "Route", "Block #",
                       "Position in block", "IRMA", "Source"]
        self._table = QTableWidget(0, len(RESULT_COLS))
        self._table.setHorizontalHeaderLabels(RESULT_COLS)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setToolTip("Double-click a row to navigate to that sheet")
        layout.addWidget(self._table, stretch=1)

        # -- Bottom buttons ---------------------------------------------------
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.setFixedWidth(80)
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        # -- Connections ------------------------------------------------------
        self._query.returnPressed.connect(self._run_search)
        self._search_btn.clicked.connect(self._run_search)
        self._table.cellDoubleClicked.connect(self._on_double_click)

    # -- Search logic ---------------------------------------------------------

    def _run_search(self):
        query = self._query.text().strip().upper()
        self._table.setRowCount(0)
        self._results = []

        if not query:
            self._status.setText("Enter an IRMA number and press Search or Enter.")
            return

        search_orig = self._chk_orig.isChecked()
        search_mod  = self._chk_mod.isChecked()

        if not search_orig and not search_mod:
            self._status.setText("Select at least one of Original / Modified.")
            return

        hits = []   # list of dicts

        for fname, sheets in self._cache.items():
            for sname, entry in sheets.items():
                if not isinstance(entry, dict):
                    continue

                orig_blocks = entry.get("blocks", [])
                mod_blocks  = self._sheet_mods.get((fname, sname))

                # Collect matching farms from each source independently so we
                # can report whether a farm appears in Original, Modified, or both.
                def _scan(blocks, source_label):
                    found = []
                    for b_idx, block in enumerate(blocks):
                        route = block.get("route", "") or ""
                        for f_idx, row in enumerate(block.get("rows", [])):
                            irma = (row.get("irma") or "").strip().upper()
                            if query in irma:
                                found.append({
                                    "fname":   fname,
                                    "sname":   sname,
                                    "route":   route,
                                    "b_idx":   b_idx,
                                    "f_idx":   f_idx,
                                    "irma":    row.get("irma", "").strip(),
                                    "source":  source_label,
                                })
                    return found

                orig_hits = _scan(orig_blocks, "Original") if search_orig else []
                mod_hits  = _scan(mod_blocks,  "Modified") if (search_mod and mod_blocks) else []

                # Merge: if the same (block, position, irma) appears in both,
                # collapse into a single "Both" row rather than duplicating.
                def _key(h):
                    return (h["b_idx"], h["f_idx"], h["irma"])

                orig_keys = {_key(h): h for h in orig_hits}
                mod_keys  = {_key(h): h for h in mod_hits}

                for k, h in orig_keys.items():
                    if k in mod_keys:
                        h_copy = dict(h); h_copy["source"] = "Both"
                        hits.append(h_copy)
                    else:
                        hits.append(h)
                for k, h in mod_keys.items():
                    if k not in orig_keys:
                        hits.append(h)

        # Sort: file -> sheet -> block -> position
        hits.sort(key=lambda h: (h["fname"], h["sname"], h["b_idx"], h["f_idx"]))

        if not hits:
            self._status.setText(
                f"No routes found containing IRMA matching '{query}'.")
            return

        self._status.setText(
            f"{len(hits)} result{'s' if len(hits) != 1 else ''} "
            f"for '{query}' - double-click a row to navigate.")

        SOURCE_COLOURS = {
            "Original": QColor("#e8f5e9"),   # light green
            "Modified": QColor("#e3f2fd"),   # light blue
            "Both":     QColor("#fff8e1"),   # light amber
        }

        self._table.setRowCount(len(hits))
        self._results = hits

        for row_idx, h in enumerate(hits):
            bg = SOURCE_COLOURS.get(h["source"], QColor("#ffffff"))
            values = [
                h["fname"],
                h["sname"],
                h["route"] or "-",
                str(h["b_idx"] + 1),
                f"Farm {h['f_idx'] + 1}",
                h["irma"],
                h["source"],
            ]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setBackground(bg)
                if col_idx == 5:   # IRMA column ? bold
                    f = item.font(); f.setBold(True); item.setFont(f)
                self._table.setItem(row_idx, col_idx, item)

    def _on_double_click(self, row, _col):
        if 0 <= row < len(self._results):
            h = self._results[row]
            self.navigate_requested.emit(h["fname"], h["sname"])


# -- Main window ---------------------------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Route Sheet Viewer")
        self.resize(1700, 960)
        self.data_root    = get_exe_dir() / "anonymized_output"
        self.dm           = load_distance_matrix(get_data_dir() / "distance_matrix.csv")
        self.dm_dur       = load_distance_matrix(get_data_dir() / "duration_matrix.csv")
        # Precompute set of all node keys known to the distance matrix for fast validation
        self._dm_keys: set = {k for pair in self.dm for k in pair}
        self._file_map    = {}
        self._month_map   = {}
        self._year_map    = {}
        self._cache       = {}
        self._mod_blocks  = None
        self._driver_start = None
        self._day_colour   = ""
        self._removed     = []      # list of farm dicts (with _from_block, _from_sheet keys)
        self._sheet_mods       = {}   # (fname, sname) -> mod_blocks (solver output)
        self._corrected_blocks = {}   # (fname, sname) -> corrected baseline (never overwritten by solver)
        self._loader      = None
        self._load_warnings = []   # per-sheet parse warnings from the current load
        # Solver UI state
        self._locked_sheet_cbs  = {}   # sname -> QCheckBox  (populated per file)
        self._demand_open_edits  = {}  # proc_key -> QLineEdit (HH:MM)
        self._demand_close_edits = {}  # proc_key -> QLineEdit (HH:MM)
        self._spin_chars  = ["|", "/", "-", "\\", "|", "/", "-", "\\"]
        self._spin_idx    = 0
        self._spin_timer  = QTimer(self)
        self._spin_timer.timeout.connect(self._tick_spinner)
        self._build_ui()
        self._scan_folders()
        # Wire the dropdown signals after the initial folder scan, so combos
        # populated during the scan don't fire their handlers. (Previously these
        # were connected from main() after construction, which had the same
        # effect; keeping them here makes MainWindow self-contained.)
        self.year_cb.currentIndexChanged.connect(self._on_year)
        self.month_cb.currentIndexChanged.connect(self._on_month)
        self.file_cb.currentIndexChanged.connect(self._on_file)
        self.sheet_cb.currentIndexChanged.connect(self._on_sheet)

    # -------------------------------------------------------------------------
    # UI construction
    # -------------------------------------------------------------------------

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(8,8,8,8); root.setSpacing(8)

        # Left control panel
        left = QFrame(); left.setFrameShape(QFrame.StyledPanel); left.setFixedWidth(220)
        ll = QVBoxLayout(left); ll.setContentsMargins(10,10,10,10); ll.setSpacing(5)
        bold = QFont(); bold.setBold(True)
        def lbl(t): l=QLabel(t); l.setFont(bold); return l

        # -- Folder browser ------------------------------------------------
        self.browse_btn = QPushButton("Browse for Folder...")
        self.browse_btn.setFixedHeight(28)
        self.browse_btn.setToolTip(
            "Choose the root data folder.\n"
            "Expected layout:\n"
            "  <root>/\n"
            "    \\__-- <year>/\n"
            "          \\__-- <month>/\n"
            "                \\__-- *.xlsx")
        self.browse_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:hover { background:#1976d2; }")
        self.browse_btn.clicked.connect(self._on_browse_folder)
        ll.addWidget(self.browse_btn)
        ll.addSpacing(4)

        for label, attr in [("Year","year_cb"),("Month","month_cb")]:
            ll.addWidget(lbl(label))
            cb = QComboBox(); setattr(self, attr, cb); ll.addWidget(cb); ll.addSpacing(2)

        # File dropdown + Load button immediately below it
        ll.addWidget(lbl("File"))
        self.file_cb = QComboBox(); ll.addWidget(self.file_cb)
        ll.addSpacing(2)

        self.load_btn = QPushButton("Load File")
        self.load_btn.setFixedHeight(28)
        self.load_btn.clicked.connect(self._on_load_clicked)
        ll.addWidget(self.load_btn)
        ll.addSpacing(6)

        # Sheet dropdown
        ll.addWidget(lbl("Sheet"))
        self.sheet_cb = QComboBox()
        # Use amber/orange selection so BLUE-route items remain distinguishable
        self.sheet_cb.setStyleSheet(
            "QComboBox QAbstractItemView::item:selected "
            "{ background: #e65100; color: white; }")
        ll.addWidget(self.sheet_cb)
        ll.addSpacing(2)

        # Day colour badge
        self.day_colour_box = QLabel("")
        self.day_colour_box.setAlignment(Qt.AlignCenter)
        self.day_colour_box.setFixedHeight(28)
        self.day_colour_box.setFont(bold)
        self.day_colour_box.setStyleSheet("border-radius: 4px; padding: 2px 6px;")
        ll.addWidget(self.day_colour_box)
        ll.addSpacing(8)

        # Timing assumptions box
        timing_title = QLabel("Timing Assumptions")
        timing_title.setFont(bold)
        ll.addWidget(timing_title)
        timing_frame = QFrame()
        timing_frame.setFrameShape(QFrame.NoFrame)
        timing_frame.setStyleSheet("""
            QFrame { background: #fafafa; border: 1px solid #e0e0e0;
                     border-radius: 6px; }
            QLabel { border: none; background: transparent; }
        """)
        tf = QVBoxLayout(timing_frame); tf.setContentsMargins(10,8,10,8); tf.setSpacing(5)
        lbl_font  = QFont()
        val_font  = QFont(); val_font.setBold(True)
        sep_color = "#e0e0e0"
        def add_timing_row(label, value, add_sep=True):
            row_w = QWidget(); row_w.setStyleSheet("background:transparent;")
            row_l = QHBoxLayout(row_w)
            row_l.setContentsMargins(0,0,0,0); row_l.setSpacing(4)
            lbl_w = QLabel(label); lbl_w.setFont(lbl_font)
            lbl_w.setStyleSheet("color:#555555;")
            val_w = QLabel(value);  val_w.setFont(val_font)
            val_w.setAlignment(Qt.AlignRight|Qt.AlignVCenter)
            val_w.setStyleSheet("color:#222222;")
            row_l.addWidget(lbl_w, stretch=1); row_l.addWidget(val_w)
            tf.addWidget(row_w)
            if add_sep:
                sep = QFrame(); sep.setFrameShape(QFrame.HLine)
                sep.setStyleSheet("color:#e0e0e0; background:#e0e0e0; max-height:1px;")
                tf.addWidget(sep)
        add_timing_row("Setup time",   f"{int(ONSITE_MIN)} min / stop")
        add_timing_row("Pump rate",    f"{int(PUMP_RATE_LPM)} L / min")
        add_timing_row("Volume limit", f"{VOL_LIMIT:,} L", add_sep=False)
        ll.addWidget(timing_frame)

        # Push everything below to the bottom
        ll.addStretch()

        # -- Bottom action buttons ----------------------------------------------
        self.export_btn = QPushButton("Export to Excel...")
        self.export_btn.setFixedHeight(30)
        self.export_btn.setStyleSheet(
            "QPushButton { background:#43a047; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#a5d6a7; }")
        self.export_btn.clicked.connect(self._on_export_excel)
        ll.addWidget(self.export_btn)
        ll.addSpacing(3)

        self.reset_sheet_btn = QPushButton("Reset This Sheet")
        self.reset_sheet_btn.setFixedHeight(26)
        self.reset_sheet_btn.setStyleSheet(
            "QPushButton { background:#29b6f6; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#81d4fa; }")
        self.reset_sheet_btn.clicked.connect(self._on_reset_current_sheet)
        ll.addWidget(self.reset_sheet_btn)
        ll.addSpacing(3)

        self.reset_btn = QPushButton("Reset All to Original")
        self.reset_btn.setFixedHeight(26)
        self.reset_btn.setStyleSheet(
            "QPushButton { background:#e57373; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#ef9a9a; }")
        self.reset_btn.clicked.connect(self._on_reset_all)
        ll.addWidget(self.reset_btn)
        ll.addSpacing(4)

        self.irma_lookup_btn = QPushButton("IRMA Farm Lookup...")
        self.irma_lookup_btn.setFixedHeight(28)
        self.irma_lookup_btn.setStyleSheet(
            "QPushButton { background:#6a1b9a; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#ce93d8; }")
        self.irma_lookup_btn.setToolTip(
            "Find which routes contain a given farm IRMA number.\n"
            "Searches all loaded files ? both original and solver-modified routes.")
        self.irma_lookup_btn.clicked.connect(self._on_irma_lookup)
        ll.addWidget(self.irma_lookup_btn)
        ll.addSpacing(4)

        root.addWidget(left)

        # Tabs
        self.tabs = QTabWidget()
        root.addWidget(self.tabs, stretch=1)
        self.tabs.currentChanged.connect(self._on_tab_changed)

        self._build_route_tab(bold)
        self._build_comparison_tab(bold)

    def _build_route_tab(self, bold):
        route_tab = QWidget()
        rt = QVBoxLayout(route_tab); rt.setContentsMargins(0,0,0,0); rt.setSpacing(4)

        # -- IRMA Search bar --------------------------------------------------
        # Sits above both tables so it searches Original and Modified together.
        search_frame = QFrame()
        search_frame.setFrameShape(QFrame.StyledPanel)
        search_frame.setMaximumHeight(36)
        sl = QHBoxLayout(search_frame)
        sl.setContentsMargins(6, 3, 6, 3); sl.setSpacing(6)

        sl.addWidget(QLabel("IRMA Search:"))
        self._search_box = QLineEdit()
        self._search_box.setPlaceholderText("e.g. 71-117  (Enter to search, Esc to clear)")
        self._search_box.setFixedWidth(230)
        sl.addWidget(self._search_box)

        self._search_prev_btn = QPushButton("< Prev")
        self._search_next_btn = QPushButton("Next >")
        self._search_clear_btn = QPushButton("Clear")
        for btn in (self._search_prev_btn, self._search_next_btn, self._search_clear_btn):
            btn.setFixedHeight(24)
            btn.setFixedWidth(64)
            btn.setEnabled(False)
        self._search_next_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; } "
            "QPushButton:disabled { background:#90caf9; }")
        self._search_prev_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; } "
            "QPushButton:disabled { background:#90caf9; }")
        self._search_clear_btn.setStyleSheet(
            "QPushButton { background:#757575; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; } "
            "QPushButton:disabled { background:#bdbdbd; }")
        sl.addWidget(self._search_prev_btn)
        sl.addWidget(self._search_next_btn)
        sl.addWidget(self._search_clear_btn)

        self._search_status = QLabel("")
        self._search_status.setMinimumWidth(120)
        sl.addWidget(self._search_status)
        sl.addStretch()

        self._search_box.returnPressed.connect(self._on_search)
        self._search_box.textChanged.connect(self._on_search_text_changed)
        self._search_next_btn.clicked.connect(self._on_search_next)
        self._search_prev_btn.clicked.connect(self._on_search_prev)
        self._search_clear_btn.clicked.connect(self._on_search_clear)
        rt.addWidget(search_frame)

        # Internal search state ? populated by _on_search()
        # List of (table, visual_row) for every match, in top-to-bottom order
        # across both tables (orig then edit).
        self._search_hits   = []   # [(table_widget, row_index), ...]
        self._search_cursor = -1   # index into _search_hits for current highlight

        top_split = QSplitter(Qt.Horizontal)

        # Original (read-only)
        orig_frame = QFrame(); orig_frame.setFrameShape(QFrame.StyledPanel)
        ol = QVBoxLayout(orig_frame); ol.setContentsMargins(4,4,4,4)
        lbl_o = QLabel("Original"); lbl_o.setFont(bold); ol.addWidget(lbl_o)
        self.orig_table = QTableWidget()
        self._init_route_table(self.orig_table, editable=False)
        ol.addWidget(self.orig_table)
        top_split.addWidget(orig_frame)

        # Editable (modified)
        edit_frame = QFrame(); edit_frame.setFrameShape(QFrame.StyledPanel)
        el = QVBoxLayout(edit_frame); el.setContentsMargins(4,4,4,4)
        lbl_e = QLabel("Modified  (drag farms or processors to reorder / drop to tray)"); lbl_e.setFont(bold)
        el.addWidget(lbl_e)
        self.edit_table = EditableRouteTable()
        self.edit_table.farm_removed.connect(self._on_farm_removed)
        self.edit_table.farm_inserted.connect(self._on_farm_inserted)
        self.edit_table.farm_reorder.connect(self._on_farm_reorder)
        self.edit_table.dest_removed.connect(self._on_dest_removed)
        self.edit_table.dest_reorder.connect(self._on_dest_reorder)
        self.edit_table.dest_inserted.connect(self._on_dest_inserted)
        self.edit_table.block_reorder.connect(self._on_block_reorder)
        self.edit_table.itemChanged.connect(self._on_mwo_changed)
        self.edit_table.itemSelectionChanged.connect(self._on_del_btn_state)
        el.addWidget(self.edit_table)
        top_split.addWidget(edit_frame)

        top_split.setSizes([800, 800])
        rt.addWidget(top_split, stretch=3)

        # Tray
        tray_frame = QFrame(); tray_frame.setFrameShape(QFrame.StyledPanel)
        tray_frame.setMaximumHeight(240)
        tl = QVBoxLayout(tray_frame); tl.setContentsMargins(4,4,4,4); tl.setSpacing(3)
        tray_hdr_row = QWidget()
        tray_hdr_l = QHBoxLayout(tray_hdr_row); tray_hdr_l.setContentsMargins(0,0,0,0); tray_hdr_l.setSpacing(6)
        lbl_t = QLabel("Removed Farms & Processors -- drag from Modified to remove | drag back to restore")
        lbl_t.setFont(bold); tray_hdr_l.addWidget(lbl_t, stretch=1)
        self._tray_del_btn = QPushButton("Delete Selected")
        self._tray_del_btn.setFixedHeight(24)
        self._tray_del_btn.setStyleSheet(
            "QPushButton { background:#e53935; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; } "
            "QPushButton:disabled { background:#ef9a9a; }")
        self._tray_del_btn.setEnabled(False)
        self._tray_del_btn.clicked.connect(self._on_tray_delete)
        tray_hdr_l.addWidget(self._tray_del_btn)

        self.add_block_btn = QPushButton("+ Add Block")
        self.add_block_btn.setFixedHeight(24)
        self.add_block_btn.setStyleSheet(
            "QPushButton { background:#7b1fa2; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; } "
            "QPushButton:disabled { background:#ce93d8; }")
        self.add_block_btn.clicked.connect(self._on_add_block)
        tray_hdr_l.addWidget(self.add_block_btn)
        tl.addWidget(tray_hdr_row)
        self.farm_tray = FarmTray()
        self.farm_tray.farm_incoming.connect(self._on_farm_removed)
        self.farm_tray.dest_incoming.connect(self._on_dest_removed)
        self.farm_tray.itemSelectionChanged.connect(self._on_del_btn_state)
        tl.addWidget(self.farm_tray)
        rt.addWidget(tray_frame, stretch=1)

        # -- Manual Add bars (Farm + Processor) -----------------------------
        add_frame = QFrame(); add_frame.setFrameShape(QFrame.StyledPanel)
        al = QVBoxLayout(add_frame); al.setContentsMargins(6,4,6,4); al.setSpacing(2)

        field_font = QFont()
        def _fe(placeholder, width=90):
            w = QLineEdit(); w.setPlaceholderText(placeholder)
            w.setFont(field_font); w.setFixedWidth(width); return w

        # -- Farm row ------------------------------------------------------
        farm_hdr = QLabel("Add Farm to Tray"); farm_hdr.setFont(bold)
        al.addWidget(farm_hdr)

        farm_row = QWidget()
        frl = QHBoxLayout(farm_row); frl.setContentsMargins(0,0,0,0); frl.setSpacing(4)

        # Editable combobox: type freely OR pick from known IRMAs.
        # Selecting a known IRMA autofills Train/Milking/EDPU/Location.
        self._add_irma = QComboBox()
        self._add_irma.setEditable(True)
        self._add_irma.setInsertPolicy(QComboBox.NoInsert)
        self._add_irma.setFixedWidth(80)
        self._add_irma.setFont(field_font)
        self._add_irma.lineEdit().setPlaceholderText("IRMA #")
        self._add_irma.lineEdit().setFont(field_font)
        self._add_irma.activated.connect(self._on_irma_autofill)

        self._add_train  = _fe("Train",  50)
        self._add_m1s    = _fe("M1 Start",62)
        self._add_m1f    = _fe("M1 Fin", 62)
        self._add_m2s    = _fe("M2 Start",62)
        self._add_m2f    = _fe("M2 Fin", 62)
        self._add_edpu   = _fe("EDPU",   50)
        self._add_loc    = _fe("Location",105)
        self._add_vol    = _fe("Vol (L)", 76)

        for w in (self._add_irma, self._add_train,
                  self._add_m1s, self._add_m1f, self._add_m2s, self._add_m2f,
                  self._add_edpu, self._add_loc, self._add_vol):
            frl.addWidget(w)
        self._add_btn = QPushButton("Add Farm")
        self._add_btn.setFixedHeight(26); self._add_btn.setFixedWidth(86)
        self._add_btn.clicked.connect(self._on_manual_add)
        frl.addWidget(self._add_btn); frl.addStretch()
        al.addWidget(farm_row)

        # -- Processor row -------------------------------------------------
        sep2 = QFrame(); sep2.setFrameShape(QFrame.HLine)
        sep2.setStyleSheet("color:#ddd; background:#ddd; max-height:1px;")
        al.addWidget(sep2)

        proc_hdr = QLabel("Add Processor to Tray"); proc_hdr.setFont(bold)
        al.addWidget(proc_hdr)

        proc_row = QWidget()
        prl = QHBoxLayout(proc_row); prl.setContentsMargins(0,0,0,0); prl.setSpacing(4)

        # Editable combobox for proc key ? populated from known processors in cache
        self._add_proc_key = QComboBox()
        self._add_proc_key.setEditable(True)
        self._add_proc_key.setInsertPolicy(QComboBox.NoInsert)
        self._add_proc_key.setMinimumWidth(220)
        self._add_proc_key.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self._add_proc_key.setFont(field_font)
        self._add_proc_key.lineEdit().setPlaceholderText("Proc Key (e.g. 901012)")
        self._add_proc_key.lineEdit().setFont(field_font)
        self._add_proc_key.activated.connect(self._on_proc_key_autofill)
        self._add_proc_key.view().setMinimumWidth(350)

        self._add_proc_name = _fe("Name (e.g. Olympic - FA)", 200)
        self._add_proc_vol  = _fe("Partial Vol L (blank=rest)", 170)
        for w in (self._add_proc_key, self._add_proc_name, self._add_proc_vol):
            prl.addWidget(w)
        self._add_proc_btn = QPushButton("Add Processor")
        self._add_proc_btn.setFixedHeight(26); self._add_proc_btn.setFixedWidth(110)
        self._add_proc_btn.clicked.connect(self._on_manual_add_proc)
        prl.addWidget(self._add_proc_btn); prl.addStretch()
        al.addWidget(proc_row)

        self._add_status = QLabel("")
        self._add_status.setStyleSheet("color: #c0392b; font-size: 8pt;")
        al.addWidget(self._add_status)
        self._add_status_timer = QTimer(self)
        self._add_status_timer.setSingleShot(True)
        self._add_status_timer.timeout.connect(lambda: self._add_status.setText(""))

        rt.addWidget(add_frame)

        self.tabs.addTab(route_tab, "Route View")

    def _build_comparison_tab(self, bold):
        comp_tab = QWidget()
        cl = QVBoxLayout(comp_tab); cl.setContentsMargins(8,8,8,8); cl.setSpacing(6)

        hdr = QLabel("File-wide Original vs Modified - all loaded sheets")
        hdr.setFont(bold); cl.addWidget(hdr)

        self._comp_tables = {}   # key -> QTableWidget

        # Top section: Processor volumes (side-by-side, full width)
        proc_split = QSplitter(Qt.Horizontal)
        for side in ("orig", "mod"):
            w = QWidget(); wl = QVBoxLayout(w); wl.setContentsMargins(0,0,0,0); wl.setSpacing(2)
            lbl = QLabel("Original ? Processor Volumes" if side=="orig"
                         else "Modified ? Processor Volumes"); lbl.setFont(bold)
            wl.addWidget(lbl)
            pt = QTableWidget(); self._init_comp_table(pt)
            self._comp_tables[f"proc_{side}"] = pt
            wl.addWidget(pt)
            proc_split.addWidget(w)
        cl.addWidget(proc_split, stretch=2)

        # Sync processor tables
        self._comp_tables["proc_orig"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["proc_mod"].verticalScrollBar().setValue(v))
        self._comp_tables["proc_mod"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["proc_orig"].verticalScrollBar().setValue(v))

        # Bottom section: Per-sheet hours & km (side-by-side)
        sheet_split = QSplitter(Qt.Horizontal)
        for side in ("orig", "mod"):
            w = QWidget(); wl = QVBoxLayout(w); wl.setContentsMargins(0,0,0,0); wl.setSpacing(2)
            lbl = QLabel("Original ? Sheet Summary" if side=="orig"
                         else "Modified ? Sheet Summary"); lbl.setFont(bold)
            wl.addWidget(lbl)
            st = QTableWidget(); self._init_comp_table(st)
            self._comp_tables[f"sheet_{side}"] = st
            wl.addWidget(st)
            sheet_split.addWidget(w)
        cl.addWidget(sheet_split, stretch=3)

        # Sync sheet summary tables
        self._comp_tables["sheet_orig"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["sheet_mod"].verticalScrollBar().setValue(v))
        self._comp_tables["sheet_mod"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["sheet_orig"].verticalScrollBar().setValue(v))

        self.tabs.addTab(comp_tab, "Comparison")

        self._build_solver_tab(bold)

    def _build_solver_tab(self, bold):
        solver_tab = QWidget()
        sl = QVBoxLayout(solver_tab)
        sl.setContentsMargins(14, 10, 14, 10)
        sl.setSpacing(8)

        # -- Title row -----------------------------------------------------
        title_row = QWidget()
        tr_l = QHBoxLayout(title_row); tr_l.setContentsMargins(0,0,0,0); tr_l.setSpacing(12)
        title = QLabel("ALNS Route Solver")
        title.setStyleSheet("color: #1a3a5c;")
        tr_l.addWidget(title)
        subtitle = QLabel(
            "Optimises farm order across all RED and BLUE routes in the loaded file. "
            "OTHER routes (Grassfed, A2, etc.) are left unchanged."
        )
        subtitle.setWordWrap(True)
        subtitle.setStyleSheet("color: #555; font-size: 9pt;")
        tr_l.addWidget(subtitle, stretch=1)
        sl.addWidget(title_row)

        # -- helper --------------------------------------------------------
        def spin_row(label_text, widget, unit=""):
            row = QWidget()
            rl  = QHBoxLayout(row)
            rl.setContentsMargins(0, 0, 0, 0)
            lbl = QLabel(label_text)
            lbl.setMinimumWidth(155)
            rl.addWidget(lbl)
            rl.addWidget(widget)
            if unit:
                rl.addWidget(QLabel(unit))
            rl.addStretch()
            return row

        # ??????????????????????????????????????????????????????????????????
        # TOP BAND ? Objective Weights | Constraints | ALNS | Run controls
        # ??????????????????????????????????????????????????????????????????
        top_band = QWidget()
        top_l = QHBoxLayout(top_band)
        top_l.setContentsMargins(0, 0, 0, 0)
        top_l.setSpacing(10)
        top_l.setAlignment(Qt.AlignTop)

        # -- Objective Weights --------------------------------------------
        obj_box = QGroupBox("Objective Weights")
        obj_box.setFont(bold)
        obj_l = QVBoxLayout(obj_box)
        obj_l.setSpacing(5)
        obj_l.setAlignment(Qt.AlignTop)

        self._sw_milking = QDoubleSpinBox()
        self._sw_milking.setRange(0.0, 20.0)
        self._sw_milking.setSingleStep(0.5)
        self._sw_milking.setValue(0.3)
        self._sw_milking.setDecimals(1)
        self._sw_milking.setToolTip(
            "Multiplier on milking-window wait time (in km-equivalent).\n"
            "0 = ignore milking windows entirely.")
        obj_l.addWidget(spin_row("Milking window x", self._sw_milking))

        self._sw_plant_win_pen = QDoubleSpinBox()
        self._sw_plant_win_pen.setRange(0.0, 10000.0)
        self._sw_plant_win_pen.setSingleStep(50.0)
        self._sw_plant_win_pen.setValue(20.0)
        self._sw_plant_win_pen.setDecimals(0)
        self._sw_plant_win_pen.setToolTip(
            "Penalty per hour the truck must wait outside a plant's receiving window.\n"
            "Expressed as km-equivalent ? same scale as routing distance.\n"
            "200 km/h: a 1-hour wait costs as much as driving 200 extra km.\n"
            "Set to 0 to ignore receiving windows entirely.")
        obj_l.addWidget(spin_row("Plant window pen", self._sw_plant_win_pen, "km/h wait"))

        self._sw_plant_margin_mins = QDoubleSpinBox()
        self._sw_plant_margin_mins.setRange(0.0, 240.0)
        self._sw_plant_margin_mins.setSingleStep(5.0)
        self._sw_plant_margin_mins.setValue(30.0)
        self._sw_plant_margin_mins.setDecimals(0)
        self._sw_plant_margin_mins.setToolTip(
            "Minutes before plant closing to start penalising arrivals.\n"
            "Creates a slope inside the window so the solver prefers\n"
            "earlier arrivals and avoids cutting it close to closing time.\n"
            "0 = no margin penalty (only penalise outside the window).")
        obj_l.addWidget(spin_row("Close margin", self._sw_plant_margin_mins, "min"))

        self._sw_plant_margin_rate = QDoubleSpinBox()
        self._sw_plant_margin_rate.setRange(0.0, 10000.0)
        self._sw_plant_margin_rate.setSingleStep(50.0)
        self._sw_plant_margin_rate.setValue(25.0)
        self._sw_plant_margin_rate.setDecimals(0)
        self._sw_plant_margin_rate.setToolTip(
            "Penalty rate inside the closing margin (per minute of depth).\n"
            "Linear ramp: 0 at margin start, full rate at closing time.\n"
            "Default 250 with 30-min margin = max 7,500 penalty at closing.")
        obj_l.addWidget(spin_row("Margin rate", self._sw_plant_margin_rate, "/min"))

        obj_l.addStretch()
        top_l.addWidget(obj_box)

        # -- Constraints ---------------------------------------------------
        con_box = QGroupBox("Constraints")
        con_box.setFont(bold)
        con_l = QVBoxLayout(con_box)
        con_l.setSpacing(5)
        con_l.setAlignment(Qt.AlignTop)

        self._sw_vol_tol = QDoubleSpinBox()
        self._sw_vol_tol.setRange(0.0, 1.0)
        self._sw_vol_tol.setSingleStep(0.01)
        self._sw_vol_tol.setValue(0.15)
        self._sw_vol_tol.setDecimals(2)
        self._sw_vol_tol.setToolTip(
            "Allowed deviation from original processor volume.\n"
            "0.15 = +/-15%.")
        con_l.addWidget(spin_row("Plant vol tolerance +/-", self._sw_vol_tol, "%*"))

        self._sw_vol_pen = QDoubleSpinBox()
        self._sw_vol_pen.setRange(0.0, 10000.0)
        self._sw_vol_pen.setSingleStep(1.0)
        self._sw_vol_pen.setValue(5.0)
        self._sw_vol_pen.setDecimals(1)
        self._sw_vol_pen.setToolTip(
            "Penalty per litre of processor volume outside the tolerance band.")
        con_l.addWidget(spin_row("Vol violation pen", self._sw_vol_pen, "/L"))

        self._sw_hard_cap = QSpinBox()
        self._sw_hard_cap.setRange(10000, 60000)
        self._sw_hard_cap.setSingleStep(500)
        self._sw_hard_cap.setValue(42000)
        self._sw_hard_cap.setToolTip(
            "Hard truck capacity limit (litres).\n"
            "Routes exceeding this get a stiff per-litre penalty.")
        con_l.addWidget(spin_row("Truck cap limit", self._sw_hard_cap, "L"))

        self._sw_cap_pen = QDoubleSpinBox()
        self._sw_cap_pen.setRange(0.0, 1000.0)
        self._sw_cap_pen.setSingleStep(5.0)
        self._sw_cap_pen.setValue(2.0)
        self._sw_cap_pen.setDecimals(1)
        self._sw_cap_pen.setToolTip(
            "Penalty per litre over the truck capacity limit.\n"
            "Expressed as km-equivalent per litre.\n"
            "10 km/L: a 1,000L overload costs 10,000 km-eq.")
        con_l.addWidget(spin_row("Over-cap penalty", self._sw_cap_pen, "km/L"))

        self._sw_max_shift = QDoubleSpinBox()
        self._sw_max_shift.setRange(1.0, 24.0)
        self._sw_max_shift.setSingleStep(0.5)
        self._sw_max_shift.setValue(12.0)
        self._sw_max_shift.setDecimals(1)
        self._sw_max_shift.setToolTip("Maximum allowed shift length in hours.")
        con_l.addWidget(spin_row("Max shift length", self._sw_max_shift, "h"))

        self._sw_shift_pen = QDoubleSpinBox()
        self._sw_shift_pen.setRange(0.0, 1000.0)
        self._sw_shift_pen.setSingleStep(10.0)
        self._sw_shift_pen.setValue(50.0)
        self._sw_shift_pen.setDecimals(0)
        self._sw_shift_pen.setToolTip("Penalty per hour over the max shift limit.")
        con_l.addWidget(spin_row("Shift overage pen", self._sw_shift_pen, "/h"))

        self._sw_shift_hours = QDoubleSpinBox()
        self._sw_shift_hours.setRange(0.0, 100.0)
        self._sw_shift_hours.setSingleStep(1.0)
        self._sw_shift_hours.setValue(1.0)
        self._sw_shift_hours.setDecimals(1)
        self._sw_shift_hours.setToolTip(
            "Cost per hour of total shift length (applied to the full shift, not just overage).\n"
            "Higher values push the solver to favour shorter days overall.\n"
            "Default 5.0 = roughly equivalent to 6 km per hour of shift time.")
        con_l.addWidget(spin_row("Shift hours weight", self._sw_shift_hours, "/h"))
        con_l.addStretch()
        top_l.addWidget(con_box)

        # -- ALNS Parameters -----------------------------------------------
        alns_box = QGroupBox("ALNS Parameters")
        alns_box.setFont(bold)
        alns_l = QVBoxLayout(alns_box)
        alns_l.setSpacing(5)
        alns_l.setAlignment(Qt.AlignTop)

        self._sw_iters = QSpinBox()
        self._sw_iters.setRange(50, 50000)
        self._sw_iters.setSingleStep(500)
        self._sw_iters.setValue(200)
        self._sw_iters.setToolTip(
            "Number of SA iterations per colour group.\n"
            "alpha is auto-computed so T decays to the cooling target over this many steps.")
        alns_l.addWidget(spin_row("Iterations per colour", self._sw_iters))

        self._sw_cool = QDoubleSpinBox()
        self._sw_cool.setRange(0.0001, 0.9999)
        self._sw_cool.setSingleStep(0.01)
        self._sw_cool.setValue(0.01)
        self._sw_cool.setDecimals(4)
        self._sw_cool.setToolTip(
            "Target temperature as a fraction of T0 at the final iteration.\n"
            "0.10 = T stays at 10% of T0 by the end ? slow, exploratory cooling.\n"
            "0.001 = T collapses to near-zero ? fast convergence.\n"
            "With T0~1,000 and 1,000 iters: 0.10 -> T_final~100, alpha~0.9977.\n"
            "alpha is auto-computed as  cool_target ^ (1 / iterations).")
        alns_l.addWidget(spin_row("Cooling target", self._sw_cool))

        self._sw_seed = QSpinBox()
        self._sw_seed.setRange(0, 2_000_000_000)
        self._sw_seed.setValue(0)
        self._sw_seed.setToolTip(
            "Random seed for the solver.\n"
            "0 = a fresh random seed each run (non-reproducible).\n"
            "Any non-zero value makes the run fully reproducible ? the same\n"
            "input, settings, and seed always produce the same result, which\n"
            "is useful for debugging and comparing parameter changes.")
        alns_l.addWidget(spin_row("Random seed (0=random)", self._sw_seed))

        alns_l.addStretch()
        top_l.addWidget(alns_box)

        # -- Run controls (buttons + progress + status + footnote) ---------
        run_box = QGroupBox("Run")
        run_box.setFont(bold)
        run_l = QVBoxLayout(run_box)
        run_l.setSpacing(6)
        run_l.setAlignment(Qt.AlignTop)

        self._solve_btn = QPushButton("Run Solver")
        self._solve_btn.setFixedHeight(36)
        self._solve_btn.setStyleSheet(
            "QPushButton { background:#1e88e5; color:white; font-weight:bold; "
            "border-radius:4px; } "
            "QPushButton:disabled { background:#90caf9; }")
        self._solve_btn.clicked.connect(self._on_solve_clicked)
        run_l.addWidget(self._solve_btn)

        self._intra_btn = QPushButton("Optimise Within Routes")
        self._intra_btn.setFixedHeight(28)
        self._intra_btn.setStyleSheet(
            "QPushButton { background:#00897b; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#80cbc4; }")
        self._intra_btn.setToolTip(
            "Reorder farms within each route using 2-opt and or-opt until\n"
            "convergence. Applies directly to the Modified panel.\n"
            "No cross-route moves ? only within-route reordering.")
        self._intra_btn.clicked.connect(self._on_intra_route_apply)
        run_l.addWidget(self._intra_btn)

        self._stop_btn = QPushButton("Stop")
        self._stop_btn.setFixedHeight(36)
        self._stop_btn.setEnabled(False)
        self._stop_btn.setStyleSheet(
            "QPushButton { background:#e53935; color:white; font-weight:bold; "
            "border-radius:4px; } "
            "QPushButton:disabled { background:#ef9a9a; }")
        self._stop_btn.clicked.connect(self._on_stop_solver)
        run_l.addWidget(self._stop_btn)

        self._solver_progress = QProgressBar()
        self._solver_progress.setRange(0, 100)
        self._solver_progress.setValue(0)
        self._solver_progress.setTextVisible(True)
        self._solver_progress.setFormat("%p%  %v / %m iters")
        run_l.addWidget(self._solver_progress)

        self._solver_status = QLabel("Ready")
        self._solver_status.setWordWrap(True)
        self._solver_status.setStyleSheet("color:#555; font-size:8pt;")
        run_l.addWidget(self._solver_status)

        self._chk_route_opt = QCheckBox("Apply route corrections on load & after solve")
        self._chk_route_opt.setChecked(False)
        self._chk_route_opt.setToolTip(
            "When checked, applies two corrections to the Modified panel:\n"
            "  1. Optimize partial-dropoff positions (e.g. Ridgecrest, Farmhouse)\n"
            "     to find the farm split that best hits the plant receiving window.\n"
            "  2. Auto-flag farms with >2h waits in the original solution, giving\n"
            "     the solver a 2-hour pickup window anchored to the original arrival.\n\n"
            "Uncheck to fully revert Modified to match Original.")
        self._chk_route_opt.stateChanged.connect(self._on_route_opt_changed)
        run_l.addWidget(self._chk_route_opt)
        # Aliases so existing code references still work
        self._chk_auto_flag = self._chk_route_opt
        self._chk_split_opt = self._chk_route_opt

        note = QLabel(
            "* % displayed; internally stored as fraction.\n"
            "Results written to Modified panel on all RED/BLUE sheets."
        )
        note.setWordWrap(True)
        note.setStyleSheet("color:#888; font-size:8pt;")
        run_l.addWidget(note)
        run_l.addStretch()
        top_l.addWidget(run_box)

        sl.addWidget(top_band)

        # ??????????????????????????????????????????????????????????????????
        # BOTTOM BAND ? Processor Demand | Locked Sheets | Solver Log
        # ??????????????????????????????????????????????????????????????????
        bottom_split = QSplitter(Qt.Horizontal)

        # -- Processor Demand targets (left) -------------------------------
        demand_box = QGroupBox("Processor Demand Targets & Receiving Windows")
        demand_box.setFont(bold)
        demand_l = QVBoxLayout(demand_box)
        demand_l.setSpacing(3)
        demand_note = QLabel(
            "Defaults loaded from current file. Edit volume or receiving window times.\n"
            "Times use HH:MM (24h). Leave both blank to treat as 24/7.")
        demand_note.setWordWrap(True)
        demand_note.setStyleSheet("color:#777; font-size:8pt;")
        demand_l.addWidget(demand_note)

        # Column header row
        hdr_font = QFont(); hdr_font.setBold(True)
        col_hdr = QWidget(); col_hdr_l = QHBoxLayout(col_hdr)
        col_hdr_l.setContentsMargins(2,0,2,0); col_hdr_l.setSpacing(4)
        for txt, w in [("Processor", 110), ("Volume", 88),
                       ("Open", 48), ("Close", 48)]:
            h = QLabel(txt); h.setFont(hdr_font); h.setFixedWidth(w)
            h.setAlignment(Qt.AlignCenter)
            col_hdr_l.addWidget(h)
        col_hdr_l.addStretch()
        demand_l.addWidget(col_hdr)

        demand_scroll = QScrollArea()
        demand_scroll.setWidgetResizable(True)
        demand_inner = QWidget()
        self._demand_layout = QVBoxLayout(demand_inner)
        self._demand_layout.setSpacing(2)
        self._demand_layout.setContentsMargins(2,2,2,2)
        demand_scroll.setWidget(demand_inner)
        demand_l.addWidget(demand_scroll)
        self._demand_spinboxes    = {}
        self._demand_open_edits   = {}
        self._demand_close_edits  = {}

        refresh_btn = QPushButton("Refresh from File")
        refresh_btn.setFixedHeight(24)
        refresh_btn.clicked.connect(self._refresh_demand_targets)
        demand_l.addWidget(refresh_btn)
        bottom_split.addWidget(demand_box)

        # -- Locked Sheets (centre) ----------------------------------------
        lock_box = QGroupBox("Locked Sheets")
        lock_box.setFont(bold)
        lock_l = QVBoxLayout(lock_box)
        lock_l.setSpacing(3)
        lock_note = QLabel(
            "Checked sheets are held constant by the solver ? their farm order\n"
            "and processor assignments will not be changed.")
        lock_note.setWordWrap(True)
        lock_note.setStyleSheet("color:#777; font-size:8pt;")
        lock_l.addWidget(lock_note)

        lock_btn_row = QWidget(); lbr_l = QHBoxLayout(lock_btn_row)
        lbr_l.setContentsMargins(0,0,0,0); lbr_l.setSpacing(4)
        sel_all_btn = QPushButton("Select All"); sel_all_btn.setFixedHeight(20)
        sel_all_btn.setStyleSheet("font-size:8pt;")
        clr_all_btn = QPushButton("Clear All");  clr_all_btn.setFixedHeight(20)
        clr_all_btn.setStyleSheet("font-size:8pt;")
        lbr_l.addWidget(sel_all_btn); lbr_l.addWidget(clr_all_btn); lbr_l.addStretch()
        lock_l.addWidget(lock_btn_row)

        lock_scroll = QScrollArea()
        lock_scroll.setWidgetResizable(True)
        lock_inner = QWidget()
        self._lock_layout = QVBoxLayout(lock_inner)
        self._lock_layout.setSpacing(1)
        self._lock_layout.setContentsMargins(4,2,4,2)
        self._lock_layout.setAlignment(Qt.AlignTop)
        lock_scroll.setWidget(lock_inner)
        lock_l.addWidget(lock_scroll)
        self._locked_sheet_cbs = {}

        def _sel_all():
            for cb in self._locked_sheet_cbs.values(): cb.setChecked(True)
        def _clr_all():
            for cb in self._locked_sheet_cbs.values(): cb.setChecked(False)
        sel_all_btn.clicked.connect(_sel_all)
        clr_all_btn.clicked.connect(_clr_all)

        bottom_split.addWidget(lock_box)

        # -- Solver Log (right) --------------------------------------------
        log_w = QWidget()
        log_l = QVBoxLayout(log_w)
        log_l.setContentsMargins(4, 0, 0, 0)
        log_lbl = QLabel("Solver Log")
        log_lbl.setFont(bold)
        log_l.addWidget(log_lbl)
        self._solver_log = QTextEdit()
        self._solver_log.setReadOnly(True)
        self._solver_log.setFont(QFont("Courier New", 8))
        self._solver_log.setStyleSheet(
            "background:#1e1e1e; color:#d4d4d4; border-radius:4px;")
        log_l.addWidget(self._solver_log)
        bottom_split.addWidget(log_w)

        bottom_split.setStretchFactor(0, 1)   # demand ? narrower
        bottom_split.setStretchFactor(1, 1)   # locked sheets
        bottom_split.setStretchFactor(2, 2)   # log ? wider
        bottom_split.setSizes([280, 180, 820])
        sl.addWidget(bottom_split, stretch=1)

        self.tabs.addTab(solver_tab, "Solver")

        self._build_debug_tab(bold)

        # Internal solver reference
        self._solver_thread = None
        self._demand_spinboxes = {}   # initialised by _build_solver_tab

    # -- Debug tab -------------------------------------------------------------

    def _build_debug_tab(self, bold):
        """Raw data inspector: shows parsed blocks, timing results, and
        distance/duration lookups for the currently displayed sheet."""
        debug_tab = QWidget()
        dl = QVBoxLayout(debug_tab)
        dl.setContentsMargins(8, 8, 8, 8)
        dl.setSpacing(6)

        # Header + refresh button
        hdr_row = QWidget()
        hrl = QHBoxLayout(hdr_row); hrl.setContentsMargins(0,0,0,0); hrl.setSpacing(8)
        hdr_lbl = QLabel("Debug - Raw Block & Timing Data")
        hdr_lbl.setFont(bold)
        hrl.addWidget(hdr_lbl, stretch=1)

        refresh_btn = QPushButton("Refresh")
        refresh_btn.setFixedHeight(24)
        refresh_btn.setStyleSheet(
            "QPushButton { background:#546e7a; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        refresh_btn.clicked.connect(self._refresh_debug_tab)
        hrl.addWidget(refresh_btn)

        copy_btn = QPushButton("Copy")
        copy_btn.setFixedHeight(24)
        copy_btn.setStyleSheet(
            "QPushButton { background:#37474f; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        copy_btn.clicked.connect(self._copy_debug_text)
        hrl.addWidget(copy_btn)

        dl.addWidget(hdr_row)

        # -- Tools row ? suppress milking --------------------------------------
        tools_row = QWidget()
        trl = QHBoxLayout(tools_row); trl.setContentsMargins(0,0,0,0); trl.setSpacing(8)

        self._suppress_no_milking_cb = QCheckBox(
            "Suppress milking windows for 37-874, 14-247, 92-545")
        self._suppress_no_milking_cb.setChecked(False)
        self._suppress_no_milking_cb.setToolTip(
            "When checked, farms 37-874, 14-247 and 92-545 have no milking-window\n"
            "constraints ? arrival times never cause waits and no conflict highlighting\n"
            "is shown for those farms.\n\n"
            "Uncheck to re-enable their windows (e.g. for verification).")
        self._suppress_no_milking_cb.stateChanged.connect(self._on_suppress_milking_changed)
        trl.addWidget(self._suppress_no_milking_cb)

        plant_win_btn = QPushButton("Plant Window Cost Report")
        plant_win_btn.setFixedHeight(24)
        plant_win_btn.setStyleSheet(
            "QPushButton { background:#6a1b9a; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        plant_win_btn.setToolTip(
            "For every route in the Modified panel, show the plant window penalty\n"
            "per processor destination ? grouped by processor across all routes.")
        plant_win_btn.clicked.connect(self._on_plant_window_report)
        trl.addWidget(plant_win_btn)

        cost_report_btn = QPushButton("Full Cost Report")
        cost_report_btn.setFixedHeight(24)
        cost_report_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        cost_report_btn.setToolTip(
            "Show complete cost breakdown for every route in the Modified panel:\n"
            "km, milking waits, shift, overtime, cap, plant window, per-block.")
        cost_report_btn.clicked.connect(self._on_full_cost_report)
        trl.addWidget(cost_report_btn)

        overtime_btn = QPushButton("Overtime Timeline")
        overtime_btn.setFixedHeight(24)
        overtime_btn.setStyleSheet(
            "QPushButton { background:#c62828; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        overtime_btn.setToolTip(
            "For every route in the Modified panel that has overtime,\n"
            "show a stop-by-stop timeline: arrival, wait reason, departure,\n"
            "cumulative shift time. Diagnose gate waits vs milking vs distance.")
        overtime_btn.clicked.connect(self._on_overtime_timeline)
        trl.addWidget(overtime_btn)

        intra_btn = QPushButton("Intra-Route Savings")
        intra_btn.setFixedHeight(24)
        intra_btn.setStyleSheet(
            "QPushButton { background:#00695c; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        intra_btn.setToolTip(
            "Exhaustively reorder farms within each route (2-opt until convergence)\n"
            "and report km and hours saved vs current Modified panel.\n"
            "Shows upper bound of what pure within-route reordering can achieve.")
        intra_btn.clicked.connect(self._on_intra_route_savings)
        trl.addWidget(intra_btn)

        block_cap_btn = QPushButton("Block Capacity Distribution")
        block_cap_btn.setFixedHeight(24)
        block_cap_btn.setStyleSheet(
            "QPushButton { background:#e65100; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        block_cap_btn.setToolTip(
            "For every non-preload block in the Modified panel, compute peak load\n"
            "(matching the cap-penalty logic exactly: total farm vol for simple blocks,\n"
            "running peak through the stop sequence for split blocks).\n\n"
            "Reports: summary statistics, percentiles, histogram, threshold-sensitivity\n"
            "table, top-N heaviest blocks, and per-route maxima.\n\n"
            "Use this to choose hard_vol_cap from data, and to spot solver gaming\n"
            "(loads piling against the threshold rather than spread naturally).")
        block_cap_btn.clicked.connect(self._on_block_capacity_distribution)
        trl.addWidget(block_cap_btn)

        trl.addStretch()
        dl.addWidget(tools_row)

        # Info strip
        self._debug_info = QLabel("")
        self._debug_info.setStyleSheet("color:#555; font-size:8pt;")
        dl.addWidget(self._debug_info)

        # Main text area (raw block/timing data)
        self._debug_text = QTextEdit()
        self._debug_text.setReadOnly(True)
        self._debug_text.setFont(QFont("Courier New", 8))
        self._debug_text.setStyleSheet(
            "background:#1e1e1e; color:#d4d4d4; border-radius:4px;")
        dl.addWidget(self._debug_text, stretch=1)

        self.tabs.addTab(debug_tab, "Debug")

    def _on_full_cost_report(self):
        """Full cost breakdown for every route ? uses same logic as solver via _sheet_cost_breakdown."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }
        lines = ["Full Cost Report - Modified panel", "=" * 70]
        grand = {k: 0.0 for k in ("km","milking","shift","overtime","cap","plant_win","total")}

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))
            frozen = all(_is_preload_block(b) or _is_fixed_vol_block(b) for b in blocks)
            tag    = " [FROZEN]" if frozen else ""

            # Detect partial dropoffs (split_after set on any dest)
            has_split = any(
                d.get("split_after") is not None
                for block in blocks
                for d in (block.get("dests") or [])
            )
            split_tag = " [SPLIT]" if has_split else ""

            bd = _sheet_cost_breakdown(blocks, self.dm, start_time, cfg)

            lines.append(f"\n{sname}{tag}{split_tag}  total={bd['total']:.1f}"
                         f"  km={bd['km']:.1f}  milk={bd['milking']:.1f}"
                         f"  shift={bd['shift']:.1f}  ot={bd['overtime']:.1f}"
                         f"  cap={bd['cap']:.1f}  pw={bd['plant_win']:.1f}")
            if not frozen:   # frozen routes are irreducible ? exclude from grand total
                for k in grand:
                    grand[k] += bd.get(k, 0.0)

        lines.append(f"\n{'='*70}")
        lines.append(f"GRAND TOTAL (non-frozen)  km={grand['km']:.1f}  milking={grand['milking']:.1f}"
                     f"  shift={grand['shift']:.1f}  overtime={grand['overtime']:.1f}"
                     f"  cap={grand['cap']:.1f}  plant_win={grand['plant_win']:.1f}"
                     f"  total={grand['total']:.1f}")
        self._debug_text.setPlainText("\n".join(lines))

    # -- Block-level capacity inspector ---------------------------------------
    def _on_block_capacity_distribution(self):
        """For every non-preload block, compute its peak load (the same value
        the cap-penalty calculation uses) and report the distribution.

        Logic mirrors the cap-penalty branches in _sheet_cost_breakdown
        (line ~2241): for split blocks we walk the stop sequence tracking a
        running load; for simple blocks the peak is the sum of farm prior_vol.

        The report has:
          1. Summary stats (count, mean, median, stdev, min/max)
          2. Percentiles (50, 75, 80, 85, 90, 95, 99)
          3. Histogram in 1000 L buckets
          4. Threshold sensitivity table (38k -> 45k, # blocks violating + L over)
          5. Top heaviest blocks (route, block index, peak)
          6. Per-route summary (max block load, # blocks, # over current cap)

        Use this to:
          - Pick hard_vol_cap based on actual loads, not a guess.
          - Detect solver gaming: a sharp spike at exactly current_cap-1 means
            the solver is piling loads against the threshold.  A clean tail
            beyond the chosen cap is what you want to see.
        """
        import statistics

        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            self._debug_text.setPlainText("(no file loaded)")
            return

        current_cap = float(self._sw_hard_cap.value())

        # blocks_data: list of dicts with peak, sheet, b_idx, kind, n_farms, n_dests, total_fv
        blocks_data = []
        skipped_preload = 0
        skipped_no_farms = 0

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict):
                continue
            start_time = entry.get("start_time")
            if not start_time:
                continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))

            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block):
                    skipped_preload += 1
                    continue

                farms = block.get("rows", [])
                if not farms:
                    skipped_no_farms += 1
                    continue

                total_fv = sum(
                    (r.get("prior_vol") or 0) for r in farms
                    if isinstance(r.get("prior_vol"), (int, float))
                )

                dests = block.get("dests") or []
                if not dests:
                    dk = block.get("dest_key", "")
                    dests = [{"key": dk, "vol_partial": None}] if dk else []

                # Mirror the peak-load logic from _sheet_cost_breakdown
                if any(d.get("split_after") is not None for d in dests):
                    is_last = (b_idx == len(blocks) - 1)
                    origin  = ("VEDDER" if b_idx == 0 else
                               (_block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"))
                    stops   = _build_block_stops(block, origin, is_last)
                    running = peak = adel = 0.0
                    for stop in stops:
                        if stop["type"] == "farm":
                            v = stop["farm"].get("prior_vol") or 0
                            if isinstance(v, (int, float)):
                                running += v
                                peak = max(peak, running)
                        elif stop["type"] == "dest":
                            off = _dest_vol_partial(stop["dest"], total_fv, adel)
                            adel += off
                            running = max(0.0, running - off)
                    kind = "SPLIT"
                else:
                    peak = total_fv
                    kind = "SIMPLE"

                blocks_data.append({
                    "peak":     peak,
                    "sheet":    sname,
                    "b_idx":    b_idx,
                    "kind":     kind,
                    "n_farms":  len(farms),
                    "n_dests":  len(dests),
                    "total_fv": total_fv,
                })

        if not blocks_data:
            self._debug_text.setPlainText(
                "(no blocks to analyse ? load a file with Modified panel data)")
            return

        peaks = [b["peak"] for b in blocks_data]
        peaks_sorted = sorted(peaks)
        n = len(peaks)

        def pct(xs_sorted, p):
            """Linear-interpolation percentile.  p in [0,100]."""
            if not xs_sorted:
                return 0.0
            if len(xs_sorted) == 1:
                return xs_sorted[0]
            k = (len(xs_sorted) - 1) * (p / 100.0)
            lo = int(k)
            hi = min(lo + 1, len(xs_sorted) - 1)
            frac = k - lo
            return xs_sorted[lo] * (1 - frac) + xs_sorted[hi] * frac

        # -- Build report -----------------------------------------------------
        lines = []
        lines.append("Block Capacity Distribution - Modified panel")
        lines.append("=" * 78)
        lines.append(f"File: {fname}")
        lines.append(f"Current hard_vol_cap setting: {current_cap:,.0f} L  (cap_penalty = {self._sw_cap_pen.value()}/L)")
        lines.append(f"Blocks analysed: {n}   skipped (preload): {skipped_preload}   skipped (empty): {skipped_no_farms}")

        # 1. Summary statistics
        mean   = statistics.mean(peaks)
        median = statistics.median(peaks)
        stdev  = statistics.stdev(peaks) if n > 1 else 0.0
        var    = statistics.variance(peaks) if n > 1 else 0.0
        lines.append("")
        lines.append("-" * 78)
        lines.append("SUMMARY STATISTICS (peak load per block, litres)")
        lines.append("-" * 78)
        lines.append(f"  count:    {n}")
        lines.append(f"  mean:     {mean:>10,.0f}")
        lines.append(f"  median:   {median:>10,.0f}")
        lines.append(f"  stdev:    {stdev:>10,.0f}")
        lines.append(f"  variance: {var:>14,.0f}")
        lines.append(f"  min:      {min(peaks):>10,.0f}")
        lines.append(f"  max:      {max(peaks):>10,.0f}")
        lines.append(f"  range:    {max(peaks) - min(peaks):>10,.0f}")

        # 2. Percentiles
        lines.append("")
        lines.append("-" * 78)
        lines.append("PERCENTILES")
        lines.append("-" * 78)
        for p in (10, 25, 50, 75, 80, 85, 90, 95, 97, 99):
            lines.append(f"  P{p:<2}: {pct(peaks_sorted, p):>10,.0f} L")

        # 3. Histogram in 1000 L buckets
        lines.append("")
        lines.append("-" * 78)
        lines.append("DISTRIBUTION HISTOGRAM (1000 L buckets, .=1 block, current cap marked)")
        lines.append("-" * 78)
        if peaks:
            bucket_size = 1000
            lo_bucket = (int(min(peaks)) // bucket_size) * bucket_size
            hi_bucket = (int(max(peaks)) // bucket_size + 1) * bucket_size
            buckets = {}
            for p in peaks:
                b = (int(p) // bucket_size) * bucket_size
                buckets[b] = buckets.get(b, 0) + 1
            max_count = max(buckets.values()) if buckets else 1
            bar_scale = max(1, max_count // 50)  # cap bar length around 50 chars
            for b in range(lo_bucket, hi_bucket + bucket_size, bucket_size):
                count = buckets.get(b, 0)
                bar_len = count // bar_scale + (1 if count > 0 else 0)
                bar = "#" * bar_len
                cap_marker = "  <- cap" if b <= current_cap < b + bucket_size else ""
                lines.append(f"  {b:>6,} ? {b + bucket_size - 1:>6,}  ({count:>3})  {bar}{cap_marker}")
            if bar_scale > 1:
                lines.append(f"  (each # ~ {bar_scale} blocks)")

        # 4. Threshold sensitivity
        lines.append("")
        lines.append("-" * 78)
        lines.append("THRESHOLD SENSITIVITY  (if hard_vol_cap = T, what penalty results?)")
        lines.append("-" * 78)
        lines.append(f"  {'Threshold':>10}  {'#Over':>6}  {'%Over':>6}  {'L over':>10}  {'Penalty $':>12}  {'Worst block':>12}")
        cap_rate = self._sw_cap_pen.value()
        for T in (38000, 39000, 40000, 41000, 41500, 42000, 42500, 43000, 44000, 45000):
            over = [p for p in peaks if p > T]
            n_over = len(over)
            l_over = sum(p - T for p in over)
            penalty = l_over * cap_rate
            worst = max(over) - T if over else 0
            mark = "  <- current" if abs(T - current_cap) < 0.5 else ""
            lines.append(f"  {T:>10,}  {n_over:>6}  {(100.0*n_over/n):>5.1f}%  {l_over:>10,.0f}  {penalty:>12,.0f}  {worst:>12,.0f}{mark}")

        # 5. Top heaviest blocks
        lines.append("")
        lines.append("-" * 78)
        lines.append("TOP 20 HEAVIEST BLOCKS")
        lines.append("-" * 78)
        lines.append(f"  {'Rank':>4}  {'Sheet':>10}  {'Block':>5}  {'Kind':>6}  {'Farms':>5}  {'Dests':>5}  {'Total fv':>10}  {'Peak':>10}  {'vs cap':>10}")
        top = sorted(blocks_data, key=lambda b: -b["peak"])[:20]
        for i, b in enumerate(top, 1):
            delta = b["peak"] - current_cap
            delta_str = f"+{delta:,.0f}" if delta > 0 else f"{delta:,.0f}"
            lines.append(f"  {i:>4}  {b['sheet']:>10}  {b['b_idx']:>5}  {b['kind']:>6}  {b['n_farms']:>5}  {b['n_dests']:>5}  {b['total_fv']:>10,.0f}  {b['peak']:>10,.0f}  {delta_str:>10}")

        # 6. Per-route summary  ? sheets sorted by max block descending
        lines.append("")
        lines.append("-" * 78)
        lines.append("PER-ROUTE SUMMARY (sheets with at least one block over current cap)")
        lines.append("-" * 78)
        per_sheet = {}
        for b in blocks_data:
            s = b["sheet"]
            d = per_sheet.setdefault(s, {"max": 0, "n_blocks": 0, "n_over": 0, "tot_over": 0.0})
            d["n_blocks"] += 1
            d["max"] = max(d["max"], b["peak"])
            if b["peak"] > current_cap:
                d["n_over"] += 1
                d["tot_over"] += (b["peak"] - current_cap)

        violating = [(s, d) for s, d in per_sheet.items() if d["n_over"] > 0]
        if not violating:
            lines.append(f"  (no routes have blocks over the current cap of {current_cap:,.0f})")
        else:
            lines.append(f"  {'Sheet':>10}  {'Max block':>10}  {'Blocks':>7}  {'#Over':>6}  {'L over':>10}  {'Penalty':>10}")
            for s, d in sorted(violating, key=lambda x: -x[1]["max"]):
                lines.append(f"  {s:>10}  {d['max']:>10,.0f}  {d['n_blocks']:>7}  {d['n_over']:>6}  {d['tot_over']:>10,.0f}  {d['tot_over'] * cap_rate:>10,.0f}")

        # 7. Diagnostic: solver gaming check
        lines.append("")
        lines.append("-" * 78)
        lines.append("GAMING DIAGNOSTIC")
        lines.append("-" * 78)
        # Count blocks within 200L below the current cap.  If the solver is gaming
        # the threshold, you'll see a noticeable spike here.
        near_under = [p for p in peaks if current_cap - 200 < p <= current_cap]
        in_500_under = [p for p in peaks if current_cap - 500 < p <= current_cap]
        in_500_over  = [p for p in peaks if current_cap < p <= current_cap + 500]
        lines.append(f"  Blocks in [cap-200, cap]:    {len(near_under):>3}")
        lines.append(f"  Blocks in [cap-500, cap]:    {len(in_500_under):>3}")
        lines.append(f"  Blocks in (cap, cap+500]:    {len(in_500_over):>3}")
        if len(in_500_under) > 0:
            ratio = len(in_500_under) / max(1, len(in_500_over))
            lines.append(f"  Ratio (under:over) within 500L of cap: {ratio:.2f}")
            if ratio > 4.0 and len(in_500_under) >= 3:
                lines.append("  (!) Sharp asymmetry ? investigate whether solver is piling loads")
                lines.append("    against the threshold rather than spreading them.")
            elif ratio < 1.5:
                lines.append("  OK Loads are spread naturally around the threshold.")
            else:
                lines.append("  Loads moderately clustered below threshold (normal).")
        # Also report std-dev of just the over-cap blocks vs all blocks
        over_peaks = [p for p in peaks if p > current_cap]
        if len(over_peaks) >= 2:
            over_stdev = statistics.stdev(over_peaks)
            lines.append(f"  Stdev of over-cap blocks: {over_stdev:>7,.0f}  "
                         f"(vs overall stdev {stdev:,.0f})")

        self._debug_text.setPlainText("\n".join(lines))

    def _on_intra_route_savings(self):
        """Exhaustively 2-opt reorder farms within each route until convergence.
        Reports km and hours saved vs current Modified panel."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }

        lines = ["Intra-Route Reordering Savings (2-opt until convergence)", "=" * 70]
        total_km_before = total_km_after = 0.0
        total_h_before  = total_h_after  = 0.0
        n_improved = 0

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = copy.deepcopy(self._sheet_mods.get(key, entry.get("blocks", [])))

            # Before: km and hours
            bd_before = _sheet_cost_breakdown(blocks, self.dm, start_time, cfg)
            km_b  = bd_before["km"]
            h_b   = bd_before["shift"]
            ot_b  = bd_before["overtime"]

            # 2-opt each non-frozen block until convergence
            improved = True
            while improved:
                improved = False
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    origin = "VEDDER" if b_idx == 0 else (
                        _block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
                    base_c = _sheet_cost(blocks, self.dm, start_time, cfg)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n - 1):
                        for j in range(i + 1, n):
                            trial_rows = rows[:i] + rows[i:j+1][::-1] + rows[j+1:]
                            blocks[b_idx] = dict(block, rows=trial_rows)
                            c = _sheet_cost(blocks, self.dm, start_time, cfg)
                            if c < best_c:
                                best_c    = c
                                best_rows = trial_rows[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True

                # Also try or-opt (single farm relocation)
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    base_c    = _sheet_cost(blocks, self.dm, start_time, cfg)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n):
                        farm = rows[i]
                        remaining = rows[:i] + rows[i+1:]
                        for j in range(len(remaining) + 1):
                            trial_rows = remaining[:j] + [farm] + remaining[j:]
                            blocks[b_idx] = dict(block, rows=trial_rows)
                            c = _sheet_cost(blocks, self.dm, start_time, cfg)
                            if c < best_c:
                                best_c    = c
                                best_rows = trial_rows[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True

            bd_after = _sheet_cost_breakdown(blocks, self.dm, start_time, cfg)
            km_a  = bd_after["km"]
            h_a   = bd_after["shift"]
            ot_a  = bd_after["overtime"]

            total_km_before += km_b;  total_km_after += km_a
            total_h_before  += h_b;   total_h_after  += h_a

            frozen = all(_is_preload_block(b) or _is_fixed_vol_block(b) for b in blocks)
            if frozen: continue

            delta_km = km_a - km_b
            delta_h  = (h_a + ot_a) - (h_b + ot_b)
            if abs(delta_km) > 0.1 or abs(delta_h) > 0.01:
                n_improved += 1
                lines.append(
                    f"  {sname}  km:{km_b:.1f}->{km_a:.1f}({delta_km:+.1f})"
                    f"  hours:{h_b:.2f}->{h_a:.2f}({delta_h:+.2f})"
                    f"  ot:{ot_b:.1f}->{ot_a:.1f}")

        delta_km_total = total_km_after - total_km_before
        delta_h_total  = total_h_after  - total_h_before
        pct_km = (delta_km_total / total_km_before * 100) if total_km_before else 0
        pct_h  = (delta_h_total  / total_h_before  * 100) if total_h_before  else 0
        lines.append(f"\n{'='*70}")
        lines.append(f"Routes improved: {n_improved}")
        lines.append(
            f"TOTAL km:    {total_km_before:.1f} -> {total_km_after:.1f}"
            f"  ({delta_km_total:+.1f},  {pct_km:+.2f}%)")
        lines.append(
            f"TOTAL hours: {total_h_before:.2f} -> {total_h_after:.2f}"
            f"  ({delta_h_total:+.2f}h,  {pct_h:+.2f}%)")
        self._debug_text.setPlainText("\n".join(lines))

    def _on_overtime_timeline(self):
        """For every route with overtime, print a stop-by-stop timeline."""
        try:
            self._on_overtime_timeline_inner()
        except Exception as ex:
            import traceback
            self._debug_text.setPlainText(
                f"Overtime Timeline crashed:\n{traceback.format_exc()}")

    def _on_overtime_timeline_inner(self):
        """For every route with overtime, print a stop-by-stop timeline."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
        }
        max_sh  = cfg["max_shift_h"]
        suppress = cfg["suppress_no_milking"]
        lines = ["Overtime Timeline ? Modified panel", "=" * 70]
        n_overtime = 0

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))

            ct = calc_times(blocks, self.dm, start_time, self.dm_dur,
                            suppress_no_milking=suppress)
            if ct is None: continue
            all_times, end_dt = ct
            base_dt = datetime.combine(date.today(), start_time)
            shift_h = (end_dt - base_dt).total_seconds() / 3600.0
            if shift_h <= max_sh:
                continue   # no overtime ? skip

            n_overtime += 1
            overtime_h = shift_h - max_sh
            lines.append(f"\n{sname}  shift={shift_h:.2f}h  overtime={overtime_h:.2f}h  "
                         f"start={fmt_hhmm(start_time)}")
            lines.append("-" * 60)

            for b_idx, block in enumerate(blocks):
                btimes  = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes: continue
                is_last = (b_idx == len(blocks) - 1)
                origin  = "VEDDER" if b_idx == 0 else (
                    _block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
                stops   = _build_block_stops(block, origin, is_last)

                if b_idx > 0:
                    lines.append(f"  --- Block {b_idx+1} ---")

                cumul_prev = 0.0
                for s_idx, stop in enumerate(stops):
                    if s_idx >= len(btimes): break
                    t = btimes[s_idx]
                    arr  = fmt_hhmm(t.get("arr"))
                    dep  = fmt_hhmm(t.get("dep"))
                    wait = t.get("wait")
                    stype = stop["type"]

                    # Cumulative shift time at this stop
                    if t.get("arr") is not None:
                        arr_dt2  = datetime.combine(date.today(), t["arr"])
                        if (arr_dt2 - base_dt).total_seconds() < -3600:
                            arr_dt2 += timedelta(days=1)
                        cumul_h = (arr_dt2 - base_dt).total_seconds() / 3600.0
                    else:
                        cumul_h = cumul_prev
                    cumul_prev = cumul_h
                    cumul_s = f"{cumul_h:.2f}h"
                    ot_flag = " <<OT" if cumul_h > max_sh else ""

                    if stype == "origin":
                        name = stop["key"]
                        lines.append(f"  ORIGIN  {name:<20}  dep={dep}  [{cumul_s}]")

                    elif stype == "farm":
                        farm  = stop["farm"]
                        irma  = farm.get("irma","?")
                        loc   = farm.get("location","")[:16]
                        vol   = farm.get("prior_vol", 0)
                        vol   = int(vol) if isinstance(vol, (int, float)) else 0
                        # Determine wait reason
                        if wait and wait > 0:
                            reason = ""
                            # MWO farms skip milking windows so any wait is
                            # from a gate or scheduling gap, not milking.
                            if farm.get("_mwo"):
                                reason = "gate-wait(MWO)"
                            else:
                                # Find which milking window fired
                                for sk, fk in [("m1_start","m1_finish"),
                                               ("m2_start","m2_finish")]:
                                    if farm.get(sk) and farm.get(fk):
                                        arr_t2 = t.get("arr")
                                        if arr_t2 is not None and time_in_window(
                                                arr_t2, farm[sk], farm[fk]):
                                            reason = f"milking({farm[sk]}-{farm[fk]})"
                                            break
                                if not reason:
                                    reason = "wait(unknown)"
                            wait_s = f"  WAIT={wait:.0f}m ({reason})"
                        else:
                            wait_s = ""
                        lines.append(
                            f"  FARM    {irma:<10} {loc:<16} {vol:>7,}L  "
                            f"arr={arr}  dep={dep}{wait_s}  [{cumul_s}{ot_flag}]")

                    elif stype == "dest":
                        d    = stop["dest"]
                        name = (d.get("name","") or d.get("key",""))[:24]
                        lines.append(
                            f"  DEST    {name:<24}  arr={arr}  dep={dep}  [{cumul_s}{ot_flag}]")

                    elif stype == "vedder":
                        lines.append(
                            f"  VEDDER  return                    arr={arr}  [{cumul_s}{ot_flag}]")

        if n_overtime == 0:
            lines.append("\nNo overtime routes found.")
        else:
            lines.append(f"\n{'='*70}")
            lines.append(f"Total routes with overtime: {n_overtime}")
        self._debug_text.setPlainText("\n".join(lines))

    def _on_plant_window_report(self):
        """Output per-processor, per-route plant window costs to the debug text area."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        cfg = {
            "plant_windows":      self._get_plant_windows(),
            "plant_win_penalty":  self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "suppress_no_milking": self._suppress_no_milking_cb.isChecked(),
        }
        plant_windows     = cfg.get("plant_windows", {})
        if not plant_windows:
            self._debug_text.setPlainText("No plant windows configured.")
            return

        plant_win_rate    = cfg.get("plant_win_penalty", 200.0)
        plant_margin_mins = cfg.get("plant_win_margin_mins", 30.0)
        plant_margin_rate = cfg.get("plant_win_margin_rate", plant_win_rate * 0.5)

        from collections import defaultdict
        dest_report = defaultdict(list)
        dest_names  = {}

        for sname, entry in sorted(self._cache[fname].items()):
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))
            ct     = calc_times(blocks, self.dm, start_time, self.dm_dur,
                                suppress_no_milking=cfg.get("suppress_no_milking", True))
            if ct is None: continue
            all_times = ct[0]

            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block): continue
                dests = block.get("dests") or []
                if not dests:
                    dk = block.get("dest_key","")
                    dn = block.get("dest_name","") or dk
                    dests = [{"key": dk, "name": dn}] if dk else []
                n_farms = len(block["rows"])
                btimes  = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes: continue

                for d_i, dest_d in enumerate(dests):
                    dn  = (dest_d.get("name","") or "").strip()
                    dk  = normalise_key(dest_d.get("key","") or "")
                    dest_names[dk] = dn or dk
                    if "yard for" in dn.lower(): continue
                    window = plant_windows.get(dk)
                    if window is None: continue

                    t_idx = _dest_stop_index(block, d_i, b_idx, blocks)
                    ft    = btimes[t_idx] if t_idx < len(btimes) else None
                    if ft is None or ft.get("arr") is None: continue

                    arr    = ft["arr"]
                    arr_s  = f"{arr.hour:02d}:{arr.minute:02d}"
                    arr_dt = datetime.combine(date.today(), arr)
                    open_str, close_str = window

                    pen    = 0.0
                    status = "OK in window"
                    if not time_in_window(arr, open_str, close_str):
                        open_t = parse_hhmm(open_str)
                        if open_t:
                            open_dt = datetime.combine(date.today(), open_t)
                            if open_t <= arr: open_dt += timedelta(days=1)
                            wait_h  = (open_dt - arr_dt).total_seconds() / 3600.0
                        else:
                            wait_h = 1.0
                        pen    = wait_h * plant_win_rate
                        status = f"OUTSIDE  wait={wait_h:.2f}h  pen={pen:.1f}"
                    else:
                        close_t = parse_hhmm(close_str)
                        if close_t:
                            close_dt = datetime.combine(date.today(), close_t)
                            open_t2  = parse_hhmm(open_str)
                            if open_t2 and close_t < open_t2: close_dt += timedelta(days=1)
                            mins_to_close = (close_dt - arr_dt).total_seconds() / 60.0
                            if 0 < mins_to_close < plant_margin_mins:
                                depth  = (plant_margin_mins - mins_to_close) / plant_margin_mins
                                pen    = depth * plant_margin_rate * (plant_margin_mins / 60.0)
                                status = f"margin  {mins_to_close:.0f}m to close  pen={pen:.1f}"
                    dest_report[dk].append((sname, b_idx+1, arr_s, pen, status))

        lines       = ["Plant Window Cost Report ? Modified panel", "=" * 60]
        grand_total = 0.0
        for dk in sorted(dest_report.keys()):
            entries    = dest_report[dk]
            dn         = dest_names.get(dk, dk)
            window     = plant_windows.get(dk, ("?","?"))
            proc_total = sum(e[3] for e in entries)
            grand_total += proc_total
            lines.append(f"\n{dn}  [{dk}]  window={window[0]}-{window[1]}  total={proc_total:.1f}")
            lines.append("-" * 50)
            for sname, b_idx, arr_s, pen, status in sorted(entries, key=lambda x: -x[3]):
                marker = "  <--" if pen > 0 else ""
                lines.append(f"  {sname} b{b_idx}  arr={arr_s}  {status}{marker}")

        lines.append(f"\n{'='*60}")
        lines.append(f"GRAND TOTAL: {grand_total:.1f} km-eq")
        self._debug_text.setPlainText("\n".join(lines))

    def _refresh_debug_tab(self):
        """Rebuild the debug text from the currently active sheet's mod_blocks."""
        fname  = self.file_cb.currentText()
        sname  = self.sheet_cb.currentText()
        blocks = self._mod_blocks
        start  = getattr(self, "_driver_start", None)
        suppress = self._suppress_no_milking_cb.isChecked() \
            if hasattr(self, "_suppress_no_milking_cb") else True

        if not blocks:
            self._debug_text.setPlainText("(no blocks loaded)")
            self._debug_info.setText("")
            return

        lines = []
        lines.append(f"File : {fname}")
        lines.append(f"Sheet: {sname}   Start: {fmt_hhmm(start) if start else '?'}")
        lines.append(f"Suppress no-milking farms: {suppress}")
        lines.append(f"THREE_WINDOW_FARMS loaded : {len(THREE_WINDOW_FARMS)} entries")
        lines.append(f"Vedder depart extra       : +{VEDDER_DEPART_EXTRA_MINS} min")
        lines.append(f"Preload wash              : +{PRELOAD_WASH_MINS} min")
        lines.append(f"Inter-processor break     : +{INTER_PROCESSOR_BREAK} min")
        lines.append("")

        # Compute timing
        ct_result = calc_times(blocks, self.dm, start, dm_dur=self.dm_dur,
                               suppress_no_milking=suppress) if start else None
        all_times, end_cursor = ct_result if ct_result else (None, None)
        all_dists = calc_distances(blocks, self.dm)
        all_durs  = calc_durations(blocks, self.dm_dur)

        total_farms = sum(len(b["rows"]) for b in blocks)
        total_vol   = sum(
            (r["prior_vol"] or 0) for b in blocks for r in b["rows"]
            if isinstance(r.get("prior_vol"), (int, float))
        )
        lines.append(f"Blocks: {len(blocks)}   Farms: {total_farms}   "
                     f"Total vol: {int(total_vol):,} L")
        if end_cursor and start:
            from datetime import datetime, date
            base  = datetime.combine(date.today(), start)
            delta = end_cursor - base
            h, rem = divmod(int(delta.total_seconds()), 3600)
            m = rem // 60
            lines.append(f"Shift end: {fmt_hhmm(end_cursor.time())}   "
                         f"Duration: {h}h {m:02d}m")
        lines.append("")
        lines.append("-" * 72)

        for b_idx, block in enumerate(blocks):
            dists = all_dists[b_idx] if b_idx < len(all_dists) else []
            durs  = all_durs[b_idx]  if b_idx < len(all_durs)  else []
            btimes = all_times[b_idx] if (all_times and b_idx < len(all_times)) else None
            farms = block["rows"]
            dests = block.get("dests") or []
            if not dests:
                dk = block.get("dest_key","")
                dests = [{"name": block.get("dest_name",""), "key": dk, "vol_partial": None}] if dk else []

            is_preload = block.get("preload") and not farms
            tag = " [PRELOAD]" if is_preload else ""
            lines.append(f"BLOCK {b_idx+1}  Route: {block.get('route','?')}{tag}")

            # Origin
            origin_key = "VEDDER" if b_idx == 0 else (
                _block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
            ot = btimes[0] if btimes else None
            origin_extra = f"  (+{VEDDER_DEPART_EXTRA_MINS}m Vedder depart)" if b_idx == 0 else ""
            lines.append(f"  Origin : {origin_key:<20} dep={fmt_hhmm(ot['dep'] if ot else None)}{origin_extra}")

            # Farm stops
            for i, farm in enumerate(farms):
                irma = farm.get("irma","?")
                vol  = farm.get("prior_vol")
                vol_s = f"{int(vol):,}L" if isinstance(vol, (int,float)) else "-"
                ft = btimes[i+1] if (btimes and i+1 < len(btimes)) else None
                arr_s = fmt_hhmm(ft["arr"] if ft else None)
                dep_s = fmt_hhmm(ft["dep"] if ft else None)
                wait  = ft["wait"] if ft else None
                wait_s = f"  wait={int(round(wait))}m" if wait else ""
                dist_s = f"{dists[i]:.1f}km" if i < len(dists) and dists[i] is not None else "?km"
                dur_s  = f"{durs[i]:.0f}m"   if i < len(durs)  and durs[i]  is not None else "?m"
                # Three windows
                w3data = THREE_WINDOW_FARMS.get(irma)
                w3_s   = f"  w3={w3data['w3'][0]}-{w3data['w3'][1]}" if w3data else ""
                no_m_s = "  [no-milking suppressed]" \
                    if (suppress and irma in NO_MILKING_WINDOW_FARMS) else ""
                lines.append(
                    f"  Farm {i+1:2d}: {irma}  {vol_s:<9}  "
                    f"arr={arr_s}  dep={dep_s}  "
                    f"({dist_s}/{dur_s}){wait_s}{w3_s}{no_m_s}"
                )

            # Processor / dest stops
            for d_i, dest_d in enumerate(dests):
                stop_idx = len(farms) + 1 + d_i
                dt = btimes[stop_idx] if (btimes and stop_idx < len(btimes)) else None
                arr_s = fmt_hhmm(dt["arr"] if dt else None)
                dep_s = fmt_hhmm(dt["dep"] if dt else None)
                vp    = dest_d.get("vol_partial")
                vp_s  = f"{int(vp):,}L partial" if isinstance(vp,(int,float)) else "rest"
                dist_v = dists[stop_idx-1] if stop_idx-1 < len(dists) and dists[stop_idx-1] is not None else None
                dur_v  = durs[stop_idx-1]  if stop_idx-1 < len(durs)  and durs[stop_idx-1]  is not None else None
                dist_s = f"{dist_v:.1f}km" if dist_v is not None else "?km"
                dur_s  = f"{dur_v:.0f}m"   if dur_v  is not None else "?m"
                is_yard = "yard for" in (dest_d.get("name","") or "").lower()
                tag2   = " [YARD]" if is_yard else ""
                lines.append(
                    f"  Proc  {d_i+1}: {dest_d.get('name') or dest_d.get('key','?'):<28}  "
                    f"arr={arr_s}  dep={dep_s}  ({dist_s}/{dur_s})  {vp_s}{tag2}"
                )

            # Vedder return (last block)
            if b_idx == len(blocks) - 1 and dests:
                v_idx = len(farms) + len(dests)
                vt = btimes[v_idx] if (btimes and v_idx < len(btimes)) else None
                lines.append(
                    f"  Return : VEDDER                      arr={fmt_hhmm(vt['arr'] if vt else None)}")

            # Preload wash note
            if is_preload:
                lines.append(f"  [wash {PRELOAD_WASH_MINS}m applied after preload offload]")

            lines.append("")

        self._debug_text.setPlainText("\n".join(lines))
        self._debug_info.setText(
            f"Sheet {sname} ? {len(blocks)} block(s), {total_farms} farm(s)  "
            f"| dm keys: {len(self.dm)}  dur keys: {len(self.dm_dur)}"
        )

    def _copy_debug_text(self):
        from PyQt5.QtWidgets import QApplication as _QApp
        _QApp.clipboard().setText(self._debug_text.toPlainText())

    def _on_route_opt_changed(self, state):
        """Single handler for the combined route-corrections checkbox.
        Check  -> split optimization then auto-flag, snapshot into _corrected_blocks.
        Uncheck -> clear _corrected_blocks and fully revert _sheet_mods to raw originals."""
        fname = self.file_cb.currentText()
        if self._chk_route_opt.isChecked():
            self._optimize_all_split_positions()
            self._on_auto_flag_waits()
            # Snapshot BEFORE display so Original panel renders from corrected blocks
            self._snapshot_corrected_blocks(fname)
            self._display_sheet()
            if self.tabs.currentIndex() == 1:
                self._refresh_comparison()
        else:
            # Full revert ? clear both stores
            if fname and fname in self._cache:
                keys = [k for k in self._sheet_mods if k[0] == fname]
                for k in keys:
                    del self._sheet_mods[k]
                corr_keys = [k for k in self._corrected_blocks if k[0] == fname]
                for k in corr_keys:
                    del self._corrected_blocks[k]
                self._display_sheet()
                if self.tabs.currentIndex() == 1:
                    self._refresh_comparison()

    def _snapshot_corrected_blocks(self, fname=None):
        """Copy current _sheet_mods into _corrected_blocks as the immutable baseline."""
        if fname is None:
            fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        for sname in self._cache[fname]:
            key = (fname, sname)
            if key in self._sheet_mods:
                self._corrected_blocks[key] = copy.deepcopy(self._sheet_mods[key])

    def _on_split_opt_checkbox_changed(self, state):
        pass  # replaced by _on_route_opt_changed

    def _on_auto_flag_checkbox_changed(self, state):
        pass  # replaced by _on_route_opt_changed

    def _on_suppress_milking_changed(self):
        """Re-render both tables when the suppress-milking option is toggled."""
        suppress = self._suppress_no_milking_cb.isChecked()
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        orig_entry = self._cache.get(fname, {}).get(sname)
        if orig_entry:
            corrections_on = (hasattr(self, "_chk_route_opt") and
                              self._chk_route_opt.isChecked())
            key = (fname, sname)
            if corrections_on and key in self._corrected_blocks:
                orig_display = copy.deepcopy(self._corrected_blocks[key])
            elif corrections_on and key in self._sheet_mods:
                orig_display = copy.deepcopy(self._sheet_mods[key])
            else:
                orig_display = orig_entry["blocks"]
            populate_table(self.orig_table, orig_display,
                           self.dm, editable=False,
                           start_time=self._driver_start, dm_dur=self.dm_dur,
                           suppress_no_milking=suppress,
                           plant_windows=self._get_plant_windows() if hasattr(self, "_get_plant_windows") else {})
        self._render_editable()
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    # -- Processor demand helpers ---------------------------------------------

    def _refresh_demand_targets(self):
        """Populate/refresh processor demand spinboxes and receiving-window fields
        from the current file totals.

        Each row now shows:
          [Processor name (140px)] [Volume spinbox] [Open HH:MM] [Close HH:MM]

        Receiving-window defaults come from PLANT_RECEIVING_WINDOWS keyed on the
        processor's numeric key, which is extracted from the processor name lookup
        against the blocks in the cache.
        """
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache: return

        # Aggregate current modified totals per processor
        proc_vols, _ = self._agg_file(fname, use_mod=True)
        totals = {p: sum(b.values()) for p, b in proc_vols.items()}

        # Build a map: processor display-name -> numeric key (from blocks)
        proc_key_map = {}   # display_name -> dest_key string
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict): continue
            for block in entry.get("blocks", []):
                for d in (block.get("dests") or []):
                    dk = normalise_key(d.get("key","") or "")
                    dn = d.get("name","") or dk
                    if dn and dk:
                        proc_key_map[dn] = dk

        # Clear old widgets
        for i in reversed(range(self._demand_layout.count())):
            w = self._demand_layout.itemAt(i).widget()
            if w: w.deleteLater()
        self._demand_spinboxes.clear()
        self._demand_open_edits.clear()
        self._demand_close_edits.clear()

        small_f = QFont()
        win_re  = re.compile(r'^\d{1,2}:\d{2}$')   # basic HH:MM check

        for proc in sorted(totals.keys()):
            row = QWidget(); rl = QHBoxLayout(row)
            rl.setContentsMargins(0,0,0,0); rl.setSpacing(4)

            lbl = QLabel(proc); lbl.setFont(small_f)
            lbl.setFixedWidth(110); lbl.setWordWrap(False)
            lbl.setToolTip(proc)

            spin = QDoubleSpinBox()
            spin.setRange(0, 10_000_000)
            spin.setSingleStep(1000)
            spin.setDecimals(0)
            spin.setValue(round(totals.get(proc, 0)))
            spin.setFont(small_f)
            spin.setFixedWidth(88)
            spin.setSuffix(" L")

            # Look up default receiving window via numeric dest key
            dk = proc_key_map.get(proc, "")
            win_default = PLANT_RECEIVING_WINDOWS.get(dk) if dk else None

            open_edit  = QLineEdit()
            close_edit = QLineEdit()
            open_edit.setFont(small_f);  open_edit.setFixedWidth(48)
            close_edit.setFont(small_f); close_edit.setFixedWidth(48)
            open_edit.setAlignment(Qt.AlignCenter)
            close_edit.setAlignment(Qt.AlignCenter)
            open_edit.setPlaceholderText("HH:MM")
            close_edit.setPlaceholderText("HH:MM")

            if win_default:
                open_str, close_str = win_default
                # Convert 00:00/23:59 sentinel back to empty (display as blank = 24/7)
                if open_str == "00:00" and close_str == "23:59":
                    open_edit.setPlaceholderText("24/7")
                    close_edit.setPlaceholderText("24/7")
                else:
                    open_edit.setText(open_str)
                    close_edit.setText(close_str)

            # Light validation colouring on edit
            def _validate(edit=open_edit):
                t = edit.text().strip()
                ok = (not t) or bool(win_re.match(t))
                edit.setStyleSheet("" if ok else "color:#c0392b;")
            open_edit.textChanged.connect(lambda _, e=open_edit: _validate(e))
            close_edit.textChanged.connect(lambda _, e=close_edit: _validate(e))

            rl.addWidget(lbl)
            rl.addWidget(spin)
            rl.addWidget(open_edit)
            rl.addWidget(close_edit)
            rl.addStretch()
            self._demand_layout.addWidget(row)

            self._demand_spinboxes[proc]    = spin
            self._demand_open_edits[proc]   = open_edit
            self._demand_close_edits[proc]  = close_edit

    def _get_demand_targets(self):
        """Return {proc_name: litres} from spinboxes, or None if empty."""
        if not self._demand_spinboxes: return None
        return {p: sb.value() for p, sb in self._demand_spinboxes.items()}

    def _get_plant_windows(self):
        """Build {dest_key: (open_str, close_str)} from the receiving-window
        fields in the Demand panel.

        Only processors where both fields are non-empty (or default 24/7) are
        included.  A processor with blank open AND blank close is treated as
        24/7 and omitted (no penalty applied).
        """
        # We need the key->name reverse map to translate display names back to keys
        fname = self.file_cb.currentText()
        proc_key_map = {}   # display_name -> dest_key
        if fname and fname in self._cache:
            for sname, entry in self._cache[fname].items():
                if not isinstance(entry, dict): continue
                for block in entry.get("blocks", []):
                    for d in (block.get("dests") or []):
                        dk = normalise_key(d.get("key","") or "")
                        dn = d.get("name","") or dk
                        if dn and dk:
                            proc_key_map[dn] = dk

        windows = {}
        win_re  = re.compile(r'^\d{1,2}:\d{2}$')
        for proc, open_edit in self._demand_open_edits.items():
            close_edit = self._demand_close_edits.get(proc)
            if close_edit is None:
                continue
            open_t  = open_edit.text().strip()
            close_t = close_edit.text().strip()
            if not open_t and not close_t:
                continue   # blank = 24/7, no window constraint
            # If only one is filled, fall back to PLANT_RECEIVING_WINDOWS default
            dk = proc_key_map.get(proc, "")
            if not dk:
                continue
            default = PLANT_RECEIVING_WINDOWS.get(dk)
            if not open_t:
                open_t  = default[0] if default else "00:00"
            if not close_t:
                close_t = default[1] if default else "23:59"
            # Validate; skip malformed entries
            open_t2  = f"{int(open_t.split(':')[0]):02d}:{open_t.split(':')[1]}" \
                        if win_re.match(open_t)  else None
            close_t2 = f"{int(close_t.split(':')[0]):02d}:{close_t.split(':')[1]}" \
                        if win_re.match(close_t) else None
            if open_t2 and close_t2:
                windows[dk] = (open_t2, close_t2)
        return windows

    def _refresh_locked_sheets_list(self):
        """Populate the Locked Sheets checkbox list from the currently loaded file."""
        # Clear old checkboxes
        for i in reversed(range(self._lock_layout.count())):
            w = self._lock_layout.itemAt(i).widget()
            if w: w.deleteLater()
        self._locked_sheet_cbs.clear()

        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        small_f = QFont()
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            dc = entry.get("day_colour", "")
            bucket = _sheet_colour_bucket(dc)
            # Only show RED/BLUE sheets ? solver doesn't touch OTHER anyway
            if bucket not in ("RED", "BLUE"):
                continue

            cb = QCheckBox(sname)
            cb.setFont(small_f)
            # Colour the checkbox label to match the sheet colour
            bg, fg, _ = day_colour_style(dc)
            if bg:
                cb.setStyleSheet(
                    f"QCheckBox {{ color: {bg.name()}; font-weight: bold; }}")
            # Pre-tick sheets that are in SOLVER_SKIP_SHEETS (always locked)
            if sname.strip() in SOLVER_SKIP_SHEETS:
                cb.setChecked(True)
                cb.setEnabled(False)
                cb.setToolTip("Always locked (SOLVER_SKIP_SHEETS).")
            self._lock_layout.addWidget(cb)
            self._locked_sheet_cbs[sname] = cb

    # -- Excel export ----------------------------------------------------------

    def _on_export_excel(self):
        fname = self.file_cb.currentText()
        fpath = self._file_map.get(fname)
        if not fpath:
            QMessageBox.warning(self, "Export", "No file loaded.")
            return
        out_path, _ = QFileDialog.getSaveFileName(
            self, "Export Modified Route Sheet", str(fpath.parent / ("modified_" + fname)),
            "Excel Files (*.xlsx)")
        if not out_path: return
        try:
            self._export_xlsx(fpath, out_path, fname)
            QMessageBox.information(self, "Export", "Saved to:\n" + out_path)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _export_xlsx(self, src_path, dst_path, fname):
        """Write a modified copy of the source workbook with _sheet_mods applied.

        Merged cells are read-only in openpyxl, so we unmerge all spans that
        touch the columns we need to write, write the values into the top-left
        cell of each former span, then re-merge everything afterwards.
        """
        wb = openpyxl.load_workbook(src_path, read_only=False, data_only=True)

        WRITE_COLS = {C_IRMA, C_TRAIN, C_M1_START, C_M1_FINISH,
                      C_M2_START, C_M2_FINISH, C_EDPU, C_LOCATION, C_PRIOR_VOL}

        for ws in wb.worksheets:
            sname = ws.title
            if sname.strip().upper() in EXCLUDE_SHEETS: continue
            key = (fname, sname)
            if key not in self._sheet_mods: continue
            mod_blocks = self._sheet_mods[key]

            # -- Collect modified farm rows in sheet order ------------------
            mod_rows = []
            for block in mod_blocks:
                mod_rows.extend(block["rows"])

            # -- Find original IRMA data row numbers ------------------------
            # Scan only up to the last row with content in IRMA col rather
            # than iter_rows() which walks every cell including phantom rows.
            irma_ws_rows = []
            for r in range(1, min(ws.max_row, 5000) + 1):
                val0 = ws.cell(r, C_IRMA).value
                if isinstance(val0, str) and IRMA_RE.match(val0.strip()):
                    irma_ws_rows.append(r)

            if not irma_ws_rows:
                continue

            # Columns that carry per-farm data: the named ones plus any extra
            # columns captured at parse time (farm name, street address, ...).
            # Including the extras in the write/unmerge set makes the *whole*
            # farm row travel when the solver swaps farms ? without this, a
            # farm's name and address stay behind and end up next to a different
            # farm's IRMA.
            # Explicitly exclude WRITE_COLS from extra_cols_in_use so that if
            # _extra_cells ever contains a milking-time or other named column
            # (e.g. from a manually-added farm or an older parsed file), the
            # extra write can never overwrite the named-field write below.
            extra_cols_in_use = set()
            for fd in mod_rows:
                extra_cols_in_use.update((fd.get("_extra_cells") or {}).keys())
            extra_cols_in_use -= WRITE_COLS   # named fields always win
            sheet_write_cols = WRITE_COLS | extra_cols_in_use

            # -- Unmerge any merged ranges that overlap our target rows/cols -
            # Build set of (row, col) tuples we intend to write so we can
            # identify which merges to temporarily dissolve.
            target_cells = set()
            for ws_row in irma_ws_rows:
                for col in sheet_write_cols:
                    target_cells.add((ws_row, col))

            merges_to_redo = []   # list of (min_row, min_col, max_row, max_col)
            for merge_range in list(ws.merged_cells.ranges):
                mr = merge_range
                # Check if any cell in this merge overlaps our targets
                overlaps = any(
                    (r, c) in target_cells
                    for r in range(mr.min_row, mr.max_row + 1)
                    for c in range(mr.min_col, mr.max_col + 1)
                )
                if overlaps:
                    merges_to_redo.append(
                        (mr.min_row, mr.min_col, mr.max_row, mr.max_col))

            for (min_r, min_c, max_r, max_c) in merges_to_redo:
                ws.unmerge_cells(
                    start_row=min_r, start_column=min_c,
                    end_row=max_r,   end_column=max_c)

            # -- Write farm data --------------------------------------------
            def _write_cell(ws_row, col, value):
                ws.cell(ws_row, col).value = value

            for i, ws_row in enumerate(irma_ws_rows):
                if i < len(mod_rows):
                    fd = mod_rows[i]
                    _write_cell(ws_row, C_IRMA,      fd.get("irma", ""))
                    _write_cell(ws_row, C_TRAIN,     fd.get("train", ""))
                    _write_cell(ws_row, C_EDPU,      fd.get("edpu", ""))
                    _write_cell(ws_row, C_LOCATION,  fd.get("location", ""))
                    _write_cell(ws_row, C_PRIOR_VOL, fd.get("prior_vol", None))
                    for col, fkey in [(C_M1_START,"m1_start"),(C_M1_FINISH,"m1_finish"),
                                      (C_M2_START,"m2_start"),(C_M2_FINISH,"m2_finish")]:
                        raw = fd.get(fkey, "")
                        # Write as datetime.time ? cells are formatted h:mm;@ so
                        # they expect a time serial, not a text string.
                        _write_cell(ws_row, col, parse_hhmm(raw) if raw else None)
                    # Extras (farm name, street address, anything else that was
                    # in the row at parse time).  Write what this farm has, and
                    # clear any extra column the new farm doesn't supply so the
                    # previous occupant's value can't bleed through.
                    extras = fd.get("_extra_cells") or {}
                    for col in extra_cols_in_use:
                        _write_cell(ws_row, col, extras.get(col))
                else:
                    for col in sheet_write_cols:
                        _write_cell(ws_row, col, None)

            # -- Write dest rows back (delivery information section) --------
            # Find all numbered delivery rows in this sheet and update volumes.
            # Same targeted scan as above ? avoid iter_rows() on phantom sheets.
            DEST_WRITE_COLS = {C_DEST_VOL, C_DEST_NAME, C_DEST_KEY}
            dest_target = set()
            deliv_data_rows = []  # rows with dest content
            in_deliv = False
            for r in range(1, min(ws.max_row, 5000) + 1):
                c2 = ws.cell(r, 2).value
                if isinstance(c2, str) and "delivery" in c2.lower():
                    in_deliv = True; deliv_data_rows = []; continue
                if in_deliv:
                    dn = ws.cell(r, C_DEST_NAME).value
                    dk = ws.cell(r, C_DEST_KEY).value
                    if dn or dk:
                        deliv_data_rows.append(r)
                    else:
                        in_deliv = False
            # Collect all dests from all mod_blocks in order
            all_mod_dests = []
            for block in mod_blocks:
                dests_b = block.get("dests") or []
                if not dests_b:
                    dk = block.get("dest_key",""); dn = block.get("dest_name","")
                    if dk: dests_b = [{"key":dk,"name":dn,"vol_partial":None}]
                all_mod_dests.extend(dests_b)
            # Unmerge dest rows
            for dr in deliv_data_rows:
                for col in DEST_WRITE_COLS:
                    dest_target.add((dr, col))
            dest_merges = []
            for mr in list(ws.merged_cells.ranges):
                if any((r,c) in dest_target
                       for r in range(mr.min_row, mr.max_row+1)
                       for c in range(mr.min_col, mr.max_col+1)):
                    dest_merges.append((mr.min_row,mr.min_col,mr.max_row,mr.max_col))
            for (min_r,min_c,max_r,max_c) in dest_merges:
                ws.unmerge_cells(start_row=min_r,start_column=min_c,
                                 end_row=max_r,end_column=max_c)
            for i, dr in enumerate(deliv_data_rows):
                if i < len(all_mod_dests):
                    d = all_mod_dests[i]
                    vp = d.get("vol_partial")
                    _write_cell(dr, C_DEST_VOL,  int(vp) if isinstance(vp,(int,float)) else None)
                    _write_cell(dr, C_DEST_NAME, d.get("name",""))
                    _write_cell(dr, C_DEST_KEY,
                                int(d["key"]) if d.get("key","").isdigit() else d.get("key",""))
                else:
                    _write_cell(dr, C_DEST_VOL, None)
                    _write_cell(dr, C_DEST_NAME, None)
                    _write_cell(dr, C_DEST_KEY,  None)
            for (min_r,min_c,max_r,max_c) in dest_merges:
                ws.merge_cells(start_row=min_r,start_column=min_c,
                               end_row=max_r,end_column=max_c)

            # -- Re-merge ---------------------------------------------------
            for (min_r, min_c, max_r, max_c) in merges_to_redo:
                ws.merge_cells(
                    start_row=min_r, start_column=min_c,
                    end_row=max_r,   end_column=max_c)

        wb.save(dst_path)

    def _on_intra_route_apply(self):
        """Launch the IntraRouteOptimiser thread."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }
        self._intra_btn.setEnabled(False)
        self._solve_btn.setEnabled(False)
        self._intra_thread = IntraRouteOptimiser(
            fname, self._cache, self.dm, cfg, self._sheet_mods, parent=self)
        self._intra_thread.progress.connect(self._on_intra_progress)
        self._intra_thread.finished.connect(self._on_intra_finished)
        self._intra_thread.log.connect(self._solver_log.append)
        self._intra_thread.start()

    def _on_intra_progress(self, cur, total, status):
        self._intra_btn.setText(f"Optimising... {cur}/{total}")
        self._solver_status.setText(status)
        if total > 0:
            self._solver_progress.setMaximum(total)
            self._solver_progress.setValue(cur)

    def _on_intra_finished(self, results):
        fname = self.file_cb.currentText()
        for key, blocks in results.items():
            self._sheet_mods[key] = blocks
        self._intra_btn.setEnabled(True)
        self._intra_btn.setText("Optimise Within Routes")
        self._solve_btn.setEnabled(True)
        self._solver_progress.setValue(self._solver_progress.maximum())
        self._solver_status.setText(f"Done ? {len(results)} route(s) improved")
        self._display_sheet()
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    # -- Solver event handlers -------------------------------------------------

    def _on_solve_clicked(self):
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            self._solver_status.setText("No file loaded.")
            return

        groups = _group_sheets_by_colour(self._cache, fname)
        n_red  = len(groups["RED"])
        n_blue = len(groups["BLUE"])
        if n_red == 0 and n_blue == 0:
            self._solver_status.setText("No RED or BLUE sheets found in this file.")
            return

        # Collect user-locked sheets (checked in the Locked Sheets panel)
        locked_sheets = {sname for sname, cb in self._locked_sheet_cbs.items()
                         if cb.isChecked()}

        # Collect plant receiving windows from the demand panel
        plant_windows = self._get_plant_windows()

        cfg = {
            "vol_tol":            self._sw_vol_tol.value(),
            "vol_penalty":        self._sw_vol_pen.value(),
            "hard_vol_cap":       self._sw_hard_cap.value(),
            "cap_penalty":        self._sw_cap_pen.value(),
            "max_shift_h":        self._sw_max_shift.value(),
            "shift_penalty":      self._sw_shift_pen.value(),
            "shift_hours_weight": self._sw_shift_hours.value(),
            "milking_weight":     self._sw_milking.value(),
            "plant_win_penalty":  self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "plant_windows":      plant_windows,
            "iterations":         self._sw_iters.value(),
            "target_cool_frac":   self._sw_cool.value(),
            "segment_size":       100,
            "seed":               (self._sw_seed.value() or None),
            "suppress_no_milking": self._suppress_no_milking_cb.isChecked(),
            "split_opt":          hasattr(self, "_chk_split_opt") and self._chk_split_opt.isChecked(),
        }

        total_iters = cfg["iterations"] * 2   # RED + BLUE
        self._solver_progress.setMaximum(total_iters)
        self._solver_progress.setValue(0)
        self._solver_log.clear()
        # compute alpha for display
        iters = cfg["iterations"]
        tcf   = cfg["target_cool_frac"]
        alpha_display = tcf ** (1.0 / iters) if iters > 1 and tcf > 0 else 0.9999

        locked_list = sorted(locked_sheets - SOLVER_SKIP_SHEETS)
        win_lines   = "\n".join(
            f"    {dk}: {v[0]}?{v[1]}" for dk, v in sorted(plant_windows.items())
        ) or "    (none active)"

        self._solver_log.append(
            f"Solving {fname}\n"
            f"  RED sheets:     {n_red}\n"
            f"  BLUE sheets:    {n_blue}\n"
            f"  Locked sheets:  {locked_list or '(none)'}\n"
            f"  Iterations:     {cfg['iterations']} per group\n"
            f"  Vol tol:        +/-{cfg['vol_tol']*100:.0f}%\n"
            f"  Truck cap:      {cfg['hard_vol_cap']:,} L  (pen {cfg['cap_penalty']:.0f}/L)\n"
            f"  Max shift:      {cfg['max_shift_h']:.1f} h\n"
            f"  Shift hrs wt:   {cfg['shift_hours_weight']:.1f}/h\n"
            f"  Milking wt:     {cfg['milking_weight']:.1f}x\n"
            f"  Plant win pen:  {cfg['plant_win_penalty']:.0f} km/h wait  "
            f"margin={cfg['plant_win_margin_mins']:.0f}min @ {cfg['plant_win_margin_rate']:.0f} km/h\n"
            f"  Plant windows:\n{win_lines}\n"
            f"  Cooling target: {tcf*100:.2f}% of T0\n"
            f"  alpha (per iter):   {alpha_display:.6f}\n"
        )

        self._solve_btn.setEnabled(False)
        self._stop_btn.setEnabled(True)
        self._solver_status.setText("Running...")

        self._solver_thread = ALNSSolver(fname, self._cache, self.dm, cfg,
                                         sheet_mods=self._sheet_mods,
                                         locked_sheets=locked_sheets)
        self._solver_thread.progress.connect(self._on_solver_progress)
        self._solver_thread.finished.connect(self._on_solver_finished)
        self._solver_thread.log.connect(self._on_solver_log)
        self._solver_thread.start()

    def _on_stop_solver(self):
        if self._solver_thread and self._solver_thread.isRunning():
            self._solver_thread.stop()
            self._solver_status.setText("Stopping...")

    def _on_solver_progress(self, cur, total, msg):
        self._solver_progress.setMaximum(total)
        self._solver_progress.setValue(cur)
        self._solver_status.setText(msg)

    def _on_solver_log(self, msg):
        self._solver_log.append(msg)
        # Auto-scroll
        sb = self._solver_log.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _on_solver_finished(self, results):
        """Write solver results into _sheet_mods and refresh the active sheet."""
        self._solve_btn.setEnabled(True)
        self._stop_btn.setEnabled(False)

        fname = self.file_cb.currentText()

        # Build a global uid_map covering every sheet in the file.
        # The solver moves farms freely between sheets of the same colour group,
        # so a farm that ends up in sheet X may have originally come from sheet Y.
        # Looking only at sheet X's orig_blocks would miss its _uid entirely and
        # leave _orig_arr unset, causing phantom milking waits.
        global_uid_map = {}   # uid -> (was_mwo, None)  -- None kept for compat with _stamp
        file_cache = self._cache.get(fname, {})
        for sname_c, entry_c in file_cache.items():
            if not isinstance(entry_c, dict):
                continue
            for block in entry_c.get("blocks", []):
                for farm in block.get("rows", []):
                    uid = farm.get("_uid")
                    if uid:
                        global_uid_map[uid] = (bool(farm.get("_mwo")), None)

        n_updated = 0
        for sname, new_blocks in results.items():
            self._sheet_mods[(fname, sname)] = new_blocks
            self._stamp_orig_arr_from_map(new_blocks, global_uid_map)
            n_updated += 1

        # Log zero-vol farms that were held in place during the solve.
        if hasattr(self._solver_thread, "zero_vol_farms") and \
                self._solver_thread.zero_vol_farms:
            n = len(self._solver_thread.zero_vol_farms)
            irmas = ', '.join(f['irma'] for _, _, _, f
                              in self._solver_thread.zero_vol_farms)
            self._solver_log.append(
                f"  {n} zero-vol farm(s) held with partner: {irmas}")

        self._solver_log.append(
            f"\nOK Done ? {n_updated} sheets updated in Modified panel.")
        self._solver_status.setText(f"OK Complete ? {n_updated} sheets updated")
        self._solver_progress.setValue(self._solver_progress.maximum())

        # Refresh currently displayed sheet if it was touched
        sname = self.sheet_cb.currentText()
        if sname in results:
            self._mod_blocks = self._sheet_mods[(fname, sname)]
            entry = self._cache.get(fname, {}).get(sname, {})
            if isinstance(entry, dict):
                self._driver_start = entry.get("start_time")
            self._render_editable()

        # Also refresh comparison tab if visible
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

        # Auto-flag 2h waits on solved results if checkbox is on
        if hasattr(self, "_chk_auto_flag") and self._chk_auto_flag.isChecked():
            self._on_auto_flag_waits()
        # Re-optimize split positions after solve
        if hasattr(self, "_chk_split_opt") and self._chk_split_opt.isChecked():
            self._optimize_all_split_positions(fname)

    def _init_route_table(self, t, editable=False):
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setAlternatingRowColors(False)
        t.verticalHeader().setVisible(False)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        t.setShowGrid(True)

    def _init_comp_table(self, t):
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.verticalHeader().setVisible(False)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        t.horizontalHeader().setStretchLastSection(True)
        t.setShowGrid(True)
        t.setAlternatingRowColors(True)

    # -------------------------------------------------------------------------
    # Spinner
    # -------------------------------------------------------------------------

    def _start_spinner(self, msg):
        self._spin_msg = msg; self._spin_idx = 0
        self.load_btn.setEnabled(False); self._spin_timer.start(80)

    def _stop_spinner(self):
        self._spin_timer.stop(); self.load_btn.setEnabled(True)

    def _tick_spinner(self):
        ch = self._spin_chars[self._spin_idx % len(self._spin_chars)]
        self.statusBar().showMessage(f"{ch}  {self._spin_msg}")
        self._spin_idx += 1

    # -------------------------------------------------------------------------
    # Folder scanning / dropdown population
    # -------------------------------------------------------------------------

    def _year_folders(self):
        result = {}
        if self.data_root.exists():
            for d in sorted(self.data_root.iterdir()):
                if d.is_dir(): result[extract_year(d.name)] = d
        return result

    def _month_folders(self, ypath):
        result = {}
        if ypath and ypath.exists():
            for d in sorted([d for d in ypath.iterdir() if d.is_dir()],
                            key=lambda d: month_key(d.name)):
                result[d.name] = d
        return result

    def _xlsx_files(self, mpath):
        result = {}
        if mpath and mpath.exists():
            for f in sorted(mpath.glob("*.xlsx")):
                result[f.name] = f
        return result

    def _block_sigs(self, b):
        for w in (self.year_cb, self.month_cb, self.file_cb, self.sheet_cb):
            w.blockSignals(b)

    def _on_browse_folder(self):
        """Let the user pick a new root data folder and rescan."""
        chosen = QFileDialog.getExistingDirectory(
            self,
            "Select root data folder",
            str(self.data_root) if self.data_root.exists() else str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks,
        )
        if not chosen:
            return   # user cancelled
        self.data_root = Path(chosen)
        self._scan_folders()
        # If no year folders were found, warn the user
        if not self._year_map:
            self.statusBar().showMessage(
                f"(!)  No year sub-folders found in '{chosen}'. "
                "Expected: <root>/<year>/<month>/<file>.xlsx")

    def _scan_folders(self):
        self._block_sigs(True)
        self._year_map = self._year_folders()
        self.year_cb.clear()
        self.year_cb.addItems(sorted(self._year_map.keys(), reverse=True))
        self._block_sigs(False)
        self._pop_months(); self._pop_files()

    def _pop_months(self, restore=None):
        self._block_sigs(True)
        self.month_cb.clear()
        self._month_map = self._month_folders(self._year_map.get(self.year_cb.currentText()))
        self.month_cb.addItems(list(self._month_map.keys()))
        if restore:
            idx = self.month_cb.findText(restore)
            if idx >= 0: self.month_cb.setCurrentIndex(idx)
        self._block_sigs(False)

    def _pop_files(self, restore=None):
        self._block_sigs(True)
        self.file_cb.clear()
        self._file_map = self._xlsx_files(self._month_map.get(self.month_cb.currentText()))
        self.file_cb.addItems(list(self._file_map.keys()))
        if restore:
            idx = self.file_cb.findText(restore)
            if idx >= 0: self.file_cb.setCurrentIndex(idx)
        self._block_sigs(False)
        self._pop_sheets()

    def _pop_sheets(self, restore=None):
        self._block_sigs(True)
        self.sheet_cb.clear()
        fname = self.file_cb.currentText()
        if fname in self._cache:
            for i, (sname, entry) in enumerate(self._cache[fname].items()):
                self.sheet_cb.addItem(sname)
                dc = entry.get("day_colour", "") if isinstance(entry, dict) else ""
                bg, fg, _ = day_colour_style(dc)
                if bg:
                    self.sheet_cb.setItemData(i, bg, Qt.BackgroundRole)
                    self.sheet_cb.setItemData(i, fg, Qt.ForegroundRole)
        if restore:
            idx = self.sheet_cb.findText(restore)
            if idx >= 0: self.sheet_cb.setCurrentIndex(idx)
        self._block_sigs(False)

    # -------------------------------------------------------------------------
    # IRMA Search
    # -------------------------------------------------------------------------

    # Highlight colours used for search results
    _SEARCH_HIT_BG    = QColor("#fff176")   # yellow  ? every match
    _SEARCH_CURSOR_BG = QColor("#f57f17")   # amber   ? the currently-navigated match

    def _on_search_text_changed(self):
        """Clear highlights as soon as the user edits the query so stale
        results don't linger.  Don't re-search on every keypress ? wait for
        Enter so partial IRMA numbers (e.g. typing '71') don't scroll the
        table around while the user is still typing."""
        if not self._search_box.text().strip():
            self._on_search_clear()

    def _on_search(self):
        """Run the search: find every farm row whose IRMA matches the query
        (case-insensitive substring) in both the Original and Modified tables,
        highlight them all, and scroll to the first hit."""
        query = self._search_box.text().strip().upper()
        self._search_clear_highlights()
        self._search_hits   = []
        self._search_cursor = -1

        if not query:
            self._search_status.setText("")
            for btn in (self._search_prev_btn, self._search_next_btn,
                        self._search_clear_btn):
                btn.setEnabled(False)
            return

        irma_col = next((i for i, (_, k) in enumerate(COLS) if k == "irma"), 0)

        for table in (self.orig_table, self.edit_table):
            for row in range(table.rowCount()):
                item = table.item(row, irma_col)
                if item is None:
                    continue
                # Only match real farm rows (they carry UserRole farm data)
                if item.data(Qt.UserRole) is None:
                    continue
                if query in item.text().upper():
                    self._search_hits.append((table, row))
                    # Apply hit highlight to every cell in this row
                    for col in range(table.columnCount()):
                        ci = table.item(row, col)
                        if ci is not None:
                            ci.setBackground(self._SEARCH_HIT_BG)

        n = len(self._search_hits)
        if n == 0:
            self._search_status.setText("No match")
            self._search_status.setStyleSheet("color: #c62828; font-weight: bold;")
            for btn in (self._search_prev_btn, self._search_next_btn):
                btn.setEnabled(False)
            self._search_clear_btn.setEnabled(True)
            return

        self._search_status.setStyleSheet("")
        for btn in (self._search_prev_btn, self._search_next_btn,
                    self._search_clear_btn):
            btn.setEnabled(True)

        # Jump to first hit
        self._search_cursor = 0
        self._search_apply_cursor()

    def _on_search_next(self):
        if not self._search_hits:
            return
        self._search_cursor = (self._search_cursor + 1) % len(self._search_hits)
        self._search_apply_cursor()

    def _on_search_prev(self):
        if not self._search_hits:
            return
        self._search_cursor = (self._search_cursor - 1) % len(self._search_hits)
        self._search_apply_cursor()

    def _search_apply_cursor(self):
        """Update the status label, re-colour all hits (current one gets
        a stronger amber), and scroll the relevant table to show it."""
        irma_col = next((i for i, (_, k) in enumerate(COLS) if k == "irma"), 0)
        n = len(self._search_hits)

        for idx, (table, row) in enumerate(self._search_hits):
            bg = self._SEARCH_CURSOR_BG if idx == self._search_cursor \
                 else self._SEARCH_HIT_BG
            for col in range(table.columnCount()):
                ci = table.item(row, col)
                if ci is not None:
                    ci.setBackground(bg)

        cur_table, cur_row = self._search_hits[self._search_cursor]
        cur_table.scrollToItem(cur_table.item(cur_row, irma_col),
                               QAbstractItemView.PositionAtCenter)
        cur_table.setCurrentCell(cur_row, irma_col)

        # Count hits per table for the label e.g. "3 / 5  (Orig: 2  Mod: 3)"
        orig_hits = sum(1 for t, _ in self._search_hits if t is self.orig_table)
        mod_hits  = sum(1 for t, _ in self._search_hits if t is self.edit_table)
        label     = f"{self._search_cursor + 1} / {n}"
        if orig_hits and mod_hits:
            label += f"  (Orig: {orig_hits}  Mod: {mod_hits})"
        elif orig_hits:
            label += "  (Original only)"
        elif mod_hits:
            label += "  (Modified only)"
        self._search_status.setText(label)

    def _search_clear_highlights(self):
        """Remove search highlight colours from all previously-highlighted rows.
        Called before a new search and on clear so the table returns to its
        normal row colours."""
        irma_col = next((i for i, (_, k) in enumerate(COLS) if k == "irma"), 0)
        for table, row in self._search_hits:
            # Re-read the natural background from the IRMA cell's current data
            # so we don't have to recompute the full row colour.  The IRMA cell
            # background was set by populate_table and doesn't need to be exact;
            # clearing to white is acceptable ? the next _display_sheet call
            # will restore proper colours anyway.  For a cleaner restore we
            # delegate back to populate_table by triggering _display_sheet,
            # but that's expensive.  Instead, restore to the standard alternating
            # block colour based on whether the b_idx stored in UserRole is even/odd.
            item = table.item(row, irma_col)
            if item is None:
                continue
            farm_data = item.data(Qt.UserRole)
            if isinstance(farm_data, tuple) and len(farm_data) >= 1:
                b_idx = farm_data[0]
                natural_bg = QColor("#e3f2fd") if b_idx % 2 == 0 else QColor("#ffffff")
            else:
                natural_bg = QColor("#ffffff")
            for col in range(table.columnCount()):
                ci = table.item(row, col)
                if ci is not None:
                    ci.setBackground(natural_bg)

    def _on_search_clear(self):
        """Clear the search box, remove highlights, and reset state."""
        self._search_clear_highlights()
        self._search_hits   = []
        self._search_cursor = -1
        self._search_box.clear()
        self._search_status.setText("")
        self._search_status.setStyleSheet("")
        for btn in (self._search_prev_btn, self._search_next_btn,
                    self._search_clear_btn):
            btn.setEnabled(False)

    def _on_irma_lookup(self):
        """Open the IRMA Farm Lookup dialog against all currently loaded data."""
        dlg = IRMALookupDialog(self._cache, self._sheet_mods, parent=self)
        dlg.navigate_requested.connect(self._navigate_to_sheet)
        # Pre-fill with the search box text if the user already has one there
        if hasattr(self, "_search_box") and self._search_box.text().strip():
            dlg._query.setText(self._search_box.text().strip())
            dlg._run_search()
        dlg.exec_()

    def _navigate_to_sheet(self, fname, sname):
        """Navigate the main window to the given file and sheet.
        Called from IRMALookupDialog when the user double-clicks a result."""
        # Select the file in the file combo
        fidx = self.file_cb.findText(fname)
        if fidx < 0:
            QMessageBox.warning(self, "Not loaded",
                f"'{fname}' is not currently loaded.\n"
                f"Use > Load File to load it first.")
            return
        if self.file_cb.currentIndex() != fidx:
            self._block_sigs(True)
            self.file_cb.setCurrentIndex(fidx)
            self._block_sigs(False)
            self._pop_sheets(restore=sname)
        # Select the sheet
        sidx = self.sheet_cb.findText(sname)
        if sidx >= 0:
            self.sheet_cb.setCurrentIndex(sidx)
        self._display_sheet()
        # Switch to the Route tab so the user can see the sheet
        self.tabs.setCurrentIndex(0)

    # -------------------------------------------------------------------------
    # File loading
    # -------------------------------------------------------------------------

    def _on_load_clicked(self):
        fname = self.file_cb.currentText()
        fpath = self._file_map.get(fname)
        if not fpath or (self._loader and self._loader.isRunning()): return
        self._start_spinner(f"Loading {fname}...")
        self._load_warnings = []   # reset warning accumulator for this load
        self._loader = FileLoader(fname, fpath)
        self._loader.done.connect(self._on_load_done)
        self._loader.failed.connect(self._on_load_failed)
        self._loader.sheet_warning.connect(self._on_sheet_warning)
        self._loader.log.connect(
            lambda msg: self._debug_text.append(msg)
            if hasattr(self, "_debug_text") else None)
        self._loader.start()

    def _on_sheet_warning(self, fname, sheet_name, message):
        """Accumulate per-sheet parse warnings ? shown as one summary after load,
        and also appended to the debug-tab text area in real time."""
        self._load_warnings.append((sheet_name, message))
        if hasattr(self, "_debug_text"):
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%H:%M:%S")
            self._debug_text.append(
                f"[{ts}] LOAD WARNING  {fname} / {sheet_name}\n  {message}\n")

    def _on_load_done(self, fname, sheets):
        self._stop_spinner()
        self._cache[fname] = sheets
        # Clear any stale modified state for this file so Modified starts
        # identical to Original on every fresh load.
        stale_keys = [k for k in self._sheet_mods if k[0] == fname]
        for k in stale_keys:
            del self._sheet_mods[k]
        corr_keys = [k for k in self._corrected_blocks if k[0] == fname]
        for k in corr_keys:
            del self._corrected_blocks[k]
        self._pop_sheets()
        self._display_sheet()
        # Must refresh demand targets first so plant windows are populated
        self._refresh_demand_targets()
        self._refresh_locked_sheets_list()
        self._populate_irma_dropdown()
        self._populate_proc_dropdown()
        # Optimize partial-dropoff split positions FIRST
        if hasattr(self, "_chk_split_opt") and self._chk_split_opt.isChecked():
            self._optimize_all_split_positions(fname)
        # Auto-flag 2h waits only if checkbox is checked (default: off)
        if hasattr(self, "_chk_auto_flag") and self._chk_auto_flag.isChecked():
            self._on_auto_flag_waits()
        # Snapshot corrected baseline so solver results don't overwrite it
        if hasattr(self, "_chk_route_opt") and self._chk_route_opt.isChecked():
            self._snapshot_corrected_blocks(fname)
            self._display_sheet()   # refresh Original panel from corrected blocks
        # Keep comparison tab fresh if it happens to be open
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

        # Show any parse warnings that accumulated during the load.
        # One dialog with all issues is less annoying than one per sheet.
        warnings = getattr(self, "_load_warnings", [])
        if warnings:
            lines = []
            for sheet_name, msg in warnings:
                lines.append(f"Sheet '{sheet_name}':\n  {msg}")
            body = "\n\n".join(lines)
            QMessageBox.warning(
                self,
                f"Load warnings ? {fname}",
                f"The following issues were found while loading {fname}.\n"
                f"Affected sheets may appear empty or have missing data.\n\n"
                f"{body}"
            )
            self._load_warnings = []

    def _optimize_all_split_positions(self, fname=None):
        """Run _optimize_split_positions on every sheet in the file,
        using current plant window config from the UI.  Updates _sheet_mods
        in place and refreshes the displayed sheet if it changed."""
        if fname is None:
            fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        plant_windows = self._get_plant_windows()
        if not plant_windows:
            return   # no windows configured ? nothing to optimise
        cfg = {
            "plant_windows":         plant_windows,
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }
        any_changed = False
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            start_time = entry.get("start_time")
            if not start_time:
                continue
            key = (fname, sname)
            if key not in self._sheet_mods:
                self._sheet_mods[key] = copy.deepcopy(entry["blocks"])
            blocks = self._sheet_mods[key]
            if _optimize_split_positions(blocks, self.dm, start_time, cfg, self.dm_dur):
                any_changed = True
        if any_changed:
            self._display_sheet()
            if self.tabs.currentIndex() == 1:
                self._refresh_comparison()

    def _on_load_failed(self, fname, err):
        self._stop_spinner()
        self.statusBar().showMessage(f"Error loading {fname}: {err}")
        if hasattr(self, "_debug_text"):
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%H:%M:%S")
            self._debug_text.append(
                f"[{ts}] LOAD ERROR  {fname}\n  {err}\n")

    def _update_day_colour_badge(self):
        bg, fg, text = day_colour_style(self._day_colour)
        if bg:
            self.day_colour_box.setText(text)
            self.day_colour_box.setStyleSheet(
                f"background-color: {bg.name()}; color: {fg.name()}; "
                f"border-radius: 4px; padding: 2px 6px; font-weight: bold;")
        else:
            self.day_colour_box.setText("")
            self.day_colour_box.setStyleSheet(
                "border-radius: 4px; padding: 2px 6px;")

    def _display_sheet(self):
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if not fname or not sname or fname not in self._cache: return
        entry = self._cache[fname].get(sname)
        if not entry: return
        # Changing sheet invalidates any previous search results ? clear them
        # silently (no UI flash) before the tables are repopulated.
        if hasattr(self, "_search_hits") and self._search_hits:
            self._search_hits   = []
            self._search_cursor = -1
            self._search_status.setText("")
            self._search_status.setStyleSheet("")
            for btn in (self._search_prev_btn, self._search_next_btn,
                        self._search_clear_btn):
                btn.setEnabled(False)
        blocks     = entry["blocks"]
        self._driver_start = entry["start_time"]
        self._day_colour   = entry.get("day_colour", "")
        self._update_day_colour_badge()
        self.farm_tray.refresh_bold_state(self._day_colour)
        suppress = self._suppress_no_milking_cb.isChecked() \
            if hasattr(self, "_suppress_no_milking_cb") else True
        # Restore saved mod_blocks for this sheet, or start fresh
        key = (fname, sname)
        if key not in self._sheet_mods:
            self._sheet_mods[key] = copy.deepcopy(blocks)
        self._mod_blocks = self._sheet_mods[key]

        # Original panel: when route corrections are active, render from the
        # corrected baseline (_corrected_blocks) ? never from solver output.
        # When corrections are off, render from raw cache blocks.
        corrections_on = (hasattr(self, "_chk_route_opt") and
                          self._chk_route_opt.isChecked())
        if corrections_on and key in self._corrected_blocks:
            orig_display = copy.deepcopy(self._corrected_blocks[key])
        elif corrections_on and key in self._sheet_mods:
            # Corrections on but not snapshotted yet ? use _sheet_mods
            orig_display = copy.deepcopy(self._sheet_mods[key])
        else:
            orig_display = copy.deepcopy(blocks)

        populate_table(self.orig_table, orig_display, self.dm, editable=False,
                       start_time=self._driver_start, dm_dur=self.dm_dur,
                       suppress_no_milking=suppress,
                       plant_windows=self._get_plant_windows() if hasattr(self, "_get_plant_windows") else {})
        self._render_editable()
        # Tray is NOT cleared ? removed farms persist across sheet switches
        total = sum(len(b["rows"]) for b in blocks)
        st = fmt_hhmm(self._driver_start) if self._driver_start else "?"
        self.statusBar().showMessage(
            f"{fname}  /  {sname}  ?  {len(blocks)} route(s), {total} farm(s)  |  Start: {st}"
            + ("" if self.dm else "  (!) distance_matrix.csv not found"))

    def _stamp_orig_arr(self, orig_blocks, mod_blocks, start_time):
        """After the solver runs: restore _mwo flags and _orig_arr
        baselines for any farm row that had them set before the solve.

        Each farm row is assigned a unique _uid at parse time (see parse_sheet).
        The solver preserves _uid through deepcopy, so we can look up any row
        in the original schedule regardless of which sheet or position it ended
        up on after optimisation ? no IRMA matching, no occurrence counting,
        no milking-signature heuristics needed.
        """
        if not orig_blocks:
            return

        # Build map: _uid -> was_mwo_checked
        uid_map = {}
        for block in orig_blocks:
            for farm in block.get("rows", []):
                uid = farm.get("_uid")
                if uid:
                    uid_map[uid] = bool(farm.get("_mwo"))

        # Restore _mwo on mod_blocks (solver deepcopies clear it)
        for block in mod_blocks:
            for farm in block.get("rows", []):
                uid = farm.get("_uid")
                if uid and uid_map.get(uid):
                    farm["_mwo"] = True

    def _stamp_orig_arr_from_map(self, mod_blocks, uid_map):
        """Restore _mwo flags on mod_blocks using a pre-built uid map.

        uid_map: {uid: (was_checked, _)} ? second element ignored (no _orig_arr).
        """
        for block in mod_blocks:
            for farm in block.get("rows", []):
                uid = farm.get("_uid")
                if not uid or uid not in uid_map:
                    continue
                was_checked, _ = uid_map[uid]
                if was_checked:
                    farm["_mwo"] = True


    def _save_mod_blocks(self):
        """Persist current mod_blocks to _sheet_mods for the active sheet."""
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if fname and sname and self._mod_blocks is not None:
            self._sheet_mods[(fname, sname)] = self._mod_blocks

    def _render_editable(self):
        self._save_mod_blocks()
        self.edit_table.blockSignals(True)
        _sup = (self._suppress_no_milking_cb.isChecked()
                if hasattr(self, "_suppress_no_milking_cb") else True)
        populate_table(self.edit_table, self._mod_blocks, self.dm, editable=True,
                       start_time=getattr(self, "_driver_start", None),
                       dm_dur=self.dm_dur,
                       suppress_no_milking=_sup,
                       plant_windows=self._get_plant_windows() if hasattr(self, "_get_plant_windows") else {})
        self.edit_table.blockSignals(False)
        # Show status warning if any block has no processor
        no_proc = [b for b in (self._mod_blocks or [])
                   if not (b.get("dests") or b.get("dest_key",""))]
        if no_proc:
            routes = ", ".join(b.get("route","?") for b in no_proc)
            self._add_status.setText(
                f"(!)  Route(s) {routes} have no processor ? drag one from the tray")
            self._add_status_timer.start(8000)

    # -------------------------------------------------------------------------
    # Farm removal / reinsertion
    # -------------------------------------------------------------------------

    def _on_manual_add(self):
        """Validate and add a manually entered farm row to the tray."""
        irma_raw = self._add_irma.currentText().strip()
        # Validate IRMA format
        if not IRMA_RE.match(irma_raw):
            self._add_irma.lineEdit().setStyleSheet("color: #c0392b; font-weight: bold;")
            self._add_status.setText(f"IRMA '{irma_raw}' is not a valid IRMA number (format: ##-###)")
            self._add_status_timer.start(4000)
            return
        # Validate IRMA is in distance matrix
        from_dm = bool(self._dm_keys) and (irma_raw in self._dm_keys)
        if not from_dm:
            self._add_irma.lineEdit().setStyleSheet("color: #c0392b; font-weight: bold;")
            self._add_status.setText(f"IRMA '{irma_raw}' not found in distance matrix")
            self._add_status_timer.start(4000)
            return
        # Valid
        self._add_irma.lineEdit().setStyleSheet("")
        self._add_status.setText("")
        vol_raw = self._add_vol.text().strip()
        try:
            vol = float(vol_raw.replace(",", "")) if vol_raw else None
        except ValueError:
            vol = None
        farm = {
            "irma":      irma_raw,
            "train":     self._add_train.text().strip(),
            "m1_start":  self._add_m1s.text().strip(),
            "m1_finish": self._add_m1f.text().strip(),
            "m2_start":  self._add_m2s.text().strip(),
            "m2_finish": self._add_m2f.text().strip(),
            "edpu":      self._add_edpu.text().strip(),
            "location":  self._add_loc.text().strip(),
            "prior_vol": vol,
        }
        farm["_from_block"]  = -1
        farm["_from_fname"]  = self.file_cb.currentText()
        farm["_from_sname"]  = self.sheet_cb.currentText()
        farm["_from_colour"] = self._day_colour
        self._removed.append(farm)
        self.farm_tray.add_farm(farm, "(manual)", self._day_colour)
        # Clear fields for fast multi-entry
        for w in (self._add_train, self._add_m1s, self._add_m1f,
                  self._add_m2s, self._add_m2f, self._add_edpu,
                  self._add_loc, self._add_vol):
            w.clear()
        self._add_irma.setCurrentIndex(-1)
        self._add_irma.lineEdit().clear()
        self._add_irma.setFocus()

    def _on_del_btn_state(self):
        """Enable Delete Selected if anything deletable is selected in either
        the Modified table or the tray."""
        edit_sel = self._edit_table_deletable_item()
        tray_sel = self.farm_tray.currentRow() >= 0
        self._tray_del_btn.setEnabled(edit_sel is not None or tray_sel)

    def _edit_table_deletable_item(self):
        """Return the currently selected QTableWidgetItem in the Modified table
        if it is a farm, dest, or block banner row ? else None."""
        item = self.edit_table.currentItem()
        if item is None:
            return None
        # Block banner
        if item.data(Qt.UserRole + 2) is not None:
            return item
        # Dest row
        dd = item.data(Qt.UserRole + 1)
        if dd is not None and dd[0] == "dest":
            return item
        # Farm row
        if item.data(Qt.UserRole) is not None:
            return item
        return None

    def _on_tray_delete(self):
        """Delete the selected item from the Modified table OR the tray,
        in that priority order (Modified table takes precedence)."""
        # -- Modified table delete -----------------------------------------
        edit_item = self._edit_table_deletable_item()
        if edit_item is not None and self._mod_blocks is not None:
            # Block banner -> delete entire block
            b_idx_banner = edit_item.data(Qt.UserRole + 2)
            if b_idx_banner is not None:
                if 0 <= b_idx_banner < len(self._mod_blocks):
                    self._mod_blocks.pop(b_idx_banner)
                    self._save_mod_blocks()
                    self._render_editable()
                return

            # Dest row -> remove dest from its block
            dd = edit_item.data(Qt.UserRole + 1)
            if dd is not None and dd[0] == "dest":
                _, b_idx, d_idx = dd
                if 0 <= b_idx < len(self._mod_blocks):
                    dests = self._mod_blocks[b_idx].get("dests") or []
                    if 0 <= d_idx < len(dests):
                        dests.pop(d_idx)
                        if dests:
                            self._mod_blocks[b_idx]["dest_name"] = dests[0]["name"]
                            self._mod_blocks[b_idx]["dest_key"]  = dests[0]["key"]
                        else:
                            self._mod_blocks[b_idx]["dest_name"] = ""
                            self._mod_blocks[b_idx]["dest_key"]  = ""
                    self._save_mod_blocks()
                    self._render_editable()
                return

            # Farm row -> remove farm from its block
            fd = edit_item.data(Qt.UserRole)
            if fd is not None:
                b_idx, f_idx = fd
                if 0 <= b_idx < len(self._mod_blocks):
                    rows = self._mod_blocks[b_idx]["rows"]
                    if 0 <= f_idx < len(rows):
                        rows.pop(f_idx)
                    self._save_mod_blocks()
                    self._render_editable()
                return

        # -- Tray delete ---------------------------------------------------
        row = self.farm_tray.currentRow()
        if row < 0:
            return
        if row < len(self._removed):
            self._removed.pop(row)
        self.farm_tray.removeRow(row)
        self._tray_del_btn.setEnabled(self.farm_tray.currentRow() >= 0)

    def _on_mwo_changed(self, item):
        """Handle the MWO (Milking Window Override) checkbox being toggled.

        When checked, the farm's milking windows are ignored entirely ?
        the truck arrives and pumps without waiting.  No _orig_arr is needed.
        """
        if item.column() != MWO_COL:
            return
        ud = item.data(Qt.UserRole)
        if ud is None:
            return
        b_idx, f_idx = ud
        if self._mod_blocks is None or b_idx >= len(self._mod_blocks):
            return
        rows = self._mod_blocks[b_idx]["rows"]
        if f_idx >= len(rows):
            return
        farm    = rows[f_idx]
        checked = (item.checkState() == Qt.Checked)
        farm["_mwo"] = checked

        # Sync to the original cache so both panels show the same state
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        orig_entry = self._cache.get(fname, {}).get(sname)
        if orig_entry:
            uid = farm.get("_uid")
            for orig_block in orig_entry.get("blocks", []):
                for orig_farm in orig_block.get("rows", []):
                    if uid and orig_farm.get("_uid") == uid:
                        orig_farm["_mwo"] = checked
                        break
        # Re-render both panels and refresh comparison
        _sup = self._suppress_no_milking_cb.isChecked()
        populate_table(self.orig_table, orig_entry["blocks"] if orig_entry else [],
                       self.dm, editable=False,
                       start_time=self._driver_start, dm_dur=self.dm_dur,
                       suppress_no_milking=_sup)
        self._render_editable()
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    def _populate_irma_dropdown(self):
        """Build _irma_lookup from all loaded cache entries and populate the
        IRMA combo with sorted IRMA numbers.  Called whenever a file is loaded.

        _irma_lookup: {irma_str: {train, m1_start, m1_finish, m2_start,
                                   m2_finish, edpu, location}}
        Only the most-recently-seen data for each IRMA is kept (all files
        in the session are merged so previous loads stay available).
        """
        if not hasattr(self, "_irma_lookup"):
            self._irma_lookup = {}

        for fname, sheets in self._cache.items():
            for sname, entry in sheets.items():
                if not isinstance(entry, dict):
                    continue
                for block in entry.get("blocks", []):
                    for row in block.get("rows", []):
                        irma = row.get("irma", "").strip()
                        if not irma:
                            continue
                        self._irma_lookup[irma] = {
                            "train":     row.get("train", ""),
                            "m1_start":  row.get("m1_start", ""),
                            "m1_finish": row.get("m1_finish", ""),
                            "m2_start":  row.get("m2_start", ""),
                            "m2_finish": row.get("m2_finish", ""),
                            "edpu":      row.get("edpu", ""),
                            "location":  row.get("location", ""),
                        }

        # Repopulate combo (block signals to avoid spurious autofill)
        self._add_irma.blockSignals(True)
        current_text = self._add_irma.currentText()
        self._add_irma.clear()
        for irma in sorted(self._irma_lookup.keys()):
            self._add_irma.addItem(irma)
        # Restore whatever the user had typed
        self._add_irma.lineEdit().setText(current_text)
        self._add_irma.blockSignals(False)

    def _on_irma_autofill(self, index):
        """When the user picks an IRMA from the dropdown, fill all known fields."""
        irma = self._add_irma.itemText(index).strip()
        data = getattr(self, "_irma_lookup", {}).get(irma)
        if not data:
            return
        self._add_train.setText(str(data.get("train", "") or ""))
        self._add_m1s.setText(str(data.get("m1_start", "") or ""))
        self._add_m1f.setText(str(data.get("m1_finish", "") or ""))
        self._add_m2s.setText(str(data.get("m2_start", "") or ""))
        self._add_m2f.setText(str(data.get("m2_finish", "") or ""))
        self._add_edpu.setText(str(data.get("edpu", "") or ""))
        self._add_loc.setText(str(data.get("location", "") or ""))
        # Leave vol blank ? user always enters that manually
        self._add_vol.setFocus()

    def _populate_proc_dropdown(self):
        """Build _proc_lookup from all loaded cache entries and populate the
        processor key combo with known processor key->name pairs."""
        if not hasattr(self, "_proc_lookup"):
            self._proc_lookup = {}   # {key: name}

        for fname, sheets in self._cache.items():
            for sname, entry in sheets.items():
                if not isinstance(entry, dict):
                    continue
                for block in entry.get("blocks", []):
                    for d in block.get("dests", []):
                        dk = str(d.get("key") or "").strip()
                        dn = str(d.get("name") or "").strip()
                        if dk:
                            self._proc_lookup[dk] = dn or dk
                    dk2 = str(block.get("dest_key") or "").strip()
                    dn2 = str(block.get("dest_name") or "").strip()
                    if dk2 and dk2 not in self._proc_lookup:
                        self._proc_lookup[dk2] = dn2 or dk2

        self._add_proc_key.blockSignals(True)
        current_text = self._add_proc_key.lineEdit().text()
        self._add_proc_key.clear()
        for key in sorted(self._proc_lookup.keys()):
            name = self._proc_lookup[key]
            self._add_proc_key.addItem(f"{key}  ?  {name}", userData=key)
        self._add_proc_key.lineEdit().setText(current_text)
        self._add_proc_key.blockSignals(False)

    def _on_proc_key_autofill(self, index):
        """When the user picks a processor from the dropdown, fill the name field."""
        key = self._add_proc_key.itemData(index)
        if key is None:
            # Fallback: parse key from display text "key  ?  name"
            text = self._add_proc_key.itemText(index)
            key = text.split("-")[0].strip() if "-" in text else text.strip()
        name = getattr(self, "_proc_lookup", {}).get(key, "")
        self._add_proc_key.lineEdit().setText(key)
        if name:
            self._add_proc_name.setText(name)
        self._add_proc_vol.setFocus()

    def _on_add_block(self):
        """
        Append a new empty block to the current sheet's modified route.

        A small dialog collects:
          - Route name / number  (free text, e.g. "1082")
          - Processor key        (numeric, e.g. 972712)
          - Processor name       (free text, e.g. "Saputo Abbotsford - 2F")
          - Optional partial volume (leave blank for "rest of load")
        """
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if not fname or not sname or self._mod_blocks is None:
            QMessageBox.warning(self, "Add Block", "No sheet loaded.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Add New Block")
        dlg.setFixedWidth(380)
        dl = QVBoxLayout(dlg)
        dl.setContentsMargins(14, 12, 14, 12)
        dl.setSpacing(8)

        bold_f = QFont(); bold_f.setBold(True)
        small_f = QFont()

        def _row(label, widget):
            rw = QWidget(); rl = QHBoxLayout(rw)
            rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(8)
            lb = QLabel(label); lb.setFont(small_f); lb.setFixedWidth(120)
            rl.addWidget(lb); rl.addWidget(widget, stretch=1)
            return rw

        route_edit = QLineEdit()
        route_edit.setPlaceholderText("e.g. 1082")
        route_edit.setFont(small_f)

        proc_key_edit = QLineEdit()
        proc_key_edit.setPlaceholderText("e.g. 972712")
        proc_key_edit.setFont(small_f)

        proc_name_edit = QLineEdit()
        proc_name_edit.setPlaceholderText("e.g. Saputo Abbotsford - 2F")
        proc_name_edit.setFont(small_f)

        vol_edit = QLineEdit()
        vol_edit.setPlaceholderText("blank = rest of load")
        vol_edit.setFont(small_f)

        status_lbl = QLabel("")
        status_lbl.setStyleSheet("color: #c0392b; font-size: 8pt;")

        dl.addWidget(QLabel("Fill in route details:"))
        dl.addWidget(_row("Route name:", route_edit))
        dl.addWidget(_row("Processor key:", proc_key_edit))
        dl.addWidget(_row("Processor name:", proc_name_edit))
        dl.addWidget(_row("Partial vol (L):", vol_edit))
        dl.addWidget(status_lbl)

        btn_row = QWidget(); br = QHBoxLayout(btn_row)
        br.setContentsMargins(0, 0, 0, 0); br.setSpacing(8)
        ok_btn = QPushButton("Add Block")
        ok_btn.setFont(bold_f)
        ok_btn.setStyleSheet(
            "QPushButton { background:#7b1fa2; color:white; font-weight:bold; "
            "border-radius:4px; padding:4px 12px; }")
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setFont(small_f)
        br.addStretch(); br.addWidget(cancel_btn); br.addWidget(ok_btn)
        dl.addWidget(btn_row)

        cancel_btn.clicked.connect(dlg.reject)

        def _accept():
            pk = proc_key_edit.text().strip()
            pn = proc_name_edit.text().strip()
            if not pk and not pn:
                status_lbl.setText("Processor key or name is required.")
                return
            vol_raw = vol_edit.text().strip()
            try:
                vol_partial = float(vol_raw.replace(",", "")) if vol_raw else None
            except ValueError:
                status_lbl.setText("Invalid volume ? enter a number or leave blank.")
                return
            dest = {"name": pn, "key": pk, "vol_partial": vol_partial}
            new_block = {
                "route":     route_edit.text().strip(),
                "dests":     [dest],
                "dest_name": pn,
                "dest_key":  pk,
                "rows":      [],
            }
            self._mod_blocks.append(new_block)
            self._save_mod_blocks()
            self._render_editable()
            self.statusBar().showMessage(
                f"Added empty block '{new_block['route'] or '(unnamed)'}' "
                f"-> {pn or pk} to sheet {sname}")
            dlg.accept()

        ok_btn.clicked.connect(_accept)
        dlg.exec_()

    def _on_auto_flag_waits(self):
        """Set MWO on every farm whose milking wait exceeds 30 minutes,
        iterating until stable so cascading waits are captured correctly.
        Also ensures farms in NO_MILKING_WINDOW_FARMS always have MWO set.
        """
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        total_flagged = 0
        sheets_updated = set()

        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            start = entry.get("start_time")
            if not start:
                continue

            key = (fname, sname)
            if key in self._sheet_mods:
                mod_blocks = self._sheet_mods[key]
            else:
                mod_blocks = copy.deepcopy(entry["blocks"])

            orig_entry = entry
            flagged = 0

            # Iterate until stable ? cascading gate waits may push subsequent
            # farms' arrivals so that they now also exceed the wait threshold.
            max_passes = 20
            for _pass in range(max_passes):
                pass_changed = False

                # Recompute timing with current flags applied
                ct = calc_times(mod_blocks, self.dm, start, self.dm_dur,
                                suppress_no_milking=True)
                if ct is None:
                    break
                all_times = ct[0]

                for b_idx, block in enumerate(mod_blocks):
                    btimes = all_times[b_idx] if b_idx < len(all_times) else None
                    if not btimes:
                        continue
                    for f_idx, farm in enumerate(block.get("rows", [])):
                        f_stop = _farm_stop_index(block, f_idx, b_idx, mod_blocks)
                        if f_stop >= len(btimes):
                            continue
                        ft   = btimes[f_stop]
                        wait = ft.get("wait")
                        arr  = ft.get("arr")
                        if not arr:
                            continue

                        # Flag if wait > 30 min and not yet flagged
                        if wait and wait > 30 and not farm.get("_mwo"):
                            farm["_mwo"] = True
                            flagged += 1
                            pass_changed = True

                if not pass_changed:
                    break  # stable ? no more cascading changes

            # Sync flags to orig panel via uid
            if flagged:
                for b_idx, block in enumerate(mod_blocks):
                    for f_idx, farm in enumerate(block.get("rows", [])):
                        if not farm.get("_mwo"):
                            continue
                        uid = farm.get("_uid")
                        for ob_i, orig_block in enumerate(orig_entry.get("blocks", [])):
                            for orig_fi, orig_farm in enumerate(orig_block.get("rows", [])):
                                if uid and orig_farm.get("_uid") == uid:
                                    orig_farm["_mwo"] = True
                                    break

                self._sheet_mods[key] = mod_blocks
                total_flagged += flagged
                sheets_updated.add(sname)

        # -- Always anchor the 3 suppressed ROBOT farms to their original arrival --
        # These farms have no valid milking window data so suppress_no_milking=True
        # means calc_times returns wait=None for them -> the normal auto-flag loop
        # never triggers -> the solver has no anchor for them -> they end up anywhere.
        # Always flag NO_MILKING_WINDOW_FARMS with MWO so the solver never
        # tries to schedule around their (unreliable) window data.
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict): continue
            key = (fname, sname)
            if key not in self._sheet_mods:
                self._sheet_mods[key] = copy.deepcopy(entry["blocks"])
            mod_blocks = self._sheet_mods[key]
            changed = False
            for block in mod_blocks:
                for farm in block.get("rows", []):
                    irma = farm.get("irma", "")
                    if irma not in NO_MILKING_WINDOW_FARMS:
                        continue
                    if farm.get("_mwo"):
                        continue  # already flagged
                    farm["_mwo"] = True
                    changed = True
                    total_flagged += 1
            if changed:
                self._sheet_mods[key] = mod_blocks
                sheets_updated.add(sname)

        # Refresh the currently displayed sheet
        cur_sname = self.sheet_cb.currentText()
        cur_entry = self._cache.get(fname, {}).get(cur_sname)
        cur_key   = (fname, cur_sname)
        if cur_key in self._sheet_mods:
            self._mod_blocks = self._sheet_mods[cur_key]
        if cur_sname in sheets_updated and cur_entry:
            populate_table(self.orig_table, cur_entry["blocks"],
                           self.dm, editable=False,
                           start_time=self._driver_start, dm_dur=self.dm_dur,
                           suppress_no_milking=self._suppress_no_milking_cb.isChecked())
        self._render_editable()
        if total_flagged and self.tabs.currentIndex() == 1:
            self._refresh_comparison()

        if total_flagged:
            self.statusBar().showMessage(
                f"MWO set on {total_flagged} farm(s) across "
                f"{len(sheets_updated)} sheet(s)", 5000)
        else:
            self.statusBar().showMessage(
                f"No farms with wait > 30min found "
                f"({'distance matrix not loaded' if not self.dm else 'across all sheets'})",
                5000)

    def _on_reset_current_sheet(self):
        """Reset only the currently displayed sheet back to original; leave others intact."""
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if not fname or not sname: return
        key = (fname, sname)
        if key in self._sheet_mods:
            del self._sheet_mods[key]
        # Remove any tray items that came from this sheet ? the reset restores
        # those farms from cache, so leaving them in the tray would cause
        # duplicates if the user dragged them back.
        rows_to_remove = [
            i for i, item in enumerate(self._removed)
            if item.get("_from_fname") == fname and item.get("_from_sname") == sname
        ]
        for i in reversed(rows_to_remove):
            self._removed.pop(i)
            self.farm_tray.removeRow(i)
        if rows_to_remove:
            self._tray_del_btn.setEnabled(self.farm_tray.currentRow() >= 0)
        # Reload display (will re-create a fresh copy from cache)
        self._display_sheet()
        # Keep comparison tab fresh if open
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    def _on_reset_all(self):
        """Reset all modified blocks back to originals and clear tray."""
        if self._cache and hasattr(self, 'file_cb'):
            fname = self.file_cb.currentText()
            if fname in self._cache:
                # Clear all sheet mods
                keys_to_clear = [k for k in self._sheet_mods if k[0] == fname]
                for k in keys_to_clear:
                    del self._sheet_mods[k]
                # Clear tray
                self.farm_tray.setRowCount(0)
                self._removed.clear()
                self._tray_del_btn.setEnabled(False)
                # Reload current sheet
                self._display_sheet()
                # Keep comparison tab fresh if open
                if self.tabs.currentIndex() == 1:
                    self._refresh_comparison()

    def _on_manual_add_proc(self):
        """Validate and add a manually entered processor destination to the tray."""
        raw_text = self._add_proc_key.lineEdit().text().strip()
        # If user selected from dropdown, text may be "key  ?  name"
        key_raw  = raw_text.split("-")[0].strip() if "-" in raw_text else raw_text
        name_raw = self._add_proc_name.text().strip()
        # Auto-fill name from lookup if not manually entered
        if not name_raw and key_raw and hasattr(self, "_proc_lookup"):
            name_raw = self._proc_lookup.get(key_raw, "")
        vol_raw  = self._add_proc_vol.text().strip()

        if not key_raw and not name_raw:
            self._add_proc_key.lineEdit().setStyleSheet("color:#c0392b; font-weight:bold;")
            self._add_status.setText("Processor key or name is required")
            self._add_status_timer.start(4000)
            return
        self._add_proc_key.lineEdit().setStyleSheet("")

        vol_partial = None
        if vol_raw:
            try:
                vol_partial = float(vol_raw.replace(",",""))
            except ValueError:
                self._add_proc_vol.setStyleSheet("color:#c0392b; font-weight:bold;")
                self._add_status.setText(f"Partial volume must be a number (got '{vol_raw}')")
                self._add_status_timer.start(4000)
                return
        self._add_proc_vol.setStyleSheet("")
        self._add_status.setText("")

        dest = {
            "key":         key_raw or name_raw,
            "name":        name_raw or key_raw,
            "vol_partial": vol_partial,
            "_from_block":  -1,
            "_from_fname":  self.file_cb.currentText(),
            "_from_sname":  self.sheet_cb.currentText(),
            "_from_colour": self._day_colour,
            "_is_dest":     True,
        }
        self._removed.append(dest)
        sheet_label = f"{self.sheet_cb.currentText()} (manual)"
        self.farm_tray.add_dest(dest, sheet_label, self._day_colour)
        for w in (self._add_proc_name, self._add_proc_vol):
            w.clear()
        self._add_proc_key.lineEdit().clear()
        self._add_proc_key.lineEdit().setFocus()

    def _on_farm_removed(self, b_idx, f_idx):
        """Remove farm from mod_blocks and add to tray."""
        if self._mod_blocks is None: return
        block = self._mod_blocks[b_idx]
        if f_idx >= len(block["rows"]): return
        farm = block["rows"].pop(f_idx)
        farm["_from_block"]  = b_idx
        farm["_from_fname"]  = self.file_cb.currentText()
        farm["_from_sname"]  = self.sheet_cb.currentText()
        farm["_from_colour"] = self._day_colour
        self._removed.append(farm)
        sheet_label = f"{self.sheet_cb.currentText()} / {block['route']}"
        self.farm_tray.add_farm(farm, sheet_label, self._day_colour)
        self._save_mod_blocks()
        self._render_editable()

    def _on_farm_reorder(self, src_b, src_f, dst_b, dst_f):
        """Move a farm within mod_blocks (internal drag reorder)."""
        if self._mod_blocks is None: return
        src_block = self._mod_blocks[src_b]
        if src_f >= len(src_block["rows"]): return
        farm = src_block["rows"].pop(src_f)
        # Strip only ephemeral tray-metadata keys, not persistent farm state
        # (_uid, _mwo, _orig_arr must survive a reorder).
        _TRAY_META = {"_from_block", "_from_fname", "_from_sname", "_from_colour"}
        restored = {k: v for k, v in farm.items() if k not in _TRAY_META}
        # Adjust dst_f if moving within same block and dst is after src
        dst_b = max(0, min(dst_b, len(self._mod_blocks)-1))
        dst_rows = self._mod_blocks[dst_b]["rows"]
        if dst_b == src_b and dst_f > src_f:
            dst_f -= 1   # list shrank by one
        dst_f = max(0, min(dst_f, len(dst_rows)))
        dst_rows.insert(dst_f, restored)
        self._render_editable()

    def _on_farm_inserted(self, tray_idx, b_idx, insert_before):
        """Return a farm from the tray into mod_blocks."""
        if self._mod_blocks is None: return
        if tray_idx < 0 or tray_idx >= len(self._removed): return
        farm = self._removed.pop(tray_idx)
        self.farm_tray.removeRow(tray_idx)
        b_idx = max(0, min(b_idx, len(self._mod_blocks)-1))
        # Strip only ephemeral tray-metadata keys, not persistent farm state
        # (_uid, _mwo, _orig_arr must survive tray round-trip).
        _TRAY_META = {"_from_block", "_from_fname", "_from_sname", "_from_colour"}
        restored = {k: v for k, v in farm.items() if k not in _TRAY_META}
        rows = self._mod_blocks[b_idx]["rows"]
        rows.insert(max(0, min(insert_before, len(rows))), restored)
        self._render_editable()

    def _on_dest_removed(self, b_idx, d_idx):
        """Remove a dest from mod_blocks and add to tray."""
        if self._mod_blocks is None: return
        block = self._mod_blocks[b_idx]
        dests = block.get("dests") or []
        if d_idx >= len(dests): return
        dest = dests.pop(d_idx)
        # Update legacy fields
        if dests:
            block["dest_name"] = dests[0]["name"]
            block["dest_key"]  = dests[0]["key"]
        else:
            block["dest_name"] = ""
            block["dest_key"]  = ""
        dest["_from_block"]  = b_idx
        dest["_from_fname"]  = self.file_cb.currentText()
        dest["_from_sname"]  = self.sheet_cb.currentText()
        dest["_from_colour"] = self._day_colour
        dest["_is_dest"]     = True
        self._removed.append(dest)
        sheet_label = f"{self.sheet_cb.currentText()} / {block['route']}"
        self.farm_tray.add_dest(dest, sheet_label, self._day_colour)
        self._save_mod_blocks()
        self._render_editable()

    def _on_dest_reorder(self, src_b, src_d, dst_b, dst_d):
        """Move a dest within (or across) blocks."""
        if self._mod_blocks is None: return
        src_block = self._mod_blocks[src_b]
        src_dests = src_block.get("dests") or []
        if src_d >= len(src_dests): return
        dest = src_dests.pop(src_d)
        # Update src block legacy fields
        if src_dests:
            src_block["dest_name"] = src_dests[0]["name"]
            src_block["dest_key"]  = src_dests[0]["key"]
        dst_b   = max(0, min(dst_b, len(self._mod_blocks) - 1))
        dst_block = self._mod_blocks[dst_b]
        dst_dests = dst_block.setdefault("dests", [])
        if dst_b == src_b and dst_d > src_d:
            dst_d -= 1
        dst_d = max(0, min(dst_d, len(dst_dests)))
        dst_dests.insert(dst_d, dest)
        dst_block["dest_name"] = dst_dests[0]["name"]
        dst_block["dest_key"]  = dst_dests[0]["key"]
        self._render_editable()

    def _on_block_reorder(self, src_b, dst_b):
        """Move an entire block (and all its farms/dests) to a new position.
        dst_b == -1 means append at end."""
        if self._mod_blocks is None: return
        n = len(self._mod_blocks)
        if src_b < 0 or src_b >= n: return
        block = self._mod_blocks.pop(src_b)
        if dst_b < 0 or dst_b >= len(self._mod_blocks):
            self._mod_blocks.append(block)
        else:
            insert_at = dst_b if dst_b <= src_b else dst_b - 1
            insert_at = max(0, min(insert_at, len(self._mod_blocks)))
            self._mod_blocks.insert(insert_at, block)
        self._save_mod_blocks()
        self._render_editable()

    def _on_dest_inserted(self, tray_idx, b_idx, insert_before):
        if self._mod_blocks is None: return
        if tray_idx < 0 or tray_idx >= len(self._removed): return
        item = self._removed[tray_idx]
        if not item.get("_is_dest"): return
        self._removed.pop(tray_idx)
        self.farm_tray.removeRow(tray_idx)
        dest = {k: v for k, v in item.items() if not k.startswith("_")}
        b_idx = max(0, min(b_idx, len(self._mod_blocks) - 1))
        dests = self._mod_blocks[b_idx].setdefault("dests", [])
        dests.insert(max(0, min(insert_before, len(dests))), dest)
        self._mod_blocks[b_idx]["dest_name"] = dests[0]["name"]
        self._mod_blocks[b_idx]["dest_key"]  = dests[0]["key"]
        self._render_editable()

    # -------------------------------------------------------------------------
    # Tab handling
    # -------------------------------------------------------------------------

    def _on_tab_changed(self, idx):
        title = self.tabs.tabText(idx)
        if title == "Comparison": self._refresh_comparison()
        if title == "Solver": self._refresh_demand_targets()
        if title == "Debug": self._refresh_debug_tab()

    # -------------------------------------------------------------------------
    # Comparison tab
    # -------------------------------------------------------------------------

    def _get_mod_blocks(self, fname, sname):
        """Get the modified blocks for a sheet, falling back to original."""
        key = (fname, sname)
        if key in self._sheet_mods:
            return self._sheet_mods[key]
        entry = self._cache.get(fname, {}).get(sname, {})
        return entry.get("blocks", []) if isinstance(entry, dict) else []

    @staticmethod
    def _normalise_proc(name):
        """Strip 'Yard for ' prefix so yard entries merge with plant entries."""
        if not name: return "Unknown"
        n = name.strip()
        if n.lower().startswith("yard for "):
            n = n[9:].strip()
        return n or "Unknown"

    def _agg_file(self, fname, use_mod=False):
        """Aggregate all sheets in fname.
        Returns:
          proc_vols: {proc_name: {"RED":v,"BLUE":v,"OTHER":v,"TOTAL":v}}
          sheet_rows: [(sname, day_colour, total_km, km_ok, shift_hours)]
        """
        if fname not in self._cache: return {}, []
        proc_vols = {}   # proc_name -> {colour_bucket: vol}
        sheet_rows = []
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict): continue
            orig_blocks = entry.get("blocks", [])
            start_time  = entry.get("start_time")
            day_colour  = entry.get("day_colour", "")
            if use_mod:
                key = (fname, sname)
                if key in self._sheet_mods:
                    blocks = self._sheet_mods[key]
                else:
                    blocks = orig_blocks
            else:
                # When route corrections are active, use the corrected baseline
                # (_corrected_blocks) ? never solver output from _sheet_mods.
                corrections_on = (hasattr(self, "_chk_route_opt") and
                                  self._chk_route_opt.isChecked())
                key = (fname, sname)
                if corrections_on and key in self._corrected_blocks:
                    blocks = self._corrected_blocks[key]
                elif corrections_on and key in self._sheet_mods:
                    blocks = self._sheet_mods[key]
                else:
                    blocks = orig_blocks
            # Only skip if there is no cache entry at all (shouldn't happen here,
            # but guard anyway). Do NOT skip empty mod_blocks ? a solver run can
            # legitimately empty a sheet by moving all its farms elsewhere, and
            # skipping it would make those litres disappear from the totals.
            if blocks is None: continue

            # Colour bucket for this sheet
            dc = day_colour.upper().strip()
            if "RED"  in dc: bucket = "RED"
            elif "BLUE" in dc: bucket = "BLUE"
            else: bucket = "OTHER"

            # Processor volumes ? split across multiple dests if present
            for block in blocks:
                dests = block.get("dests") or []
                if not dests:
                    dk = block.get("dest_key","") or block.get("dest_name","") or "Unknown"
                    dests = [{"name": block.get("dest_name",""), "key": dk, "vol_partial": None}]
                total_farm_vol = sum((r["prior_vol"] or 0) for r in block["rows"]
                                     if isinstance(r.get("prior_vol"), (int, float)))
                already = 0.0
                for dest_d in dests:
                    raw = dest_d.get("name") or dest_d.get("key") or "Unknown"
                    dest = self._normalise_proc(raw)
                    vp = dest_d.get("vol_partial")
                    remaining = max(0.0, total_farm_vol - already)
                    offload = min(float(vp), remaining) if vp is not None else remaining
                    already += offload
                    if dest not in proc_vols:
                        proc_vols[dest] = {"RED": 0, "BLUE": 0, "OTHER": 0}
                    proc_vols[dest][bucket] += offload

            # Distance
            total_km = 0.0; km_ok = True
            for dists in calc_distances(blocks, self.dm):
                for d in dists[:-1]:
                    if d is None: km_ok = False
                    else: total_km += d

            # Shift hours
            shift_hours = None
            if start_time:
                _ct3 = calc_times(blocks, self.dm, start_time, dm_dur=self.dm_dur)
                if _ct3 is not None:
                    from datetime import datetime, date
                    base = datetime.combine(date.today(), start_time)
                    shift_hours = (_ct3[1] - base).total_seconds() / 3600.0

            sheet_rows.append((sname, day_colour, total_km, km_ok, shift_hours))
        return proc_vols, sheet_rows

    def _refresh_comparison(self):
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache: return

        orig_pv, orig_sr = self._agg_file(fname, use_mod=False)
        mod_pv,  mod_sr  = self._agg_file(fname, use_mod=True)

        # Compute tray volume ? farms currently sitting in the tray are excluded
        # from modified routes and will make the modified total appear lower.
        tray_vol = sum(
            f.get("prior_vol") or 0
            for f in self._removed
            if not f.get("_is_dest") and isinstance(f.get("prior_vol"), (int, float))
        )

        # Warn in the status bar if tray farms are causing a volume discrepancy
        orig_total = sum(sum(b.values()) for b in orig_pv.values())
        mod_total  = sum(sum(b.values()) for b in mod_pv.values())
        diff = orig_total - mod_total
        if tray_vol > 0:
            self.statusBar().showMessage(
                f"(!)  {tray_vol:,.0f} L in tray (not on any route) - "
                f"Modified total is {diff:,.0f} L less than Original."
            )
        elif diff > 100:
            # Unexplained discrepancy ? shouldn't normally happen
            self.statusBar().showMessage(
                f"(!)  Modified total is {diff:,.0f} L less than Original "
                f"with no farms in tray ? check for blocks without a processor."
            )
        else:
            self.statusBar().clearMessage()

        # Processor volumes
        all_procs = sorted(set(orig_pv.keys()) | set(mod_pv.keys()))
        changed_procs = {p for p in all_procs
                         if orig_pv.get(p,{}) != mod_pv.get(p,{})}
        self._fill_proc_comp(self._comp_tables["proc_orig"], orig_pv, all_procs, changed_procs, bold_changed=False)
        self._fill_proc_comp(self._comp_tables["proc_mod"],  mod_pv,  all_procs, changed_procs, bold_changed=True)

        # Sheet summaries ? build full sheet info dict keyed by sname
        def sr_to_map(sr):
            return {s: (dc, km, ok, h) for s, dc, km, ok, h in sr}
        orig_sm = sr_to_map(orig_sr)
        mod_sm  = sr_to_map(mod_sr)
        all_sheets = list(dict.fromkeys([s for s,*_ in orig_sr] + [s for s,*_ in mod_sr]))
        changed_sheets = {s for s in all_sheets
                          if (orig_sm.get(s,("",0,True,None))[1:3]
                              != mod_sm.get(s,("",0,True,None))[1:3])}
        self._fill_sheet_comp(self._comp_tables["sheet_orig"], orig_sm, all_sheets, changed_sheets, bold_changed=False)
        self._fill_sheet_comp(self._comp_tables["sheet_mod"],  mod_sm,  all_sheets, changed_sheets, bold_changed=True)

    def _comp_cell(self, text, bg, font, align=Qt.AlignRight|Qt.AlignVCenter, fg=None):
        item = QTableWidgetItem(str(text))
        item.setBackground(bg); item.setFont(font)
        item.setTextAlignment(align)
        item.setFlags(Qt.ItemIsEnabled|Qt.ItemIsSelectable)
        if fg: item.setForeground(fg)
        return item

    def _fill_proc_comp(self, table, proc_vols, all_procs, changed_procs, bold_changed):
        # Columns: Processor | Red | Blue | Other | Total
        table.clearSpans(); table.clear()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Processor", "Red (L)", "Blue (L)", "Other (L)", "Total (L)"])
        table.setRowCount(len(all_procs) + 1)
        bold_font = QFont(); bold_font.setBold(True)
        norm_font = QFont()
        tot_r = tot_b = tot_o = tot_t = 0
        for i, proc in enumerate(all_procs):
            buckets  = proc_vols.get(proc, {"RED":0,"BLUE":0,"OTHER":0})
            r_vol = buckets.get("RED",   0)
            b_vol = buckets.get("BLUE",  0)
            o_vol = buckets.get("OTHER", 0)
            total = r_vol + b_vol + o_vol
            changed = proc in changed_procs
            bg   = CLR_CHANGED if changed else (CLR_ALT if i%2==0 else CLR_WHITE)
            font = bold_font if (changed and bold_changed) else norm_font
            table.setItem(i, 0, self._comp_cell(proc,            bg, font, Qt.AlignLeft|Qt.AlignVCenter))
            table.setItem(i, 1, self._comp_cell(f"{int(r_vol):,}" if r_vol else "-", bg, font))
            table.setItem(i, 2, self._comp_cell(f"{int(b_vol):,}" if b_vol else "-", bg, font))
            table.setItem(i, 3, self._comp_cell(f"{int(o_vol):,}" if o_vol else "-", bg, font))
            table.setItem(i, 4, self._comp_cell(f"{int(total):,}", bg, font))
            tot_r += r_vol; tot_b += b_vol; tot_o += o_vol; tot_t += total
        r = len(all_procs)
        table.setItem(r, 0, make_header_item("TOTAL", bg=CLR_TOTAL, fg=QColor("#000000")))
        for col, val in [(1,tot_r),(2,tot_b),(3,tot_o),(4,tot_t)]:
            tv = make_header_item(f"{int(val):,}", bg=CLR_TOTAL, fg=QColor("#000000"))
            tv.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
            table.setItem(r, col, tv)

    def _fill_sheet_comp(self, table, sheet_map, all_sheets, changed_sheets, bold_changed):
        # Columns: Sheet | ? Red | ? Blue | ? Other | km | h
        table.clearSpans(); table.clear()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Sheet", "Red", "Blue", "Other", "km", "Hours"])
        bold_font = QFont(); bold_font.setBold(True)
        norm_font = QFont()

        # Group by colour bucket for subtotals: RED rows, BLUE rows, OTHER rows
        # Order: RED sheets, then BLUE, then OTHER, then totals
        def bucket_of(dc):
            dc = dc.upper()
            if "RED"  in dc: return "RED"
            if "BLUE" in dc: return "BLUE"
            return "OTHER"

        grouped = {"RED": [], "BLUE": [], "OTHER": []}
        for s in all_sheets:
            dc = sheet_map.get(s, ("","",0,True,None))[0] if isinstance(sheet_map.get(s), tuple) else ""
            grouped[bucket_of(dc)].append(s)

        rows_ordered = []
        subtotals = {}  # bucket -> (km_tot, h_tot, ok_all)
        for bucket in ("RED", "BLUE", "OTHER"):
            sheets_in = grouped[bucket]
            k_tot = 0.0; h_tot = 0.0; ok_all = True
            for s in sheets_in:
                rows_ordered.append((s, bucket))
                if s in sheet_map:
                    _, km, ok, h = sheet_map[s]
                    k_tot += km; ok_all = ok_all and ok
                    if h: h_tot += h
            subtotals[bucket] = (k_tot, h_tot, ok_all, len(sheets_in))

        # Total row count: data rows + 3 subtotal rows + 1 grand total
        n_data = len(all_sheets)
        n_sub  = sum(1 for b in ("RED","BLUE","OTHER") if subtotals[b][3] > 0)
        table.setRowCount(n_data + n_sub + 1)

        CLR_RED_SUB  = QColor("#ffcdd2")
        CLR_BLUE_SUB = QColor("#bbdefb")
        CLR_OTH_SUB  = QColor("#e0e0e0")
        sub_colours  = {"RED": CLR_RED_SUB, "BLUE": CLR_BLUE_SUB, "OTHER": CLR_OTH_SUB}
        dot_colours  = {"RED": "[R]", "BLUE": "[B]", "OTHER": "[O]"}

        row_i = 0
        grand_km = 0.0; grand_h = 0.0; grand_ok = True
        for bucket in ("RED", "BLUE", "OTHER"):
            sheets_in = grouped[bucket]
            if not sheets_in: continue
            for s in sheets_in:
                changed = s in changed_sheets
                bg   = CLR_CHANGED if changed else (CLR_ALT if row_i%2==0 else CLR_WHITE)
                font = bold_font if (changed and bold_changed) else norm_font
                table.setItem(row_i, 0, self._comp_cell(s, bg, font, Qt.AlignLeft|Qt.AlignVCenter))
                # Colour dot columns
                for col, b in enumerate(("RED","BLUE","OTHER"), start=1):
                    dot = "[stop]" if b == bucket else ""
                    dot_bg = sub_colours[b] if b == bucket else bg
                    table.setItem(row_i, col, self._comp_cell(dot, dot_bg, font, Qt.AlignCenter))
                if s in sheet_map:
                    _, km, ok, h = sheet_map[s]
                    km_str = f"{km:.1f}" if ok else f"~{km:.1f}*"
                    h_str  = f"{h:.2f}" if h is not None else "-"
                    grand_km += km; grand_ok = grand_ok and ok
                    if h: grand_h += h
                else:
                    km_str = "-"; h_str = "-"
                table.setItem(row_i, 4, self._comp_cell(km_str, bg, font))
                table.setItem(row_i, 5, self._comp_cell(h_str,  bg, font))
                row_i += 1
            # Subtotal row for this colour bucket
            k_tot, h_tot, ok_sub, _ = subtotals[bucket]
            sub_bg = sub_colours[bucket]
            label = f"{bucket.title()} Subtotal"
            table.setItem(row_i, 0, make_header_item(label, bg=sub_bg, fg=QColor("#000000")))
            for col in (1,2,3):
                table.setItem(row_i, col, make_header_item("", bg=sub_bg, fg=QColor("#000000")))
            km_s = f"{k_tot:.1f}" if ok_sub else f"~{k_tot:.1f}*"
            table.setItem(row_i, 4, make_header_item(km_s, bg=sub_bg, fg=QColor("#000000")))
            table.setItem(row_i, 5, make_header_item(f"{h_tot:.2f}", bg=sub_bg, fg=QColor("#000000")))
            row_i += 1

        # Grand total
        km_grand = f"{grand_km:.1f}" if grand_ok else f"~{grand_km:.1f}*"
        table.setItem(row_i, 0, make_header_item("TOTAL", bg=CLR_TOTAL, fg=QColor("#000000")))
        for col in (1,2,3):
            table.setItem(row_i, col, make_header_item("", bg=CLR_TOTAL, fg=QColor("#000000")))
        table.setItem(row_i, 4, make_header_item(km_grand,          bg=CLR_TOTAL, fg=QColor("#000000")))
        table.setItem(row_i, 5, make_header_item(f"{grand_h:.2f}",  bg=CLR_TOTAL, fg=QColor("#000000")))

    # -------------------------------------------------------------------------
    # Dropdown signals
    # -------------------------------------------------------------------------

    def _on_year(self):  self._pop_months(); self._pop_files()
    def _on_month(self): self._pop_files()
    def _on_file(self):  self._pop_sheets()
    def _on_sheet(self): self._display_sheet()


def _run_selftests():
    """Lightweight checks for the pure domain helpers. Run with:
        python viewer158.py --selftest
    These pin down the behaviour of the core time/volume functions so a future
    refactor can't silently change them. No GUI or data files required.
    """
    import math as _math
    fails = []

    def check(name, cond):
        if cond:
            print(f"  ok  {name}")
        else:
            print(f"  FAIL {name}")
            fails.append(name)

    def approx(a, b, tol=1e-6):
        return abs(a - b) <= tol

    # parse_hhmm
    check("parse_hhmm valid",   parse_hhmm("06:30") == dt_time(6, 30))
    check("parse_hhmm empty",   parse_hhmm("") is None)
    check("parse_hhmm dash",    parse_hhmm("-") is None)
    check("parse_hhmm garbage", parse_hhmm("not-a-time") is None)

    # time_in_window (normal + overnight + boundaries)
    check("window inside",      time_in_window(dt_time(7, 0),  "06:00", "08:00") is True)
    check("window outside",     time_in_window(dt_time(9, 0),  "06:00", "08:00") is False)
    check("window open bound",  time_in_window(dt_time(6, 0),  "06:00", "08:00") is True)
    check("window overnight a", time_in_window(dt_time(23, 0), "22:00", "02:00") is True)
    check("window overnight b", time_in_window(dt_time(1, 0),  "22:00", "02:00") is True)
    check("window overnight c", time_in_window(dt_time(12, 0), "22:00", "02:00") is False)
    check("window none",        time_in_window(None,           "06:00", "08:00") is False)

    # stop_duration / drive_mins
    check("stop_duration zero", approx(stop_duration(0),   ONSITE_MIN))
    check("stop_duration pump", approx(stop_duration(PUMP_RATE_LPM), ONSITE_MIN + 1.0))
    check("drive_mins",         approx(drive_mins(DRIVE_SPEED_KMH), 60.0))
    check("drive_mins none",    drive_mins(None) is None)

    # _dest_vol_partial
    check("dest_partial rest",  approx(_dest_vol_partial({"vol_partial": None}, 1000, 200), 800))
    check("dest_partial cap",   approx(_dest_vol_partial({"vol_partial": 500}, 1000, 200), 500))
    check("dest_partial short", approx(_dest_vol_partial({"vol_partial": 500}, 600, 200), 400))

    # _block_has_split
    check("has_split true",  _block_has_split({"dests": [{"split_after": 1}]}) is True)
    check("has_split false", _block_has_split({"dests": [{"key": "x"}]}) is False)
    check("has_split empty", _block_has_split({}) is False)

    # _block_dest_offloads
    o1 = _block_dest_offloads({"rows": [{"prior_vol": 1000}],
                               "dests": [{"key": "A", "vol_partial": None}]})
    check("offload single", o1 == {"A": 1000.0})
    o2 = _block_dest_offloads({"rows": [{"prior_vol": 1000}],
                               "dests": [{"key": "A", "vol_partial": 600},
                                         {"key": "B", "vol_partial": None}]})
    check("offload split", approx(o2.get("A", 0), 600) and approx(o2.get("B", 0), 400))
    o3 = _block_dest_offloads({"rows": [{"prior_vol": 500}], "dests": [], "dest_key": "K"})
    check("offload dest_key fallback", o3 == {"K": 500.0})
    o4 = _block_dest_offloads({"rows": [{"prior_vol": 500}]})
    check("offload no-key fallback", o4 == {"?": 500.0})

    # _build_block_stops sequence + split interleaving
    stops = _build_block_stops(
        {"rows": [{"irma": "01-001"}], "dests": [{"key": "A", "name": "A"}]},
        "VEDDER", True)
    check("stops types", [s["type"] for s in stops] == ["origin", "farm", "dest", "vedder"])
    stops2 = _build_block_stops(
        {"rows": [{"irma": "01-001"}, {"irma": "02-002"}],
         "dests": [{"key": "B", "name": "B", "split_after": 1}]},
        "VEDDER", False)
    check("stops split interleave",
          [s["type"] for s in stops2] == ["origin", "farm", "dest", "farm"])

    # _apply_prob_floors_ceilings (ceiling clamp + floor raise, sum preserved)
    p = _apply_prob_floors_ceilings([0.5, 0.5], [0.0, 0.0], [0.3, 1.0])
    check("floors ceiling clamp", approx(p[0], 0.3) and approx(p[1], 0.7))
    p2 = _apply_prob_floors_ceilings([0.1, 0.9], [0.25, 0.0], [1.0, 1.0])
    check("floors floor raise", approx(p2[0], 0.25) and approx(p2[1], 0.75))
    check("floors sum preserved", approx(sum(p), 1.0) and approx(sum(p2), 1.0))

    print()
    if fails:
        print(f"SELFTEST FAILED ? {len(fails)} failing check(s): {', '.join(fails)}")
        return 1
    print("SELFTEST PASSED ? all checks ok")
    return 0


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Populate THREE_WINDOW_FARMS before constructing MainWindow so that
    # calc_times and arrives_during_milking can use it from the start.
    global THREE_WINDOW_FARMS
    THREE_WINDOW_FARMS = _load_three_window_farms()
    win = MainWindow()   # dropdown signals are wired inside MainWindow.__init__
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    # Configure logging once, here, so library/import code stays quiet but
    # parsing and data-load warnings reach the console when the app runs.
    logging.basicConfig(level=logging.INFO,
                        format="%(levelname)s %(name)s: %(message)s")
    if "--selftest" in sys.argv:
        sys.exit(_run_selftests())
    main()
