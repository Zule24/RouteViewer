"""
viewer.py  -  Vedder D100 Route Manager
"""

import re, sys, csv, copy, random, math, uuid, logging
from functools import lru_cache
from collections import defaultdict
from pathlib import Path
from datetime import time as dt_time, datetime, date, timedelta

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QComboBox, QPushButton, QTableWidget, QTableWidgetItem,
    QFrame, QHeaderView, QAbstractItemView, QTabWidget, QSplitter,
    QScrollArea, QDoubleSpinBox, QSpinBox, QProgressBar,
    QGroupBox, QCheckBox, QTextEdit, QDialog,
    QLineEdit, QMessageBox, QFileDialog, QToolTip, QButtonGroup,
    QGraphicsScene, QGraphicsView, QGraphicsEllipseItem,
    QGraphicsRectItem, QGraphicsTextItem, QDialogButtonBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QMimeData, QByteArray, QObject, QRectF
from PyQt5.QtGui import (QFont, QColor, QDrag, QPainter, QPen, QBrush,
                         QTransform, QPixmap, QPainterPath, QTextDocument)
from PyQt5.QtPrintSupport import QPrinter

import openpyxl
from openpyxl.styles import Font, Border, Side

# -- Logging -------------------------------------------------------------------
# Module logger. A NullHandler keeps libraries quiet by default; main() attaches
# a real handler so parsing/data-load problems surface instead of vanishing.
logger = logging.getLogger("viewer")
logger.addHandler(logging.NullHandler())

# -- Constants -----------------------------------------------------------------

IRMA_RE     = re.compile(r"^\d{2}-\d{3}$")
SHEET_RE    = re.compile(r"^\d{4}$")  # kept for display logic
EXCLUDE_SHEETS = {"INFO", "SHEET1", "SHEET2"}  # sheets with no route data
MONTH_ORDER = ["january","february","march","april","may","june",
               "july","august","september","october","november","december"]
VOL_LIMIT   = 41200   # soft target / penalty threshold (L)
HARD_CAP    = 44000   # absolute maximum a route may carry (L)

# Litres assumed to be on a preload trailer at the start of the day. This is the
# starting load that gets offloaded before the day's collection begins.
# NOTE: intentionally distinct from VOL_LIMIT (the truck's collection capacity).
PRELOAD_VOL = 40000

C_IRMA=1; C_TRAIN=4; C_M1_START=5; C_M1_FINISH=8; C_M2_START=11
C_M2_FINISH=14; C_EDPU=17; C_ROUTE=21; C_LOCATION=32; C_PRIOR_VOL=51
C_DRIVER_START = 52   # AZ1 - driver start time (datetime.time)
C_DAY_COLOUR   = 62   # BJ1 - day colour string (RED/BLUE/GRASSFED/A2 etc.)
C_SHEET_DATE   = 35   # AI1 - sheet date string (merged AI1:AU1, e.g. "Monday July 14, 2025")

DRIVE_SPEED_KMH = 50.0   # km/h average
ONSITE_MIN      = 15.0   # fixed on-site setup minutes per stop
PUMP_RATE_LPM   = 750.0  # litres per minute

VEDDER_DEPART_EXTRA_MINS = 40   # extra minutes added to shift start (Vedder departure)
PRELOAD_WASH_MINS        = 75   # wash time added after a preload offload (1h 15m)
INTER_PROCESSOR_BREAK    = 10   # break minutes inserted between processor stops

# Farms whose milking windows can be suppressed (e.g. robots / continuous milking)
NO_MILKING_WINDOW_FARMS = {"37-874", "14-247", "92-545", "21-132"}

# Regulatory milking buffer constants
# 2025 BC regs: 2 h pre-window (wash cycle) + 1 h post-window (milk cooling)
# No pickup is permitted within these extended exclusion zones.
MILKING_PRE_BUFFER_MINS  = 120.0   # hours before window start -> no pickup
MILKING_POST_BUFFER_MINS = 60.0    # hours after  window end   -> no pickup

# Three-window farm data loaded from JSON at startup
def _load_three_window_farms():
    import json
    p = get_data_dir() / "three_window_farms.json"
    if p.exists():
        try:
            with open(p, encoding="utf-8") as _f:
                return json.load(_f)
        except Exception as ex:
            logger.warning("Could not load %s: %s", p, ex)
    return {}

THREE_WINDOW_FARMS: dict = {}   # populated in main() after get_data_dir() is available

COLS = [
    ("IRMA #",   "irma"),
    ("Tr",       "train"),
    ("M1 Beg",   "m1_start"),
    ("M1 End",   "m1_finish"),
    ("M2 Beg",   "m2_start"),
    ("M2 End",   "m2_finish"),
    ("E",        "edpu"),
    ("Name",     "location"),
    ("Vol (L)",  "prior_vol"),
    ("Dist",     "dist"),
    ("Arr.",     "arr_time"),
    ("Wait",     "wait_time"),
    ("Dep.",     "dep_time"),
    ("M",        "_mwo"),
]

# Full names shown as tooltips on column headers where the label is abbreviated
_COL_TIPS = {
    "train":     "Train",
    "edpu":      "EDPU",
    "dep_time":  "Depart",
    "_mwo":      "Milking Window Override",
}

# Routes starting before this hour (24h clock) are "Day" shifts; from this
# hour onwards they are "Night" shifts used for the truck-availability constraint.
DAY_NIGHT_CUTOFF_H = 12

def _is_day_sheet(start_time):
    """True when the route's start_time is a day shift (before noon)."""
    if start_time is None:
        return True
    t = start_time.time() if isinstance(start_time, datetime) else start_time
    return t.hour < DAY_NIGHT_CUTOFF_H

MWO_COL = next(i for i, (_, k) in enumerate(COLS) if k == "_mwo")

# Sheet names (exact, case-insensitive) that the solver leaves untouched.
SOLVER_SKIP_SHEETS = {"1603", "1604",
                      "1531", "1021", "1031", "1071", "1125",
                      "1081", "1451", "1281", "1441", "1121", "1561", "1211",
                      "1421", "1431", "1023",
                      "1123", "1521", "1551", "1381"}

# Mennonite farms — no pickup on Sundays.  Highlighted in the route viewer
# for planner awareness; solver logic is not affected.
MENNONITE_FARMS = {
    "92-590", "07-400", "99-020", "64-386",
    "92-808", "53-530", "92-904", "59-218",
}

# Default plant receiving windows (open HH:MM, close HH:MM).
# None means 24/7 - no restriction.  Overnight windows (close < open) are
# handled by time_in_window().  These are hard-coded from the
# "Plant_Receiving_Windows" reference sheet; the user can override them in
# the Solver tab at run time.
PLANT_RECEIVING_WINDOWS = {
    "909312": ("00:00", "23:59"),  # Agropur Burnaby        - 24/7
    "972711": ("05:00", "23:00"),  # Saputo Port Coquitlam  - 5am–11pm
    "902011": ("06:00", "18:00"),  # Avalon Dairy           - 6am–6pm
    "907011": ("10:00", "16:00"),  # Birchwood Dairy        - 10am–4pm
    "906011": ("06:00", "18:00"),  # Dhaliwal Dairy         - 6am–6pm
    "965713": ("06:00", "18:00"),  # First Choice           - 6am–6pm
    "911011": ("18:00", "20:00"),  # Golden Ears PM         - 6pm–8pm
    "918011": ("06:00", "18:00"),  # Khalsa FY              - 6am–6pm
    "901012": ("06:00", "18:00"),  # Olympic Dairy          - 6am–6pm
    "905011": ("06:00", "18:00"),  # Pinnacle Dairy         - 6am–6pm
    "916011": ("06:00", "18:00"),  # Prabu Foods            - 6am–6pm
    "951305": ("06:00", "18:00"),  # Reva Foods             - 6am–6pm
    "917011": ("08:00", "11:00"),  # Ridgecrest             - 8am–11am
    "981301": ("08:00", "11:00"),  # WOW Foods              - 8am–11am
    "912011": ("05:00", "17:00"),  # Meadowfresh (regular)  - 5am–5pm
    "908011": ("05:00", "00:30"),  # Punjab                 - 5am–12:30am (overnight)
    "915011": ("00:00", "23:59"),  # Vitalus Abbotsford     - 24/7
    "972712": ("06:00", "23:59"),  # Saputo Abbotsford      - 6am–midnight
    "902013": ("06:00", "18:00"),  # GRASSFED Avalon        - 6am–6pm
    "907012": ("06:00", "18:00"),  # GRASSFED Birchwood     - 6am–6pm
    "929011": ("06:00", "18:00"),  # Earth's Own / A2       - 6am–6pm
    "912015": ("05:00", "17:00"),  # GRASSFED Meadowfresh   - 5am–5pm
    "913011": ("06:00", "08:00"),  # Farmhouse Agassiz      - 6am–8am
}

# Time windows where a delivery is still allowed (the plant is open per
# PLANT_RECEIVING_WINDOWS above) but heavily discouraged - e.g. another
# division's trucks need the same dock during this slot, so ours should use
# it outside this window whenever the route can reasonably avoid it.
# {dest_key: [(start_str, end_str), ...]} - a dest can have more than one
# avoid-window if ever needed, though normally just one.  Separate from
# PLANT_RECEIVING_WINDOWS: a dest can be open the whole time and still carry
# an avoid-window inside that open period.
AVOID_WINDOWS = {
    "972712": [("19:00", "22:00")],  # Saputo Abbotsford - shared dock, another
                                       # division needs 7-10pm
}

# Most processors can only receive one truck at a time - the overlap penalty
# and the Processor Schedule chart's red-border highlighting both treat any
# 2+ simultaneous trucks as a violation by default.  A few processors have
# multiple unloading bays and can genuinely take more than one truck at once;
# {dest_key: capacity}.  Any dest_key not listed here defaults to capacity 1.
# Note this is a CAPACITY, not a blanket exemption - e.g. with capacity 2,
# two trucks overlapping is fine, but a third truck overlapping both of them
# still gets flagged and penalized.
PROCESSOR_DOCK_CAPACITY = {
    "972711": 2,   # Saputo Port Coquitlam - 2 unloading bays
    "972712": 2,   # Saputo Abbotsford - 2 unloading bays
}

# Tray uses the same columns plus a "From Route" column
TRAY_COLS = COLS + [("Sheet / Route", "_from_route"), ("Type", "_day_colour")]

CLR_HEADER    = QColor("#1a3a5c")
CLR_HEADER_FG = QColor("#ffffff")
CLR_ROUTE_HDR = QColor("#2e6da4")
CLR_ROUTE_FG  = QColor("#ffffff")
CLR_SUBTOTAL  = QColor("#d0e4f7")
CLR_TOTAL     = QColor("#a8c8f0")
CLR_ALT       = QColor("#f0f6fc")
CLR_WHITE     = QColor("#ffffff")
CLR_DEPOT     = QColor("#e8f5e9")
CLR_DEST      = QColor("#fff3e0")
CLR_DEST_WARN = QColor("#ffccbc")   # orange-red: dest arrival outside plant window
CLR_REMOVED   = QColor("#fff8e1")
CLR_RED       = QColor("#c0392b")
CLR_RED_BG    = QColor("#fdecea")
CLR_CHANGED   = QColor("#fff9c4")   # highlight for changed farms in comparison

# -- Helpers -------------------------------------------------------------------

def get_exe_dir():
    """Directory of the executable (or script). Used for the browse-default root."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent

def get_data_dir():
    """Directory where bundled data files live.
    When frozen by PyInstaller --onefile, data files are unpacked to
    sys._MEIPASS at runtime, not next to the exe."""
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            return Path(meipass)
        return Path(sys.executable).parent
    return Path(__file__).parent

def get_matrix_dir():
    """Directory where the distance/duration CSVs live.

    Always returns the folder containing the exe (or this script when running
    from source).  The matrices are kept OUTSIDE the PyInstaller bundle so
    they can be updated with add_farm.py without rebuilding the exe.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def fmt_time(v):
    if isinstance(v, dt_time): return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, str): return _sanitise_time_str(v)
    return ""

def _sanitise_time_str(s):
    """Normalise common malformed milking-time strings to HH:MM.

    Handles:
      "5;00"  -> "5:00"   semicolon separator
      "5.00"  -> "5:00"   period separator
      "500"   -> "5:00"   3-digit bare number  (H + MM)
      "1500"  -> "15:00"  4-digit bare number  (HH + MM)
      "05:00" -> "05:00"  already correct, pass through
      "ROBOT" -> "ROBOT"  special keyword, pass through
      "abc"   -> ""       non-numeric garbage, blanked
      "5:"    -> ""       incomplete, blanked
    """
    if not isinstance(s, str):
        return s
    s = s.strip()
    if not s or s == "-":
        return s
    # Special keyword - pass through unchanged
    if s.upper() == "ROBOT":
        return s
    # Semicolon -> colon
    if ";" in s:
        s = s.replace(";", ":")
    # Period -> colon only when both sides are purely digits
    elif "." in s and ":" not in s:
        parts = s.split(".")
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            s = f"{parts[0]}:{parts[1]}"
    # Bare digits with no separator
    if ":" not in s and s.isdigit():
        if len(s) == 3:
            s = f"{s[0]}:{s[1:]}"      # "500"  -> "5:00"
        elif len(s) == 4:
            s = f"{s[:2]}:{s[2:]}"     # "1500" -> "15:00"
    # Final validation: must be HH:MM with numeric parts on both sides
    if ":" in s:
        parts = s.split(":")
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return s
    # Anything that doesn't look like a valid time gets blanked
    return ""

def extract_year(name):
    m = re.search(r"\d{4}", name)
    return m.group(0) if m else name

def month_key(name):
    n = name.lower()
    for i, m in enumerate(MONTH_ORDER):
        if m in n: return i
    return 99

# -- Milking window conflict --------------------------------------------------

@lru_cache(maxsize=4096)
def _parse_hhmm_cached(s):
    return _parse_hhmm_impl(s)

def parse_hhmm(s):
    """Cached front-end for _parse_hhmm_impl (hot path: millions of calls
    per solve on a small set of distinct window strings)."""
    try:
        return _parse_hhmm_cached(s)
    except TypeError:          # unhashable input - fall back uncached
        return _parse_hhmm_impl(s)

def _parse_hhmm_impl(s):
    """Parse 'HH:MM' string -> datetime.time, or return None.

    Also accepts:
    - datetime.time directly (pass-through) - handles cells that openpyxl
      returns as time objects rather than strings on some Excel files.
    - float Excel time serial (fraction of a day) - converts to time.
    - Common malformed strings normalised via _sanitise_time_str before
      parsing (e.g. "5;00", "500", "1500").
    """
    if s is None: return None
    if isinstance(s, dt_time): return s
    if isinstance(s, float):
        # Excel serial: fraction of 24h
        total_mins = round(s * 24 * 60)
        return dt_time(total_mins // 60 % 24, total_mins % 60)
    if not s or s == "-": return None
    s = _sanitise_time_str(str(s))
    try:
        h, m = s.split(":")
        return dt_time(int(h), int(m))
    except (ValueError, AttributeError):
        return None

def time_in_window(t, start_s, finish_s):
    """Return True if time t falls within [start_s, finish_s] (HH:MM strings).
    Handles overnight windows (finish < start)."""
    if t is None: return False
    ts = parse_hhmm(start_s); tf = parse_hhmm(finish_s)
    if ts is None or tf is None: return False
    if ts <= tf:
        return ts <= t <= tf
    else:  # overnight
        return t >= ts or t <= tf

_MONTH_NAMES = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
}


def _parse_sheet_date_str(val):
    """Parse date strings from AI1 into a datetime.date.

    Handles the format variants seen in real workbooks:
      'Sunday June 7 2026'        – no commas
      'Monday June 8, 2026'       – comma after day
      'Saturday, May 23, 2026'    – comma after weekday too
      'Tuesday, April 28 th'      – ordinal suffix, no year (uses current year)
    Returns None if the value cannot be parsed.
    """
    if not isinstance(val, str):
        return None
    # Strip ordinal suffixes ('28 th', '1st', '2nd', '3rd') before matching
    s = re.sub(r'(\d)\s*(?:st|nd|rd|th)\b', r'\1', val.strip(), flags=re.IGNORECASE)
    # Match: <Month> <Day> [,] [<Year>]  — weekday prefix is ignored by re.search
    m = re.search(r'(\w+)\s+(\d{1,2}),?\s*(\d{4})?', s)
    if not m:
        return None
    mon = _MONTH_NAMES.get(m.group(1).lower())
    if not mon:
        return None
    try:
        yr = int(m.group(3)) if m.group(3) else date.today().year
        return date(yr, mon, int(m.group(2)))
    except ValueError:
        return None


def _extract_row1_date(ws):
    """Read the sheet date from the merged cell AI1 (C_SHEET_DATE = 35).

    The cell contains a string formatted as '[Day of week] [Month] [Day#], [Year]'
    e.g. 'Monday July 14, 2025'.  openpyxl returns the merged region's value
    from the top-left cell only; the rest of the merged range is None.
    Also handles the (unlikely) case where openpyxl returns an actual
    datetime.date/datetime.datetime object instead of a string.
    """
    from datetime import datetime as _dt
    val = ws.cell(1, C_SHEET_DATE).value
    if isinstance(val, date) and not isinstance(val, _dt):
        return val
    if isinstance(val, _dt):
        return val.date()
    return _parse_sheet_date_str(val)


def _sheets_date_str(cache, fname):
    """Return a compact date string from the sheet_date fields in the cache.

    Single date  -> "Jul_14"    (used as PDF filename prefix)
    Date range   -> "Jul_14-16" if same month, "Jul_14-Aug_02" across months
    No dates     -> "" (caller falls back to workbook filename stem)
    """
    dates = sorted(
        {entry["sheet_date"]
         for entry in cache.get(fname, {}).values()
         if isinstance(entry, dict) and entry.get("sheet_date")
         and ("RED" in (entry.get("day_colour") or "")
              or "BLUE" in (entry.get("day_colour") or ""))}
    )
    if not dates:
        return ""
    lo, hi = dates[0], dates[-1]
    if lo == hi:
        return f"{lo.strftime('%b')}_{lo.day}"              # "Jul_14"
    if lo.month == hi.month:
        return f"{lo.strftime('%b')}_{lo.day}-{hi.day}"     # "Jul_14-16"
    return f"{lo.strftime('%b')}_{lo.day}-{hi.strftime('%b')}_{hi.day}"  # "Jul_14-Aug_2"


def _extended_milking_window(start_str, finish_str):
    """Return (ext_start, ext_finish) as datetime.time with regulatory buffers.

    Extends the raw milking window [start_str, finish_str] by
    MILKING_PRE_BUFFER_MINS before the start and MILKING_POST_BUFFER_MINS
    after the finish.  The extended zone is treated as a complete no-pickup
    window by the solver and cost functions.  Returns (None, None) if either
    boundary cannot be parsed.  time_in_window() already handles overnight
    windows, so wrapping-past-midnight extended times are safe.
    """
    ts = parse_hhmm(start_str)
    tf = parse_hhmm(finish_str)
    if ts is None or tf is None:
        return None, None
    today = date.today()
    ext_s = (datetime.combine(today, ts) - timedelta(minutes=MILKING_PRE_BUFFER_MINS)).time()
    ext_f = (datetime.combine(today, tf) + timedelta(minutes=MILKING_POST_BUFFER_MINS)).time()
    return ext_s, ext_f


def arrives_during_milking(arr_time, row_data, suppress_no_milking=True):
    """Return True if arr_time (datetime.time) falls within any milking window.
    Supports w1/w2 from the sheet and w3 from THREE_WINDOW_FARMS.
    If suppress_no_milking is True, farms in NO_MILKING_WINDOW_FARMS always return False."""
    if arr_time is None: return False
    irma = row_data.get("irma", "")
    if suppress_no_milking and irma in NO_MILKING_WINDOW_FARMS:
        return False
    for start_key, finish_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
        ext_s, ext_f = _extended_milking_window(
            row_data.get(start_key,""), row_data.get(finish_key,""))
        if ext_s is not None and time_in_window(arr_time, ext_s, ext_f):
            return True
    # Third window from THREE_WINDOW_FARMS
    w3 = THREE_WINDOW_FARMS.get(irma)
    if w3:
        w3_start = w3.get("w3", [None, None])[0]
        w3_finish = w3.get("w3", [None, None])[1]
        ext_s3, ext_f3 = _extended_milking_window(w3_start, w3_finish)
        if ext_s3 is not None and time_in_window(arr_time, ext_s3, ext_f3):
            return True
    return False


# -- Distance matrix -----------------------------------------------------------

@lru_cache(maxsize=None)
def normalise_key(k):
    s = str(k).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): s = s[:-2]
    return s

def load_distance_matrix(path):
    dm = {}
    if not path.exists():
        logger.warning("Distance matrix not found at %s - distances unavailable", path)
        return dm
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        if not headers: return dm
        col_keys = [normalise_key(h) for h in headers[1:]]
        for row in reader:
            if not row: continue
            rk = normalise_key(row[0])
            for j, ck in enumerate(col_keys):
                try:
                    raw = row[j+1].strip()
                    if not raw: continue
                    val = float(raw)
                except (ValueError, IndexError):
                    continue
                dm[(rk, ck)] = val
                dm[(ck, rk)] = val
    return dm

def lookup(dm, a, b):
    # Fast path: solver-state keys are already normalised, so try the raw pair
    # first and only pay for normalise_key on a miss (raw Excel values etc.).
    key = (a, b)
    if key in dm:
        return dm[key]
    return dm.get((normalise_key(a), normalise_key(b)))

def _block_dest_keys(block):
    """Return ordered list of destination keys for a block (may be multiple)."""
    dests = block.get("dests") or []
    if dests:
        return [d["key"] for d in dests if d.get("key")]
    # Legacy fallback
    dk = block.get("dest_key", "")
    return [dk] if dk else []


def _block_last_dest_key(block):
    """The final destination key of a block (used as origin of next block)."""
    keys = _block_dest_keys(block)
    return keys[-1] if keys else ""


def _dest_stop_index(block, d_i, b_idx, blocks):
    """Return the btimes/dists index for the d_i-th dest in block,
    respecting split_after. Preload blocks always put dest at index 1+d_i."""
    if block.get("preload") and not block.get("rows"):
        return 1 + d_i
    is_last = (b_idx == len(blocks) - 1)
    origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
    stops   = _build_block_stops(block, origin, is_last)
    dest_stops = [s for s in stops if s["type"] == "dest"]
    if d_i < len(dest_stops):
        return dest_stops[d_i]["_si"]
    return len(block.get("rows",[])) + 1 + d_i   # fallback


def _farm_stop_index(block, f_i, b_idx, blocks):
    """Return the btimes/dists index for the f_i-th farm in block,
    respecting any split_after dests that appear before this farm."""
    is_last = (b_idx == len(blocks) - 1)
    origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
    stops   = _build_block_stops(block, origin, is_last)
    farm_stops = [s for s in stops if s["type"] == "farm"]
    if f_i < len(farm_stops):
        return farm_stops[f_i]["_si"]
    return f_i + 1   # fallback


def _block_has_split(block):
    """Return True if any dest in this block has a split_after index set,
    meaning it should be visited mid-route rather than after all farms."""
    dests = block.get("dests") or []
    return any(d.get("split_after") is not None for d in dests)


def _build_block_stops(block, origin, is_last):
    """Build the ordered stop sequence for a block, respecting split_after.

    Returns a list of dicts, each with:
      type:   'origin' | 'farm' | 'dest' | 'vedder'
      key:    irma or dest_key string (for distance lookup)
      farm:   farm row dict (type=='farm' only)
      dest:   dest dict (type=='dest' only)

    split_after on a dest dict means "insert this dest after farm index N".
    None / missing means after all farms (legacy behaviour).
    """
    farms = block.get("rows", [])
    dests = block.get("dests") or []
    if not dests:
        dk = block.get("dest_key","")
        dn = block.get("dest_name","") or dk
        dests = [{"key": dk, "name": dn, "vol_partial": None}] if dk else []

    stops = [{"type": "origin", "key": origin}]

    # Group dests by their split_after index.
    # split_after=None means after all farms (index = len(farms)).
    dests_by_split = defaultdict(list)
    for d in dests:
        sa = d.get("split_after")
        if sa is None:
            sa = len(farms)   # after all farms
        dests_by_split[sa].append(d)

    for f_idx, farm in enumerate(farms):
        stops.append({"type": "farm", "key": farm.get("irma",""), "farm": farm})
        # Insert any dests that split after this farm.
        # Skip f_idx+1 == len(farms) - those are handled by the "after all farms"
        # loop below to avoid double-insertion.
        if f_idx + 1 < len(farms):
            for d in dests_by_split.get(f_idx + 1, []):
                stops.append({"type": "dest", "key": d.get("key",""), "dest": d})

    # Dests with split_after=len(farms) (after all farms)
    for d in dests_by_split.get(len(farms), []):
        stops.append({"type": "dest", "key": d.get("key",""), "dest": d})

    if is_last:
        stops.append({"type": "vedder", "key": "VEDDER"})

    # Tag each stop with its sequence index for btimes/dists lookups
    for i, s in enumerate(stops):
        s["_si"] = i

    return stops


def _pdf_from_text(text, title, parent=None, fname="", date_str=""):
    """Save monospace text content to a portrait A4 PDF via QPrinter."""
    _prefix = (date_str + "_") if date_str else (Path(fname).stem + "_" if fname else "")
    path, _ = QFileDialog.getSaveFileName(
        parent, f"Export {title}",
        _prefix + title.replace(" ", "_").replace("/", "-") + ".pdf",
        "PDF Files (*.pdf)")
    if not path:
        return
    printer = QPrinter(QPrinter.HighResolution)
    printer.setOutputFormat(QPrinter.PdfFormat)
    printer.setOutputFileName(path)
    printer.setPageSize(QPrinter.A4)
    printer.setOrientation(QPrinter.Portrait)
    doc = QTextDocument()
    doc.setDefaultFont(QFont("Courier New", 7))
    doc.setPlainText(text)
    doc.print_(printer)


def _pdf_from_widget(widget, title, parent=None, landscape=True, fname="", date_str=""):
    """Render a QWidget directly to a landscape (default) A4 PDF at printer resolution."""
    _prefix = (date_str + "_") if date_str else (Path(fname).stem + "_" if fname else "")
    path, _ = QFileDialog.getSaveFileName(
        parent, f"Export {title}",
        _prefix + title.replace(" ", "_").replace("/", "-") + ".pdf",
        "PDF Files (*.pdf)")
    if not path:
        return

    printer = QPrinter(QPrinter.HighResolution)
    printer.setOutputFormat(QPrinter.PdfFormat)
    printer.setOutputFileName(path)
    printer.setPageSize(QPrinter.A4)
    printer.setOrientation(
        QPrinter.Landscape if landscape else QPrinter.Portrait)

    painter = QPainter(printer)
    page    = painter.viewport()

    # Use the widget's full content size (minimumSize captures the full canvas
    # even if parts are scrolled out of view in the parent scroll area)
    ww = max(widget.minimumWidth(),  widget.width(),  100)
    wh = max(widget.minimumHeight(), widget.height(), 100)

    # Scale to fit page, preserving aspect ratio, centred
    s  = min(page.width() / ww, page.height() / wh)
    ox = (page.width()  - ww * s) / 2
    oy = (page.height() - wh * s) / 2

    # White background
    painter.fillRect(page, Qt.white)
    painter.translate(ox, oy)
    painter.scale(s, s)

    # Render widget directly to printer painter (vector, no pixmap intermediary)
    widget.render(painter, flags=QWidget.DrawChildren)
    painter.end()



def _route_stop_segments(blocks, all_times, start_mins):
    """Extract per-stop timing segments from calc_times output.

    Returns list of (kind, arr_m, dep_m) where kind is 'farm' or 'processor',
    arr_m / dep_m are absolute minutes since midnight (with past-midnight
    correction so values are monotonically increasing).
    """
    segments = []
    prev_dep = start_mins
    for b_idx, (block, b_times) in enumerate(zip(blocks, all_times)):
        is_last = (b_idx == len(blocks) - 1)
        origin  = ("VEDDER" if b_idx == 0
                   else (_block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"))
        stops   = _build_block_stops(block, origin, is_last)
        for stop, timing in zip(stops, b_times):
            if stop["type"] in ("origin", "vedder"):
                continue
            arr_t = timing.get("arr")
            dep_t = timing.get("dep")
            if arr_t is None or dep_t is None:
                continue
            arr_m = arr_t.hour * 60 + arr_t.minute
            dep_m = dep_t.hour * 60 + dep_t.minute
            # Correct for midnight wrap
            while arr_m < prev_dep - 5:
                arr_m += 24 * 60
            while dep_m < arr_m:
                dep_m += 24 * 60
            prev_dep = dep_m
            kind = "farm" if stop["type"] == "farm" else "processor"
            segments.append((kind, arr_m, dep_m))
    return segments


def _is_holdover_block(block):
    """Return True if every destination in this block is a 'Yard for ...' location.

    These blocks end at a yard trailer - the farms are collected the same day
    but the milk sits overnight before being picked up.  The FARM ORDER is
    still optimisable; only the final yard destination must stay constant.
    Use _is_preload_block() to detect the next-morning pickup blocks that must
    be held completely frozen.
    """
    dests = block.get("dests") or []
    if not dests:
        dk = block.get("dest_key", "")
        dn = block.get("dest_name", "")
        dests = [{"name": dn, "key": dk}] if (dk or dn) else []
    if not dests:
        return False
    return all("yard for" in (d.get("name", "") or "").lower() for d in dests)


def _is_preload_block(block):
    """Return True if this block is a preload offload - no farms, just a
    delivery to a processor at the start of the day (previous night's load).

    These blocks must be held completely constant: no farms added, destination
    not changed.  Identified by the 'preload' flag set during parsing.
    """
    return bool(block.get("preload")) and not block.get("rows")


def _is_fixed_vol_block(block):
    """Return True if this block has a fixed partial-volume delivery on its
    last destination (no catch-all remainder dest).

    When every destination in a block has an explicit vol_partial, the total
    deliverable volume is capped at their sum.  Any farm volume above that cap
    is silently dropped from accounting.  Rather than trying to optimise around
    this, the solver holds these blocks completely frozen - the dispatcher has
    explicitly specified the volumes and the solver has no business changing the
    farm assignment.
    """
    dests = block.get("dests") or []
    if not dests:
        return False
    return dests[-1].get("vol_partial") is not None


def calc_distances(blocks, dm):
    """
    Routing rules (multi-destination aware, split_after supported):
      - First block starts from VEDDER.
      - Each subsequent block starts from the last destination of the previous block.
      - Within a block: stops follow _build_block_stops order, which interleaves
        partial-dropoff dests at their split_after position.
      - Last block appends a return leg to VEDDER after the final destination.
    Returns list of per-block distance lists (one entry per stop, None sentinel last).
    """
    n = len(blocks)
    all_dists = []
    for b_idx, block in enumerate(blocks):
        is_last = (b_idx == n - 1)
        origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        stops   = _build_block_stops(block, origin, is_last)
        keys    = [s["key"] for s in stops]
        dists   = [lookup(dm, keys[i], keys[i+1]) if i < len(keys)-1 else None
                   for i in range(len(keys))]
        all_dists.append(dists)
    return all_dists

def calc_durations(blocks, dm_dur):
    """
    Parallel to calc_distances but returns drive-time minutes from the duration
    matrix.  Respects split_after via _build_block_stops.
    """
    n = len(blocks)
    all_durs = []
    for b_idx, block in enumerate(blocks):
        is_last = (b_idx == n - 1)
        origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        stops   = _build_block_stops(block, origin, is_last)
        keys    = [s["key"] for s in stops]
        durs    = [lookup(dm_dur, keys[i], keys[i+1]) if i < len(keys)-1 else None
                   for i in range(len(keys))]
        all_durs.append(durs)
    return all_durs


# -- Time estimation ----------------------------------------------------------

def mins_to_time(start_time, offset_mins):
    """Add offset_mins (float) to a datetime.time. Returns datetime.time."""
    dt = datetime.combine(date.today(), start_time) + timedelta(minutes=offset_mins)
    return dt.time()

def fmt_hhmm(t):
    if t is None: return "-"
    return f"{t.hour:02d}:{t.minute:02d}"

def _continuous_minutes(t, start_time):
    """Convert a wall-clock datetime.time to minutes since the reference
    midnight of the day a sheet's route starts on, adding 24h if t is
    clearly "earlier" than the sheet's own start time - which can only mean
    it's actually the next calendar day (a route running into overtime past
    midnight).  Used to lay out arrival/departure times on one continuous,
    shared axis for the processor schedule chart.
    """
    if t is None or start_time is None:
        return None
    base_min = start_time.hour * 60 + start_time.minute
    t_min    = t.hour * 60 + t.minute
    if t_min < base_min - 60:   # more than an hour "before" start -> wrapped
        t_min += 24 * 60
    return t_min

def _min_to_hhmm(m):
    """Format continuous minutes back to a real wall-clock HH:MM (wraps at
    24h, since this is for display of an actual clock time, not the
    internal chart coordinate)."""
    if m is None:
        return "-"
    m = int(m) % (24 * 60)
    return f"{m // 60:02d}:{m % 60:02d}"

def day_colour_style(day_colour):
    """Return (bg QColor, fg QColor, display_text) for a day colour string."""
    dc = day_colour.upper().replace("\n", " ").strip()
    if "GRASSFED"  in dc or "GRASS FED" in dc or "GRASS" in dc:
        return QColor("#43a047"), QColor("#ffffff"), day_colour.replace("\n"," ")
    if "A2"        in dc: return QColor("#b3e5fc"), QColor("#0d47a1"), day_colour
    if "RED"       in dc: return QColor("#e53935"), QColor("#ffffff"), day_colour
    if "BLUE"      in dc: return QColor("#1e88e5"), QColor("#ffffff"), day_colour
    if "CREAM"     in dc: return QColor("#fff9c4"), QColor("#5d4037"), day_colour
    if "SKIM"      in dc: return QColor("#e0f7fa"), QColor("#006064"), day_colour
    if "WPC"       in dc: return QColor("#f3e5f5"), QColor("#6a1b9a"), day_colour
    if "HOLD"      in dc: return QColor("#ffccbc"), QColor("#bf360c"), day_colour.replace("\n"," ")
    if dc:                return QColor("#9e9e9e"), QColor("#ffffff"), day_colour
    return None, None, ""


def stop_duration(vol_litres):
    """Minutes spent at a stop: setup + pump time."""
    pump = (vol_litres / PUMP_RATE_LPM) if vol_litres else 0.0
    return ONSITE_MIN + pump

def drive_mins(km):
    """Minutes to drive km at DRIVE_SPEED_KMH."""
    if km is None: return None
    return (km / DRIVE_SPEED_KMH) * 60.0

def _dest_vol_partial(dest_dict, total_farm_vol, already_delivered):
    """
    Work out how many litres are offloaded at this destination stop.
    dest_dict: {name, key, vol_partial}  vol_partial=None means "rest of load".
    already_delivered: total already dropped at earlier dests in same block.
    """
    remaining = max(0.0, total_farm_vol - already_delivered)
    vp = dest_dict.get("vol_partial")
    if vp is None:
        return remaining
    return min(float(vp), remaining)


def _block_dest_offloads(block):
    """Litres offloaded per destination key for a single block -> {key: litres}.

    Walks the block's destinations in order, assigning each its vol_partial
    (or the remaining farm volume when vol_partial is None), and aggregates by
    destination key.  When the block has no explicit dests, falls back to
    dest_key / dest_name.  This is the single source of truth for the
    volume-accounting loop shared by the cost, penalty, and diagnostic code.
    """
    dests_b = block.get("dests") or []
    if not dests_b:
        dk = block.get("dest_key") or block.get("dest_name") or "?"
        dests_b = [{"key": dk, "vol_partial": None}]
    farm_vol = sum((r.get("prior_vol") or 0) for r in block.get("rows", [])
                   if isinstance(r.get("prior_vol"), (int, float)))
    out = {}
    already = 0.0
    for d in dests_b:
        dk = d.get("key") or "?"
        vp = d.get("vol_partial")
        rem = max(0.0, farm_vol - already)
        offload = min(float(vp), rem) if vp is not None else rem
        already += offload
        out[dk] = out.get(dk, 0.0) + offload
    return out


def calc_times(blocks, dm, start_time, dm_dur=None, suppress_no_milking=True,
               precomputed_dists=None):
    """
    Returns list of per-block time lists in parallel with calc_distances.
    Each block: list of dicts {arr, dep, wait} per stop, ordered by
    _build_block_stops (which respects split_after for mid-route partial dropoffs).

    Multi-destination with split_after: a dest with split_after=N is visited
    after farm N in the sequence, allowing partial dropoff mid-route.

    precomputed_dists: optional output of calc_distances(blocks, dm) for the same
    blocks.  Passing it avoids recomputing the distance legs when the caller has
    already done so (e.g. _sheet_cost), which roughly halves the distance work in
    the solver's hot path.
    """
    if start_time is None:
        return None

    all_dists    = precomputed_dists if precomputed_dists is not None \
                   else calc_distances(blocks, dm)
    all_dur_lists = calc_durations(blocks, dm_dur) if dm_dur else None
    all_times    = []
    base   = datetime.combine(date.today(), start_time)
    cursor = base

    for b_idx, block in enumerate(blocks):
        dists  = all_dists[b_idx]
        durs   = all_dur_lists[b_idx] if all_dur_lists else None
        is_last = (b_idx == len(blocks) - 1)
        origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        stops   = _build_block_stops(block, origin, is_last)
        block_times = []

        def _leg_mins(s_idx):
            if durs is not None and s_idx < len(durs) and durs[s_idx] is not None:
                return durs[s_idx]
            d = dists[s_idx] if s_idx < len(dists) else None
            return drive_mins(d)

        # Preload block: no farms, drive straight to dests then wash
        if block.get("preload") and not block.get("rows"):
            already_del  = 0.0
            real_dests   = [s for s in stops if s["type"] == "dest"
                            and "yard for" not in (s["dest"].get("name","") or "").lower()]
            n_real       = len(real_dests)
            real_seen    = 0
            for s_idx, stop in enumerate(stops):
                if stop["type"] == "origin":
                    if b_idx == 0:
                        dep_dt = cursor + timedelta(minutes=VEDDER_DEPART_EXTRA_MINS)
                        block_times.append({"arr": cursor.time(), "dep": dep_dt.time(), "wait": None})
                        cursor = dep_dt
                    else:
                        block_times.append({"arr": cursor.time(), "dep": cursor.time(), "wait": None})
                    continue
                if stop["type"] == "dest":
                    d      = stop["dest"]
                    is_yard = "yard for" in (d.get("name","") or "").lower()
                    dm_m   = 0.0 if is_yard else _leg_mins(s_idx - 1)
                    if dm_m is None:
                        block_times.append({"arr": None, "dep": None, "wait": None})
                        continue
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    vp     = d.get("vol_partial")
                    rem    = max(0.0, PRELOAD_VOL - already_del)
                    offload = min(float(vp), rem) if vp is not None else rem
                    already_del += offload
                    dep_dt = arr_dt + timedelta(minutes=stop_duration(offload))
                    if not is_yard:
                        real_seen += 1
                        if real_seen < n_real:
                            dep_dt += timedelta(minutes=INTER_PROCESSOR_BREAK)
                    block_times.append({"arr": arr_dt.time(), "dep": dep_dt.time(), "wait": None})
                    cursor = dep_dt
            cursor += timedelta(minutes=PRELOAD_WASH_MINS)
            all_times.append(block_times)
            continue

        # Normal block: farms possibly interleaved with partial-dropoff dests
        total_farm_vol  = sum((r.get("prior_vol") or 0) for r in block.get("rows",[])
                               if isinstance(r.get("prior_vol"), (int, float)))
        already_del     = 0.0
        real_dests      = [s for s in stops if s["type"] == "dest"
                           and "yard for" not in (s["dest"].get("name","") or "").lower()]
        n_real_dests    = len(real_dests)
        real_dest_seen  = 0

        for s_idx, stop in enumerate(stops):
            stype = stop["type"]

            if stype == "origin":
                if b_idx == 0:
                    dep_dt = cursor + timedelta(minutes=VEDDER_DEPART_EXTRA_MINS)
                    block_times.append({"arr": cursor.time(), "dep": dep_dt.time(), "wait": None})
                    cursor = dep_dt
                else:
                    block_times.append({"arr": cursor.time(), "dep": cursor.time(), "wait": None})

            elif stype == "farm":
                farm   = stop["farm"]
                dm_m   = _leg_mins(s_idx - 1)
                if dm_m is None or cursor is None:
                    block_times.append({"arr": None, "dep": None, "wait": None})
                else:
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    vol    = farm.get("prior_vol")
                    # Zero-vol farms in a paired set (same IRMA as another farm
                    # in the block with non-zero vol) contribute 0 minutes -
                    # the truck arrives, connects the trailer, and leaves with
                    # no pump time and no setup time.  arr == dep, no wait check.
                    if isinstance(vol, (int, float)) and vol == 0:
                        block_times.append({"arr": arr_dt.time(), "dep": arr_dt.time(),
                                            "wait": None})
                        cursor = arr_dt
                        continue
                    vol    = vol or 0
                    dur    = stop_duration(vol if isinstance(vol, (int, float)) else 0)
                    irma   = farm.get("irma","")
                    wait_mins = 0.0

                    # MWO (Milking Window Override): skip all milking window
                    # checks for this farm - truck arrives and pumps immediately.
                    if farm.get("_mwo"):
                        pass
                    elif suppress_no_milking and irma in NO_MILKING_WINDOW_FARMS:
                        pass
                    else:
                        for s_key, f_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
                            ext_s, ext_f = _extended_milking_window(
                                farm.get(s_key,""), farm.get(f_key,""))
                            if ext_s is not None and time_in_window(arr_dt.time(), ext_s, ext_f):
                                # Wait until extended finish (window end + post buffer)
                                end_w = datetime.combine(arr_dt.date(), ext_f)
                                if end_w <= arr_dt: end_w += timedelta(days=1)
                                wait_mins = (end_w - arr_dt).total_seconds() / 60.0
                                break
                        if wait_mins == 0.0:
                            w3data = THREE_WINDOW_FARMS.get(irma)
                            if w3data:
                                w3s, w3f = w3data.get("w3",[None,None])
                                ext_s3, ext_f3 = _extended_milking_window(w3s, w3f)
                                if ext_s3 is not None and time_in_window(arr_dt.time(), ext_s3, ext_f3):
                                    end_w3 = datetime.combine(arr_dt.date(), ext_f3)
                                    if end_w3 <= arr_dt: end_w3 += timedelta(days=1)
                                    wait_mins = (end_w3 - arr_dt).total_seconds() / 60.0

                    dep_dt = arr_dt + timedelta(minutes=wait_mins + dur)
                    block_times.append({"arr": arr_dt.time(), "dep": dep_dt.time(),
                                        "wait": wait_mins if wait_mins > 0 else None})
                    cursor = dep_dt

            elif stype == "dest":
                d       = stop["dest"]
                is_yard = "yard for" in (d.get("name","") or "").lower()
                dm_m    = 0.0 if is_yard else _leg_mins(s_idx - 1)
                # Volume collected so far up to this split point
                if is_yard:
                    offload = 0.0
                else:
                    offload = _dest_vol_partial(d, total_farm_vol, already_del)
                already_del += offload
                if dm_m is None or cursor is None:
                    block_times.append({"arr": None, "dep": None, "wait": None})
                else:
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    dep_dt = arr_dt + timedelta(minutes=stop_duration(offload))
                    if not is_yard:
                        real_dest_seen += 1
                        if real_dest_seen < n_real_dests:
                            dep_dt += timedelta(minutes=INTER_PROCESSOR_BREAK)
                    block_times.append({"arr": arr_dt.time(), "dep": dep_dt.time(), "wait": None})
                    cursor = dep_dt

            elif stype == "vedder":
                dm_m = _leg_mins(s_idx - 1)
                if dm_m is None or cursor is None:
                    block_times.append({"arr": None, "dep": None, "wait": None})
                else:
                    arr_dt = cursor + timedelta(minutes=dm_m)
                    block_times.append({"arr": arr_dt.time(), "dep": arr_dt.time(), "wait": None})

        all_times.append(block_times)

    if not all_times:
        return None
    return all_times, cursor if cursor is not None else base


def total_route_dist(blocks, dm):
    """Sum all known distances across blocks. Returns (total_km, all_known bool)."""
    total = 0.0; ok = True
    for dists in calc_distances(blocks, dm):
        for d in dists[:-1]:
            if d is None: ok = False
            else: total += d
    return total, ok

# -- Excel parsing -------------------------------------------------------------

# -- Delivery-info column constants (1-based) ---------------------------------
# Col 1 = row number label (1., 2., 3.)
# Col 2 = partial volume (e.g. 6900) or "Full Load" or empty
# Col 6 = processor name
# Col 20 = processor key (numeric)
C_DEST_VOL  = 2    # partial volume column in delivery rows
C_DEST_NAME = 6    # processor name column
C_DEST_KEY  = 20   # processor key column


def _parse_dest_row(ws, r, ws_formula=None):
    """
    Parse one numbered delivery row.  Returns a dest dict:
      {name, key, vol_partial}
    vol_partial is a float if a specific partial volume was given,
    or None meaning "rest of load" (blank or "Full Load").
    Returns None if the row has no meaningful content.
    """
    dn_raw = ws.cell(r, C_DEST_NAME).value
    dk_raw = ws.cell(r, C_DEST_KEY).value
    # Must have at least a name or a key to be a real dest row
    dn = str(dn_raw).strip() if dn_raw else ""
    dk = (str(int(dk_raw)).strip()
          if isinstance(dk_raw, (int, float))
          else str(dk_raw or "").strip())
    if not dn and not dk:
        return None
    # "WASH AT VTL" and similar wash instructions are operational notes,
    # not processor destinations - exclude them.
    if "wash" in dn.lower():
        return None
    vol_raw = ws.cell(r, C_DEST_VOL).value
    if vol_raw is None and ws_formula is not None:
        vol_raw = _try_eval_formula(ws_formula.cell(r, C_DEST_VOL).value)
    if vol_raw is None or (isinstance(vol_raw, str)
                           and vol_raw.strip().lower() in ("", "full load")):
        vol_partial = None
    else:
        try:
            vol_partial = float(vol_raw)
        except (ValueError, TypeError):
            vol_partial = None
    return {"name": dn, "key": dk, "vol_partial": vol_partial}


class _FastCellCache:
    """Wraps a read_only openpyxl worksheet to make random .cell(r, c) access
    fast.

    Background: openpyxl's read_only mode is a streaming reader with no
    persistent in-memory model - every call to ws.cell(r, c) re-parses the
    sheet's XML from the start internally (via _cells_by_row / _reader.parse).
    That's fine for the sequential access pattern read_only mode was designed
    for (iter_rows top to bottom, once), but parse_sheet does repeated random
    access (checking several columns per row, jumping around for delivery
    sections), which is exactly the pattern read_only mode handles by
    redundantly re-parsing - turning a sub-second parse into several seconds
    per sheet, multiplied across every sheet in the workbook.

    This wrapper does ONE iter_rows() pass up front (the access pattern
    read_only mode is actually fast at) and caches every cell value in a
    plain dict, then answers all subsequent .cell(r, c).value calls from that
    dict - O(1) lookups instead of O(sheet size) re-parses.  Exposes just
    enough of the worksheet interface (.cell(), .max_row, .max_column) for
    parse_sheet's needs; not a general-purpose worksheet replacement.
    """

    class _CellProxy:
        __slots__ = ("value",)
        def __init__(self, value):
            self.value = value

    _EMPTY = None  # sentinel set below to a cached _CellProxy(None)

    def __init__(self, ws):
        self._values = {}
        max_r = 0
        max_c = 0
        for row in ws.iter_rows():
            for cell in row:
                v = getattr(cell, "value", None)
                r = getattr(cell, "row", None)
                c = getattr(cell, "column", None)
                if r is None or c is None:
                    continue   # EmptyCell in read_only mode - skip
                if v is not None:
                    self._values[(r, c)] = v
                if r > max_r: max_r = r
                if c > max_c: max_c = c
        self.max_row    = max_r
        self.max_column = max_c
        self._empty = _FastCellCache._CellProxy(None)

    def cell(self, row, column):
        v = self._values.get((row, column))
        if v is None:
            return self._empty
        return _FastCellCache._CellProxy(v)


def parse_sheet(ws, ws_formula=None):
    """Parse one worksheet.  ws is the data_only workbook sheet;
    ws_formula (optional) is the same sheet from a formula workbook - used
    to recover numeric values when the cached data value is None (e.g. the
    file was saved without recalculating formulas)."""

    # In read_only mode openpyxl yields EmptyCell objects for empty cells.
    # EmptyCell has .value=None but no .row/.column attributes - normalise
    # them to None so the rest of the parser never sees them.
    def _cell_value(cell):
        try:
            return cell.value
        except AttributeError:
            return None

    def _cell(r, c):
        """Read a cell, falling back to formula evaluation if the cached value is None."""
        val = ws.cell(r, c).value
        if val is None and ws_formula is not None:
            val = _try_eval_formula(ws_formula.cell(r, c).value)
        return val
    # Read driver start time from AZ1 (col 52, row 1)
    raw_start = ws.cell(1, C_DRIVER_START).value
    if isinstance(raw_start, dt_time):
        driver_start = raw_start
    else:
        driver_start = None

    # Read day colour from BJ1 (col 62, row 1)
    raw_colour = ws.cell(1, C_DAY_COLOUR).value
    day_colour = str(raw_colour).strip().upper() if raw_colour else ""

    # Determine the effective row range without walking every cell.
    # ws.max_row can be inflated by phantom rows (e.g. from Excel scrolling
    # or accidental edits far down the sheet).  Instead of iter_rows() which
    # materialises every cell, scan only the columns the parser cares about
    # to find the last row that actually has content.  This is O(real_data)
    # rather than O(max_row * max_col) and avoids stalls on large June files.
    MAX_SCAN_ROWS = min(ws.max_row, 5000)   # safety cap - no route sheet has 5000 rows
    SCAN_COLS = {C_IRMA, 2, 6}             # IRMA, col-2 (delivery/route headers), col-6 (dest)
    last_data_row = 0
    for r in range(1, MAX_SCAN_ROWS + 1):
        for c in SCAN_COLS:
            if ws.cell(r, c).value is not None:
                last_data_row = r
                break
    all_row_nums = list(range(1, last_data_row + 1))
    irma_header_rows = []
    # dest_list_rows maps irma_header_row -> [dest_dict, ...]
    dest_list_rows = {}
    # Delivery info found before any IRMA# row -> preload dests
    preload_dests = []

    i = 0
    while i < len(all_row_nums):
        r    = all_row_nums[i]
        val0 = ws.cell(r, C_IRMA).value
        c2   = ws.cell(r, 2).value

        if isinstance(val0, str) and val0.strip().upper() == "IRMA#":
            irma_header_rows.append(r)

        # Detect "Delivery Information:" block - grab numbered rows that follow
        if isinstance(c2, str) and "delivery" in c2.lower():
            hdr_key = irma_header_rows[-1] if irma_header_rows else None
            if hdr_key is not None:
                dest_list_rows.setdefault(hdr_key, [])
                j = i + 1
                while j < len(all_row_nums):
                    dr = all_row_nums[j]
                    # Stop at empty row or another section header
                    c2j = ws.cell(dr, 2).value
                    if c2j is None and ws.cell(dr, C_DEST_NAME).value is None:
                        break
                    d = _parse_dest_row(ws, dr, ws_formula=ws_formula)
                    if d:
                        dest_list_rows[hdr_key].append(d)
                    j += 1
                i = j
                continue
            else:
                # Delivery info before any IRMA# row - this is a preload offload block.
                # Capture dests into preload_dests (only the first such section).
                if not preload_dests:
                    j = i + 1
                    while j < len(all_row_nums):
                        dr  = all_row_nums[j]
                        c2j = ws.cell(dr, 2).value
                        if c2j is None and ws.cell(dr, C_DEST_NAME).value is None:
                            break
                        d = _parse_dest_row(ws, dr, ws_formula=ws_formula)
                        if d:
                            preload_dests.append(d)
                        j += 1
                    i = j
                    continue

        # Legacy "Destination:" label in col 6 (single-dest fallback)
        c6 = ws.cell(r, 6).value
        if isinstance(c6, str) and "destination" in c6.lower():
            hdr_key = irma_header_rows[-1] if irma_header_rows else None
            if hdr_key is not None and hdr_key not in dest_list_rows:
                nxt = i + 1
                if nxt < len(all_row_nums):
                    dr = all_row_nums[nxt]
                    d = _parse_dest_row(ws, dr, ws_formula=ws_formula)
                    if d:
                        dest_list_rows[hdr_key] = [d]

        i += 1

    blocks  = []
    current = None

    # If we found a preload delivery section before any IRMA# block, prepend it.
    if preload_dests:
        first_p = preload_dests[0]
        blocks.append({
            "route":     "",
            "dests":     preload_dests,
            "dest_name": first_p["name"],
            "dest_key":  first_p["key"],
            "rows":      [],
            "preload":   True,
        })
    for r in all_row_nums:
        val0 = ws.cell(r, C_IRMA).value
        if isinstance(val0, str) and val0.strip().upper() == "IRMA#":
            route_val = ws.cell(r, C_ROUTE).value or ""
            dests = dest_list_rows.get(r, [])
            # Legacy compat: if dests is empty check old dest_rows approach
            # Provide backward-compat fields dest_name/dest_key from first dest
            first = dests[0] if dests else {"name": "", "key": "", "vol_partial": None}
            current = {
                "route":     str(route_val).strip(),
                "dests":     dests,
                # Legacy aliases kept for compatibility with existing code:
                "dest_name": first["name"],
                "dest_key":  first["key"],
                "rows":      [],
            }
            blocks.append(current)
            continue
        if current is not None and isinstance(val0, str) and IRMA_RE.match(val0.strip()):
            # Named per-farm fields the rest of the code knows about. Anything
            # else non-empty on this row is captured into _extra_cells so it
            # travels with the farm when the solver moves it to a different row.
            # Without this, columns like farm name (R) and street address (AM)
            # would stay put while the IRMA in column A moved, leaving the
            # exported sheet showing farm A's name next to farm B's IRMA.
            NAMED_COLS = {C_IRMA, C_TRAIN, C_M1_START, C_M1_FINISH,
                          C_M2_START, C_M2_FINISH, C_EDPU, C_LOCATION,
                          C_PRIOR_VOL, C_ROUTE}
            extras = {}
            for c in range(1, ws.max_column + 1):
                if c in NAMED_COLS:
                    continue
                cv = ws.cell(r, c).value
                if cv is None:
                    continue
                if isinstance(cv, str) and cv.strip() == "":
                    continue
                extras[c] = cv
            current["rows"].append({
                "_uid":         str(uuid.uuid4()),
                "irma":         val0.strip(),
                "train":        ws.cell(r, C_TRAIN).value or "",
                "m1_start":     fmt_time(ws.cell(r, C_M1_START).value),
                "m1_finish":    fmt_time(ws.cell(r, C_M1_FINISH).value),
                "m2_start":     fmt_time(ws.cell(r, C_M2_START).value),
                "m2_finish":    fmt_time(ws.cell(r, C_M2_FINISH).value),
                "edpu":         ws.cell(r, C_EDPU).value or "",
                "location":     ws.cell(r, C_LOCATION).value or "",
                "prior_vol":    _cell(r, C_PRIOR_VOL),
                # Farms in NO_MILKING_WINDOW_FARMS are exempt from milking
                # window checks everywhere, but several different functions
                # (calc_times, _sheet_cost, _sheet_cost_breakdown, the table
                # display) each separately re-check IRMA membership gated by
                # a suppress_no_milking flag that has to be correctly threaded
                # through every call site.  That's fragile - miss one call
                # site and the farm isn't actually exempt there.  Setting
                # _mwo=True here instead means every one of those places
                # picks up the exemption automatically through the single,
                # already-correct MWO mechanism, with no separate threading
                # needed.  The checkbox shows as ticked in the UI immediately
                # on load, matching what's actually happening internally.
                "_mwo":         val0.strip() in NO_MILKING_WINDOW_FARMS,
                "_extra_cells": extras,
            })
    # Mark any remaining zero-farm blocks with a dest as preload
    # (handles edge case where IRMA# exists but no farms follow before next block)
    for block in blocks:
        if not block["rows"] and block.get("dests") and not block.get("preload"):
            block["preload"] = True

    return blocks, driver_start, day_colour

# -- Table item factories ------------------------------------------------------

def make_header_item(text, bg=CLR_HEADER, fg=CLR_HEADER_FG, bold=True, draggable=False):
    item = QTableWidgetItem(str(text))
    item.setBackground(bg); item.setForeground(fg)
    if bold:
        f = item.font(); f.setBold(True); item.setFont(f)
    flags = Qt.ItemIsEnabled
    if draggable:
        flags |= Qt.ItemIsDragEnabled | Qt.ItemIsSelectable
    item.setFlags(flags)
    item.setTextAlignment(Qt.AlignCenter)
    return item

def make_data_item(text, bg=CLR_WHITE, align=Qt.AlignCenter, fg=None, draggable=False):
    item = QTableWidgetItem(str(text) if text is not None else "")
    item.setBackground(bg)
    flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
    if draggable: flags |= Qt.ItemIsDragEnabled
    item.setFlags(flags)
    item.setTextAlignment(align)
    if fg: item.setForeground(fg)
    return item

# -- Route table renderer ------------------------------------------------------

def populate_table(table, blocks, dm, editable=False, start_time=None, dm_dur=None,
                   suppress_no_milking=True, plant_windows=None):
    table.clearSpans()
    table.clear()
    table.setColumnCount(len(COLS))
    table.setHorizontalHeaderLabels([c[0] for c in COLS])

    if not blocks:
        table.setRowCount(0)
        return

    all_dists  = calc_distances(blocks, dm)
    _ct = calc_times(blocks, dm, start_time, dm_dur=dm_dur, suppress_no_milking=suppress_no_milking)   # may be None
    all_times, _end_cursor = _ct if _ct is not None else (None, None)
    n_blocks   = len(blocks)
    # per block: banner + col_hdr + origin + farms + dest + subtotal
    # last block gets one extra row for VEDDER return
    # per block: banner(1) + col_hdr(1) + origin(1) + farms + dests + subtotal(1)
    # last block: +1 for VEDDER return
    def _ndests(b):
        d = b.get("dests") or []
        return max(len(d), 1)  # always at least one dest row
    total_rows = sum(2 + 1 + len(b["rows"]) + _ndests(b) + 1 for b in blocks) + 2
    if n_blocks > 0: total_rows += 1  # VEDDER return on last block
    table.setRowCount(total_rows)

    r = 0; day_dist = 0.0; day_vol = 0.0; day_ok = True

    for b_idx, block in enumerate(blocks):
        dists     = all_dists[b_idx]
        farms     = block["rows"]
        dests     = block.get("dests") or []
        if not dests:
            dk = block.get("dest_key",""); dn = block.get("dest_name","") or dk or "Destination"
            dests = [{"name": dn, "key": dk, "vol_partial": None}] if dk else []
        route_dist = 0.0; route_ok = True
        for d in dists[:-1]:
            if d is None: route_ok = False
            else: route_dist += d

        route_vol = sum((row["prior_vol"] or 0) for row in farms
                        if isinstance(row.get("prior_vol"), (int, float)))

        # Banner
        table.setSpan(r, 0, 1, len(COLS))
        banner_item = make_header_item(f"  Route: {block['route']}",
                                       bg=CLR_ROUTE_HDR, fg=CLR_ROUTE_FG,
                                       draggable=editable)
        banner_item.setData(Qt.UserRole + 2, b_idx)
        table.setItem(r, 0, banner_item)
        r += 1

        # Column sub-headers
        for c_idx, (hdr, _) in enumerate(COLS):
            table.setItem(r, c_idx, make_header_item(hdr))
        r += 1

        # Origin row (VEDDER for first block; previous processor for subsequent)
        if b_idx == 0:
            origin_key = "VEDDER"; origin_sub = "Depot"
        else:
            origin_key = _block_last_dest_key(blocks[b_idx-1]) or "VEDDER"
            prev_dests = blocks[b_idx-1].get("dests") or []
            origin_sub = (prev_dests[-1]["name"] if prev_dests else
                          blocks[b_idx-1].get("dest_name","")) or origin_key
        od = dists[0]
        od_str = f"{od:.1f}" if od is not None else "-"
        # Time entries for this block: [origin, farm0, farm1, ..., dest]
        btimes = all_times[b_idx] if all_times else None
        origin_dep_str = fmt_hhmm(btimes[0]["dep"]) if btimes else "-"
        for c_idx, (_, key) in enumerate(COLS):
            if key == "irma":       item = make_data_item(origin_key, bg=CLR_DEPOT)
            elif key == "location": item = make_data_item(origin_sub, bg=CLR_DEPOT)
            elif key == "dist":     item = make_data_item(od_str, bg=CLR_DEPOT,
                                                          align=Qt.AlignRight|Qt.AlignVCenter)
            elif key == "dep_time": item = make_data_item(origin_dep_str, bg=CLR_DEPOT,
                                                          align=Qt.AlignCenter)
            else:                   item = make_data_item("", bg=CLR_DEPOT)
            table.setItem(r, c_idx, item)
        r += 1

        # -- Stop rows in actual visit order ------------------------------
        # Walk _build_block_stops to interleave farms and mid-route dest dropoffs
        # in the order the truck actually visits them.
        is_last_block = (b_idx == len(blocks) - 1)
        origin_key_for_stops = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
        block_stops = _build_block_stops(block, origin_key_for_stops, is_last_block)

        # For preload blocks (no farms) use PRELOAD_VOL as the notional total
        PRELOAD_VOL_DISPLAY = 40000
        total_farm_vol = (PRELOAD_VOL_DISPLAY if block.get("preload") and not farms
                          else route_vol)
        already_del    = 0.0
        last_dest_dep_s = "-"   # for subtotal end-time
        farm_alt_idx   = 0       # for alternating row colors

        # Map dest dict id -> position in block.dests list (for d_i lookups)
        dests_by_id = {id(d): i for i, d in enumerate(dests)}
        # Track farm position in block.rows for f_i lookups
        farms_by_id = {id(f): i for i, f in enumerate(farms)}

        # Build list of (dest_d, partial_offload) by walking dests in stop order to
        # properly compute "remaining" and split tagging
        for stop in block_stops:
            si = stop["_si"]
            stype = stop["type"]
            # Skip origin and vedder - handled separately above and below
            if stype in ("origin", "vedder"):
                continue

            ft = btimes[si] if btimes and si < len(btimes) else None
            arr_t = ft["arr"] if ft else None
            dep_t = ft["dep"] if ft else None
            arr_s = fmt_hhmm(arr_t)
            dep_s = fmt_hhmm(dep_t)
            # Distance leg from THIS stop to the next stop
            leg = dists[si] if si < len(dists) else None
            ds  = f"{leg:.1f}" if leg is not None else "-"

            if stype == "farm":
                row_data = stop["farm"]
                f_i      = farms_by_id.get(id(row_data), 0)
                bg       = CLR_ALT if farm_alt_idx % 2 == 0 else CLR_WHITE
                farm_alt_idx += 1
                is_robot     = str(row_data.get("m1_start", "")).strip().upper() == "ROBOT"
                is_mennonite = str(row_data.get("irma", "")).strip() in MENNONITE_FARMS
                if is_mennonite:
                    bg = QColor("#e1bee7")   # light purple
                MENNONITE_TIP = "Mennonite farm — no pickup on Sunday."
                milking_conflict = arrives_during_milking(arr_t, row_data, suppress_no_milking=suppress_no_milking)
                for c_idx, (_, key) in enumerate(COLS):
                    if key == "dist":
                        item = make_data_item(ds, bg=bg, align=Qt.AlignRight|Qt.AlignVCenter,
                                             draggable=editable)
                    elif key == "prior_vol":
                        v = row_data.get(key)
                        item = make_data_item(f"{int(v):,}" if isinstance(v,(int,float)) else "",
                                             bg=bg, align=Qt.AlignRight|Qt.AlignVCenter,
                                             draggable=editable)
                    elif key == "arr_time":
                        item = make_data_item(arr_s,
                                             bg=CLR_RED_BG if milking_conflict else bg,
                                             fg=CLR_RED    if milking_conflict else None,
                                             align=Qt.AlignCenter, draggable=editable)
                    elif key == "wait_time":
                        wait = ft["wait"] if ft else None
                        if wait and wait > 0:
                            wm = int(round(wait))
                            wait_s = f"{wm}m" if wm < 60 else f"{wm//60}h{wm%60:02d}m"
                            item = make_data_item(wait_s, bg=QColor("#fff3e0"),
                                                 fg=QColor("#e65100"),
                                                 align=Qt.AlignCenter, draggable=editable)
                        else:
                            item = make_data_item("", bg=bg, align=Qt.AlignCenter, draggable=editable)
                    elif key == "dep_time":
                        item = make_data_item(dep_s, bg=bg, align=Qt.AlignCenter, draggable=editable)
                    elif key == "_mwo":
                        checked = bool(row_data.get("_mwo", False))
                        item = QTableWidgetItem()
                        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                        item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
                        item.setBackground(QColor("#bbdefb") if checked else bg)
                        if editable:
                            item.setToolTip(
                                "Check to give this farm a 2-hour pickup window\n"
                                "starting from its current arrival time in this schedule.\n"
                                "Reflects across both Original and Modified views.")
                        item.setData(Qt.UserRole, (b_idx, f_i))
                    else:
                        # For farm rows, show the farm name (from _extra_cells
                        # col R=18) in the Location column instead of the raw
                        # location string.  Processors keep their name as-is.
                        if key == "location":
                            farm_name = (row_data.get("_extra_cells") or {}).get(18, "")
                            display_val = farm_name if farm_name else row_data.get(key, "")
                        else:
                            display_val = row_data.get(key, "")
                        item = make_data_item(display_val, bg=bg, draggable=editable)
                    item.setData(Qt.UserRole, (b_idx, f_i))
                    if is_mennonite:
                        item.setToolTip(MENNONITE_TIP)
                    table.setItem(r, c_idx, item)
                if is_robot:
                    m1s_col = next(i for i, (_, k) in enumerate(COLS) if k == "m1_start")
                    robot_bg = QColor("#e1bee7") if is_mennonite else QColor("#e8f5e9")
                    robot_item = make_data_item(
                        "ROBOT", bg=robot_bg, align=Qt.AlignCenter)
                    rf = robot_item.font(); rf.setBold(True)
                    robot_item.setFont(rf)
                    robot_item.setForeground(QColor("#2e7d32"))
                    robot_item.setData(Qt.UserRole, (b_idx, f_i))
                    if is_mennonite:
                        robot_item.setToolTip(MENNONITE_TIP)
                    table.setItem(r, m1s_col, robot_item)
                    # Clear the 3 hidden cells so Qt doesn't show stale data
                    for span_key in ("m1_finish", "m2_start", "m2_finish"):
                        sc = next(i for i, (_, k) in enumerate(COLS) if k == span_key)
                        empty = QTableWidgetItem("")
                        empty.setData(Qt.UserRole, (b_idx, f_i))
                        table.setItem(r, sc, empty)
                    table.setSpan(r, m1s_col, 1, 4)

                r += 1

            elif stype == "dest":
                dest_d = stop["dest"]
                d_i    = dests_by_id.get(id(dest_d), 0)
                dest_arr_s   = arr_s
                dest_dep_s_d = dep_s
                last_dest_dep_s = dest_dep_s_d
                # partial vol for this dest
                vp = dest_d.get("vol_partial")
                remaining = max(0.0, total_farm_vol - already_del)
                offload = min(float(vp), remaining) if vp is not None else remaining
                already_del += offload
                vol_s = f"{int(offload):,}" if offload else ""
                # Detect if this is a mid-route partial dropoff (truck continues after)
                is_yard_dest = "yard for" in (dest_d.get("name","") or "").lower()
                # A dest is a "mid-route partial" if it has split_after set (not None)
                # AND there are more stops of type 'farm' after it in block_stops
                is_partial_split = (dest_d.get("split_after") is not None)

                # Check if arrival is outside the plant's receiving window
                dest_win_conflict = False
                if plant_windows and not is_yard_dest and ft and ft.get("arr"):
                    dk_chk = normalise_key(dest_d.get("key","") or "")
                    win_chk = plant_windows.get(dk_chk)
                    if win_chk:
                        dest_win_conflict = not time_in_window(ft["arr"], win_chk[0], win_chk[1])

                # Use a distinct background for partial mid-route dropoffs so they're visually obvious
                if is_partial_split:
                    dest_bg = QColor("#fff8e1")   # soft amber - "partial dropoff"
                elif dest_win_conflict:
                    dest_bg = CLR_DEST_WARN
                else:
                    dest_bg = CLR_DEST

                # Compose location string with PARTIAL tag where appropriate
                loc_str = dest_d.get("name","") or dest_d.get("key","")
                if is_partial_split:
                    loc_str = f"v PARTIAL - {loc_str}"

                for c_idx, (_, key) in enumerate(COLS):
                    if key == "irma":
                        item = make_data_item(dest_d.get("key",""), bg=dest_bg,
                                             align=Qt.AlignCenter, draggable=editable)
                    elif key == "location":
                        item = make_data_item(loc_str, bg=dest_bg, draggable=editable)
                    elif key == "prior_vol":
                        item = make_data_item(vol_s, bg=dest_bg,
                                              align=Qt.AlignRight|Qt.AlignVCenter, draggable=editable)
                    elif key == "dist":
                        item = make_data_item("0.0" if is_yard_dest else ds,
                                              bg=dest_bg, align=Qt.AlignRight|Qt.AlignVCenter)
                    elif key == "arr_time":
                        item = make_data_item(dest_arr_s, bg=dest_bg, align=Qt.AlignCenter)
                    elif key == "dep_time":
                        item = make_data_item(dest_dep_s_d, bg=dest_bg, align=Qt.AlignCenter)
                    else:
                        item = make_data_item("", bg=dest_bg)
                    # tag dest rows so drag resolver knows they're dest rows
                    item.setData(Qt.UserRole+1, ("dest", b_idx, d_i))
                    table.setItem(r, c_idx, item)
                r += 1

        # No-processor warning row if block has no dests
        if not dests:
            CLR_WARN = QColor("#fff3cd")
            for c_idx, (_, key) in enumerate(COLS):
                if c_idx == 0:
                    item = make_header_item("(!) No processor assigned",
                                            bg=CLR_WARN, fg=QColor("#856404"))
                else:
                    item = make_data_item("", bg=CLR_WARN)
                    item.setFlags(Qt.ItemIsEnabled)
                table.setItem(r, c_idx, item)
            r += 1

        # VEDDER return row - only on the last block
        if is_last_block:
            vedder_stops = [s for s in block_stops if s["type"] == "vedder"]
            vedder_idx   = vedder_stops[0]["_si"] if vedder_stops else (len(farms) + len(dests) + 1)
            vedder_ret_t = btimes[vedder_idx] if btimes and vedder_idx < len(btimes) else None
            vedder_arr_s = fmt_hhmm(vedder_ret_t["arr"]) if vedder_ret_t else "-"
            for c_idx, (_, key) in enumerate(COLS):
                if key == "irma":       item = make_data_item("VEDDER", bg=CLR_DEPOT)
                elif key == "location": item = make_data_item("Depot", bg=CLR_DEPOT)
                elif key == "dist":     item = make_data_item("-", bg=CLR_DEPOT,
                                                              align=Qt.AlignRight|Qt.AlignVCenter)
                elif key == "arr_time": item = make_data_item(vedder_arr_s, bg=CLR_DEPOT,
                                                              align=Qt.AlignCenter)
                else:                   item = make_data_item("", bg=CLR_DEPOT)
                table.setItem(r, c_idx, item)
            r += 1

        # Route subtotal
        vol_over = route_vol > VOL_LIMIT
        ds2 = f"{route_dist:.1f} km" if route_ok else f"~{route_dist:.1f} km*"
        end_s = last_dest_dep_s
        for c_idx, (_, key) in enumerate(COLS):
            if c_idx == 0:
                item = make_header_item("Subtotal",
                                        bg=CLR_SUBTOTAL, fg=QColor("#000000"))
            elif key == "dist":
                item = make_header_item(ds2, bg=CLR_SUBTOTAL, fg=QColor("#000000"))
            elif key == "prior_vol":
                item = make_header_item(f"{int(route_vol):,} L",
                                        bg=CLR_RED_BG if vol_over else CLR_SUBTOTAL,
                                        fg=CLR_RED    if vol_over else QColor("#000000"))
            elif key == "dep_time":
                item = make_header_item(f"End: {end_s}", bg=CLR_SUBTOTAL, fg=QColor("#000000"))
            else:
                item = make_data_item("", bg=CLR_SUBTOTAL)
                item.setFlags(Qt.ItemIsEnabled)
            table.setItem(r, c_idx, item)
        r += 1

        day_dist += route_dist; day_vol += route_vol
        if not route_ok: day_ok = False

    # Day total - no red colouring here
    ds3 = f"{day_dist:.1f} km" if day_ok else f"~{day_dist:.1f} km*"
    # Shift end = arrival back at VEDDER (last entry in last block's times)
    shift_end = fmt_hhmm(_end_cursor.time()) if _end_cursor is not None else "-"
    for c_idx, (_, key) in enumerate(COLS):
        if c_idx == 0:
            item = make_header_item("Total", bg=CLR_TOTAL, fg=QColor("#000000"))
        elif key == "dist":
            item = make_header_item(ds3, bg=CLR_TOTAL, fg=QColor("#000000"))
        elif key == "prior_vol":
            item = make_header_item(f"{int(day_vol):,} L", bg=CLR_TOTAL, fg=QColor("#000000"))
        elif key == "dep_time":
            item = make_header_item(f"Shift end: {shift_end}", bg=CLR_TOTAL, fg=QColor("#000000"))
        else:
            item = make_data_item("", bg=CLR_TOTAL)
            item.setFlags(Qt.ItemIsEnabled)
        table.setItem(r, c_idx, item)
    r += 1

    # Shift length row
    shift_len_str = "-"
    if start_time and _end_cursor is not None:
        from datetime import datetime, date
        base  = datetime.combine(date.today(), start_time)
        delta = _end_cursor - base
        total_mins = int(delta.total_seconds() / 60)
        hours, mins = divmod(total_mins, 60)
        shift_len_str = f"{hours}h {mins:02d}m"
    CLR_SHIFT = QColor("#e8eaf6")
    table.setSpan(r, 0, 1, len(COLS))
    lbl = f"  Estimated Shift Length:  {fmt_hhmm(start_time)} -> {shift_end}  =  {shift_len_str}"
    item = make_header_item(lbl, bg=CLR_SHIFT, fg=QColor("#1a237e"))
    table.setItem(r, 0, item)

    table.resizeRowsToContents()
    table.setRowHeight(r, 22)   # cap shift-length row to normal row height
    hh = table.horizontalHeader()
    _COL_W = {
        "irma":      64,
        "train":     24,   # "Tr"  - just Y/N/number
        "m1_start":  46,
        "m1_finish": 46,
        "m2_start":  46,
        "m2_finish": 46,
        "edpu":      20,   # "E"   - just N/Y
        "location":  None, # Stretch - fills all remaining width
        "prior_vol": 58,
        "dist":      58,
        "arr_time":  46,
        "wait_time": 38,
        "dep_time":  56,   # "Dep."
        "_mwo":      20,   # "M"   - checkbox
    }
    for c_idx, (_, key) in enumerate(COLS):
        w = _COL_W.get(key)
        if w is None:
            hh.setSectionResizeMode(c_idx, QHeaderView.Stretch)
        else:
            hh.setSectionResizeMode(c_idx, QHeaderView.Interactive)
            table.setColumnWidth(c_idx, w)
        tip = _COL_TIPS.get(key)
        if tip:
            item = table.horizontalHeaderItem(c_idx)
            if item:
                item.setToolTip(tip)

# -- File loader thread --------------------------------------------------------

def _try_eval_formula(formula_str):
    """
    Attempt to evaluate a simple Excel formula string to a float.
    Handles bare numbers, simple arithmetic (+,-,*,/,()), and
    SUM() / sum() with only numeric literals.
    Returns a float, or None if the formula is too complex
    (contains cell references or unsupported functions).
    """
    if not isinstance(formula_str, str):
        return None
    s = formula_str.strip()
    if s.startswith("="):
        s = s[1:].strip()
    # Strip a leading SUM() wrapper with no cell refs
    import re as _re
    sum_m = _re.match(r'^[Ss][Uu][Mm]\(([^)]+)\)$', s)
    if sum_m:
        s = sum_m.group(1).replace(",", "+")
    # Reject anything that still looks like a cell reference (letters followed by digits)
    if _re.search(r'[A-Za-z]', s):
        return None
    # Only allow digits, spaces, arithmetic ops, dots, parens
    if not _re.match(r'^[\d\s\+\-\*\/\.\(\)]+$', s):
        return None
    try:
        result = eval(s, {"__builtins__": {}}, {})  # no builtins - safe for pure arithmetic
        return float(result)
    except Exception:
        return None


class FileLoader(QThread):
    done   = pyqtSignal(str, dict)
    failed = pyqtSignal(str, str)
    # Non-fatal per-sheet parse warnings - accumulated and shown in a dialog
    # after load completes, and also written to the debug log.
    sheet_warning = pyqtSignal(str, str, str)
    # Informational/timing log messages - written to the debug log only,
    # never shown in the warning dialog.
    log = pyqtSignal(str)

    def __init__(self, fname, fpath):
        super().__init__()
        self.fname = fname; self.fpath = fpath

    def run(self):
        try:
            import time as _time
            _t0 = _time.time()
            self.log.emit(f"[{self.fname}] Opening workbook (data pass)...")
            # read_only=True streams rows lazily instead of building a full
            # in-memory style/merge cache for every sheet.  On a large
            # multi-sheet workbook (e.g. 79 sheets) this is the difference
            # between ~0.5s and ~38s - read_only=False was the root cause of
            # "loading stalls" on real route files.  parse_sheet only reads
            # cell values by coordinate, which works identically in read-only
            # mode (verified: parsed output is byte-identical aside from the
            # random _uid field).
            wb_data = openpyxl.load_workbook(self.fpath, read_only=True, data_only=True)
            self.log.emit(
                f"[{self.fname}] Data workbook opened in {_time.time()-_t0:.1f}s, "
                f"opening formula pass...")
            _t1 = _time.time()
            wb_form = openpyxl.load_workbook(self.fpath, read_only=True, data_only=False)
            self.log.emit(
                f"[{self.fname}] Formula workbook opened in {_time.time()-_t1:.1f}s")
            sheets = {}
            for n in wb_data.sheetnames:
                if n.strip().upper() in EXCLUDE_SHEETS:
                    continue
                try:
                    _ts = _time.time()
                    self.log.emit(f"[{self.fname} / {n}] Parsing...")
                    # Wrap the read_only worksheet in a fast cache before
                    # parsing.  parse_sheet does repeated random .cell(r, c)
                    # access, which in read_only mode re-parses the sheet's
                    # XML from scratch on every call - the cache does one
                    # sequential pass up front and serves the rest from a
                    # dict, which is what makes parsing genuinely fast rather
                    # than just the workbook *open* being fast.
                    ws_cached      = _FastCellCache(wb_data[n])
                    ws_form_cached = _FastCellCache(wb_form[n])
                    blocks, start_time, day_colour = parse_sheet(
                        ws_cached, ws_formula=ws_form_cached)
                    self.log.emit(
                        f"[{self.fname} / {n}] Done in {_time.time()-_ts:.1f}s - "
                        f"{len(blocks)} block(s), start={start_time}, colour={day_colour!r}")

                    # -- Post-parse diagnostics --------------------------------
                    # Each check emits a specific warning so the user knows
                    # *why* a sheet looks empty or wrong, rather than just
                    # seeing a blank tab.
                    warnings = []

                    if not blocks:
                        warnings.append(
                            "No blocks found - no 'IRMA#' header rows were "
                            "detected.  Check that the sheet follows the "
                            "expected layout (IRMA# in column A).")

                    if not start_time:
                        warnings.append(
                            f"No driver start time found in cell "
                            f"{openpyxl.utils.get_column_letter(C_DRIVER_START)}1"
                            f" (column {C_DRIVER_START}).  Times will not be "
                            f"calculated for this sheet.")

                    if not day_colour:
                        warnings.append(
                            f"No day colour found in cell "
                            f"{openpyxl.utils.get_column_letter(C_DAY_COLOUR)}1"
                            f" (column {C_DAY_COLOUR}).  Expected RED, BLUE, "
                            f"or GRASSFED.  The solver may skip this sheet.")

                    # Check for formula cells whose cached value is None -
                    # the most common cause of 'looks fine in Excel, missing
                    # data in the viewer'.  Report the first few occurrences.
                    formula_nones = []
                    for block in blocks:
                        for row in block.get("rows", []):
                            if row.get("prior_vol") is None:
                                formula_nones.append(row.get("irma", "?"))
                    if formula_nones:
                        sample = ", ".join(formula_nones[:5])
                        extra  = f" (and {len(formula_nones)-5} more)" \
                                 if len(formula_nones) > 5 else ""
                        warnings.append(
                            f"prior_vol is None for farm(s): {sample}{extra}.  "
                            f"These cells may contain uncalculated formulas - "
                            f"open the file in Excel, press Ctrl+Alt+F9 to "
                            f"force-recalculate, then save and reload.")

                    # Check for blocks with no destination
                    no_dest = [str(i+1) for i, b in enumerate(blocks)
                               if not b.get("dests") and not b.get("dest_key")
                               and not b.get("preload")]
                    if no_dest:
                        warnings.append(
                            f"Block(s) {', '.join(no_dest)} have no destination.  "
                            f"The 'Delivery Information:' section may be missing "
                            f"or in an unexpected column.")

                    for w in warnings:
                        logger.warning("[%s / %s] %s", self.fname, n, w)
                        self.sheet_warning.emit(self.fname, n, w)

                    if blocks or start_time:
                        _raw_ai1    = ws_cached.cell(1, C_SHEET_DATE).value
                        _sheet_date = _extract_row1_date(ws_cached)
                        if _raw_ai1 is not None and _sheet_date is None:
                            logger.debug("[%s / %s] AI1 value not parsed as date: %r",
                                         self.fname, n, _raw_ai1)
                        sheets[n] = {"blocks": blocks, "start_time": start_time,
                                     "day_colour": day_colour,
                                     "sheet_date": _sheet_date}
                    else:
                        # Emit a warning so it's visible even when no blocks
                        # or start_time were found (sheet would be silently
                        # dropped from the cache without this).
                        if not warnings:
                            msg = ("Sheet produced no usable data and no "
                                   "specific warnings - the layout may not "
                                   "match the expected format.")
                            logger.warning("[%s / %s] %s", self.fname, n, msg)
                            self.sheet_warning.emit(self.fname, n, msg)

                except Exception as sheet_err:
                    # Per-sheet failure: log it and emit a warning rather than
                    # aborting the whole file load.  The sheet is skipped but
                    # other sheets in the same workbook still load.
                    import traceback
                    detail = traceback.format_exc()
                    msg = (f"Failed to parse sheet '{n}': {sheet_err}\n{detail}")
                    logger.error("[%s / %s] %s", self.fname, n, msg)
                    self.sheet_warning.emit(self.fname, n,
                        f"Parse error - {sheet_err}  "
                        f"(see console / log for full traceback)")

            wb_data.close(); wb_form.close()
            self.done.emit(self.fname, sheets)
        except Exception as e:
            self.failed.emit(self.fname, str(e))

# -- Drag-aware route table ----------------------------------------------------

MIME_FARM      = "application/x-farm"
MIME_TRAY_FARM = "application/x-tray-farm"
MIME_DEST      = "application/x-dest"
MIME_TRAY_DEST = "application/x-tray-dest"
MIME_BLOCK     = "application/x-block"     # drag entire block by its banner row

class EditableRouteTable(QTableWidget):
    """Editable route table: drag farms OUT (to tray) or accept drags IN (from tray).
    Also supports internal row reordering via drag."""
    farm_removed  = pyqtSignal(int, int)        # b_idx, f_idx  - removed to tray
    farm_inserted = pyqtSignal(int, int, int)   # tray_idx, b_idx, insert_before_f_idx
    farm_reorder  = pyqtSignal(int, int, int, int)  # src_b, src_f, dst_b, dst_f
    dest_removed  = pyqtSignal(int, int)        # b_idx, d_idx  - dest removed to tray
    dest_reorder  = pyqtSignal(int, int, int, int)  # src_b, src_d, dst_b, dst_d
    dest_inserted = pyqtSignal(int, int, int)   # tray_idx, b_idx, insert_before_d_idx
    block_reorder = pyqtSignal(int, int)        # src_b_idx, insert_before_b_idx

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.setShowGrid(True)
        self.setDropIndicatorShown(False)   # we draw our own
        self._drop_indicator_row = -1       # visual row to draw indicator above

    def dragMoveEvent(self, e):
        m = e.mimeData()
        if (m.hasFormat(MIME_FARM) or m.hasFormat(MIME_TRAY_FARM) or
                m.hasFormat(MIME_DEST) or m.hasFormat(MIME_TRAY_DEST) or
                m.hasFormat(MIME_BLOCK)):
            row = self.rowAt(e.pos().y())
            if m.hasFormat(MIME_BLOCK):
                # Snap indicator to the nearest banner boundary
                target_b = self._resolve_block_drop(row)
                if target_b >= 0:
                    # Find the visual row of that banner
                    snap_row = row  # fallback
                    for ri in range(self.rowCount()):
                        it = self.item(ri, 0)
                        if it is not None and it.data(Qt.UserRole + 2) == target_b:
                            snap_row = ri
                            break
                    self._drop_indicator_row = snap_row
                else:
                    # Append at very end
                    self._drop_indicator_row = self.rowCount()
            else:
                self._drop_indicator_row = row
            self.viewport().update()
            e.acceptProposedAction()
        else:
            e.ignore()

    def dragLeaveEvent(self, e):
        self._drop_indicator_row = -1
        self.viewport().update()
        super().dragLeaveEvent(e)

    def paintEvent(self, e):
        super().paintEvent(e)
        if self._drop_indicator_row < 0: return
        painter = QPainter(self.viewport())
        pen = QPen(QColor("#000000"), 2)
        painter.setPen(pen)
        row = self._drop_indicator_row
        if row < self.rowCount():
            rect = self.visualRect(self.model().index(row, 0))
            y = rect.top()
        else:
            rect = self.visualRect(self.model().index(self.rowCount()-1, 0))
            y = rect.bottom()
        painter.drawLine(0, y, self.viewport().width(), y)
        # Draw small arrow indicators at both ends
        size = 6
        for x in [0, self.viewport().width() - size]:
            painter.setBrush(QColor("#000000"))
            painter.setPen(Qt.NoPen)
            from PyQt5.QtGui import QPolygon
            from PyQt5.QtCore import QPoint
            tri = QPolygon([QPoint(x, y-size//2), QPoint(x+size, y), QPoint(x, y+size//2)])
            painter.drawPolygon(tri)
        painter.end()

    def startDrag(self, actions):
        item = self.currentItem()
        if item is None: return

        # Block banner drag - UserRole+2 holds b_idx
        block_idx = item.data(Qt.UserRole + 2)
        if block_idx is not None:
            mime = QMimeData()
            mime.setData(MIME_BLOCK, QByteArray(str(block_idx).encode()))
            drag = QDrag(self)
            drag.setMimeData(mime)
            drag.exec_(Qt.MoveAction)
            return

        # Dest drag - UserRole+1 holds ("dest", b_idx, d_idx)
        dest_data = item.data(Qt.UserRole + 1)
        if dest_data is not None and dest_data[0] == "dest":
            _, b_idx, d_idx = dest_data
            mime = QMimeData()
            mime.setData(MIME_DEST, QByteArray(f"{b_idx},{d_idx}".encode()))
            drag = QDrag(self)
            drag.setMimeData(mime)
            drag.exec_(Qt.MoveAction)
            return

        # Farm drag - UserRole holds (b_idx, f_idx)
        farm_data = item.data(Qt.UserRole)
        if farm_data is None: return
        b_idx, f_idx = farm_data
        mime = QMimeData()
        mime.setData(MIME_FARM, QByteArray(f"{b_idx},{f_idx}".encode()))
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec_(Qt.MoveAction)

    def dragEnterEvent(self, e):
        m = e.mimeData()
        if (m.hasFormat(MIME_FARM) or m.hasFormat(MIME_TRAY_FARM) or
                m.hasFormat(MIME_DEST) or m.hasFormat(MIME_TRAY_DEST) or
                m.hasFormat(MIME_BLOCK)):
            e.acceptProposedAction()
        else:
            e.ignore()

    def dropEvent(self, e):
        self._drop_indicator_row = -1
        self.viewport().update()
        mime = e.mimeData()
        row_at_drop = self.rowAt(e.pos().y())

        if mime.hasFormat(MIME_BLOCK):
            src_b = int(bytes(mime.data(MIME_BLOCK)).decode())
            dst_b = self._resolve_block_drop(row_at_drop)
            if dst_b != src_b:
                self.block_reorder.emit(src_b, dst_b)
            e.acceptProposedAction()
            return

        if mime.hasFormat(MIME_DEST):
            raw = bytes(mime.data(MIME_DEST)).decode()
            src_b, src_d = map(int, raw.split(","))
            target_b, target_d = self._resolve_dest_drop(row_at_drop)
            self.dest_reorder.emit(src_b, src_d, target_b, target_d)
            e.acceptProposedAction()
            return

        if mime.hasFormat(MIME_TRAY_DEST):
            raw = bytes(mime.data(MIME_TRAY_DEST)).decode()
            tray_idx = int(raw)
            target_b, target_d = self._resolve_dest_drop(row_at_drop)
            self.dest_inserted.emit(tray_idx, target_b, target_d)
            e.acceptProposedAction()
            return

        # Resolve what block/farm-position the drop landed in
        target_b_idx, target_f_idx = self._resolve_drop_target(row_at_drop)

        if mime.hasFormat(MIME_FARM):
            raw = bytes(mime.data(MIME_FARM)).decode()
            src_b, src_f = map(int, raw.split(","))
            self.farm_reorder.emit(src_b, src_f, target_b_idx, target_f_idx)
            e.acceptProposedAction()

        elif mime.hasFormat(MIME_TRAY_FARM):
            raw = bytes(mime.data(MIME_TRAY_FARM)).decode()
            tray_idx = int(raw)
            self.farm_inserted.emit(tray_idx, target_b_idx, target_f_idx)
            e.acceptProposedAction()

    def _resolve_drop_target(self, row_at_drop):
        """Return (b_idx, insert_before_f_idx) for the given visual row.

        When a block has no farms every row in it is a header/origin/dest row
        and carries no Qt.UserRole farm data.  We detect the enclosing block by
        scanning backward for the nearest banner row (UserRole+2) so that drops
        on empty blocks are correctly attributed to that block at position 0.
        """
        if row_at_drop < 0:
            if self.rowCount() == 0:
                return (0, 0)
            # Dropped below all rows - find the last block
            for r in range(self.rowCount() - 1, -1, -1):
                it = self.item(r, 0)
                if it is not None:
                    bd = it.data(Qt.UserRole + 2)
                    if bd is not None:
                        return (bd, 0)
                    fd = it.data(Qt.UserRole)
                    if fd is not None:
                        return (fd[0], fd[1] + 1)
            return (0, 0)

        item = self.item(row_at_drop, 0)
        if item is None:
            return (0, 0)

        # Landed directly on a farm row
        data = item.data(Qt.UserRole)
        if data is not None:
            return (data[0], data[1])

        # Dropped onto a dest (processor) row - insert AFTER the last farm
        # in that block, i.e. immediately before the processor.
        dest_tag = item.data(Qt.UserRole + 1)
        if dest_tag is not None and isinstance(dest_tag, tuple) and dest_tag[0] == "dest":
            b_idx = dest_tag[1]
            # Scan backward for the last farm row in this block
            for rb in range(row_at_drop - 1, -1, -1):
                itb = self.item(rb, 0)
                if itb is None:
                    continue
                fdb = itb.data(Qt.UserRole)
                if fdb is not None and fdb[0] == b_idx:
                    return (b_idx, fdb[1] + 1)   # after last farm
                bdb = itb.data(Qt.UserRole + 2)
                if bdb is not None:
                    return (b_idx, 0)             # empty block, insert at 0
            return (b_idx, 0)

        # Scan forward for the next farm row in this block
        for r in range(row_at_drop, self.rowCount()):
            it = self.item(r, 0)
            if it is None:
                continue
            fd = it.data(Qt.UserRole)
            if fd is not None:
                return (fd[0], fd[1])
            # Hit the next block's banner - stop; drop is at start of that block
            bd = it.data(Qt.UserRole + 2)
            if bd is not None and r > row_at_drop:
                # Insert at position 0 of whichever block we're currently in.
                # Find that block by scanning back for nearest banner.
                for rb in range(row_at_drop, -1, -1):
                    itb = self.item(rb, 0)
                    if itb is not None:
                        bdb = itb.data(Qt.UserRole + 2)
                        if bdb is not None:
                            return (bdb, 0)
                return (bd, 0)

        # Nothing found forward - scan backward for containing block banner
        for r in range(row_at_drop, -1, -1):
            it = self.item(r, 0)
            if it is None:
                continue
            fd = it.data(Qt.UserRole)
            if fd is not None:
                return (fd[0], fd[1] + 1)
            bd = it.data(Qt.UserRole + 2)
            if bd is not None:
                return (bd, 0)

        return (0, 0)

    def _resolve_dest_drop(self, row_at_drop):
        """Return (b_idx, insert_before_d_idx) for a dest drop.
        Falls back to the enclosing block (via banner) when no dest rows exist."""
        if row_at_drop < 0:
            return (0, 0)
        item = self.item(row_at_drop, 0)
        if item is None: return (0, 0)
        dd = item.data(Qt.UserRole + 1)
        if dd is not None and dd[0] == "dest":
            return (dd[1], dd[2])
        # Scan forward for a dest row
        for ri in range(row_at_drop, self.rowCount()):
            it = self.item(ri, 0)
            if it:
                dd2 = it.data(Qt.UserRole + 1)
                if dd2 and dd2[0] == "dest":
                    return (dd2[1], dd2[2])
                # Hit next block banner - drop is in current block
                bd = it.data(Qt.UserRole + 2)
                if bd is not None and ri > row_at_drop:
                    break
        # Scan backward for nearest dest or banner
        for ri in range(row_at_drop, -1, -1):
            it = self.item(ri, 0)
            if it:
                dd2 = it.data(Qt.UserRole + 1)
                if dd2 and dd2[0] == "dest":
                    return (dd2[1], dd2[2] + 1)
                # Found the banner for this block - append as first dest
                bd = it.data(Qt.UserRole + 2)
                if bd is not None:
                    return (bd, 0)
        return (0, 0)

    def _resolve_block_drop(self, row_at_drop):
        """Return the b_idx to insert the dragged block *before*.
        Scans for the nearest banner row (UserRole+2 set) to use as the target."""
        if row_at_drop < 0:
            return -1   # caller interprets as "append at end"
        # Scan forward for a banner row
        for ri in range(row_at_drop, self.rowCount()):
            it = self.item(ri, 0)
            if it is not None:
                bd = it.data(Qt.UserRole + 2)
                if bd is not None:
                    return bd
        # Nothing found forward - append after last block
        return -1

# -- Farm tray (removed farms, table-based, drag back) ------------------------

class FarmTray(QTableWidget):
    """Holds removed farms. Farms can be dragged back to the editable route table."""
    farm_incoming = pyqtSignal(int, int)   # b_idx, f_idx - farm dropped onto tray from route
    dest_incoming = pyqtSignal(int, int)   # b_idx, d_idx - dest dropped onto tray from route

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self.setColumnCount(len(TRAY_COLS))
        self.setHorizontalHeaderLabels([c[0] for c in TRAY_COLS])
        self.setRowCount(0)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.setShowGrid(True)
        self.setAlternatingRowColors(True)

    def keyPressEvent(self, event):
        from PyQt5.QtCore import Qt as _Qt
        if event.key() in (_Qt.Key_Delete, _Qt.Key_Backspace):
            # Bubble up to MainWindow via parent chain
            parent = self.parent()
            while parent:
                if hasattr(parent, '_on_tray_delete'):
                    parent._on_tray_delete()
                    return
                parent = parent.parent()
        super().keyPressEvent(event)

    def startDrag(self, actions):
        row = self.currentRow()
        if row < 0: return
        # Check if this tray row is a dest or a farm
        item0 = self.item(row, 0)
        is_dest = item0 and item0.data(Qt.UserRole) == "dest"
        mime = QMimeData()
        mime_type = MIME_TRAY_DEST if is_dest else MIME_TRAY_FARM
        mime.setData(mime_type, QByteArray(str(row).encode()))
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec_(Qt.MoveAction)

    def dragEnterEvent(self, e):
        m = e.mimeData()
        if m.hasFormat(MIME_FARM) or m.hasFormat(MIME_DEST): e.acceptProposedAction()
        else: e.ignore()

    def dragMoveEvent(self, e):
        m = e.mimeData()
        if m.hasFormat(MIME_FARM) or m.hasFormat(MIME_DEST): e.acceptProposedAction()
        else: e.ignore()

    def dropEvent(self, e):
        m = e.mimeData()
        if m.hasFormat(MIME_FARM):
            raw = bytes(m.data(MIME_FARM)).decode()
            b_idx, f_idx = map(int, raw.split(","))
            self.farm_incoming.emit(b_idx, f_idx)
            e.acceptProposedAction()
        elif m.hasFormat(MIME_DEST):
            raw = bytes(m.data(MIME_DEST)).decode()
            b_idx, d_idx = map(int, raw.split(","))
            self.dest_incoming.emit(b_idx, d_idx)
            e.acceptProposedAction()

    def add_farm(self, farm_dict, route_label, farm_colour="", current_colour=""):
        """Add a farm row. farm_colour is the day type it was removed from.
        current_colour is the active sheet's day type - if different, row is bold."""
        r = self.rowCount(); self.insertRow(r)
        mismatch = bool(farm_colour and current_colour and
                        farm_colour.upper().strip() != current_colour.upper().strip())
        bold_font = QFont(); bold_font.setBold(True)
        normal_font = QFont()
        row_font = bold_font if mismatch else normal_font

        for c_idx, (_, key) in enumerate(TRAY_COLS):
            if key == "_from_route":
                val = route_label
                bg  = CLR_REMOVED
                fg  = None
            elif key == "_day_colour":
                bg2, fg2, label = day_colour_style(farm_colour)
                val = label or farm_colour or "-"
                bg  = bg2 if bg2 else CLR_REMOVED
                fg  = fg2
            elif key == "prior_vol":
                v = farm_dict.get(key)
                val = f"{int(v):,}" if isinstance(v, (int,float)) else ""
                bg  = CLR_REMOVED; fg = None
            elif key == "dist":
                val = ""; bg = CLR_REMOVED; fg = None
            elif key == "_mwo":
                val = "OK" if farm_dict.get("_mwo") else ""
                bg = CLR_REMOVED; fg = None
            else:
                if key == "location":
                    farm_name = (farm_dict.get("_extra_cells") or {}).get(18, "")
                    val = farm_name if farm_name else farm_dict.get(key, "")
                else:
                    val = farm_dict.get(key, "")
                bg  = CLR_REMOVED; fg = None

            item = QTableWidgetItem(str(val) if val is not None else "")
            item.setBackground(bg)
            if fg: item.setForeground(fg)
            item.setFont(row_font)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsDragEnabled)
            align = Qt.AlignRight|Qt.AlignVCenter if key == "prior_vol" else Qt.AlignCenter
            item.setTextAlignment(align)
            self.setItem(r, c_idx, item)
        self.setColumnWidth(0, 68)

    def add_dest(self, dest_dict, route_label, farm_colour="", current_colour=""):
        """Add a destination row to the tray. Marked with UserRole='dest'."""
        r = self.rowCount(); self.insertRow(r)
        mismatch = bool(farm_colour and current_colour and
                        farm_colour.upper().strip() != current_colour.upper().strip())
        bold_font = QFont(); bold_font.setBold(True)
        normal_font = QFont()
        row_font = bold_font if mismatch else normal_font
        CLR_DEST_TRAY = QColor("#e8f5e9")  # greenish to distinguish from farm rows
        for c_idx, (_, key) in enumerate(TRAY_COLS):
            if key == "_from_route":
                val = route_label; bg = CLR_DEST_TRAY; fg = None
            elif key == "_day_colour":
                bg2, fg2, label = day_colour_style(farm_colour)
                val = label or farm_colour or "-"
                bg = bg2 if bg2 else CLR_DEST_TRAY; fg = fg2
            elif key == "irma":
                val = dest_dict.get("key",""); bg = CLR_DEST_TRAY; fg = None
            elif key == "location":
                val = dest_dict.get("name","") or dest_dict.get("key","")
                bg = CLR_DEST_TRAY; fg = None
            elif key == "prior_vol":
                vp = dest_dict.get("vol_partial")
                val = f"{int(vp):,}" if isinstance(vp,(int,float)) else "rest"
                bg = CLR_DEST_TRAY; fg = None
            else:
                val = ""; bg = CLR_DEST_TRAY; fg = None
            item = QTableWidgetItem(str(val) if val is not None else "")
            item.setBackground(bg)
            if fg: item.setForeground(fg)
            item.setFont(row_font)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsDragEnabled)
            item.setTextAlignment(Qt.AlignCenter)
            # Tag first cell so we know it's a dest row
            if c_idx == 0: item.setData(Qt.UserRole, "dest")
            self.setItem(r, c_idx, item)

    def refresh_bold_state(self, current_colour):
        """Bold rows whose farm type differs from current_colour."""
        bold_font   = QFont(); bold_font.setBold(True)
        normal_font = QFont()
        type_col = next((i for i,(_, k) in enumerate(TRAY_COLS) if k == "_day_colour"), None)
        for r in range(self.rowCount()):
            farm_colour = ""
            if type_col is not None:
                it = self.item(r, type_col)
                if it: farm_colour = it.text()
            mismatch = bool(farm_colour and current_colour and
                            farm_colour.upper().strip() != current_colour.upper().strip())
            font = bold_font if mismatch else normal_font
            for c in range(self.columnCount()):
                it = self.item(r, c)
                if it: it.setFont(font)

# -- Synchronised scroll helper ------------------------------------------------

def sync_scroll(src_bar, dst_bar, value):
    dst_bar.setValue(value)

# ══════════════════════════════════════════════════════════════════════════════
# ALNS Solver
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_colour_bucket(day_colour):
    dc = day_colour.upper().strip()
    if "RED"  in dc: return "RED"
    if "BLUE" in dc: return "BLUE"
    return "OTHER"


def _route_km_simple(block, dm, origin="VEDDER"):
    """Quick distance estimate: origin -> farms -> all dests (no VEDDER return).
    Pure km only - used for 2-opt / or-opt intra-block resequencing where
    milking windows don't change between candidate orderings of the same farms."""
    farms     = [r["irma"] for r in block["rows"]]
    dest_keys = _block_dest_keys(block)
    stops     = [origin] + farms + dest_keys
    total     = 0.0
    for i in range(len(stops) - 1):
        d = lookup(dm, stops[i], stops[i + 1])
        if d is not None:
            total += d
    return total


def _route_cost_with_milking(block, dm, origin="VEDDER",
                              milking_weight=1.0, suppress_no_milking=True,
                              shift_start_mins=300):
    """Distance cost + milking-wait penalty for a block.

    Used by _best_insert_cost so the greedy repair heuristic avoids placing
    farms into slots that cause long waits, not just long drives.

    Milking wait estimation (fast, no full calc_times):
      Tracks an absolute datetime cursor starting at shift_start_mins past
      midnight on 'today'.  Each leg advances the cursor by drive time.
      If arrival falls inside a milking window (w1, w2, or w3), the wait
      until window-end is added as a penalty of wait_mins * milking_weight
      (in km-equivalent units, since 1 km-eq ~ 1.2 min at 50 km/h - close
      enough for ranking purposes).  The cursor advances by the wait so
      downstream farms see a realistic arrival time, correctly handling
      midnight-crossing routes without modular wrap errors.

    shift_start_mins: minutes past midnight for departure from the block origin.
      When no start time is available the sheet is excluded from insertion
      entirely (see _best_insert_cost), so this value always comes from a
      real entry.start_time.
    """
    if milking_weight <= 0.0:
        return _route_km_simple(block, dm, origin=origin)

    from datetime import timedelta as _td, datetime as _dt, date as _d

    farms     = block["rows"]
    dest_keys = _block_dest_keys(block)
    stops     = [origin] + [r["irma"] for r in farms] + dest_keys

    total_km = 0.0
    wait_pen = 0.0
    # Absolute datetime cursor - no modular wrap, handles overnight routes correctly
    cursor   = _dt.combine(_d.today(), _dt.min.time()) + _td(minutes=shift_start_mins)

    for i in range(len(stops) - 1):
        leg = lookup(dm, stops[i], stops[i + 1])
        if leg is not None:
            total_km += leg
            cursor   += _td(minutes=(leg / DRIVE_SPEED_KMH) * 60.0)

        # Check milking window for farm stops only (not origin or dest)
        if i < len(farms):
            farm = farms[i]
            irma = farm.get("irma", "")
            if suppress_no_milking and irma in NO_MILKING_WINDOW_FARMS:
                continue

            arr_t  = cursor.time()
            wait_m = 0.0

            # w1 / w2 (extended by regulatory pre/post buffers)
            for s_key, f_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
                ext_s, ext_f = _extended_milking_window(farm.get(s_key,""), farm.get(f_key,""))
                if ext_s is not None and time_in_window(arr_t, ext_s, ext_f):
                    end_w = _dt.combine(cursor.date(), ext_f)
                    if end_w <= cursor:
                        end_w += _td(days=1)
                    wait_m = (end_w - cursor).total_seconds() / 60.0
                    break
            # w3 (extended)
            if wait_m == 0.0:
                w3data = THREE_WINDOW_FARMS.get(irma)
                if w3data:
                    w3s, w3f = w3data.get("w3", [None, None])
                    ext_s3, ext_f3 = _extended_milking_window(w3s, w3f)
                    if ext_s3 is not None and time_in_window(arr_t, ext_s3, ext_f3):
                        end_w3 = _dt.combine(cursor.date(), ext_f3)
                        if end_w3 <= cursor:
                            end_w3 += _td(days=1)
                        wait_m = (end_w3 - cursor).total_seconds() / 60.0

            # Penalty: wait_mins * milking_weight (same units as km at ~1.2 min/km)
            wait_pen += wait_m * milking_weight
            # Advance cursor by the wait so downstream farms see a later arrival
            cursor   += _td(minutes=wait_m)

    return total_km + wait_pen


def _group_dest_catalogue(sheets):
    """Return {dest_key: dest_name} for every processor in the colour group."""
    cat = {}
    for _sname, entry in sheets:
        for block in entry.get("blocks", []):
            for d in (block.get("dests") or []):
                dk = d.get("key"); dn = d.get("name","")
                if dk: cat[dk] = dn or dk
            # legacy fallback
            dk = block.get("dest_key")
            if dk and dk not in cat:
                cat[dk] = block.get("dest_name","") or dk
    return cat


def _sheet_cost(blocks, dm, start_time, cfg, dm_dur=None):
    """
    Scalar cost for one truck's day (one sheet's worth of blocks).

    dm_dur: optional duration-matrix lookup (minutes between nodes).  When
    supplied, calc_times uses real recorded travel durations for legs that
    have data, falling back to the flat-speed distance estimate only for
    legs missing duration data.  Without it (dm_dur=None, the default), every
    leg uses the flat-speed estimate - this is what every caller did before
    duration data was wired into the solver, so omitting dm_dur preserves
    old behaviour exactly.

    cfg keys used:
      orig_dest_vols     – {dest_key: original_litres}  (group-wide)
      vol_tol            – fractional tolerance  (0.15 -> +/-15 %)
      vol_penalty        – penalty per litre outside tolerance
      milking_weight     – multiplier on milking-wait km-equivalent
      max_shift_h        – maximum shift hours before penalty
      shift_penalty      – penalty per hour over max_shift_h
      min_shift_h        – minimum shift hours before penalty
      shift_under_penalty – penalty per hour under min_shift_h
    """
    # -- distance --------------------------------------------------------------
    # NOTE: total_km is the literal distance driven - a separate cost axis from
    # travel TIME (fuel/wear vs schedule).  It deliberately stays distance-based
    # even when dm_dur is supplied; only the TIME-derived components below
    # (shift_hours, milking wait, plant-window arrival) use real durations.
    total_km  = 0.0
    all_dists = calc_distances(blocks, dm)
    for dists in all_dists:
        for d in dists[:-1]:
            if d is not None:
                total_km += d

    # -- shift time & milking-window waits -------------------------------------
    shift_hours   = 0.0
    milking_mins  = 0.0

    if start_time:
        _suppress = cfg.get("suppress_no_milking", True)
        _ct2 = calc_times(blocks, dm, start_time, dm_dur=dm_dur,
                          suppress_no_milking=_suppress,
                          precomputed_dists=all_dists)
        all_times  = _ct2[0] if _ct2 is not None else None
        _end_cur2  = _ct2[1] if _ct2 is not None else None
        if _end_cur2 is not None:
            base = datetime.combine(date.today(), start_time)
            shift_hours = (_end_cur2 - base).total_seconds() / 3600.0
        if all_times:

            for b_idx, block in enumerate(blocks):
                btimes = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes:
                    continue
                # Build the stop sequence once per block instead of once per
                # farm (_farm_stop_index rebuilds it internally) - identical
                # indices, ~len(rows)x fewer _build_block_stops calls.
                _is_last_b = (b_idx == len(blocks) - 1)
                _origin_b  = "VEDDER" if b_idx == 0 else (
                    _block_last_dest_key(blocks[b_idx - 1]) or "VEDDER")
                _stops_b   = _build_block_stops(block, _origin_b, _is_last_b)
                _farm_sis  = [s["_si"] for s in _stops_b if s["type"] == "farm"]
                for f_i, farm in enumerate(block["rows"]):
                    f_stop = _farm_sis[f_i] if f_i < len(_farm_sis) else f_i + 1
                    ft = btimes[f_stop] if f_stop < len(btimes) else None
                    if ft is None or ft["arr"] is None:
                        continue
                    arr = ft["arr"]
                    arr_dt = datetime.combine(date.today(), arr)

                    # Skip milking penalty for suppressed farms
                    if _suppress and farm.get("irma","") in NO_MILKING_WINDOW_FARMS:
                        continue

                    # MWO: skip milking penalty entirely for flagged farms
                    if farm.get("_mwo"):
                        continue

                    # Extended window: raw [m_start, m_finish] ± regulatory buffers
                    for s_key, f_key in [("m1_start", "m1_finish"),
                                         ("m2_start", "m2_finish")]:
                        ext_s, ext_f = _extended_milking_window(
                            farm.get(s_key, ""), farm.get(f_key, ""))
                        if ext_s is not None and time_in_window(arr, ext_s, ext_f):
                            end_w = datetime.combine(date.today(), ext_f)
                            if end_w < arr_dt:
                                end_w += timedelta(days=1)
                            milking_mins += (end_w - arr_dt).total_seconds() / 60.0
                            break
                    # w3 penalty (extended)
                    w3data = THREE_WINDOW_FARMS.get(farm.get("irma",""))
                    if w3data:
                        w3pair = w3data.get("w3", [None, None])
                        w3s, w3f = w3pair[0], w3pair[1]
                        ext_s3, ext_f3 = _extended_milking_window(w3s, w3f)
                        if ext_s3 is not None and time_in_window(arr, ext_s3, ext_f3):
                            end_w3 = datetime.combine(date.today(), ext_f3)
                            if end_w3 < arr_dt:
                                end_w3 += timedelta(days=1)
                            milking_mins += (end_w3 - arr_dt).total_seconds() / 60.0

    # -- plant volume penalty (group-wide vols supplied via cfg) ---------------
    # NOTE: vol penalty is computed at group level in _group_cost to avoid
    # double-counting; here we just return the non-volume portion so that
    # _best_insert_pos can use a fast per-sheet cost without needing group state.
    # When called from _group_cost, vol_pen is added there instead.
    vol_pen = 0.0
    if cfg.get("_include_vol_pen", False):
        vol_tol          = cfg.get("vol_tol", 0.15)
        vol_penalty_rate = cfg.get("vol_penalty", 1.0)
        orig_dest_vols   = cfg.get("orig_dest_vols", {})
        dest_vols        = {}
        for block in blocks:
            for dk, off in _block_dest_offloads(block).items():
                dest_vols[dk] = dest_vols.get(dk, 0.0) + off
        for dk, orig_vol in orig_dest_vols.items():
            cur_vol = dest_vols.get(dk, 0.0)
            lo = orig_vol * (1.0 - vol_tol)
            hi = orig_vol * (1.0 + vol_tol)
            if cur_vol < lo:
                vol_pen += (lo - cur_vol) * vol_penalty_rate
            elif cur_vol > hi:
                vol_pen += (cur_vol - hi) * vol_penalty_rate

    # -- truck capacity penalty ------------------------------------------------
    # For multi-dest routes with split_after, the truck offloads mid-route so
    # the total farm volume overstates the actual peak load.  We walk the stop
    # sequence and track the running load: +farm vol at farm stops, -offload vol
    # at dest stops.  The peak load is what actually matters for capacity.
    hard_cap      = cfg.get("hard_vol_cap", HARD_CAP)
    cap_pen_rate  = cfg.get("cap_penalty", 2.0)
    cap_pen       = 0.0
    for b_idx, block in enumerate(blocks):
        if _is_preload_block(block):
            continue   # preload blocks start empty - no cap issue
        dests = block.get("dests") or []
        if not dests:
            dk = block.get("dest_key","")
            dests = [{"key": dk, "vol_partial": None}] if dk else []
        farms = block.get("rows", [])
        total_farm_vol = sum((r.get("prior_vol") or 0) for r in farms
                             if isinstance(r.get("prior_vol"), (int, float)))

        # If any dest has a split_after, compute peak load along the sequence
        has_split = _block_has_split(block)
        if has_split:
            is_last  = (b_idx == len(blocks) - 1)
            origin   = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
            stops    = _build_block_stops(block, origin, is_last)
            running  = 0.0
            peak     = 0.0
            already_delivered = 0.0
            for stop in stops:
                if stop["type"] == "farm":
                    vol = stop["farm"].get("prior_vol") or 0
                    running += vol if isinstance(vol, (int, float)) else 0
                    peak = max(peak, running)
                elif stop["type"] == "dest":
                    offload = _dest_vol_partial(stop["dest"], total_farm_vol, already_delivered)
                    already_delivered += offload
                    running = max(0.0, running - offload)
            if peak > hard_cap:
                cap_pen += (peak - hard_cap) * cap_pen_rate
        else:
            # No mid-route dropoff - total farm vol is the peak load
            if total_farm_vol > hard_cap:
                cap_pen += (total_farm_vol - hard_cap) * cap_pen_rate

    # -- shift overage penalty -------------------------------------------------
    max_shift      = cfg.get("max_shift_h", 14.0)
    shift_pen_rate = cfg.get("shift_penalty", 200.0)
    shift_pen      = max(0.0, shift_hours - max_shift) * shift_pen_rate

    # -- shift shortfall penalty -------------------------------------------------
    # Mirrors the overage penalty above but in the other direction: routes
    # finishing meaningfully short of a minimum shift length are discouraged
    # too.  Guarded the same way shift_hours_cost is below - shift_hours
    # defaults to 0.0 when there's no start_time (no timing data at all for
    # this sheet), which would otherwise look like a zero-length shift and
    # spuriously trigger the full shortfall penalty for every untimed sheet.
    min_shift            = cfg.get("min_shift_h", 8.0)
    shift_under_pen_rate = cfg.get("shift_under_penalty", 30.0)
    shift_under_pen      = (max(0.0, min_shift - shift_hours) * shift_under_pen_rate
                            if shift_hours > 0 else 0.0)

    # -- total shift hours cost ------------------------------------------------
    # A small per-hour cost on the full shift duration (not just the overage).
    # This gives the solver a continuous gradient toward shorter days - without
    # it, any route that stays under max_shift_h looks equally good regardless
    # of whether it finishes in 10 hours or 13.5.
    # Default: 5.0 km-equivalent per hour (small enough not to override distance
    # or milking objectives, large enough to break ties in favour of earlier ends).
    shift_hours_rate = cfg.get("shift_hours_weight", 0.0)
    shift_hours_cost = shift_hours * shift_hours_rate if shift_hours > 0 else 0.0

    # -- milking wait penalty --------------------------------------------------
    # milking_mins * milking_weight: 1 min wait = 1 km-equivalent at weight=1.
    milking_equiv = milking_mins * cfg.get("milking_weight", 1.0)

    # -- plant receiving-window penalty ----------------------------------------
    # Both components expressed in km-equivalent per hour - same scale as routing.
    # 1. OUTSIDE penalty: arrival before open or after close.
    #    Rate: plant_win_penalty (default 200 km/h).
    # 2. MARGIN penalty: arrival inside the window but within the last
    #    plant_win_margin_mins minutes before close - gradient toward earlier arrivals.
    #    Rate: plant_win_margin_rate (default = plant_win_penalty * 0.5 km/h).
    plant_windows       = cfg.get("plant_windows", {})
    plant_win_rate      = cfg.get("plant_win_penalty", 200.0)          # km per hour outside
    plant_margin_mins   = cfg.get("plant_win_margin_mins", 30.0)
    plant_margin_rate   = cfg.get("plant_win_margin_rate",
                                  plant_win_rate * 0.5)                # km per hour inside margin
    # Avoid-windows: a flat penalty for arriving at a dest during a window
    # where the dock is wanted by another truck/division, even though the
    # plant itself is open.  Independent of the open/close check above - a
    # dest can have no plant_windows entry at all and still carry an
    # avoid-window, or vice versa.
    avoid_windows       = cfg.get("avoid_windows", AVOID_WINDOWS)
    avoid_win_rate       = cfg.get("avoid_window_penalty", 0.0)
    plant_win_cost = 0.0
    avoid_win_cost = 0.0
    if (plant_windows or avoid_windows) and start_time and all_times:
        for b_idx3, block3 in enumerate(blocks):
            btimes3 = all_times[b_idx3] if b_idx3 < len(all_times) else None
            if not btimes3:
                continue
            dests3 = block3.get("dests") or []
            if not dests3:
                dk3 = block3.get("dest_key", "")
                dests3 = [{"key": dk3}] if dk3 else []
            for d_i3, dest_d3 in enumerate(dests3):
                # Yard-for destinations are overnight parking - no receiving window
                if "yard for" in (dest_d3.get("name","") or "").lower():
                    continue
                dk3 = normalise_key(dest_d3.get("key", "") or "")
                window3     = plant_windows.get(dk3)
                avoid_list3 = avoid_windows.get(dk3)
                if window3 is None and not avoid_list3:
                    continue
                if block3.get("preload"):
                    t_idx3 = 1
                else:
                    t_idx3 = _dest_stop_index(block3, d_i3, b_idx3, blocks)
                ft3 = btimes3[t_idx3] if t_idx3 < len(btimes3) else None
                if ft3 is None or ft3.get("arr") is None:
                    continue
                arr3     = ft3["arr"]
                arr_dt3  = datetime.combine(date.today(), arr3)

                if window3 is not None:
                    open_str3, close_str3 = window3
                    close_t3 = parse_hhmm(close_str3)
                    open_t3  = parse_hhmm(open_str3)

                    if not time_in_window(arr3, open_str3, close_str3):
                        # Outside window - penalise by hours until next open
                        if open_t3 is not None:
                            open_dt3 = datetime.combine(date.today(), open_t3)
                            if open_t3 > arr3:
                                wait_h = (open_dt3 - arr_dt3).total_seconds() / 3600.0
                            else:
                                open_dt3 += timedelta(days=1)
                                wait_h = (open_dt3 - arr_dt3).total_seconds() / 3600.0
                            plant_win_cost += wait_h * plant_win_rate
                        else:
                            plant_win_cost += 1.0 * plant_win_rate   # flat 1-hour penalty
                    elif plant_margin_mins > 0 and close_t3 is not None and plant_margin_rate > 0:
                        close_dt3 = datetime.combine(date.today(), close_t3)
                        if open_t3 is not None and close_t3 < open_t3:
                            close_dt3 += timedelta(days=1)
                        mins_to_close = (close_dt3 - arr_dt3).total_seconds() / 60.0
                        if mins_to_close < 0:
                            close_dt3 += timedelta(days=1)
                            mins_to_close = (close_dt3 - arr_dt3).total_seconds() / 60.0
                        if mins_to_close < plant_margin_mins:
                            depth = (plant_margin_mins - mins_to_close) / plant_margin_mins
                            # depth * margin expressed as hours * rate
                            plant_win_cost += depth * plant_margin_rate * (plant_margin_mins / 60.0)

                if avoid_list3 and avoid_win_rate > 0:
                    for (av_start3, av_end3) in avoid_list3:
                        if time_in_window(arr3, av_start3, av_end3):
                            avoid_win_cost += avoid_win_rate
                            break   # one flat hit per stop even if windows overlap

    return (total_km + milking_equiv + vol_pen + shift_pen + shift_under_pen
            + shift_hours_cost + cap_pen + plant_win_cost + avoid_win_cost)


def _sheet_cost_breakdown(blocks, dm, start_time, cfg, dm_dur=None):
    """Same as _sheet_cost but returns a dict of cost components instead of a scalar.
    Used by the Full Cost Report so it always uses exactly the same logic as the solver."""
    # Re-use _sheet_cost internals by running both and computing the breakdown.
    # We compute each component independently using the same logic as _sheet_cost.

    # -- distance -------------------------------------------------------------
    total_km = 0.0
    for dists in calc_distances(blocks, dm):
        for d in dists[:-1]:
            if d is not None:
                total_km += d

    # -- shift time & milking -------------------------------------------------
    shift_hours  = 0.0
    milking_mins = 0.0
    all_times    = None
    _suppress    = cfg.get("suppress_no_milking", True)
    if start_time:
        _ct = calc_times(blocks, dm, start_time, dm_dur=dm_dur,
                         suppress_no_milking=_suppress)
        if _ct is not None:
            all_times = _ct[0]
            base = datetime.combine(date.today(), start_time)
            shift_hours = (_ct[1] - base).total_seconds() / 3600.0
        if all_times:
            for b_idx, block in enumerate(blocks):
                btimes = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes: continue
                for f_i, farm in enumerate(block["rows"]):
                    f_stop = _farm_stop_index(block, f_i, b_idx, blocks)
                    ft = btimes[f_stop] if f_stop < len(btimes) else None
                    if ft is None or ft.get("arr") is None: continue
                    arr = ft["arr"]
                    arr_dt = datetime.combine(date.today(), arr)
                    if _suppress and farm.get("irma","") in NO_MILKING_WINDOW_FARMS:
                        continue
                    if farm.get("_mwo"):
                        continue
                    # Extended window: raw [m_start, m_finish] ± regulatory buffers
                    for s_key, f_key in [("m1_start","m1_finish"),("m2_start","m2_finish")]:
                        ext_s, ext_f = _extended_milking_window(
                            farm.get(s_key,""), farm.get(f_key,""))
                        if ext_s is not None and time_in_window(arr, ext_s, ext_f):
                            end_w = datetime.combine(date.today(), ext_f)
                            if end_w < arr_dt: end_w += timedelta(days=1)
                            milking_mins += (end_w - arr_dt).total_seconds() / 60.0
                            break
                    # w3 (extended)
                    w3data = THREE_WINDOW_FARMS.get(farm.get("irma",""))
                    if w3data:
                        w3s, w3f = w3data.get("w3",[None,None])
                        ext_s3, ext_f3 = _extended_milking_window(w3s, w3f)
                        if ext_s3 is not None and time_in_window(arr, ext_s3, ext_f3):
                            end_w3 = datetime.combine(date.today(), ext_f3)
                            if end_w3 < arr_dt: end_w3 += timedelta(days=1)
                            milking_mins += (end_w3 - arr_dt).total_seconds() / 60.0

    # -- cap -------------------------------------------------------------------
    hard_cap     = cfg.get("hard_vol_cap", HARD_CAP)
    cap_pen_rate = cfg.get("cap_penalty", 2.0)
    cap_pen      = 0.0
    for b_idx, block in enumerate(blocks):
        if _is_preload_block(block): continue
        dests = block.get("dests") or []
        if not dests:
            dk = block.get("dest_key","")
            dests = [{"key": dk, "vol_partial": None}] if dk else []
        farms = block.get("rows", [])
        total_fv = sum((r.get("prior_vol") or 0) for r in farms
                       if isinstance(r.get("prior_vol"),(int,float)))
        if any(d.get("split_after") is not None for d in dests):
            is_last = (b_idx == len(blocks)-1)
            origin  = "VEDDER" if b_idx == 0 else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
            stops   = _build_block_stops(block, origin, is_last)
            running = peak = adel = 0.0
            for stop in stops:
                if stop["type"] == "farm":
                    v = stop["farm"].get("prior_vol") or 0
                    running += v if isinstance(v,(int,float)) else 0
                    peak = max(peak, running)
                elif stop["type"] == "dest":
                    off = _dest_vol_partial(stop["dest"], total_fv, adel)
                    adel += off; running = max(0.0, running - off)
            if peak > hard_cap: cap_pen += (peak - hard_cap) * cap_pen_rate
        else:
            if total_fv > hard_cap: cap_pen += (total_fv - hard_cap) * cap_pen_rate

    # -- overtime --------------------------------------------------------------
    max_shift    = cfg.get("max_shift_h", 14.0)
    shift_pen    = max(0.0, shift_hours - max_shift) * cfg.get("shift_penalty", 200.0)
    shift_cost   = shift_hours * cfg.get("shift_hours_weight", 0.0)

    # -- shift shortfall ---------------------------------------------------------
    # Mirrors the overtime penalty above but in the other direction.  Guarded
    # against shift_hours==0 (no timing data for this sheet) the same way
    # _sheet_cost is, so an untimed sheet isn't mistaken for a zero-length shift.
    min_shift        = cfg.get("min_shift_h", 8.0)
    shift_under_pen  = (max(0.0, min_shift - shift_hours) * cfg.get("shift_under_penalty", 30.0)
                        if shift_hours > 0 else 0.0)

    # -- milking ---------------------------------------------------------------
    milking_equiv = milking_mins * cfg.get("milking_weight", 1.0)

    # -- plant window ----------------------------------------------------------
    plant_windows     = cfg.get("plant_windows", {})
    plant_win_rate    = cfg.get("plant_win_penalty", 200.0)
    plant_margin_mins = cfg.get("plant_win_margin_mins", 30.0)
    plant_margin_rate = cfg.get("plant_win_margin_rate", plant_win_rate * 0.5)
    avoid_windows     = cfg.get("avoid_windows", AVOID_WINDOWS)
    avoid_win_rate    = cfg.get("avoid_window_penalty", 0.0)
    plant_win_cost    = 0.0
    avoid_win_cost    = 0.0
    if (plant_windows or avoid_windows) and all_times:
        for b_idx3, block3 in enumerate(blocks):
            btimes3 = all_times[b_idx3] if b_idx3 < len(all_times) else None
            if not btimes3: continue
            dests3 = block3.get("dests") or []
            if not dests3:
                dk3 = block3.get("dest_key","")
                dests3 = [{"key": dk3}] if dk3 else []
            for d_i3, dest_d3 in enumerate(dests3):
                if "yard for" in (dest_d3.get("name","") or "").lower(): continue
                dk3 = normalise_key(dest_d3.get("key","") or "")
                window3     = plant_windows.get(dk3)
                avoid_list3 = avoid_windows.get(dk3)
                if window3 is None and not avoid_list3:
                    continue
                t_idx3 = (1 if block3.get("preload")
                          else _dest_stop_index(block3, d_i3, b_idx3, blocks))
                ft3 = btimes3[t_idx3] if t_idx3 < len(btimes3) else None
                if ft3 is None or ft3.get("arr") is None: continue
                arr3 = ft3["arr"]
                arr_dt3 = datetime.combine(date.today(), arr3)

                if window3 is not None:
                    open_str3, close_str3 = window3
                    open_t3 = parse_hhmm(open_str3)
                    if not time_in_window(arr3, open_str3, close_str3):
                        if open_t3:
                            open_dt3 = datetime.combine(date.today(), open_t3)
                            if open_t3 <= arr3: open_dt3 += timedelta(days=1)
                            plant_win_cost += (open_dt3 - arr_dt3).total_seconds()/3600.0 * plant_win_rate
                        else:
                            plant_win_cost += plant_win_rate
                    else:
                        close_t3 = parse_hhmm(close_str3)
                        if close_t3:
                            close_dt3 = datetime.combine(date.today(), close_t3)
                            if open_t3 and close_t3 < open_t3: close_dt3 += timedelta(days=1)
                            mtc = (close_dt3 - arr_dt3).total_seconds()/60.0
                            if mtc < 0:
                                close_dt3 += timedelta(days=1)
                                mtc = (close_dt3 - arr_dt3).total_seconds()/60.0
                            if mtc < plant_margin_mins:
                                depth = (plant_margin_mins - mtc) / plant_margin_mins
                                plant_win_cost += depth * plant_margin_rate * (plant_margin_mins / 60.0)

                if avoid_list3 and avoid_win_rate > 0:
                    for (av_start3, av_end3) in avoid_list3:
                        if time_in_window(arr3, av_start3, av_end3):
                            avoid_win_cost += avoid_win_rate
                            break

    return {
        "km":        total_km,
        "milking":   milking_equiv,
        "shift":     shift_cost,
        "overtime":  shift_pen,
        "shortfall": shift_under_pen,
        "cap":       cap_pen,
        "plant_win": plant_win_cost,
        "avoid_win": avoid_win_cost,
        "total":     total_km + milking_equiv + shift_cost + shift_pen + shift_under_pen
                     + cap_pen + plant_win_cost + avoid_win_cost,
    }


def _optimize_split_positions(blocks, dm, start_time, cfg, dm_dur=None):
    """For each block with a mid-route partial-dropoff dest (fixed vol_partial,
    not the last dest, has a plant_window constraint), find the farm insertion
    position that minimises _sheet_cost.  Mutates blocks in-place.
    Returns True if any position changed."""
    plant_windows = cfg.get("plant_windows", {})
    changed = False
    for b_idx, block in enumerate(blocks):
        if _is_preload_block(block) or _is_fixed_vol_block(block):
            continue
        dests   = block.get("dests") or []
        n_farms = len(block["rows"])
        if n_farms == 0 or len(dests) < 2:
            continue
        for d_i, dest_d in enumerate(dests):
            if d_i == len(dests) - 1:
                continue   # last dest is never split
            if dest_d.get("vol_partial") is None:
                continue   # not a fixed partial volume
            dk = normalise_key(dest_d.get("key","") or "")
            if dk not in plant_windows:
                continue   # only optimise window-constrained dests
            best_pos  = dest_d.get("split_after") if dest_d.get("split_after") is not None else n_farms
            dest_d["split_after"] = best_pos
            best_cost = _sheet_cost(blocks, dm, start_time, cfg, dm_dur=dm_dur)
            for pos in range(n_farms + 1):
                if pos == best_pos:
                    continue
                dest_d["split_after"] = pos
                c = _sheet_cost(blocks, dm, start_time, cfg, dm_dur=dm_dur)
                if c < best_cost:
                    best_cost = c
                    best_pos  = pos
            if dest_d.get("split_after") != best_pos:
                changed = True
            dest_d["split_after"] = best_pos
    return changed


def _sheet_cost_breakdown_state(state, dm, cache, fname, cfg, dm_dur=None):
    """Aggregate _sheet_cost_breakdown across all sheets in a solver state."""
    totals = {"km":0.0,"milking":0.0,"shift":0.0,"overtime":0.0,"shortfall":0.0,"cap":0.0,"plant_win":0.0,"avoid_win":0.0,"total":0.0}
    for sname, blocks in state:
        entry = cache.get(fname, {}).get(sname, {})
        st    = entry.get("start_time") if isinstance(entry, dict) else None
        if not st: continue
        bd = _sheet_cost_breakdown(blocks, dm, st, cfg, dm_dur=dm_dur)
        for k in totals:
            totals[k] += bd.get(k, 0.0)
    return totals


def _group_vol_penalty(state, orig_dest_vols, cfg):
    """
    Compute the volume penalty for the entire colour group at once.
    Sums delivered litres to each processor across ALL sheets, then
    penalises deviations from orig_dest_vols.
    """
    vol_tol          = cfg.get("vol_tol", 0.15)
    vol_penalty_rate = cfg.get("vol_penalty", 1.0)
    dest_vols = {}
    for _sname, blocks in state:
        for block in blocks:
            for dk, off in _block_dest_offloads(block).items():
                dest_vols[dk] = dest_vols.get(dk, 0.0) + off
    pen = 0.0
    for dk, orig_vol in orig_dest_vols.items():
        cur_vol = dest_vols.get(dk, 0.0)
        lo = orig_vol * (1.0 - vol_tol)
        hi = orig_vol * (1.0 + vol_tol)
        if cur_vol < lo:
            pen += (lo - cur_vol) * vol_penalty_rate
        elif cur_vol > hi:
            pen += (cur_vol - hi) * vol_penalty_rate
    return pen



def _copy_blocks(blocks):
    """Fast structural copy of a list of block dicts for solver states.

    Copies each block dict, its rows list, and its dest dicts, but SHARES the
    individual row (farm) dicts.  Rows are treated as immutable during solving:
    operators move rows between lists or insert fresh dict() copies, and never
    edit a row's fields in place (T1/T2 volume folding mutates copies before
    the loop; the post-solve restore writes back original values).  Dest dicts
    ARE copied because _optimize_split_positions mutates "split_after" in
    place mid-loop.  This is much faster than copy.deepcopy on large states.
    """
    out = []
    for b in blocks:
        nb = dict(b)
        nb["rows"]  = list(b.get("rows") or [])
        nb["dests"] = [dict(d) for d in (b.get("dests") or [])]
        out.append(nb)
    return out


def _copy_state(state):
    """Structural copy of an entire solver state (see _copy_blocks)."""
    return [(sn, _copy_blocks(blocks)) for sn, blocks in state]


def _group_sheets_by_colour(cache, fname):
    """Return {"RED": [(sname, entry), ...], "BLUE": [...]} for the loaded file."""
    groups = {"RED": [], "BLUE": []}
    if fname not in cache:
        return groups
    for sname, entry in cache[fname].items():
        if not isinstance(entry, dict):
            continue
        bucket = _sheet_colour_bucket(entry.get("day_colour", ""))
        if bucket in groups:
            groups[bucket].append((sname, entry))
    return groups


def _highs_verify_processor_assignment(colour, sheets, state, dm, cfg, log_fn):
    """
    Post-optimality check: given the ALNS result (fixed farm sequences, fixed
    processor assignments), ask HiGHS whether reassigning processors across
    routes would reduce total cost.

    The MIP variables:
      x[i,j] in {0,1}  -  route i is assigned to processor j

    Objective: minimise  sum_{i,j} x[i,j] * last_leg_cost(i,j)
                       + vol_deviation_penalties
                       + shift_overage_penalties   (encoded as big-M terms)
                       + cap_overage_penalties

    Constraints:
      Each route assigned to exactly one processor (from that route's
      original candidate set - processors that appeared in the file for
      this colour group).
      Processor volume balance within +/-vol_tol of original.
      Truck capacity hard cap (as a big-M penalty rather than a hard cut,
      to keep the MIP feasible when data itself exceeds cap).

    Returns a string summary (multi-line) to be appended to the solver log.
    """
    try:
        from scipy.optimize import milp, LinearConstraint, Bounds  # noqa: F401 (availability probe)
        import numpy as np
    except ImportError:
        return "  [HiGHS check] scipy not available - skipping verification."

    # -- gather route data from ALNS result -----------------------------------
    # state is list of (sname, blocks); map to flat list of route descriptors
    route_records = []   # {sname, block_idx, block, vol, start_time}
    for sname, blocks in state:
        entry = {}
        for _sn, e in sheets:
            if _sn == sname: entry = e; break
        st = entry.get("start_time") if isinstance(entry, dict) else None
        for bi, block in enumerate(blocks):
            vol = sum((r["prior_vol"] or 0) for r in block.get("rows", [])
                      if isinstance(r.get("prior_vol"), (int, float)))
            route_records.append({"sname": sname, "bi": bi, "block": block,
                                  "vol": vol, "start_time": st})

    # catalogue of all processors seen in this colour group
    dest_keys = sorted({
        d.get("key") or "?"
        for _sn, blocks in state
        for block in blocks
        for d in (block.get("dests") or [{"key": block.get("dest_key","?")}])
        if d.get("key")
    })
    if not dest_keys or not route_records:
        return f"  [{colour}] HiGHS: no routes or processors - skipping."

    n_routes = len(route_records)
    n_procs  = len(dest_keys)
    proc_idx = {dk: j for j, dk in enumerate(dest_keys)}

    # original processor assignment from ALNS result
    def _route_current_dest(rec):
        dests = rec["block"].get("dests") or []
        if dests: return dests[0].get("key") or "?"
        return rec["block"].get("dest_key") or "?"

    # original processor volumes (from ALNS result - used for tolerance bounds)
    orig_dest_vols = {}
    for sname, entry_orig in sheets:
        for block in entry_orig.get("blocks", []):
            dests_b = block.get("dests") or []
            if not dests_b:
                dk = block.get("dest_key") or "?"
                dests_b = [{"key": dk, "vol_partial": None}]
            fv = sum((r["prior_vol"] or 0) for r in block.get("rows", [])
                     if isinstance(r.get("prior_vol"), (int, float)))
            already = 0.0
            for d in dests_b:
                dk = d.get("key") or "?"
                vp = d.get("vol_partial")
                rem = max(0.0, fv - already)
                off = min(float(vp), rem) if vp is not None else rem
                already += off
                orig_dest_vols[dk] = orig_dest_vols.get(dk, 0.0) + off

    # -- cost coefficients c[i*n_procs + j] ----------------------------------
    # Cost = last-farm-to-processor leg distance (km via dm, or duration if dm_dur)
    # We use the same units as _sheet_cost (km-equivalent).
    # Shift penalty and cap penalty encoded as additive per-assignment costs.
    vol_tol      = cfg.get("vol_tol", 0.15)
    hard_cap     = cfg.get("hard_vol_cap", HARD_CAP)
    cap_pen_rate = cfg.get("cap_penalty", 2.0)

    c = []
    for rec in route_records:
        block = rec["block"]
        farms = [r["irma"] for r in block.get("rows", [])]
        last_farm = farms[-1] if farms else "VEDDER"
        for dk in dest_keys:
            leg = lookup(dm, last_farm, dk)
            leg_cost = leg if leg is not None else 9999.0
            # Cap penalty contribution for this route if vol > hard_cap
            cap_contrib = max(0.0, rec["vol"] - hard_cap) * cap_pen_rate
            c.append(leg_cost + cap_contrib)

    # -- integrality: all binary -----------------------------------------------
    integrality = [1] * (n_routes * n_procs)  # 1 = integer

    # -- equality constraint: each route assigned to exactly one processor -----
    # sum_j x[i,j] = 1  for each i
    A_eq_rows, b_eq = [], []
    for i in range(n_routes):
        row = [0.0] * (n_routes * n_procs)
        for j in range(n_procs):
            row[i * n_procs + j] = 1.0
        A_eq_rows.append(row)
        b_eq.append(1.0)

    # -- inequality constraints: volume tolerance per processor ----------------
    # sum_i vol_i * x[i,j]  <=  orig_vol_j * (1 + vol_tol)
    # sum_i vol_i * x[i,j]  >=  orig_vol_j * (1 - vol_tol)   [as <= with negation]
    A_ineq_rows, b_ineq = [], []
    for j, dk in enumerate(dest_keys):
        orig_v = orig_dest_vols.get(dk, 0.0)
        if orig_v <= 0:
            continue
        hi = orig_v * (1.0 + vol_tol)
        lo = orig_v * (1.0 - vol_tol)
        # upper bound
        row_hi = [0.0] * (n_routes * n_procs)
        for i, rec in enumerate(route_records):
            row_hi[i * n_procs + j] = rec["vol"]
        A_ineq_rows.append(row_hi); b_ineq.append(hi)
        # lower bound (negate)
        row_lo = [-x for x in row_hi]
        A_ineq_rows.append(row_lo); b_ineq.append(-lo)

    c_arr     = np.array(c, dtype=float)
    bounds_   = Bounds(lb=0.0, ub=1.0)

    constraints = []
    if A_eq_rows:
        A_eq = np.array(A_eq_rows, dtype=float)
        constraints.append(LinearConstraint(A_eq, lb=np.array(b_eq), ub=np.array(b_eq)))
    if A_ineq_rows:
        A_iq = np.array(A_ineq_rows, dtype=float)
        neg_inf = np.full(len(b_ineq), -np.inf)
        constraints.append(LinearConstraint(A_iq, lb=neg_inf, ub=np.array(b_ineq)))

    try:
        result = milp(c_arr, constraints=constraints, integrality=np.array(integrality),
                      bounds=bounds_, options={"disp": False, "time_limit": 30.0})
    except Exception as ex:
        return f"  [{colour}] HiGHS MIP failed: {ex}"

    if not result.success:
        return f"  [{colour}] HiGHS: no feasible processor assignment found ({result.message})."

    # -- compute current (ALNS) last-leg cost for comparison ------------------
    current_cost = 0.0
    for i, rec in enumerate(route_records):
        cur_dk = _route_current_dest(rec)
        j_cur  = proc_idx.get(cur_dk, 0)
        current_cost += c[i * n_procs + j_cur]

    mip_cost = result.fun
    improvement = current_cost - mip_cost

    lines = [f"\n-- [{colour}] HiGHS Processor Assignment Verification --"]
    lines.append(f"   ALNS last-leg + cap cost : {current_cost:,.1f} km-eq")
    lines.append(f"   MIP optimal cost          : {mip_cost:,.1f} km-eq")

    THRESHOLD = 0.5   # ignore sub-km differences (floating point noise)
    if improvement <= THRESHOLD:
        lines.append("   OK Processor assignment is OPTIMAL - no improvement possible.")
    else:
        lines.append(f"   * Improvement available   : {improvement:,.1f} km-eq")
        lines.append("   Suggested reassignments:")
        x_sol = result.x
        for i, rec in enumerate(route_records):
            cur_dk   = _route_current_dest(rec)
            best_j   = int(np.argmax(x_sol[i*n_procs:(i+1)*n_procs]))
            best_dk  = dest_keys[best_j]
            if best_dk != cur_dk:
                farms = [r["irma"] for r in rec["block"].get("rows", [])]
                last  = farms[-1] if farms else "?"
                cur_cost_leg  = c[i * n_procs + proc_idx.get(cur_dk, 0)]
                best_cost_leg = c[i * n_procs + best_j]
                lines.append(
                    f"     Sheet {rec['sname']!r} block {rec['bi']+1}: "
                    f"{cur_dk} -> {best_dk}  "
                    f"(last farm {last}, "
                    f"leg {cur_cost_leg:.1f} -> {best_cost_leg:.1f} km-eq)")
    return "\n".join(lines)


class IntraRouteOptimiser(QThread):
    """Applies 2-opt + or-opt within every route until convergence.
    No cross-route moves - pure within-block reordering."""
    progress = pyqtSignal(int, int, str)   # cur, total, status
    finished = pyqtSignal(dict)            # {(fname,sname): improved_blocks}
    log      = pyqtSignal(str)

    def __init__(self, fname, cache, dm, cfg, sheet_mods, parent=None,
                 dm_dur=None, locked_sheets=None):
        super().__init__(parent)
        self.fname         = fname
        self.cache         = cache
        self.dm            = dm
        self.dm_dur        = dm_dur
        self.cfg           = cfg
        self.sheet_mods    = dict(sheet_mods)
        self.locked_sheets = {str(s).strip() for s in (locked_sheets or set())}

    def run(self):
        results     = {}
        all_snames  = sorted(self.cache.get(self.fname, {}).keys())
        skip        = SOLVER_SKIP_SHEETS | self.locked_sheets
        all_snames  = [s for s in all_snames if s.strip() not in skip]
        total       = len(all_snames)
        n_improved  = 0

        for cur, sname in enumerate(all_snames):
            entry = self.cache[self.fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue

            key    = (self.fname, sname)
            blocks = copy.deepcopy(
                self.sheet_mods.get(key, entry.get("blocks", [])))

            self.progress.emit(cur, total, f"Optimising {sname}...")

            improved = True
            changed  = False
            while improved:
                improved = False
                # 2-opt
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    base_c    = _sheet_cost(blocks, self.dm, start_time, self.cfg, dm_dur=self.dm_dur)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n - 1):
                        for j in range(i + 1, n):
                            trial = rows[:i] + rows[i:j+1][::-1] + rows[j+1:]
                            blocks[b_idx] = dict(block, rows=trial)
                            c = _sheet_cost(blocks, self.dm, start_time, self.cfg, dm_dur=self.dm_dur)
                            if c < best_c:
                                best_c = c; best_rows = trial[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True; changed = True

                # or-opt (single farm relocation)
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    base_c    = _sheet_cost(blocks, self.dm, start_time, self.cfg, dm_dur=self.dm_dur)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n):
                        farm = rows[i]
                        rest = rows[:i] + rows[i+1:]
                        for j in range(len(rest) + 1):
                            trial = rest[:j] + [farm] + rest[j:]
                            blocks[b_idx] = dict(block, rows=trial)
                            c = _sheet_cost(blocks, self.dm, start_time, self.cfg, dm_dur=self.dm_dur)
                            if c < best_c:
                                best_c = c; best_rows = trial[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True; changed = True

            if changed:
                results[key] = blocks
                n_improved += 1

        self.progress.emit(total, total, "Done")
        self.log.emit(f"Intra-route optimisation complete - {n_improved} route(s) improved")
        self.finished.emit(results)


def _apply_prob_floors_ceilings(probs, key_floors, key_ceilings):
    """Redistribute a probability vector in place so each entry respects its
    per-index floor and ceiling, iterating until stable.

    Below-floor entries are raised to their floor and the deficit is taken
    proportionally from entries above their floor; above-ceiling entries are
    clamped and the surplus is redistributed proportionally to those below
    their ceiling.  Returns the same list (mutated).  Shared by _roulette
    (selection) and the solver's probability-display logging so the two can
    never drift.
    """
    n = len(probs)
    if not (any(f > 0.0 for f in key_floors) or any(c < 1.0 for c in key_ceilings)):
        return probs
    for _ in range(n * 2):   # extra iterations for combined floor+ceiling cascades
        changed = False
        # Raise below-floor keys
        below = [i for i, p in enumerate(probs) if p < key_floors[i]]
        above_floor = [i for i, p in enumerate(probs) if p >= key_floors[i]]
        if below and above_floor:
            deficit   = sum(key_floors[i] - probs[i] for i in below)
            above_sum = sum(probs[i] for i in above_floor)
            for i in below:
                probs[i] = key_floors[i]
            if above_sum > 0:
                for i in above_floor:
                    probs[i] -= deficit * (probs[i] / above_sum)
            changed = True
        # Clamp above-ceiling keys
        over = [i for i, p in enumerate(probs) if p > key_ceilings[i] + 1e-9]
        over_set = set(over)
        under_ceil = [i for i, p in enumerate(probs)
                      if p <= key_ceilings[i] and i not in over_set]
        if over and under_ceil:
            surplus   = sum(probs[i] - key_ceilings[i] for i in over)
            under_sum = sum(probs[i] for i in under_ceil)
            for i in over:
                probs[i] = key_ceilings[i]
            if under_sum > 0:
                for i in under_ceil:
                    probs[i] += surplus * (probs[i] / under_sum)
            changed = True
        if not changed:
            break
    return probs


class ALNSSolver(QThread):
    """
    Adaptive Large Neighbourhood Search.

    Operates on RED and BLUE colour groups independently.
    Within each group, farms may move between any sheets/routes of the same
    colour, and block destinations (dest_key) may also be reassigned.

    Three paired move types, each with its own adaptive weight:
      1. Farm move   - destroy (random | worst) + repair (best-insert | regret)
      2. Dest move   - strip dest_keys from n blocks + regret-order reassignment
      3. Combined    - both farm and dest moves together

    Simulated-annealing acceptance; alpha auto-computed from user's
    target_cool_frac and iteration count.

    Signals
    -------
    progress(cur_iter, total_iters, status_str)
    finished({sname: new_blocks, ...})
    log(message_str)
    """

    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(dict)
    log      = pyqtSignal(str)

    def __init__(self, fname, cache, dm, cfg, parent=None, sheet_mods=None,
                 locked_sheets=None, dm_dur=None):
        super().__init__(parent)
        self.fname         = fname
        self.cache         = cache
        self.dm            = dm
        self.dm_dur        = dm_dur
        self.cfg           = cfg
        self._stop         = False
        self.sheet_mods    = sheet_mods or {}
        # locked_sheets: set of sname strings the solver must not modify
        self.locked_sheets = {str(s).strip() for s in (locked_sheets or set())}
        # Fixed dock-visit background for locked sheets in the colour group
        # currently being solved - recomputed once per _solve_group_inner
        # call (locked sheets never change during a solve, so this is cheap
        # to compute once rather than per-iteration).  Read by
        # _group_overlap_penalty so active routes get penalized for
        # colliding with a locked truck's known, fixed dock time.
        self._locked_dest_visits = {}   # {dest_key: [(arr_min, dep_min), ...]}
        # Accumulated across all colour groups during run(); read by MainWindow
        # after solving for logging paired trailers held with their lead.
        self.paired_followers = []   # (sname, b_idx, lead_uid, follower_dict, order)

    def stop(self):
        self._stop = True

    # -- group-level cost ------------------------------------------------------

    def _extract_dest_visits(self, blocks, start_time):
        """
        Return [(dest_key, arr_min, dep_min), ...] for one sheet's blocks -
        every real dock visit (yard-for overnight parking excluded), with
        arr/dep as continuous minutes on the shared axis (see
        _continuous_minutes).  Shared by _group_overlap_penalty for the
        active solver state and by the locked-sheet background precomputed
        in _solve_group_inner, so both use identical extraction logic.
        """
        if not start_time:
            return []
        suppress = self.cfg.get("suppress_no_milking", True)
        ct = calc_times(blocks, self.dm, start_time, self.dm_dur,
                        suppress_no_milking=suppress)
        if ct is None:
            return []
        all_times, _ = ct
        out = []
        for b_idx, block in enumerate(blocks):
            btimes = all_times[b_idx] if b_idx < len(all_times) else None
            if not btimes:
                continue
            dests = block.get("dests") or []
            if not dests:
                dk0 = block.get("dest_key", "")
                dests = [{"key": dk0}] if dk0 else []
            for d_i, dest_d in enumerate(dests):
                dn = (dest_d.get("name", "") or "")
                if "yard for" in dn.lower():
                    continue   # overnight parking, not a real dock visit
                dk = normalise_key(dest_d.get("key", "") or "")
                if not dk:
                    continue
                t_idx = _dest_stop_index(block, d_i, b_idx, blocks)
                ft = btimes[t_idx] if t_idx < len(btimes) else None
                if ft is None or ft.get("arr") is None or ft.get("dep") is None:
                    continue
                arr_m = _continuous_minutes(ft["arr"], start_time)
                dep_m = _continuous_minutes(ft["dep"], start_time)
                if dep_m < arr_m:
                    dep_m += 24 * 60
                out.append((dk, arr_m, dep_m))
        return out

    def _group_overlap_penalty(self, state):
        """
        Penalize truck visits at the same processor exceeding that
        processor's dock capacity within this colour group.

        The solver only ever operates on one colour group's sheets at a time
        (RED, BLUE, and GRASSFED run on entirely separate days and are solved
        independently - see _solve_group_inner), so every pair of sheets
        compared here is automatically from the same day.  There is no need
        to check colour explicitly: a cross-colour pair can never appear
        together in `state` in the first place, since `state` only ever
        contains sheets from the one colour group currently being solved.

        Most processors can only take one truck at a time, but a few (see
        PROCESSOR_DOCK_CAPACITY) have multiple bays and can take 2+ at once
        without it being a real problem - this is a CAPACITY check via a
        sweep-line over arrival/departure events, not a simple "any overlap
        is bad" pairwise check, so e.g. capacity 2 means two trucks
        overlapping is free, but a third overlapping both of them is not.

        Locked sheets (SOLVER_SKIP_SHEETS or user-locked) are excluded from
        `state` entirely before solving even starts, so without extra help
        this function would have zero visibility into where a locked truck
        is parked - the solver could happily schedule an active truck right
        on top of a locked one's known, fixed dock time.  self._locked_dest_visits
        (precomputed once per colour group in _solve_group_inner, since
        locked sheets never change during the solve) supplies that fixed
        background: it's merged in here as additional occupied time at each
        dock, so active routes get penalized for colliding with a locked
        route, even though the locked route itself is never touched or
        re-evaluated.

        cfg["overlap_penalty"]: km-equivalent cost per minute of capacity
        EXCESS (concurrent trucks beyond what the dock can take), summed
        across every processor.  Default 0.0 - disabled until explicitly
        turned on, same as the avoid-window penalty.
        cfg["dock_capacity"]: optional override of PROCESSOR_DOCK_CAPACITY.
        """
        rate = self.cfg.get("overlap_penalty", 0.0)
        if rate <= 0:
            return 0.0

        dock_capacity = self.cfg.get("dock_capacity", PROCESSOR_DOCK_CAPACITY)
        visits_by_dest = {}   # dest_key -> [(arr_min, dep_min), ...]

        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            start_time = entry.get("start_time") if isinstance(entry, dict) else None
            for (dk, arr_m, dep_m) in self._extract_dest_visits(blocks, start_time):
                visits_by_dest.setdefault(dk, []).append((arr_m, dep_m))

        # Merge in the fixed locked-sheet background, if any was precomputed.
        for dk, locked_vs in getattr(self, "_locked_dest_visits", {}).items():
            if locked_vs:
                visits_by_dest.setdefault(dk, []).extend(locked_vs)

        total_excess_min = 0.0
        for dk, vs in visits_by_dest.items():
            if len(vs) < 2:
                continue
            capacity = dock_capacity.get(dk, 1)
            # Sweep-line: +1 at each arrival, -1 at each departure.  Sort by
            # time, processing departures (-1) before arrivals (+1) at an
            # exact tie so a truck leaving the instant another arrives isn't
            # counted as an overlap.  Between consecutive distinct event
            # times, the concurrent count is constant; whenever it exceeds
            # capacity, that interval's length times the excess count is
            # added to the penalty.
            events = []
            for (a, d) in vs:
                events.append((a, 1))
                events.append((d, -1))
            events.sort(key=lambda e: (e[0], e[1]))
            count   = 0
            prev_t  = None
            for (t, delta) in events:
                if prev_t is not None and t > prev_t and count > capacity:
                    total_excess_min += (count - capacity) * (t - prev_t)
                count += delta
                prev_t = t

        return total_excess_min * rate

    def _auto_night_start_mins(self):
        """Return the earliest start time of any night-shift route in this file
        as minutes since midnight, or None if no night sheets exist."""
        earliest = None
        for entry in self.cache.get(self.fname, {}).values():
            if not isinstance(entry, dict):
                continue
            st = entry.get("start_time")
            if st is None or _is_day_sheet(st):
                continue
            t = st.time() if isinstance(st, datetime) else st
            m = t.hour * 60 + t.minute
            if earliest is None or m < earliest:
                earliest = m
        return earliest

    def _sheet_end_minutes(self, blocks, start_time):
        """Return a route's depot-return time as minutes since midnight, or None.

        Values > 1440 indicate the route runs past midnight.  Uses calc_times'
        final cursor so the result is consistent with the shift-cost accounting.
        """
        if not start_time:
            return None
        suppress = self.cfg.get("suppress_no_milking", True)
        ct = calc_times(blocks, self.dm, start_time, self.dm_dur,
                        suppress_no_milking=suppress)
        if ct is None:
            return None
        _, end_cursor = ct
        if end_cursor is None:
            return None
        et = end_cursor.time() if isinstance(end_cursor, datetime) else end_cursor
        end_m = et.hour * 60 + et.minute
        # Detect overnight: end clock time < start clock time means +24h
        st = start_time.time() if isinstance(start_time, datetime) else start_time
        start_m = st.hour * 60 + st.minute
        if end_m < start_m:
            end_m += 24 * 60
        return end_m

    def _truck_avail_penalty(self, state):
        """Penalty when too few day-shift trucks finish before the night shift starts.

        cfg keys consumed:
          truck_avail_enabled       : bool  - master switch
          truck_avail_night_mins    : int   - night-shift start (minutes, auto-detected)
          _truck_avail_group_needed : int   - pre-scaled min trucks back for this group
          truck_avail_penalty       : float - cost per truck short of minimum
        """
        if not self.cfg.get("truck_avail_enabled", False):
            return 0.0
        night_mins = self.cfg.get("truck_avail_night_mins")
        needed     = self.cfg.get("_truck_avail_group_needed", 0)
        pen_rate   = self.cfg.get("truck_avail_penalty", 3000.0)
        if night_mins is None or needed <= 0:
            return 0.0

        on_time = 0
        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            if not isinstance(entry, dict):
                continue
            st = entry.get("start_time")
            if not _is_day_sheet(st):
                continue
            end_m = self._sheet_end_minutes(blocks, st)
            if end_m is not None and end_m <= night_mins:
                on_time += 1

        shortage = max(0, needed - on_time)
        return shortage * pen_rate

    def _group_cost(self, state, orig_dest_vols, sheet_cost_cache=None):
        """Full cost: sum per-sheet costs + group-wide volume penalty.

        sheet_cost_cache: optional dict {sname: cost} - if provided, sheets
        present in the cache are not recomputed.  Callers that know which sheets
        changed can pass a partial cache to skip unchanged sheets entirely.
        """
        total = 0.0
        for sname, blocks in state:
            if sheet_cost_cache is not None and sname in sheet_cost_cache:
                total += sheet_cost_cache[sname]
            else:
                entry = self.cache[self.fname].get(sname, {})
                st    = entry.get("start_time") if isinstance(entry, dict) else None
                c     = _sheet_cost(blocks, self.dm, st, self.cfg, dm_dur=self.dm_dur)
                total += c
                if sheet_cost_cache is not None:
                    sheet_cost_cache[sname] = c
        total += _group_vol_penalty(state, orig_dest_vols, self.cfg)
        total += self._group_overlap_penalty(state)
        total += self._truck_avail_penalty(state)
        return total

    def _make_sheet_cost_cache(self, state):
        """Compute and return a full per-sheet cost cache for the given state."""
        cache = {}
        for sname, blocks in state:
            entry = self.cache[self.fname].get(sname, {})
            st    = entry.get("start_time") if isinstance(entry, dict) else None
            cache[sname] = _sheet_cost(blocks, self.dm, st, self.cfg, dm_dur=self.dm_dur)
        return cache

    # -- flat farm list --------------------------------------------------------

    def _flatten_farms(self, state):
        out = []
        for s_idx, (sname, blocks) in enumerate(state):
            for b_idx, block in enumerate(blocks):
                for farm in block["rows"]:
                    out.append((s_idx, b_idx, farm))
        return out

    # ══════════════════════════════════════════════════════════════════════════
    # FARM destroy / repair
    # ══════════════════════════════════════════════════════════════════════════

    def _destroy_random(self, state, n_remove):
        """Remove n_remove random farms.
        - Preload blocks (previous-day delivery, no farms): completely frozen.
        - Holdover blocks (yard-for dest): farms CAN be moved; dest stays fixed.
        """
        flat = []
        for s_idx, (sname, blocks) in enumerate(state):
            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block):
                    continue   # frozen - previous-day load, no farms anyway
                if _is_fixed_vol_block(block):
                    continue   # frozen - explicit vol_partial on every dest
                for f_idx in range(len(block["rows"])):
                    flat.append((s_idx, b_idx, f_idx))
        if not flat:
            return _copy_state(state), []
        n_remove = min(n_remove, len(flat))
        chosen   = set(map(tuple, random.sample(flat, n_remove)))

        removed   = []
        new_state = []
        for s_idx, (sname, blocks) in enumerate(state):
            new_blocks = []
            for b_idx, block in enumerate(blocks):
                keep = []
                for f_idx, farm in enumerate(block["rows"]):
                    if (s_idx, b_idx, f_idx) in chosen:
                        removed.append((s_idx, b_idx, dict(farm)))
                    else:
                        keep.append(dict(farm))
                new_blocks.append(dict(block, rows=keep))
            new_state.append((sname, new_blocks))
        return new_state, removed

    def _destroy_worst(self, state, n_remove):
        """Remove the n farms whose individual removal saves the most sheet-km.
        Preload blocks are completely frozen; holdover-block farms are fair game."""
        savings = []
        for s_idx, (sname, blocks) in enumerate(state):
            entry = self.cache[self.fname].get(sname, {})
            st    = entry.get("start_time") if isinstance(entry, dict) else None
            base  = _sheet_cost(blocks, self.dm, st, self.cfg, dm_dur=self.dm_dur)
            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block):
                    continue   # frozen
                if _is_fixed_vol_block(block):
                    continue   # frozen - explicit vol_partial on every dest
                for f_idx in range(len(block["rows"])):
                    trial = _copy_blocks(blocks)
                    trial[b_idx]["rows"].pop(f_idx)
                    new_c = _sheet_cost(trial, self.dm, st, self.cfg, dm_dur=self.dm_dur)
                    savings.append((base - new_c, s_idx, b_idx, f_idx))
        savings.sort(reverse=True)

        # Collect top-n by (s_idx, b_idx, f_idx), at most one f_idx per block
        # to avoid index-shifting conflicts
        seen    = {}
        for saving, s_idx, b_idx, f_idx in savings:
            key = (s_idx, b_idx)
            seen.setdefault(key, []).append(f_idx)
            if sum(len(v) for v in seen.values()) >= n_remove:
                break

        # Now deepcopy state and remove by index (high-to-low to preserve indices)
        new_state = _copy_state(state)
        removed   = []
        for (s_idx, b_idx), idxs in seen.items():
            block = new_state[s_idx][1][b_idx]
            for f_idx in sorted(set(idxs), reverse=True):
                if f_idx < len(block["rows"]):
                    removed.append((s_idx, b_idx, block["rows"].pop(f_idx)))
        return new_state, removed

    def _best_insert_cost(self, blocks, farm, dm, shift_start=None,
                          baseline=None, dm_dur=None):
        """
        Return (b_idx, pos, marginal_cost) for cheapest insertion of farm.

        Computes baseline _sheet_cost once (or accepts a pre-computed one),
        then evaluates only the delta for each candidate position.  Saves one
        _sheet_cost call per (farm, sheet) pair when the caller caches baselines
        across multiple farms on the same sheet.

        Skips preload blocks and fixed-vol blocks.
        """
        if baseline is None:
            baseline = _sheet_cost(blocks, dm, shift_start, self.cfg, dm_dur=dm_dur)

        best_b, best_pos, best_cost = None, None, float("inf")

        for b_idx, block in enumerate(blocks):
            if _is_preload_block(block):
                continue
            if _is_fixed_vol_block(block):
                continue
            rows = block["rows"]
            for pos in range(len(rows) + 1):
                trial_rows   = rows[:pos] + [farm] + rows[pos:]
                trial_block  = dict(block, rows=trial_rows)
                trial_blocks = blocks[:b_idx] + [trial_block] + blocks[b_idx+1:]
                delta = _sheet_cost(trial_blocks, dm, shift_start, self.cfg, dm_dur=dm_dur) - baseline
                if delta < best_cost:
                    best_b, best_pos, best_cost = b_idx, pos, delta

        return best_b, best_pos, best_cost

    def _repair_best(self, state, removed):
        """Greedy best-insertion: each farm -> cheapest slot across whole group.

        Sheets with no start_time are frozen - reliable arrival estimates
        require a real shift start; without one the sheet is skipped entirely.
        Farms are inserted most-constrained-first (highest minimum insertion
        cost) so hard-to-place farms claim their preferred slot before easier
        ones, reducing cascades of suboptimal placements.
        """
        state = _copy_state(state)
        # Per-sheet start times. None -> no start time -> sheet is frozen.
        start_map = {}
        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            start_map[sname] = entry.get("start_time") if isinstance(entry, dict) else None

        # Only sheets with a real start time are eligible for insertion.
        # When day_night_lock is on, also build a per-sheet day/night tag so
        # the repair can restrict each farm to its origin shift class.
        lock = self.cfg.get("day_night_lock", False)
        day_map = {sname: _is_day_sheet(st)
                   for sname, st in start_map.items()} if lock else {}

        eligible = [(s_idx, sname, blocks)
                    for s_idx, (sname, blocks) in enumerate(state)
                    if start_map.get(sname) is not None]

        # Pre-compute baseline _sheet_cost for each eligible sheet.
        # This saves one _sheet_cost call per (farm x sheet) combination -
        # with 20 removed farms and 27 sheets that's 540 saved calls per repair.
        baseline_cache = {
            sname: _sheet_cost(blocks, self.dm, start_map[sname], self.cfg, dm_dur=self.dm_dur)
            for _, sname, blocks in eligible
        }

        # (farm, sheet, generation) memo - see _repair_regret for rationale.
        # Results computed during the _min_cost ordering pass are reused by
        # the insertion loop below for every sheet not modified in between.
        sheet_gen   = {sname: 0 for sname, _ in state}
        insert_memo = {}

        def _memo_insert_cost(farm, sname, blocks):
            mkey = (id(farm), sname, sheet_gen[sname])
            res  = insert_memo.get(mkey)
            if res is None:
                res = self._best_insert_cost(blocks, farm, self.dm,
                                             shift_start=start_map[sname],
                                             baseline=baseline_cache[sname],
                                             dm_dur=self.dm_dur)
                insert_memo[mkey] = res
            return res

        def _min_cost(item):
            _, _, farm = item
            best_c = float("inf")
            for _, sname, blocks in eligible:
                b, pos, c = _memo_insert_cost(farm, sname, blocks)
                if b is not None and c < best_c:
                    best_c = c
            return best_c

        ordered = sorted(removed, key=_min_cost, reverse=True)
        cross_route = 0
        for s_hint, b_hint, farm in ordered:
            best_s, best_b, best_pos, best_c = None, None, None, float("inf")
            # When locked, only consider sheets in the same day/night class
            # as the farm's origin sheet.
            if lock and s_hint < len(state):
                origin_is_day = day_map.get(state[s_hint][0], True)
                candidates = [(si, sn, bl) for si, sn, bl in eligible
                              if day_map.get(sn, True) == origin_is_day]
            else:
                candidates = eligible
            for s_idx, sname, blocks in candidates:
                b_idx, pos, c = _memo_insert_cost(farm, sname, blocks)
                if b_idx is not None and c < best_c:
                    best_s, best_b, best_pos, best_c = s_idx, b_idx, pos, c
            if best_s is not None:
                if best_s != s_hint:
                    cross_route += 1
                state[best_s][1][best_b]["rows"].insert(best_pos, dict(farm))
                # Invalidate baseline and memoised insertions for the modified
                # sheet so subsequent farms see the updated cost.
                sname_mod = state[best_s][0]
                sheet_gen[sname_mod] += 1
                baseline_cache[sname_mod] = _sheet_cost(
                    state[best_s][1], self.dm, start_map[sname_mod], self.cfg,
                    dm_dur=self.dm_dur)
        return state, cross_route

    def _repair_regret(self, state, removed, k=2):
        """k-regret insertion: always insert the farm with the highest regret first.

        Sheets with no start_time are frozen and excluded from candidate slots.
        """
        state   = _copy_state(state)
        pending = list(removed)
        # Per-sheet start times. None -> frozen.
        start_map = {}
        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            start_map[sname] = entry.get("start_time") if isinstance(entry, dict) else None

        # Only sheets with a real start time are eligible for insertion.
        lock = self.cfg.get("day_night_lock", False)
        day_map = {sname: _is_day_sheet(start_map[sname])
                   for sname in start_map if start_map[sname] is not None} if lock else {}

        eligible_idxs = {s_idx for s_idx, (sname, _) in enumerate(state)
                         if start_map.get(sname) is not None}

        # Pre-compute baseline _sheet_cost for each eligible sheet.
        baseline_cache = {
            sname: _sheet_cost(blocks, self.dm, start_map[sname], self.cfg, dm_dur=self.dm_dur)
            for s_idx, (sname, blocks) in enumerate(state)
            if s_idx in eligible_idxs
        }

        # (farm, sheet, generation) -> _best_insert_cost result.  A sheet's
        # generation bumps only when a farm is inserted into it, so results
        # for untouched sheets are reused verbatim across regret rounds
        # instead of being recomputed (they are deterministic and identical).
        # Keyed by id(farm): farm dicts are pinned by the caller's `removed`
        # list for the lifetime of this memo, so ids are stable and unique here.
        sheet_gen   = {sname: 0 for sname, _ in state}
        insert_memo = {}

        cross_route = 0
        while pending:
            best_farm_i = None
            best_regret = -float("inf")
            best_slot   = None
            for i, (s_hint, b_hint, farm) in enumerate(pending):
                # When locked, restrict target sheets to same day/night class
                if lock and s_hint < len(state):
                    origin_is_day = day_map.get(state[s_hint][0], True)
                    allowed = {si for si in eligible_idxs
                               if day_map.get(state[si][0], True) == origin_is_day}
                else:
                    allowed = eligible_idxs
                slot_costs = []
                for s_idx, (sname, blocks) in enumerate(state):
                    if s_idx not in allowed:
                        continue   # sheet has no start time or wrong shift class - frozen
                    mkey = (id(farm), sname, sheet_gen[sname])
                    res  = insert_memo.get(mkey)
                    if res is None:
                        res = self._best_insert_cost(blocks, farm, self.dm,
                                                     shift_start=start_map[sname],
                                                     baseline=baseline_cache[sname],
                                                     dm_dur=self.dm_dur)
                        insert_memo[mkey] = res
                    b_idx, pos, c = res
                    if b_idx is not None:
                        slot_costs.append((c, s_idx, b_idx, pos))
                slot_costs.sort()
                if not slot_costs:
                    continue
                regret = (slot_costs[1][0] - slot_costs[0][0]) if len(slot_costs) >= k else 0.0
                if regret > best_regret:
                    best_regret = regret
                    best_farm_i = i
                    best_slot   = slot_costs[0]
            if best_farm_i is None:
                break
            _, s_idx, b_idx, pos = best_slot
            s_hint, b_hint, farm = pending.pop(best_farm_i)
            if s_idx != s_hint:
                cross_route += 1
            state[s_idx][1][b_idx]["rows"].insert(pos, dict(farm))
            # Invalidate baseline and memoised insertions for the modified sheet
            sname_mod = state[s_idx][0]
            sheet_gen[sname_mod] += 1
            baseline_cache[sname_mod] = _sheet_cost(
                state[s_idx][1], self.dm, start_map[sname_mod], self.cfg,
                dm_dur=self.dm_dur)
        return state, cross_route

    # ══════════════════════════════════════════════════════════════════════════
    # DEST destroy / repair
    # ══════════════════════════════════════════════════════════════════════════

    @staticmethod
    def _acc_block_vols(block, vols_dict):
        """Accumulate delivered litres per dest-key from one block into vols_dict."""
        dests_b  = block.get("dests") or []
        farm_vol = sum((r.get("prior_vol") or 0) for r in block["rows"]
                       if isinstance(r.get("prior_vol"), (int, float)))
        already  = 0.0
        for d in dests_b:
            dk  = d.get("key") or "?"
            vp  = d.get("vol_partial")
            rem = max(0.0, farm_vol - already)
            off = min(float(vp), rem) if vp is not None else rem
            already += off
            vols_dict[dk] = vols_dict.get(dk, 0.0) + off

    def _destroy_dest(self, state, n_strip, orig_dest_vols):
        """
        Randomly select n_strip blocks and strip their entire dest list.
        Works even when every block only has a single destination.
        Returns (new_state, stripped: [(s_idx, b_idx, dests_list), ...]).
        """
        state = _copy_state(state)
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sname, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if (block.get("dests") or block.get("dest_key"))
               and not _is_holdover_block(block)    # yard-for dest stays fixed
               and not _is_preload_block(block)      # preload block fully frozen
               and not _is_fixed_vol_block(block)    # fixed vol_partial - operational instruction
        ]
        if not candidates:
            return state, []
        n_strip  = min(n_strip, len(candidates))
        chosen   = random.sample(candidates, n_strip)
        stripped = []
        for s_idx, b_idx in chosen:
            block = state[s_idx][1][b_idx]
            dests = block.get("dests") or []
            if not dests:
                dk = block.get("dest_key",""); dn = block.get("dest_name","")
                dests = [{"key": dk, "name": dn, "vol_partial": None}] if dk else []
            stripped.append((s_idx, b_idx, [dict(d) for d in dests]))
            block["dests"]     = []
            block["dest_key"]  = ""
            block["dest_name"] = ""
        return state, stripped

    def _repair_dest(self, state, stripped, dest_catalogue, orig_dest_vols):
        """
        Re-assign stripped dest lists to blocks using regret-order greedy.
        stripped: [(s_idx, b_idx, dests_list), ...]

        For each stripped item we find the cheapest block (any block, not just
        original) to attach the dest list to. This lets the solver redistribute
        processors across routes.

        Cost of attaching dests_list to block (s2, b2) =
          route_km(block with dests_list appended) + volume_penalty_delta
        """
        state   = _copy_state(state)
        pending = list(stripped)   # [(s_idx, b_idx, dests_list)]

        # running vol totals for all already-assigned dests
        cur_vols = {}
        for _sname, blocks in state:
            for block in blocks:
                ALNSSolver._acc_block_vols(block, cur_vols)

        vol_tol  = self.cfg.get("vol_tol", 0.15)
        vol_rate = self.cfg.get("vol_penalty", 1.0)

        def _block_origin(s2, b2):
            """True origin of block (s2,b2): VEDDER if first, else last dest of prev block."""
            if b2 == 0:
                return "VEDDER"
            prev = state[s2][1][b2 - 1]
            prev_dests = prev.get("dests") or []
            if prev_dests:
                return prev_dests[-1].get("key","") or "VEDDER"
            return prev.get("dest_key","") or "VEDDER"

        plant_windows       = self.cfg.get("plant_windows", {})
        plant_win_rate      = self.cfg.get("plant_win_penalty", 500.0)
        plant_margin_mins   = self.cfg.get("plant_win_margin_mins", 30.0)
        plant_margin_rate   = self.cfg.get("plant_win_margin_rate", plant_win_rate * 0.5)
        avoid_windows       = self.cfg.get("avoid_windows", AVOID_WINDOWS)
        avoid_win_rate      = self.cfg.get("avoid_window_penalty", 0.0)

        def _attach_cost(s2, b2, dests_list):
            """Cost of giving dests_list to block (s2,b2), using correct block origin."""
            block     = state[s2][1][b2]

            # Guard: if dests_list is fully capped (every dest has fixed vol_partial,
            # i.e. no catch-all remainder), the total deliverable volume is fixed at
            # the sum of all vol_partials.  If the block's farm_vol exceeds that cap,
            # assigning this dest list here silently drops the overflow - identical to
            # the _is_fixed_vol_block problem.  Apply a prohibitive overflow penalty.
            if dests_list and dests_list[-1].get("vol_partial") is not None:
                dest_cap = sum(float(d.get("vol_partial") or 0) for d in dests_list)
                farm_vol = sum((r.get("prior_vol") or 0) for r in block["rows"]
                               if isinstance(r.get("prior_vol"), (int, float)))
                overflow = max(0.0, farm_vol - dest_cap)
                if overflow > 0:
                    cap_pen_rate = self.cfg.get("cap_penalty", 2.0)
                    return overflow * cap_pen_rate * 1000  # prohibitive

            farm_keys = [r["irma"] for r in block["rows"]]
            dest_keys = [d.get("key","") for d in dests_list if d.get("key")]
            blocks_in_sheet = state[s2][1]
            is_last   = (b2 == len(blocks_in_sheet) - 1)
            origin    = _block_origin(s2, b2)
            stops     = [origin] + farm_keys + dest_keys
            if is_last:
                stops = stops + ["VEDDER"]
            leg_km    = sum(lookup(self.dm, stops[i], stops[i+1]) or 0.0
                            for i in range(len(stops)-1))
            # volume penalty delta
            farm_vol  = sum((r.get("prior_vol") or 0) for r in block["rows"]
                            if isinstance(r.get("prior_vol"), (int, float)))
            already   = 0.0
            vol_pen   = 0.0
            for d in dests_list:
                dk  = d.get("key","")
                vp  = d.get("vol_partial")
                rem = max(0.0, farm_vol - already)
                off = min(float(vp), rem) if vp is not None else rem
                already += off
                new_v = cur_vols.get(dk, 0.0) + off
                orig  = orig_dest_vols.get(dk, 0.0)
                lo = orig * (1.0 - vol_tol); hi = orig * (1.0 + vol_tol)
                vol_pen += (max(0.0, lo - new_v) + max(0.0, new_v - hi)) * vol_rate

            # Plant window penalty - skip for yard-for destinations (24/7 parking)
            win_pen = 0.0
            avoid_pen = 0.0
            is_yard_dests = all("yard for" in (d.get("name","") or "").lower()
                                for d in dests_list if d.get("name"))
            if (plant_windows or avoid_windows) and not is_yard_dests:
                sname_s2 = state[s2][0]
                entry_s2 = self.cache.get(self.fname, {}).get(sname_s2, {})
                st_s2    = entry_s2.get("start_time") if isinstance(entry_s2, dict) else None
                if st_s2:
                    # Estimate time to reach first dest: drive from origin through
                    # all farms then to each dest in sequence.
                    cursor = datetime.combine(date.today(), st_s2)
                    for i in range(len(stops) - 1):
                        leg = lookup(self.dm, stops[i], stops[i+1])
                        if leg is None:
                            break
                        cursor += timedelta(minutes=(leg / DRIVE_SPEED_KMH) * 60.0)
                        stop = stops[i + 1]
                        dk_stop     = normalise_key(stop)
                        window      = plant_windows.get(dk_stop)
                        avoid_list  = avoid_windows.get(dk_stop)
                        if window is None and not avoid_list:
                            continue
                        arr_t = cursor.time()

                        if window is not None:
                            if not time_in_window(arr_t, window[0], window[1]):
                                open_t = parse_hhmm(window[0])
                                if open_t:
                                    open_dt = datetime.combine(date.today(), open_t)
                                    if open_t <= arr_t:
                                        open_dt += timedelta(days=1)
                                    wait_h = (open_dt - cursor).total_seconds() / 3600.0
                                else:
                                    wait_h = 1.0
                                win_pen += wait_h * plant_win_rate
                            elif plant_margin_mins > 0 and plant_margin_rate > 0:
                                close_t = parse_hhmm(window[1])
                                if close_t:
                                    close_dt = datetime.combine(date.today(), close_t)
                                    open_t2 = parse_hhmm(window[0])
                                    if open_t2 and close_t < open_t2:
                                        close_dt += timedelta(days=1)
                                    mins_to_close = (close_dt - cursor).total_seconds() / 60.0
                                    if 0 < mins_to_close < plant_margin_mins:
                                        depth = (plant_margin_mins - mins_to_close) / plant_margin_mins
                                        win_pen += depth * plant_margin_rate * (plant_margin_mins / 60.0)

                        if avoid_list and avoid_win_rate > 0:
                            for (av_s, av_e) in avoid_list:
                                if time_in_window(arr_t, av_s, av_e):
                                    avoid_pen += avoid_win_rate
                                    break

            return leg_km + vol_pen + win_pen + avoid_pen

        # flat list of all (s_idx, b_idx) slots
        all_slots = [(s_idx, b_idx)
                     for s_idx, (_sname, blocks) in enumerate(state)
                     for b_idx in range(len(blocks))]

        while pending:
            best_pi   = None
            best_regret = -float("inf")
            best_slot = None

            for pi, (_, _, dests_list) in enumerate(pending):
                costs = sorted(
                    ((_attach_cost(s2, b2, dests_list), (s2, b2))
                     for s2, b2 in all_slots
                     if not (state[s2][1][b2].get("dests") or
                             state[s2][1][b2].get("dest_key"))),
                    key=lambda x: x[0]
                )
                if not costs:
                    # Every slot already has dests - just pick cheapest overall
                    costs = sorted(
                        ((_attach_cost(s2, b2, dests_list), (s2, b2))
                         for s2, b2 in all_slots),
                        key=lambda x: x[0]
                    )
                if not costs:
                    continue
                regret = (costs[1][0] - costs[0][0]) if len(costs) >= 2 else 0.0
                if regret > best_regret:
                    best_regret = regret; best_pi = pi; best_slot = costs[0][1]

            if best_pi is None:
                break

            s_src, b_src, dests_list = pending.pop(best_pi)
            s2, b2 = best_slot
            block = state[s2][1][b2]
            # Subtract any volumes the slot already had before overwriting
            old_vols = {}
            ALNSSolver._acc_block_vols(block, old_vols)
            for dk, v in old_vols.items():
                cur_vols[dk] = cur_vols.get(dk, 0.0) - v
            block["dests"]     = [dict(d) for d in dests_list]
            block["dest_key"]  = dests_list[0].get("key","")  if dests_list else ""
            block["dest_name"] = dests_list[0].get("name","") if dests_list else ""
            # Add new volumes
            ALNSSolver._acc_block_vols(block, cur_vols)

        # Any still-pending items (no empty slots): put back on original block
        for s_src, b_src, dests_list in pending:
            block = state[s_src][1][b_src]
            block["dests"]     = [dict(d) for d in dests_list]
            block["dest_key"]  = dests_list[0].get("key","")  if dests_list else ""
            block["dest_name"] = dests_list[0].get("name","") if dests_list else ""

        return state

    # ══════════════════════════════════════════════════════════════════════════
    # DEST permutation & volume-split operators
    # ══════════════════════════════════════════════════════════════════════════

    def _shuffle_dests(self, state, orig_dest_vols):
        """
        Pick one random multi-dest block and try all permutations of its dest
        list.  Returns the state with the best permutation applied to that
        block only - SA in the main loop decides whether to accept.
        Single-block change keeps the move small and well-defined for SA.
        """
        from itertools import permutations as _perms
        vol_tol  = self.cfg.get("vol_tol", 0.15)
        vol_rate = self.cfg.get("vol_penalty", 1.0)

        # Collect all blocks with >=2 dests
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sn, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if len(block.get("dests") or []) >= 2
               and len(block.get("dests") or []) <= 5
        ]
        if not candidates:
            return _copy_state(state)

        s_idx, b_idx = random.choice(candidates)
        state = _copy_state(state)
        blocks   = state[s_idx][1]
        block    = blocks[b_idx]
        dests    = block["dests"]
        farm_keys = [r["irma"] for r in block["rows"]]
        farm_vol  = sum((r.get("prior_vol") or 0) for r in block["rows"]
                        if isinstance(r.get("prior_vol"), (int, float)))
        is_last   = (b_idx == len(blocks) - 1)
        origin    = ("VEDDER" if b_idx == 0 else
                     ((blocks[b_idx-1].get("dests") or [{}])[-1].get("key","") or "VEDDER"))

        # Group vol snapshot excluding this block
        base_vols = {}
        for si2, (_sn2, blks2) in enumerate(state):
            for bi2, blk2 in enumerate(blks2):
                if si2 == s_idx and bi2 == b_idx:
                    continue
                ALNSSolver._acc_block_vols(blk2, base_vols)

        def _perm_cost(perm):
            dest_keys = [d.get("key","") for d in perm if d.get("key")]
            stops = [origin] + farm_keys + dest_keys + (["VEDDER"] if is_last else [])
            km = sum(lookup(self.dm, stops[i], stops[i+1]) or 0.0
                     for i in range(len(stops)-1))
            vol_pen = 0.0; already = 0.0; vols = dict(base_vols)
            for d in perm:
                dk = d.get("key",""); vp = d.get("vol_partial")
                rem = max(0.0, farm_vol - already)
                off = min(float(vp), rem) if vp is not None else rem
                already += off
                new_v = vols.get(dk, 0.0) + off
                orig_v = orig_dest_vols.get(dk, 0.0)
                lo = orig_v * (1.0 - vol_tol); hi = orig_v * (1.0 + vol_tol)
                vol_pen += (max(0.0, lo-new_v) + max(0.0, new_v-hi)) * vol_rate
            return km + vol_pen

        best_perm = min(_perms(dests), key=_perm_cost)
        block["dests"]     = list(best_perm)
        block["dest_key"]  = block["dests"][0].get("key","")
        block["dest_name"] = block["dests"][0].get("name","")
        return state

    # ══════════════════════════════════════════════════════════════════════════
    # Intra-block resequencing operators  (2-opt and Or-opt)
    # ══════════════════════════════════════════════════════════════════════════

    def _two_opt_single(self, state):
        """Pick one random eligible block; find and apply its best 2-opt swap.

        2-opt reverses a sub-sequence of the farm list.  This reorders farms
        already assigned to a block, which the destroy/repair operators cannot
        do.  SA in the main loop decides whether to accept the result.

        Returns a candidate state (may be identical to input if no swap helps).
        """
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sn, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if not _is_preload_block(block) and not _is_fixed_vol_block(block)
            and len(block["rows"]) >= 2
        ]
        if not candidates:
            return _copy_state(state)

        s_idx, b_idx = random.choice(candidates)
        state = _copy_state(state)
        blocks = state[s_idx][1]
        block  = blocks[b_idx]
        rows   = block["rows"]
        n      = len(rows)
        origin = "VEDDER" if b_idx == 0 else (
            _block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"
        )

        base_cost = _route_km_simple(block, self.dm, origin=origin)
        best_rows = rows[:]
        best_cost = base_cost

        for i in range(n - 1):
            for j in range(i + 1, n):
                new_rows = rows[:i] + rows[i:j + 1][::-1] + rows[j + 1:]
                trial    = dict(block, rows=new_rows)
                c = _route_km_simple(trial, self.dm, origin=origin)
                if c < best_cost:
                    best_cost = c
                    best_rows = new_rows[:]

        state[s_idx][1][b_idx] = dict(block, rows=best_rows)
        return state

    def _or_opt_single(self, state, seg_len=2):
        """Pick one random eligible block; relocate the best segment of seg_len
        consecutive farms to a cheaper position in the same block.

        Or-opt with seg_len=2 moves pairs of farms; with seg_len=1 it is a
        simple relocation.  SA in the main loop decides whether to accept.
        """
        candidates = [
            (s_idx, b_idx)
            for s_idx, (_sn, blocks) in enumerate(state)
            for b_idx, block in enumerate(blocks)
            if not _is_preload_block(block) and not _is_fixed_vol_block(block)
            and len(block["rows"]) >= seg_len + 1
        ]
        if not candidates:
            return _copy_state(state)

        s_idx, b_idx = random.choice(candidates)
        state = copy.deepcopy(state)
        blocks = state[s_idx][1]
        block  = blocks[b_idx]
        rows   = block["rows"]
        n      = len(rows)
        origin = "VEDDER" if b_idx == 0 else (
            _block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"
        )

        base_cost = _route_km_simple(block, self.dm, origin=origin)
        best_rows = rows[:]
        best_cost = base_cost

        for i in range(n - seg_len + 1):
            seg  = rows[i:i + seg_len]
            rest = rows[:i] + rows[i + seg_len:]
            for j in range(len(rest) + 1):
                new_rows = rest[:j] + seg + rest[j:]
                trial    = dict(block, rows=new_rows)
                c = _route_km_simple(trial, self.dm, origin=origin)
                if c < best_cost:
                    best_cost = c
                    best_rows = new_rows[:]

        state[s_idx][1][b_idx] = dict(block, rows=best_rows)
        return state

    def _two_opt_all_blocks(self, state):
        """Exhaustive 2-opt polish applied to every block until convergence.

        Called once after the ALNS loop on best_state.  Guaranteed to return a
        solution at least as good as the input; never worsens.
        """
        state   = _copy_state(state)
        changed = True
        while changed:
            changed = False
            for s_idx, (sname, blocks) in enumerate(state):
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block):
                        continue
                    if _is_fixed_vol_block(block):
                        continue
                    rows   = block["rows"]
                    n      = len(rows)
                    if n < 2:
                        continue
                    origin = "VEDDER" if b_idx == 0 else (
                        _block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"
                    )
                    improved = True
                    while improved:
                        improved = False
                        base_cost = _route_km_simple(
                            dict(block, rows=rows), self.dm, origin=origin
                        )
                        for i in range(n - 1):
                            for j in range(i + 1, n):
                                new_rows = rows[:i] + rows[i:j + 1][::-1] + rows[j + 1:]
                                c = _route_km_simple(
                                    dict(block, rows=new_rows), self.dm, origin=origin
                                )
                                if c < base_cost - 0.001:
                                    rows      = new_rows
                                    base_cost = c
                                    improved  = True
                                    changed   = True
                    state[s_idx][1][b_idx] = dict(block, rows=rows)
                    # Keep local blocks reference in sync for subsequent b_idx origin lookups
                    blocks[b_idx] = state[s_idx][1][b_idx]
        return state

    # ══════════════════════════════════════════════════════════════════════════
    # Top-level roulette
    # ══════════════════════════════════════════════════════════════════════════

    def _roulette(self, scores, min_prob=0.0, floors=None, ceilings=None):
        """
        Weighted random selection from scores dict.

        floors (dict, optional): per-key minimum probability.  Takes precedence
          over min_prob.  Any key not in floors falls back to min_prob.

        ceilings (dict, optional): per-key maximum probability.  Keys exceeding
          their ceiling are clamped and the surplus is redistributed to keys
          below their ceiling.  Applied after floors.

        Both floors and ceilings can be combined.  The floor/ceiling pass
        iterates until stable to handle cascading interactions.
        """
        keys   = list(scores.keys())
        n      = len(keys)
        raw    = [max(scores[k], 1e-9) for k in keys]
        total  = sum(raw)
        probs  = [v / total for v in raw]

        # Build per-key floor and ceiling arrays
        key_floors   = [floors.get(k, min_prob) for k in keys]   if floors   else [min_prob] * n
        key_ceilings = [ceilings.get(k, 1.0)    for k in keys]   if ceilings else [1.0]      * n

        _apply_prob_floors_ceilings(probs, key_floors, key_ceilings)

        r   = random.random()
        cum = 0.0
        for k, p in zip(keys, probs):
            cum += p
            if r <= cum:
                return k
        return keys[-1]

    # ══════════════════════════════════════════════════════════════════════════
    # Main ALNS loop
    # ══════════════════════════════════════════════════════════════════════════

    def _solve_group(self, colour, sheets, total_iters, iter_offset):
        if not sheets:
            return {}
        try:
            return self._solve_group_inner(colour, sheets, total_iters, iter_offset)
        except Exception:
            import traceback
            tb = traceback.format_exc()
            self.log.emit(f"\n[{colour}] SOLVER CRASHED:\n{tb}")
            return {}

    def _solve_group_inner(self, colour, sheets, total_iters, iter_offset):

        dest_catalogue = _group_dest_catalogue(sheets)

        # initial state - start from mod_blocks if available (preserves _mwo
        # and _orig_arr flags set by the user); fall back to cache blocks.
        def _initial_blocks(sname, entry):
            mod_key = (self.fname, sname)
            if mod_key in self.sheet_mods:
                return copy.deepcopy(self.sheet_mods[mod_key])
            return copy.deepcopy(entry.get("blocks", []))

        # Filter out sheets that must not be touched by the solver FIRST,
        # then build orig_dest_vols only from sheets that are actually in the
        # solver's control.  Including skip/locked sheets in the target inflates
        # it with volumes the solver can never deliver, making targets impossible
        # and causing the vol penalty to dominate and never improve.
        all_skip = SOLVER_SKIP_SHEETS | self.locked_sheets
        locked_in_group = [sn for sn, _ in sheets if sn.strip() in self.locked_sheets]

        # Precompute locked sheets' fixed dock-visit times BEFORE they're
        # filtered out of `sheets` below.  Locked sheets never change during
        # this solve, so this only needs to happen once here rather than
        # inside the per-iteration cost evaluation - _group_overlap_penalty
        # reads self._locked_dest_visits to penalize an active route for
        # colliding with a locked truck's known, fixed dock time, even though
        # the locked truck itself is never touched.
        self._locked_dest_visits = {}
        skipped_entries = [(sn, e) for sn, e in sheets if sn.strip() in all_skip]
        for sn, e in skipped_entries:
            if not isinstance(e, dict):
                continue
            st_locked = e.get("start_time")
            blocks_locked = _initial_blocks(sn, e)
            for (dk, arr_m, dep_m) in self._extract_dest_visits(blocks_locked, st_locked):
                self._locked_dest_visits.setdefault(dk, []).append((arr_m, dep_m))

        sheets = [(sn, e) for sn, e in sheets
                  if sn.strip() not in all_skip]
        if locked_in_group:
            self.log.emit(f"  [{colour}] Locked (held constant): {locked_in_group}")
        if not sheets:
            return {}

        # -- Truck availability: precompute per-group parameters ---------------
        if self.cfg.get("truck_avail_enabled", False):
            night_mins = self._auto_night_start_mins()
            self.cfg["truck_avail_night_mins"] = night_mins
            # Day sheets in this group actually being optimised
            group_day = sum(
                1 for _, e in sheets
                if isinstance(e, dict) and _is_day_sheet(e.get("start_time"))
            )
            # All day sheets in the file (for proportional scaling across groups)
            total_day = sum(
                1 for e in self.cache.get(self.fname, {}).values()
                if isinstance(e, dict) and _is_day_sheet(e.get("start_time"))
            )
            min_back = self.cfg.get("truck_avail_min_back", 8)
            group_needed = (max(0, round(min_back * group_day / total_day))
                           if total_day > 0 else 0)
            self.cfg["_truck_avail_group_needed"] = group_needed
            if night_mins is not None and group_day > 0:
                self.log.emit(
                    f"  [{colour}] Truck avail: night starts "
                    f"{night_mins // 60:02d}:{night_mins % 60:02d}, "
                    f"need {group_needed}/{group_day} day trucks back in this group"
                )
            elif group_day == 0:
                self.log.emit(f"  [{colour}] Truck avail: no day shifts in this group")
        else:
            self.cfg["_truck_avail_group_needed"] = 0
        # Farms with prior_vol == 0 (or falsy numeric) contribute nothing to
        # collection volume and skew the solver's cost function.  They are
        # -- Adjacent same-IRMA farm pairing -----------------------------------
        # Any run of two or more farms that share the same IRMA AND are
        # adjacent in the original sheet (e.g. T1 / T2 trailers at the same
        # farm) is a locked unit: the solver must keep them together, in the
        # same block, adjacent, and in their original order.  This holds
        # regardless of volume - both trailers may carry milk, or one may be
        # zero-vol; either way they travel as one.
        #
        # Strategy: keep only the LEAD farm of each run in the working state
        # the solver sees; strip the FOLLOWERS (every farm after the first in
        # the run) before solving, recording the lead's _uid.  After solving,
        # re-insert each follower immediately after its lead - wherever the
        # solver moved the lead to - preserving original run order.
        #
        # paired_followers: list of (sname, b_idx, lead_uid, follower_dict,
        #                            order_within_run)
        # paired_followers entries: (sname, b_idx, lead_uid, follower_dict, order)
        paired_followers = []
        # lead_uid -> original lead prior_vol (before folding followers in),
        # so we can restore it on reinsert.
        lead_orig_vol = {}

        def _strip_pairs(sname, blocks):
            """Strip followers of adjacent same-IRMA runs, keeping the lead.

            Critically, the follower's volume is FOLDED INTO the lead while
            the followers are sidelined.  The solver only sees the lead during
            optimisation, so if we left the lead carrying only its own volume,
            the capacity and volume-balance penalties would undercount the
            real load of the pair and the solver would happily place the pair
            on a route that actually overflows once the followers are added
            back.  By making the lead temporarily carry the COMBINED pair
            volume, every capacity/balance decision the solver makes reflects
            the true load.  Original volumes are restored on reinsert.
            """
            stripped = []
            for b_idx, block in enumerate(blocks):
                rows = block.get("rows", [])
                new_rows = []
                i = 0
                while i < len(rows):
                    lead = rows[i]
                    lead_irma = lead.get("irma", "")
                    lead_uid  = lead.get("_uid")
                    # Gather the adjacent run of identical IRMAs after the lead
                    j = i + 1
                    order = 1
                    folded_vol = 0.0
                    have_followers = False
                    while (j < len(rows)
                           and rows[j].get("irma", "") == lead_irma
                           and lead_irma != ""):
                        have_followers = True
                        follower = copy.deepcopy(rows[j])
                        fv = follower.get("prior_vol")
                        if isinstance(fv, (int, float)):
                            folded_vol += fv
                        paired_followers.append(
                            (sname, b_idx, lead_uid, follower, order))
                        order += 1
                        j += 1
                    if have_followers:
                        # Fold follower volume into a COPY of the lead so the
                        # solver sees the combined load; remember the original
                        # so we can restore it on reinsert.
                        lead_copy = copy.deepcopy(lead)
                        lv = lead_copy.get("prior_vol")
                        lv = lv if isinstance(lv, (int, float)) else 0.0
                        if lead_uid is not None:
                            lead_orig_vol[lead_uid] = lead.get("prior_vol")
                        lead_copy["prior_vol"] = lv + folded_vol
                        new_rows.append(lead_copy)
                    else:
                        new_rows.append(lead)
                    i = j
                stripped.append(dict(block, rows=new_rows))
            return stripped

        def _reinsert_pairs(result):
            """Re-insert followers immediately after their lead in the solved state."""
            if not paired_followers:
                return result
            # First restore every lead's original (un-folded) volume.
            for sname, blocks in result.items():
                for block in blocks:
                    for farm in block.get("rows", []):
                        uid = farm.get("_uid")
                        if uid in lead_orig_vol:
                            farm["prior_vol"] = lead_orig_vol[uid]
            # Build a uid -> (sname, b_idx, f_idx) map across the solved state
            uid_loc = {}
            for sname, blocks in result.items():
                for b_idx, block in enumerate(blocks):
                    for f_idx, farm in enumerate(block.get("rows", [])):
                        uid = farm.get("_uid")
                        if uid:
                            uid_loc[uid] = (sname, b_idx, f_idx)

            # Insert followers in ascending run-order so that order 1 lands
            # right after the lead, order 2 right after order 1, etc.
            for _sname, _b_idx, lead_uid, follower, order in sorted(
                    paired_followers, key=lambda t: t[4]):
                if lead_uid and lead_uid in uid_loc:
                    sname, b_idx, f_idx = uid_loc[lead_uid]
                    # f_idx is the lead's current position; insert after the
                    # last already-placed member of this run.  Walk forward
                    # past any followers of the same lead already inserted.
                    insert_at = f_idx + 1
                    rows = result[sname][b_idx]["rows"]
                    while (insert_at < len(rows)
                           and rows[insert_at].get("irma", "")
                               == follower.get("irma", "")):
                        insert_at += 1
                    rows.insert(insert_at, follower)
                    # Shift cached indices at/after the insertion point
                    for uid, (s, b, fi) in list(uid_loc.items()):
                        if s == sname and b == b_idx and fi >= insert_at:
                            uid_loc[uid] = (s, b, fi + 1)
                    fu = follower.get("_uid")
                    if fu:
                        uid_loc[fu] = (sname, b_idx, insert_at)
                else:
                    # Lead not found (shouldn't happen) - fall back to original
                    sname = _sname
                    if sname in result and _b_idx < len(result[sname]):
                        result[sname][_b_idx]["rows"].append(follower)
            return result

        # Build volume targets from the non-skipped sheets only
        orig_dest_vols = {}
        for _sname, entry in sheets:
            for block in entry.get("blocks", []):
                for dk, off in _block_dest_offloads(block).items():
                    orig_dest_vols[dk] = orig_dest_vols.get(dk, 0.0) + off

        state = [(sname, _strip_pairs(sname, _initial_blocks(sname, entry)))
                 for sname, entry in sheets]

        if paired_followers:
            pairs_desc = ", ".join(
                f"{f['irma']}" for _, _, _, f, _ in paired_followers)
            self.log.emit(
                f"  [{colour}] {len(paired_followers)} paired trailer(s) held "
                f"with their lead (solver keeps adjacent same-IRMA farms "
                f"together): {pairs_desc}")

        # Optimize split positions for partial-dropoff dests on initial state
        if self.cfg.get("split_opt", False):
            for sname, blocks in state:
                entry = self.cache.get(self.fname, {}).get(sname, {})
                st    = entry.get("start_time") if isinstance(entry, dict) else None
                if st:
                    _optimize_split_positions(blocks, self.dm, st, self.cfg, dm_dur=self.dm_dur)

        cfg_no_win = dict(self.cfg, plant_win_penalty=0.0, plant_windows={},
                          win_miss_penalty=0.0, cap_penalty=0.0)

        # Compute frozen block cost offset BEFORE using it in best_cost
        frozen_cost_offset = 0.0
        frozen_cost_offset_no_win = 0.0
        for sname, blocks in state:
            entry = self.cache.get(self.fname, {}).get(sname, {})
            st    = entry.get("start_time") if isinstance(entry, dict) else None
            for block in blocks:
                if _is_preload_block(block) or _is_fixed_vol_block(block):
                    frozen_cost_offset        += _sheet_cost([block], self.dm, st, self.cfg, dm_dur=self.dm_dur)
                    frozen_cost_offset_no_win += _sheet_cost([block], self.dm, st, cfg_no_win, dm_dur=self.dm_dur)

        best_state = _copy_state(state)
        cur_sheet_cache  = self._make_sheet_cost_cache(state)
        best_cost  = sum(cur_sheet_cache.values()) + _group_vol_penalty(state, orig_dest_vols, self.cfg) + self._group_overlap_penalty(state) - frozen_cost_offset
        cur_cost   = best_cost

        cost_no_win = sum(
            _sheet_cost(blocks, self.dm,
                        (self.cache.get(self.fname, {}).get(sname, {}) or {}).get("start_time"),
                        cfg_no_win, dm_dur=self.dm_dur)
            for sname, blocks in state
        ) + _group_vol_penalty(state, orig_dest_vols, self.cfg)
        cost_no_win -= frozen_cost_offset_no_win
        T0               = max(cost_no_win * 0.05, 10.0)
        T                = T0
        target_cool_frac = self.cfg.get("target_cool_frac", 0.001)
        if total_iters > 1 and target_cool_frac > 0:
            alpha = target_cool_frac ** (1.0 / total_iters)
        else:
            alpha = 0.9999

        # Move-type scores with hard probability floors.
        #
        # The adaptive roulette learns which moves improve cost and rewards them.
        # Without a floor, farm/combined moves get penalised whenever they fail
        # to beat the plant-window penalty (which dominates early on), causing
        # intra-block 2-opt/or-opt to dominate by iteration 200 - exactly the
        # behaviour we observed.  Hard floors guarantee a minimum budget for
        # cross-route moves regardless of adaptive history.
        #
        # Floors (min_prob passed to _roulette):
        #   farm:     0.25  - guaranteed 25% of iterations move farms cross-route
        #   combined: 0.15
        #   dest:     0.10
        #   2opt:     0.05
        #   or_opt:   0.05
        FARM_FLOOR     = 0.25
        COMBINED_FLOOR = 0.15
        DEST_FLOOR     = 0.10
        INTRA_FLOOR    = 0.05
        MOVE_FLOORS = {
            "farm":     FARM_FLOOR,
            "combined": COMBINED_FLOOR,
            "dest":     DEST_FLOOR,
            "2opt":     INTRA_FLOOR,
            "or_opt":   INTRA_FLOOR,
        }
        # Ceilings prevent intra-block ops from hoarding probability even when
        # their adaptive scores are high (they improve easily but do nothing
        # for cross-route optimisation).
        MOVE_CEILINGS = {
            "2opt":   0.12,
            "or_opt": 0.12,
        }
        OP_MIN_PROB   = 0.25   # floor for destroy/repair sub-roulette (2 options each)

        move_scores = {
            "farm":     3.0,   # primary cross-route move
            "combined": 3.0,   # cross-route + dest reassignment
            "dest":     1.5,   # dest-only reassignment
            "2opt":     0.5,   # intra-block only - useful but deprioritised
            "or_opt":   0.5,
        }

        # Farm sub-operator scores
        d_scores = {"random": 1.0, "worst": 1.0}
        r_scores = {"best":   1.0, "regret": 1.0}

        DECAY    = 0.97
        REWARD   = 2.0
        ACCEPT_REWARD = 0.5
        seg_size = self.cfg.get("segment_size", 100)

        n_farms  = sum(len(b["rows"]) for _, blocks in state for b in blocks)
        n_blocks = sum(len(blocks)    for _, blocks in state)
        n_remove_dests = max(1, min(int(n_blocks * 0.20), 5))

        self.log.emit(
            f"[{colour}]  sheets={len(sheets)}  farms={n_farms}  "
            f"blocks={n_blocks}  cost0={best_cost:.1f}  (cost_no_win={cost_no_win:.1f})\n"
            f"          alpha={alpha:.6f}  T0={T0:.1f} (from cost_no_win)  "
            f"cool_target={target_cool_frac*100:.4f}%\n"
            f"          remove/move: farms=T-scaled[3->30]  dests={n_remove_dests}  "
            f"floors: farm={FARM_FLOOR:.0%} combined={COMBINED_FLOOR:.0%} "
            f"dest={DEST_FLOOR:.0%} intra={INTRA_FLOOR:.0%}"
        )

        iters_no_improve  = 0
        diag_accepted  = {k: 0 for k in move_scores}
        diag_tried     = {k: 0 for k in move_scores}
        diag_cross_route = 0

        for it in range(total_iters):
            if self._stop:
                break

            # Yield the GIL periodically so the UI thread stays responsive
            if it % 50 == 0:
                import time as _t; _t.sleep(0)

            # Decay all scores each segment - just multiply, no hard floor
            # on raw scores since the probability floor handles balance.
            if it % seg_size == 0 and it > 0:
                for d in (move_scores, d_scores, r_scores):
                    for k in d:
                        d[k] = max(0.01, d[k] * DECAY)

            move_type = self._roulette(move_scores, floors=MOVE_FLOORS,
                                        ceilings=MOVE_CEILINGS)

            # n_remove scales with temperature: large disruptions when hot
            # (broad exploration), small when cold (fine-tuning).
            # Range: [3, 30] linearly interpolated by T/T0.
            t_frac = min(1.0, T / T0) if T0 > 0 else 0.0
            n_remove_farms = max(3, int(3 + (30 - 3) * t_frac))

            # Track which sheets are modified so we only recompute their costs.
            changed_sheets = set()

            # -- execute chosen move -------------------------------------------
            if move_type == "farm":
                d_op = self._roulette(d_scores, min_prob=OP_MIN_PROB)
                r_op = self._roulette(r_scores, min_prob=OP_MIN_PROB)
                if d_op == "random":
                    new_state, removed = self._destroy_random(state, n_remove_farms)
                else:
                    new_state, removed = self._destroy_worst(state, n_remove_farms)
                changed_sheets = {state[s][0] for s, _, _ in removed}
                if r_op == "best":
                    new_state, _cr = self._repair_best(new_state, removed)
                else:
                    new_state, _cr = self._repair_regret(new_state, removed)
                # Repair may place farms on any sheet - track destinations too
                for s_idx, (sn, _) in enumerate(new_state):
                    if any(sn == state[s][0] for s, _, _ in removed):
                        changed_sheets.add(sn)
                    # Also detect sheets that gained farms from other routes
                    orig_count = sum(len(b["rows"]) for b in state[s_idx][1])
                    new_count  = sum(len(b["rows"]) for b in new_state[s_idx][1])
                    if orig_count != new_count:
                        changed_sheets.add(sn)
                diag_cross_route += _cr

            elif move_type == "dest":
                new_state, stripped = self._destroy_dest(state, n_remove_dests,
                                                          orig_dest_vols)
                new_state = self._repair_dest(new_state, stripped,
                                              dest_catalogue, orig_dest_vols)
                changed_sheets = {state[s][0] for s, _, _ in stripped}

            elif move_type == "2opt":
                new_state = self._two_opt_single(state)
                # 2-opt touches one block on one sheet - detect by row count change
                for s_idx, (sn, blocks) in enumerate(new_state):
                    if blocks != state[s_idx][1]:
                        changed_sheets.add(sn)

            elif move_type == "or_opt":
                seg = 2 if random.random() < 0.7 else 3
                new_state = self._or_opt_single(state, seg_len=seg)
                for s_idx, (sn, blocks) in enumerate(new_state):
                    if blocks != state[s_idx][1]:
                        changed_sheets.add(sn)

            else:  # combined
                d_op = self._roulette(d_scores, min_prob=OP_MIN_PROB)
                r_op = self._roulette(r_scores, min_prob=OP_MIN_PROB)
                if d_op == "random":
                    new_state, removed = self._destroy_random(state, n_remove_farms)
                else:
                    new_state, removed = self._destroy_worst(state, n_remove_farms)
                changed_sheets = {state[s][0] for s, _, _ in removed}
                new_state, stripped = self._destroy_dest(new_state, n_remove_dests,
                                                          orig_dest_vols)
                changed_sheets |= {state[s][0] for s, _, _ in stripped}
                if r_op == "best":
                    new_state, _cr = self._repair_best(new_state, removed)
                else:
                    new_state, _cr = self._repair_regret(new_state, removed)
                for s_idx, (sn, _) in enumerate(new_state):
                    orig_count = sum(len(b["rows"]) for b in state[s_idx][1])
                    new_count  = sum(len(b["rows"]) for b in new_state[s_idx][1])
                    if orig_count != new_count:
                        changed_sheets.add(sn)
                diag_cross_route += _cr
                new_state = self._repair_dest(new_state, stripped,
                                              dest_catalogue, orig_dest_vols)

            # Recompute costs only for changed sheets; reuse cache for the rest.
            new_sheet_cache = dict(cur_sheet_cache)
            for s_idx, (sn, blocks) in enumerate(new_state):
                if sn in changed_sheets:
                    entry = self.cache[self.fname].get(sn, {})
                    st    = entry.get("start_time") if isinstance(entry, dict) else None
                    new_sheet_cache[sn] = _sheet_cost(blocks, self.dm, st, self.cfg, dm_dur=self.dm_dur)
            new_cost = sum(new_sheet_cache.values()) + _group_vol_penalty(new_state, orig_dest_vols, self.cfg) + self._group_overlap_penalty(new_state) - frozen_cost_offset
            delta    = new_cost - cur_cost

            # SA acceptance - compute probability explicitly for diagnostics
            if delta < 0:
                accept_prob = 1.0
            elif T > 1e-12:
                accept_prob = math.exp(-delta / T)
            else:
                accept_prob = 0.0
            accepted = accept_prob >= 1.0 or random.random() < accept_prob

            diag_tried[move_type] = diag_tried.get(move_type, 0) + 1

            if accepted:
                diag_accepted[move_type] = diag_accepted.get(move_type, 0) + 1
                state           = new_state
                cur_cost        = new_cost
                cur_sheet_cache = new_sheet_cache

                # Re-optimize split positions on any changed sheet (if enabled)
                split_changed = False
                if self.cfg.get("split_opt", False):
                    for s_idx, (sname, blocks) in enumerate(state):
                        if sname in changed_sheets:
                            entry = self.cache.get(self.fname, {}).get(sname, {})
                            st    = entry.get("start_time") if isinstance(entry, dict) else None
                            if st and _optimize_split_positions(blocks, self.dm, st, self.cfg, dm_dur=self.dm_dur):
                                split_changed = True
                                new_sheet_cache[sname] = _sheet_cost(blocks, self.dm, st, self.cfg, dm_dur=self.dm_dur)
                if split_changed:
                    cur_cost        = sum(new_sheet_cache.values()) + _group_vol_penalty(state, orig_dest_vols, self.cfg) + self._group_overlap_penalty(state) - frozen_cost_offset
                    cur_sheet_cache = new_sheet_cache

                improved = cur_cost < best_cost
                if improved:
                    reward = REWARD
                    best_cost        = cur_cost
                    best_state       = _copy_state(state)
                    iters_no_improve = 0
                else:
                    reward = ACCEPT_REWARD
                    iters_no_improve += 1
                move_scores[move_type] = move_scores[move_type] * DECAY + reward
                if move_type in ("farm", "combined"):
                    d_scores[d_op] = d_scores[d_op] * DECAY + reward
                    r_scores[r_op] = r_scores[r_op] * DECAY + reward
            else:
                iters_no_improve += 1

            T *= alpha

            # -- Every 50 iterations: detailed diagnostic output --------------
            if it % 50 == 49:
                bd_best = _sheet_cost_breakdown_state(best_state, self.dm, self.cache, self.fname, self.cfg, dm_dur=self.dm_dur)
                vol_pen_d = _group_vol_penalty(best_state, orig_dest_vols, self.cfg)
                overlap_pen_d = self._group_overlap_penalty(best_state)

                # Per-processor vol deviation (uses the same offload accounting
                # as the volume penalty so the displayed numbers always match).
                dest_vols_d = {}
                for _, blocks_d in best_state:
                    for block_d in blocks_d:
                        for dk_d, off_d in _block_dest_offloads(block_d).items():
                            dest_vols_d[dk_d] = dest_vols_d.get(dk_d, 0.0) + off_d

                vol_tol_d = self.cfg.get("vol_tol", 0.15)
                vol_lines = []
                for dk_d, orig_v in sorted(orig_dest_vols.items()):
                    cur_v = dest_vols_d.get(dk_d, 0.0)
                    pct = (cur_v / orig_v * 100) if orig_v else 0
                    flag = " (!)" if abs(cur_v - orig_v) > orig_v * vol_tol_d else ""
                    vol_lines.append(
                        f"      {dk_d}: {cur_v:>8,.0f} / {orig_v:>8,.0f} L  ({pct:5.1f}%){flag}")

                t50 = T * math.log(2)
                t10 = T * math.log(10)
                acc_parts = []
                for k in move_scores:
                    tried = diag_tried.get(k, 0)
                    acc   = diag_accepted.get(k, 0)
                    acc_parts.append(f"{k}={acc}/{tried}" if tried else f"{k}=-")

                self.log.emit(
                    f"  [{colour}] -- it={it+1} --  best={best_cost:.1f}\n"
                    f"    Cost breakdown: km={bd_best['km']:.1f}  shift={bd_best['shift']:.1f}"
                    f"  shift_pen={bd_best['overtime']:.1f}  shortfall={bd_best['shortfall']:.1f}  milking={bd_best['milking']:.1f}"
                    f"  cap={bd_best['cap']:.1f}  vol_pen={vol_pen_d:.1f}  overlap_pen={overlap_pen_d:.1f}\n"
                    f"    T={T:.2f}  50% accept if deltaZ<{t50:.1f}  10% if deltaZ<{t10:.1f}\n"
                    f"    Accepted/tried (last 50): {('  '.join(acc_parts))}\n"
                    f"    Cross-route placements (last 50 farm/combined moves): {diag_cross_route}\n"
                    f"    Processor volumes:\n" + "\n".join(vol_lines)
                )
                diag_accepted    = {k: 0 for k in move_scores}
                diag_tried       = {k: 0 for k in move_scores}
                diag_cross_route = 0

            # progress report
            if it % 25 == 0 or it == total_iters - 1:
                global_it = iter_offset + it
                total_all = iter_offset + total_iters
                self.progress.emit(
                    global_it, total_all,
                    f"{colour}: {it+1}/{total_iters}  "
                    f"best={best_cost:.1f}  cur={cur_cost:.1f}  T={T:.4f}"
                )
                if it % 100 == 0:
                    # Show effective probabilities using the same floor/ceiling
                    # redistribution _roulette applies, so the display matches.
                    _raw = [max(move_scores[k], 1e-9) for k in move_scores]
                    _tot = sum(_raw)
                    _keys = list(move_scores.keys())
                    _pv   = [m / _tot for m in _raw]
                    _fl   = [MOVE_FLOORS.get(k, 0.0) for k in _keys]
                    _cl   = [MOVE_CEILINGS.get(k, 1.0) for k in _keys]
                    _apply_prob_floors_ceilings(_pv, _fl, _cl)
                    _pt = sum(_pv)
                    prob_str = "  ".join(f"{k}={_pv[i]/_pt:.0%}"
                                         for i, k in enumerate(_keys))
                    self.log.emit(f"  [{colour}] it={it+1:4d}  best={best_cost:.1f}"
                                  f"  T={T:.4f}  prob: {prob_str}")

        def _state_vol(st):
            return sum(
                (f.get("prior_vol") or 0)
                for _, blks in st for b in blks
                for f in b["rows"]
                if isinstance(f.get("prior_vol"), (int, float))
            )
        input_farms = sum(len(b["rows"]) for _, blks in state for b in blks)
        input_vol   = _state_vol(state)

        output_farms = sum(len(b["rows"]) for _, blocks in best_state for b in blocks)
        output_vol   = _state_vol(best_state)
        farm_ok = "OK" if output_farms == input_farms else f"(!) LOST {input_farms - output_farms}"
        vol_ok  = "OK" if abs(output_vol - input_vol) < 1 else f"(!) LOST {input_vol - output_vol:,.0f}L"

        # Full cost breakdown for the final best state
        _km_f = _shift_f = _shift_pen_f = _shift_under_f = _milking_f = _cap_f = 0.0
        for sname_f, blocks_f in best_state:
            entry_f = self.cache.get(self.fname, {}).get(sname_f, {})
            st_f    = entry_f.get("start_time") if isinstance(entry_f, dict) else None
            for dists_f in calc_distances(blocks_f, self.dm):
                for d_f in dists_f[:-1]:
                    if d_f is not None: _km_f += d_f
            if st_f:
                ct_f = calc_times(blocks_f, self.dm, st_f,
                                  suppress_no_milking=self.cfg.get("suppress_no_milking", True))
                if ct_f:
                    _base_f = datetime.combine(date.today(), st_f)
                    _sh_f   = (ct_f[1] - _base_f).total_seconds() / 3600.0
                    _max_sh = self.cfg.get("max_shift_h", 12.0)
                    _min_sh = self.cfg.get("min_shift_h", 8.0)
                    _shift_f       += _sh_f * self.cfg.get("shift_hours_weight", 0.0)
                    _shift_pen_f   += max(0.0, _sh_f - _max_sh) * self.cfg.get("shift_penalty", 200.0)
                    _shift_under_f += max(0.0, _min_sh - _sh_f) * self.cfg.get("shift_under_penalty", 30.0)
                    for b_f, block_f in enumerate(blocks_f):
                        bt_f = ct_f[0][b_f] if b_f < len(ct_f[0]) else None
                        if not bt_f: continue
                        for fi_f, farm_f in enumerate(block_f["rows"]):
                            ft_f = bt_f[fi_f+1] if (fi_f+1) < len(bt_f) else None
                            if ft_f and ft_f.get("wait"):
                                _milking_f += ft_f["wait"] * self.cfg.get("milking_weight", 1.0)
            _hc_f = self.cfg.get("hard_vol_cap", HARD_CAP)
            _cr_f = self.cfg.get("cap_penalty", 2.0)
            for block_f in blocks_f:
                _rv_f = sum((r.get("prior_vol") or 0) for r in block_f["rows"]
                            if isinstance(r.get("prior_vol"), (int, float)))
                if _rv_f > _hc_f:
                    _cap_f += (_rv_f - _hc_f) * _cr_f
        _vol_pen_f   = _group_vol_penalty(best_state, orig_dest_vols, self.cfg)
        _plant_win_f = (best_cost - _km_f - _shift_f - _shift_pen_f - _shift_under_f
                        - _milking_f - _cap_f - _vol_pen_f)

        self.log.emit(
            f"[{colour}] Done - best={best_cost:.1f}  "
            f"(started {self._group_cost([(sname, copy.deepcopy(entry.get('blocks',[]))) for sname,entry in sheets], orig_dest_vols):.1f})\n"
            f"  farms: {input_farms}->{output_farms} {farm_ok}  "
            f"vol: {input_vol:,.0f}->{output_vol:,.0f}L {vol_ok}\n"
            f"  Cost breakdown:\n"
            f"    km={_km_f:.1f}  milking={_milking_f:.1f}  shift={_shift_f:.1f}"
            f"  overtime={_shift_pen_f:.1f}  shortfall={_shift_under_f:.1f}  cap={_cap_f:.1f}"
            f"  plant_win={_plant_win_f:.1f}  vol_pen={_vol_pen_f:.1f}"
        )

        # -- Farm displacement summary -----------------------------------------
        # Compare each farm's original (sname, b_idx) placement against the
        # solved best_state.  A farm is identified by its _uid if available,
        # otherwise by (irma, prior_vol) as a fallback key.
        def _farm_key(farm):
            uid = farm.get("_uid")
            if uid:
                return uid
            return (farm.get("irma",""), farm.get("prior_vol", 0))

        # Build original placement map: farm_key -> (sname, b_idx)
        orig_placement = {}
        for sname, entry in sheets:
            for b_idx, block in enumerate(entry.get("blocks", [])):
                for farm in block.get("rows", []):
                    orig_placement[_farm_key(farm)] = (sname, b_idx)

        # Build solved placement map: farm_key -> (sname, b_idx)
        solved_placement = {}
        for sname, blocks in best_state:
            for b_idx, block in enumerate(blocks):
                for farm in block.get("rows", []):
                    solved_placement[_farm_key(farm)] = (sname, b_idx)

        n_total = len(orig_placement)
        n_diff_block  = 0  # different block (may be same route)
        n_diff_route  = 0  # different route entirely
        for key, (orig_sname, orig_b) in orig_placement.items():
            solved = solved_placement.get(key)
            if solved is None:
                continue
            sol_sname, sol_b = solved
            if sol_sname != orig_sname:
                n_diff_route += 1
                n_diff_block += 1
            elif sol_b != orig_b:
                n_diff_block += 1

        pct_block = n_diff_block / n_total * 100 if n_total else 0
        pct_route = n_diff_route / n_total * 100 if n_total else 0
        self.log.emit(
            f"  [{colour}] Farm displacement vs original:\n"
            f"    Different block (incl. different route): {n_diff_block:3d} / {n_total}  ({pct_block:.1f}%)\n"
            f"    Different route only:                   {n_diff_route:3d} / {n_total}  ({pct_route:.1f}%)"
        )

        # Post-solve: apply exhaustive 2-opt to every block in best_state.
        # This is deterministic, never worsens the solution, and recovers
        # any intra-block ordering improvements the ALNS loop may have missed.
        self.log.emit(f"[{colour}] Applying 2-opt polish pass...")
        best_state_list = [(sn, blks) for sn, blks in best_state]
        polished = self._two_opt_all_blocks(best_state_list)
        polished_cost = self._group_cost(polished, orig_dest_vols)
        if polished_cost < best_cost:
            self.log.emit(
                f"[{colour}] 2-opt improved {best_cost:.1f} -> {polished_cost:.1f}"
            )
            best_state = polished
            best_cost  = polished_cost

        # Post-solve: exhaustively optimise destination order on every multi-dest
        # block.  _shuffle_dests was previously a stochastic SA move but is
        # deterministic - it always returns the best permutation.  Running it
        # in the SA loop just burned iterations for free improvements and inflated
        # the adaptive scores of non-exploratory moves.  Here it runs once over
        # every eligible block after the search is complete.
        self.log.emit(f"[{colour}] Applying destination-order polish pass...")
        shuffled = _copy_state(best_state)
        for s_idx, (sn, blocks) in enumerate(shuffled):
            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block) or _is_fixed_vol_block(block):
                    continue
                dests = block.get("dests") or []
                if len(dests) < 2 or len(dests) > 5:
                    continue
                # _shuffle_dests picks a random block - call it directly per block
                trial_state = shuffled
                improved_state = self._shuffle_dests(trial_state, orig_dest_vols)
                # Only keep if it improved
                shuffled_cost = self._group_cost(improved_state, orig_dest_vols)
                if shuffled_cost < best_cost:
                    shuffled = improved_state
                    best_cost = shuffled_cost
        shuffled_cost = self._group_cost(shuffled, orig_dest_vols)
        if shuffled_cost < self._group_cost([(sn, blks) for sn, blks in best_state],
                                             orig_dest_vols):
            self.log.emit(
                f"[{colour}] Dest-order polish improved -> {shuffled_cost:.1f}"
            )
            best_state = shuffled

        # Accumulate paired followers for logging; they are already re-inserted
        # into the result below so they appear on their original route sheets.
        self.paired_followers.extend(paired_followers)

        result = {sname: blocks for sname, blocks in best_state}
        _reinsert_pairs(result)
        return result

    def run(self):
        fname  = self.fname
        cache  = self.cache
        # Seed the RNG when the user supplied one, so a given (input, settings,
        # seed) triple reproduces exactly. None -> leave the global RNG alone.
        seed = self.cfg.get("seed")
        if seed is not None:
            random.seed(seed)
            self.log.emit(f"Random seed: {seed} (reproducible run)")
        groups = _group_sheets_by_colour(cache, fname)
        iters  = self.cfg.get("iterations", 300)

        # Log holdover blocks found in the file so the user can verify them
        holdover_info = []
        for colour_g, sheet_list in groups.items():
            for sname, entry in sheet_list:
                for bi, block in enumerate(entry.get("blocks", [])):
                    dnames = [d.get("name","?") for d in (block.get("dests") or [])]
                    if _is_preload_block(block):
                        holdover_info.append(
                            f"  [{colour_g}] Sheet {sname} block {bi+1}: "
                            f"PRELOAD (fully frozen) -> {', '.join(dnames)}")
                    elif _is_holdover_block(block):
                        holdover_info.append(
                            f"  [{colour_g}] Sheet {sname} block {bi+1}: "
                            f"holdover (farms free, dest fixed) -> {', '.join(dnames)}")
                    elif _is_fixed_vol_block(block):
                        vps = ", ".join(
                            f"{d.get('name','?')}={d.get('vol_partial',0):,.0f}L"
                            for d in (block.get("dests") or [])
                        )
                        holdover_info.append(
                            f"  [{colour_g}] Sheet {sname} block {bi+1}: "
                            f"fixed-vol (fully frozen) -> {vps}")
        if holdover_info:
            self.log.emit("Preload / holdover blocks detected:\n" +
                          "\n".join(holdover_info))
        else:
            self.log.emit("No preload or holdover blocks found.")

        red_result  = self._solve_group("RED",  groups["RED"],
                                        iters, iter_offset=0)
        blue_result = self._solve_group("BLUE", groups["BLUE"],
                                        iters, iter_offset=iters)

        results = {}
        results.update(red_result)
        results.update(blue_result)

        # -- HiGHS post-optimality processor assignment check -----------------
        self.log.emit("\nRunning HiGHS processor-assignment verification...")
        for colour, sheets in [("RED", groups["RED"]), ("BLUE", groups["BLUE"])]:
            if not sheets:
                continue
            # Build state list from the solved results
            solved_state = [(sn, results[sn]) for sn, _e in sheets if sn in results]
            if not solved_state:
                continue
            summary = _highs_verify_processor_assignment(
                colour, sheets, solved_state, self.dm, self.cfg, self.log.emit)
            self.log.emit(summary)

        self.finished.emit(results)


# -- IRMA Lookup Dialog --------------------------------------------------------

class IRMALookupDialog(QDialog):
    """Cross-file, cross-sheet IRMA farm and processor lookup.

    Given a full or partial IRMA number or processor key, searches every loaded
    file and sheet (both original parsed data and solver-modified routes) and
    lists every block that contains a match.  Double-clicking a result navigates
    the main window to that file/sheet.

    Results table columns:
        File | Sheet | Route | Block # | Type | IRMA / Key | Source
    where Source is 'Original', 'Modified', or 'Both'.
    """

    navigate_requested = pyqtSignal(str, str)

    def __init__(self, cache, sheet_mods, parent=None):
        super().__init__(parent)
        self.setWindowTitle("IRMA / Processor Lookup")
        self.setMinimumSize(880, 480)
        self._cache      = cache
        self._sheet_mods = sheet_mods
        self._results    = []

        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 10, 10, 10)

        # -- Search bar -------------------------------------------------------
        bar = QHBoxLayout()
        bar.addWidget(QLabel("IRMA / Processor:"))
        self._query = QLineEdit()
        self._query.setPlaceholderText(
            "Farm IRMA (71-117) or processor key (905011) - partial match OK")
        self._query.setMinimumWidth(220)
        bar.addWidget(self._query, stretch=1)

        self._search_btn = QPushButton("Search")
        self._search_btn.setFixedHeight(28)
        self._search_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:3px; padding: 0 12px; } "
            "QPushButton:disabled { background:#90caf9; }")
        bar.addWidget(self._search_btn)

        self._chk_orig = QCheckBox("Original")
        self._chk_orig.setChecked(True)
        self._chk_mod  = QCheckBox("Modified")
        self._chk_mod.setChecked(True)
        bar.addWidget(self._chk_orig)
        bar.addWidget(self._chk_mod)
        layout.addLayout(bar)

        # -- Status label ----------------------------------------------------
        self._status = QLabel("Enter an IRMA or processor key and press Search or Enter.")
        layout.addWidget(self._status)

        # -- Results table ----------------------------------------------------
        RESULT_COLS = ["File", "Sheet", "Route", "Block #", "Type", "IRMA / Key", "Source"]
        self._table = QTableWidget(0, len(RESULT_COLS))
        self._table.setHorizontalHeaderLabels(RESULT_COLS)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setToolTip("Double-click a row to navigate to that sheet")
        layout.addWidget(self._table, stretch=1)

        # -- Bottom buttons ---------------------------------------------------
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.setFixedWidth(80)
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        # -- Connections ------------------------------------------------------
        self._query.returnPressed.connect(self._run_search)
        self._search_btn.clicked.connect(self._run_search)
        self._table.cellDoubleClicked.connect(self._on_double_click)

    # -- Search logic ---------------------------------------------------------

    def _run_search(self):
        query = self._query.text().strip().upper()
        self._table.setRowCount(0)
        self._results = []

        if not query:
            self._status.setText("Enter an IRMA or processor key and press Search or Enter.")
            return

        search_orig = self._chk_orig.isChecked()
        search_mod  = self._chk_mod.isChecked()

        if not search_orig and not search_mod:
            self._status.setText("Select at least one of Original / Modified.")
            return

        hits = []

        for fname, sheets in self._cache.items():
            for sname, entry in sheets.items():
                if not isinstance(entry, dict):
                    continue

                orig_blocks = entry.get("blocks", [])
                mod_blocks  = self._sheet_mods.get((fname, sname))

                def _scan(blocks, source_label):
                    found = []
                    for b_idx, block in enumerate(blocks):
                        route = block.get("route", "") or ""

                        # -- Farm rows --
                        for f_idx, row in enumerate(block.get("rows", [])):
                            irma = (row.get("irma") or "").strip().upper()
                            if query in irma:
                                found.append({
                                    "fname":  fname,
                                    "sname":  sname,
                                    "route":  route,
                                    "b_idx":  b_idx,
                                    "f_idx":  f_idx,
                                    "key":    row.get("irma", "").strip(),
                                    "kind":   "Farm",
                                    "source": source_label,
                                })

                        # -- Processor destinations --
                        for d_idx, dest in enumerate(block.get("dests", [])):
                            proc_key = (dest.get("key") or "").strip().upper()
                            if query in proc_key:
                                found.append({
                                    "fname":  fname,
                                    "sname":  sname,
                                    "route":  route,
                                    "b_idx":  b_idx,
                                    "f_idx":  -(d_idx + 1),   # negative = dest slot
                                    "key":    dest.get("key", "").strip(),
                                    "kind":   "Processor",
                                    "source": source_label,
                                })
                    return found

                orig_hits = _scan(orig_blocks, "Original") if search_orig else []
                mod_hits  = _scan(mod_blocks,  "Modified") if (search_mod and mod_blocks) else []

                def _key(h):
                    return (h["b_idx"], h["f_idx"], h["key"], h["kind"])

                orig_keys = {_key(h): h for h in orig_hits}
                mod_keys  = {_key(h): h for h in mod_hits}

                for k, h in orig_keys.items():
                    if k in mod_keys:
                        h_copy = dict(h); h_copy["source"] = "Both"
                        hits.append(h_copy)
                    else:
                        hits.append(h)
                for k, h in mod_keys.items():
                    if k not in orig_keys:
                        hits.append(h)

        hits.sort(key=lambda h: (h["fname"], h["sname"], h["b_idx"], h["f_idx"]))

        if not hits:
            self._status.setText(
                f"No results found matching '{query}'.")
            return

        self._status.setText(
            f"{len(hits)} result{'s' if len(hits) != 1 else ''} "
            f"for '{query}' - double-click a row to navigate.")

        SOURCE_COLOURS = {
            "Original": QColor("#e8f5e9"),
            "Modified": QColor("#e3f2fd"),
            "Both":     QColor("#fff8e1"),
        }
        KIND_COLOURS = {
            "Processor": QColor("#fce4ec"),   # light red tint
            "Farm":      None,                # use source colour
        }

        self._table.setRowCount(len(hits))
        self._results = hits

        for row_idx, h in enumerate(hits):
            src_bg  = SOURCE_COLOURS.get(h["source"], QColor("#ffffff"))
            kind_bg = KIND_COLOURS.get(h["kind"])
            bg      = kind_bg if kind_bg else src_bg

            values = [
                h["fname"],
                h["sname"],
                h["route"] or "-",
                str(h["b_idx"] + 1),
                h["kind"],
                h["key"],
                h["source"],
            ]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setBackground(bg)
                if col_idx == 5:   # key column - bold
                    f = item.font(); f.setBold(True); item.setFont(f)
                self._table.setItem(row_idx, col_idx, item)

    def _on_double_click(self, row, _col):
        if 0 <= row < len(self._results):
            h = self._results[row]
            self.navigate_requested.emit(h["fname"], h["sname"])


class ProcessorScheduleWidget(QWidget):
    """Paints a Gantt-style chart: processors on the Y axis, time of day on
    the X axis, and a coloured box for every truck's visit (arrival to
    departure).  Boxes that overlap another truck at the same processor are
    drawn with a bold red border.  Any configured avoid-window for that
    processor is shaded behind the boxes on its row.

    Visits whose time ranges overlap are stacked into separate sub-lanes
    within their processor's row (like overlapping meetings in a calendar
    app), rather than drawn on top of each other.  Without this, two
    overlapping boxes would occupy the exact same screen space and the
    later-drawn one would completely hide the earlier one's route label -
    the opposite of what a chart meant to show overlap should do.  Rows
    with overlap lanes expand vertically so each lane has at least
    MIN_LANE_H pixels, keeping bars readable at the cost of a taller
    chart.  Processor name labels are drawn in a separate frozen column
    (ProcessorLabelWidget) outside the scroll area so they remain visible
    during horizontal panning.

    visits: list of dicts {dest_key, dest_name, sname, colour, arr_min, dep_min}
            arr_min/dep_min are continuous minutes since a shared reference
            midnight (see _continuous_minutes) - not wrapped at 24h, so two
            trucks visiting at the same real clock time always line up at
            the same X position even if one route's day technically started
            "yesterday" relative to the other's wrap point.
    """
    PX_PER_MIN       = 4
    ROW_HEIGHT_BASE  = 32   # minimum row height; expands when lanes need more space
    MIN_LANE_H       = 20   # minimum pixels per lane (rows grow beyond ROW_HEIGHT_BASE)
    LABEL_WIDTH      = 230
    HEADER_HEIGHT    = 28
    TOP_MARGIN       = 14
    MIN_LABEL_LANE_H = 14   # below this lane height, skip the inline text label

    def __init__(self, visits, avoid_windows, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)
        self._boxes = []   # [((x,y,w,h), visit_dict), ...] for hit-testing
        self._avoid_windows = avoid_windows or {}
        self._build(visits)

    def _build(self, visits):
        by_proc = {}
        for v in visits:
            by_proc.setdefault(v["dest_key"], []).append(v)
        # Sort processors alphabetically by display name for predictable lookup
        self._procs = sorted(by_proc.items(),
                             key=lambda kv: kv[1][0]["dest_name"].lower())

        # Overlap detection: within each processor's own visits, grouped by
        # colour bucket (RED-RED, BLUE-BLUE only - RED and BLUE run on
        # entirely separate calendar days, so a RED visit and a BLUE visit
        # sharing a clock time never actually collide in reality even if
        # both happen to be visible on the same chart), flag any visit that
        # was present at a moment when the concurrent count at that
        # processor exceeded its dock capacity.  Most processors take one
        # truck at a time; a few (PROCESSOR_DOCK_CAPACITY) have multiple
        # bays, so e.g. capacity 2 means two trucks overlapping is fine but
        # a third overlapping both of them is flagged.  This mirrors
        # ALNSSolver._group_overlap_penalty exactly, so the chart's red
        # borders always match what the solver actually penalizes.
        for dk, vs in self._procs:
            capacity = PROCESSOR_DOCK_CAPACITY.get(dk, 1)
            by_colour = {}
            for v in vs:
                by_colour.setdefault(_sheet_colour_bucket(v.get("colour", "")), []).append(v)
            for _bucket, bucket_vs in by_colour.items():
                if len(bucket_vs) < 2:
                    continue
                events = []
                for idx, v in enumerate(bucket_vs):
                    events.append((v["arr_min"], 1, idx))
                    events.append((v["dep_min"], -1, idx))
                events.sort(key=lambda e: (e[0], e[1]))
                active = set()
                for (_t, delta, idx) in events:
                    if delta == 1:
                        active.add(idx)
                        if len(active) > capacity:
                            for a_idx in active:
                                bucket_vs[a_idx]["overlap"] = True
                    else:
                        active.discard(idx)
            vs.sort(key=lambda v: v["arr_min"])

        # Lane assignment - purely a layout concern, independent of colour
        # bucket or capacity: any two visits whose time ranges intersect at
        # all must never share a lane, full stop, regardless of whether
        # that overlap is a "real" same-day collision or just two different
        # colours happening to share a clock time in an "All" view.  Greedy
        # interval-graph colouring: sorted by arrival, reuse the first lane
        # whose previous occupant has already departed, otherwise open a
        # new lane.  Row height expands when more lanes are needed so each
        # lane has at least MIN_LANE_H pixels.
        self._lane_counts = {}   # dest_key -> lanes needed for that row
        for dk, vs in self._procs:
            lane_end = []   # lane_end[i] = departure time of that lane's current occupant
            for v in vs:    # already sorted by arr_min above
                placed = False
                for lane_idx in range(len(lane_end)):
                    if lane_end[lane_idx] <= v["arr_min"]:
                        v["_lane"] = lane_idx
                        lane_end[lane_idx] = v["dep_min"]
                        placed = True
                        break
                if not placed:
                    v["_lane"] = len(lane_end)
                    lane_end.append(v["dep_min"])
            self._lane_counts[dk] = max(1, len(lane_end))

        # Per-row heights: expand when lanes need more than ROW_HEIGHT_BASE
        cumulative = 0
        self._row_tops    = {}   # dest_key -> y offset from chart_top
        self._row_heights = {}   # dest_key -> row height in pixels
        for dk, _vs in self._procs:
            lc = self._lane_counts.get(dk, 1)
            rh = max(self.ROW_HEIGHT_BASE, lc * self.MIN_LANE_H)
            self._row_tops[dk]    = cumulative
            self._row_heights[dk] = rh
            cumulative += rh
        self._total_rows_h = max(1, cumulative)

        all_mins = [v["arr_min"] for _dk, vs in self._procs for v in vs] + \
                   [v["dep_min"] for _dk, vs in self._procs for v in vs]
        if all_mins:
            self._t_min = max(0, (min(all_mins) // 60) * 60 - 30)
            self._t_max = (max(all_mins) // 60 + 1) * 60 + 30
        else:
            self._t_min, self._t_max = 0, 24 * 60

        w = int((self._t_max - self._t_min) * self.PX_PER_MIN) + 20
        h = self._total_rows_h + 10
        self.setMinimumSize(w, h)

    def _x_for(self, minute):
        return int((minute - self._t_min) * self.PX_PER_MIN)

    def paintEvent(self, _event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, True)
        self._boxes = []

        font_label = QFont("Calibri", 9)

        if not self._procs:
            painter.setFont(QFont("Calibri", 11))
            painter.drawText(20, 30, "No processor visits found for this file.")
            return

        chart_top    = 0
        chart_bottom = self._total_rows_h
        chart_right  = self._x_for(self._t_max)

        # -- hour gridlines (labels drawn by frozen ProcessorHeaderWidget) ---
        t = (self._t_min // 60) * 60
        while t <= self._t_max:
            x = self._x_for(t)
            painter.setPen(QPen(QColor("#dddddd"), 1))
            painter.drawLine(x, chart_top, x, chart_bottom)
            t += 60

        # -- avoid-window shading (behind the visit boxes) ---------------
        for dk, _vs in self._procs:
            row_top = chart_top + self._row_tops[dk]
            rh      = self._row_heights[dk]
            for (av_start, av_end) in self._avoid_windows.get(dk, []):
                av_s_min = _hhmm_minutes(av_start)
                av_e_min = _hhmm_minutes(av_end)
                if av_s_min is None or av_e_min is None:
                    continue
                base = (self._t_min // (24 * 60)) * (24 * 60)
                for day_off in (0, 24 * 60):
                    s = base + day_off + av_s_min
                    e = base + day_off + av_e_min
                    if e <= self._t_min or s >= self._t_max:
                        continue
                    x1 = self._x_for(max(s, self._t_min))
                    x2 = self._x_for(min(e, self._t_max))
                    painter.fillRect(x1, row_top, max(1, x2 - x1), rh,
                                     QColor(244, 67, 54, 45))

        # -- row separators and visit boxes (labels drawn by ProcessorLabelWidget) --
        for dk, vs in self._procs:
            row_top    = chart_top + self._row_tops[dk]
            rh         = self._row_heights[dk]
            lane_count = self._lane_counts.get(dk, 1)
            lane_h     = rh / lane_count

            painter.setPen(QPen(QColor("#cccccc"), 1))
            painter.drawLine(0, row_top, chart_right, row_top)

            pad = 4 if lane_count == 1 else 2

            for v in vs:
                x1   = self._x_for(v["arr_min"])
                x2   = self._x_for(v["dep_min"])
                bw   = max(3, x2 - x1)
                lane = v.get("_lane", 0)
                by   = int(round(row_top + lane * lane_h + pad))
                bh   = max(3, int(round(lane_h - 2 * pad)))
                bg, fg, _ = day_colour_style(v.get("colour", ""))
                if bg is None:
                    bg = QColor("#90a4ae")
                    fg = QColor("#ffffff")
                painter.setBrush(bg)
                if v.get("overlap"):
                    painter.setPen(QPen(QColor("#d32f2f"), 2 if lane_count > 2 else 3))
                else:
                    painter.setPen(QPen(QColor("#37474f"), 1))
                painter.drawRoundedRect(x1, by, bw, bh, 3, 3)
                self._boxes.append(((x1, by, bw, bh), v))

                if bw > 36 and lane_h >= self.MIN_LABEL_LANE_H:
                    painter.setPen(QPen(fg, 1))
                    painter.setFont(font_label)
                    painter.drawText(x1 + 3, by, bw - 6, bh,
                                    Qt.AlignVCenter | Qt.AlignLeft, v["sname"])

        painter.setPen(QPen(QColor("#cccccc"), 1))
        painter.drawLine(0, chart_bottom, chart_right, chart_bottom)

    def mouseMoveEvent(self, event):
        pos = event.pos()
        for (x, y, w, h), v in self._boxes:
            if x <= pos.x() <= x + w and y <= pos.y() <= y + h:
                txt = (f"{v['dest_name']}\n{v['sname']}\n"
                      f"{_min_to_hhmm(v['arr_min'])} - {_min_to_hhmm(v['dep_min'])}")
                if v.get("overlap"):
                    txt += "\n** exceeds this dock's capacity at this time **"
                QToolTip.showText(event.globalPos(), txt, self)
                return
        QToolTip.hideText()

    def leaveEvent(self, _event):
        QToolTip.hideText()


class ProcessorLabelWidget(QWidget):
    """Frozen left column showing processor names, kept in vertical sync
    with ProcessorScheduleWidget via set_v_offset().

    Placed outside the chart's QScrollArea so it never scrolls horizontally.
    When the chart scroll area scrolls vertically, its verticalScrollBar
    valueChanged signal calls set_v_offset() here to shift the label
    painting by the same amount, keeping each label aligned with its row.
    """
    LABEL_WIDTH   = ProcessorScheduleWidget.LABEL_WIDTH
    HEADER_HEIGHT = ProcessorScheduleWidget.HEADER_HEIGHT
    TOP_MARGIN    = ProcessorScheduleWidget.TOP_MARGIN

    def __init__(self, procs, row_tops, row_heights, total_rows_h, parent=None):
        super().__init__(parent)
        self._procs        = procs
        self._row_tops     = row_tops
        self._row_heights  = row_heights
        self._total_rows_h = total_rows_h
        self._v_offset     = 0
        self.setFixedWidth(self.LABEL_WIDTH)

    def set_v_offset(self, v):
        self._v_offset = v
        self.update()

    def paintEvent(self, _event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, False)
        painter.fillRect(self.rect(), QColor("#ffffff"))

        font_label = QFont("Calibri", 9)

        # y coordinate in widget space where the first row's top edge lands
        y0 = -self._v_offset

        painter.setClipRect(self.rect())

        for dk, vs in self._procs:
            rh    = self._row_heights.get(dk, 32)
            row_y = y0 + self._row_tops.get(dk, 0)

            # Skip rows entirely outside the visible area
            if row_y + rh < 0 or row_y > self.height():
                continue

            dest_name = vs[0]["dest_name"]

            # Row separator
            painter.setPen(QPen(QColor("#cccccc"), 1))
            painter.drawLine(0, row_y, self.LABEL_WIDTH, row_y)

            # Label - vertically centred within the row
            painter.setFont(font_label)
            painter.setPen(QPen(QColor("#222222"), 1))
            painter.drawText(8, row_y, self.LABEL_WIDTH - 16, rh,
                             Qt.AlignVCenter | Qt.AlignLeft, dest_name)

        # Bottom border of the last row
        bottom_y = y0 + self._total_rows_h
        if 0 <= bottom_y <= self.height():
            painter.setPen(QPen(QColor("#cccccc"), 1))
            painter.drawLine(0, bottom_y, self.LABEL_WIDTH, bottom_y)

        # Right border - visual divider between frozen column and chart
        painter.setClipping(False)
        painter.setPen(QPen(QColor("#aaaaaa"), 1))
        painter.drawLine(self.LABEL_WIDTH - 1, 0, self.LABEL_WIDTH - 1, self.height())


class ProcessorHeaderWidget(QWidget):
    """Frozen time-axis header for ProcessorScheduleDialog.

    Placed above the chart scroll area in a horizontally-synced scroll
    area so the HH:MM tick labels always stay aligned with the chart's
    vertical gridlines, regardless of horizontal scroll position, and
    remain visible when the user scrolls the chart vertically.
    """
    PX_PER_MIN    = ProcessorScheduleWidget.PX_PER_MIN
    HEADER_HEIGHT = ProcessorScheduleWidget.HEADER_HEIGHT
    TOP_MARGIN    = ProcessorScheduleWidget.TOP_MARGIN

    def __init__(self, t_min, t_max, parent=None):
        super().__init__(parent)
        self._t_min = t_min
        self._t_max = t_max
        self.setFixedHeight(self.HEADER_HEIGHT + self.TOP_MARGIN)
        self.setMinimumWidth(int((t_max - t_min) * self.PX_PER_MIN) + 20)

    def _x_for(self, minute):
        return int((minute - self._t_min) * self.PX_PER_MIN)

    def paintEvent(self, _event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, False)
        painter.fillRect(self.rect(), QColor("#f5f5f5"))

        font_axis = QFont("Calibri", 8)
        painter.setFont(font_axis)
        h = self.height()

        t = (self._t_min // 60) * 60
        while t <= self._t_max:
            x = self._x_for(t)
            # Tick mark at the bottom edge
            painter.setPen(QPen(QColor("#aaaaaa"), 1))
            painter.drawLine(x, h - 5, x, h)
            # HH:MM label
            painter.setPen(QPen(QColor("#555555"), 1))
            hh = (t // 60) % 24
            mm = t % 60
            painter.drawText(x + 2, 0, 60, h - 4,
                             Qt.AlignVCenter | Qt.AlignLeft, f"{hh:02d}:{mm:02d}")
            t += 60

        # Bottom border
        painter.setPen(QPen(QColor("#aaaaaa"), 1))
        painter.drawLine(0, h - 1, self.width(), h - 1)


def _hhmm_minutes(s):
    """HH:MM string -> minutes since midnight, or None if unparseable."""
    t = parse_hhmm(s)
    if t is None:
        return None
    return t.hour * 60 + t.minute


class TruckAvailWidget(QWidget):
    """Gantt-style timeline.

    Day section:   one row per day route — bar from start to depot-return.
    Night section: one row per night route — start-time marker only.
    A vertical dashed line marks the earliest night-shift start.
    """
    PX_PER_MIN = 4
    ROW_H      = 30
    LABEL_W    = 70
    HEADER_H   = 28
    PAD        = 4
    SEP_H      = 6    # gap between day and night sections
    SECT_HDR_H = 20   # section header label height

    def __init__(self, day_routes, night_routes, night_start_mins, parent=None):
        super().__init__(parent)
        self._day    = day_routes
        self._night  = night_routes
        self._night_start = night_start_mins

        all_starts = ([r["start_mins"] for r in day_routes] +
                      [r["start_mins"] for r in night_routes])
        all_ends   = ([r["end_mins"] for r in day_routes] +
                      [r["end_mins"] for r in night_routes if r.get("end_mins") is not None])

        if all_starts:
            t_min = (min(all_starts) // 60) * 60 - 15
        else:
            t_min = 0

        candidates = []
        if all_ends:
            candidates.append((max(all_ends) // 60 + 1) * 60 + 30)
        if all_starts:
            candidates.append((max(all_starts) // 60 + 1) * 60 + 30)
        if night_start_mins is not None:
            candidates.append((night_start_mins // 60 + 1) * 60 + 30)
        t_max = max(candidates) if candidates else 24 * 60

        self._t_min = max(0, t_min)
        self._t_max = t_max

        n_day   = len(day_routes)
        n_night = len(night_routes)
        night_block = (self.SEP_H + self.SECT_HDR_H + n_night * self.ROW_H
                       if n_night else 0)
        w = self.LABEL_W + int((self._t_max - self._t_min) * self.PX_PER_MIN) + 20
        h = self.HEADER_H + n_day * self.ROW_H + night_block + 10
        self.setMinimumSize(w, max(h, 60))

    def _x(self, minute):
        return self.LABEL_W + int((minute - self._t_min) * self.PX_PER_MIN)

    def paintEvent(self, _event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, False)
        painter.fillRect(self.rect(), QColor("#ffffff"))

        font_axis  = QFont("Calibri", 8)
        font_label = QFont("Calibri", 9)
        font_bold  = QFont("Calibri", 9); font_bold.setBold(True)
        font_sect  = QFont("Calibri", 8); font_sect.setBold(True)

        chart_top = self.HEADER_H
        n_day     = len(self._day)
        day_bot   = chart_top + n_day * self.ROW_H

        CLR_GREEN  = QColor("#43a047")
        CLR_RED    = QColor("#e53935")
        CLR_GBORD  = QColor("#2e7d32")
        CLR_RBORD  = QColor("#b71c1c")
        CLR_RED_BG = QColor("#ffcdd2")
        CLR_BLU_BG = QColor("#bbdefb")
        CLR_OTH_BG = QColor("#e0e0e0")
        CLR_NIGHT_BG  = QColor("#ede7f6")   # pale indigo for night rows
        CLR_NIGHT_MRK = QColor("#5e35b1")   # indigo marker

        # ── Hour gridlines + axis labels ──────────────────────────────────
        painter.setFont(font_axis)
        total_bot = (day_bot + self.SEP_H + self.SECT_HDR_H +
                     len(self._night) * self.ROW_H if self._night else day_bot)
        t = (self._t_min // 60) * 60
        while t <= self._t_max:
            x = self._x(t)
            painter.setPen(QPen(QColor("#e0e0e0"), 1))
            painter.drawLine(x, chart_top, x, total_bot)
            painter.setPen(QPen(QColor("#555555"), 1))
            hh = (t // 60) % 24
            mm = t % 60
            painter.drawText(x + 2, 2, 52, self.HEADER_H - 4,
                             Qt.AlignVCenter | Qt.AlignLeft, f"{hh:02d}:{mm:02d}")
            t += 60

        # ── Night-shift start deadline line ───────────────────────────────
        if self._night_start is not None:
            xn = self._x(self._night_start)
            painter.setPen(QPen(QColor("#1565c0"), 2, Qt.DashLine))
            painter.drawLine(xn, 0, xn, total_bot)
            painter.setFont(font_bold)
            painter.setPen(QPen(QColor("#1565c0"), 1))
            nh = (self._night_start // 60) % 24
            nm = self._night_start % 60
            painter.drawText(xn + 3, 2, 80, self.HEADER_H - 4,
                             Qt.AlignVCenter | Qt.AlignLeft,
                             f"Night {nh:02d}:{nm:02d}")

        # ── Day route rows ────────────────────────────────────────────────
        for i, r in enumerate(self._day):
            row_top = chart_top + i * self.ROW_H
            on_time = r["on_time"]

            bg = QColor("#f8f8f8") if i % 2 == 0 else QColor("#ffffff")
            painter.fillRect(0, row_top, self.width(), self.ROW_H, bg)

            lbl_bg = (CLR_RED_BG if r["colour"] == "RED" else
                      CLR_BLU_BG if r["colour"] == "BLUE" else CLR_OTH_BG)
            painter.fillRect(0, row_top, self.LABEL_W, self.ROW_H, lbl_bg)
            painter.setFont(font_bold)
            painter.setPen(QPen(QColor("#222222"), 1))
            painter.drawText(4, row_top, self.LABEL_W - 8, self.ROW_H,
                             Qt.AlignVCenter | Qt.AlignLeft, r["sname"])

            x1 = self._x(r["start_mins"])
            x2 = self._x(r["end_mins"])
            bw = max(4, x2 - x1)
            by = row_top + self.PAD
            bh = self.ROW_H - 2 * self.PAD

            painter.setBrush(CLR_GREEN if on_time else CLR_RED)
            painter.setPen(QPen(CLR_GBORD if on_time else CLR_RBORD, 1))
            painter.drawRoundedRect(x1, by, bw, bh, 3, 3)

            # Stop segments overlaid on bar: farm=teal, processor=orange
            for kind, arr_m, dep_m in r.get("segments", []):
                xs = self._x(arr_m)
                xe = self._x(dep_m)
                sw = max(2, xe - xs)
                painter.fillRect(xs, by + 2, sw, bh - 4,
                                 QColor("#00bfa5") if kind == "farm"
                                 else QColor("#ff6d00"))

            if bw > 44:
                end_hh = (r["end_mins"] // 60) % 24
                end_mm = r["end_mins"] % 60
                painter.setPen(QPen(QColor("#ffffff"), 1))
                painter.setFont(font_label)
                painter.drawText(x1 + 3, by, bw - 6, bh,
                                 Qt.AlignVCenter | Qt.AlignRight,
                                 f"{end_hh:02d}:{end_mm:02d}")

            painter.setPen(QPen(QColor("#cccccc"), 1))
            painter.drawLine(0, row_top, self.width(), row_top)

        # Bottom of day section
        painter.setPen(QPen(QColor("#999999"), 1))
        painter.drawLine(0, day_bot, self.width(), day_bot)

        # ── Night section ─────────────────────────────────────────────────
        if self._night:
            sect_top = day_bot + self.SEP_H
            # Section header
            painter.fillRect(0, sect_top, self.width(), self.SECT_HDR_H,
                             QColor("#f0f0f0"))
            painter.setFont(font_sect)
            painter.setPen(QPen(QColor("#444444"), 1))
            painter.drawText(4, sect_top, self.width() - 8, self.SECT_HDR_H,
                             Qt.AlignVCenter | Qt.AlignLeft, "Night shift starts")

            night_top = sect_top + self.SECT_HDR_H
            for i, r in enumerate(self._night):
                row_top = night_top + i * self.ROW_H

                painter.fillRect(0, row_top, self.width(), self.ROW_H,
                                 CLR_NIGHT_BG)

                # Label
                lbl_bg = (CLR_RED_BG if r["colour"] == "RED" else
                          CLR_BLU_BG if r["colour"] == "BLUE" else CLR_OTH_BG)
                painter.fillRect(0, row_top, self.LABEL_W, self.ROW_H, lbl_bg)
                painter.setFont(font_bold)
                painter.setPen(QPen(QColor("#222222"), 1))
                painter.drawText(4, row_top, self.LABEL_W - 8, self.ROW_H,
                                 Qt.AlignVCenter | Qt.AlignLeft, r["sname"])

                x1 = self._x(r["start_mins"])
                by = row_top + self.PAD
                bh = self.ROW_H - 2 * self.PAD

                if r.get("end_mins") is not None:
                    # Full bar from start to end
                    x2 = self._x(r["end_mins"])
                    bw = max(4, x2 - x1)
                    painter.setBrush(QBrush(CLR_NIGHT_MRK))
                    painter.setPen(QPen(QColor("#311b92"), 1))
                    painter.drawRoundedRect(x1, by, bw, bh, 3, 3)
                    # Stop segments
                    for kind, arr_m, dep_m in r.get("segments", []):
                        xs = self._x(arr_m)
                        xe = self._x(dep_m)
                        sw = max(2, xe - xs)
                        painter.fillRect(xs, by + 2, sw, bh - 4,
                                         QColor("#00bfa5") if kind == "farm"
                                         else QColor("#ff6d00"))
                    if bw > 44:
                        end_hh = (r["end_mins"] // 60) % 24
                        end_mm = r["end_mins"] % 60
                        painter.setPen(QPen(QColor("#ffffff"), 1))
                        painter.setFont(font_label)
                        painter.drawText(x1 + 3, by, bw - 6, bh,
                                         Qt.AlignVCenter | Qt.AlignRight,
                                         f"{end_hh:02d}:{end_mm:02d}")
                else:
                    # Fallback: narrow marker if end time unavailable
                    painter.setBrush(QBrush(CLR_NIGHT_MRK))
                    painter.setPen(QPen(QColor("#311b92"), 1))
                    painter.drawRoundedRect(x1 - 2, by, 5, bh, 2, 2)
                    sh = (r["start_mins"] // 60) % 24
                    sm = r["start_mins"] % 60
                    painter.setFont(font_label)
                    painter.setPen(QPen(QColor("#311b92"), 1))
                    painter.drawText(x1 + 6, row_top, 60, self.ROW_H,
                                     Qt.AlignVCenter | Qt.AlignLeft,
                                     f"{sh:02d}:{sm:02d}")

                painter.setPen(QPen(QColor("#cccccc"), 1))
                painter.drawLine(0, row_top, self.width(), row_top)


class TruckAvailDialog(QDialog):
    """Dialog visualising day-route return times and night-route start times.
    Includes a RED / BLUE / All filter.
    """

    def __init__(self, day_routes, night_routes, night_start_mins, parent=None, fname="", date_str=""):
        super().__init__(parent)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setWindowTitle("Truck Availability - Return Times")
        self.setMinimumSize(900, 500)
        self.resize(1200, 620)

        self._all_day     = day_routes
        self._all_night   = night_routes
        self._night_start = night_start_mins
        self._fname       = fname
        self._date_str    = date_str

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        # Filter buttons
        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("Show:"))
        self._filter_btns = {}
        for label in ("All", "RED", "BLUE"):
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedWidth(60)
            btn.clicked.connect(lambda _, l=label: self._set_filter(l))
            self._filter_btns[label] = btn
            top_row.addWidget(btn)
        top_row.addSpacing(20)

        # Legend
        for hex_c, text in [("#43a047", "On time"), ("#e53935", "Late"),
                             ("#1565c0", "Night deadline"), ("#5e35b1", "Night shift"),
                             ("#00bfa5", "Farm stop"), ("#ff6d00", "Processor stop")]:
            dot = QLabel()
            dot.setFixedSize(12, 12)
            dot.setStyleSheet(f"background:{hex_c}; border-radius:2px;")
            top_row.addWidget(dot)
            top_row.addWidget(QLabel(text))
            top_row.addSpacing(8)
        top_row.addStretch()
        layout.addLayout(top_row)

        self._summary_lbl = QLabel()
        self._summary_lbl.setWordWrap(True)
        layout.addWidget(self._summary_lbl)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(False)
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        layout.addWidget(self._scroll, stretch=1)

        close_row = QHBoxLayout()
        close_row.addStretch()
        pdf_btn = QPushButton("Export PDF")
        pdf_btn.setStyleSheet(
            "QPushButton{background:#4a148c;color:white;font-weight:bold;"
            "border-radius:3px;padding:0 10px;}")
        pdf_btn.setToolTip("Export the truck availability timeline to a landscape A4 PDF.")
        pdf_btn.clicked.connect(lambda: _pdf_from_widget(
            self._scroll.widget(), "Truck_Availability_Timeline",
            parent=self, landscape=True, fname=self._fname, date_str=self._date_str))
        close_row.addWidget(pdf_btn)
        btn = QPushButton("Close")
        btn.clicked.connect(self.close)
        close_row.addWidget(btn)
        layout.addLayout(close_row)

        self._set_filter("All")

    def update_routes(self, day_routes, night_routes, night_start_mins):
        """Refresh with new data (called when solver finishes)."""
        self._all_day     = day_routes
        self._all_night   = night_routes
        self._night_start = night_start_mins
        current = next((l for l, b in self._filter_btns.items()
                        if b.isChecked()), "All")
        self._set_filter(current)

    def _set_filter(self, bucket):
        for label, btn in self._filter_btns.items():
            btn.setChecked(label == bucket)

        if bucket == "All":
            day   = self._all_day
            night = self._all_night
        else:
            day   = [r for r in self._all_day   if r["colour"] == bucket]
            night = [r for r in self._all_night if r["colour"] == bucket]

        on_time = sum(1 for r in day if r["on_time"])
        total   = len(day)
        if self._night_start is not None:
            nh, nm = self._night_start // 60, self._night_start % 60
            self._summary_lbl.setText(
                f"{on_time}/{total} day routes back before night deadline "
                f"({nh:02d}:{nm:02d})  •  {total - on_time} run over  •  "
                f"{len(night)} night routes shown")
        else:
            self._summary_lbl.setText(
                f"{total} day routes  •  {len(night)} night routes  •  "
                "No night deadline detected.")

        widget = TruckAvailWidget(day, night, self._night_start)
        self._scroll.setWidget(widget)


class ProcessorScheduleDialog(QDialog):
    """Shows every truck's visit to every processor across the currently
    loaded file as a Gantt-style chart, built from whatever blocks are
    currently active (solver output if present, original parse otherwise).

    Includes a RED / BLUE / All toggle.  RED and BLUE run on entirely
    separate calendar days, but a single loaded file's sheets can span both
    (e.g. a combined "this week" route sheet), so without a filter the chart
    would show two unrelated days' trucks on one shared axis.  The toggle
    lets the user view one day's schedule in isolation.  Overlap highlighting
    already only ever compares visits within the same colour bucket
    regardless of which toggle is selected (see ProcessorScheduleWidget),
    so "All" is still safe to use - it just shows more at once.
    """

    def __init__(self, visits, avoid_windows, fname, parent=None, date_str=""):
        super().__init__(parent)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setWindowTitle(f"Processor Schedule - {fname}")
        self.setMinimumSize(900, 640)
        self.resize(1400, 800)
        self._fname        = fname
        self._date_str     = date_str
        self._all_visits   = visits
        self._avoid_windows = avoid_windows

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        legend = QLabel(
            "Each box is one truck's time at a processor (arrival to departure). "
            "Overlapping visits split into taller rows with separate lanes so "
            "every route number stays readable. "
            "Thick red border = exceeds dock capacity at that moment, "
            "same day only (most docks take 1 truck at a time; a few take 2 -- "
            "see PROCESSOR_DOCK_CAPACITY). Pink shading = a configured "
            "avoid-window for that processor. Hover a box for details. "
            "Processor names and the time axis stay frozen while scrolling.")
        legend.setWordWrap(True)
        layout.addWidget(legend)

        # -- RED / BLUE / All toggle ----------------------------------------
        toggle_row = QHBoxLayout()
        toggle_row.addWidget(QLabel("Show:"))

        def _count(bucket):
            if bucket is None:
                return len(visits)
            return sum(1 for v in visits
                      if _sheet_colour_bucket(v.get("colour", "")) == bucket)

        self._toggle_group = QButtonGroup(self)
        self._toggle_group.setExclusive(True)
        self._btn_all  = QPushButton(f"All ({_count(None)})")
        self._btn_red  = QPushButton(f"RED ({_count('RED')})")
        self._btn_blue = QPushButton(f"BLUE ({_count('BLUE')})")
        for b in (self._btn_all, self._btn_red, self._btn_blue):
            b.setCheckable(True)
            b.setFixedHeight(24)
            self._toggle_group.addButton(b)
            toggle_row.addWidget(b)
        self._btn_all.setChecked(True)
        self._btn_all.clicked.connect(lambda: self._set_filter(None))
        self._btn_red.clicked.connect(lambda: self._set_filter("RED"))
        self._btn_blue.clicked.connect(lambda: self._set_filter("BLUE"))
        toggle_row.addStretch()
        layout.addLayout(toggle_row)

        # Frozen header row (time axis): blank spacer aligned with label column,
        # then a horizontally-synced scroll area showing ProcessorHeaderWidget.
        header_band = QHBoxLayout()
        header_band.setContentsMargins(0, 0, 0, 0)
        header_band.setSpacing(0)
        layout.addLayout(header_band)

        header_gap = QFrame()
        header_gap.setFixedWidth(ProcessorScheduleWidget.LABEL_WIDTH)
        header_gap.setFrameShape(QFrame.NoFrame)
        header_band.addWidget(header_gap)

        self._header_scroll = QScrollArea()
        self._header_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._header_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._header_scroll.setWidgetResizable(False)
        self._header_scroll.setFrameShape(QFrame.NoFrame)
        header_band.addWidget(self._header_scroll)

        # Two-panel chart area: frozen label column left, scrollable chart right
        chart_panel = QHBoxLayout()
        chart_panel.setContentsMargins(0, 0, 0, 0)
        chart_panel.setSpacing(0)
        layout.addLayout(chart_panel)

        # Left: label column holder (fixed width, not inside a scroll area)
        self._label_frame = QFrame()
        self._label_frame.setFixedWidth(ProcessorScheduleWidget.LABEL_WIDTH)
        self._label_frame_layout = QVBoxLayout(self._label_frame)
        self._label_frame_layout.setContentsMargins(0, 0, 0, 0)
        self._label_frame_layout.setSpacing(0)
        chart_panel.addWidget(self._label_frame)

        # Right: scroll area for the chart content only (no labels)
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(False)
        chart_panel.addWidget(self._scroll)

        self._chart = None
        self._label_widget = None
        self._header_widget = None
        self._set_filter(None)   # initial build: All

        close_row = QHBoxLayout()
        close_row.addStretch()
        pdf_btn = QPushButton("Export PDF")
        pdf_btn.setStyleSheet(
            "QPushButton{background:#4a148c;color:white;font-weight:bold;"
            "border-radius:3px;padding:0 10px;}")
        pdf_btn.setToolTip("Export the processor schedule to a landscape A4 PDF.")
        pdf_btn.clicked.connect(lambda: _pdf_from_widget(
            self._chart, "Processor_Schedule", parent=self, landscape=True,
            fname=self._fname, date_str=self._date_str))
        close_row.addWidget(pdf_btn)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.close)
        close_row.addWidget(close_btn)
        layout.addLayout(close_row)

    def _set_filter(self, bucket):
        """Rebuild the chart for the chosen colour bucket (None = All)."""
        if bucket is None:
            filtered = self._all_visits
        else:
            filtered = [v for v in self._all_visits
                       if _sheet_colour_bucket(v.get("colour", "")) == bucket]
        self._chart = ProcessorScheduleWidget(filtered, self._avoid_windows)
        self._scroll.setWidget(self._chart)

        # Rebuild the frozen label column to match the new chart
        while self._label_frame_layout.count():
            item = self._label_frame_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._label_widget = ProcessorLabelWidget(
            self._chart._procs,
            self._chart._row_tops,
            self._chart._row_heights,
            self._chart._total_rows_h,
        )
        self._label_frame_layout.addWidget(self._label_widget)

        # Rebuild the frozen time-axis header to match the new chart
        self._header_widget = ProcessorHeaderWidget(
            self._chart._t_min, self._chart._t_max)
        self._header_scroll.setWidget(self._header_widget)
        self._header_scroll.setFixedHeight(self._header_widget.height())

        # Sync vertical scroll: chart drives label column offset
        self._scroll.verticalScrollBar().valueChanged.connect(
            self._label_widget.set_v_offset)
        # Sync horizontal scroll: chart drives header position
        self._scroll.horizontalScrollBar().valueChanged.connect(
            self._header_scroll.horizontalScrollBar().setValue)
        self._header_scroll.horizontalScrollBar().valueChanged.connect(
            self._scroll.horizontalScrollBar().setValue)
        # Apply current scroll positions immediately (handles filter-switch case)
        self._label_widget.set_v_offset(
            self._scroll.verticalScrollBar().value())
        self._header_scroll.horizontalScrollBar().setValue(
            self._scroll.horizontalScrollBar().value())


# -- Main window ---------------------------------------------------------------

class _VolLegend(QWidget):
    """Small overlay widget showing the volume→colour gradient legend."""
    _MAX_VOL = 40_000

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(110, 185)
        self.setAttribute(Qt.WA_TransparentForMouseEvents, False)

    @staticmethod
    def vol_color(vol, max_vol=40_000):
        """Blue (0 L) → Amber (20 k) → Red (40 k)."""
        t = max(0.0, min(1.0, vol / max_vol))
        if t < 0.5:
            s = t * 2
            r = int(25  + (255 - 25)  * s)
            g = int(118 + (193 - 118) * s)
            b = int(210 + (7   - 210) * s)
        else:
            s = (t - 0.5) * 2
            r = int(255 + (211 - 255) * s)
            g = int(193 + (47  - 193) * s)
            b = int(7   + (47  - 7)   * s)
        return QColor(r, g, b)

    def paintEvent(self, _ev):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, False)

        # Background pill
        p.setBrush(QBrush(QColor(0, 0, 0, 185)))
        p.setPen(QPen(QColor("#555555"), 1))
        p.drawRoundedRect(self.rect().adjusted(1, 1, -1, -1), 6, 6)

        # Title
        p.setPen(QColor("#ffffff"))
        p.setFont(QFont("Calibri", 8, QFont.Bold))
        p.drawText(0, 5, 110, 16, Qt.AlignCenter, "Volume (L)")

        # Gradient bar
        bx, by, bw, bh = 18, 26, 18, 120
        for py in range(bh):
            vol = self._MAX_VOL * (1 - py / bh)
            p.fillRect(bx, by + py, bw, 1, self.vol_color(vol))
        p.setPen(QPen(QColor("#888888"), 1))
        p.drawRect(bx, by, bw, bh)

        # Tick labels
        p.setPen(QColor("#ffffff"))
        p.setFont(QFont("Calibri", 7))
        for frac, label in [(1.0, "0"), (0.75, "10k"),
                             (0.5, "20k"), (0.25, "30k"), (0.0, "40k")]:
            py = by + int(bh * frac)
            p.drawLine(bx + bw, py, bx + bw + 3, py)
            p.drawText(bx + bw + 5, py - 6, 38, 13,
                       Qt.AlignLeft | Qt.AlignVCenter, label)


class MapDialog(QDialog):
    """Modal map dialog showing one block of the current sheet at a time.

    Reads road geometry from routes.db (searched for beside the exe).
    Uses the same Web Mercator projection and OSM tile system as map_tester.py.
    """

    _TILE_URL    = "https://tile.openstreetmap.org/{z}/{x}/{y}.png"
    _TILE_CACHE  = Path.home() / ".cache" / "map_tester_tiles"
    _TILE_ZOOM   = 13
    _TILE_BOUNDS = (49.00, 49.38, -123.30, -121.40)

    # ── tiny helpers ──────────────────────────────────────────────────────────

    @staticmethod
    def _find_db():
        """Return path to routes.db beside the exe (or script in dev mode)."""
        here = (Path(sys.executable).parent
                if getattr(sys, "frozen", False)
                else Path(__file__).parent)
        return here / "routes.db"

    @staticmethod
    def _merc(lat, lon):
        x =  math.radians(lon)
        y = -math.log(math.tan(math.pi / 4 + math.radians(lat) / 2))
        return x, y

    @staticmethod
    def _tile_xy(lat, lon, z):
        n  = 2 ** z
        tx = int((lon + 180) / 360 * n)
        lr = math.radians(lat)
        ty = int((1 - math.log(math.tan(lr) + 1 / math.cos(lr)) / math.pi) / 2 * n)
        return tx, ty

    @staticmethod
    def _tile_nw(tx, ty, z):
        n   = 2 ** z
        lon = tx / n * 360 - 180
        lat = math.degrees(math.atan(math.sinh(math.pi * (1 - 2 * ty / n))))
        return lat, lon

    def _to_scene(self, lat, lon):
        mx, my = self._merc(lat, lon)
        return mx * self._scale + self._ox, my * self._scale + self._oy

    # ── construction ──────────────────────────────────────────────────────────

    def __init__(self, blocks, sname, irma_lookup, all_sheets=None, parent=None):
        """
        blocks      : list of block dicts for this sheet
        sname       : sheet name (for window title)
        irma_lookup : {irma: (lat, lon)} for all known locations
        all_sheets  : {sname: {"blocks": [...], "day_colour": ...}} for overlay picker
        """
        super().__init__(parent)
        self.setModal(True)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setWindowTitle(f"Route Map — {sname}")
        self.setMinimumSize(1000, 680)
        self.resize(1200, 780)

        self._blocks        = blocks
        self._sname         = sname
        self._irma_locs     = irma_lookup
        self._all_sheets    = all_sheets or {}
        self._block_idx     = 0
        self._tile_items    = {}
        self._route_items   = []
        self._overlay_items = {}   # sname -> [scene items]
        self._overlay_cols  = {}   # sname -> QColor
        self._tile_loader   = None
        self._overlay_palette = [
            QColor("#FF6B6B"), QColor("#4ECDC4"), QColor("#45B7D1"),
            QColor("#FECA57"), QColor("#FF9FF3"), QColor("#96CEB4"),
            QColor("#5F27CD"), QColor("#00D2D3"),
        ]
        self._palette_idx = 0

        # DB
        db_path = self._find_db()
        self._db_ok = db_path.exists()
        self._db    = None
        if self._db_ok:
            import sqlite3 as _sq
            self._con = _sq.connect(str(db_path), check_same_thread=False)
        else:
            self._con = None

        # Projection from all known coords
        lats = [v[0] for v in irma_lookup.values()]
        lons = [v[1] for v in irma_lookup.values()]
        if lats:
            lat_min = min(lats) - 0.05; lat_max = max(lats) + 0.05
            lon_min = min(lons) - 0.05; lon_max = max(lons) + 0.05
        else:
            lat_min, lat_max, lon_min, lon_max = 49.0, 49.4, -123.3, -121.5

        mW, mH, margin = 1160, 720, 50
        mx0, my0 = self._merc(lat_max, lon_min)
        mx1, my1 = self._merc(lat_min, lon_max)
        self._scale = min((mW - 2*margin) / (mx1 - mx0),
                          (mH - 2*margin) / (my1 - my0))
        self._ox = margin - mx0 * self._scale
        self._oy = margin - my0 * self._scale

        # ── Layout ────────────────────────────────────────────────────────────
        lay = QVBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(4)

        # Top bar: block nav + status
        top = QHBoxLayout()
        self._prev_btn = QPushButton("◀ Prev Block")
        self._next_btn = QPushButton("Next Block ▶")
        for b in (self._prev_btn, self._next_btn):
            b.setFixedHeight(26)
            b.setStyleSheet(
                "QPushButton{background:#1565c0;color:white;font-weight:bold;"
                "border-radius:3px;padding:0 10px;}"
                "QPushButton:disabled{background:#90caf9;}")
        self._prev_btn.clicked.connect(lambda: self._show_block(self._block_idx - 1))
        self._next_btn.clicked.connect(lambda: self._show_block(self._block_idx + 1))
        self._block_lbl = QLabel("")
        self._block_lbl.setStyleSheet("font-weight:bold;")
        top.addWidget(self._prev_btn)
        top.addWidget(self._block_lbl)
        top.addWidget(self._next_btn)
        top.addSpacing(20)
        self._status_lbl = QLabel("Loading tiles…" if self._db_ok
                                  else "⚠ routes.db not found — geometry unavailable")
        self._status_lbl.setStyleSheet(
            "color:#555;" if self._db_ok else "color:#c62828;font-weight:bold;")
        top.addWidget(self._status_lbl)
        top.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        top.addWidget(close_btn)
        lay.addLayout(top)

        # Overlay row
        ov_row = QHBoxLayout()
        ov_row.addWidget(QLabel("Overlay:"))
        self._ov_combo = QComboBox()
        self._ov_combo.setMinimumWidth(160)
        self._ov_combo.setToolTip("Select a sheet to overlay on this map")
        other_sheets = sorted(s for s in self._all_sheets if s != sname)
        self._ov_combo.addItems(other_sheets)
        ov_row.addWidget(self._ov_combo, stretch=1)

        add_ov_btn = QPushButton("Add")
        add_ov_btn.setFixedWidth(50)
        add_ov_btn.setStyleSheet(
            "QPushButton{background:#37474f;color:white;border-radius:3px;padding:0 6px;}"
            "QPushButton:hover{background:#546e7a;}")
        add_ov_btn.clicked.connect(self._on_add_overlay)
        ov_row.addWidget(add_ov_btn)

        clr_ov_btn = QPushButton("Clear Overlays")
        clr_ov_btn.setFixedWidth(100)
        clr_ov_btn.setStyleSheet(
            "QPushButton{background:#37474f;color:white;border-radius:3px;padding:0 6px;}"
            "QPushButton:hover{background:#546e7a;}")
        clr_ov_btn.clicked.connect(self._clear_overlays)
        ov_row.addWidget(clr_ov_btn)

        ov_row.addSpacing(10)
        self._ov_chips_layout = QHBoxLayout()
        self._ov_chips_layout.setSpacing(4)
        ov_row.addLayout(self._ov_chips_layout)
        ov_row.addStretch()
        lay.addLayout(ov_row)

        # Map view
        self._scene = QGraphicsScene(self)
        self._view  = QGraphicsView(self._scene)
        self._view.setRenderHint(QPainter.Antialiasing)
        self._view.setRenderHint(QPainter.SmoothPixmapTransform)
        self._view.setDragMode(QGraphicsView.ScrollHandDrag)
        self._view.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self._view.setBackgroundBrush(QBrush(QColor("#1a2332")))
        self._view.setStyleSheet("border:none;")
        self._view.wheelEvent = self._wheel
        lay.addWidget(self._view, stretch=1)

        # Volume legend — overlaid in the top-right corner of the map view
        self._legend = _VolLegend(self._view)
        self._legend.move(self._view.width() - self._legend.width() - 10, 10)
        self._legend.raise_()
        self._legend.show()

        # Build base scene + load tiles
        self._build_base()
        self._load_tiles()
        # _show_block(0) deferred to showEvent so the view is properly sized first

    def resizeEvent(self, ev):
        super().resizeEvent(ev)
        if hasattr(self, "_legend"):
            self._legend.move(
                self._view.width() - self._legend.width() - 10, 10)

    def showEvent(self, ev):
        super().showEvent(ev)
        # Re-fit to Lower Mainland then draw block 0 now that the view has a real size
        self._fit_to_bounds()
        self._show_block(0)

    def _wheel(self, ev):
        f = 1.15 if ev.angleDelta().y() > 0 else 1 / 1.15
        self._view.scale(f, f)

    def closeEvent(self, ev):
        if self._con:
            self._con.close()
        super().closeEvent(ev)

    # ── Scene base (all dots) ─────────────────────────────────────────────────

    def _fit_to_bounds(self):
        """Fit the view to the Lower Mainland tile area."""
        x0, y0 = self._to_scene(49.38, -123.30)
        x1, y1 = self._to_scene(49.00, -121.40)
        self._view.fitInView(QRectF(x0, y0, x1 - x0, y1 - y0),
                             Qt.KeepAspectRatio)

    def _build_base(self):
        self._scene.clear()
        self._tile_items.clear()
        x0, y0 = self._to_scene(49.40, -123.35)
        x1, y1 = self._to_scene(48.95, -121.35)
        self._scene.setSceneRect(QRectF(x0, y0, x1-x0, y1-y0))

        # Dim dots for every known location
        pen = QPen(QColor("#445566"), 0.8)
        for irma, (lat, lon) in self._irma_locs.items():
            x, y = self._to_scene(lat, lon)
            r    = 2.5
            dot  = QGraphicsEllipseItem(-r, -r, r*2, r*2)
            dot.setPen(pen)
            if irma.isdigit():
                dot.setBrush(QBrush(QColor("#e53935")))
            elif irma == "VEDDER":
                dot.setBrush(QBrush(QColor("#ffd600")))
            else:
                dot.setBrush(QBrush(QColor("#334455")))
            dot.setFlag(dot.ItemIgnoresTransformations, True)
            dot.setPos(x, y)
            dot.setZValue(2)
            dot.setToolTip(irma)
            self._scene.addItem(dot)

    # ── OSM tile loading ──────────────────────────────────────────────────────

    def _load_tiles(self):
        lat_min, lat_max, lon_min, lon_max = self._TILE_BOUNDS
        z   = self._TILE_ZOOM
        tx0, ty0 = self._tile_xy(lat_max, lon_min, z)
        tx1, ty1 = self._tile_xy(lat_min, lon_max, z)
        tiles = [(tx, ty) for ty in range(ty0, ty1+1) for tx in range(tx0, tx1+1)]

        self._TILE_CACHE.mkdir(parents=True, exist_ok=True)
        self._sig  = _TileSignaller()
        self._sig.tile_ready.connect(self._on_tile)
        self._loader = _TileLoaderThread(tiles, z, self._sig, self._TILE_URL,
                                         self._TILE_CACHE)
        self._loader.finished.connect(lambda: self._status_lbl.setText(
            "Tiles loaded." if self._db_ok else self._status_lbl.text()))
        self._loader.start()

    def _on_tile(self, z, tx, ty, pixmap):
        if z != self._TILE_ZOOM:
            return
        lat_nw, lon_nw = self._tile_nw(tx,   ty,   z)
        lat_se, lon_se = self._tile_nw(tx+1, ty+1, z)
        x0, y0 = self._to_scene(lat_nw, lon_nw)
        x1, y1 = self._to_scene(lat_se, lon_se)
        item = self._scene.addPixmap(pixmap)
        item.setPos(x0, y0)
        sx = (x1-x0) / 256
        sy = (y1-y0) / 256
        item.setTransform(QTransform.fromScale(sx, sy))
        item.setOpacity(0.85)
        item.setZValue(0)
        self._tile_items[(tx, ty)] = item

    # ── Route drawing ─────────────────────────────────────────────────────────

    def _clear_route(self):
        for item in self._route_items:
            self._scene.removeItem(item)
        self._route_items = []

    # ── Overlay methods ───────────────────────────────────────────────────────

    def _on_add_overlay(self):
        sname = self._ov_combo.currentText()
        if not sname or sname == self._sname:
            return
        if sname in self._overlay_items:
            return   # already shown
        entry = self._all_sheets.get(sname)
        if not isinstance(entry, dict):
            return
        blocks = entry.get("blocks", [])
        if not blocks:
            return
        colour = self._overlay_palette[self._palette_idx % len(self._overlay_palette)]
        self._palette_idx += 1
        self._overlay_cols[sname]  = colour
        self._overlay_items[sname] = []
        self._draw_overlay(sname, blocks, colour)
        self._add_overlay_chip(sname, colour)

    def _draw_overlay(self, sname, blocks, colour):
        """Draw all blocks of an overlay sheet as thin semi-transparent lines."""
        items = self._overlay_items[sname]
        ov_pen = QPen(QColor(colour.red(), colour.green(), colour.blue(), 190), 1.8)
        ov_pen.setCosmetic(True)
        ov_pen.setCapStyle(Qt.RoundCap)
        ov_pen.setJoinStyle(Qt.RoundJoin)

        for b_idx, block in enumerate(blocks):
            is_last = (b_idx == len(blocks) - 1)
            origin  = ("VEDDER" if b_idx == 0
                       else (_block_last_dest_key(blocks[b_idx-1]) or "VEDDER"))
            stops   = _build_block_stops(block, origin, is_last)
            valid   = [(s["key"], s) for s in stops if s.get("key")]

            for (a_key, _a), (b_key, _b) in zip(valid, valid[1:]):
                if not a_key or not b_key:
                    continue
                geom = self._get_geom(a_key, b_key)
                if geom:
                    pts = geom
                elif a_key in self._irma_locs and b_key in self._irma_locs:
                    pts = [self._irma_locs[a_key], self._irma_locs[b_key]]
                else:
                    continue
                path  = QPainterPath()
                first = True
                for lat, lon in pts:
                    x, y = self._to_scene(lat, lon)
                    if first: path.moveTo(x, y); first = False
                    else:     path.lineTo(x, y)
                item = self._scene.addPath(path, ov_pen)
                item.setZValue(2)   # below main route (z=3-4)
                items.append(item)

    def _remove_overlay(self, sname):
        for item in self._overlay_items.pop(sname, []):
            self._scene.removeItem(item)
        self._overlay_cols.pop(sname, None)

    def _clear_overlays(self):
        for sname in list(self._overlay_items.keys()):
            self._remove_overlay(sname)
        # Remove all chips
        while self._ov_chips_layout.count():
            w = self._ov_chips_layout.takeAt(0).widget()
            if w:
                w.deleteLater()

    def _add_overlay_chip(self, sname, colour):
        """Add a small coloured chip with an × to remove the overlay."""
        chip = QPushButton(f"× {sname}")
        chip.setFlat(True)
        chip.setStyleSheet(
            f"QPushButton{{"
            f"background:{colour.name()}; color:{'#000' if colour.lightness() > 128 else '#fff'};"
            f"border-radius:3px; padding:1px 5px; font-size:7pt; font-weight:bold;}}"
            f"QPushButton:hover{{opacity:0.8;}}")
        chip.setToolTip(f"Remove overlay: {sname}")
        chip.clicked.connect(lambda _, s=sname, c=chip: self._remove_chip(s, c))
        self._ov_chips_layout.addWidget(chip)

    def _remove_chip(self, sname, chip):
        self._remove_overlay(sname)
        self._ov_chips_layout.removeWidget(chip)
        chip.deleteLater()

    def _get_geom(self, origin, dest):
        """Return [(lat,lon),...] from routes.db, or None."""
        if not self._con:
            return None
        import struct as _st
        row = self._con.execute(
            "SELECT geometry FROM routes WHERE origin=? AND dest=?",
            (origin, dest)).fetchone()
        if not row or not row[0]:
            return None
        data = row[0]
        n    = len(data) // 4
        vals = _st.unpack(f"{n}f", data)
        return [(vals[i], vals[i+1]) for i in range(0, n, 2)]

    def _show_block(self, idx):
        idx = max(0, min(idx, len(self._blocks) - 1))
        self._block_idx = idx
        self._prev_btn.setEnabled(idx > 0)
        self._next_btn.setEnabled(idx < len(self._blocks) - 1)
        n = len(self._blocks)
        self._block_lbl.setText(f"Block {idx+1} of {n}")
        self._clear_route()

        block   = self._blocks[idx]
        is_last = (idx == len(self._blocks) - 1)
        origin  = ("VEDDER" if idx == 0
                   else (_block_last_dest_key(self._blocks[idx-1]) or "VEDDER"))
        stops   = _build_block_stops(block, origin, is_last)

        # Preload = block with no farm stops → treat as full (40 k) throughout
        is_preload  = not any(s["type"] == "farm" for s in stops)
        current_vol = 40_000.0 if is_preload else 0.0

        # Filter to stops that have a key and are in irma_locs
        valid = [(s["key"], s) for s in stops if s.get("key")]

        total_vol = 0.0

        for (a_key, a_stop), (b_key, b_stop) in zip(valid, valid[1:]):
            if not a_key or not b_key:
                continue

            colour = _VolLegend.vol_color(current_vol)
            geom   = self._get_geom(a_key, b_key)

            if geom:
                pts = geom
            elif a_key in self._irma_locs and b_key in self._irma_locs:
                pts    = [self._irma_locs[a_key], self._irma_locs[b_key]]
                colour = QColor(colour.red(), colour.green(), colour.blue(), 100)
            else:
                # Update volume even if we can't draw the segment
                if b_stop["type"] == "farm" and b_stop.get("farm"):
                    try:
                        current_vol = min(40_000.0,
                            current_vol + float(b_stop["farm"].get("prior_vol") or 0))
                    except (TypeError, ValueError):
                        pass
                elif b_stop["type"] == "dest":
                    current_vol = 0.0
                continue

            path  = QPainterPath()
            first = True
            for lat, lon in pts:
                x, y = self._to_scene(lat, lon)
                if first: path.moveTo(x, y); first = False
                else:     path.lineTo(x, y)

            shp = QPen(QColor(0, 0, 0, 100), 3); shp.setCosmetic(True)
            shadow = self._scene.addPath(path, shp)
            shadow.setZValue(3)
            self._route_items.append(shadow)

            lp = QPen(colour, 2.5); lp.setCosmetic(True)
            lp.setCapStyle(Qt.RoundCap); lp.setJoinStyle(Qt.RoundJoin)
            line = self._scene.addPath(path, lp)
            line.setZValue(4)
            self._route_items.append(line)

            # Update volume on arrival at b
            if b_stop["type"] == "farm" and b_stop.get("farm"):
                try:
                    fv = float(b_stop["farm"].get("prior_vol") or 0)
                    current_vol = min(40_000.0, current_vol + fv)
                    total_vol  += fv
                except (TypeError, ValueError):
                    pass
            elif b_stop["type"] == "dest":
                current_vol = 0.0

        # Highlight stop dots + tooltips
        for seq, stop in enumerate(stops):
            key = stop["key"]
            if not key or key not in self._irma_locs:
                continue
            lat, lon = self._irma_locs[key]
            x, y = self._to_scene(lat, lon)
            r = 5.5

            if stop["type"] == "dest":
                fill = QColor("#ff6d00"); tip_prefix = "Processor"
            elif stop["type"] == "vedder":
                fill = QColor("#ffd600"); tip_prefix = "Depot"
            elif stop["type"] == "origin":
                fill = QColor("#ffd600"); tip_prefix = "Depot"
            else:
                fill = (QColor("#e1bee7") if key in MENNONITE_FARMS
                        else QColor("#43a047"))
                tip_prefix = "Farm"

            # Farm name from irma_lookup dict (stored as tooltip on base dot)
            name = ""
            if stop["type"] == "farm" and stop.get("farm"):
                ec = stop["farm"].get("_extra_cells") or {}
                name = ec.get(18, "") or ""
            tip = f"{tip_prefix}: {key}" + (f"\n{name}" if name else "")

            dot = QGraphicsEllipseItem(-r, -r, r*2, r*2)
            dot.setPen(QPen(QColor("#222"), 1.5))
            dot.setBrush(QBrush(fill))
            dot.setFlag(dot.ItemIgnoresTransformations, True)
            dot.setPos(x, y)
            dot.setZValue(6)
            dot.setToolTip(tip)
            self._scene.addItem(dot)
            self._route_items.append(dot)

            # Number label affixed to dot (child item — moves with it)
            font = QFont("Calibri", 7); font.setBold(True)
            num = QGraphicsTextItem(str(seq), dot)   # parent = dot
            num.setFont(font)
            num.setDefaultTextColor(QColor("#ffffff"))
            num.setZValue(7)
            num.setToolTip(tip)
            br = num.boundingRect()
            # Dark pill background behind the number
            bg = QGraphicsRectItem(
                r + 1, -br.height() / 2,
                br.width() + 2, br.height(),
                dot)
            bg.setBrush(QBrush(QColor(0, 0, 0, 200)))
            bg.setPen(QPen(Qt.NoPen))
            bg.setZValue(6)
            num.setPos(r + 2, -br.height() / 2)
            # num and bg are children of dot — removed automatically when dot is removed

        if not self._db_ok:
            self._status_lbl.setText("⚠ routes.db not found — place it beside the exe")
        else:
            pre_tag = "  [PRELOAD]" if is_preload else ""
            vol_tag = f"  ·  {int(total_vol):,} L collected" if not is_preload else "  ·  40,000 L (full)"
            self._status_lbl.setText(
                f"Block {idx+1}/{n}  ·  {len(stops)-2} stops{pre_tag}{vol_tag}")


# ── Tile loader helpers (shared with map_tester style) ─────────────────────

class _TileSignaller(QObject):
    tile_ready = pyqtSignal(int, int, int, object)


class _TileLoaderThread(QThread):
    def __init__(self, tiles, zoom, sig, url_tmpl, cache_dir, parent=None):
        super().__init__(parent)
        self._tiles    = tiles
        self._zoom     = zoom
        self._sig      = sig
        self._url_tmpl = url_tmpl
        self._cache    = cache_dir

    def run(self):
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=16) as pool:
            list(pool.map(self._one, self._tiles))

    def _one(self, tile):
        tx, ty = tile
        z      = self._zoom
        cache_path = self._cache / f"{z}_{tx}_{ty}.png"
        if cache_path.exists():
            pm = QPixmap(str(cache_path))
        else:
            url = self._url_tmpl.format(z=z, x=tx, y=ty)
            try:
                import urllib.request as _ur
                req  = _ur.Request(url, headers={"User-Agent": "VedderD100RouteManager/1.0"})
                data = _ur.urlopen(req, timeout=10).read()
                cache_path.write_bytes(data)
                pm = QPixmap(); pm.loadFromData(data)
            except Exception:
                return
        if not pm.isNull():
            self._sig.tile_ready.emit(z, tx, ty, pm)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Vedder D100 Route Manager")
        self.resize(1700, 960)
        self.data_root    = get_exe_dir() / "anonymized_output"
        self.dm           = load_distance_matrix(get_matrix_dir() / "distance_matrix.csv")
        self.dm_dur       = load_distance_matrix(get_matrix_dir() / "duration_matrix.csv")
        # Precompute set of all node keys known to the distance matrix for fast validation
        self._dm_keys: set = {k for pair in self.dm for k in pair}
        self._file_map    = {}
        self._month_map   = {}
        self._year_map    = {}
        self._cache       = {}
        self._mod_blocks  = None
        self._driver_start = None
        self._day_colour   = ""
        self._removed     = []      # list of farm dicts (with _from_block, _from_sheet keys)
        self._sheet_mods       = {}   # (fname, sname) -> mod_blocks (solver output)
        self._corrected_blocks = {}   # (fname, sname) -> corrected baseline (never overwritten by solver)
        self._loader      = None
        self._load_warnings = []   # per-sheet parse warnings from the current load
        # Solver UI state
        self._locked_sheet_cbs  = {}   # sname -> QCheckBox  (populated per file)
        self._demand_open_edits  = {}  # proc_key -> QLineEdit (HH:MM)
        self._demand_close_edits = {}  # proc_key -> QLineEdit (HH:MM)
        self._spin_chars  = ["|", "/", "-", "\\", "|", "/", "-", "\\"]
        self._spin_idx    = 0
        self._spin_timer  = QTimer(self)
        self._spin_timer.timeout.connect(self._tick_spinner)
        self._build_ui()
        self._scan_folders()
        # Wire the dropdown signals after the initial folder scan, so combos
        # populated during the scan don't fire their handlers. (Previously these
        # were connected from main() after construction, which had the same
        # effect; keeping them here makes MainWindow self-contained.)
        self.year_cb.currentIndexChanged.connect(self._on_year)
        self.month_cb.currentIndexChanged.connect(self._on_month)
        self.file_cb.currentIndexChanged.connect(self._on_file)
        self.sheet_cb.currentIndexChanged.connect(self._on_sheet)

    # -------------------------------------------------------------------------
    # UI construction
    # -------------------------------------------------------------------------

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(8,8,8,8); root.setSpacing(8)

        # Left control panel
        left = QFrame(); left.setFrameShape(QFrame.StyledPanel); left.setFixedWidth(170)
        ll = QVBoxLayout(left); ll.setContentsMargins(6,8,6,8); ll.setSpacing(4)
        bold = QFont(); bold.setBold(True)
        def lbl(t): l=QLabel(t); l.setFont(bold); return l

        # -- Folder browser ------------------------------------------------
        self.browse_btn = QPushButton("Browse for Folder...")
        self.browse_btn.setFixedHeight(28)
        self.browse_btn.setToolTip(
            "Choose the root data folder.\n"
            "Expected layout:\n"
            "  <root>/\n"
            "    \\__-- <year>/\n"
            "          \\__-- <month>/\n"
            "                \\__-- *.xlsx")
        self.browse_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:hover { background:#1976d2; }")
        self.browse_btn.clicked.connect(self._on_browse_folder)
        ll.addWidget(self.browse_btn)
        ll.addSpacing(4)

        for label, attr in [("Year","year_cb"),("Month","month_cb")]:
            ll.addWidget(lbl(label))
            cb = QComboBox(); setattr(self, attr, cb); ll.addWidget(cb); ll.addSpacing(2)

        # File dropdown + Load button immediately below it
        ll.addWidget(lbl("File"))
        self.file_cb = QComboBox(); ll.addWidget(self.file_cb)
        ll.addSpacing(2)

        self.load_btn = QPushButton("Load File")
        self.load_btn.setFixedHeight(28)
        self.load_btn.clicked.connect(self._on_load_clicked)
        ll.addWidget(self.load_btn)
        ll.addSpacing(6)

        # Sheet dropdown
        ll.addWidget(lbl("Sheet"))
        self.sheet_cb = QComboBox()
        # Use amber/orange selection so BLUE-route items remain distinguishable
        self.sheet_cb.setStyleSheet(
            "QComboBox QAbstractItemView::item:selected "
            "{ background: #e65100; color: white; }")
        ll.addWidget(self.sheet_cb)
        ll.addSpacing(2)

        # Day colour badge
        self.day_colour_box = QLabel("")
        self.day_colour_box.setAlignment(Qt.AlignCenter)
        self.day_colour_box.setFixedHeight(28)
        self.day_colour_box.setFont(bold)
        self.day_colour_box.setStyleSheet("border-radius: 4px; padding: 2px 6px;")
        ll.addWidget(self.day_colour_box)

        # Day / Night shift badge
        self.shift_type_box = QLabel("")
        self.shift_type_box.setAlignment(Qt.AlignCenter)
        self.shift_type_box.setFixedHeight(22)
        self.shift_type_box.setStyleSheet("border-radius: 3px; padding: 1px 6px; font-size: 8pt;")
        ll.addWidget(self.shift_type_box)
        ll.addSpacing(6)

        # Sheet date label (e.g. "Monday July 14, 2025")
        self.sheet_date_lbl = QLabel("")
        self.sheet_date_lbl.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        self.sheet_date_lbl.setWordWrap(True)
        self.sheet_date_lbl.setStyleSheet(
            "color: #555; font-size: 8pt; padding: 0 4px;")
        ll.addWidget(self.sheet_date_lbl)
        ll.addSpacing(8)

        # Timing assumptions box
        timing_title = QLabel("Timing Assumptions")
        timing_title.setFont(bold)
        ll.addWidget(timing_title)
        timing_frame = QFrame()
        timing_frame.setFrameShape(QFrame.NoFrame)
        timing_frame.setStyleSheet("""
            QFrame { background: #fafafa; border: 1px solid #e0e0e0;
                     border-radius: 6px; }
            QLabel { border: none; background: transparent; }
        """)
        tf = QVBoxLayout(timing_frame); tf.setContentsMargins(10,8,10,8); tf.setSpacing(5)
        lbl_font  = QFont()
        val_font  = QFont(); val_font.setBold(True)
        def add_timing_row(label, value, add_sep=True):
            row_w = QWidget(); row_w.setStyleSheet("background:transparent;")
            row_l = QHBoxLayout(row_w)
            row_l.setContentsMargins(0,0,0,0); row_l.setSpacing(4)
            lbl_w = QLabel(label); lbl_w.setFont(lbl_font)
            lbl_w.setStyleSheet("color:#555555;")
            lbl_w.setWordWrap(True)
            val_w = QLabel(value);  val_w.setFont(val_font)
            val_w.setAlignment(Qt.AlignRight|Qt.AlignVCenter)
            val_w.setStyleSheet("color:#222222;")
            row_l.addWidget(lbl_w, stretch=1); row_l.addWidget(val_w)
            tf.addWidget(row_w)
            if add_sep:
                sep = QFrame(); sep.setFrameShape(QFrame.HLine)
                sep.setStyleSheet("color:#e0e0e0; background:#e0e0e0; max-height:1px;")
                tf.addWidget(sep)
        add_timing_row("Setup time",   f"{int(ONSITE_MIN)} min / stop")
        add_timing_row("Pump rate",    f"{int(PUMP_RATE_LPM)} L / min")
        add_timing_row("Volume limit", f"{VOL_LIMIT:,} L", add_sep=False)
        ll.addWidget(timing_frame)

        # Push everything below to the bottom
        ll.addStretch()

        # -- Bottom action buttons ----------------------------------------------
        self.export_btn = QPushButton("Export to Excel...")
        self.export_btn.setFixedHeight(30)
        self.export_btn.setStyleSheet(
            "QPushButton { background:#43a047; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#a5d6a7; }")
        self.export_btn.clicked.connect(self._on_export_excel)
        ll.addWidget(self.export_btn)
        ll.addSpacing(3)

        self.reset_sheet_btn = QPushButton("Reset This Sheet")
        self.reset_sheet_btn.setFixedHeight(26)
        self.reset_sheet_btn.setStyleSheet(
            "QPushButton { background:#29b6f6; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#81d4fa; }")
        self.reset_sheet_btn.clicked.connect(self._on_reset_current_sheet)
        ll.addWidget(self.reset_sheet_btn)
        ll.addSpacing(3)

        self.reset_btn = QPushButton("Reset All to Original")
        self.reset_btn.setFixedHeight(26)
        self.reset_btn.setStyleSheet(
            "QPushButton { background:#e57373; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#ef9a9a; }")
        self.reset_btn.clicked.connect(self._on_reset_all)
        ll.addWidget(self.reset_btn)
        ll.addSpacing(4)

        self.irma_lookup_btn = QPushButton("Lookup...")
        self.irma_lookup_btn.setFixedHeight(28)
        self.irma_lookup_btn.setStyleSheet(
            "QPushButton { background:#6a1b9a; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#ce93d8; }")
        self.irma_lookup_btn.setToolTip(
            "Find which routes contain a given farm IRMA number.\n"
            "Searches all loaded files - both original and solver-modified routes.")
        self.irma_lookup_btn.clicked.connect(self._on_irma_lookup)
        ll.addWidget(self.irma_lookup_btn)
        ll.addSpacing(4)

        root.addWidget(left)

        # Tabs
        self.tabs = QTabWidget()
        root.addWidget(self.tabs, stretch=1)
        self.tabs.currentChanged.connect(self._on_tab_changed)

        self._build_route_tab(bold)
        self._build_comparison_tab(bold)

    def _build_route_tab(self, bold):
        route_tab = QWidget()
        rt = QVBoxLayout(route_tab); rt.setContentsMargins(0,0,0,0); rt.setSpacing(4)

        # -- IRMA Search bar --------------------------------------------------
        # Sits above both tables so it searches Original and Modified together.
        search_frame = QFrame()
        search_frame.setFrameShape(QFrame.StyledPanel)
        search_frame.setMaximumHeight(36)
        sl = QHBoxLayout(search_frame)
        sl.setContentsMargins(6, 3, 6, 3); sl.setSpacing(6)

        sl.addWidget(QLabel("IRMA Search:"))
        self._search_box = QLineEdit()
        self._search_box.setPlaceholderText("e.g. 71-117  (Enter to search, Esc to clear)")
        self._search_box.setFixedWidth(230)
        sl.addWidget(self._search_box)

        self._search_prev_btn = QPushButton("< Prev")
        self._search_next_btn = QPushButton("Next >")
        self._search_clear_btn = QPushButton("Clear")
        for btn in (self._search_prev_btn, self._search_next_btn, self._search_clear_btn):
            btn.setFixedHeight(24)
            btn.setFixedWidth(64)
            btn.setEnabled(False)
        self._search_next_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; } "
            "QPushButton:disabled { background:#90caf9; }")
        self._search_prev_btn.setStyleSheet(
            "QPushButton { background:#1565c0; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; } "
            "QPushButton:disabled { background:#90caf9; }")
        self._search_clear_btn.setStyleSheet(
            "QPushButton { background:#757575; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; } "
            "QPushButton:disabled { background:#bdbdbd; }")
        sl.addWidget(self._search_prev_btn)
        sl.addWidget(self._search_next_btn)
        sl.addWidget(self._search_clear_btn)

        self._search_status = QLabel("")
        self._search_status.setMinimumWidth(120)
        sl.addWidget(self._search_status)
        sl.addStretch()

        self._search_box.returnPressed.connect(self._on_search)
        self._search_box.textChanged.connect(self._on_search_text_changed)
        self._search_next_btn.clicked.connect(self._on_search_next)
        self._search_prev_btn.clicked.connect(self._on_search_prev)
        self._search_clear_btn.clicked.connect(self._on_search_clear)
        rt.addWidget(search_frame)

        # "View on Map" button — lives in the search bar row
        self._map_btn = QPushButton("🗺 View on Map")
        self._map_btn.setFixedHeight(28)
        self._map_btn.setStyleSheet(
            "QPushButton{background:#2e7d32;color:white;font-weight:bold;"
            "border-radius:4px;font-size:8pt;padding:0 10px;}"
            "QPushButton:disabled{background:#a5d6a7;}")
        self._map_btn.setToolTip(
            "Open a map showing the current sheet's route.\n"
            "Requires routes.db beside the exe.")
        self._map_btn.clicked.connect(self._on_view_on_map)
        search_frame.setMaximumHeight(999)  # allow taller if needed
        sl.addWidget(self._map_btn)

        # Internal search state - populated by _on_search()
        # List of (table, visual_row) for every match, in top-to-bottom order
        # across both tables (orig then edit).
        self._search_hits   = []   # [(table_widget, row_index), ...]
        self._search_cursor = -1   # index into _search_hits for current highlight

        top_split = QSplitter(Qt.Horizontal)

        # Original (read-only)
        orig_frame = QFrame(); orig_frame.setFrameShape(QFrame.StyledPanel)
        ol = QVBoxLayout(orig_frame); ol.setContentsMargins(4,4,4,4)
        lbl_o = QLabel("Original"); lbl_o.setFont(bold); ol.addWidget(lbl_o)
        self.orig_table = QTableWidget()
        self._init_route_table(self.orig_table, editable=False)
        ol.addWidget(self.orig_table)
        top_split.addWidget(orig_frame)

        # Editable (modified)
        edit_frame = QFrame(); edit_frame.setFrameShape(QFrame.StyledPanel)
        el = QVBoxLayout(edit_frame); el.setContentsMargins(4,4,4,4)
        lbl_e = QLabel("Modified  (drag farms or processors to reorder / drop to tray)"); lbl_e.setFont(bold)
        el.addWidget(lbl_e)
        self.edit_table = EditableRouteTable()
        self.edit_table.farm_removed.connect(self._on_farm_removed)
        self.edit_table.farm_inserted.connect(self._on_farm_inserted)
        self.edit_table.farm_reorder.connect(self._on_farm_reorder)
        self.edit_table.dest_removed.connect(self._on_dest_removed)
        self.edit_table.dest_reorder.connect(self._on_dest_reorder)
        self.edit_table.dest_inserted.connect(self._on_dest_inserted)
        self.edit_table.block_reorder.connect(self._on_block_reorder)
        self.edit_table.itemChanged.connect(self._on_mwo_changed)
        self.edit_table.itemSelectionChanged.connect(self._on_del_btn_state)
        el.addWidget(self.edit_table)
        top_split.addWidget(edit_frame)

        top_split.setSizes([800, 800])
        rt.addWidget(top_split, stretch=3)

        # Tray
        tray_frame = QFrame(); tray_frame.setFrameShape(QFrame.StyledPanel)
        tray_frame.setMaximumHeight(240)
        tl = QVBoxLayout(tray_frame); tl.setContentsMargins(4,4,4,4); tl.setSpacing(3)
        tray_hdr_row = QWidget()
        tray_hdr_l = QHBoxLayout(tray_hdr_row); tray_hdr_l.setContentsMargins(0,0,0,0); tray_hdr_l.setSpacing(6)
        lbl_t = QLabel("Removed Farms & Processors -- drag from Modified to remove | drag back to restore")
        lbl_t.setFont(bold); tray_hdr_l.addWidget(lbl_t, stretch=1)
        self._tray_del_btn = QPushButton("Delete Selected")
        self._tray_del_btn.setFixedHeight(24)
        self._tray_del_btn.setStyleSheet(
            "QPushButton { background:#e53935; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; } "
            "QPushButton:disabled { background:#ef9a9a; }")
        self._tray_del_btn.setEnabled(False)
        self._tray_del_btn.clicked.connect(self._on_tray_delete)
        tray_hdr_l.addWidget(self._tray_del_btn)

        self.add_block_btn = QPushButton("+ Add Block")
        self.add_block_btn.setFixedHeight(24)
        self.add_block_btn.setStyleSheet(
            "QPushButton { background:#7b1fa2; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; } "
            "QPushButton:disabled { background:#ce93d8; }")
        self.add_block_btn.clicked.connect(self._on_add_block)
        tray_hdr_l.addWidget(self.add_block_btn)
        tl.addWidget(tray_hdr_row)
        self.farm_tray = FarmTray()
        self.farm_tray.farm_incoming.connect(self._on_farm_removed)
        self.farm_tray.dest_incoming.connect(self._on_dest_removed)
        self.farm_tray.itemSelectionChanged.connect(self._on_del_btn_state)
        tl.addWidget(self.farm_tray)
        rt.addWidget(tray_frame, stretch=1)

        # -- Manual Add bars (Farm + Processor) -----------------------------
        add_frame = QFrame(); add_frame.setFrameShape(QFrame.StyledPanel)
        al = QVBoxLayout(add_frame); al.setContentsMargins(6,4,6,4); al.setSpacing(2)

        field_font = QFont()
        def _fe(placeholder, width=90):
            w = QLineEdit(); w.setPlaceholderText(placeholder)
            w.setFont(field_font); w.setFixedWidth(width); return w

        # -- Farm row ------------------------------------------------------
        farm_hdr = QLabel("Add Farm to Tray"); farm_hdr.setFont(bold)
        al.addWidget(farm_hdr)

        farm_row = QWidget()
        frl = QHBoxLayout(farm_row); frl.setContentsMargins(0,0,0,0); frl.setSpacing(4)

        # Editable combobox: type freely OR pick from known IRMAs.
        # Selecting a known IRMA autofills Train/Milking/EDPU/Location.
        self._add_irma = QComboBox()
        self._add_irma.setEditable(True)
        self._add_irma.setInsertPolicy(QComboBox.NoInsert)
        self._add_irma.setFixedWidth(80)
        self._add_irma.setFont(field_font)
        self._add_irma.lineEdit().setPlaceholderText("IRMA #")
        self._add_irma.lineEdit().setFont(field_font)
        self._add_irma.activated.connect(self._on_irma_autofill)

        self._add_train  = _fe("Train",  50)
        self._add_m1s    = _fe("M1 Start",62)
        self._add_m1f    = _fe("M1 Fin", 62)
        self._add_m2s    = _fe("M2 Start",62)
        self._add_m2f    = _fe("M2 Fin", 62)
        self._add_edpu   = _fe("EDPU",   50)
        self._add_loc    = _fe("Name",    105)
        self._add_vol    = _fe("Vol (L)", 76)

        for w in (self._add_irma, self._add_train,
                  self._add_m1s, self._add_m1f, self._add_m2s, self._add_m2f,
                  self._add_edpu, self._add_loc, self._add_vol):
            frl.addWidget(w)
        self._add_btn = QPushButton("Add Farm")
        self._add_btn.setFixedHeight(26); self._add_btn.setFixedWidth(86)
        self._add_btn.clicked.connect(self._on_manual_add)
        frl.addWidget(self._add_btn); frl.addStretch()
        al.addWidget(farm_row)

        # -- Processor row -------------------------------------------------
        sep2 = QFrame(); sep2.setFrameShape(QFrame.HLine)
        sep2.setStyleSheet("color:#ddd; background:#ddd; max-height:1px;")
        al.addWidget(sep2)

        proc_hdr = QLabel("Add Processor to Tray"); proc_hdr.setFont(bold)
        al.addWidget(proc_hdr)

        proc_row = QWidget()
        prl = QHBoxLayout(proc_row); prl.setContentsMargins(0,0,0,0); prl.setSpacing(4)

        # Editable combobox for proc key - populated from known processors in cache
        self._add_proc_key = QComboBox()
        self._add_proc_key.setEditable(True)
        self._add_proc_key.setInsertPolicy(QComboBox.NoInsert)
        self._add_proc_key.setMinimumWidth(220)
        self._add_proc_key.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self._add_proc_key.setFont(field_font)
        self._add_proc_key.lineEdit().setPlaceholderText("Proc Key (e.g. 901012)")
        self._add_proc_key.lineEdit().setFont(field_font)
        self._add_proc_key.activated.connect(self._on_proc_key_autofill)
        self._add_proc_key.lineEdit().editingFinished.connect(self._on_proc_key_editing_finished)
        self._add_proc_key.view().setMinimumWidth(350)

        self._add_proc_name = _fe("Name (e.g. Olympic - FA)", 200)
        self._add_proc_vol  = _fe("Partial Vol L (blank=rest)", 170)
        for w in (self._add_proc_key, self._add_proc_name, self._add_proc_vol):
            prl.addWidget(w)
        self._add_proc_btn = QPushButton("Add Processor")
        self._add_proc_btn.setFixedHeight(26); self._add_proc_btn.setFixedWidth(110)
        self._add_proc_btn.clicked.connect(self._on_manual_add_proc)
        prl.addWidget(self._add_proc_btn); prl.addStretch()
        al.addWidget(proc_row)

        self._add_status = QLabel("")
        self._add_status.setStyleSheet("color: #c0392b; font-size: 8pt;")
        al.addWidget(self._add_status)
        self._add_status_timer = QTimer(self)
        self._add_status_timer.setSingleShot(True)
        self._add_status_timer.timeout.connect(lambda: self._add_status.setText(""))

        rt.addWidget(add_frame)

        self.tabs.addTab(route_tab, "Route View")

    def _build_comparison_tab(self, bold):
        comp_tab = QWidget()
        cl = QVBoxLayout(comp_tab); cl.setContentsMargins(8,8,8,8); cl.setSpacing(6)

        hdr = QLabel("File-wide Original vs Modified - all loaded sheets")
        hdr.setFont(bold); cl.addWidget(hdr)

        self._comp_tables = {}   # key -> QTableWidget

        # Top section: Processor volumes (side-by-side, full width)
        proc_split = QSplitter(Qt.Horizontal)
        for side in ("orig", "mod"):
            w = QWidget(); wl = QVBoxLayout(w); wl.setContentsMargins(0,0,0,0); wl.setSpacing(2)
            lbl = QLabel("Original - Processor Volumes" if side=="orig"
                         else "Modified - Processor Volumes"); lbl.setFont(bold)
            wl.addWidget(lbl)
            pt = QTableWidget(); self._init_comp_table(pt)
            self._comp_tables[f"proc_{side}"] = pt
            wl.addWidget(pt)
            proc_split.addWidget(w)
        cl.addWidget(proc_split, stretch=2)

        # Sync processor tables
        self._comp_tables["proc_orig"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["proc_mod"].verticalScrollBar().setValue(v))
        self._comp_tables["proc_mod"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["proc_orig"].verticalScrollBar().setValue(v))

        # Bottom section: Per-sheet hours & km (side-by-side)
        sheet_split = QSplitter(Qt.Horizontal)
        for side in ("orig", "mod"):
            w = QWidget(); wl = QVBoxLayout(w); wl.setContentsMargins(0,0,0,0); wl.setSpacing(2)
            lbl = QLabel("Original - Sheet Summary" if side=="orig"
                         else "Modified - Sheet Summary"); lbl.setFont(bold)
            wl.addWidget(lbl)
            st = QTableWidget(); self._init_comp_table(st)
            self._comp_tables[f"sheet_{side}"] = st
            wl.addWidget(st)
            sheet_split.addWidget(w)
        cl.addWidget(sheet_split, stretch=3)

        # Sync sheet summary tables
        self._comp_tables["sheet_orig"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["sheet_mod"].verticalScrollBar().setValue(v))
        self._comp_tables["sheet_mod"].verticalScrollBar().valueChanged.connect(
            lambda v: self._comp_tables["sheet_orig"].verticalScrollBar().setValue(v))

        self.tabs.addTab(comp_tab, "Comparison")

        self._build_solver_tab(bold)

    def _build_solver_tab(self, bold):
        solver_tab = QWidget()
        sl = QVBoxLayout(solver_tab)
        sl.setContentsMargins(14, 10, 14, 10)
        sl.setSpacing(8)

        # -- Title row -----------------------------------------------------
        title_row = QWidget()
        tr_l = QHBoxLayout(title_row); tr_l.setContentsMargins(0,0,0,0); tr_l.setSpacing(12)
        title = QLabel("ALNS Route Solver")
        title.setStyleSheet("color: #1a3a5c;")
        tr_l.addWidget(title)
        subtitle = QLabel(
            "Optimises farm order across all RED and BLUE routes in the loaded file. "
            "OTHER routes (Grassfed, A2, etc.) are left unchanged."
        )
        subtitle.setWordWrap(True)
        subtitle.setStyleSheet("color: #555; font-size: 9pt;")
        tr_l.addWidget(subtitle, stretch=1)
        sl.addWidget(title_row)

        # -- helper --------------------------------------------------------
        def spin_row(label_text, widget, unit=""):
            row = QWidget()
            rl  = QHBoxLayout(row)
            rl.setContentsMargins(0, 0, 0, 0)
            lbl = QLabel(label_text)
            lbl.setMinimumWidth(155)
            rl.addWidget(lbl)
            rl.addWidget(widget)
            if unit:
                rl.addWidget(QLabel(unit))
            rl.addStretch()
            return row

        # ══════════════════════════════════════════════════════════════════
        # TOP BAND - Objective Weights | Constraints | ALNS | Run controls
        # ══════════════════════════════════════════════════════════════════
        top_band = QWidget()
        top_l = QHBoxLayout(top_band)
        top_l.setContentsMargins(0, 0, 0, 0)
        top_l.setSpacing(10)
        top_l.setAlignment(Qt.AlignTop)

        # -- Objective Weights --------------------------------------------
        obj_box = QGroupBox("Objective Weights")
        obj_box.setFont(bold)
        obj_l = QVBoxLayout(obj_box)
        obj_l.setSpacing(5)
        obj_l.setAlignment(Qt.AlignTop)

        self._sw_milking = QDoubleSpinBox()
        self._sw_milking.setRange(0.0, 20.0)
        self._sw_milking.setSingleStep(0.5)
        self._sw_milking.setValue(0.3)
        self._sw_milking.setDecimals(1)
        self._sw_milking.setToolTip(
            "Multiplier on milking-window wait time (in km-equivalent).\n"
            "0 = ignore milking windows entirely.")
        obj_l.addWidget(spin_row("Milking window x", self._sw_milking))

        self._sw_plant_win_pen = QDoubleSpinBox()
        self._sw_plant_win_pen.setRange(0.0, 10000.0)
        self._sw_plant_win_pen.setSingleStep(50.0)
        self._sw_plant_win_pen.setValue(20.0)
        self._sw_plant_win_pen.setDecimals(0)
        self._sw_plant_win_pen.setToolTip(
            "Penalty per hour the truck must wait outside a plant's receiving window.\n"
            "Expressed as km-equivalent - same scale as routing distance.\n"
            "200 km/h: a 1-hour wait costs as much as driving 200 extra km.\n"
            "Set to 0 to ignore receiving windows entirely.")
        obj_l.addWidget(spin_row("Plant window pen", self._sw_plant_win_pen, "km/h wait"))

        self._sw_plant_margin_mins = QDoubleSpinBox()
        self._sw_plant_margin_mins.setRange(0.0, 240.0)
        self._sw_plant_margin_mins.setSingleStep(5.0)
        self._sw_plant_margin_mins.setValue(30.0)
        self._sw_plant_margin_mins.setDecimals(0)
        self._sw_plant_margin_mins.setToolTip(
            "Minutes before plant closing to start penalising arrivals.\n"
            "Creates a slope inside the window so the solver prefers\n"
            "earlier arrivals and avoids cutting it close to closing time.\n"
            "0 = no margin penalty (only penalise outside the window).")
        obj_l.addWidget(spin_row("Close margin", self._sw_plant_margin_mins, "min"))

        self._sw_plant_margin_rate = QDoubleSpinBox()
        self._sw_plant_margin_rate.setRange(0.0, 10000.0)
        self._sw_plant_margin_rate.setSingleStep(50.0)
        self._sw_plant_margin_rate.setValue(25.0)
        self._sw_plant_margin_rate.setDecimals(0)
        self._sw_plant_margin_rate.setToolTip(
            "Penalty rate inside the closing margin (per minute of depth).\n"
            "Linear ramp: 0 at margin start, full rate at closing time.\n"
            "Default 250 with 30-min margin = max 7,500 penalty at closing.")
        obj_l.addWidget(spin_row("Margin rate", self._sw_plant_margin_rate, "/min"))

        self._sw_avoid_win_pen = QDoubleSpinBox()
        self._sw_avoid_win_pen.setRange(0.0, 10000.0)
        self._sw_avoid_win_pen.setSingleStep(50.0)
        self._sw_avoid_win_pen.setValue(0.0)
        self._sw_avoid_win_pen.setDecimals(0)
        self._sw_avoid_win_pen.setToolTip(
            "Flat penalty for arriving at a dest during one of its configured\n"
            "avoid-windows (currently: Saputo Abbotsford / 972712, 7-10pm -\n"
            "another division needs that dock during this slot).\n"
            "Expressed as km-equivalent, same scale as routing distance.\n"
            "Applies even though the plant is otherwise open during this time -\n"
            "independent of the receiving-window penalty above.\n"
            "Set to 0 to disable (default) - the solver will not try to avoid\n"
            "the window at all until this is raised above 0.")
        obj_l.addWidget(spin_row("Avoid-window pen", self._sw_avoid_win_pen, "km flat"))

        self._sw_overlap_pen = QDoubleSpinBox()
        self._sw_overlap_pen.setRange(0.0, 1000.0)
        self._sw_overlap_pen.setSingleStep(1.0)
        self._sw_overlap_pen.setValue(0.0)
        self._sw_overlap_pen.setDecimals(1)
        self._sw_overlap_pen.setToolTip(
            "Penalty per minute that a processor's dock capacity is exceeded\n"
            "by trucks running the SAME day's routes (both RED, or both BLUE).\n"
            "Most docks take 1 truck at a time; a couple (Saputo Port\n"
            "Coquitlam / Abbotsford) take 2 - see PROCESSOR_DOCK_CAPACITY.\n"
            "This is a capacity check, not a blanket no-overlap rule: at a\n"
            "2-bay dock, two trucks overlapping is free, only a third\n"
            "overlapping both of them gets penalized.\n"
            "Since RED/BLUE/GRASSFED run on entirely separate days and are\n"
            "solved independently, this can never fire between different\n"
            "colours - only between routes of the same colour group, which\n"
            "is the only case where it's physically possible.\n"
            "Expressed as km-equivalent per minute of capacity excess.\n"
            "Set to 0 to disable (default).")
        obj_l.addWidget(spin_row("Truck overlap pen", self._sw_overlap_pen, "km/min"))

        obj_l.addStretch()
        top_l.addWidget(obj_box)

        # -- Constraints ---------------------------------------------------
        con_box = QGroupBox("Constraints")
        con_box.setFont(bold)
        con_l = QVBoxLayout(con_box)
        con_l.setSpacing(5)
        con_l.setAlignment(Qt.AlignTop)

        self._sw_vol_tol = QDoubleSpinBox()
        self._sw_vol_tol.setRange(0.0, 1.0)
        self._sw_vol_tol.setSingleStep(0.01)
        self._sw_vol_tol.setValue(0.15)
        self._sw_vol_tol.setDecimals(2)
        self._sw_vol_tol.setToolTip(
            "Allowed deviation from original processor volume.\n"
            "0.15 = +/-15%.")
        con_l.addWidget(spin_row("Plant vol tolerance +/-", self._sw_vol_tol, "%*"))

        self._sw_vol_pen = QDoubleSpinBox()
        self._sw_vol_pen.setRange(0.0, 10000.0)
        self._sw_vol_pen.setSingleStep(1.0)
        self._sw_vol_pen.setValue(5.0)
        self._sw_vol_pen.setDecimals(1)
        self._sw_vol_pen.setToolTip(
            "Penalty per litre of processor volume outside the tolerance band.")
        con_l.addWidget(spin_row("Vol violation pen", self._sw_vol_pen, "/L"))

        self._sw_hard_cap = QSpinBox()
        self._sw_hard_cap.setRange(10000, 60000)
        self._sw_hard_cap.setSingleStep(500)
        self._sw_hard_cap.setValue(HARD_CAP)
        self._sw_hard_cap.setToolTip(
            "Hard truck capacity limit (litres).\n"
            "Routes exceeding this get a stiff per-litre penalty.")
        con_l.addWidget(spin_row("Truck cap limit", self._sw_hard_cap, "L"))

        self._sw_cap_pen = QDoubleSpinBox()
        self._sw_cap_pen.setRange(0.0, 1000.0)
        self._sw_cap_pen.setSingleStep(5.0)
        self._sw_cap_pen.setValue(2.0)
        self._sw_cap_pen.setDecimals(1)
        self._sw_cap_pen.setToolTip(
            "Penalty per litre over the truck capacity limit.\n"
            "Expressed as km-equivalent per litre.\n"
            "10 km/L: a 1,000L overload costs 10,000 km-eq.")
        con_l.addWidget(spin_row("Over-cap penalty", self._sw_cap_pen, "km/L"))

        self._sw_max_shift = QDoubleSpinBox()
        self._sw_max_shift.setRange(1.0, 24.0)
        self._sw_max_shift.setSingleStep(0.5)
        self._sw_max_shift.setValue(12.0)
        self._sw_max_shift.setDecimals(1)
        self._sw_max_shift.setToolTip("Maximum allowed shift length in hours.")
        con_l.addWidget(spin_row("Max shift length", self._sw_max_shift, "h"))

        self._sw_shift_pen = QDoubleSpinBox()
        self._sw_shift_pen.setRange(0.0, 1000.0)
        self._sw_shift_pen.setSingleStep(10.0)
        self._sw_shift_pen.setValue(50.0)
        self._sw_shift_pen.setDecimals(0)
        self._sw_shift_pen.setToolTip("Penalty per hour over the max shift limit.")
        con_l.addWidget(spin_row("Shift overage pen", self._sw_shift_pen, "/h"))

        self._sw_min_shift = QDoubleSpinBox()
        self._sw_min_shift.setRange(0.0, 24.0)
        self._sw_min_shift.setSingleStep(0.5)
        self._sw_min_shift.setValue(8.0)
        self._sw_min_shift.setDecimals(1)
        self._sw_min_shift.setToolTip("Minimum allowed shift length in hours.")
        con_l.addWidget(spin_row("Min shift length", self._sw_min_shift, "h"))

        self._sw_shift_under_pen = QDoubleSpinBox()
        self._sw_shift_under_pen.setRange(0.0, 1000.0)
        self._sw_shift_under_pen.setSingleStep(10.0)
        self._sw_shift_under_pen.setValue(30.0)
        self._sw_shift_under_pen.setDecimals(0)
        self._sw_shift_under_pen.setToolTip("Penalty per hour under the min shift limit.")
        con_l.addWidget(spin_row("Shift shortfall pen", self._sw_shift_under_pen, "/h"))

        self._sw_shift_hours = QDoubleSpinBox()
        self._sw_shift_hours.setRange(0.0, 100.0)
        self._sw_shift_hours.setSingleStep(1.0)
        self._sw_shift_hours.setValue(1.0)
        self._sw_shift_hours.setDecimals(1)
        self._sw_shift_hours.setToolTip(
            "Cost per hour of total shift length (applied to the full shift, not just overage).\n"
            "Higher values push the solver to favour shorter days overall.\n"
            "Default 5.0 = roughly equivalent to 6 km per hour of shift time.")
        con_l.addWidget(spin_row("Shift hours weight", self._sw_shift_hours, "/h"))

        self._sw_day_night_lock = QCheckBox("Lock farms to current day / night shift")
        self._sw_day_night_lock.setToolTip(
            "When enabled, the solver will not move a farm between a day route\n"
            "(start before noon) and a night route (start at noon or later).\n"
            "Preserves pickup timing consistency so prior-volume estimates\n"
            "remain reliable.  Routes starting before noon = Day; noon+ = Night.")
        con_l.addWidget(self._sw_day_night_lock)

        con_l.addStretch()
        top_l.addWidget(con_box)

        # -- Truck Availability (Day -> Night) ---------------------------------
        trk_box = QGroupBox("Truck Availability")
        trk_box.setFont(bold)
        trk_l = QVBoxLayout(trk_box)
        trk_l.setSpacing(5)
        trk_l.setAlignment(Qt.AlignTop)

        self._sw_truck_avail_chk = QCheckBox("Enforce day -> night truck return")
        self._sw_truck_avail_chk.setToolTip(
            "Penalise solutions where too few day-shift trucks return\n"
            "to depot before the earliest night-shift start time.\n"
            "Night start is auto-detected from the loaded file.\n"
            "Cutoff: routes starting before noon = Day, noon or later = Night.")
        trk_l.addWidget(self._sw_truck_avail_chk)

        self._sw_truck_avail_night_lbl = QLabel("Night start: (load a file)")
        self._sw_truck_avail_night_lbl.setStyleSheet("color: #555555; font-size: 8pt;")
        trk_l.addWidget(self._sw_truck_avail_night_lbl)

        self._sw_truck_avail_min_back = QSpinBox()
        self._sw_truck_avail_min_back.setRange(1, 50)
        self._sw_truck_avail_min_back.setValue(8)
        self._sw_truck_avail_min_back.setToolTip(
            "Total number of day-shift trucks that must return before\n"
            "the night shift starts.  Applied proportionally if RED and\n"
            "BLUE groups have different numbers of day routes.")
        trk_l.addWidget(spin_row("Min trucks back", self._sw_truck_avail_min_back))

        self._sw_truck_avail_pen = QDoubleSpinBox()
        self._sw_truck_avail_pen.setRange(100.0, 99999.0)
        self._sw_truck_avail_pen.setSingleStep(500.0)
        self._sw_truck_avail_pen.setValue(3000.0)
        self._sw_truck_avail_pen.setDecimals(0)
        self._sw_truck_avail_pen.setToolTip(
            "Penalty per truck short of the minimum return count.\n"
            "Set large enough to outweigh typical route-cost differences.")
        trk_l.addWidget(spin_row("Penalty / missing truck", self._sw_truck_avail_pen))

        self._btn_truck_avail_viz = QPushButton("View Return Timeline...")
        self._btn_truck_avail_viz.setToolTip(
            "Open a Gantt chart showing when each day-shift route returns\n"
            "to depot vs the night-shift start deadline.")
        self._btn_truck_avail_viz.clicked.connect(self._on_truck_avail_visualize)
        trk_l.addWidget(self._btn_truck_avail_viz)

        trk_l.addStretch()
        top_l.addWidget(trk_box)
        alns_box = QGroupBox("ALNS Parameters")
        alns_box.setFont(bold)
        alns_l = QVBoxLayout(alns_box)
        alns_l.setSpacing(5)
        alns_l.setAlignment(Qt.AlignTop)

        self._sw_iters = QSpinBox()
        self._sw_iters.setRange(50, 50000)
        self._sw_iters.setSingleStep(500)
        self._sw_iters.setValue(200)
        self._sw_iters.setToolTip(
            "Number of SA iterations per colour group.\n"
            "alpha is auto-computed so T decays to the cooling target over this many steps.")
        alns_l.addWidget(spin_row("Iterations per colour", self._sw_iters))

        self._sw_cool = QDoubleSpinBox()
        self._sw_cool.setRange(0.0001, 0.9999)
        self._sw_cool.setSingleStep(0.01)
        self._sw_cool.setValue(0.01)
        self._sw_cool.setDecimals(4)
        self._sw_cool.setToolTip(
            "Target temperature as a fraction of T0 at the final iteration.\n"
            "0.10 = T stays at 10% of T0 by the end - slow, exploratory cooling.\n"
            "0.001 = T collapses to near-zero - fast convergence.\n"
            "With T0~1,000 and 1,000 iters: 0.10 -> T_final~100, alpha~0.9977.\n"
            "alpha is auto-computed as  cool_target ^ (1 / iterations).")
        alns_l.addWidget(spin_row("Cooling target", self._sw_cool))

        self._sw_seed = QSpinBox()
        self._sw_seed.setRange(0, 2_000_000_000)
        self._sw_seed.setValue(0)
        self._sw_seed.setToolTip(
            "Random seed for the solver.\n"
            "0 = a fresh random seed each run (non-reproducible).\n"
            "Any non-zero value makes the run fully reproducible - the same\n"
            "input, settings, and seed always produce the same result, which\n"
            "is useful for debugging and comparing parameter changes.")
        alns_l.addWidget(spin_row("Random seed (0=random)", self._sw_seed))

        alns_l.addStretch()
        top_l.addWidget(alns_box)

        # -- Run controls (buttons + progress + status + footnote) ---------
        run_box = QGroupBox("Run")
        run_box.setFont(bold)
        run_l = QVBoxLayout(run_box)
        run_l.setSpacing(6)
        run_l.setAlignment(Qt.AlignTop)

        self._solve_btn = QPushButton("Run Solver")
        self._solve_btn.setFixedHeight(36)
        self._solve_btn.setStyleSheet(
            "QPushButton { background:#1e88e5; color:white; font-weight:bold; "
            "border-radius:4px; } "
            "QPushButton:disabled { background:#90caf9; }")
        self._solve_btn.clicked.connect(self._on_solve_clicked)
        run_l.addWidget(self._solve_btn)

        self._intra_btn = QPushButton("Optimise Within Routes")
        self._intra_btn.setFixedHeight(28)
        self._intra_btn.setStyleSheet(
            "QPushButton { background:#00897b; color:white; font-weight:bold; "
            "border-radius:4px; font-size:8pt; } "
            "QPushButton:disabled { background:#80cbc4; }")
        self._intra_btn.setToolTip(
            "Reorder farms within each route using 2-opt and or-opt until\n"
            "convergence. Applies directly to the Modified panel.\n"
            "No cross-route moves - only within-route reordering.")
        self._intra_btn.clicked.connect(self._on_intra_route_apply)
        run_l.addWidget(self._intra_btn)

        self._stop_btn = QPushButton("Stop")
        self._stop_btn.setFixedHeight(36)
        self._stop_btn.setEnabled(False)
        self._stop_btn.setStyleSheet(
            "QPushButton { background:#e53935; color:white; font-weight:bold; "
            "border-radius:4px; } "
            "QPushButton:disabled { background:#ef9a9a; }")
        self._stop_btn.clicked.connect(self._on_stop_solver)
        run_l.addWidget(self._stop_btn)

        self._solver_progress = QProgressBar()
        self._solver_progress.setRange(0, 100)
        self._solver_progress.setValue(0)
        self._solver_progress.setTextVisible(True)
        self._solver_progress.setFormat("%p%  %v / %m iters")
        run_l.addWidget(self._solver_progress)

        self._solver_status = QLabel("Ready")
        self._solver_status.setWordWrap(True)
        self._solver_status.setStyleSheet("color:#555; font-size:8pt;")
        run_l.addWidget(self._solver_status)

        note = QLabel(
            "* % displayed; internally stored as fraction.\n"
            "Results written to Modified panel on all RED/BLUE sheets."
        )
        note.setWordWrap(True)
        note.setStyleSheet("color:#888; font-size:8pt;")
        run_l.addWidget(note)
        run_l.addStretch()
        top_l.addWidget(run_box)

        sl.addWidget(top_band)

        # ══════════════════════════════════════════════════════════════════
        # BOTTOM BAND - Processor Demand | Locked Sheets | Solver Log
        # ══════════════════════════════════════════════════════════════════
        bottom_split = QSplitter(Qt.Horizontal)

        # -- Processor Demand targets (left) -------------------------------
        demand_box = QGroupBox("Processor Demand Targets & Receiving Windows")
        demand_box.setFont(bold)
        demand_l = QVBoxLayout(demand_box)
        demand_l.setSpacing(3)
        demand_note = QLabel(
            "Defaults loaded from current file. Edit volume or receiving window times.\n"
            "Times use HH:MM (24h). Leave both blank to treat as 24/7.")
        demand_note.setWordWrap(True)
        demand_note.setStyleSheet("color:#777; font-size:8pt;")
        demand_l.addWidget(demand_note)

        # Column header row
        hdr_font = QFont(); hdr_font.setBold(True)
        col_hdr = QWidget(); col_hdr_l = QHBoxLayout(col_hdr)
        col_hdr_l.setContentsMargins(2,0,2,0); col_hdr_l.setSpacing(4)
        for txt, w in [("Processor", 110), ("Volume", 88),
                       ("Open", 48), ("Close", 48)]:
            h = QLabel(txt); h.setFont(hdr_font); h.setFixedWidth(w)
            h.setAlignment(Qt.AlignCenter)
            col_hdr_l.addWidget(h)
        col_hdr_l.addStretch()
        demand_l.addWidget(col_hdr)

        demand_scroll = QScrollArea()
        demand_scroll.setWidgetResizable(True)
        demand_inner = QWidget()
        self._demand_layout = QVBoxLayout(demand_inner)
        self._demand_layout.setSpacing(2)
        self._demand_layout.setContentsMargins(2,2,2,2)
        demand_scroll.setWidget(demand_inner)
        demand_l.addWidget(demand_scroll)
        self._demand_spinboxes    = {}
        self._demand_open_edits   = {}
        self._demand_close_edits  = {}

        refresh_btn = QPushButton("Refresh from File")
        refresh_btn.setFixedHeight(24)
        refresh_btn.clicked.connect(self._refresh_demand_targets)
        demand_l.addWidget(refresh_btn)
        bottom_split.addWidget(demand_box)

        # -- Locked Sheets (centre) ----------------------------------------
        lock_box = QGroupBox("Locked Sheets")
        lock_box.setFont(bold)
        lock_l = QVBoxLayout(lock_box)
        lock_l.setSpacing(3)
        lock_note = QLabel(
            "Checked sheets are held constant by the solver - their farm order\n"
            "and processor assignments will not be changed.")
        lock_note.setWordWrap(True)
        lock_note.setStyleSheet("color:#777; font-size:8pt;")
        lock_l.addWidget(lock_note)

        lock_btn_row = QWidget(); lbr_l = QHBoxLayout(lock_btn_row)
        lbr_l.setContentsMargins(0,0,0,0); lbr_l.setSpacing(4)
        sel_all_btn  = QPushButton("Select All");      sel_all_btn.setFixedHeight(20)
        sel_all_btn.setStyleSheet("font-size:8pt;")
        clr_all_btn  = QPushButton("Clear All");       clr_all_btn.setFixedHeight(20)
        clr_all_btn.setStyleSheet("font-size:8pt;")
        rst_def_btn  = QPushButton("Reset Defaults");  rst_def_btn.setFixedHeight(20)
        rst_def_btn.setStyleSheet("font-size:8pt;")
        rst_def_btn.setToolTip(
            "Re-lock all default sheets (SOLVER_SKIP_SHEETS) and unlock any\n"
            "manually locked sheets, restoring the out-of-box state.")
        lbr_l.addWidget(sel_all_btn)
        lbr_l.addWidget(clr_all_btn)
        lbr_l.addWidget(rst_def_btn)
        lbr_l.addStretch()
        lock_l.addWidget(lock_btn_row)

        lock_scroll = QScrollArea()
        lock_scroll.setWidgetResizable(True)
        lock_inner = QWidget()
        self._lock_layout = QVBoxLayout(lock_inner)
        self._lock_layout.setSpacing(1)
        self._lock_layout.setContentsMargins(4,2,4,2)
        self._lock_layout.setAlignment(Qt.AlignTop)
        lock_scroll.setWidget(lock_inner)
        lock_l.addWidget(lock_scroll)
        self._locked_sheet_cbs = {}

        def _sel_all():
            for cb in self._locked_sheet_cbs.values(): cb.setChecked(True)
        def _clr_all():
            for cb in self._locked_sheet_cbs.values(): cb.setChecked(False)
        def _rst_def():
            for sname, cb in self._locked_sheet_cbs.items():
                cb.setChecked(sname.strip() in SOLVER_SKIP_SHEETS)
        sel_all_btn.clicked.connect(_sel_all)
        clr_all_btn.clicked.connect(_clr_all)
        rst_def_btn.clicked.connect(_rst_def)

        bottom_split.addWidget(lock_box)

        # -- Solver Log (right) --------------------------------------------
        log_w = QWidget()
        log_l = QVBoxLayout(log_w)
        log_l.setContentsMargins(4, 0, 0, 0)
        log_lbl = QLabel("Solver Log")
        log_lbl.setFont(bold)
        log_l.addWidget(log_lbl)
        self._solver_log = QTextEdit()
        self._solver_log.setReadOnly(True)
        self._solver_log.setFont(QFont("Courier New", 8))
        self._solver_log.setStyleSheet(
            "background:#1e1e1e; color:#d4d4d4; border-radius:4px;")
        log_l.addWidget(self._solver_log)
        bottom_split.addWidget(log_w)

        bottom_split.setStretchFactor(0, 1)   # demand - narrower
        bottom_split.setStretchFactor(1, 1)   # locked sheets
        bottom_split.setStretchFactor(2, 2)   # log - wider
        bottom_split.setSizes([280, 180, 820])
        sl.addWidget(bottom_split, stretch=1)

        self.tabs.addTab(solver_tab, "Solver")

        self._build_debug_tab(bold)

        # Internal solver reference
        self._solver_thread = None
        self._demand_spinboxes = {}   # initialised by _build_solver_tab

    # -- Debug tab -------------------------------------------------------------

    def _build_debug_tab(self, bold):
        """Raw data inspector: shows parsed blocks, timing results, and
        distance/duration lookups for the currently displayed sheet."""
        debug_tab = QWidget()
        dl = QVBoxLayout(debug_tab)
        dl.setContentsMargins(8, 8, 8, 8)
        dl.setSpacing(6)

        # Header + refresh button
        hdr_row = QWidget()
        hrl = QHBoxLayout(hdr_row); hrl.setContentsMargins(0,0,0,0); hrl.setSpacing(8)
        hdr_lbl = QLabel("Debug - Raw Block & Timing Data")
        hdr_lbl.setFont(bold)
        hrl.addWidget(hdr_lbl, stretch=1)

        refresh_btn = QPushButton("Refresh")
        refresh_btn.setFixedHeight(24)
        refresh_btn.setStyleSheet(
            "QPushButton { background:#546e7a; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        refresh_btn.clicked.connect(self._refresh_debug_tab)
        hrl.addWidget(refresh_btn)

        copy_btn = QPushButton("Copy")
        copy_btn.setFixedHeight(24)
        copy_btn.setStyleSheet(
            "QPushButton { background:#37474f; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        copy_btn.clicked.connect(self._copy_debug_text)
        hrl.addWidget(copy_btn)

        pdf_btn = QPushButton("Export PDF")
        pdf_btn.setFixedHeight(24)
        pdf_btn.setStyleSheet(
            "QPushButton { background:#4a148c; color:white; font-weight:bold; "
            "border-radius:3px; font-size:8pt; padding: 0 8px; }")
        pdf_btn.setToolTip("Export the current report to a print-ready PDF.")
        pdf_btn.clicked.connect(self._on_export_debug_pdf)
        hrl.addWidget(pdf_btn)

        dl.addWidget(hdr_row)

        # -- Tools: two rows of buttons ----------------------------------------
        tools_widget = QWidget()
        tools_vl = QVBoxLayout(tools_widget)
        tools_vl.setContentsMargins(0, 0, 0, 0)
        tools_vl.setSpacing(4)

        # ── Row 1: checkboxes + text reports ──────────────────────────────────
        row1 = QWidget()
        trl  = QHBoxLayout(row1); trl.setContentsMargins(0,0,0,0); trl.setSpacing(6)

        self._suppress_no_milking_cb = QCheckBox(
            "Suppress milking windows for 37-874, 14-247, 92-545")
        self._suppress_no_milking_cb.setChecked(False)
        self._suppress_no_milking_cb.setToolTip(
            "When checked, farms 37-874, 14-247 and 92-545 have no milking-window\n"
            "constraints - arrival times never cause waits and no conflict highlighting\n"
            "is shown for those farms.\n\n"
            "Uncheck to re-enable their windows (e.g. for verification).")
        self._suppress_no_milking_cb.stateChanged.connect(self._on_suppress_milking_changed)
        trl.addWidget(self._suppress_no_milking_cb)

        self._chk_route_opt = QCheckBox("Apply route corrections on load & after solve")
        self._chk_route_opt.setChecked(False)
        self._chk_route_opt.setToolTip(
            "When checked, applies two corrections to the Modified panel:\n"
            "  1. Optimize partial-dropoff positions (e.g. Ridgecrest, Farmhouse)\n"
            "     to find the farm split that best hits the plant receiving window.\n"
            "  2. Auto-flag farms with >2h waits in the original solution, giving\n"
            "     the solver a 2-hour pickup window anchored to the original arrival.\n\n"
            "Uncheck to fully revert Modified to match Original.")
        self._chk_route_opt.stateChanged.connect(self._on_route_opt_changed)
        trl.addWidget(self._chk_route_opt)
        self._chk_auto_flag = self._chk_route_opt
        self._chk_split_opt = self._chk_route_opt

        for label, style, tip, slot in [
            ("Plant Window Cost Report", "#6a1b9a",
             "For every route in the Modified panel, show the plant window penalty\n"
             "per processor destination - grouped by processor across all routes.",
             self._on_plant_window_report),
            ("Full Cost Report", "#1565c0",
             "Show complete cost breakdown for every route in the Modified panel:\n"
             "km, milking waits, shift, overtime, cap, plant window, per-block.",
             self._on_full_cost_report),
            ("Capacity Report", "#e65100",
             "List every route block whose total farm volume exceeds 41,500 L.\n"
             "Shows block-by-block volumes and how far over capacity each is.",
             self._on_capacity_report),
            ("Changelog", "#00695c",
             "Show which farms were added, removed, or moved between\n"
             "the Original and Modified panels for each route.",
             self._on_changelog_report),
            ("All Routes", "#37474f",
             "Concise one-line summary of every route: colour, start time,\n"
             "block count, farm count, and total volume.",
             self._on_all_routes_report),
            ("Route Listing", "#37474f",
             "Comprehensive printable listing: every route with each farm\n"
             "and processor stop listed per block.",
             self._on_route_listing_report),
        ]:
            btn = QPushButton(label)
            btn.setFixedHeight(24)
            btn.setStyleSheet(
                f"QPushButton {{ background:{style}; color:white; font-weight:bold; "
                f"border-radius:3px; font-size:8pt; padding: 0 8px; }}")
            btn.setToolTip(tip)
            btn.clicked.connect(slot)
            trl.addWidget(btn)
        trl.addStretch()
        tools_vl.addWidget(row1)

        # ── Row 2: visual / analysis tools ────────────────────────────────────
        row2 = QWidget()
        tr2  = QHBoxLayout(row2); tr2.setContentsMargins(0,0,0,0); tr2.setSpacing(6)

        for label, style, tip, slot in [
            ("Processor Schedule", "#00695c",
             "Visual chart: every truck's arrival-to-departure time at every\n"
             "processor across the loaded file, on one shared time axis.\n"
             "Highlights overlapping trucks at the same dock and shows any\n"
             "configured avoid-windows.",
             self._on_processor_schedule),
            ("Overtime Timeline", "#c62828",
             "For every route in the Modified panel that has overtime,\n"
             "show a stop-by-stop timeline: arrival, wait reason, departure,\n"
             "cumulative shift time. Diagnose gate waits vs milking vs distance.",
             self._on_overtime_timeline),
            ("Intra-Route Savings", "#00695c",
             "Exhaustively reorder farms within each route (2-opt until convergence)\n"
             "and report km and hours saved vs current Modified panel.\n"
             "Shows upper bound of what pure within-route reordering can achieve.",
             self._on_intra_route_savings),
            ("Block Capacity Distribution", "#e65100",
             "For every non-preload block in the Modified panel, compute peak load\n"
             "(matching the cap-penalty logic exactly: total farm vol for simple blocks,\n"
             "running peak through the stop sequence for split blocks).\n\n"
             "Reports: summary statistics, percentiles, histogram, threshold-sensitivity\n"
             "table, top-N heaviest blocks, and per-route maxima.\n\n"
             "Use this to choose hard_vol_cap from data, and to spot solver gaming\n"
             "(loads piling against the threshold rather than spread naturally).",
             self._on_block_capacity_distribution),
        ]:
            btn = QPushButton(label)
            btn.setFixedHeight(24)
            btn.setStyleSheet(
                f"QPushButton {{ background:{style}; color:white; font-weight:bold; "
                f"border-radius:3px; font-size:8pt; padding: 0 8px; }}")
            btn.setToolTip(tip)
            btn.clicked.connect(slot)
            tr2.addWidget(btn)
        tr2.addStretch()
        tools_vl.addWidget(row2)

        dl.addWidget(tools_widget)

        # Info strip
        self._debug_info = QLabel("")
        self._debug_info.setStyleSheet("color:#555; font-size:8pt;")
        dl.addWidget(self._debug_info)

        # Main text area (raw block/timing data)
        self._debug_text = QTextEdit()
        self._debug_text.setReadOnly(True)
        self._debug_text.setFont(QFont("Courier New", 8))
        self._debug_text.setStyleSheet(
            "background:#1e1e1e; color:#d4d4d4; border-radius:4px;")
        dl.addWidget(self._debug_text, stretch=1)

        self.tabs.addTab(debug_tab, "Debug")

        # ── Farm Summary tab ─────────────────────────────────────────────────
        farm_sum_tab = QWidget()
        fs_l = QVBoxLayout(farm_sum_tab)
        fs_l.setContentsMargins(8, 8, 8, 8)
        fs_l.setSpacing(6)

        # Filter row
        fs_top = QHBoxLayout()
        fs_top.addWidget(QLabel("Show:"))
        self._fs_filter = "ALL"
        self._fs_btns   = {}
        for lbl in ("All", "RED", "BLUE"):
            b = QPushButton(lbl)
            b.setCheckable(True)
            b.setChecked(lbl == "All")
            b.setFixedWidth(64)
            b.clicked.connect(lambda _, l=lbl: self._set_farm_summary_filter(l.upper()
                                                                              if l != "All" else "ALL"))
            self._fs_btns[lbl] = b
            fs_top.addWidget(b)
        fs_top.addSpacing(16)
        self._fs_count_lbl = QLabel("")
        self._fs_count_lbl.setStyleSheet("color:#555555; font-size:8pt;")
        fs_top.addWidget(self._fs_count_lbl)
        fs_top.addStretch()
        fs_l.addLayout(fs_top)

        # Table
        FS_COLS = ["IRMA", "Name", "M1 Start", "M1 End",
                   "M2 Start", "M2 End", "Volume (L)", "Sheets"]
        self._fs_table = QTableWidget(0, len(FS_COLS))
        self._fs_table.setHorizontalHeaderLabels(FS_COLS)
        self._fs_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._fs_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._fs_table.setAlternatingRowColors(True)
        self._fs_table.verticalHeader().setVisible(False)
        hdr = self._fs_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)   # Name stretches
        hdr.setSectionResizeMode(7, QHeaderView.Stretch)   # Sheets stretches
        self._fs_table.setStyleSheet("font-size: 8pt;")
        fs_l.addWidget(self._fs_table, stretch=1)

        self.tabs.insertTab(2, farm_sum_tab, "Farm Summary")

    def _set_farm_summary_filter(self, bucket):
        self._fs_filter = bucket
        for lbl, btn in self._fs_btns.items():
            btn.setChecked((lbl.upper() if lbl != "All" else "ALL") == bucket)
        self._refresh_farm_summary()

    def _refresh_farm_summary(self):
        """Rebuild the Farm Summary table from the currently loaded file."""
        fname = self.file_cb.currentText()
        self._fs_table.setRowCount(0)
        if not fname or fname not in self._cache:
            self._fs_count_lbl.setText("(no file loaded)")
            return

        bucket = self._fs_filter   # "ALL", "RED", or "BLUE"

        # Aggregate: {irma -> {name, m1_start, m1_finish, m2_start, m2_finish,
        #                       total_vol, sheets: set}}
        farms = {}
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            colour = entry.get("day_colour", "")
            if bucket != "ALL" and colour != bucket:
                continue
            for block in entry.get("blocks", []):
                for row in block.get("rows", []):
                    irma = str(row.get("irma") or "").strip()
                    if not irma:
                        continue
                    if irma not in farms:
                        name = (row.get("_extra_cells") or {}).get(18, "") or ""
                        farms[irma] = {
                            "name":      name,
                            "m1_start":  str(row.get("m1_start")  or "").strip(),
                            "m1_finish": str(row.get("m1_finish") or "").strip(),
                            "m2_start":  str(row.get("m2_start")  or "").strip(),
                            "m2_finish": str(row.get("m2_finish") or "").strip(),
                            "total_vol": 0.0,
                            "sheets":    set(),
                        }
                    vol = row.get("prior_vol") or 0
                    try:
                        farms[irma]["total_vol"] += float(vol)
                    except (TypeError, ValueError):
                        pass
                    farms[irma]["sheets"].add(sname)

        if not farms:
            self._fs_count_lbl.setText("No farms found.")
            return

        # Sort by IRMA (numeric part then alpha)
        def irma_sort_key(k):
            parts = k.replace("-", " ").split()
            try:    return (int(parts[0]), int(parts[1]) if len(parts) > 1 else 0)
            except: return (999999, 0)

        sorted_irmas = sorted(farms.keys(), key=irma_sort_key)

        CLR_MENN   = QColor("#e1bee7")
        CLR_WHITE  = QColor("#ffffff")
        CLR_ALT    = QColor("#f5f5f5")
        MENN_TIP   = "Mennonite farm — no pickup on Sunday."

        self._fs_table.setRowCount(len(sorted_irmas))
        for row_idx, irma in enumerate(sorted_irmas):
            d  = farms[irma]
            is_menn = irma in MENNONITE_FARMS
            bg = CLR_MENN if is_menn else (CLR_ALT if row_idx % 2 == 0 else CLR_WHITE)

            sheets_str = ", ".join(sorted(d["sheets"]))
            vol_str    = f"{int(d['total_vol']):,}" if d["total_vol"] else ""

            values = [
                irma,
                d["name"],
                d["m1_start"],
                d["m1_finish"],
                d["m2_start"],
                d["m2_finish"],
                vol_str,
                sheets_str,
            ]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem(str(val))
                item.setBackground(bg)
                if col_idx == 6:   # volume — right-align
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if col_idx in (2, 3, 4, 5):   # milking times — center
                    item.setTextAlignment(Qt.AlignCenter)
                if is_menn:
                    item.setToolTip(MENN_TIP)
                # IRMA bold
                if col_idx == 0:
                    f = item.font(); f.setBold(True); item.setFont(f)
                # ROBOT milking — italic grey
                if col_idx in (2, 3, 4, 5) and val.upper() == "ROBOT":
                    item.setForeground(QColor("#2e7d32"))
                    f = item.font(); f.setBold(True); item.setFont(f)
                self._fs_table.setItem(row_idx, col_idx, item)

        self._fs_count_lbl.setText(
            f"{len(sorted_irmas)} farm{'s' if len(sorted_irmas) != 1 else ''}  "
            f"({sum(1 for i in sorted_irmas if i in MENNONITE_FARMS)} Mennonite)"
        )

    def _collect_processor_schedule(self):
        """Build a list of processor visits across every sheet in the
        loaded file, using whatever blocks are currently active (solver
        output in self._sheet_mods if present, else the original parse).

        Returns list of dicts: {dest_key, dest_name, sname, colour,
        arr_min, dep_min} where arr_min/dep_min are continuous minutes on
        a shared axis (see _continuous_minutes) so every truck's visit
        lines up correctly regardless of which sheet it came from.
        """
        fname = self.file_cb.currentText()
        visits = []
        if not fname or fname not in self._cache:
            return visits
        suppress = self._suppress_no_milking_cb.isChecked() \
            if hasattr(self, "_suppress_no_milking_cb") else True

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict):
                continue
            start_time = entry.get("start_time")
            if not start_time:
                continue
            colour = entry.get("day_colour", "")
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))

            ct = calc_times(blocks, self.dm, start_time, self.dm_dur,
                            suppress_no_milking=suppress)
            if ct is None:
                continue
            all_times, _ = ct

            for b_idx, block in enumerate(blocks):
                btimes = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes:
                    continue
                dests = block.get("dests") or []
                if not dests:
                    dk0 = block.get("dest_key", "")
                    dn0 = block.get("dest_name", "") or dk0
                    dests = [{"key": dk0, "name": dn0}] if dk0 else []
                for d_i, dest_d in enumerate(dests):
                    dn = (dest_d.get("name", "") or "").strip()
                    if "yard for" in dn.lower():
                        continue   # overnight parking, not a real processor visit
                    dk = normalise_key(dest_d.get("key", "") or "")
                    if not dk:
                        continue
                    t_idx = _dest_stop_index(block, d_i, b_idx, blocks)
                    ft = btimes[t_idx] if t_idx < len(btimes) else None
                    if ft is None or ft.get("arr") is None or ft.get("dep") is None:
                        continue
                    arr_m = _continuous_minutes(ft["arr"], start_time)
                    dep_m = _continuous_minutes(ft["dep"], start_time)
                    if dep_m < arr_m:   # the dwell itself wrapped past midnight
                        dep_m += 24 * 60
                    visits.append({
                        "dest_key":  dk,
                        "dest_name": dn or dk,
                        "sname":     sname,
                        "colour":    colour,
                        "arr_min":   arr_m,
                        "dep_min":   dep_m,
                    })
        return visits

    def _on_processor_schedule(self):
        """Open the Processor Schedule chart for the currently loaded file."""
        try:
            self._on_processor_schedule_inner()
        except Exception:
            import traceback
            self._debug_text.setPlainText(
                f"Processor Schedule crashed:\n{traceback.format_exc()}")

    def _on_processor_schedule_inner(self):
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            QMessageBox.information(self, "Processor Schedule", "No file loaded.")
            return
        visits = self._collect_processor_schedule()
        if not visits:
            QMessageBox.information(
                self, "Processor Schedule",
                "No processor visits with usable arrival/departure times "
                "were found for this file.")
            return
        dlg = ProcessorScheduleDialog(visits, AVOID_WINDOWS, fname, parent=self,
                                      date_str=_sheets_date_str(self._cache, fname))
        dlg.exec_()

    def _on_full_cost_report(self):
        """Full cost breakdown for every route - uses same logic as solver via _sheet_cost_breakdown."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "avoid_window_penalty":  self._sw_avoid_win_pen.value(),
            "overlap_penalty":       self._sw_overlap_pen.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "min_shift_h":           self._sw_min_shift.value(),
            "shift_under_penalty":   self._sw_shift_under_pen.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }
        lines = ["Full Cost Report - Modified panel",
                 f"File: {fname}",
                 "=" * 70]
        grand = {k: 0.0 for k in ("km","milking","shift","overtime","cap","plant_win","total")}

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))
            frozen = all(_is_preload_block(b) or _is_fixed_vol_block(b) for b in blocks)
            tag    = " [FROZEN]" if frozen else ""

            # Detect partial dropoffs (split_after set on any dest)
            has_split = any(
                d.get("split_after") is not None
                for block in blocks
                for d in (block.get("dests") or [])
            )
            split_tag = " [SPLIT]" if has_split else ""

            bd = _sheet_cost_breakdown(blocks, self.dm, start_time, cfg, dm_dur=self.dm_dur)

            lines.append(f"\n{sname}{tag}{split_tag}  total={bd['total']:.1f}"
                         f"  km={bd['km']:.1f}  milk={bd['milking']:.1f}"
                         f"  shift={bd['shift']:.1f}  ot={bd['overtime']:.1f}"
                         f"  cap={bd['cap']:.1f}  pw={bd['plant_win']:.1f}  "
                         f"avoid={bd['avoid_win']:.1f}")
            if not frozen:   # frozen routes are irreducible - exclude from grand total
                for k in grand:
                    grand[k] += bd.get(k, 0.0)

        lines.append(f"\n{'='*70}")
        lines.append(f"GRAND TOTAL (non-frozen)  km={grand['km']:.1f}  milking={grand['milking']:.1f}"
                     f"  shift={grand['shift']:.1f}  overtime={grand['overtime']:.1f}"
                     f"  cap={grand['cap']:.1f}  plant_win={grand['plant_win']:.1f}"
                     f"  total={grand['total']:.1f}")
        self._debug_text.setPlainText("\n".join(lines))

    # -- Block-level capacity inspector ---------------------------------------
    def _on_block_capacity_distribution(self):
        """For every non-preload block, compute its peak load (the same value
        the cap-penalty calculation uses) and report the distribution.

        Logic mirrors the cap-penalty branches in _sheet_cost_breakdown
        (line ~2241): for split blocks we walk the stop sequence tracking a
        running load; for simple blocks the peak is the sum of farm prior_vol.

        The report has:
          1. Summary stats (count, mean, median, stdev, min/max)
          2. Percentiles (50, 75, 80, 85, 90, 95, 99)
          3. Histogram in 1000 L buckets
          4. Threshold sensitivity table (38k -> 45k, # blocks violating + L over)
          5. Top heaviest blocks (route, block index, peak)
          6. Per-route summary (max block load, # blocks, # over current cap)

        Use this to:
          - Pick hard_vol_cap based on actual loads, not a guess.
          - Detect solver gaming: a sharp spike at exactly current_cap-1 means
            the solver is piling loads against the threshold.  A clean tail
            beyond the chosen cap is what you want to see.
        """
        import statistics

        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            self._debug_text.setPlainText("(no file loaded)")
            return

        current_cap = float(self._sw_hard_cap.value())

        # blocks_data: list of dicts with peak, sheet, b_idx, kind, n_farms, n_dests, total_fv
        blocks_data = []
        skipped_preload = 0
        skipped_no_farms = 0

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict):
                continue
            start_time = entry.get("start_time")
            if not start_time:
                continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))

            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block):
                    skipped_preload += 1
                    continue

                farms = block.get("rows", [])
                if not farms:
                    skipped_no_farms += 1
                    continue

                total_fv = sum(
                    (r.get("prior_vol") or 0) for r in farms
                    if isinstance(r.get("prior_vol"), (int, float))
                )

                dests = block.get("dests") or []
                if not dests:
                    dk = block.get("dest_key", "")
                    dests = [{"key": dk, "vol_partial": None}] if dk else []

                # Mirror the peak-load logic from _sheet_cost_breakdown
                if any(d.get("split_after") is not None for d in dests):
                    is_last = (b_idx == len(blocks) - 1)
                    origin  = ("VEDDER" if b_idx == 0 else
                               (_block_last_dest_key(blocks[b_idx - 1]) or "VEDDER"))
                    stops   = _build_block_stops(block, origin, is_last)
                    running = peak = adel = 0.0
                    for stop in stops:
                        if stop["type"] == "farm":
                            v = stop["farm"].get("prior_vol") or 0
                            if isinstance(v, (int, float)):
                                running += v
                                peak = max(peak, running)
                        elif stop["type"] == "dest":
                            off = _dest_vol_partial(stop["dest"], total_fv, adel)
                            adel += off
                            running = max(0.0, running - off)
                    kind = "SPLIT"
                else:
                    peak = total_fv
                    kind = "SIMPLE"

                blocks_data.append({
                    "peak":     peak,
                    "sheet":    sname,
                    "b_idx":    b_idx,
                    "kind":     kind,
                    "n_farms":  len(farms),
                    "n_dests":  len(dests),
                    "total_fv": total_fv,
                })

        if not blocks_data:
            self._debug_text.setPlainText(
                "(no blocks to analyse - load a file with Modified panel data)")
            return

        peaks = [b["peak"] for b in blocks_data]
        peaks_sorted = sorted(peaks)
        n = len(peaks)

        def pct(xs_sorted, p):
            """Linear-interpolation percentile.  p in [0,100]."""
            if not xs_sorted:
                return 0.0
            if len(xs_sorted) == 1:
                return xs_sorted[0]
            k = (len(xs_sorted) - 1) * (p / 100.0)
            lo = int(k)
            hi = min(lo + 1, len(xs_sorted) - 1)
            frac = k - lo
            return xs_sorted[lo] * (1 - frac) + xs_sorted[hi] * frac

        # -- Build report -----------------------------------------------------
        lines = []
        lines.append("Block Capacity Distribution - Modified panel")
        lines.append("=" * 78)
        lines.append(f"File: {fname}")
        lines.append(f"Current hard_vol_cap setting: {current_cap:,.0f} L  (cap_penalty = {self._sw_cap_pen.value()}/L)")
        lines.append(f"Blocks analysed: {n}   skipped (preload): {skipped_preload}   skipped (empty): {skipped_no_farms}")

        # 1. Summary statistics
        mean   = statistics.mean(peaks)
        median = statistics.median(peaks)
        stdev  = statistics.stdev(peaks) if n > 1 else 0.0
        var    = statistics.variance(peaks) if n > 1 else 0.0
        lines.append("")
        lines.append("-" * 78)
        lines.append("SUMMARY STATISTICS (peak load per block, litres)")
        lines.append("-" * 78)
        lines.append(f"  count:    {n}")
        lines.append(f"  mean:     {mean:>10,.0f}")
        lines.append(f"  median:   {median:>10,.0f}")
        lines.append(f"  stdev:    {stdev:>10,.0f}")
        lines.append(f"  variance: {var:>14,.0f}")
        lines.append(f"  min:      {min(peaks):>10,.0f}")
        lines.append(f"  max:      {max(peaks):>10,.0f}")
        lines.append(f"  range:    {max(peaks) - min(peaks):>10,.0f}")

        # 2. Percentiles
        lines.append("")
        lines.append("-" * 78)
        lines.append("PERCENTILES")
        lines.append("-" * 78)
        for p in (10, 25, 50, 75, 80, 85, 90, 95, 97, 99):
            lines.append(f"  P{p:<2}: {pct(peaks_sorted, p):>10,.0f} L")

        # 3. Histogram in 1000 L buckets
        lines.append("")
        lines.append("-" * 78)
        lines.append("DISTRIBUTION HISTOGRAM (1000 L buckets, .=1 block, current cap marked)")
        lines.append("-" * 78)
        if peaks:
            bucket_size = 1000
            lo_bucket = (int(min(peaks)) // bucket_size) * bucket_size
            hi_bucket = (int(max(peaks)) // bucket_size + 1) * bucket_size
            buckets = {}
            for p in peaks:
                b = (int(p) // bucket_size) * bucket_size
                buckets[b] = buckets.get(b, 0) + 1
            max_count = max(buckets.values()) if buckets else 1
            bar_scale = max(1, max_count // 50)  # cap bar length around 50 chars
            for b in range(lo_bucket, hi_bucket + bucket_size, bucket_size):
                count = buckets.get(b, 0)
                bar_len = count // bar_scale + (1 if count > 0 else 0)
                bar = "#" * bar_len
                cap_marker = "  <- cap" if b <= current_cap < b + bucket_size else ""
                lines.append(f"  {b:>6,} – {b + bucket_size - 1:>6,}  ({count:>3})  {bar}{cap_marker}")
            if bar_scale > 1:
                lines.append(f"  (each # ~ {bar_scale} blocks)")

        # 4. Threshold sensitivity
        lines.append("")
        lines.append("-" * 78)
        lines.append("THRESHOLD SENSITIVITY  (if hard_vol_cap = T, what penalty results?)")
        lines.append("-" * 78)
        lines.append(f"  {'Threshold':>10}  {'#Over':>6}  {'%Over':>6}  {'L over':>10}  {'Penalty $':>12}  {'Worst block':>12}")
        cap_rate = self._sw_cap_pen.value()
        for T in (38000, 39000, 40000, 41000, 41500, 42000, 42500, 43000, 44000, 45000):
            over = [p for p in peaks if p > T]
            n_over = len(over)
            l_over = sum(p - T for p in over)
            penalty = l_over * cap_rate
            worst = max(over) - T if over else 0
            mark = "  <- current" if abs(T - current_cap) < 0.5 else ""
            lines.append(f"  {T:>10,}  {n_over:>6}  {(100.0*n_over/n):>5.1f}%  {l_over:>10,.0f}  {penalty:>12,.0f}  {worst:>12,.0f}{mark}")

        # 5. Top heaviest blocks
        lines.append("")
        lines.append("-" * 78)
        lines.append("TOP 20 HEAVIEST BLOCKS")
        lines.append("-" * 78)
        lines.append(f"  {'Rank':>4}  {'Sheet':>10}  {'Block':>5}  {'Kind':>6}  {'Farms':>5}  {'Dests':>5}  {'Total fv':>10}  {'Peak':>10}  {'vs cap':>10}")
        top = sorted(blocks_data, key=lambda b: -b["peak"])[:20]
        for i, b in enumerate(top, 1):
            delta = b["peak"] - current_cap
            delta_str = f"+{delta:,.0f}" if delta > 0 else f"{delta:,.0f}"
            lines.append(f"  {i:>4}  {b['sheet']:>10}  {b['b_idx']:>5}  {b['kind']:>6}  {b['n_farms']:>5}  {b['n_dests']:>5}  {b['total_fv']:>10,.0f}  {b['peak']:>10,.0f}  {delta_str:>10}")

        # 6. Per-route summary  - sheets sorted by max block descending
        lines.append("")
        lines.append("-" * 78)
        lines.append("PER-ROUTE SUMMARY (sheets with at least one block over current cap)")
        lines.append("-" * 78)
        per_sheet = {}
        for b in blocks_data:
            s = b["sheet"]
            d = per_sheet.setdefault(s, {"max": 0, "n_blocks": 0, "n_over": 0, "tot_over": 0.0})
            d["n_blocks"] += 1
            d["max"] = max(d["max"], b["peak"])
            if b["peak"] > current_cap:
                d["n_over"] += 1
                d["tot_over"] += (b["peak"] - current_cap)

        violating = [(s, d) for s, d in per_sheet.items() if d["n_over"] > 0]
        if not violating:
            lines.append(f"  (no routes have blocks over the current cap of {current_cap:,.0f})")
        else:
            lines.append(f"  {'Sheet':>10}  {'Max block':>10}  {'Blocks':>7}  {'#Over':>6}  {'L over':>10}  {'Penalty':>10}")
            for s, d in sorted(violating, key=lambda x: -x[1]["max"]):
                lines.append(f"  {s:>10}  {d['max']:>10,.0f}  {d['n_blocks']:>7}  {d['n_over']:>6}  {d['tot_over']:>10,.0f}  {d['tot_over'] * cap_rate:>10,.0f}")

        # 7. Diagnostic: solver gaming check
        lines.append("")
        lines.append("-" * 78)
        lines.append("GAMING DIAGNOSTIC")
        lines.append("-" * 78)
        # Count blocks within 200L below the current cap.  If the solver is gaming
        # the threshold, you'll see a noticeable spike here.
        near_under = [p for p in peaks if current_cap - 200 < p <= current_cap]
        in_500_under = [p for p in peaks if current_cap - 500 < p <= current_cap]
        in_500_over  = [p for p in peaks if current_cap < p <= current_cap + 500]
        lines.append(f"  Blocks in [cap-200, cap]:    {len(near_under):>3}")
        lines.append(f"  Blocks in [cap-500, cap]:    {len(in_500_under):>3}")
        lines.append(f"  Blocks in (cap, cap+500]:    {len(in_500_over):>3}")
        if len(in_500_under) > 0:
            ratio = len(in_500_under) / max(1, len(in_500_over))
            lines.append(f"  Ratio (under:over) within 500L of cap: {ratio:.2f}")
            if ratio > 4.0 and len(in_500_under) >= 3:
                lines.append("  (!) Sharp asymmetry - investigate whether solver is piling loads")
                lines.append("    against the threshold rather than spreading them.")
            elif ratio < 1.5:
                lines.append("  OK Loads are spread naturally around the threshold.")
            else:
                lines.append("  Loads moderately clustered below threshold (normal).")
        # Also report std-dev of just the over-cap blocks vs all blocks
        over_peaks = [p for p in peaks if p > current_cap]
        if len(over_peaks) >= 2:
            over_stdev = statistics.stdev(over_peaks)
            lines.append(f"  Stdev of over-cap blocks: {over_stdev:>7,.0f}  "
                         f"(vs overall stdev {stdev:,.0f})")

        self._debug_text.setPlainText("\n".join(lines))

    def _on_intra_route_savings(self):
        """Exhaustively 2-opt reorder farms within each route until convergence.
        Reports km and hours saved vs current Modified panel."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "avoid_window_penalty":  self._sw_avoid_win_pen.value(),
            "overlap_penalty":       self._sw_overlap_pen.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "min_shift_h":           self._sw_min_shift.value(),
            "shift_under_penalty":   self._sw_shift_under_pen.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }

        lines = ["Intra-Route Reordering Savings (2-opt until convergence)", "=" * 70]
        total_km_before = total_km_after = 0.0
        total_h_before  = total_h_after  = 0.0
        n_improved = 0

        locked_sheets = {sname for sname, cb in self._locked_sheet_cbs.items()
                         if cb.isChecked()}
        skip = SOLVER_SKIP_SHEETS | locked_sheets

        for sname in sorted(self._cache[fname].keys()):
            if sname.strip() in skip:
                continue
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = copy.deepcopy(self._sheet_mods.get(key, entry.get("blocks", [])))

            # Before: km and hours
            bd_before = _sheet_cost_breakdown(blocks, self.dm, start_time, cfg, dm_dur=self.dm_dur)
            km_b  = bd_before["km"]
            h_b   = bd_before["shift"]
            ot_b  = bd_before["overtime"]

            # 2-opt each non-frozen block until convergence
            improved = True
            while improved:
                improved = False
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    base_c = _sheet_cost(blocks, self.dm, start_time, cfg, dm_dur=self.dm_dur)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n - 1):
                        for j in range(i + 1, n):
                            trial_rows = rows[:i] + rows[i:j+1][::-1] + rows[j+1:]
                            blocks[b_idx] = dict(block, rows=trial_rows)
                            c = _sheet_cost(blocks, self.dm, start_time, cfg, dm_dur=self.dm_dur)
                            if c < best_c:
                                best_c    = c
                                best_rows = trial_rows[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True

                # Also try or-opt (single farm relocation)
                for b_idx, block in enumerate(blocks):
                    if _is_preload_block(block) or _is_fixed_vol_block(block):
                        continue
                    rows = block["rows"]
                    n    = len(rows)
                    if n < 2: continue
                    base_c    = _sheet_cost(blocks, self.dm, start_time, cfg, dm_dur=self.dm_dur)
                    best_rows = rows[:]
                    best_c    = base_c
                    for i in range(n):
                        farm = rows[i]
                        remaining = rows[:i] + rows[i+1:]
                        for j in range(len(remaining) + 1):
                            trial_rows = remaining[:j] + [farm] + remaining[j:]
                            blocks[b_idx] = dict(block, rows=trial_rows)
                            c = _sheet_cost(blocks, self.dm, start_time, cfg, dm_dur=self.dm_dur)
                            if c < best_c:
                                best_c    = c
                                best_rows = trial_rows[:]
                    blocks[b_idx] = dict(block, rows=best_rows)
                    if best_rows != rows:
                        improved = True

            bd_after = _sheet_cost_breakdown(blocks, self.dm, start_time, cfg, dm_dur=self.dm_dur)
            km_a  = bd_after["km"]
            h_a   = bd_after["shift"]
            ot_a  = bd_after["overtime"]

            total_km_before += km_b;  total_km_after += km_a
            total_h_before  += h_b;   total_h_after  += h_a

            frozen = all(_is_preload_block(b) or _is_fixed_vol_block(b) for b in blocks)
            if frozen: continue

            delta_km = km_a - km_b
            delta_h  = (h_a + ot_a) - (h_b + ot_b)
            if abs(delta_km) > 0.1 or abs(delta_h) > 0.01:
                n_improved += 1
                lines.append(
                    f"  {sname}  km:{km_b:.1f}->{km_a:.1f}({delta_km:+.1f})"
                    f"  hours:{h_b:.2f}->{h_a:.2f}({delta_h:+.2f})"
                    f"  ot:{ot_b:.1f}->{ot_a:.1f}")

        delta_km_total = total_km_after - total_km_before
        delta_h_total  = total_h_after  - total_h_before
        pct_km = (delta_km_total / total_km_before * 100) if total_km_before else 0
        pct_h  = (delta_h_total  / total_h_before  * 100) if total_h_before  else 0
        lines.append(f"\n{'='*70}")
        lines.append(f"Routes improved: {n_improved}")
        lines.append(
            f"TOTAL km:    {total_km_before:.1f} -> {total_km_after:.1f}"
            f"  ({delta_km_total:+.1f},  {pct_km:+.2f}%)")
        lines.append(
            f"TOTAL hours: {total_h_before:.2f} -> {total_h_after:.2f}"
            f"  ({delta_h_total:+.2f}h,  {pct_h:+.2f}%)")
        self._debug_text.setPlainText("\n".join(lines))

    def _on_overtime_timeline(self):
        """For every route with overtime, print a stop-by-stop timeline."""
        try:
            self._on_overtime_timeline_inner()
        except Exception:
            import traceback
            self._debug_text.setPlainText(
                f"Overtime Timeline crashed:\n{traceback.format_exc()}")

    def _on_overtime_timeline_inner(self):
        """For every route with overtime, print a stop-by-stop timeline."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "min_shift_h":           self._sw_min_shift.value(),
            "shift_under_penalty":   self._sw_shift_under_pen.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
        }
        max_sh  = cfg["max_shift_h"]
        suppress = cfg["suppress_no_milking"]
        lines = ["Overtime Timeline - Modified panel",
                 f"File: {fname}",
                 "=" * 70]
        n_overtime = 0

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))

            ct = calc_times(blocks, self.dm, start_time, self.dm_dur,
                            suppress_no_milking=suppress)
            if ct is None: continue
            all_times, end_dt = ct
            base_dt = datetime.combine(date.today(), start_time)
            shift_h = (end_dt - base_dt).total_seconds() / 3600.0
            if shift_h <= max_sh:
                continue   # no overtime - skip

            n_overtime += 1
            overtime_h = shift_h - max_sh
            lines.append(f"\n{sname}  shift={shift_h:.2f}h  overtime={overtime_h:.2f}h  "
                         f"start={fmt_hhmm(start_time)}")
            lines.append("-" * 60)

            for b_idx, block in enumerate(blocks):
                btimes  = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes: continue
                is_last = (b_idx == len(blocks) - 1)
                origin  = "VEDDER" if b_idx == 0 else (
                    _block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
                stops   = _build_block_stops(block, origin, is_last)

                if b_idx > 0:
                    lines.append(f"  --- Block {b_idx+1} ---")

                cumul_prev = 0.0
                for s_idx, stop in enumerate(stops):
                    if s_idx >= len(btimes): break
                    t = btimes[s_idx]
                    arr  = fmt_hhmm(t.get("arr"))
                    dep  = fmt_hhmm(t.get("dep"))
                    wait = t.get("wait")
                    stype = stop["type"]

                    # Cumulative shift time at this stop
                    if t.get("arr") is not None:
                        arr_dt2  = datetime.combine(date.today(), t["arr"])
                        if (arr_dt2 - base_dt).total_seconds() < -3600:
                            arr_dt2 += timedelta(days=1)
                        cumul_h = (arr_dt2 - base_dt).total_seconds() / 3600.0
                    else:
                        cumul_h = cumul_prev
                    cumul_prev = cumul_h
                    cumul_s = f"{cumul_h:.2f}h"
                    ot_flag = " <<OT" if cumul_h > max_sh else ""

                    if stype == "origin":
                        name = stop["key"]
                        lines.append(f"  ORIGIN  {name:<20}  dep={dep}  [{cumul_s}]")

                    elif stype == "farm":
                        farm  = stop["farm"]
                        irma  = farm.get("irma","?")
                        loc   = farm.get("location","")[:16]
                        vol   = farm.get("prior_vol", 0)
                        vol   = int(vol) if isinstance(vol, (int, float)) else 0
                        # Determine wait reason
                        if wait and wait > 0:
                            reason = ""
                            # MWO farms skip milking windows so any wait is
                            # from a gate or scheduling gap, not milking.
                            if farm.get("_mwo"):
                                reason = "gate-wait(MWO)"
                            else:
                                # Find which milking window fired
                                for sk, fk in [("m1_start","m1_finish"),
                                               ("m2_start","m2_finish")]:
                                    if farm.get(sk) and farm.get(fk):
                                        arr_t2 = t.get("arr")
                                        if arr_t2 is not None and time_in_window(
                                                arr_t2, farm[sk], farm[fk]):
                                            reason = f"milking({farm[sk]}-{farm[fk]})"
                                            break
                                if not reason:
                                    reason = "wait(unknown)"
                            wait_s = f"  WAIT={wait:.0f}m ({reason})"
                        else:
                            wait_s = ""
                        lines.append(
                            f"  FARM    {irma:<10} {loc:<16} {vol:>7,}L  "
                            f"arr={arr}  dep={dep}{wait_s}  [{cumul_s}{ot_flag}]")

                    elif stype == "dest":
                        d    = stop["dest"]
                        name = (d.get("name","") or d.get("key",""))[:24]
                        lines.append(
                            f"  DEST    {name:<24}  arr={arr}  dep={dep}  [{cumul_s}{ot_flag}]")

                    elif stype == "vedder":
                        lines.append(
                            f"  VEDDER  return                    arr={arr}  [{cumul_s}{ot_flag}]")

        if n_overtime == 0:
            lines.append("\nNo overtime routes found.")
        else:
            lines.append(f"\n{'='*70}")
            lines.append(f"Total routes with overtime: {n_overtime}")
        self._debug_text.setPlainText("\n".join(lines))

    def _on_capacity_report(self):
        """Report every block whose farm volume exceeds the 41,500 L tanker capacity."""
        CAPACITY = VOL_LIMIT
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        lines      = [f"Capacity Report  (limit: {CAPACITY:,} L) — Modified panel",
                      f"File: {fname}",
                      "=" * 65]
        over_count = 0
        ok_count   = 0

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict):
                continue
            blocks = self._sheet_mods.get((fname, sname),
                                          entry.get("blocks", []))
            sheet_header_written = False

            for b_idx, block in enumerate(blocks):
                # Sum prior_vol for all farm rows in this block
                block_vol = 0.0
                for row in block.get("rows", []):
                    try:
                        block_vol += float(row.get("prior_vol") or 0)
                    except (TypeError, ValueError):
                        pass

                if block_vol > CAPACITY:
                    over_count += 1
                    if not sheet_header_written:
                        colour = entry.get("day_colour", "")
                        lines.append(f"\n{sname}  [{colour}]")
                        sheet_header_written = True
                    over_by = block_vol - CAPACITY
                    farms_n = len(block.get("rows", []))
                    lines.append(
                        f"  Block {b_idx + 1:>2}:  {block_vol:>8,.0f} L  "
                        f"(+{over_by:,.0f} L over,  {farms_n} farms)"
                    )
                else:
                    ok_count += 1

        lines.append("")
        lines.append("─" * 65)
        if over_count == 0:
            lines.append(f"✓  All {ok_count} blocks are within capacity ({CAPACITY:,} L).")
        else:
            lines.append(
                f"⚠  {over_count} block{'s' if over_count != 1 else ''} over capacity  |  "
                f"{ok_count} within capacity.")

        self._debug_text.setPlainText("\n".join(lines))
        self.tabs.setCurrentWidget(self.tabs.widget(self.tabs.count() - 1))

    def _on_all_routes_report(self):
        """One-line-per-route summary of every sheet in the Modified panel."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        rows = []
        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict):
                continue
            colour = entry.get("day_colour", "")
            if colour not in ("RED", "BLUE"):
                continue
            st = entry.get("start_time")
            start_str = fmt_hhmm(st) if st else "--:--"
            blocks = self._sheet_mods.get((fname, sname), entry.get("blocks", []))
            n_blocks = len(blocks)
            n_farms  = sum(len(b.get("rows", [])) for b in blocks)
            vol = 0.0
            for b in blocks:
                for r in b.get("rows", []):
                    try:   vol += float(r.get("prior_vol") or 0)
                    except (TypeError, ValueError): pass
            rows.append((sname, colour, start_str, n_blocks, n_farms, vol))

        lines = ["All Routes Summary — Modified panel",
                 f"File: {fname}",
                 "=" * 62]

        if not rows:
            lines.append("No RED or BLUE routes found.")
            self._debug_text.setPlainText("\n".join(lines))
            return

        # Column header
        lines.append(
            f"  {'Sheet':<8}  {'Colour':<6}  {'Start':<7}  "
            f"{'Blk':>3}  {'Farms':>5}  {'Volume (L)':>12}")
        lines.append(
            f"  {'─'*8}  {'─'*6}  {'─'*7}  "
            f"{'─'*3}  {'─'*5}  {'─'*12}")

        red_rows  = [r for r in rows if r[1] == "RED"]
        blue_rows = [r for r in rows if r[1] == "BLUE"]

        for colour_label, group in [("RED", red_rows), ("BLUE", blue_rows)]:
            if not group:
                continue
            lines.append(f"\n  ── {colour_label} ──")
            for sname, colour, start, n_blk, n_farms, vol in group:
                over = "  ⚠" if vol > HARD_CAP else ("  !" if vol > VOL_LIMIT else "")
                lines.append(
                    f"  {sname:<8}  {colour:<6}  {start:<7}  "
                    f"{n_blk:>3}  {n_farms:>5}  {vol:>12,.0f}{over}")

        # Totals
        total_farms  = sum(r[4] for r in rows)
        total_vol    = sum(r[5] for r in rows)
        over_limit   = sum(1 for r in rows if r[5] > VOL_LIMIT)
        over_hard    = sum(1 for r in rows if r[5] > HARD_CAP)

        lines.append("")
        lines.append("─" * 62)
        lines.append(
            f"  Routes: {len(red_rows)} RED  {len(blue_rows)} BLUE  "
            f"({len(rows)} total)   "
            f"Farms: {total_farms}   "
            f"Volume: {total_vol:,.0f} L")
        if over_limit or over_hard:
            lines.append(
                f"  ⚠  {over_hard} route(s) over hard cap ({HARD_CAP:,} L)   "
                f"!  {over_limit} route(s) over soft limit ({VOL_LIMIT:,} L)")
        lines.append(
            f"  Avg volume/route: {total_vol/len(rows):,.0f} L   "
            f"Avg farms/route: {total_farms/len(rows):.1f}")

        self._debug_text.setPlainText("\n".join(lines))
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == "Debug":
                self.tabs.setCurrentIndex(i)
                break

    def _on_route_listing_report(self):
        """Comprehensive per-route farm and processor listing."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        def _vol(row):
            try:    return float(row.get("prior_vol") or 0)
            except: return 0.0

        lines = ["Route Listing — Modified panel",
                 f"File: {fname}",
                 "=" * 65]

        any_route = False
        red_sheets  = []
        blue_sheets = []

        for sname in sorted(self._cache[fname].keys()):
            entry = self._cache[fname].get(sname)
            if not isinstance(entry, dict):
                continue
            colour = entry.get("day_colour", "")
            if colour == "RED":
                red_sheets.append((sname, entry))
            elif colour == "BLUE":
                blue_sheets.append((sname, entry))

        for group_label, group in [("RED", red_sheets), ("BLUE", blue_sheets)]:
            if not group:
                continue
            lines.append(f"\n{'─'*65}")
            lines.append(f"  {group_label} ROUTES")
            lines.append(f"{'─'*65}")

            for sname, entry in group:
                any_route = True
                st = entry.get("start_time")
                start_str = fmt_hhmm(st) if st else "--:--"
                blocks    = self._sheet_mods.get(
                    (fname, sname), entry.get("blocks", []))

                n_farms = sum(len(b.get("rows", [])) for b in blocks)
                vol     = sum(
                    _vol(r)
                    for b in blocks for r in b.get("rows", [])
                    if r.get("prior_vol") is not None)

                vol_flag = "  ⚠" if vol > HARD_CAP else ("  !" if vol > VOL_LIMIT else "")
                lines.append(
                    f"\n  {sname}  ·  {start_str}  ·  {len(blocks)} block(s)  "
                    f"·  {n_farms} farms  ·  {vol:,.0f} L{vol_flag}")

                for b_idx, block in enumerate(blocks):
                    rows  = block.get("rows", [])
                    dests = block.get("dests", [])
                    is_preload = len(rows) == 0

                    b_vol = sum(
                        _vol(r) for r in rows
                        if r.get("prior_vol") is not None)

                    lines.append(
                        f"\n    Block {b_idx + 1}"
                        + (f"  [PRELOAD  {b_vol:,.0f} L]" if is_preload
                           else f"  [{b_vol:,.0f} L]"))

                    if rows:
                        for f_idx, row in enumerate(rows):
                            irma  = str(row.get("irma") or "").strip()
                            ec    = row.get("_extra_cells") or {}
                            name  = str(ec.get(18, "") or "").strip()
                            fvol  = row.get("prior_vol")
                            try:    fvol_str = f"{float(fvol):,.0f} L"
                            except: fvol_str = ""
                            m_flag = "  ★" if irma in MENNONITE_FARMS else ""
                            name_col = f"  {name}" if name else ""
                            lines.append(
                                f"      {f_idx+1:>2}.  {irma:<10}{name_col:<35}"
                                f"  {fvol_str:>10}{m_flag}")
                    else:
                        lines.append("      (no farms — preload block)")

                    if dests:
                        for dest in dests:
                            key  = dest.get("key") or dest.get("dest_key") or "?"
                            name = (dest.get("name") or "").strip()
                            proc_str = f"{name}  ({key})" if name else key
                            lines.append(f"      →  Processor: {proc_str}")

                lines.append("")

        if not any_route:
            lines.append("No RED or BLUE routes found.")

        # Totals
        all_sheets  = red_sheets + blue_sheets
        total_farms = sum(
            sum(len(b.get("rows", [])) for b in
                self._sheet_mods.get((fname, sn), e.get("blocks", [])))
            for sn, e in all_sheets)
        total_vol = sum(
            _vol(r)
            for sn, e in all_sheets
            for b in self._sheet_mods.get((fname, sn), e.get("blocks", []))
            for r in b.get("rows", [])
            if r.get("prior_vol") is not None)

        lines.append("═" * 65)
        lines.append(
            f"  {len(red_sheets)} RED  ·  {len(blue_sheets)} BLUE  ·  "
            f"{len(all_sheets)} routes total  ·  "
            f"{total_farms} farms  ·  {total_vol:,.0f} L")
        lines.append(
            f"  ★ = Mennonite (no Sunday pickup)   "
            f"! = over {VOL_LIMIT:,} L   ⚠ = over {HARD_CAP:,} L")

        self._debug_text.setPlainText("\n".join(lines))
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == "Debug":
                self.tabs.setCurrentIndex(i)
                break

    def _on_changelog_report(self):
        """Compare Original vs Modified panels and report per-route farm changes."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        # ── Build original and modified farm location maps ────────────────────
        # {irma: [(sname, block_idx, row_idx)]}
        orig_locs = {}
        mod_locs  = {}
        irma_names = {}   # {irma: name}

        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            orig_blocks = entry.get("blocks", [])
            mod_blocks  = self._sheet_mods.get((fname, sname), orig_blocks)

            for b_idx, block in enumerate(orig_blocks):
                for f_idx, row in enumerate(block.get("rows", [])):
                    irma = str(row.get("irma") or "").strip()
                    if irma:
                        orig_locs.setdefault(irma, []).append((sname, b_idx, f_idx))
                        if irma not in irma_names:
                            ec   = row.get("_extra_cells") or {}
                            name = str(ec.get(18, "") or "").strip()
                            if name:
                                irma_names[irma] = name

            for b_idx, block in enumerate(mod_blocks):
                for f_idx, row in enumerate(block.get("rows", [])):
                    irma = str(row.get("irma") or "").strip()
                    if irma:
                        mod_locs.setdefault(irma, []).append((sname, b_idx, f_idx))
                        if irma not in irma_names:
                            ec   = row.get("_extra_cells") or {}
                            name = str(ec.get(18, "") or "").strip()
                            if name:
                                irma_names[irma] = name

        # ── Diff ──────────────────────────────────────────────────────────────
        # changes[sname] = {"added": [...], "removed": [...], "moved": [...]}
        changes = {}

        def _entry(sname):
            changes.setdefault(sname, {"added": [], "removed": [], "moved": []})
            return changes[sname]

        def _fmt(irma):
            name = irma_names.get(irma, "")
            return f"{irma}  {name}" if name else irma

        all_irmas = sorted(set(orig_locs) | set(mod_locs))

        for irma in all_irmas:
            orig = orig_locs.get(irma, [])
            mod  = mod_locs.get(irma,  [])
            orig_sheets = {s for s, _, _ in orig}
            mod_sheets  = {s for s, _, _ in mod}

            # Removed entirely from a sheet
            for sname in orig_sheets - mod_sheets:
                if mod_sheets:
                    dest = next(iter(mod_sheets))
                    _entry(sname)["removed"].append(f"{_fmt(irma)}  → moved to {dest}")
                else:
                    _entry(sname)["removed"].append(_fmt(irma))

            # Added to a sheet
            for sname in mod_sheets - orig_sheets:
                if orig_sheets:
                    src = next(iter(orig_sheets))
                    _entry(sname)["added"].append(f"{_fmt(irma)}  ← from {src}")
                else:
                    _entry(sname)["added"].append(_fmt(irma))

            # Same sheet — check if block or position changed
            for sname in orig_sheets & mod_sheets:
                op = [(b, f) for s, b, f in orig if s == sname]
                mp = [(b, f) for s, b, f in mod  if s == sname]
                if op and mp and op != mp:
                    ob, of_ = op[0];  mb, mf = mp[0]
                    desc = (f"{_fmt(irma)}  "
                            f"block {ob+1} position {of_+1} → block {mb+1} position {mf+1}")
                    _entry(sname)["moved"].append(desc)

        # ── Render ────────────────────────────────────────────────────────────
        lines = ["Route Changelog — Original vs Modified",
                 f"File: {fname}",
                 "=" * 65]
        total_a = total_r = total_m = 0
        any_change = False

        for sname in sorted(changes.keys()):
            c = changes[sname]
            if not (c["removed"] or c["added"] or c["moved"]):
                continue
            any_change = True
            entry  = self._cache[fname].get(sname, {})
            colour = entry.get("day_colour", "") if isinstance(entry, dict) else ""
            lines.append(f"\n{sname}  [{colour}]")

            for desc in sorted(c["removed"]):
                lines.append(f"  -  {desc}")
                total_r += 1
            for desc in sorted(c["added"]):
                lines.append(f"  +  {desc}")
                total_a += 1
            for desc in sorted(c["moved"]):
                lines.append(f"  ↕  {desc}")
                total_m += 1

        lines.append("")
        lines.append("─" * 65)
        if not any_change:
            lines.append("No changes detected — Modified matches Original.")
        else:
            lines.append(
                f"Total:  +{total_a} added   -{total_r} removed   "
                f"↕{total_m} reordered")

        self._debug_text.setPlainText("\n".join(lines))
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == "Debug":
                self.tabs.setCurrentIndex(i)
                break

    def _on_plant_window_report(self):
        """Output per-processor, per-route plant window costs to the debug text area."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        cfg = {
            "plant_windows":      self._get_plant_windows(),
            "plant_win_penalty":  self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "avoid_window_penalty":  self._sw_avoid_win_pen.value(),
            "overlap_penalty":       self._sw_overlap_pen.value(),
            "suppress_no_milking": self._suppress_no_milking_cb.isChecked(),
        }
        plant_windows     = cfg.get("plant_windows", {})
        if not plant_windows:
            self._debug_text.setPlainText("No plant windows configured.")
            return

        plant_win_rate    = cfg.get("plant_win_penalty", 200.0)
        plant_margin_mins = cfg.get("plant_win_margin_mins", 30.0)
        plant_margin_rate = cfg.get("plant_win_margin_rate", plant_win_rate * 0.5)

        from collections import defaultdict
        dest_report = defaultdict(list)
        dest_names  = {}

        for sname, entry in sorted(self._cache[fname].items()):
            if not isinstance(entry, dict): continue
            start_time = entry.get("start_time")
            if not start_time: continue
            key    = (fname, sname)
            blocks = self._sheet_mods.get(key, entry.get("blocks", []))
            ct     = calc_times(blocks, self.dm, start_time, self.dm_dur,
                                suppress_no_milking=cfg.get("suppress_no_milking", True))
            if ct is None: continue
            all_times = ct[0]

            for b_idx, block in enumerate(blocks):
                if _is_preload_block(block): continue
                dests = block.get("dests") or []
                if not dests:
                    dk = block.get("dest_key","")
                    dn = block.get("dest_name","") or dk
                    dests = [{"key": dk, "name": dn}] if dk else []
                btimes  = all_times[b_idx] if b_idx < len(all_times) else None
                if not btimes: continue

                for d_i, dest_d in enumerate(dests):
                    dn  = (dest_d.get("name","") or "").strip()
                    dk  = normalise_key(dest_d.get("key","") or "")
                    dest_names[dk] = dn or dk
                    if "yard for" in dn.lower(): continue
                    window = plant_windows.get(dk)
                    if window is None: continue

                    t_idx = _dest_stop_index(block, d_i, b_idx, blocks)
                    ft    = btimes[t_idx] if t_idx < len(btimes) else None
                    if ft is None or ft.get("arr") is None: continue

                    arr    = ft["arr"]
                    arr_s  = f"{arr.hour:02d}:{arr.minute:02d}"
                    arr_dt = datetime.combine(date.today(), arr)
                    open_str, close_str = window

                    pen    = 0.0
                    status = "OK in window"
                    if not time_in_window(arr, open_str, close_str):
                        open_t = parse_hhmm(open_str)
                        if open_t:
                            open_dt = datetime.combine(date.today(), open_t)
                            if open_t <= arr: open_dt += timedelta(days=1)
                            wait_h  = (open_dt - arr_dt).total_seconds() / 3600.0
                        else:
                            wait_h = 1.0
                        pen    = wait_h * plant_win_rate
                        status = f"OUTSIDE  wait={wait_h:.2f}h  pen={pen:.1f}"
                    else:
                        close_t = parse_hhmm(close_str)
                        if close_t:
                            close_dt = datetime.combine(date.today(), close_t)
                            open_t2  = parse_hhmm(open_str)
                            if open_t2 and close_t < open_t2: close_dt += timedelta(days=1)
                            mins_to_close = (close_dt - arr_dt).total_seconds() / 60.0
                            if 0 < mins_to_close < plant_margin_mins:
                                depth  = (plant_margin_mins - mins_to_close) / plant_margin_mins
                                pen    = depth * plant_margin_rate * (plant_margin_mins / 60.0)
                                status = f"margin  {mins_to_close:.0f}m to close  pen={pen:.1f}"
                    dest_report[dk].append((sname, b_idx+1, arr_s, pen, status))

        lines       = ["Plant Window Cost Report - Modified panel",
                       f"File: {fname}",
                       "=" * 60]
        grand_total = 0.0
        for dk in sorted(dest_report.keys()):
            entries    = dest_report[dk]
            dn         = dest_names.get(dk, dk)
            window     = plant_windows.get(dk, ("?","?"))
            proc_total = sum(e[3] for e in entries)
            grand_total += proc_total
            lines.append(f"\n{dn}  [{dk}]  window={window[0]}-{window[1]}  total={proc_total:.1f}")
            lines.append("-" * 50)
            for sname, b_idx, arr_s, pen, status in sorted(entries, key=lambda x: -x[3]):
                marker = "  <--" if pen > 0 else ""
                lines.append(f"  {sname} b{b_idx}  arr={arr_s}  {status}{marker}")

        lines.append(f"\n{'='*60}")
        lines.append(f"GRAND TOTAL: {grand_total:.1f} km-eq")
        self._debug_text.setPlainText("\n".join(lines))

    def _refresh_debug_tab(self):
        """Rebuild the debug text from the currently active sheet's mod_blocks."""
        fname  = self.file_cb.currentText()
        sname  = self.sheet_cb.currentText()
        blocks = self._mod_blocks
        start  = getattr(self, "_driver_start", None)
        suppress = self._suppress_no_milking_cb.isChecked() \
            if hasattr(self, "_suppress_no_milking_cb") else True

        if not blocks:
            self._debug_text.setPlainText("(no blocks loaded)")
            self._debug_info.setText("")
            return

        lines = []
        lines.append(f"File : {fname}")
        lines.append(f"Sheet: {sname}   Start: {fmt_hhmm(start) if start else '?'}")
        lines.append(f"Suppress no-milking farms: {suppress}")
        lines.append(f"THREE_WINDOW_FARMS loaded : {len(THREE_WINDOW_FARMS)} entries")
        lines.append(f"Vedder depart extra       : +{VEDDER_DEPART_EXTRA_MINS} min")
        lines.append(f"Preload wash              : +{PRELOAD_WASH_MINS} min")
        lines.append(f"Inter-processor break     : +{INTER_PROCESSOR_BREAK} min")
        lines.append("")

        # Compute timing
        ct_result = calc_times(blocks, self.dm, start, dm_dur=self.dm_dur,
                               suppress_no_milking=suppress) if start else None
        all_times, end_cursor = ct_result if ct_result else (None, None)
        all_dists = calc_distances(blocks, self.dm)
        all_durs  = calc_durations(blocks, self.dm_dur)

        total_farms = sum(len(b["rows"]) for b in blocks)
        total_vol   = sum(
            (r["prior_vol"] or 0) for b in blocks for r in b["rows"]
            if isinstance(r.get("prior_vol"), (int, float))
        )
        lines.append(f"Blocks: {len(blocks)}   Farms: {total_farms}   "
                     f"Total vol: {int(total_vol):,} L")
        if end_cursor and start:
            from datetime import datetime, date
            base  = datetime.combine(date.today(), start)
            delta = end_cursor - base
            h, rem = divmod(int(delta.total_seconds()), 3600)
            m = rem // 60
            lines.append(f"Shift end: {fmt_hhmm(end_cursor.time())}   "
                         f"Duration: {h}h {m:02d}m")
        lines.append("")
        lines.append("-" * 72)

        for b_idx, block in enumerate(blocks):
            dists = all_dists[b_idx] if b_idx < len(all_dists) else []
            durs  = all_durs[b_idx]  if b_idx < len(all_durs)  else []
            btimes = all_times[b_idx] if (all_times and b_idx < len(all_times)) else None
            farms = block["rows"]
            dests = block.get("dests") or []
            if not dests:
                dk = block.get("dest_key","")
                dests = [{"name": block.get("dest_name",""), "key": dk, "vol_partial": None}] if dk else []

            is_preload = block.get("preload") and not farms
            tag = " [PRELOAD]" if is_preload else ""
            lines.append(f"BLOCK {b_idx+1}  Route: {block.get('route','?')}{tag}")

            # Origin
            origin_key = "VEDDER" if b_idx == 0 else (
                _block_last_dest_key(blocks[b_idx-1]) or "VEDDER")
            ot = btimes[0] if btimes else None
            origin_extra = f"  (+{VEDDER_DEPART_EXTRA_MINS}m Vedder depart)" if b_idx == 0 else ""
            lines.append(f"  Origin : {origin_key:<20} dep={fmt_hhmm(ot['dep'] if ot else None)}{origin_extra}")

            # Farm stops
            for i, farm in enumerate(farms):
                irma = farm.get("irma","?")
                vol  = farm.get("prior_vol")
                vol_s = f"{int(vol):,}L" if isinstance(vol, (int,float)) else "-"
                ft = btimes[i+1] if (btimes and i+1 < len(btimes)) else None
                arr_s = fmt_hhmm(ft["arr"] if ft else None)
                dep_s = fmt_hhmm(ft["dep"] if ft else None)
                wait  = ft["wait"] if ft else None
                wait_s = f"  wait={int(round(wait))}m" if wait else ""
                dist_s = f"{dists[i]:.1f}km" if i < len(dists) and dists[i] is not None else "-km"
                dur_s  = f"{durs[i]:.0f}m"   if i < len(durs)  and durs[i]  is not None else "-m"
                # Three windows
                w3data = THREE_WINDOW_FARMS.get(irma)
                w3_s   = f"  w3={w3data['w3'][0]}-{w3data['w3'][1]}" if w3data else ""
                no_m_s = "  [no-milking suppressed]" \
                    if (suppress and irma in NO_MILKING_WINDOW_FARMS) else ""
                lines.append(
                    f"  Farm {i+1:2d}: {irma}  {vol_s:<9}  "
                    f"arr={arr_s}  dep={dep_s}  "
                    f"({dist_s}/{dur_s}){wait_s}{w3_s}{no_m_s}"
                )

            # Processor / dest stops
            for d_i, dest_d in enumerate(dests):
                stop_idx = len(farms) + 1 + d_i
                dt = btimes[stop_idx] if (btimes and stop_idx < len(btimes)) else None
                arr_s = fmt_hhmm(dt["arr"] if dt else None)
                dep_s = fmt_hhmm(dt["dep"] if dt else None)
                vp    = dest_d.get("vol_partial")
                vp_s  = f"{int(vp):,}L partial" if isinstance(vp,(int,float)) else "rest"
                dist_v = dists[stop_idx-1] if stop_idx-1 < len(dists) and dists[stop_idx-1] is not None else None
                dur_v  = durs[stop_idx-1]  if stop_idx-1 < len(durs)  and durs[stop_idx-1]  is not None else None
                dist_s = f"{dist_v:.1f}km" if dist_v is not None else "-km"
                dur_s  = f"{dur_v:.0f}m"   if dur_v  is not None else "-m"
                is_yard = "yard for" in (dest_d.get("name","") or "").lower()
                tag2   = " [YARD]" if is_yard else ""
                lines.append(
                    f"  Proc  {d_i+1}: {dest_d.get('name') or dest_d.get('key','?'):<28}  "
                    f"arr={arr_s}  dep={dep_s}  ({dist_s}/{dur_s})  {vp_s}{tag2}"
                )

            # Vedder return (last block)
            if b_idx == len(blocks) - 1 and dests:
                v_idx = len(farms) + len(dests)
                vt = btimes[v_idx] if (btimes and v_idx < len(btimes)) else None
                lines.append(
                    f"  Return : VEDDER                      arr={fmt_hhmm(vt['arr'] if vt else None)}")

            # Preload wash note
            if is_preload:
                lines.append(f"  [wash {PRELOAD_WASH_MINS}m applied after preload offload]")

            lines.append("")

        self._debug_text.setPlainText("\n".join(lines))
        self._debug_info.setText(
            f"Sheet {sname} - {len(blocks)} block(s), {total_farms} farm(s)  "
            f"| dm keys: {len(self.dm)}  dur keys: {len(self.dm_dur)}"
        )

    def _on_export_debug_pdf(self):
        """Export the current Debug tab text to a print-ready PDF."""
        text = self._debug_text.toPlainText().strip()
        if not text:
            QMessageBox.information(self, "Export PDF",
                                    "Nothing to export — run a report first.")
            return
        # Use first line as title
        title = text.split("\n")[0][:60].strip() or "Report"
        _cur_fname = self.file_cb.currentText()
        _pdf_from_text(text, title, parent=self, fname=_cur_fname,
                       date_str=_sheets_date_str(self._cache, _cur_fname))

    def _copy_debug_text(self):
        from PyQt5.QtWidgets import QApplication as _QApp
        _QApp.clipboard().setText(self._debug_text.toPlainText())

    def _on_route_opt_changed(self, state):
        """Single handler for the combined route-corrections checkbox.
        Check  -> split optimization then auto-flag, snapshot into _corrected_blocks.
        Uncheck -> clear _corrected_blocks and fully revert _sheet_mods to raw originals."""
        fname = self.file_cb.currentText()
        if self._chk_route_opt.isChecked():
            self._optimize_all_split_positions()
            self._on_auto_flag_waits()
            # Snapshot BEFORE display so Original panel renders from corrected blocks
            self._snapshot_corrected_blocks(fname)
            self._display_sheet()
            if self.tabs.currentIndex() == 1:
                self._refresh_comparison()
        else:
            # Full revert - clear both stores
            if fname and fname in self._cache:
                keys = [k for k in self._sheet_mods if k[0] == fname]
                for k in keys:
                    del self._sheet_mods[k]
                corr_keys = [k for k in self._corrected_blocks if k[0] == fname]
                for k in corr_keys:
                    del self._corrected_blocks[k]
                self._display_sheet()
                if self.tabs.currentIndex() == 1:
                    self._refresh_comparison()

    def _snapshot_corrected_blocks(self, fname=None):
        """Copy current _sheet_mods into _corrected_blocks as the immutable baseline."""
        if fname is None:
            fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        for sname in self._cache[fname]:
            key = (fname, sname)
            if key in self._sheet_mods:
                self._corrected_blocks[key] = copy.deepcopy(self._sheet_mods[key])

    def _on_split_opt_checkbox_changed(self, state):
        pass  # replaced by _on_route_opt_changed

    def _on_auto_flag_checkbox_changed(self, state):
        pass  # replaced by _on_route_opt_changed

    def _on_suppress_milking_changed(self):
        """Re-render both tables when the suppress-milking option is toggled."""
        suppress = self._suppress_no_milking_cb.isChecked()
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        orig_entry = self._cache.get(fname, {}).get(sname)
        if orig_entry:
            corrections_on = (hasattr(self, "_chk_route_opt") and
                              self._chk_route_opt.isChecked())
            key = (fname, sname)
            if corrections_on and key in self._corrected_blocks:
                orig_display = copy.deepcopy(self._corrected_blocks[key])
            elif corrections_on and key in self._sheet_mods:
                orig_display = copy.deepcopy(self._sheet_mods[key])
            else:
                orig_display = orig_entry["blocks"]
            populate_table(self.orig_table, orig_display,
                           self.dm, editable=False,
                           start_time=self._driver_start, dm_dur=self.dm_dur,
                           suppress_no_milking=suppress,
                           plant_windows=self._get_plant_windows() if hasattr(self, "_get_plant_windows") else {})
        self._render_editable()
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    # -- Processor demand helpers ---------------------------------------------

    def _refresh_demand_targets(self):
        """Populate/refresh processor demand spinboxes and receiving-window fields
        from the current file totals.

        Each row now shows:
          [Processor name (140px)] [Volume spinbox] [Open HH:MM] [Close HH:MM]

        Receiving-window defaults come from PLANT_RECEIVING_WINDOWS keyed on the
        processor's numeric key, which is extracted from the processor name lookup
        against the blocks in the cache.
        """
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache: return

        # Aggregate current modified totals per processor
        proc_vols, _ = self._agg_file(fname, use_mod=True)
        totals = {p: sum(b.values()) for p, b in proc_vols.items()}

        # Build a map: processor display-name -> numeric key (from blocks)
        proc_key_map = {}   # display_name -> dest_key string
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict): continue
            for block in entry.get("blocks", []):
                for d in (block.get("dests") or []):
                    dk = normalise_key(d.get("key","") or "")
                    dn = d.get("name","") or dk
                    if dn and dk:
                        proc_key_map[dn] = dk

        # Clear old widgets
        for i in reversed(range(self._demand_layout.count())):
            w = self._demand_layout.itemAt(i).widget()
            if w: w.deleteLater()
        self._demand_spinboxes.clear()
        self._demand_open_edits.clear()
        self._demand_close_edits.clear()

        small_f = QFont()
        win_re  = re.compile(r'^\d{1,2}:\d{2}$')   # basic HH:MM check

        for proc in sorted(totals.keys()):
            row = QWidget(); rl = QHBoxLayout(row)
            rl.setContentsMargins(0,0,0,0); rl.setSpacing(4)

            lbl = QLabel(proc); lbl.setFont(small_f)
            lbl.setFixedWidth(110); lbl.setWordWrap(False)
            lbl.setToolTip(proc)

            spin = QDoubleSpinBox()
            spin.setRange(0, 10_000_000)
            spin.setSingleStep(1000)
            spin.setDecimals(0)
            spin.setValue(round(totals.get(proc, 0)))
            spin.setFont(small_f)
            spin.setFixedWidth(88)
            spin.setSuffix(" L")

            # Look up default receiving window via numeric dest key
            dk = proc_key_map.get(proc, "")
            win_default = PLANT_RECEIVING_WINDOWS.get(dk) if dk else None

            open_edit  = QLineEdit()
            close_edit = QLineEdit()
            open_edit.setFont(small_f);  open_edit.setFixedWidth(48)
            close_edit.setFont(small_f); close_edit.setFixedWidth(48)
            open_edit.setAlignment(Qt.AlignCenter)
            close_edit.setAlignment(Qt.AlignCenter)
            open_edit.setPlaceholderText("HH:MM")
            close_edit.setPlaceholderText("HH:MM")

            if win_default:
                open_str, close_str = win_default
                # Convert 00:00/23:59 sentinel back to empty (display as blank = 24/7)
                if open_str == "00:00" and close_str == "23:59":
                    open_edit.setPlaceholderText("24/7")
                    close_edit.setPlaceholderText("24/7")
                else:
                    open_edit.setText(open_str)
                    close_edit.setText(close_str)

            # Light validation colouring on edit
            def _validate(edit=open_edit):
                t = edit.text().strip()
                ok = (not t) or bool(win_re.match(t))
                edit.setStyleSheet("" if ok else "color:#c0392b;")
            open_edit.textChanged.connect(lambda _, e=open_edit: _validate(e))
            close_edit.textChanged.connect(lambda _, e=close_edit: _validate(e))

            rl.addWidget(lbl)
            rl.addWidget(spin)
            rl.addWidget(open_edit)
            rl.addWidget(close_edit)
            rl.addStretch()
            self._demand_layout.addWidget(row)

            self._demand_spinboxes[proc]    = spin
            self._demand_open_edits[proc]   = open_edit
            self._demand_close_edits[proc]  = close_edit

    def _get_demand_targets(self):
        """Return {proc_name: litres} from spinboxes, or None if empty."""
        if not self._demand_spinboxes: return None
        return {p: sb.value() for p, sb in self._demand_spinboxes.items()}

    def _get_plant_windows(self):
        """Build {dest_key: (open_str, close_str)} from the receiving-window
        fields in the Demand panel.

        Only processors where both fields are non-empty (or default 24/7) are
        included.  A processor with blank open AND blank close is treated as
        24/7 and omitted (no penalty applied).
        """
        # We need the key->name reverse map to translate display names back to keys
        fname = self.file_cb.currentText()
        proc_key_map = {}   # display_name -> dest_key
        if fname and fname in self._cache:
            for sname, entry in self._cache[fname].items():
                if not isinstance(entry, dict): continue
                for block in entry.get("blocks", []):
                    for d in (block.get("dests") or []):
                        dk = normalise_key(d.get("key","") or "")
                        dn = d.get("name","") or dk
                        if dn and dk:
                            proc_key_map[dn] = dk

        windows = {}
        win_re  = re.compile(r'^\d{1,2}:\d{2}$')
        for proc, open_edit in self._demand_open_edits.items():
            close_edit = self._demand_close_edits.get(proc)
            if close_edit is None:
                continue
            open_t  = open_edit.text().strip()
            close_t = close_edit.text().strip()
            if not open_t and not close_t:
                continue   # blank = 24/7, no window constraint
            # If only one is filled, fall back to PLANT_RECEIVING_WINDOWS default
            dk = proc_key_map.get(proc, "")
            if not dk:
                continue
            default = PLANT_RECEIVING_WINDOWS.get(dk)
            if not open_t:
                open_t  = default[0] if default else "00:00"
            if not close_t:
                close_t = default[1] if default else "23:59"
            # Validate; skip malformed entries
            open_t2  = f"{int(open_t.split(':')[0]):02d}:{open_t.split(':')[1]}" \
                        if win_re.match(open_t)  else None
            close_t2 = f"{int(close_t.split(':')[0]):02d}:{close_t.split(':')[1]}" \
                        if win_re.match(close_t) else None
            if open_t2 and close_t2:
                windows[dk] = (open_t2, close_t2)
        return windows

    def _refresh_locked_sheets_list(self):
        """Populate the Locked Sheets checkbox list from the currently loaded file."""
        # Clear old checkboxes
        for i in reversed(range(self._lock_layout.count())):
            w = self._lock_layout.itemAt(i).widget()
            if w: w.deleteLater()
        self._locked_sheet_cbs.clear()

        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        small_f = QFont()
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            dc = entry.get("day_colour", "")
            bucket = _sheet_colour_bucket(dc)
            # Only show RED/BLUE sheets - solver doesn't touch OTHER anyway
            if bucket not in ("RED", "BLUE"):
                continue

            cb = QCheckBox(sname)
            cb.setFont(small_f)
            # Colour the checkbox label to match the sheet colour
            bg, fg, _ = day_colour_style(dc)
            if bg:
                cb.setStyleSheet(
                    f"QCheckBox {{ color: {bg.name()}; font-weight: bold; }}")
            # Pre-tick sheets that are in SOLVER_SKIP_SHEETS (locked by default)
            if sname.strip() in SOLVER_SKIP_SHEETS:
                cb.setChecked(True)
                cb.setToolTip("Locked by default - uncheck to allow solver to modify.")
            self._lock_layout.addWidget(cb)
            self._locked_sheet_cbs[sname] = cb

    # -- Excel export ----------------------------------------------------------

    def _on_export_excel(self):
        fname = self.file_cb.currentText()
        fpath = self._file_map.get(fname)
        if not fpath:
            QMessageBox.warning(self, "Export", "No file loaded.")
            return
        out_path, _ = QFileDialog.getSaveFileName(
            self, "Export Modified Route Sheet", str(fpath.parent / ("modified_" + fname)),
            "Excel Files (*.xlsx)")
        if not out_path: return
        try:
            self.statusBar().showMessage(f"Exporting {fname}  ...")
            QApplication.processEvents()
            self._export_xlsx(fpath, out_path, fname)
            self.statusBar().showMessage(f"Export complete  ->  {out_path}", 6000)
            QMessageBox.information(self, "Export", "Saved to:\n" + out_path)
        except Exception as e:
            self.statusBar().showMessage(f"Export failed: {e}", 6000)
            QMessageBox.critical(self, "Export Error", str(e))

    def _export_xlsx(self, src_path, dst_path, fname):
        """Write a modified copy of the source workbook with _sheet_mods applied.

        Merged cells are read-only in openpyxl, so we unmerge all spans that
        touch the columns we need to write, write the values into the top-left
        cell of each former span, then re-merge everything afterwards.

        NOTE on load time: openpyxl has no API to load only specific sheets in
        writable mode - read_only=False always parses every sheet's full XML
        plus the shared style table for the entire workbook, regardless of how
        many sheets we actually modify.  On a large multi-sheet file (e.g. 79
        route sheets with a heavy shared style table) this load alone can take
        30+ seconds.  There is no supported way around this short of editing
        the underlying XML/zip directly, which is out of scope here.  We do
        take the one optimization that *is* available: data_only=False, since
        export only writes values and never needs the cached formula results
        that data_only=True computes.  We also surface progress so a slow
        export reads as "working" rather than "hung".
        """
        if getattr(self, "_debug_text", None) is not None:
            self._debug_text.append(
                f"[{fname}] Opening workbook for export (this can take 20-40s "
                f"on large multi-sheet files)...")
        import time as _time
        _t0 = _time.time()
        # data_only=False: export never reads cached values, only writes -
        # skipping data_only avoids computing the formula-value cache openpyxl
        # builds when data_only=True, which is pure overhead here.
        self.statusBar().showMessage(f"Exporting {fname}  -  loading workbook ...")
        QApplication.processEvents()
        wb = openpyxl.load_workbook(src_path, read_only=False, data_only=False)
        self.statusBar().showMessage(f"Exporting {fname}  -  writing changes ...")
        QApplication.processEvents()
        if getattr(self, "_debug_text", None) is not None:
            self._debug_text.append(
                f"[{fname}] Workbook opened in {_time.time()-_t0:.1f}s, writing changes...")

        WRITE_COLS = {C_IRMA, C_TRAIN, C_M1_START, C_M1_FINISH,
                      C_M2_START, C_M2_FINISH, C_EDPU, C_LOCATION, C_PRIOR_VOL}

        for ws in wb.worksheets:
            sname = ws.title
            if sname.strip().upper() in EXCLUDE_SHEETS: continue
            key = (fname, sname)
            if key not in self._sheet_mods: continue
            mod_blocks = self._sheet_mods[key]

            # -- Map farm row numbers to blocks ---------------------------------
            # Each IRMA# header in the sheet starts a new block (leg) that HAS
            # farm rows.  But mod_blocks may have a preload block at index 0
            # that has NO farm rows (just a dest) and appears before the first
            # IRMA# header.  If we don't account for that, farm_rows_by_block[0]
            # ends up holding the FIRST REAL block's farm rows, which then gets
            # matched against mod_blocks[0] (the preload block) - an off-by-one
            # that silently drops every farm's data on export.
            #
            # Detect this the same way the dest scan does: if there's a
            # "Delivery Information" section before the first IRMA# header,
            # that's a preload block with zero farm rows, and needs a
            # placeholder entry so list indices line up with mod_blocks.
            has_preload_block = False
            for r in range(1, min(ws.max_row, 5000) + 1):
                val0 = ws.cell(r, C_IRMA).value
                if isinstance(val0, str) and val0.strip().upper() == "IRMA#":
                    break   # reached the first real block - stop looking
                val2 = ws.cell(r, 2).value
                if isinstance(val2, str) and "delivery" in val2.lower():
                    has_preload_block = True
                    break

            farm_rows_by_block = []   # [[excel_row, ...], ...] per block
            if has_preload_block:
                farm_rows_by_block.append([])   # placeholder: preload has no farm rows

            current_farm_rows  = []
            in_block = False
            for r in range(1, min(ws.max_row, 5000) + 1):
                val0 = ws.cell(r, C_IRMA).value
                if isinstance(val0, str) and val0.strip().upper() == "IRMA#":
                    if in_block:
                        farm_rows_by_block.append(current_farm_rows)
                    current_farm_rows = []
                    in_block = True
                elif in_block and isinstance(val0, str) and IRMA_RE.match(val0.strip()):
                    current_farm_rows.append(r)
            if in_block:
                farm_rows_by_block.append(current_farm_rows)

            if not farm_rows_by_block:
                continue

            irma_ws_rows_all = [r for rows in farm_rows_by_block for r in rows]

            # -- Style donor row (font/border for new or newly-revealed cells) --
            # Two situations create cells with no trustworthy existing style:
            #   1. Rows we INSERT for overflow farms start out as bare openpyxl
            #      cells with no font/border at all.
            #   2. Cells revealed by splitting a wide ROBOT-style milking merge
            #      (e.g. E14:P14) into the standard four fields can carry stale
            #      leftover formatting from whatever the cell looked like before
            #      it was ever merged - observed on a real file as "MS Sans
            #      Serif" 10pt with no border, instead of the sheet's actual
            #      Calibri 13pt with thin/medium table borders.
            # Fix: scan for a normal (non-ROBOT, standard-merge) farm row in
            # this sheet and use its per-column font/border as the style
            # source for both cases below, so new/revealed cells match the
            # rest of the table instead of looking visually broken.  Font name
            # is always forced to Calibri (the sheet's actual font) even if a
            # donor cell's stored font name is something else, while size/
            # bold/italic/colour are preserved from the donor.
            style_donor_row = None
            for _r in irma_ws_rows_all:
                _is_wide = any(
                    mr.min_row == _r and mr.max_row == _r and
                    mr.min_col <= C_M1_START and mr.max_col >= C_M2_START and
                    (mr.max_col - mr.min_col + 1) > 3
                    for mr in ws.merged_cells.ranges)
                if not _is_wide:
                    style_donor_row = _r
                    break

            def _calibri_font(donor_font):
                return Font(name="Calibri",
                           size=donor_font.size or 13,
                           bold=donor_font.bold, italic=donor_font.italic,
                           color=donor_font.color, underline=donor_font.underline)

            def _copy_border(donor_border):
                """Rebuild a plain Border from a donor cell's border.

                openpyxl returns border (and other style) attributes as a
                StyleProxy wrapper in some access paths, which isn't directly
                assignable to another cell (it's unhashable, so the shared
                style table rejects it).  Reading each side's style/colour and
                constructing a fresh Border avoids that entirely.
                """
                def _side(s):
                    if s is None or s.style is None:
                        return Side(style=None)
                    return Side(style=s.style, color=s.color)
                return Border(left=_side(donor_border.left),
                              right=_side(donor_border.right),
                              top=_side(donor_border.top),
                              bottom=_side(donor_border.bottom))

            def _apply_table_style(dst_row, col, src_row=None):
                """Copy font (forced to Calibri) + border from the style donor
                row (or src_row if given) onto (dst_row, col).  No-op if no
                donor row was found (extremely unlikely - would require every
                farm row on the sheet to be a ROBOT placeholder)."""
                donor_row = src_row if src_row is not None else style_donor_row
                if donor_row is None:
                    return
                donor_cell = ws.cell(donor_row, col)
                dst_cell   = ws.cell(dst_row, col)
                dst_cell.font   = _calibri_font(donor_cell.font)
                dst_cell.border = _copy_border(donor_cell.border)

            # extra cols across all mod blocks
            mod_rows_all = [f for b in mod_blocks for f in b.get("rows", [])]
            extra_cols_in_use = set()
            for fd in mod_rows_all:
                extra_cols_in_use.update((fd.get("_extra_cells") or {}).keys())
            extra_cols_in_use -= WRITE_COLS
            sheet_write_cols = WRITE_COLS | extra_cols_in_use

            # -- Unmerge ----------------------------------------------------
            target_cells = set()
            for ws_row in irma_ws_rows_all:
                for col in sheet_write_cols:
                    target_cells.add((ws_row, col))

            # Build row -> farm dict so we know, before re-merging, whether
            # each farm row will end up holding a real farm with four
            # independent milking times, or a "ROBOT" placeholder.
            # Some sheets use a single wide merge (e.g. E14:P14) spanning all
            # four milking columns to show one "ROBOT" label instead of four
            # separate time cells.  If the solver slots a real farm with real
            # milking windows into that row, re-merging back to the wide span
            # would silently hide three of its four time values (Excel only
            # shows the top-left cell of a merge) even though they were
            # correctly written underneath.  We detect this case and split
            # the wide merge into the standard four 3-column merges instead,
            # but only for rows that will hold a real (non-ROBOT) farm.
            MILKING_COL_LO = C_M1_START
            row_to_farm = {}
            for b_idx, block in enumerate(mod_blocks):
                if b_idx >= len(farm_rows_by_block):
                    break
                slot_rows   = farm_rows_by_block[b_idx]
                block_farms = block.get("rows", [])
                for i, ws_row in enumerate(slot_rows):
                    if i < len(block_farms):
                        row_to_farm[ws_row] = block_farms[i]

            merges_to_redo = []          # restored as-is: (min_r,min_c,max_r,max_c)
            milking_split_rows = []      # wide merge -> split to standard 4-field
            milking_combine_rows = set() # standard 4-field -> combine to wide merge
            for merge_range in list(ws.merged_cells.ranges):
                mr = merge_range
                overlaps = any(
                    (r, c) in target_cells
                    for r in range(mr.min_row, mr.max_row + 1)
                    for c in range(mr.min_col, mr.max_col + 1)
                )
                if not overlaps:
                    continue

                # Detect a wide ROBOT-style milking merge: single row, spans
                # more than the standard 3-column width and covers the whole
                # milking-time region (M1 start through M2 area).
                is_wide_milking_merge = (
                    mr.min_row == mr.max_row and
                    mr.min_col <= MILKING_COL_LO and
                    mr.max_col >= C_M2_START and
                    (mr.max_col - mr.min_col + 1) > 3
                )
                if is_wide_milking_merge:
                    farm = row_to_farm.get(mr.min_row)
                    is_robot_placeholder = (
                        farm is not None and
                        str(farm.get("m1_start", "")).strip().upper() == "ROBOT"
                    )
                    if farm is not None and not is_robot_placeholder:
                        milking_split_rows.append(mr.min_row)
                        continue   # split instead of restoring - see below

                # Detect one of the standard 3-column milking merges (the
                # normal per-field shape).  If a ROBOT farm is being slotted
                # into this row, we need to combine all four standard merges
                # into one wide span instead of restoring four separate ones -
                # the inverse of the wide-merge-split case above.
                is_standard_milking_merge = (
                    mr.min_row == mr.max_row and
                    mr.min_col in (C_M1_START, C_M1_FINISH, C_M2_START, C_M2_FINISH) and
                    (mr.max_col - mr.min_col + 1) == 3
                )
                if is_standard_milking_merge:
                    farm = row_to_farm.get(mr.min_row)
                    if farm is not None and \
                            str(farm.get("m1_start", "")).strip().upper() == "ROBOT":
                        milking_combine_rows.add(mr.min_row)
                        continue   # combine instead of restoring - see below

                merges_to_redo.append(
                    (mr.min_row, mr.min_col, mr.max_row, mr.max_col))

            # Snapshot every original single-row merge BEFORE we start
            # unmerging.  Inserted rows (below) copy their non-milking merge
            # layout (IRMA / name / location / prior_vol / etc.) from their
            # block's template row - but by the time the insertion loop runs,
            # Phase A has already unmerged all the write-column merges, so
            # reading them off the live worksheet would miss exactly the ones
            # we need.  Capture them here while they still exist.
            orig_row_merges = {}
            for mr in ws.merged_cells.ranges:
                if mr.min_row == mr.max_row:
                    orig_row_merges.setdefault(mr.min_row, []).append(
                        (mr.min_col, mr.max_col))

            for (min_r, min_c, max_r, max_c) in merges_to_redo:
                ws.unmerge_cells(
                    start_row=min_r, start_column=min_c,
                    end_row=max_r,   end_column=max_c)
            for row in milking_split_rows:
                # The wide merge spanned this whole row's milking region -
                # unmerging it (already done above via the overlap check would
                # have skipped it, so do it explicitly here) leaves four plain
                # cells; we'll re-merge them as four standard 3-column pairs
                # after writing, instead of restoring the wide span.
                for mr in list(ws.merged_cells.ranges):
                    if (mr.min_row == row and mr.max_row == row and
                            mr.min_col <= MILKING_COL_LO and
                            mr.max_col >= C_M2_START):
                        ws.unmerge_cells(start_row=row, start_column=mr.min_col,
                                         end_row=row,   end_column=mr.max_col)
                        break
            for row in milking_combine_rows:
                # The four standard 3-column merges were skipped above (we
                # need to combine them into one wide span instead) - unmerge
                # each of them explicitly here.
                for mr in list(ws.merged_cells.ranges):
                    if (mr.min_row == row and mr.max_row == row and
                            mr.min_col in (C_M1_START, C_M1_FINISH,
                                          C_M2_START, C_M2_FINISH) and
                            (mr.max_col - mr.min_col + 1) == 3):
                        ws.unmerge_cells(start_row=row, start_column=mr.min_col,
                                         end_row=row,   end_column=mr.max_col)

            # -- Insert rows if any block has more farms than slots -----------
            # Process last-to-first so that earlier blocks' row numbers don't
            # need updating yet.  Track cumulative offset to shift earlier
            # blocks' rows after all insertions are done.
            total_inserted = [0] * len(farm_rows_by_block)   # rows inserted per block
            insertions = []   # [(first_inserted_row, count), ...] for row shifting
            for b_idx in range(len(farm_rows_by_block) - 1, -1, -1):
                if b_idx >= len(mod_blocks):
                    continue
                slot_rows   = farm_rows_by_block[b_idx]
                block_farms = mod_blocks[b_idx].get("rows", [])
                extra = len(block_farms) - len(slot_rows)
                if extra <= 0:
                    continue
                if not slot_rows:
                    # No existing farm rows to use as a template or insertion
                    # anchor (e.g. a preload block that never had farm rows in
                    # the original sheet).  Can't safely fabricate a farm
                    # section here - flag it and skip rather than crash.
                    self._export_warnings = getattr(self, '_export_warnings', [])
                    self._export_warnings.append(
                        f"Sheet '{sname}' block {b_idx+1}: solver added "
                        f"{len(block_farms)} farm(s) to a block that had no "
                        f"farm rows in the original sheet - could not write "
                        f"(no template row available).")
                    continue

                total_inserted[b_idx] = extra
                insert_after = slot_rows[-1]
                template_row = slot_rows[-1]

                # CRITICAL: openpyxl's insert_rows() shifts cell *values* down
                # but does NOT move merged-cell ranges.  Every merge below the
                # insertion point therefore stays `extra` rows too high and now
                # sits on top of the rows we just inserted.  On save those stale
                # merges silently clobber the farm data we write into the new
                # rows - a full-width section merge (e.g. A:AX on the "TOTAL
                # VOLUME" row) blanks the entire row, and a milking merge hides
                # cells under it.  This is the root of "milking times disappear
                # when the solver adds farms to a block."
                #
                # Fix: capture every merge below the insertion point and unmerge
                # it BEFORE inserting (the worksheet is still internally
                # consistent here - unmerging AFTER insert_rows raises KeyError
                # on the shifted merged-cell stubs), insert the rows, then
                # re-merge each captured range `extra` rows lower so the merge
                # model matches the shifted data.  The template row's
                # non-milking merges come from the snapshot taken before Phase A
                # unmerged them (reading live ranges here would miss every
                # write-column merge: IRMA, name, location, prior_vol).
                non_milking_tmpl_merges = [
                    (c1, c2)
                    for (c1, c2) in orig_row_merges.get(template_row, [])
                    if c2 < C_M1_START or c1 > C_M2_FINISH + 2
                ]
                below = [
                    (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
                    for mr in list(ws.merged_cells.ranges)
                    if mr.min_row > insert_after
                ]
                for (mnr, mnc, mxr, mxc) in below:
                    ws.unmerge_cells(start_row=mnr, start_column=mnc,
                                     end_row=mxr, end_column=mxc)

                ws.insert_rows(insert_after + 1, extra)
                insertions.append((insert_after + 1, extra))

                for (mnr, mnc, mxr, mxc) in below:
                    ws.merge_cells(start_row=mnr + extra, start_column=mnc,
                                   end_row=mxr + extra, end_column=mxc)

                # Number formats: copy column-by-column from the template row.
                # Font + border: copy from the style donor row (a normal,
                # non-ROBOT farm row) so inserted rows match the sheet's
                # actual Calibri table styling instead of showing as bare,
                # unbordered openpyxl-default cells.  Falls back to the
                # template row's own styling if no donor was found anywhere
                # on the sheet (would require every farm row to be ROBOT).
                _style_src_row = style_donor_row if style_donor_row is not None \
                                 else template_row
                for offset in range(extra):
                    new_row = insert_after + 1 + offset
                    for col in range(1, ws.max_column + 1):
                        tmpl_cell = ws.cell(template_row, col)
                        new_cell  = ws.cell(new_row, col)
                        if tmpl_cell.number_format and \
                                tmpl_cell.number_format != "General":
                            new_cell.number_format = tmpl_cell.number_format
                        _apply_table_style(new_row, col, src_row=_style_src_row)

                # Merges for the inserted rows.  We can't blindly copy the
                # template row's merges: the template (the block's last original
                # slot) might be a blank or ROBOT row whose milking region is a
                # single wide merge (E:P), which would hide three of the four
                # time values for any real farm placed in a new row.  So copy
                # only the template's NON-milking merges (captured above), and
                # build the milking region to match the farm that will actually
                # occupy each new row: the standard four 3-column fields for a
                # real farm, or one wide span for a ROBOT placeholder.  The
                # generic unmerge/re-write/re-merge passes below then handle
                # these the same as pre-existing rows.
                MK_LO = C_M1_START
                MK_HI = C_M2_FINISH + 2
                for offset in range(extra):
                    new_row  = insert_after + 1 + offset
                    farm_idx = len(slot_rows) + offset
                    farm = (block_farms[farm_idx]
                            if farm_idx < len(block_farms) else None)
                    is_robot = (farm is not None and
                                str(farm.get("m1_start", "")).strip().upper()
                                == "ROBOT")
                    for (min_c, max_c) in non_milking_tmpl_merges:
                        try:
                            ws.merge_cells(start_row=new_row, start_column=min_c,
                                           end_row=new_row,   end_column=max_c)
                        except Exception:
                            pass
                    try:
                        if is_robot:
                            ws.merge_cells(start_row=new_row, start_column=MK_LO,
                                           end_row=new_row,   end_column=MK_HI)
                        else:
                            for c0 in (C_M1_START, C_M1_FINISH,
                                       C_M2_START, C_M2_FINISH):
                                ws.merge_cells(start_row=new_row, start_column=c0,
                                               end_row=new_row, end_column=c0 + 2)
                    except Exception:
                        pass
                # Extend this block's slots with the new rows (insert_after is
                # still valid since we inserted below it)
                farm_rows_by_block[b_idx] = slot_rows + [
                    insert_after + 1 + o for o in range(extra)]

            # Now shift row numbers for blocks that come AFTER each insertion.
            # Each block's insertion shifts all subsequent blocks' rows down.
            # Compute cumulative offset from all insertions that precede each block.
            cumulative = 0
            for b_idx in range(len(farm_rows_by_block)):
                if cumulative > 0:
                    farm_rows_by_block[b_idx] = [
                        r + cumulative for r in farm_rows_by_block[b_idx]]
                cumulative += total_inserted[b_idx]

            irma_ws_rows_all = [r for rows in farm_rows_by_block for r in rows]

            # The milking_split_rows / milking_combine_rows lists (ROBOT wide-
            # merge handling) were captured BEFORE any rows were inserted, so an
            # entry for a row that sits below an insertion point is now stale.
            # Shift each by the number of rows inserted strictly above it so the
            # end-of-loop re-merge lands on the row the data actually moved to.
            def _shift_row(orig_row):
                s = 0
                for (pos, n) in insertions:
                    if orig_row >= pos:
                        s += n
                return orig_row + s
            if insertions:
                milking_split_rows   = [_shift_row(r) for r in milking_split_rows]
                milking_combine_rows = {_shift_row(r) for r in milking_combine_rows}

            target_cells = set()
            for ws_row in irma_ws_rows_all:
                for col in sheet_write_cols:
                    target_cells.add((ws_row, col))

            # merges_to_redo was captured by Phase A at the ORIGINAL row numbers
            # (before any rows were inserted) and Phase A already unmerged those
            # ranges.  Shift each entry to its post-insertion row so it is
            # restored in the right place.
            #
            # NOTE: this block previously did `merges_to_redo = []` and REBUILT
            # the list from the live merged ranges.  But Phase A had already
            # removed every farm-row merge that overlaps a write column, so the
            # rebuilt list captured none of them - they were never re-merged,
            # and every exported value (IRMA, milking times, name, location,
            # ...) collapsed into a single unmerged cell.  We now keep Phase A's
            # list and only shift it.
            merges_to_redo = [
                (_shift_row(r1), c1, _shift_row(r2), c2)
                for (r1, c1, r2, c2) in merges_to_redo
            ]
            # Whatever merges are still live AND overlap a write cell are the
            # ones we built for the inserted rows (Phase A never saw them).
            # Capture them at their already-final positions and unmerge so that
            # writing - including ROBOT rows that write None into the non-anchor
            # milking cells of a wide merge - never targets a read-only merged
            # cell.  They are restored together with everything else below.
            for merge_range in list(ws.merged_cells.ranges):
                mr = merge_range
                if any((r, c) in target_cells
                       for r in range(mr.min_row, mr.max_row + 1)
                       for c in range(mr.min_col, mr.max_col + 1)):
                    merges_to_redo.append(
                        (mr.min_row, mr.min_col, mr.max_row, mr.max_col))
            for (min_r, min_c, max_r, max_c) in merges_to_redo:
                # Phase A entries are already unmerged (unmerging again raises);
                # live inserted-row merges get unmerged here.
                try:
                    ws.unmerge_cells(start_row=min_r, start_column=min_c,
                                     end_row=max_r,   end_column=max_c)
                except Exception:
                    pass

            written_rows = set()
            if not hasattr(self, '_export_warnings'):
                self._export_warnings = []

            def _write_cell(ws_row, col, value):
                ws.cell(ws_row, col).value = value

            for b_idx, block in enumerate(mod_blocks):
                if b_idx >= len(farm_rows_by_block):
                    break
                slot_rows   = farm_rows_by_block[b_idx]
                block_farms = block.get("rows", [])
                for i, ws_row in enumerate(slot_rows):
                    written_rows.add(ws_row)
                    if i < len(block_farms):
                        fd = block_farms[i]
                        _write_cell(ws_row, C_IRMA,      fd.get("irma", ""))
                        _write_cell(ws_row, C_TRAIN,     fd.get("train", ""))
                        _write_cell(ws_row, C_EDPU,      fd.get("edpu", ""))
                        _write_cell(ws_row, C_LOCATION,  fd.get("location", ""))
                        _write_cell(ws_row, C_PRIOR_VOL, fd.get("prior_vol", None))
                        # ROBOT farms store the literal text "ROBOT" in
                        # m1_start (with the other three fields blank) to
                        # represent automated milking with no fixed window.
                        # parse_hhmm("ROBOT") safely returns None rather than
                        # crashing, but writing None would silently drop the
                        # label - so write it through as plain text instead
                        # of trying to parse it as a time.
                        #
                        # Number format: always force "h:mm;@" on these four
                        # columns at write time, regardless of whatever format
                        # the cell already had.  Cells revealed by splitting a
                        # wide ROBOT merge, or newly inserted rows whose
                        # template happened to be a ROBOT row, can carry
                        # number_format="General" (confirmed on a real file -
                        # the cell was hidden under the merge and never needed
                        # a time format before).  Writing a real time value
                        # into a General-formatted cell makes Excel fall back
                        # to a long default time display ("10:00:00 AM"),
                        # which often doesn't fit the column and renders as
                        # "####".  "h:mm;@" is the format every normal time
                        # cell on the sheet already uses - it displays a real
                        # time compactly as "10:00" and the "@" component lets
                        # the same format hold literal text (e.g. "ROBOT")
                        # without breaking, so it's safe to apply universally
                        # rather than only when we can find a good donor cell.
                        if str(fd.get("m1_start", "")).strip().upper() == "ROBOT":
                            _write_cell(ws_row, C_M1_START, "ROBOT")
                            ws.cell(ws_row, C_M1_START).number_format = "h:mm;@"
                            for col in (C_M1_FINISH, C_M2_START, C_M2_FINISH):
                                _write_cell(ws_row, col, None)
                                ws.cell(ws_row, col).number_format = "h:mm;@"
                        else:
                            for col, fkey in [(C_M1_START,"m1_start"),(C_M1_FINISH,"m1_finish"),
                                              (C_M2_START,"m2_start"),(C_M2_FINISH,"m2_finish")]:
                                raw = fd.get(fkey, "")
                                _write_cell(ws_row, col, parse_hhmm(raw) if raw else None)
                                ws.cell(ws_row, col).number_format = "h:mm;@"
                        extras = fd.get("_extra_cells") or {}
                        for col in extra_cols_in_use:
                            _write_cell(ws_row, col, extras.get(col))
                    else:
                        # More slots than farms in this block - clear leftover
                        for col in sheet_write_cols:
                            _write_cell(ws_row, col, None)

            # Clear any unreachable rows (extra blocks in sheet vs mod_blocks)
            for ws_row in irma_ws_rows_all:
                if ws_row not in written_rows:
                    for col in sheet_write_cols:
                        _write_cell(ws_row, col, None)
            # -- Write dest rows back (delivery information section) --------
            # Each delivery section belongs to a specific block.  We map
            # delivery section row numbers to their block index so each
            # block's modified dests land in the correct section.
            DEST_WRITE_COLS = {C_DEST_VOL, C_DEST_NAME, C_DEST_KEY}

            def _scan_dest_rows():
                """(Re)scan the live worksheet and return {block_idx: [row,...]}
                of the delivery-section rows belonging to each block.

                Delivery sections use NUMBERED slots ("1.", "2.", "3." in
                column 1) - a section can have empty numbered slots (a number
                but no name/key yet) that are still valid write targets.  The
                section ends at a non-numbered row (e.g. the "CIP Wash:" row or
                a blank spacer), NOT at the first empty numbered slot.  This is
                what gives the export room to write a destination the solver
                moved onto a block: the empty numbered slots are already there.

                Re-run after inserting rows so the mapping reflects new layout.
                """
                out = {}
                cur_block = -1
                saw_hdr   = False
                in_dv     = False
                acc       = []

                def _flush(bidx, rows):
                    if rows:
                        out.setdefault(bidx, []).extend(rows)

                def _is_numbered_slot(v):
                    # Column-1 slot markers look like "1.", "2.", "10." etc.
                    if v is None:
                        return False
                    s = str(v).strip().rstrip(".")
                    return s.isdigit()

                for r in range(1, min(ws.max_row, 5000) + 1):
                    v0 = ws.cell(r, 1).value
                    v2 = ws.cell(r, 2).value
                    if isinstance(v0, str) and v0.strip().upper() == "IRMA#":
                        _flush(cur_block, acc); acc = []; in_dv = False
                        cur_block += 1; saw_hdr = True
                    elif isinstance(v2, str) and "delivery" in v2.lower():
                        _flush(cur_block, acc); acc = []; in_dv = True
                        if not saw_hdr:
                            cur_block = 0   # preload block before first IRMA#
                    elif in_dv:
                        dn = ws.cell(r, C_DEST_NAME).value
                        dk = ws.cell(r, C_DEST_KEY).value
                        # A row belongs to the section if it has dest content OR
                        # is a numbered (possibly empty) slot.  Anything else
                        # (wash note, blank spacer) ends the section.
                        if dn or dk or _is_numbered_slot(v0):
                            # Skip wash notes that happen to carry a name/key
                            # (e.g. "WASH AT VTL") - not a real dest slot.
                            nm = str(dn or "").lower()
                            if "wash" in nm:
                                _flush(cur_block, acc); acc = []; in_dv = False
                            else:
                                acc.append(r)
                        else:
                            _flush(cur_block, acc); acc = []; in_dv = False
                _flush(cur_block, acc)
                return out

            dest_rows_by_block = _scan_dest_rows()

            # -- Insert delivery rows for blocks with more dests than slots ----
            # When the solver moves a route's end destination onto a block whose
            # original delivery section has fewer rows than the block now needs,
            # there is no row to write the extra dest into and it silently
            # vanishes from the export.  Mirror the farm-overflow fix: insert
            # the missing delivery rows (cloning the section's last row for
            # styling/format), handling openpyxl's "insert_rows shifts values
            # but not merges" quirk, then re-scan so the write loop sees the
            # complete, correctly-positioned section.  Process sections
            # last-to-first so earlier row numbers stay valid mid-insertion.
            dest_insert_plan = []   # (insert_after_row, extra, template_row)
            for b_idx, block in enumerate(mod_blocks):
                slot_rows = dest_rows_by_block.get(b_idx, [])
                if not slot_rows:
                    continue
                dests_b = block.get("dests") or []
                if not dests_b:
                    dk0 = block.get("dest_key", ""); dn0 = block.get("dest_name", "")
                    if dk0:
                        dests_b = [{"key": dk0, "name": dn0, "vol_partial": None}]
                extra = len(dests_b) - len(slot_rows)
                if extra > 0:
                    dest_insert_plan.append((slot_rows[-1], extra, slot_rows[-1]))

            if dest_insert_plan:
                # Insert from the bottom of the sheet upward so each insertion's
                # anchor row number is unaffected by later (higher-row) ones.
                for insert_after, extra, template_row in sorted(
                        dest_insert_plan, key=lambda t: t[0], reverse=True):
                    # Snapshot + unmerge every merge below the insertion point,
                    # insert, then re-merge each shifted down by `extra`.
                    below = [
                        (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
                        for mr in list(ws.merged_cells.ranges)
                        if mr.min_row > insert_after
                    ]
                    # Also capture the template delivery row's own merges so the
                    # new rows reproduce the section's column layout (name span,
                    # key span, etc.) rather than being bare cells.
                    tmpl_merges = [
                        (mr.min_col, mr.max_col)
                        for mr in list(ws.merged_cells.ranges)
                        if mr.min_row == template_row and mr.max_row == template_row
                    ]
                    for (mnr, mnc, mxr, mxc) in below:
                        ws.unmerge_cells(start_row=mnr, start_column=mnc,
                                         end_row=mxr, end_column=mxc)

                    ws.insert_rows(insert_after + 1, extra)

                    for (mnr, mnc, mxr, mxc) in below:
                        ws.merge_cells(start_row=mnr + extra, start_column=mnc,
                                       end_row=mxr + extra, end_column=mxc)

                    # Style/format the new rows from the template delivery row,
                    # then reproduce its merges on each new row.  Also write a
                    # numbered-slot marker into column 1 ("4.", "5." ...) so the
                    # re-scan below recognises these as real delivery slots -
                    # otherwise a blank-column-1 inserted row terminates the
                    # section scan and the new dests are dropped again.
                    tmpl_slot = ws.cell(template_row, 1).value
                    try:
                        base_n = int(str(tmpl_slot).strip().rstrip("."))
                    except (ValueError, TypeError):
                        base_n = len(dest_rows_by_block.get(
                            next((bi for bi, rows in dest_rows_by_block.items()
                                  if template_row in rows), -1), []))
                    for offset in range(extra):
                        new_row = insert_after + 1 + offset
                        for col in range(1, ws.max_column + 1):
                            tmpl_cell = ws.cell(template_row, col)
                            new_cell  = ws.cell(new_row, col)
                            if tmpl_cell.number_format and \
                                    tmpl_cell.number_format != "General":
                                new_cell.number_format = tmpl_cell.number_format
                            _apply_table_style(new_row, col, src_row=template_row)
                        # Clear any value cloned styling left and set the slot no.
                        ws.cell(new_row, 1).value = f"{base_n + offset + 1}."
                        for (mnc, mxc) in tmpl_merges:
                            try:
                                ws.merge_cells(start_row=new_row, start_column=mnc,
                                               end_row=new_row,   end_column=mxc)
                            except Exception:
                                pass

                # Re-scan so dest_rows_by_block reflects every inserted row at
                # its final position - no manual offset bookkeeping needed.
                dest_rows_by_block = _scan_dest_rows()

            dest_target = set()
            all_dest_rows = [r for rows in dest_rows_by_block.values() for r in rows]
            for dr in all_dest_rows:
                for col in DEST_WRITE_COLS:
                    dest_target.add((dr, col))
            dest_merges = []
            for mr in list(ws.merged_cells.ranges):
                if any((r,c) in dest_target
                       for r in range(mr.min_row, mr.max_row+1)
                       for c in range(mr.min_col, mr.max_col+1)):
                    dest_merges.append((mr.min_row,mr.min_col,mr.max_row,mr.max_col))
            for (min_r,min_c,max_r,max_c) in dest_merges:
                ws.unmerge_cells(start_row=min_r,start_column=min_c,
                                 end_row=max_r,end_column=max_c)

            for b_idx, block in enumerate(mod_blocks):
                slot_rows = dest_rows_by_block.get(b_idx, [])
                if not slot_rows:
                    continue
                dests_b = block.get("dests") or []
                if not dests_b:
                    dk = block.get("dest_key",""); dn = block.get("dest_name","")
                    if dk: dests_b = [{"key":dk,"name":dn,"vol_partial":None}]
                for i, dr in enumerate(slot_rows):
                    if i < len(dests_b):
                        d = dests_b[i]
                        vp = d.get("vol_partial")
                        _write_cell(dr, C_DEST_VOL,  int(vp) if isinstance(vp,(int,float)) else None)
                        _write_cell(dr, C_DEST_NAME, d.get("name",""))
                        _write_cell(dr, C_DEST_KEY,
                                    int(d["key"]) if str(d.get("key","")).isdigit()
                                    else d.get("key",""))
                    else:
                        _write_cell(dr, C_DEST_VOL,  None)
                        _write_cell(dr, C_DEST_NAME, None)
                        _write_cell(dr, C_DEST_KEY,  None)

            for (min_r,min_c,max_r,max_c) in dest_merges:
                ws.merge_cells(start_row=min_r,start_column=min_c,
                               end_row=max_r,end_column=max_c)

            # -- Re-merge ---------------------------------------------------
            for (min_r, min_c, max_r, max_c) in merges_to_redo:
                ws.merge_cells(
                    start_row=min_r, start_column=min_c,
                    end_row=max_r,   end_column=max_c)

            # Rows where a wide ROBOT-style milking merge was split: re-merge
            # using the standard four 3-column fields (M1 start/finish,
            # M2 start/finish) so the real farm's four time values are each
            # independently visible, instead of restoring the wide span that
            # would hide three of them under Excel's "only show top-left of
            # a merge" behaviour.
            #
            # The three newly-revealed top-left cells (M1 finish, M2 start,
            # M2 finish) were hidden under the wide merge and can carry stale
            # leftover formatting from before the merge ever existed -
            # observed on a real file as "MS Sans Serif" 10pt with no border,
            # rather than the sheet's actual Calibri 13pt table styling.  Style
            # all four from the donor row so they match the rest of the table
            # regardless of what was hiding underneath.
            for row in milking_split_rows:
                ws.merge_cells(start_row=row, start_column=C_M1_START,
                               end_row=row,   end_column=C_M1_START + 2)
                ws.merge_cells(start_row=row, start_column=C_M1_FINISH,
                               end_row=row,   end_column=C_M1_FINISH + 2)
                ws.merge_cells(start_row=row, start_column=C_M2_START,
                               end_row=row,   end_column=C_M2_START + 2)
                ws.merge_cells(start_row=row, start_column=C_M2_FINISH,
                               end_row=row,   end_column=C_M2_FINISH + 2)
                # Only the three cells that were genuinely hidden under the
                # wide merge need restyling - M1_START was the merge's visible
                # top-left cell and already had its own correct (and possibly
                # row-specific) styling, which we don't want to overwrite.
                for col in (C_M1_FINISH, C_M2_START, C_M2_FINISH):
                    _apply_table_style(row, col)

            # Rows where a ROBOT farm was slotted into a row that previously
            # had the standard four-field merge layout: combine into one wide
            # merge matching the sheet's ROBOT convention (observed as
            # C_M1_START through C_M2_FINISH+2, e.g. E:P) so the single
            # "ROBOT" label displays across the whole milking-time region
            # instead of leaving four separate (mostly blank) cells.
            for row in milking_combine_rows:
                ws.merge_cells(start_row=row, start_column=C_M1_START,
                               end_row=row,   end_column=C_M2_FINISH + 2)

        self.statusBar().showMessage(f"Exporting {fname}  -  saving file ...")
        QApplication.processEvents()
        wb.save(dst_path)

        # Surface any per-block slot warnings
        if hasattr(self, '_export_warnings') and self._export_warnings:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Export warnings",
                "\n\n".join(self._export_warnings))
            self._export_warnings = []

    def _on_intra_route_apply(self):
        """Launch the IntraRouteOptimiser thread."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        cfg = {
            "plant_windows":         self._get_plant_windows(),
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "avoid_window_penalty":  self._sw_avoid_win_pen.value(),
            "overlap_penalty":       self._sw_overlap_pen.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "min_shift_h":           self._sw_min_shift.value(),
            "shift_under_penalty":   self._sw_shift_under_pen.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }
        self._intra_btn.setEnabled(False)
        self._solve_btn.setEnabled(False)
        locked_sheets = {sname for sname, cb in self._locked_sheet_cbs.items()
                         if cb.isChecked()}
        self._intra_thread = IntraRouteOptimiser(
            fname, self._cache, self.dm, cfg, self._sheet_mods, parent=self,
            dm_dur=self.dm_dur, locked_sheets=locked_sheets)
        self._intra_thread.progress.connect(self._on_intra_progress)
        self._intra_thread.finished.connect(self._on_intra_finished)
        self._intra_thread.log.connect(self._solver_log.append)
        self._intra_thread.start()

    def _on_intra_progress(self, cur, total, status):
        self._intra_btn.setText(f"Optimising... {cur}/{total}")
        self._solver_status.setText(status)
        if total > 0:
            self._solver_progress.setMaximum(total)
            self._solver_progress.setValue(cur)

    def _on_intra_finished(self, results):
        for key, blocks in results.items():
            self._sheet_mods[key] = blocks
        self._intra_btn.setEnabled(True)
        self._intra_btn.setText("Optimise Within Routes")
        self._solve_btn.setEnabled(True)
        self._solver_progress.setValue(self._solver_progress.maximum())
        self._solver_status.setText(f"Done - {len(results)} route(s) improved")
        self._display_sheet()
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    # -- Solver event handlers -------------------------------------------------

    def _on_solve_clicked(self):
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            self._solver_status.setText("No file loaded.")
            return

        groups = _group_sheets_by_colour(self._cache, fname)
        n_red  = len(groups["RED"])
        n_blue = len(groups["BLUE"])
        if n_red == 0 and n_blue == 0:
            self._solver_status.setText("No RED or BLUE sheets found in this file.")
            return

        # Collect user-locked sheets (checked in the Locked Sheets panel)
        locked_sheets = {sname for sname, cb in self._locked_sheet_cbs.items()
                         if cb.isChecked()}

        # Collect plant receiving windows from the demand panel
        plant_windows = self._get_plant_windows()

        cfg = {
            "vol_tol":            self._sw_vol_tol.value(),
            "vol_penalty":        self._sw_vol_pen.value(),
            "hard_vol_cap":       self._sw_hard_cap.value(),
            "cap_penalty":        self._sw_cap_pen.value(),
            "max_shift_h":        self._sw_max_shift.value(),
            "min_shift_h":        self._sw_min_shift.value(),
            "shift_under_penalty":self._sw_shift_under_pen.value(),
            "shift_penalty":      self._sw_shift_pen.value(),
            "shift_hours_weight": self._sw_shift_hours.value(),
            "milking_weight":     self._sw_milking.value(),
            "plant_win_penalty":  self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "avoid_window_penalty":  self._sw_avoid_win_pen.value(),
            "overlap_penalty":       self._sw_overlap_pen.value(),
            "plant_windows":      plant_windows,
            "iterations":         self._sw_iters.value(),
            "target_cool_frac":   self._sw_cool.value(),
            "segment_size":       100,
            "seed":               (self._sw_seed.value() or None),
            "suppress_no_milking": self._suppress_no_milking_cb.isChecked(),
            "split_opt":          hasattr(self, "_chk_split_opt") and self._chk_split_opt.isChecked(),
            "truck_avail_enabled":  self._sw_truck_avail_chk.isChecked(),
            "truck_avail_min_back": self._sw_truck_avail_min_back.value(),
            "truck_avail_penalty":  self._sw_truck_avail_pen.value(),
            "day_night_lock":       self._sw_day_night_lock.isChecked(),
        }

        # Refresh night-start label before solver runs
        self._refresh_truck_avail_label(cfg)

        total_iters = cfg["iterations"] * 2   # RED + BLUE
        self._solver_progress.setMaximum(total_iters)
        self._solver_progress.setValue(0)
        self._solver_log.clear()
        # compute alpha for display
        iters = cfg["iterations"]
        tcf   = cfg["target_cool_frac"]
        alpha_display = tcf ** (1.0 / iters) if iters > 1 and tcf > 0 else 0.9999

        locked_list = sorted(locked_sheets - SOLVER_SKIP_SHEETS)
        win_lines   = "\n".join(
            f"    {dk}: {v[0]}–{v[1]}" for dk, v in sorted(plant_windows.items())
        ) or "    (none active)"

        self._solver_log.append(
            f"Solving {fname}\n"
            f"  RED sheets:     {n_red}\n"
            f"  BLUE sheets:    {n_blue}\n"
            f"  Locked sheets:  {locked_list or '(none)'}\n"
            f"  Iterations:     {cfg['iterations']} per group\n"
            f"  Vol tol:        +/-{cfg['vol_tol']*100:.0f}%\n"
            f"  Truck cap:      {cfg['hard_vol_cap']:,} L  (pen {cfg['cap_penalty']:.0f}/L)\n"
            f"  Max shift:      {cfg['max_shift_h']:.1f} h  (pen {cfg['shift_penalty']:.0f}/h over)\n"
            f"  Min shift:      {cfg['min_shift_h']:.1f} h  (pen {cfg['shift_under_penalty']:.0f}/h under)\n"
            f"  Shift hrs wt:   {cfg['shift_hours_weight']:.1f}/h\n"
            f"  Milking wt:     {cfg['milking_weight']:.1f}x\n"
            f"  Plant win pen:  {cfg['plant_win_penalty']:.0f} km/h wait  "
            f"margin={cfg['plant_win_margin_mins']:.0f}min @ {cfg['plant_win_margin_rate']:.0f} km/h\n"
            f"  Plant windows:\n{win_lines}\n"
            f"  Cooling target: {tcf*100:.2f}% of T0\n"
            f"  alpha (per iter):   {alpha_display:.6f}\n"
        )

        self._solve_btn.setEnabled(False)
        self._stop_btn.setEnabled(True)
        self._solver_status.setText("Running...")

        self._solver_thread = ALNSSolver(fname, self._cache, self.dm, cfg,
                                         sheet_mods=self._sheet_mods,
                                         locked_sheets=locked_sheets,
                                         dm_dur=self.dm_dur)
        self._solver_thread.progress.connect(self._on_solver_progress)
        self._solver_thread.finished.connect(self._on_solver_finished)
        self._solver_thread.log.connect(self._on_solver_log)
        self._solver_thread.start()

    def _on_stop_solver(self):
        if self._solver_thread and self._solver_thread.isRunning():
            self._solver_thread.stop()
            self._solver_status.setText("Stopping...")

    def _on_solver_progress(self, cur, total, msg):
        self._solver_progress.setMaximum(total)
        self._solver_progress.setValue(cur)
        self._solver_status.setText(msg)

    def _on_solver_log(self, msg):
        self._solver_log.append(msg)
        # Auto-scroll
        sb = self._solver_log.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _on_solver_finished(self, results):
        """Write solver results into _sheet_mods and refresh the active sheet."""
        self._solve_btn.setEnabled(True)
        self._stop_btn.setEnabled(False)

        fname = self.file_cb.currentText()

        # Build a global uid_map covering every sheet in the file.
        # The solver moves farms freely between sheets of the same colour group,
        # so a farm that ends up in sheet X may have originally come from sheet Y.
        # Looking only at sheet X's orig_blocks would miss its _uid entirely and
        # leave _orig_arr unset, causing phantom milking waits.
        global_uid_map = {}   # uid -> (was_mwo, None)  -- None kept for compat with _stamp
        file_cache = self._cache.get(fname, {})
        for sname_c, entry_c in file_cache.items():
            if not isinstance(entry_c, dict):
                continue
            for block in entry_c.get("blocks", []):
                for farm in block.get("rows", []):
                    uid = farm.get("_uid")
                    if uid:
                        global_uid_map[uid] = (bool(farm.get("_mwo")), None)

        n_updated = 0
        for sname, new_blocks in results.items():
            self._sheet_mods[(fname, sname)] = new_blocks
            self._stamp_orig_arr_from_map(new_blocks, global_uid_map)
            n_updated += 1

        # Log paired trailers that were held with their lead during the solve.
        if hasattr(self._solver_thread, "paired_followers") and \
                self._solver_thread.paired_followers:
            n = len(self._solver_thread.paired_followers)
            irmas = ', '.join(f['irma'] for _, _, _, f, _
                              in self._solver_thread.paired_followers)
            self._solver_log.append(
                f"  {n} paired trailer(s) held adjacent to their lead: {irmas}")

        self._solver_log.append(
            f"\nOK Done - {n_updated} sheets updated in Modified panel.")
        self._solver_status.setText(f"OK Complete - {n_updated} sheets updated")
        self._solver_progress.setValue(self._solver_progress.maximum())

        # Refresh currently displayed sheet if it was touched
        sname = self.sheet_cb.currentText()
        if sname in results:
            self._mod_blocks = self._sheet_mods[(fname, sname)]
            entry = self._cache.get(fname, {}).get(sname, {})
            if isinstance(entry, dict):
                self._driver_start = entry.get("start_time")
            self._render_editable()

        # Also refresh comparison tab if visible
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

        # Auto-flag 2h waits on solved results if checkbox is on
        if hasattr(self, "_chk_auto_flag") and self._chk_auto_flag.isChecked():
            self._on_auto_flag_waits()
        # Re-optimize split positions after solve
        if hasattr(self, "_chk_split_opt") and self._chk_split_opt.isChecked():
            self._optimize_all_split_positions(fname)

        # Refresh truck availability timeline if it is open
        dlg = getattr(self, "_truck_avail_dlg", None)
        if dlg and dlg.isVisible():
            try:
                day_routes, night_routes, night_start_mins = \
                    self._compute_truck_avail_routes(fname)
                dlg.update_routes(day_routes, night_routes, night_start_mins)
            except Exception:
                pass

    def _init_route_table(self, t, editable=False):
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setAlternatingRowColors(False)
        t.verticalHeader().setVisible(False)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        t.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        t.setShowGrid(True)

    def _init_comp_table(self, t):
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.verticalHeader().setVisible(False)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        t.horizontalHeader().setStretchLastSection(True)
        t.setShowGrid(True)
        t.setAlternatingRowColors(True)

    # -------------------------------------------------------------------------
    # Spinner
    # -------------------------------------------------------------------------

    def _start_spinner(self, msg):
        self._spin_msg = msg; self._spin_idx = 0
        self.load_btn.setEnabled(False); self._spin_timer.start(80)

    def _stop_spinner(self):
        self._spin_timer.stop(); self.load_btn.setEnabled(True)

    def _tick_spinner(self):
        ch = self._spin_chars[self._spin_idx % len(self._spin_chars)]
        self.statusBar().showMessage(f"{ch}  {self._spin_msg}")
        self._spin_idx += 1

    # -------------------------------------------------------------------------
    # Folder scanning / dropdown population
    # -------------------------------------------------------------------------

    def _year_folders(self):
        result = {}
        if self.data_root.exists():
            try:
                entries = sorted(self.data_root.iterdir())
            except PermissionError:
                self.statusBar().showMessage(
                    f"Permission denied reading '{self.data_root}'")
                return result
            for d in entries:
                if d.is_dir(): result[extract_year(d.name)] = d
        return result

    def _month_folders(self, ypath):
        result = {}
        if ypath and ypath.exists():
            try:
                entries = [d for d in ypath.iterdir() if d.is_dir()]
            except PermissionError:
                self.statusBar().showMessage(
                    f"Permission denied reading '{ypath}'")
                return result
            for d in sorted(entries, key=lambda d: month_key(d.name)):
                result[d.name] = d
        return result

    def _xlsx_files(self, mpath):
        result = {}
        if mpath and mpath.exists():
            try:
                files = sorted(mpath.glob("*.xlsx"))
            except PermissionError:
                self.statusBar().showMessage(
                    f"Permission denied reading '{mpath}'")
                return result
            for f in files:
                result[f.name] = f
        return result

    def _block_sigs(self, b):
        for w in (self.year_cb, self.month_cb, self.file_cb, self.sheet_cb):
            w.blockSignals(b)

    def _on_browse_folder(self):
        """Let the user pick a new root data folder and rescan."""
        chosen = QFileDialog.getExistingDirectory(
            self,
            "Select root data folder",
            str(self.data_root) if self.data_root.exists() else str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks,
        )
        if not chosen:
            return   # user cancelled
        self.data_root = Path(chosen)
        try:
            self._scan_folders()
        except OSError as exc:
            self.statusBar().showMessage(
                f"Cannot read folder: {exc}")
            return
        # If no year folders were found, warn the user
        if not self._year_map:
            self.statusBar().showMessage(
                f"(!)  No year sub-folders found in '{chosen}'. "
                "Expected: <root>/<year>/<month>/<file>.xlsx")

    def _scan_folders(self):
        self._block_sigs(True)
        self._year_map = self._year_folders()
        self.year_cb.clear()
        self.year_cb.addItems(sorted(self._year_map.keys(), reverse=True))
        self._block_sigs(False)
        self._pop_months(); self._pop_files()

    def _pop_months(self, restore=None):
        self._block_sigs(True)
        self.month_cb.clear()
        self._month_map = self._month_folders(self._year_map.get(self.year_cb.currentText()))
        self.month_cb.addItems(list(self._month_map.keys()))
        if restore:
            idx = self.month_cb.findText(restore)
            if idx >= 0: self.month_cb.setCurrentIndex(idx)
        self._block_sigs(False)

    def _pop_files(self, restore=None):
        self._block_sigs(True)
        self.file_cb.clear()
        self._file_map = self._xlsx_files(self._month_map.get(self.month_cb.currentText()))
        self.file_cb.addItems(list(self._file_map.keys()))
        if restore:
            idx = self.file_cb.findText(restore)
            if idx >= 0: self.file_cb.setCurrentIndex(idx)
        self._block_sigs(False)
        self._pop_sheets()

    def _pop_sheets(self, restore=None):
        self._block_sigs(True)
        self.sheet_cb.clear()
        fname = self.file_cb.currentText()
        if fname in self._cache:
            for i, (sname, entry) in enumerate(self._cache[fname].items()):
                self.sheet_cb.addItem(sname)
                dc = entry.get("day_colour", "") if isinstance(entry, dict) else ""
                bg, fg, _ = day_colour_style(dc)
                if bg:
                    self.sheet_cb.setItemData(i, bg, Qt.BackgroundRole)
                    self.sheet_cb.setItemData(i, fg, Qt.ForegroundRole)
        if restore:
            idx = self.sheet_cb.findText(restore)
            if idx >= 0: self.sheet_cb.setCurrentIndex(idx)
        self._block_sigs(False)

    # -------------------------------------------------------------------------
    # IRMA Search
    # -------------------------------------------------------------------------

    # Highlight colours used for search results
    _SEARCH_HIT_BG    = QColor("#fff176")   # yellow  - every match
    _SEARCH_CURSOR_BG = QColor("#f57f17")   # amber   - the currently-navigated match

    def _on_search_text_changed(self):
        """Clear highlights as soon as the user edits the query so stale
        results don't linger.  Don't re-search on every keypress - wait for
        Enter so partial IRMA numbers (e.g. typing '71') don't scroll the
        table around while the user is still typing."""
        if not self._search_box.text().strip():
            self._on_search_clear()

    def _on_search(self):
        """Run the search: find every farm row whose IRMA matches the query
        (case-insensitive substring) in both the Original and Modified tables,
        highlight them all, and scroll to the first hit."""
        query = self._search_box.text().strip().upper()
        self._search_clear_highlights()
        self._search_hits   = []
        self._search_cursor = -1

        if not query:
            self._search_status.setText("")
            for btn in (self._search_prev_btn, self._search_next_btn,
                        self._search_clear_btn):
                btn.setEnabled(False)
            return

        irma_col = next((i for i, (_, k) in enumerate(COLS) if k == "irma"), 0)

        for table in (self.orig_table, self.edit_table):
            for row in range(table.rowCount()):
                item = table.item(row, irma_col)
                if item is None:
                    continue
                # Only match real farm rows (they carry UserRole farm data)
                if item.data(Qt.UserRole) is None:
                    continue
                if query in item.text().upper():
                    self._search_hits.append((table, row))
                    # Apply hit highlight to every cell in this row
                    for col in range(table.columnCount()):
                        ci = table.item(row, col)
                        if ci is not None:
                            ci.setBackground(self._SEARCH_HIT_BG)

        n = len(self._search_hits)
        if n == 0:
            self._search_status.setText("No match")
            self._search_status.setStyleSheet("color: #c62828; font-weight: bold;")
            for btn in (self._search_prev_btn, self._search_next_btn):
                btn.setEnabled(False)
            self._search_clear_btn.setEnabled(True)
            return

        self._search_status.setStyleSheet("")
        for btn in (self._search_prev_btn, self._search_next_btn,
                    self._search_clear_btn):
            btn.setEnabled(True)

        # Jump to first hit
        self._search_cursor = 0
        self._search_apply_cursor()

    def _on_search_next(self):
        if not self._search_hits:
            return
        self._search_cursor = (self._search_cursor + 1) % len(self._search_hits)
        self._search_apply_cursor()

    def _on_search_prev(self):
        if not self._search_hits:
            return
        self._search_cursor = (self._search_cursor - 1) % len(self._search_hits)
        self._search_apply_cursor()

    def _search_apply_cursor(self):
        """Update the status label, re-colour all hits (current one gets
        a stronger amber), and scroll the relevant table to show it."""
        irma_col = next((i for i, (_, k) in enumerate(COLS) if k == "irma"), 0)
        n = len(self._search_hits)

        for idx, (table, row) in enumerate(self._search_hits):
            bg = self._SEARCH_CURSOR_BG if idx == self._search_cursor \
                 else self._SEARCH_HIT_BG
            for col in range(table.columnCount()):
                ci = table.item(row, col)
                if ci is not None:
                    ci.setBackground(bg)

        cur_table, cur_row = self._search_hits[self._search_cursor]
        cur_table.scrollToItem(cur_table.item(cur_row, irma_col),
                               QAbstractItemView.PositionAtCenter)
        cur_table.setCurrentCell(cur_row, irma_col)

        # Count hits per table for the label e.g. "3 / 5  (Orig: 2  Mod: 3)"
        orig_hits = sum(1 for t, _ in self._search_hits if t is self.orig_table)
        mod_hits  = sum(1 for t, _ in self._search_hits if t is self.edit_table)
        label     = f"{self._search_cursor + 1} / {n}"
        if orig_hits and mod_hits:
            label += f"  (Orig: {orig_hits}  Mod: {mod_hits})"
        elif orig_hits:
            label += "  (Original only)"
        elif mod_hits:
            label += "  (Modified only)"
        self._search_status.setText(label)

    def _search_clear_highlights(self):
        """Remove search highlight colours from all previously-highlighted rows.
        Called before a new search and on clear so the table returns to its
        normal row colours."""
        irma_col = next((i for i, (_, k) in enumerate(COLS) if k == "irma"), 0)
        for table, row in self._search_hits:
            # Re-read the natural background from the IRMA cell's current data
            # so we don't have to recompute the full row colour.  The IRMA cell
            # background was set by populate_table and doesn't need to be exact;
            # clearing to white is acceptable - the next _display_sheet call
            # will restore proper colours anyway.  For a cleaner restore we
            # delegate back to populate_table by triggering _display_sheet,
            # but that's expensive.  Instead, restore to the standard alternating
            # block colour based on whether the b_idx stored in UserRole is even/odd.
            item = table.item(row, irma_col)
            if item is None:
                continue
            farm_data = item.data(Qt.UserRole)
            if isinstance(farm_data, tuple) and len(farm_data) >= 1:
                b_idx = farm_data[0]
                natural_bg = QColor("#e3f2fd") if b_idx % 2 == 0 else QColor("#ffffff")
            else:
                natural_bg = QColor("#ffffff")
            for col in range(table.columnCount()):
                ci = table.item(row, col)
                if ci is not None:
                    ci.setBackground(natural_bg)

    def _on_search_clear(self):
        """Clear the search box, remove highlights, and reset state."""
        self._search_clear_highlights()
        self._search_hits   = []
        self._search_cursor = -1
        self._search_box.clear()
        self._search_status.setText("")
        self._search_status.setStyleSheet("")
        for btn in (self._search_prev_btn, self._search_next_btn,
                    self._search_clear_btn):
            btn.setEnabled(False)

    def _on_irma_lookup(self):
        """Open the IRMA Farm Lookup dialog against all currently loaded data."""
        dlg = IRMALookupDialog(self._cache, self._sheet_mods, parent=self)
        dlg.navigate_requested.connect(self._navigate_to_sheet)
        # Pre-fill with the search box text if the user already has one there
        if hasattr(self, "_search_box") and self._search_box.text().strip():
            dlg._query.setText(self._search_box.text().strip())
            dlg._run_search()
        dlg.exec_()

    def _navigate_to_sheet(self, fname, sname):
        """Navigate the main window to the given file and sheet.
        Called from IRMALookupDialog when the user double-clicks a result."""
        # Select the file in the file combo
        fidx = self.file_cb.findText(fname)
        if fidx < 0:
            QMessageBox.warning(self, "Not loaded",
                f"'{fname}' is not currently loaded.\n"
                f"Use > Load File to load it first.")
            return
        if self.file_cb.currentIndex() != fidx:
            self._block_sigs(True)
            self.file_cb.setCurrentIndex(fidx)
            self._block_sigs(False)
            self._pop_sheets(restore=sname)
        # Select the sheet
        sidx = self.sheet_cb.findText(sname)
        if sidx >= 0:
            self.sheet_cb.setCurrentIndex(sidx)
        self._display_sheet()
        # Switch to the Route tab so the user can see the sheet
        self.tabs.setCurrentIndex(0)

    # -------------------------------------------------------------------------
    # File loading
    # -------------------------------------------------------------------------

    def _on_load_clicked(self):
        fname = self.file_cb.currentText()
        fpath = self._file_map.get(fname)
        if not fpath or (self._loader and self._loader.isRunning()): return
        self._start_spinner(f"Loading {fname}...")
        self._load_warnings = []   # reset warning accumulator for this load
        self._loader = FileLoader(fname, fpath)
        self._loader.done.connect(self._on_load_done)
        self._loader.failed.connect(self._on_load_failed)
        self._loader.sheet_warning.connect(self._on_sheet_warning)
        self._loader.log.connect(
            lambda msg: self._debug_text.append(msg)
            if hasattr(self, "_debug_text") else None)
        self._loader.start()

    def _on_sheet_warning(self, fname, sheet_name, message):
        """Accumulate per-sheet parse warnings - shown as one summary after load,
        and also appended to the debug-tab text area in real time."""
        self._load_warnings.append((sheet_name, message))
        if hasattr(self, "_debug_text"):
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%H:%M:%S")
            self._debug_text.append(
                f"[{ts}] LOAD WARNING  {fname} / {sheet_name}\n  {message}\n")

    def _on_load_done(self, fname, sheets):
        self._stop_spinner()
        self._cache[fname] = sheets
        # Clear any stale modified state for this file so Modified starts
        # identical to Original on every fresh load.
        stale_keys = [k for k in self._sheet_mods if k[0] == fname]
        for k in stale_keys:
            del self._sheet_mods[k]
        corr_keys = [k for k in self._corrected_blocks if k[0] == fname]
        for k in corr_keys:
            del self._corrected_blocks[k]
        self._pop_sheets()
        self._display_sheet()
        # Must refresh demand targets first so plant windows are populated
        self._refresh_demand_targets()
        self._refresh_locked_sheets_list()
        self._populate_irma_dropdown()
        self._refresh_truck_avail_label()
        self._populate_proc_dropdown()
        self._refresh_farm_summary()
        # Optimize partial-dropoff split positions FIRST
        if hasattr(self, "_chk_split_opt") and self._chk_split_opt.isChecked():
            self._optimize_all_split_positions(fname)
        # Auto-flag 2h waits only if checkbox is checked (default: off)
        if hasattr(self, "_chk_auto_flag") and self._chk_auto_flag.isChecked():
            self._on_auto_flag_waits()
        # Snapshot corrected baseline so solver results don't overwrite it
        if hasattr(self, "_chk_route_opt") and self._chk_route_opt.isChecked():
            self._snapshot_corrected_blocks(fname)
            self._display_sheet()   # refresh Original panel from corrected blocks
        # Keep comparison tab fresh if it happens to be open
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

        # Show any parse warnings that accumulated during the load.
        # One dialog with all issues is less annoying than one per sheet.
        warnings = getattr(self, "_load_warnings", [])
        if warnings:
            lines = []
            for sheet_name, msg in warnings:
                lines.append(f"Sheet '{sheet_name}':\n  {msg}")
            body = "\n\n".join(lines)

            dlg = QDialog(self)
            dlg.setWindowTitle(f"Load warnings - {fname}")
            dlg.setWindowFlags(dlg.windowFlags() & ~Qt.WindowContextHelpButtonHint)
            dlg.setMinimumWidth(520)
            dlg.resize(560, 400)
            lay = QVBoxLayout(dlg)
            lay.setContentsMargins(12, 12, 12, 12)
            lay.setSpacing(8)

            intro = QLabel(
                f"The following issues were found while loading {fname}.\n"
                "Affected sheets may appear empty or have missing data."
            )
            intro.setWordWrap(True)
            lay.addWidget(intro)

            from PyQt5.QtWidgets import QTextEdit
            txt = QTextEdit()
            txt.setReadOnly(True)
            txt.setPlainText(body)
            txt.setStyleSheet("font-family: monospace; font-size: 8pt;")
            lay.addWidget(txt, stretch=1)

            btns = QDialogButtonBox(QDialogButtonBox.Ok)
            btns.accepted.connect(dlg.accept)
            lay.addWidget(btns)

            dlg.exec_()
            self._load_warnings = []

    def _optimize_all_split_positions(self, fname=None):
        """Run _optimize_split_positions on every sheet in the file,
        using current plant window config from the UI.  Updates _sheet_mods
        in place and refreshes the displayed sheet if it changed."""
        if fname is None:
            fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return
        plant_windows = self._get_plant_windows()
        if not plant_windows:
            return   # no windows configured - nothing to optimise
        cfg = {
            "plant_windows":         plant_windows,
            "plant_win_penalty":     self._sw_plant_win_pen.value(),
            "plant_win_margin_mins": self._sw_plant_margin_mins.value(),
            "plant_win_margin_rate": self._sw_plant_margin_rate.value(),
            "avoid_window_penalty":  self._sw_avoid_win_pen.value(),
            "overlap_penalty":       self._sw_overlap_pen.value(),
            "suppress_no_milking":   self._suppress_no_milking_cb.isChecked(),
            "milking_weight":        self._sw_milking.value(),
            "shift_hours_weight":    self._sw_shift_hours.value(),
            "shift_penalty":         self._sw_shift_pen.value(),
            "max_shift_h":           self._sw_max_shift.value(),
            "min_shift_h":           self._sw_min_shift.value(),
            "shift_under_penalty":   self._sw_shift_under_pen.value(),
            "cap_penalty":           self._sw_cap_pen.value(),
            "hard_vol_cap":          self._sw_hard_cap.value(),
            "vol_tol":               self._sw_vol_tol.value(),
            "vol_penalty":           self._sw_vol_pen.value(),
        }
        any_changed = False
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            start_time = entry.get("start_time")
            if not start_time:
                continue
            key = (fname, sname)
            if key not in self._sheet_mods:
                self._sheet_mods[key] = copy.deepcopy(entry["blocks"])
            blocks = self._sheet_mods[key]
            if _optimize_split_positions(blocks, self.dm, start_time, cfg, self.dm_dur):
                any_changed = True
        if any_changed:
            self._display_sheet()
            if self.tabs.currentIndex() == 1:
                self._refresh_comparison()

    def _on_load_failed(self, fname, err):
        self._stop_spinner()
        self.statusBar().showMessage(f"Error loading {fname}: {err}")
        if hasattr(self, "_debug_text"):
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%H:%M:%S")
            self._debug_text.append(
                f"[{ts}] LOAD ERROR  {fname}\n  {err}\n")

    def _update_day_colour_badge(self):
        bg, fg, text = day_colour_style(self._day_colour)
        if bg:
            self.day_colour_box.setText(text)
            self.day_colour_box.setStyleSheet(
                f"background-color: {bg.name()}; color: {fg.name()}; "
                f"border-radius: 4px; padding: 2px 6px; font-weight: bold;")
        else:
            self.day_colour_box.setText("")
            self.day_colour_box.setStyleSheet(
                "border-radius: 4px; padding: 2px 6px;")
        self._update_shift_type_badge()
        # Sheet date — read from cache for the current sheet
        _fname = self.file_cb.currentText() if hasattr(self, "file_cb") else ""
        _sname = self.sheet_cb.currentText() if hasattr(self, "sheet_cb") else ""
        _entry = self._cache.get(_fname, {}).get(_sname, {})
        _sd    = _entry.get("sheet_date") if isinstance(_entry, dict) else None
        self.sheet_date_lbl.setText(
            f"{_sd.strftime('%A %B')} {_sd.day}, {_sd.year}" if _sd else "")

    def _update_shift_type_badge(self):
        """Show DAY or NIGHT label based on the current sheet's start time."""
        st = getattr(self, "_driver_start", None)
        if st is None:
            self.shift_type_box.setText("")
            self.shift_type_box.setStyleSheet("border-radius: 3px; padding: 1px 6px;")
        elif _is_day_sheet(st):
            self.shift_type_box.setText("DAY")
            self.shift_type_box.setStyleSheet(
                "background-color: #f57c00; color: white; font-weight: bold; "
                "border-radius: 3px; padding: 1px 6px; font-size: 8pt;")
        else:
            self.shift_type_box.setText("NIGHT")
            self.shift_type_box.setStyleSheet(
                "background-color: #3949ab; color: white; font-weight: bold; "
                "border-radius: 3px; padding: 1px 6px; font-size: 8pt;")

    def _on_view_on_map(self):
        """Open the MapDialog for the current sheet."""
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText() if hasattr(self, "sheet_cb") else ""
        if not fname or fname not in self._cache or not sname:
            QMessageBox.information(self, "Map", "Load a sheet first.")
            return
        entry  = self._cache[fname].get(sname)
        if not isinstance(entry, dict):
            QMessageBox.information(self, "Map", "No route data for this sheet.")
            return
        blocks = self._sheet_mods.get((fname, sname), entry.get("blocks", []))
        if not blocks:
            QMessageBox.information(self, "Map", "No blocks in this sheet.")
            return

        # Load irma -> (lat, lon) from routes.db locations table
        db_path = MapDialog._find_db()
        irma_locs = {}
        if db_path.exists():
            import sqlite3 as _sq
            con = _sq.connect(str(db_path))
            for row in con.execute("SELECT irma, lat, lon FROM locations"):
                irma_locs[row[0]] = (row[1], row[2])
            con.close()
        else:
            QMessageBox.warning(
                self, "Map",
                "routes.db not found beside the exe.\n"
                "The map will open without road geometry or background tiles.")

        dlg = MapDialog(blocks, sname, irma_locs,
                        all_sheets=self._cache.get(fname, {}),
                        parent=self)
        dlg.exec_()

    def _compute_truck_avail_routes(self, fname):
        """Compute day and night route timing data for the truck avail timeline.
        Returns (day_routes, night_routes, night_start_mins).
        """
        suppress = (self._suppress_no_milking_cb.isChecked()
                    if hasattr(self, "_suppress_no_milking_cb") else True)
        dm     = self.dm
        dm_dur = getattr(self, "dm_dur", None)
        cache  = self._cache

        night_start_mins = None
        for entry in cache.get(fname, {}).values():
            if not isinstance(entry, dict):
                continue
            st = entry.get("start_time")
            if st is None or _is_day_sheet(st):
                continue
            t = st.time() if isinstance(st, datetime) else st
            m = t.hour * 60 + t.minute
            if night_start_mins is None or m < night_start_mins:
                night_start_mins = m

        day_routes = []
        for sname, entry in cache.get(fname, {}).items():
            if not isinstance(entry, dict):
                continue
            st     = entry.get("start_time")
            colour = entry.get("day_colour", "")
            if st is None or not _is_day_sheet(st):
                continue
            if colour not in ("RED", "BLUE"):
                continue
            blocks = self._sheet_mods.get((fname, sname), entry.get("blocks", []))
            try:
                ct = calc_times(blocks, dm, st, dm_dur, suppress_no_milking=suppress)
            except Exception:
                continue
            if ct is None:
                continue
            all_times, end_cursor = ct
            if end_cursor is None:
                continue
            t_st  = st.time()         if isinstance(st,         datetime) else st
            t_end = end_cursor.time() if isinstance(end_cursor, datetime) else end_cursor
            sm = t_st.hour  * 60 + t_st.minute
            em = t_end.hour * 60 + t_end.minute
            if em < sm:
                em += 24 * 60
            on_time = night_start_mins is None or em <= night_start_mins
            segs = _route_stop_segments(blocks, all_times, sm)
            day_routes.append({
                "sname":      sname,
                "colour":     colour,
                "start_mins": sm,
                "end_mins":   em,
                "on_time":    on_time,
                "segments":   segs,
            })

        night_routes = []
        for sname, entry in cache.get(fname, {}).items():
            if not isinstance(entry, dict):
                continue
            st     = entry.get("start_time")
            colour = entry.get("day_colour", "")
            if st is None or _is_day_sheet(st):
                continue
            if colour not in ("RED", "BLUE"):
                continue
            t_st = st.time() if isinstance(st, datetime) else st
            sm   = t_st.hour * 60 + t_st.minute
            blocks = self._sheet_mods.get((fname, sname), entry.get("blocks", []))
            em   = None
            segs = []
            try:
                ct = calc_times(blocks, dm, st, dm_dur, suppress_no_milking=suppress)
                if ct is not None:
                    all_times, end_cursor = ct
                    if end_cursor is not None:
                        t_end = end_cursor.time() if isinstance(end_cursor, datetime) else end_cursor
                        em = t_end.hour * 60 + t_end.minute
                        if em < sm:
                            em += 24 * 60
                        segs = _route_stop_segments(blocks, all_times, sm)
            except Exception:
                pass
            night_routes.append({
                "sname":      sname,
                "colour":     colour,
                "start_mins": sm,
                "end_mins":   em,
                "segments":   segs,
            })

        day_routes.sort(  key=lambda r: r["start_mins"])
        night_routes.sort(key=lambda r: r["start_mins"])
        return day_routes, night_routes, night_start_mins

    def _on_truck_avail_visualize(self):
        """Open (or refresh) the TruckAvailDialog showing full route timelines."""
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            QMessageBox.information(self, "No file", "Load a file first.")
            return
        try:
            day_routes, night_routes, night_start_mins = \
                self._compute_truck_avail_routes(fname)
            dlg = getattr(self, "_truck_avail_dlg", None)
            if dlg and dlg.isVisible():
                dlg.update_routes(day_routes, night_routes, night_start_mins)
                dlg.raise_()
            else:
                dlg = TruckAvailDialog(day_routes, night_routes, night_start_mins, self,
                                       fname=fname, date_str=_sheets_date_str(self._cache, fname))
                self._truck_avail_dlg = dlg
                dlg.show()
        except Exception as _exc:
            import traceback as _tb
            QMessageBox.critical(
                self, "Truck Availability",
                f"Error opening timeline:\n{_exc}\n\n{_tb.format_exc()}")

    def _refresh_truck_avail_label(self, cfg=None):
        """Update the night-start display in the Truck Availability solver panel."""
        if not hasattr(self, "_sw_truck_avail_night_lbl"):
            return
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            self._sw_truck_avail_night_lbl.setText("Night start: (load a file)")
            return
        earliest = None
        day_count = night_count = 0
        for entry in self._cache[fname].values():
            if not isinstance(entry, dict):
                continue
            st = entry.get("start_time")
            if st is None:
                continue
            if _is_day_sheet(st):
                day_count += 1
            else:
                night_count += 1
                t = st.time() if isinstance(st, datetime) else st
                m = t.hour * 60 + t.minute
                if earliest is None or m < earliest:
                    earliest = m
        parts = []
        if earliest is not None:
            parts.append(f"Night starts {earliest // 60:02d}:{earliest % 60:02d}")
        parts.append(f"{day_count} day / {night_count} night routes")
        self._sw_truck_avail_night_lbl.setText("  ".join(parts))

    def _display_sheet(self):
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if not fname or not sname or fname not in self._cache: return
        entry = self._cache[fname].get(sname)
        if not entry: return
        # Changing sheet invalidates any previous search results - clear them
        # silently (no UI flash) before the tables are repopulated.
        if hasattr(self, "_search_hits") and self._search_hits:
            self._search_hits   = []
            self._search_cursor = -1
            self._search_status.setText("")
            self._search_status.setStyleSheet("")
            for btn in (self._search_prev_btn, self._search_next_btn,
                        self._search_clear_btn):
                btn.setEnabled(False)
        blocks     = entry["blocks"]
        self._driver_start = entry["start_time"]
        self._day_colour   = entry.get("day_colour", "")
        self._update_day_colour_badge()
        self.farm_tray.refresh_bold_state(self._day_colour)
        suppress = self._suppress_no_milking_cb.isChecked() \
            if hasattr(self, "_suppress_no_milking_cb") else True
        # Restore saved mod_blocks for this sheet, or start fresh
        key = (fname, sname)
        if key not in self._sheet_mods:
            self._sheet_mods[key] = copy.deepcopy(blocks)
        self._mod_blocks = self._sheet_mods[key]

        # Original panel: when route corrections are active, render from the
        # corrected baseline (_corrected_blocks) - never from solver output.
        # When corrections are off, render from raw cache blocks.
        corrections_on = (hasattr(self, "_chk_route_opt") and
                          self._chk_route_opt.isChecked())
        if corrections_on and key in self._corrected_blocks:
            orig_display = copy.deepcopy(self._corrected_blocks[key])
        elif corrections_on and key in self._sheet_mods:
            # Corrections on but not snapshotted yet - use _sheet_mods
            orig_display = copy.deepcopy(self._sheet_mods[key])
        else:
            orig_display = copy.deepcopy(blocks)

        populate_table(self.orig_table, orig_display, self.dm, editable=False,
                       start_time=self._driver_start, dm_dur=self.dm_dur,
                       suppress_no_milking=suppress,
                       plant_windows=self._get_plant_windows() if hasattr(self, "_get_plant_windows") else {})
        self._render_editable()
        # Tray is NOT cleared - removed farms persist across sheet switches
        total = sum(len(b["rows"]) for b in blocks)
        st = fmt_hhmm(self._driver_start) if self._driver_start else "?"
        self.statusBar().showMessage(
            f"{fname}  /  {sname}  -  {len(blocks)} route(s), {total} farm(s)  |  Start: {st}"
            + ("" if self.dm else "  (!) distance_matrix.csv not found"))

    def _stamp_orig_arr(self, orig_blocks, mod_blocks, start_time):
        """After the solver runs: restore _mwo flags and _orig_arr
        baselines for any farm row that had them set before the solve.

        Each farm row is assigned a unique _uid at parse time (see parse_sheet).
        The solver preserves _uid through deepcopy, so we can look up any row
        in the original schedule regardless of which sheet or position it ended
        up on after optimisation - no IRMA matching, no occurrence counting,
        no milking-signature heuristics needed.
        """
        if not orig_blocks:
            return

        # Build map: _uid -> was_mwo_checked
        uid_map = {}
        for block in orig_blocks:
            for farm in block.get("rows", []):
                uid = farm.get("_uid")
                if uid:
                    uid_map[uid] = bool(farm.get("_mwo"))

        # Restore _mwo on mod_blocks (solver deepcopies clear it)
        for block in mod_blocks:
            for farm in block.get("rows", []):
                uid = farm.get("_uid")
                if uid and uid_map.get(uid):
                    farm["_mwo"] = True

    def _stamp_orig_arr_from_map(self, mod_blocks, uid_map):
        """Restore _mwo flags on mod_blocks using a pre-built uid map.

        uid_map: {uid: (was_checked, _)} - second element ignored (no _orig_arr).
        """
        for block in mod_blocks:
            for farm in block.get("rows", []):
                uid = farm.get("_uid")
                if not uid or uid not in uid_map:
                    continue
                was_checked, _ = uid_map[uid]
                if was_checked:
                    farm["_mwo"] = True


    def _save_mod_blocks(self):
        """Persist current mod_blocks to _sheet_mods for the active sheet."""
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if fname and sname and self._mod_blocks is not None:
            self._sheet_mods[(fname, sname)] = self._mod_blocks

    def _render_editable(self):
        self._save_mod_blocks()
        self.edit_table.blockSignals(True)
        _sup = (self._suppress_no_milking_cb.isChecked()
                if hasattr(self, "_suppress_no_milking_cb") else True)
        populate_table(self.edit_table, self._mod_blocks, self.dm, editable=True,
                       start_time=getattr(self, "_driver_start", None),
                       dm_dur=self.dm_dur,
                       suppress_no_milking=_sup,
                       plant_windows=self._get_plant_windows() if hasattr(self, "_get_plant_windows") else {})
        self.edit_table.blockSignals(False)
        # Show status warning if any block has no processor
        no_proc = [b for b in (self._mod_blocks or [])
                   if not (b.get("dests") or b.get("dest_key",""))]
        if no_proc:
            routes = ", ".join(b.get("route","?") for b in no_proc)
            self._add_status.setText(
                f"(!)  Route(s) {routes} have no processor - drag one from the tray")
            self._add_status_timer.start(8000)

    # -------------------------------------------------------------------------
    # Farm removal / reinsertion
    # -------------------------------------------------------------------------

    def _on_manual_add(self):
        """Validate and add a manually entered farm row to the tray."""
        irma_raw = self._add_irma.currentText().strip()
        # Validate IRMA format
        if not IRMA_RE.match(irma_raw):
            self._add_irma.lineEdit().setStyleSheet("color: #c0392b; font-weight: bold;")
            self._add_status.setText(f"IRMA '{irma_raw}' is not a valid IRMA number (format: ##-###)")
            self._add_status_timer.start(4000)
            return
        # Validate IRMA is in distance matrix
        from_dm = bool(self._dm_keys) and (irma_raw in self._dm_keys)
        if not from_dm:
            self._add_irma.lineEdit().setStyleSheet("color: #c0392b; font-weight: bold;")
            self._add_status.setText(f"IRMA '{irma_raw}' not found in distance matrix")
            self._add_status_timer.start(4000)
            return
        # Valid
        self._add_irma.lineEdit().setStyleSheet("")
        self._add_status.setText("")
        vol_raw = self._add_vol.text().strip()
        try:
            vol = float(vol_raw.replace(",", "")) if vol_raw else None
        except ValueError:
            vol = None
        farm = {
            "irma":      irma_raw,
            "train":     self._add_train.text().strip(),
            "m1_start":  self._add_m1s.text().strip(),
            "m1_finish": self._add_m1f.text().strip(),
            "m2_start":  self._add_m2s.text().strip(),
            "m2_finish": self._add_m2f.text().strip(),
            "edpu":      self._add_edpu.text().strip(),
            "location":  self._add_loc.text().strip(),
            "prior_vol": vol,
        }
        farm["_from_block"]  = -1
        farm["_from_fname"]  = self.file_cb.currentText()
        farm["_from_sname"]  = self.sheet_cb.currentText()
        farm["_from_colour"] = self._day_colour
        self._removed.append(farm)
        self.farm_tray.add_farm(farm, "(manual)", self._day_colour)
        # Clear fields for fast multi-entry
        for w in (self._add_train, self._add_m1s, self._add_m1f,
                  self._add_m2s, self._add_m2f, self._add_edpu,
                  self._add_loc, self._add_vol):
            w.clear()
        self._add_irma.setCurrentIndex(-1)
        self._add_irma.lineEdit().clear()
        self._add_irma.setFocus()

    def _on_del_btn_state(self):
        """Enable Delete Selected if anything deletable is selected in either
        the Modified table or the tray."""
        edit_sel = self._edit_table_deletable_item()
        tray_sel = self.farm_tray.currentRow() >= 0
        self._tray_del_btn.setEnabled(edit_sel is not None or tray_sel)

    def _edit_table_deletable_item(self):
        """Return the currently selected QTableWidgetItem in the Modified table
        if it is a farm, dest, or block banner row - else None."""
        item = self.edit_table.currentItem()
        if item is None:
            return None
        # Block banner
        if item.data(Qt.UserRole + 2) is not None:
            return item
        # Dest row
        dd = item.data(Qt.UserRole + 1)
        if dd is not None and dd[0] == "dest":
            return item
        # Farm row
        if item.data(Qt.UserRole) is not None:
            return item
        return None

    def _on_tray_delete(self):
        """Delete the selected item from the Modified table OR the tray,
        in that priority order (Modified table takes precedence)."""
        # -- Modified table delete -----------------------------------------
        edit_item = self._edit_table_deletable_item()
        if edit_item is not None and self._mod_blocks is not None:
            # Block banner -> delete entire block
            b_idx_banner = edit_item.data(Qt.UserRole + 2)
            if b_idx_banner is not None:
                if 0 <= b_idx_banner < len(self._mod_blocks):
                    self._mod_blocks.pop(b_idx_banner)
                    self._save_mod_blocks()
                    self._render_editable()
                return

            # Dest row -> remove dest from its block
            dd = edit_item.data(Qt.UserRole + 1)
            if dd is not None and dd[0] == "dest":
                _, b_idx, d_idx = dd
                if 0 <= b_idx < len(self._mod_blocks):
                    dests = self._mod_blocks[b_idx].get("dests") or []
                    if 0 <= d_idx < len(dests):
                        dests.pop(d_idx)
                        if dests:
                            self._mod_blocks[b_idx]["dest_name"] = dests[0]["name"]
                            self._mod_blocks[b_idx]["dest_key"]  = dests[0]["key"]
                        else:
                            self._mod_blocks[b_idx]["dest_name"] = ""
                            self._mod_blocks[b_idx]["dest_key"]  = ""
                    self._save_mod_blocks()
                    self._render_editable()
                return

            # Farm row -> remove farm from its block
            fd = edit_item.data(Qt.UserRole)
            if fd is not None:
                b_idx, f_idx = fd
                if 0 <= b_idx < len(self._mod_blocks):
                    rows = self._mod_blocks[b_idx]["rows"]
                    if 0 <= f_idx < len(rows):
                        rows.pop(f_idx)
                    self._save_mod_blocks()
                    self._render_editable()
                return

        # -- Tray delete ---------------------------------------------------
        row = self.farm_tray.currentRow()
        if row < 0:
            return
        if row < len(self._removed):
            self._removed.pop(row)
        self.farm_tray.removeRow(row)
        self._tray_del_btn.setEnabled(self.farm_tray.currentRow() >= 0)

    def _on_mwo_changed(self, item):
        """Handle the MWO (Milking Window Override) checkbox being toggled.

        When checked, the farm's milking windows are ignored entirely -
        the truck arrives and pumps without waiting.  No _orig_arr is needed.
        """
        if item.column() != MWO_COL:
            return
        ud = item.data(Qt.UserRole)
        if ud is None:
            return
        b_idx, f_idx = ud
        if self._mod_blocks is None or b_idx >= len(self._mod_blocks):
            return
        rows = self._mod_blocks[b_idx]["rows"]
        if f_idx >= len(rows):
            return
        farm    = rows[f_idx]
        checked = (item.checkState() == Qt.Checked)
        farm["_mwo"] = checked

        # Sync to the original cache so both panels show the same state
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        orig_entry = self._cache.get(fname, {}).get(sname)
        if orig_entry:
            uid = farm.get("_uid")
            for orig_block in orig_entry.get("blocks", []):
                for orig_farm in orig_block.get("rows", []):
                    if uid and orig_farm.get("_uid") == uid:
                        orig_farm["_mwo"] = checked
                        break
        # Re-render both panels and refresh comparison
        _sup = self._suppress_no_milking_cb.isChecked()
        populate_table(self.orig_table, orig_entry["blocks"] if orig_entry else [],
                       self.dm, editable=False,
                       start_time=self._driver_start, dm_dur=self.dm_dur,
                       suppress_no_milking=_sup)
        self._render_editable()
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    def _populate_irma_dropdown(self):
        """Build _irma_lookup from all loaded cache entries and populate the
        IRMA combo with sorted IRMA numbers.  Called whenever a file is loaded.

        _irma_lookup: {irma_str: {train, m1_start, m1_finish, m2_start,
                                   m2_finish, edpu, location, name}}
        name is taken from _extra_cells[18] (column R - farm name), matching
        how the route view table displays it.
        Only the most-recently-seen data for each IRMA is kept (all files
        in the session are merged so previous loads stay available).
        """
        if not hasattr(self, "_irma_lookup"):
            self._irma_lookup = {}

        for fname, sheets in self._cache.items():
            for sname, entry in sheets.items():
                if not isinstance(entry, dict):
                    continue
                for block in entry.get("blocks", []):
                    for row in block.get("rows", []):
                        irma = row.get("irma", "").strip()
                        if not irma:
                            continue
                        farm_name = (row.get("_extra_cells") or {}).get(18, "")
                        existing  = self._irma_lookup.get(irma, {})
                        self._irma_lookup[irma] = {
                            "train":     row.get("train", ""),
                            "m1_start":  row.get("m1_start", ""),
                            "m1_finish": row.get("m1_finish", ""),
                            "m2_start":  row.get("m2_start", ""),
                            "m2_finish": row.get("m2_finish", ""),
                            "edpu":      row.get("edpu", ""),
                            "location":  row.get("location", ""),
                            "name":      farm_name or existing.get("name", ""),
                        }

        # Repopulate combo (block signals to avoid spurious autofill)
        self._add_irma.blockSignals(True)
        current_text = self._add_irma.currentText()
        self._add_irma.clear()
        for irma in sorted(self._irma_lookup.keys()):
            self._add_irma.addItem(irma)
        # Restore whatever the user had typed
        self._add_irma.lineEdit().setText(current_text)
        self._add_irma.blockSignals(False)

    def _on_irma_autofill(self, index):
        """When the user picks an IRMA from the dropdown, fill all known fields."""
        irma = self._add_irma.itemText(index).strip()
        data = getattr(self, "_irma_lookup", {}).get(irma)
        if not data:
            return
        self._add_train.setText(str(data.get("train", "") or ""))
        self._add_m1s.setText(str(data.get("m1_start", "") or ""))
        self._add_m1f.setText(str(data.get("m1_finish", "") or ""))
        self._add_m2s.setText(str(data.get("m2_start", "") or ""))
        self._add_m2f.setText(str(data.get("m2_finish", "") or ""))
        self._add_edpu.setText(str(data.get("edpu", "") or ""))
        self._add_loc.setText(str(data.get("name", "") or data.get("location", "") or ""))
        # Leave vol blank - user always enters that manually
        self._add_vol.setFocus()

    def _populate_proc_dropdown(self):
        """Build _proc_lookup from all loaded cache entries and populate the
        processor key combo with known processor key->name pairs."""
        if not hasattr(self, "_proc_lookup"):
            self._proc_lookup = {}   # {key: name}

        for fname, sheets in self._cache.items():
            for sname, entry in sheets.items():
                if not isinstance(entry, dict):
                    continue
                for block in entry.get("blocks", []):
                    for d in block.get("dests", []):
                        dk = str(d.get("key") or "").strip()
                        dn = str(d.get("name") or "").strip()
                        if dk:
                            self._proc_lookup[dk] = dn or dk
                    dk2 = str(block.get("dest_key") or "").strip()
                    dn2 = str(block.get("dest_name") or "").strip()
                    if dk2 and dk2 not in self._proc_lookup:
                        self._proc_lookup[dk2] = dn2 or dk2

        self._add_proc_key.blockSignals(True)
        current_text = self._add_proc_key.lineEdit().text()
        self._add_proc_key.clear()
        for key in sorted(self._proc_lookup.keys()):
            name = self._proc_lookup[key]
            self._add_proc_key.addItem(f"{key}  -  {name}", userData=key)
        self._add_proc_key.lineEdit().setText(current_text)
        self._add_proc_key.blockSignals(False)

    def _on_proc_key_autofill(self, index):
        """When the user picks a processor from the dropdown, fill the name field."""
        key = self._add_proc_key.itemData(index)
        if key is None:
            # Fallback: parse key from display text "key  -  name"
            text = self._add_proc_key.itemText(index)
            key = text.split("-")[0].strip() if "-" in text else text.strip()
        name = getattr(self, "_proc_lookup", {}).get(key, "")
        self._add_proc_key.lineEdit().setText(key)
        if name:
            self._add_proc_name.setText(name)
        self._add_proc_vol.setFocus()

    def _on_proc_key_editing_finished(self):
        """Clean up after the completer fills in 'key  -  name' on Enter keypress.

        QComboBox.activated only fires when the user selects from the visible
        dropdown.  When the user types a partial key and presses Enter, the
        inline completer sets the line edit text to the full display string
        (e.g. '901012  -  Olympic Dairy') without firing activated.  This
        handler detects that case, strips the name back out of the key field,
        and fills the name field correctly.
        """
        raw = self._add_proc_key.lineEdit().text().strip()
        if not raw:
            return
        # If the completer pasted "key  -  name" into the key field, split it
        if "  -  " in raw:
            key, _, name_from_text = raw.partition("  -  ")
            key  = key.strip()
            name = name_from_text.strip()
        else:
            key  = raw
            name = getattr(self, "_proc_lookup", {}).get(key, "")
        # Always reduce the key field to just the bare key
        self._add_proc_key.lineEdit().setText(key)
        # Fill the name field only if it is currently empty
        if name and not self._add_proc_name.text().strip():
            self._add_proc_name.setText(name)

    def _on_add_block(self):
        """
        Append a new empty block to the current sheet's modified route.

        A small dialog collects:
          - Route name / number  (free text, e.g. "1082")
          - Processor key        (numeric, e.g. 972712)
          - Processor name       (free text, e.g. "Saputo Abbotsford - 2F")
          - Optional partial volume (leave blank for "rest of load")
        """
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if not fname or not sname or self._mod_blocks is None:
            QMessageBox.warning(self, "Add Block", "No sheet loaded.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Add New Block")
        dlg.setFixedWidth(380)
        dl = QVBoxLayout(dlg)
        dl.setContentsMargins(14, 12, 14, 12)
        dl.setSpacing(8)

        bold_f = QFont(); bold_f.setBold(True)
        small_f = QFont()

        def _row(label, widget):
            rw = QWidget(); rl = QHBoxLayout(rw)
            rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(8)
            lb = QLabel(label); lb.setFont(small_f); lb.setFixedWidth(120)
            rl.addWidget(lb); rl.addWidget(widget, stretch=1)
            return rw

        route_edit = QLineEdit()
        route_edit.setPlaceholderText("e.g. 1082")
        route_edit.setFont(small_f)

        proc_key_edit = QLineEdit()
        proc_key_edit.setPlaceholderText("e.g. 972712")
        proc_key_edit.setFont(small_f)

        proc_name_edit = QLineEdit()
        proc_name_edit.setPlaceholderText("e.g. Saputo Abbotsford - 2F")
        proc_name_edit.setFont(small_f)

        vol_edit = QLineEdit()
        vol_edit.setPlaceholderText("blank = rest of load")
        vol_edit.setFont(small_f)

        status_lbl = QLabel("")
        status_lbl.setStyleSheet("color: #c0392b; font-size: 8pt;")

        dl.addWidget(QLabel("Fill in route details:"))
        dl.addWidget(_row("Route name:", route_edit))
        dl.addWidget(_row("Processor key:", proc_key_edit))
        dl.addWidget(_row("Processor name:", proc_name_edit))
        dl.addWidget(_row("Partial vol (L):", vol_edit))
        dl.addWidget(status_lbl)

        btn_row = QWidget(); br = QHBoxLayout(btn_row)
        br.setContentsMargins(0, 0, 0, 0); br.setSpacing(8)
        ok_btn = QPushButton("Add Block")
        ok_btn.setFont(bold_f)
        ok_btn.setStyleSheet(
            "QPushButton { background:#7b1fa2; color:white; font-weight:bold; "
            "border-radius:4px; padding:4px 12px; }")
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setFont(small_f)
        br.addStretch(); br.addWidget(cancel_btn); br.addWidget(ok_btn)
        dl.addWidget(btn_row)

        cancel_btn.clicked.connect(dlg.reject)

        def _accept():
            pk = proc_key_edit.text().strip()
            pn = proc_name_edit.text().strip()
            if not pk and not pn:
                status_lbl.setText("Processor key or name is required.")
                return
            vol_raw = vol_edit.text().strip()
            try:
                vol_partial = float(vol_raw.replace(",", "")) if vol_raw else None
            except ValueError:
                status_lbl.setText("Invalid volume - enter a number or leave blank.")
                return
            dest = {"name": pn, "key": pk, "vol_partial": vol_partial}
            new_block = {
                "route":     route_edit.text().strip(),
                "dests":     [dest],
                "dest_name": pn,
                "dest_key":  pk,
                "rows":      [],
            }
            self._mod_blocks.append(new_block)
            self._save_mod_blocks()
            self._render_editable()
            self.statusBar().showMessage(
                f"Added empty block '{new_block['route'] or '(unnamed)'}' "
                f"-> {pn or pk} to sheet {sname}")
            dlg.accept()

        ok_btn.clicked.connect(_accept)
        dlg.exec_()

    def _on_auto_flag_waits(self):
        """Set MWO on every farm whose milking wait exceeds 30 minutes,
        iterating until stable so cascading waits are captured correctly.
        Also ensures farms in NO_MILKING_WINDOW_FARMS always have MWO set.
        """
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache:
            return

        total_flagged = 0
        sheets_updated = set()

        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict):
                continue
            start = entry.get("start_time")
            if not start:
                continue

            key = (fname, sname)
            if key in self._sheet_mods:
                mod_blocks = self._sheet_mods[key]
            else:
                mod_blocks = copy.deepcopy(entry["blocks"])

            orig_entry = entry
            flagged = 0

            # Iterate until stable - cascading gate waits may push subsequent
            # farms' arrivals so that they now also exceed the wait threshold.
            max_passes = 20
            for _pass in range(max_passes):
                pass_changed = False

                # Recompute timing with current flags applied
                ct = calc_times(mod_blocks, self.dm, start, self.dm_dur,
                                suppress_no_milking=True)
                if ct is None:
                    break
                all_times = ct[0]

                for b_idx, block in enumerate(mod_blocks):
                    btimes = all_times[b_idx] if b_idx < len(all_times) else None
                    if not btimes:
                        continue
                    for f_idx, farm in enumerate(block.get("rows", [])):
                        f_stop = _farm_stop_index(block, f_idx, b_idx, mod_blocks)
                        if f_stop >= len(btimes):
                            continue
                        ft   = btimes[f_stop]
                        wait = ft.get("wait")
                        arr  = ft.get("arr")
                        if not arr:
                            continue

                        # Flag if wait > 30 min and not yet flagged
                        if wait and wait > 30 and not farm.get("_mwo"):
                            farm["_mwo"] = True
                            flagged += 1
                            pass_changed = True

                if not pass_changed:
                    break  # stable - no more cascading changes

            # Sync flags to orig panel via uid
            if flagged:
                for b_idx, block in enumerate(mod_blocks):
                    for f_idx, farm in enumerate(block.get("rows", [])):
                        if not farm.get("_mwo"):
                            continue
                        uid = farm.get("_uid")
                        for ob_i, orig_block in enumerate(orig_entry.get("blocks", [])):
                            for orig_fi, orig_farm in enumerate(orig_block.get("rows", [])):
                                if uid and orig_farm.get("_uid") == uid:
                                    orig_farm["_mwo"] = True
                                    break

                self._sheet_mods[key] = mod_blocks
                total_flagged += flagged
                sheets_updated.add(sname)

        # -- Always anchor the 3 suppressed ROBOT farms to their original arrival --
        # These farms have no valid milking window data so suppress_no_milking=True
        # means calc_times returns wait=None for them -> the normal auto-flag loop
        # never triggers -> the solver has no anchor for them -> they end up anywhere.
        # Always flag NO_MILKING_WINDOW_FARMS with MWO so the solver never
        # tries to schedule around their (unreliable) window data.
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict): continue
            key = (fname, sname)
            if key not in self._sheet_mods:
                self._sheet_mods[key] = copy.deepcopy(entry["blocks"])
            mod_blocks = self._sheet_mods[key]
            changed = False
            for block in mod_blocks:
                for farm in block.get("rows", []):
                    irma = farm.get("irma", "")
                    if irma not in NO_MILKING_WINDOW_FARMS:
                        continue
                    if farm.get("_mwo"):
                        continue  # already flagged
                    farm["_mwo"] = True
                    changed = True
                    total_flagged += 1
            if changed:
                self._sheet_mods[key] = mod_blocks
                sheets_updated.add(sname)

        # Refresh the currently displayed sheet
        cur_sname = self.sheet_cb.currentText()
        cur_entry = self._cache.get(fname, {}).get(cur_sname)
        cur_key   = (fname, cur_sname)
        if cur_key in self._sheet_mods:
            self._mod_blocks = self._sheet_mods[cur_key]
        if cur_sname in sheets_updated and cur_entry:
            populate_table(self.orig_table, cur_entry["blocks"],
                           self.dm, editable=False,
                           start_time=self._driver_start, dm_dur=self.dm_dur,
                           suppress_no_milking=self._suppress_no_milking_cb.isChecked())
        self._render_editable()
        if total_flagged and self.tabs.currentIndex() == 1:
            self._refresh_comparison()

        if total_flagged:
            self.statusBar().showMessage(
                f"MWO set on {total_flagged} farm(s) across "
                f"{len(sheets_updated)} sheet(s)", 5000)
        else:
            self.statusBar().showMessage(
                f"No farms with wait > 30min found "
                f"({'distance matrix not loaded' if not self.dm else 'across all sheets'})",
                5000)

    def _on_reset_current_sheet(self):
        """Reset only the currently displayed sheet back to original; leave others intact."""
        fname = self.file_cb.currentText()
        sname = self.sheet_cb.currentText()
        if not fname or not sname: return
        key = (fname, sname)
        if key in self._sheet_mods:
            del self._sheet_mods[key]
        # Remove any tray items that came from this sheet - the reset restores
        # those farms from cache, so leaving them in the tray would cause
        # duplicates if the user dragged them back.
        rows_to_remove = [
            i for i, item in enumerate(self._removed)
            if item.get("_from_fname") == fname and item.get("_from_sname") == sname
        ]
        for i in reversed(rows_to_remove):
            self._removed.pop(i)
            self.farm_tray.removeRow(i)
        if rows_to_remove:
            self._tray_del_btn.setEnabled(self.farm_tray.currentRow() >= 0)
        # Reload display (will re-create a fresh copy from cache)
        self._display_sheet()
        # Keep comparison tab fresh if open
        if self.tabs.currentIndex() == 1:
            self._refresh_comparison()

    def _on_reset_all(self):
        """Reset all modified blocks back to originals and clear tray."""
        if self._cache and hasattr(self, 'file_cb'):
            fname = self.file_cb.currentText()
            if fname in self._cache:
                # Clear all sheet mods
                keys_to_clear = [k for k in self._sheet_mods if k[0] == fname]
                for k in keys_to_clear:
                    del self._sheet_mods[k]
                # Clear tray
                self.farm_tray.setRowCount(0)
                self._removed.clear()
                self._tray_del_btn.setEnabled(False)
                # Reload current sheet
                self._display_sheet()
                # Keep comparison tab fresh if open
                if self.tabs.currentIndex() == 1:
                    self._refresh_comparison()

    def _on_manual_add_proc(self):
        """Validate and add a manually entered processor destination to the tray."""
        raw_text = self._add_proc_key.lineEdit().text().strip()
        # If user selected from dropdown, text may be "key  -  name"
        key_raw  = raw_text.split("-")[0].strip() if "-" in raw_text else raw_text
        name_raw = self._add_proc_name.text().strip()
        # Auto-fill name from lookup if not manually entered
        if not name_raw and key_raw and hasattr(self, "_proc_lookup"):
            name_raw = self._proc_lookup.get(key_raw, "")
        vol_raw  = self._add_proc_vol.text().strip()

        if not key_raw and not name_raw:
            self._add_proc_key.lineEdit().setStyleSheet("color:#c0392b; font-weight:bold;")
            self._add_status.setText("Processor key or name is required")
            self._add_status_timer.start(4000)
            return
        self._add_proc_key.lineEdit().setStyleSheet("")

        vol_partial = None
        if vol_raw:
            try:
                vol_partial = float(vol_raw.replace(",",""))
            except ValueError:
                self._add_proc_vol.setStyleSheet("color:#c0392b; font-weight:bold;")
                self._add_status.setText(f"Partial volume must be a number (got '{vol_raw}')")
                self._add_status_timer.start(4000)
                return
        self._add_proc_vol.setStyleSheet("")
        self._add_status.setText("")

        dest = {
            "key":         key_raw or name_raw,
            "name":        name_raw or key_raw,
            "vol_partial": vol_partial,
            "_from_block":  -1,
            "_from_fname":  self.file_cb.currentText(),
            "_from_sname":  self.sheet_cb.currentText(),
            "_from_colour": self._day_colour,
            "_is_dest":     True,
        }
        self._removed.append(dest)
        sheet_label = f"{self.sheet_cb.currentText()} (manual)"
        self.farm_tray.add_dest(dest, sheet_label, self._day_colour)
        for w in (self._add_proc_name, self._add_proc_vol):
            w.clear()
        self._add_proc_key.lineEdit().clear()
        self._add_proc_key.lineEdit().setFocus()

    def _on_farm_removed(self, b_idx, f_idx):
        """Remove farm from mod_blocks and add to tray."""
        if self._mod_blocks is None: return
        block = self._mod_blocks[b_idx]
        if f_idx >= len(block["rows"]): return
        farm = block["rows"].pop(f_idx)
        farm["_from_block"]  = b_idx
        farm["_from_fname"]  = self.file_cb.currentText()
        farm["_from_sname"]  = self.sheet_cb.currentText()
        farm["_from_colour"] = self._day_colour
        self._removed.append(farm)
        sheet_label = f"{self.sheet_cb.currentText()} / {block['route']}"
        self.farm_tray.add_farm(farm, sheet_label, self._day_colour)
        self._save_mod_blocks()
        self._render_editable()

    def _on_farm_reorder(self, src_b, src_f, dst_b, dst_f):
        """Move a farm within mod_blocks (internal drag reorder)."""
        if self._mod_blocks is None: return
        src_block = self._mod_blocks[src_b]
        if src_f >= len(src_block["rows"]): return
        farm = src_block["rows"].pop(src_f)
        # Strip only ephemeral tray-metadata keys, not persistent farm state
        # (_uid, _mwo, _orig_arr must survive a reorder).
        _TRAY_META = {"_from_block", "_from_fname", "_from_sname", "_from_colour"}
        restored = {k: v for k, v in farm.items() if k not in _TRAY_META}
        # Adjust dst_f if moving within same block and dst is after src
        dst_b = max(0, min(dst_b, len(self._mod_blocks)-1))
        dst_rows = self._mod_blocks[dst_b]["rows"]
        if dst_b == src_b and dst_f > src_f:
            dst_f -= 1   # list shrank by one
        dst_f = max(0, min(dst_f, len(dst_rows)))
        dst_rows.insert(dst_f, restored)
        self._render_editable()

    def _on_farm_inserted(self, tray_idx, b_idx, insert_before):
        """Return a farm from the tray into mod_blocks."""
        if self._mod_blocks is None: return
        if tray_idx < 0 or tray_idx >= len(self._removed): return
        farm = self._removed.pop(tray_idx)
        self.farm_tray.removeRow(tray_idx)
        b_idx = max(0, min(b_idx, len(self._mod_blocks)-1))
        # Strip only ephemeral tray-metadata keys, not persistent farm state
        # (_uid, _mwo, _orig_arr must survive tray round-trip).
        _TRAY_META = {"_from_block", "_from_fname", "_from_sname", "_from_colour"}
        restored = {k: v for k, v in farm.items() if k not in _TRAY_META}
        rows = self._mod_blocks[b_idx]["rows"]
        rows.insert(max(0, min(insert_before, len(rows))), restored)
        self._render_editable()

    def _on_dest_removed(self, b_idx, d_idx):
        """Remove a dest from mod_blocks and add to tray."""
        if self._mod_blocks is None: return
        block = self._mod_blocks[b_idx]
        dests = block.get("dests") or []
        if d_idx >= len(dests): return
        dest = dests.pop(d_idx)
        # Update legacy fields
        if dests:
            block["dest_name"] = dests[0]["name"]
            block["dest_key"]  = dests[0]["key"]
        else:
            block["dest_name"] = ""
            block["dest_key"]  = ""
        dest["_from_block"]  = b_idx
        dest["_from_fname"]  = self.file_cb.currentText()
        dest["_from_sname"]  = self.sheet_cb.currentText()
        dest["_from_colour"] = self._day_colour
        dest["_is_dest"]     = True
        self._removed.append(dest)
        sheet_label = f"{self.sheet_cb.currentText()} / {block['route']}"
        self.farm_tray.add_dest(dest, sheet_label, self._day_colour)
        self._save_mod_blocks()
        self._render_editable()

    def _on_dest_reorder(self, src_b, src_d, dst_b, dst_d):
        """Move a dest within (or across) blocks."""
        if self._mod_blocks is None: return
        src_block = self._mod_blocks[src_b]
        src_dests = src_block.get("dests") or []
        if src_d >= len(src_dests): return
        dest = src_dests.pop(src_d)
        # Update src block legacy fields
        if src_dests:
            src_block["dest_name"] = src_dests[0]["name"]
            src_block["dest_key"]  = src_dests[0]["key"]
        dst_b   = max(0, min(dst_b, len(self._mod_blocks) - 1))
        dst_block = self._mod_blocks[dst_b]
        dst_dests = dst_block.setdefault("dests", [])
        if dst_b == src_b and dst_d > src_d:
            dst_d -= 1
        dst_d = max(0, min(dst_d, len(dst_dests)))
        dst_dests.insert(dst_d, dest)
        dst_block["dest_name"] = dst_dests[0]["name"]
        dst_block["dest_key"]  = dst_dests[0]["key"]
        self._render_editable()

    def _on_block_reorder(self, src_b, dst_b):
        """Move an entire block (and all its farms/dests) to a new position.
        dst_b == -1 means append at end."""
        if self._mod_blocks is None: return
        n = len(self._mod_blocks)
        if src_b < 0 or src_b >= n: return
        block = self._mod_blocks.pop(src_b)
        if dst_b < 0 or dst_b >= len(self._mod_blocks):
            self._mod_blocks.append(block)
        else:
            insert_at = dst_b if dst_b <= src_b else dst_b - 1
            insert_at = max(0, min(insert_at, len(self._mod_blocks)))
            self._mod_blocks.insert(insert_at, block)
        self._save_mod_blocks()
        self._render_editable()

    def _on_dest_inserted(self, tray_idx, b_idx, insert_before):
        if self._mod_blocks is None: return
        if tray_idx < 0 or tray_idx >= len(self._removed): return
        item = self._removed[tray_idx]
        if not item.get("_is_dest"): return
        self._removed.pop(tray_idx)
        self.farm_tray.removeRow(tray_idx)
        dest = {k: v for k, v in item.items() if not k.startswith("_")}
        b_idx = max(0, min(b_idx, len(self._mod_blocks) - 1))
        dests = self._mod_blocks[b_idx].setdefault("dests", [])
        dests.insert(max(0, min(insert_before, len(dests))), dest)
        self._mod_blocks[b_idx]["dest_name"] = dests[0]["name"]
        self._mod_blocks[b_idx]["dest_key"]  = dests[0]["key"]
        self._render_editable()

    # -------------------------------------------------------------------------
    # Tab handling
    # -------------------------------------------------------------------------

    def _on_tab_changed(self, idx):
        title = self.tabs.tabText(idx)
        if title == "Comparison": self._refresh_comparison()
        if title == "Solver": self._refresh_demand_targets()
        if title == "Debug": self._refresh_debug_tab()

    # -------------------------------------------------------------------------
    # Comparison tab
    # -------------------------------------------------------------------------

    def _get_mod_blocks(self, fname, sname):
        """Get the modified blocks for a sheet, falling back to original."""
        key = (fname, sname)
        if key in self._sheet_mods:
            return self._sheet_mods[key]
        entry = self._cache.get(fname, {}).get(sname, {})
        return entry.get("blocks", []) if isinstance(entry, dict) else []

    @staticmethod
    def _normalise_proc(name):
        """Strip 'Yard for ' prefix so yard entries merge with plant entries."""
        if not name: return "Unknown"
        n = name.strip()
        if n.lower().startswith("yard for "):
            n = n[9:].strip()
        return n or "Unknown"

    def _agg_file(self, fname, use_mod=False):
        """Aggregate all sheets in fname.
        Returns:
          proc_vols: {proc_name: {"RED":v,"BLUE":v,"OTHER":v,"TOTAL":v}}
          sheet_rows: [(sname, day_colour, total_km, km_ok, shift_hours)]
        """
        if fname not in self._cache: return {}, []
        proc_vols = {}   # proc_name -> {colour_bucket: vol}
        sheet_rows = []
        for sname, entry in self._cache[fname].items():
            if not isinstance(entry, dict): continue
            orig_blocks = entry.get("blocks", [])
            start_time  = entry.get("start_time")
            day_colour  = entry.get("day_colour", "")
            if use_mod:
                key = (fname, sname)
                if key in self._sheet_mods:
                    blocks = self._sheet_mods[key]
                else:
                    blocks = orig_blocks
            else:
                # When route corrections are active, use the corrected baseline
                # (_corrected_blocks) - never solver output from _sheet_mods.
                corrections_on = (hasattr(self, "_chk_route_opt") and
                                  self._chk_route_opt.isChecked())
                key = (fname, sname)
                if corrections_on and key in self._corrected_blocks:
                    blocks = self._corrected_blocks[key]
                elif corrections_on and key in self._sheet_mods:
                    blocks = self._sheet_mods[key]
                else:
                    blocks = orig_blocks
            # Only skip if there is no cache entry at all (shouldn't happen here,
            # but guard anyway). Do NOT skip empty mod_blocks - a solver run can
            # legitimately empty a sheet by moving all its farms elsewhere, and
            # skipping it would make those litres disappear from the totals.
            if blocks is None: continue

            # Colour bucket for this sheet
            dc = day_colour.upper().strip()
            if "RED"  in dc: bucket = "RED"
            elif "BLUE" in dc: bucket = "BLUE"
            else: bucket = "OTHER"

            # Processor volumes - split across multiple dests if present
            for block in blocks:
                dests = block.get("dests") or []
                if not dests:
                    dk = block.get("dest_key","") or block.get("dest_name","") or "Unknown"
                    dests = [{"name": block.get("dest_name",""), "key": dk, "vol_partial": None}]
                total_farm_vol = sum((r["prior_vol"] or 0) for r in block["rows"]
                                     if isinstance(r.get("prior_vol"), (int, float)))
                already = 0.0
                for dest_d in dests:
                    raw = dest_d.get("name") or dest_d.get("key") or "Unknown"
                    dest = self._normalise_proc(raw)
                    vp = dest_d.get("vol_partial")
                    remaining = max(0.0, total_farm_vol - already)
                    offload = min(float(vp), remaining) if vp is not None else remaining
                    already += offload
                    if dest not in proc_vols:
                        proc_vols[dest] = {"RED": 0, "BLUE": 0, "OTHER": 0}
                    proc_vols[dest][bucket] += offload

            # Distance
            total_km = 0.0; km_ok = True
            for dists in calc_distances(blocks, self.dm):
                for d in dists[:-1]:
                    if d is None: km_ok = False
                    else: total_km += d

            # Shift hours
            shift_hours = None
            if start_time:
                _ct3 = calc_times(blocks, self.dm, start_time, dm_dur=self.dm_dur)
                if _ct3 is not None:
                    from datetime import datetime, date
                    base = datetime.combine(date.today(), start_time)
                    shift_hours = (_ct3[1] - base).total_seconds() / 3600.0

            sheet_rows.append((sname, day_colour, total_km, km_ok, shift_hours))
        return proc_vols, sheet_rows

    def _refresh_comparison(self):
        fname = self.file_cb.currentText()
        if not fname or fname not in self._cache: return

        orig_pv, orig_sr = self._agg_file(fname, use_mod=False)
        mod_pv,  mod_sr  = self._agg_file(fname, use_mod=True)

        # Compute tray volume - farms currently sitting in the tray are excluded
        # from modified routes and will make the modified total appear lower.
        tray_vol = sum(
            f.get("prior_vol") or 0
            for f in self._removed
            if not f.get("_is_dest") and isinstance(f.get("prior_vol"), (int, float))
        )

        # Warn in the status bar if tray farms are causing a volume discrepancy
        orig_total = sum(sum(b.values()) for b in orig_pv.values())
        mod_total  = sum(sum(b.values()) for b in mod_pv.values())
        diff = orig_total - mod_total
        if tray_vol > 0:
            self.statusBar().showMessage(
                f"(!)  {tray_vol:,.0f} L in tray (not on any route) - "
                f"Modified total is {diff:,.0f} L less than Original."
            )
        elif diff > 100:
            # Unexplained discrepancy - shouldn't normally happen
            self.statusBar().showMessage(
                f"(!)  Modified total is {diff:,.0f} L less than Original "
                f"with no farms in tray - check for blocks without a processor."
            )
        else:
            self.statusBar().clearMessage()

        # Processor volumes
        all_procs = sorted(set(orig_pv.keys()) | set(mod_pv.keys()))
        changed_procs = {p for p in all_procs
                         if orig_pv.get(p,{}) != mod_pv.get(p,{})}
        self._fill_proc_comp(self._comp_tables["proc_orig"], orig_pv, all_procs, changed_procs, bold_changed=False)
        self._fill_proc_comp(self._comp_tables["proc_mod"],  mod_pv,  all_procs, changed_procs, bold_changed=True)

        # Sheet summaries - build full sheet info dict keyed by sname
        def sr_to_map(sr):
            return {s: (dc, km, ok, h) for s, dc, km, ok, h in sr}
        orig_sm = sr_to_map(orig_sr)
        mod_sm  = sr_to_map(mod_sr)
        all_sheets = list(dict.fromkeys([s for s,*_ in orig_sr] + [s for s,*_ in mod_sr]))
        changed_sheets = {s for s in all_sheets
                          if (orig_sm.get(s,("",0,True,None))[1:3]
                              != mod_sm.get(s,("",0,True,None))[1:3])}
        self._fill_sheet_comp(self._comp_tables["sheet_orig"], orig_sm, all_sheets, changed_sheets, bold_changed=False)
        self._fill_sheet_comp(self._comp_tables["sheet_mod"],  mod_sm,  all_sheets, changed_sheets, bold_changed=True)

    def _comp_cell(self, text, bg, font, align=Qt.AlignRight|Qt.AlignVCenter, fg=None):
        item = QTableWidgetItem(str(text))
        item.setBackground(bg); item.setFont(font)
        item.setTextAlignment(align)
        item.setFlags(Qt.ItemIsEnabled|Qt.ItemIsSelectable)
        if fg: item.setForeground(fg)
        return item

    def _fill_proc_comp(self, table, proc_vols, all_procs, changed_procs, bold_changed):
        # Columns: Processor | Red | Blue | Other | Total
        table.clearSpans(); table.clear()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Processor", "Red (L)", "Blue (L)", "Other (L)", "Total (L)"])
        table.setRowCount(len(all_procs) + 1)
        bold_font = QFont(); bold_font.setBold(True)
        norm_font = QFont()
        tot_r = tot_b = tot_o = tot_t = 0
        for i, proc in enumerate(all_procs):
            buckets  = proc_vols.get(proc, {"RED":0,"BLUE":0,"OTHER":0})
            r_vol = buckets.get("RED",   0)
            b_vol = buckets.get("BLUE",  0)
            o_vol = buckets.get("OTHER", 0)
            total = r_vol + b_vol + o_vol
            changed = proc in changed_procs
            bg   = CLR_CHANGED if changed else (CLR_ALT if i%2==0 else CLR_WHITE)
            font = bold_font if (changed and bold_changed) else norm_font
            table.setItem(i, 0, self._comp_cell(proc,            bg, font, Qt.AlignLeft|Qt.AlignVCenter))
            table.setItem(i, 1, self._comp_cell(f"{int(r_vol):,}" if r_vol else "-", bg, font))
            table.setItem(i, 2, self._comp_cell(f"{int(b_vol):,}" if b_vol else "-", bg, font))
            table.setItem(i, 3, self._comp_cell(f"{int(o_vol):,}" if o_vol else "-", bg, font))
            table.setItem(i, 4, self._comp_cell(f"{int(total):,}", bg, font))
            tot_r += r_vol; tot_b += b_vol; tot_o += o_vol; tot_t += total
        r = len(all_procs)
        table.setItem(r, 0, make_header_item("TOTAL", bg=CLR_TOTAL, fg=QColor("#000000")))
        for col, val in [(1,tot_r),(2,tot_b),(3,tot_o),(4,tot_t)]:
            tv = make_header_item(f"{int(val):,}", bg=CLR_TOTAL, fg=QColor("#000000"))
            tv.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
            table.setItem(r, col, tv)

    def _fill_sheet_comp(self, table, sheet_map, all_sheets, changed_sheets, bold_changed):
        # Columns: Sheet | ● Red | ● Blue | ● Other | km | h
        table.clearSpans(); table.clear()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Sheet", "Red", "Blue", "Other", "km", "Hours"])
        bold_font = QFont(); bold_font.setBold(True)
        norm_font = QFont()

        # Group by colour bucket for subtotals: RED rows, BLUE rows, OTHER rows
        # Order: RED sheets, then BLUE, then OTHER, then totals
        def bucket_of(dc):
            dc = dc.upper()
            if "RED"  in dc: return "RED"
            if "BLUE" in dc: return "BLUE"
            return "OTHER"

        grouped = {"RED": [], "BLUE": [], "OTHER": []}
        for s in all_sheets:
            dc = sheet_map.get(s, ("","",0,True,None))[0] if isinstance(sheet_map.get(s), tuple) else ""
            grouped[bucket_of(dc)].append(s)

        rows_ordered = []
        subtotals = {}  # bucket -> (km_tot, h_tot, ok_all)
        for bucket in ("RED", "BLUE", "OTHER"):
            sheets_in = grouped[bucket]
            k_tot = 0.0; h_tot = 0.0; ok_all = True
            for s in sheets_in:
                rows_ordered.append((s, bucket))
                if s in sheet_map:
                    _, km, ok, h = sheet_map[s]
                    k_tot += km; ok_all = ok_all and ok
                    if h: h_tot += h
            subtotals[bucket] = (k_tot, h_tot, ok_all, len(sheets_in))

        # Total row count: data rows + 3 subtotal rows + 1 grand total
        n_data = len(all_sheets)
        n_sub  = sum(1 for b in ("RED","BLUE","OTHER") if subtotals[b][3] > 0)
        table.setRowCount(n_data + n_sub + 1)

        CLR_RED_SUB  = QColor("#ffcdd2")
        CLR_BLUE_SUB = QColor("#bbdefb")
        CLR_OTH_SUB  = QColor("#e0e0e0")
        sub_colours  = {"RED": CLR_RED_SUB, "BLUE": CLR_BLUE_SUB, "OTHER": CLR_OTH_SUB}

        row_i = 0
        grand_km = 0.0; grand_h = 0.0; grand_ok = True
        for bucket in ("RED", "BLUE", "OTHER"):
            sheets_in = grouped[bucket]
            if not sheets_in: continue
            for s in sheets_in:
                changed = s in changed_sheets
                bg   = CLR_CHANGED if changed else (CLR_ALT if row_i%2==0 else CLR_WHITE)
                font = bold_font if (changed and bold_changed) else norm_font
                table.setItem(row_i, 0, self._comp_cell(s, bg, font, Qt.AlignLeft|Qt.AlignVCenter))
                # Colour dot columns
                for col, b in enumerate(("RED","BLUE","OTHER"), start=1):
                    dot = "*" if b == bucket else ""
                    dot_bg = sub_colours[b] if b == bucket else bg
                    table.setItem(row_i, col, self._comp_cell(dot, dot_bg, font, Qt.AlignCenter))
                if s in sheet_map:
                    _, km, ok, h = sheet_map[s]
                    km_str = f"{km:.1f}" if ok else f"~{km:.1f}*"
                    h_str  = f"{h:.2f}" if h is not None else "-"
                    grand_km += km; grand_ok = grand_ok and ok
                    if h: grand_h += h
                else:
                    km_str = "-"; h_str = "-"
                table.setItem(row_i, 4, self._comp_cell(km_str, bg, font))
                table.setItem(row_i, 5, self._comp_cell(h_str,  bg, font))
                row_i += 1
            # Subtotal row for this colour bucket
            k_tot, h_tot, ok_sub, _ = subtotals[bucket]
            sub_bg = sub_colours[bucket]
            label = f"{bucket.title()} Subtotal"
            table.setItem(row_i, 0, make_header_item(label, bg=sub_bg, fg=QColor("#000000")))
            for col in (1,2,3):
                table.setItem(row_i, col, make_header_item("", bg=sub_bg, fg=QColor("#000000")))
            km_s = f"{k_tot:.1f}" if ok_sub else f"~{k_tot:.1f}*"
            table.setItem(row_i, 4, make_header_item(km_s, bg=sub_bg, fg=QColor("#000000")))
            table.setItem(row_i, 5, make_header_item(f"{h_tot:.2f}", bg=sub_bg, fg=QColor("#000000")))
            row_i += 1

        # Grand total
        km_grand = f"{grand_km:.1f}" if grand_ok else f"~{grand_km:.1f}*"
        table.setItem(row_i, 0, make_header_item("TOTAL", bg=CLR_TOTAL, fg=QColor("#000000")))
        for col in (1,2,3):
            table.setItem(row_i, col, make_header_item("", bg=CLR_TOTAL, fg=QColor("#000000")))
        table.setItem(row_i, 4, make_header_item(km_grand,          bg=CLR_TOTAL, fg=QColor("#000000")))
        table.setItem(row_i, 5, make_header_item(f"{grand_h:.2f}",  bg=CLR_TOTAL, fg=QColor("#000000")))

    # -------------------------------------------------------------------------
    # Dropdown signals
    # -------------------------------------------------------------------------

    def _on_year(self):  self._pop_months(); self._pop_files()
    def _on_month(self): self._pop_files()
    def _on_file(self):  self._pop_sheets()
    def _on_sheet(self): self._display_sheet()


def _run_selftests():
    """Lightweight checks for the pure domain helpers. Run with:
        python viewer158.py --selftest
    These pin down the behaviour of the core time/volume functions so a future
    refactor can't silently change them. No GUI or data files required.
    """
    fails = []

    def check(name, cond):
        if cond:
            print(f"  ok  {name}")
        else:
            print(f"  FAIL {name}")
            fails.append(name)

    def approx(a, b, tol=1e-6):
        return abs(a - b) <= tol

    # parse_hhmm
    check("parse_hhmm valid",   parse_hhmm("06:30") == dt_time(6, 30))
    check("parse_hhmm empty",   parse_hhmm("") is None)
    check("parse_hhmm dash",    parse_hhmm("-") is None)
    check("parse_hhmm garbage", parse_hhmm("not-a-time") is None)

    # time_in_window (normal + overnight + boundaries)
    check("window inside",      time_in_window(dt_time(7, 0),  "06:00", "08:00") is True)
    check("window outside",     time_in_window(dt_time(9, 0),  "06:00", "08:00") is False)
    check("window open bound",  time_in_window(dt_time(6, 0),  "06:00", "08:00") is True)
    check("window overnight a", time_in_window(dt_time(23, 0), "22:00", "02:00") is True)
    check("window overnight b", time_in_window(dt_time(1, 0),  "22:00", "02:00") is True)
    check("window overnight c", time_in_window(dt_time(12, 0), "22:00", "02:00") is False)
    check("window none",        time_in_window(None,           "06:00", "08:00") is False)

    # Extended milking window helper (2 h pre, 1 h post)
    # Raw window 06:00–08:00 → extended 04:00–09:00
    def ext(s_str, f_str): return _extended_milking_window(s_str, f_str)
    es, ef = ext("06:00", "08:00")
    check("ext window start",   es == dt_time(4, 0))
    check("ext window finish",  ef == dt_time(9, 0))
    check("ext window pre-zone in",    time_in_window(dt_time(5, 0),  es, ef) is True)
    check("ext window pre-zone out",   time_in_window(dt_time(3, 59), es, ef) is False)
    check("ext window post-zone in",   time_in_window(dt_time(8, 59), es, ef) is True)
    check("ext window post-zone out",  time_in_window(dt_time(9, 1),  es, ef) is False)
    # Overnight base window 22:00–02:00 → extended 20:00–03:00
    es2, ef2 = ext("22:00", "02:00")
    check("ext overnight start",  es2 == dt_time(20, 0))
    check("ext overnight finish", ef2 == dt_time(3, 0))
    check("ext overnight pre in",  time_in_window(dt_time(21, 0), es2, ef2) is True)
    check("ext overnight post in", time_in_window(dt_time(2, 30), es2, ef2) is True)
    check("ext overnight mid out", time_in_window(dt_time(12, 0), es2, ef2) is False)
    check("ext window none",       ext("", "06:00") == (None, None))

    # stop_duration / drive_mins
    check("stop_duration zero", approx(stop_duration(0),   ONSITE_MIN))
    check("stop_duration pump", approx(stop_duration(PUMP_RATE_LPM), ONSITE_MIN + 1.0))
    check("drive_mins",         approx(drive_mins(DRIVE_SPEED_KMH), 60.0))
    check("drive_mins none",    drive_mins(None) is None)

    # _dest_vol_partial
    check("dest_partial rest",  approx(_dest_vol_partial({"vol_partial": None}, 1000, 200), 800))
    check("dest_partial cap",   approx(_dest_vol_partial({"vol_partial": 500}, 1000, 200), 500))
    check("dest_partial short", approx(_dest_vol_partial({"vol_partial": 500}, 600, 200), 400))

    # _block_has_split
    check("has_split true",  _block_has_split({"dests": [{"split_after": 1}]}) is True)
    check("has_split false", _block_has_split({"dests": [{"key": "x"}]}) is False)
    check("has_split empty", _block_has_split({}) is False)

    # _block_dest_offloads
    o1 = _block_dest_offloads({"rows": [{"prior_vol": 1000}],
                               "dests": [{"key": "A", "vol_partial": None}]})
    check("offload single", o1 == {"A": 1000.0})
    o2 = _block_dest_offloads({"rows": [{"prior_vol": 1000}],
                               "dests": [{"key": "A", "vol_partial": 600},
                                         {"key": "B", "vol_partial": None}]})
    check("offload split", approx(o2.get("A", 0), 600) and approx(o2.get("B", 0), 400))
    o3 = _block_dest_offloads({"rows": [{"prior_vol": 500}], "dests": [], "dest_key": "K"})
    check("offload dest_key fallback", o3 == {"K": 500.0})
    o4 = _block_dest_offloads({"rows": [{"prior_vol": 500}]})
    check("offload no-key fallback", o4 == {"?": 500.0})

    # _build_block_stops sequence + split interleaving
    stops = _build_block_stops(
        {"rows": [{"irma": "01-001"}], "dests": [{"key": "A", "name": "A"}]},
        "VEDDER", True)
    check("stops types", [s["type"] for s in stops] == ["origin", "farm", "dest", "vedder"])
    stops2 = _build_block_stops(
        {"rows": [{"irma": "01-001"}, {"irma": "02-002"}],
         "dests": [{"key": "B", "name": "B", "split_after": 1}]},
        "VEDDER", False)
    check("stops split interleave",
          [s["type"] for s in stops2] == ["origin", "farm", "dest", "farm"])

    # _apply_prob_floors_ceilings (ceiling clamp + floor raise, sum preserved)
    p = _apply_prob_floors_ceilings([0.5, 0.5], [0.0, 0.0], [0.3, 1.0])
    check("floors ceiling clamp", approx(p[0], 0.3) and approx(p[1], 0.7))
    p2 = _apply_prob_floors_ceilings([0.1, 0.9], [0.25, 0.0], [1.0, 1.0])
    check("floors floor raise", approx(p2[0], 0.25) and approx(p2[1], 0.75))
    check("floors sum preserved", approx(sum(p), 1.0) and approx(sum(p2), 1.0))

    print()
    if fails:
        print(f"SELFTEST FAILED - {len(fails)} failing check(s): {', '.join(fails)}")
        return 1
    print("SELFTEST PASSED - all checks ok")
    return 0


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setApplicationName("Vedder D100 Route Manager")

    # Populate THREE_WINDOW_FARMS before constructing MainWindow so that
    # calc_times and arrives_during_milking can use it from the start.
    global THREE_WINDOW_FARMS
    THREE_WINDOW_FARMS = _load_three_window_farms()
    win = MainWindow()   # dropdown signals are wired inside MainWindow.__init__
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    # Configure logging once, here, so library/import code stays quiet but
    # parsing and data-load warnings reach the console when the app runs.
    logging.basicConfig(level=logging.INFO,
                        format="%(levelname)s %(name)s: %(message)s")
    if "--selftest" in sys.argv:
        sys.exit(_run_selftests())
    main()
